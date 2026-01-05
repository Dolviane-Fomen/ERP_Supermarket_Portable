"""
Script de débogage pour l'export du rapport ACM
Ce script permet de tester l'export et identifier les problèmes
"""

import os
import sys
import django

# Configuration de Django
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'erp_project.settings')
django.setup()

from django.contrib.sessions.models import Session
from django.contrib.auth.models import User
from supermarket.models import Agence, Client, SuiviClientAction, Commande
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from collections import defaultdict

def test_export_rapport():
    """Test de l'export du rapport ACM"""
    
    print("=" * 80)
    print("SCRIPT DE DÉBOGAGE - EXPORT RAPPORT ACM")
    print("=" * 80)
    print()
    
    # 1. Vérifier les données de base
    print("1. Vérification des données de base...")
    try:
        agences = Agence.objects.all()
        print(f"   ✓ Nombre d'agences: {agences.count()}")
        if agences.exists():
            agence = agences.first()
            print(f"   ✓ Première agence: {agence.nom_agence}")
        else:
            print("   ✗ Aucune agence trouvée!")
            return
        
        clients = Client.objects.filter(agence=agence)
        print(f"   ✓ Nombre de clients: {clients.count()}")
        
        # Vérifier les actions
        date_debut = datetime.now() - timedelta(days=7)
        date_fin = datetime.now()
        
        actions = SuiviClientAction.objects.filter(
            client__agence=agence,
            date_action__gte=date_debut,
            date_action__lte=date_fin
        )
        print(f"   ✓ Nombre d'actions (7 derniers jours): {actions.count()}")
        
        commandes = Commande.objects.filter(
            client__agence=agence,
            date__gte=date_debut.date(),
            date__lte=date_fin.date()
        )
        print(f"   ✓ Nombre de commandes (7 derniers jours): {commandes.count()}")
        
    except Exception as e:
        print(f"   ✗ Erreur lors de la vérification: {str(e)}")
        import traceback
        traceback.print_exc()
        return
    
    print()
    
    # 2. Simuler la structure de données du rapport
    print("2. Simulation de la structure de données du rapport...")
    
    PLAGES_HORAIRES = [
        ('06h30-08h30', '06h30-08h30'),
        ('08h30-10h00', '08h30-10h00'),
        ('10h00-11h30', '10h00-11h30'),
        ('14h30-16h00', '14h30-16h00'),
        ('16h00-17h30', '16h00-17h30'),
    ]
    
    # Créer des données de test
    clients_data = []
    stats_par_vague = {}
    
    for plage_code, plage_label in PLAGES_HORAIRES:
        stats_par_vague[plage_code] = {
            'nb_appels': 0,
            'nb_reponses': 0,
            'nb_commandes': 0,
            'total_commandes': 0.0
        }
    
    # Récupérer quelques clients avec leurs données
    test_clients = clients[:5]  # Prendre les 5 premiers clients
    
    for client in test_clients:
        try:
            client_actions = actions.filter(client=client)
            client_commandes = commandes.filter(client=client)
            
            # Grouper par plage horaire
            actions_par_plage = {}
            commandes_par_plage = {}
            totaux_par_plage = {}
            
            for plage_code, plage_label in PLAGES_HORAIRES:
                actions_par_plage[plage_code] = []
                commandes_par_plage[plage_code] = []
                totaux_par_plage[plage_code] = 0.0
            
            # Traiter les actions
            for action in client_actions:
                heure_appel = action.heure_appel
                if heure_appel:
                    heure_str = heure_appel.strftime('%H:%M')
                    # Déterminer la plage horaire
                    plage_code = None
                    for code, label in PLAGES_HORAIRES:
                        debut, fin = label.split('-')
                        debut_h, debut_m = map(int, debut.replace('h', ':').split(':'))
                        fin_h, fin_m = map(int, fin.replace('h', ':').split(':'))
                        heure_int = heure_appel.hour * 60 + heure_appel.minute
                        debut_int = debut_h * 60 + debut_m
                        fin_int = fin_h * 60 + fin_m
                        if debut_int <= heure_int < fin_int:
                            plage_code = code
                            break
                    
                    if plage_code:
                        actions_par_plage[plage_code].append({
                            'heure_appel': heure_str,
                            'action': action.action or '',
                        })
            
            # Traiter les commandes
            for cmd in client_commandes:
                heure_cmd = cmd.heure
                if heure_cmd:
                    heure_str = heure_cmd.strftime('%H:%M')
                    # Déterminer la plage horaire
                    plage_code = None
                    for code, label in PLAGES_HORAIRES:
                        debut, fin = label.split('-')
                        debut_h, debut_m = map(int, debut.replace('h', ':').split(':'))
                        fin_h, fin_m = map(int, fin.replace('h', ':').split(':'))
                        heure_int = heure_cmd.hour * 60 + heure_cmd.minute
                        debut_int = debut_h * 60 + debut_m
                        fin_int = fin_h * 60 + fin_m
                        if debut_int <= heure_int < fin_int:
                            plage_code = code
                            break
                    
                    if plage_code:
                        # Récupérer les articles de la commande
                        articles_str = ''
                        if hasattr(cmd, 'lignes_commande'):
                            articles_list = [f"{l.article.nom_article} ({l.quantite})" for l in cmd.lignes_commande.all()]
                            articles_str = ', '.join(articles_list) if articles_list else f"{cmd.article.nom_article} ({cmd.quantite})"
                        else:
                            articles_str = f"{cmd.article.nom_article} ({cmd.quantite})"
                        
                        total_cmd = float(cmd.prix_total or 0)
                        commandes_par_plage[plage_code].append({
                            'date': cmd.date.strftime('%d/%m/%Y') if cmd.date else '',
                            'heure': heure_str,
                            'articles': articles_str,
                            'total': total_cmd,
                        })
                        totaux_par_plage[plage_code] += total_cmd
            
            total_client = sum(totaux_par_plage.values())
            
            client_data = {
                'id': client.id,
                'detail': client.detail or '',
                'ref': client.ref or '',
                'potentiel': client.potentiel or '',
                'nom': client.intitule or '',
                'telephone': client.telephone or '',
                'actions_par_plage': actions_par_plage,
                'commandes_par_plage': commandes_par_plage,
                'totaux_par_plage': totaux_par_plage,
                'total': total_client,
            }
            
            clients_data.append(client_data)
            print(f"   ✓ Client {client.intitule}: {len(client_actions)} actions, {len(client_commandes)} commandes")
            
        except Exception as e:
            print(f"   ✗ Erreur pour le client {client.id}: {str(e)}")
            import traceback
            traceback.print_exc()
            continue
    
    print(f"   ✓ Total clients traités: {len(clients_data)}")
    print()
    
    # 3. Tester la création du fichier Excel
    print("3. Test de création du fichier Excel...")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport ACM Test"
        
        # Styles
        header_fill = PatternFill(start_color="06beb6", end_color="48b1bf", fill_type="solid")
        header_font_bold = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-tête
        ws.merge_cells('A1:H1')
        ws['A1'] = f"RAPPORT ACM - {agence.nom_agence} (TEST)"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Période : {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # En-têtes de colonnes
        start_row = 4
        headers_fixes = ['DETAIL', 'REF', 'POTENTIEL /5', 'NOMS', 'TELEPHONES']
        
        col = 1
        for header in headers_fixes:
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            col += 1
        
        # En-têtes des plages horaires
        for plage_code, plage_label in PLAGES_HORAIRES:
            ws.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col+2)}{start_row}')
            cell = ws.cell(row=start_row, column=col)
            cell.value = plage_label
            cell.font = header_font_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            col += 3
        
        # Colonne Total
        cell = ws.cell(row=start_row, column=col)
        cell.value = 'TOTAL'
        cell.font = header_font_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Deuxième ligne : sous-en-têtes
        start_row2 = start_row + 1
        col = len(headers_fixes) + 1
        for _ in PLAGES_HORAIRES:
            ws.cell(row=start_row2, column=col, value='Heure').font = header_font_bold
            ws.cell(row=start_row2, column=col).fill = header_fill
            ws.cell(row=start_row2, column=col).border = border
            col += 1
            
            ws.cell(row=start_row2, column=col, value='Commandes').font = header_font_bold
            ws.cell(row=start_row2, column=col).fill = header_fill
            ws.cell(row=start_row2, column=col).border = border
            col += 1
            
            ws.cell(row=start_row2, column=col, value='Total').font = header_font_bold
            ws.cell(row=start_row2, column=col).fill = header_fill
            ws.cell(row=start_row2, column=col).border = border
            col += 1
        
        # Données des clients
        row = start_row2 + 1
        data_font = Font(size=10, name='Calibri')
        fill_blanc = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for idx, client_data in enumerate(clients_data):
            try:
                col = 1
                
                # Colonnes fixes
                ws.cell(row=row, column=col, value=client_data.get('detail', '')).border = border
                ws.cell(row=row, column=col).font = data_font
                col += 1
                
                ws.cell(row=row, column=col, value=client_data.get('ref', '')).border = border
                ws.cell(row=row, column=col).font = data_font
                col += 1
                
                ws.cell(row=row, column=col, value=client_data.get('potentiel', '')).border = border
                ws.cell(row=row, column=col).font = data_font
                col += 1
                
                ws.cell(row=row, column=col, value=client_data.get('nom', '')).border = border
                ws.cell(row=row, column=col).font = Font(bold=True, size=10, name='Calibri')
                col += 1
                
                ws.cell(row=row, column=col, value=client_data.get('telephone', '')).border = border
                ws.cell(row=row, column=col).font = data_font
                col += 1
                
                # Plages horaires
                actions_par_plage = client_data.get('actions_par_plage', {})
                commandes_par_plage = client_data.get('commandes_par_plage', {})
                
                for plage_code, plage_label in PLAGES_HORAIRES:
                    actions_list = actions_par_plage.get(plage_code, [])
                    commandes_plage = commandes_par_plage.get(plage_code, [])
                    
                    # Heure
                    all_heures = []
                    for action in actions_list:
                        if action.get('heure_appel'):
                            all_heures.append(action['heure_appel'])
                    for cmd in commandes_plage:
                        if cmd.get('heure') and cmd['heure'] not in all_heures:
                            all_heures.append(cmd['heure'])
                    
                    heures_str = '\n'.join(all_heures) if all_heures else ''
                    ws.cell(row=row, column=col, value=heures_str).border = border
                    ws.cell(row=row, column=col).font = data_font
                    col += 1
                    
                    # Commandes
                    all_texts = []
                    for action in actions_list:
                        if action.get('action'):
                            all_texts.append(action['action'])
                    for cmd in commandes_plage:
                        if cmd.get('articles'):
                            all_texts.append(cmd['articles'])
                    
                    actions_str = '\n'.join(all_texts) if all_texts else ''
                    ws.cell(row=row, column=col, value=actions_str).border = border
                    ws.cell(row=row, column=col).font = data_font
                    col += 1
                    
                    # Total
                    total_plage = client_data.get('totaux_par_plage', {}).get(plage_code, 0)
                    total_plage_str = f"{total_plage:.0f} FCFA" if total_plage > 0 else ''
                    ws.cell(row=row, column=col, value=total_plage_str).border = border
                    ws.cell(row=row, column=col).font = data_font
                    col += 1
                
                # Total général
                total_client = client_data.get('total', 0)
                total_str = f"{total_client:.0f} FCFA" if total_client > 0 else ''
                ws.cell(row=row, column=col, value=total_str).border = border
                ws.cell(row=row, column=col).font = Font(bold=True, size=10, name='Calibri')
                
                row += 1
                print(f"   ✓ Ligne client {idx + 1} ajoutée")
                
            except Exception as e:
                print(f"   ✗ Erreur ligne client {idx + 1}: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        # Sauvegarder le fichier
        filename = f"rapport_acm_test_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        print(f"   ✓ Fichier Excel créé: {filename}")
        print(f"   ✓ Chemin complet: {os.path.abspath(filename)}")
        
    except Exception as e:
        print(f"   ✗ Erreur lors de la création du fichier Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return
    
    print()
    print("=" * 80)
    print("TEST TERMINÉ")
    print("=" * 80)
    print()
    print("Vérifications à faire:")
    print("1. Ouvrir le fichier Excel généré")
    print("2. Vérifier que toutes les colonnes sont présentes")
    print("3. Vérifier que les données des clients sont correctes")
    print("4. Vérifier les statistiques")
    print()

if __name__ == '__main__':
    test_export_rapport()

