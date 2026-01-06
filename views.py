from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Max, Q, F
from django.db import models, connection
from django.utils import timezone
from decimal import Decimal, InvalidOperation
import re
from .models import Agence, Compte, Employe, Caisse, SessionCaisse, Client, InventaireStock, LigneInventaireStock, StatistiqueVente, MouvementStock, Article, TypeVente, FactureVente, LigneFactureVente, DocumentVente, FactureTemporaire, Famille, Fournisseur, MouvementStock, PlanComptable, PlanTiers, CodeJournaux, TauxTaxe, FactureAchat, LigneFactureAchat, FactureTransfert, LigneFactureTransfert

def normalize_decimal_input(value):
    """
    Normalise les entrées décimales pour éviter les erreurs de conversion
    Utilisé dans les vues de facturation pour corriger le problème des virgules déplacées
    Version améliorée pour gérer les formats français et internationaux
    """
    if value is None:
        return Decimal('0')
    
    # Convertir en chaîne
    value_str = str(value).strip()
    
    # Remplacer les virgules par des points (format français vers format international)
    value_str = value_str.replace(',', '.')
    
    # Supprimer les caractères non numériques sauf point et moins
    # Garder aussi les espaces pour les séparateurs de milliers français
    value_str = re.sub(r'[^\d.\-\s]', '', value_str)
    
    # Supprimer les espaces (séparateurs de milliers français)
    value_str = value_str.replace(' ', '')
    
    # Gérer les cas vides
    if not value_str or value_str in ['-', '.']:
        return Decimal('0')
    
    # S'assurer qu'il n'y a qu'un seul point décimal
    parts = value_str.split('.')
    if len(parts) > 2:
        value_str = parts[0] + '.' + ''.join(parts[1:])
    
    try:
        return Decimal(value_str)
    except (InvalidOperation, ValueError):
        print(f"[WARNING] Erreur conversion Decimal: {value} -> valeur par défaut 0")
        return Decimal('0')

def safe_decimal_calculation(value1, value2, operation='multiply'):
    """
    Effectue des calculs décimaux sécurisés pour éviter les erreurs de précision
    """
    try:
        dec1 = normalize_decimal_input(value1)
        dec2 = normalize_decimal_input(value2)
        
        if operation == 'multiply':
            return dec1 * dec2
        elif operation == 'add':
            return dec1 + dec2
        elif operation == 'subtract':
            return dec1 - dec2
        elif operation == 'divide':
            if dec2 == 0:
                return Decimal('0')
            return dec1 / dec2
        else:
            return Decimal('0')
    except Exception as e:
        print(f"[WARNING] Erreur calcul décimal: {e}")
        return Decimal('0')

def safe_quantity_conversion(quantity_value):
    """
    Conversion sécurisée des quantités avec gestion des décimales
    """
    return normalize_decimal_input(quantity_value)

def safe_price_conversion(price_value):
    """
    Conversion sécurisée des prix avec gestion des décimales
    """
    return normalize_decimal_input(price_value)

def get_user_agence(request):
    """Récupérer l'agence de l'utilisateur connecté"""
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
        print(f"[ALERTE] get_user_agence: {request.user.username} -> {compte.agence.nom_agence}")
        return compte.agence
    except Compte.DoesNotExist:
        print(f"[ALERTE] get_user_agence: {request.user.username} -> AUCUN COMPTE")
        return None

def login_caisse(request):
    """Page de connexion pour la gestion de caisse"""
    if request.user.is_authenticated:
        return redirect('dashboard_caisse')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                # Vérifier que l'utilisateur a un compte lié à une agence
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        login(request, user)
                        # Stocker l'agence dans la session
                        request.session['agence_id'] = compte.agence.id_agence
                        request.session['agence_nom'] = compte.agence.nom_agence
                        messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                        return redirect('dashboard_caisse')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Compte non trouvé ou inactif.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')
    
    # Récupérer les agences disponibles pour affichage
    agences = Agence.objects.all()
    context = {
        'agences': agences
    }
    return render(request, 'supermarket/caisse/login.html', context)

def logout_caisse(request):
    """Déconnexion de la gestion de caisse"""
    if request.user.is_authenticated:
        logout(request)
        # Nettoyer la session
        request.session.flush()
        messages.info(request, 'Vous avez été déconnecté avec succès.')
    return redirect('index')

def index(request):
    return render(request, 'supermarket/index.html')



@login_required
def dashboard_caisse(request):
    # Récupérer l'agence de l'utilisateur connecté
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    # Récupérer le compte de l'utilisateur
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
        return redirect('logout_caisse')
    
    caisse_ouverte = False

    try:
        aujourd_hui = timezone.now().date()

        session_caisse = SessionCaisse.objects.filter(

            agence=agence, 

            date_ouverture__date=aujourd_hui, 

            statut='ouverte'

        ).first()

        caisse_ouverte = session_caisse is not None

    except:
        pass
    
    chiffre_affaires = 0
    nombre_ventes = 0
    tickets_attente = 0

    try:

        # Récupérer les ventes du jour liées aux sessions de caisse
        ventes_jour = FactureVente.objects.filter(
            agence=agence, 
            date=aujourd_hui,
            session_caisse__isnull=False
        )
        chiffre_affaires = ventes_jour.aggregate(total=Sum('nette_a_payer'))['total'] or 0

        nombre_ventes = ventes_jour.count()

        # Compter les factures temporaires UNIQUEMENT pour la session courante ouverte
        if session_caisse:
            tickets_attente = FactureTemporaire.objects.filter(session_caisse=session_caisse).count()
        else:
            tickets_attente = 0

    except:

        pass
    

    
    
    # Récupérer l'employé associé
    employe = None
    try:
        employe = Employe.objects.get(compte=compte)
    except Employe.DoesNotExist:
        pass
    
    return render(request, 'supermarket/caisse/dashboard.html', {
        'agence': agence, 
        'compte': compte,
        'employe': employe,
        'caisse_ouverte': caisse_ouverte, 
        'chiffre_affaires': chiffre_affaires,
        'nombre_ventes': nombre_ventes, 
        'tickets_attente': tickets_attente, 
        'session_caisse': session_caisse
    })

@login_required
def dashboard_kpis_api(request):
    print(f"[ALERTE] DASHBOARD_KPIS_API: {request.user.username}")
    
    agence = get_user_agence(request)
    if not agence:
        print(f"[ALERTE] DASHBOARD_KPIS_API: Aucune agence pour {request.user.username}")

        return JsonResponse({

            'chiffre_affaires': 0, 

            'nombre_ventes': 0, 

            'tickets_attente': 0, 

            'caisse_ouverte': False,

            'sessions_info': [],

            'premiere_ouverture': None

        })
    
    print(f"[ALERTE] DASHBOARD_KPIS_API: Agence {agence.nom_agence}")
    
    
    
    aujourd_hui = timezone.now().date()

    
    
    # Calculer les KPIs de manière simple d'abord

    chiffre_affaires = 0

    nombre_ventes = 0

    tickets_attente = 0
    
    caisse_ouverte = False

    

    try:

        # Récupérer la session de caisse active
        session_ouverte = SessionCaisse.objects.filter(
            agence=agence, 
            date_ouverture__date=aujourd_hui, 
            statut='ouverte'
        ).first()
        
        if session_ouverte:
            print(f"[ALERTE] DASHBOARD_KPIS_API: Session trouvée {session_ouverte.id}")
            # Récupérer seulement les factures de la session active
            ventes_jour = FactureVente.objects.filter(
                agence=agence, 
                date=aujourd_hui,
                session_caisse=session_ouverte
            )
            print(f"[ALERTE] DASHBOARD_KPIS_API: {ventes_jour.count()} factures trouvées")
            
            # Si aucune facture liée à la session, récupérer toutes les factures du jour et les lier
            if ventes_jour.count() == 0:
                print(f"[ALERTE] DASHBOARD_KPIS_API: Aucune facture liée à la session, recherche des factures orphelines")
                factures_orphelines = FactureVente.objects.filter(
                    agence=agence,
                    date=aujourd_hui,
                    session_caisse__isnull=True
                )
                
                if factures_orphelines.count() > 0:
                    print(f"[ALERTE] DASHBOARD_KPIS_API: {factures_orphelines.count()} factures orphelines trouvées, liaison à la session")
                    # Lier ces factures à la session actuelle
                    factures_orphelines.update(session_caisse=session_ouverte)
                    
                    # Récupérer maintenant toutes les factures de la session
                    ventes_jour = FactureVente.objects.filter(
                        agence=agence, 
                        date=aujourd_hui,
                        session_caisse=session_ouverte
                    )
                    print(f"[ALERTE] DASHBOARD_KPIS_API: {ventes_jour.count()} factures après liaison")
        else:
            print(f"[ALERTE] DASHBOARD_KPIS_API: Aucune session active")
            # Aucune session active, donc pas de ventes
            ventes_jour = FactureVente.objects.none()
        
        
        chiffre_affaires = ventes_jour.aggregate(total=Sum('nette_a_payer'))['total'] or 0

        nombre_ventes = ventes_jour.count()

        # Compter les factures temporaires en attente UNIQUEMENT pour la session courante
        if session_ouverte:
            tickets_attente = FactureTemporaire.objects.filter(session_caisse=session_ouverte).count()
        else:
            tickets_attente = 0
        
        
    except Exception as e:
    

        pass
    

    # Vérifier le statut de la caisse (session_ouverte déjà récupérée plus haut)
    try:

        caisse_ouverte = session_ouverte is not None
        
        # Récupérer toutes les sessions pour l'affichage
        sessions_caisse = SessionCaisse.objects.filter(

            agence=agence, 

            date_ouverture__date=aujourd_hui

        ).order_by('date_ouverture')
        
        
        
    except Exception as e:

        pass
    
    
    # Préparer les informations des sessions

    sessions_info = []

    try:

        for i, session in enumerate(sessions_caisse[:2]):  # Max 2 sessions


            session_data = {

                'numero': i + 1,

                'ouverture': session.date_ouverture.strftime('%H:%M') if session.date_ouverture else '',

                'fermeture': session.date_fermeture.strftime('%H:%M') if session.date_fermeture else '',

                'ouverte_par': session.employe.compte.nom_complet if session.employe and hasattr(session.employe, 'compte') else 'Non spécifié',

                'fermee_par': session.employe.compte.nom_complet if session.employe and session.date_fermeture and hasattr(session.employe, 'compte') else '',


                'statut': 'Ouverte' if session.statut == 'ouverte' else 'Fermée'

            }

            sessions_info.append(session_data)

    except Exception as e:

        print(f"DEBUG: Erreur sessions_info: {str(e)}")
    

    
    
    result = {

        'chiffre_affaires': float(chiffre_affaires),

        'nombre_ventes': nombre_ventes,

        'tickets_attente': tickets_attente,

        'caisse_ouverte': caisse_ouverte,

        'sessions_info': sessions_info,


        'premiere_ouverture': None,

        'nombre_sessions': len(sessions_info)

    }

    
    
    print(f"DEBUG: Résultat final: {result}")

    
    
    return JsonResponse(result)



def generate_ticket_number(agence):

    """Générer un numéro de ticket auto-incrémenté"""

    try:

        # Obtenir la date actuelle

        aujourd_hui = timezone.now().date()

        
        
        # Chercher le dernier numéro de ticket pour aujourd'hui

        dernier_ticket = FactureVente.objects.filter(

            agence=agence,

            date=aujourd_hui

        ).aggregate(

            max_numero=Max('numero_ticket')

        )['max_numero']

        
        
        if dernier_ticket:

            # Extraire le numéro du dernier ticket (format: TKT20250923001)

            try:

                # Le format est TKT + YYYYMMDD + 3 chiffres

                numero_sequence = int(dernier_ticket[-3:])

                nouveau_numero = numero_sequence + 1

            except (ValueError, IndexError):

                nouveau_numero = 1

        else:

            nouveau_numero = 1
        
        
        
        # Formater le nouveau numéro de ticket

        date_str = aujourd_hui.strftime('%Y%m%d')

        numero_ticket = f"TKT{date_str}{nouveau_numero:03d}"

        # Vérifier que le numéro n'existe pas déjà
        while FactureVente.objects.filter(numero_ticket=numero_ticket).exists():
            nouveau_numero += 1
            numero_ticket = f"TKT{date_str}{nouveau_numero:03d}"

        print(f"[ALERTE] NUMÉRO TICKET GÉNÉRÉ: {numero_ticket}")
        
        return numero_ticket
        
        
        
    except Exception as e:

        print(f"Erreur lors de la génération du numéro de ticket: {e}")

        # En cas d'erreur, utiliser un format simple avec timestamp

        timestamp = timezone.now().strftime('%Y%m%d%H%M%S')

        return f"TKT{timestamp}"

    

@login_required
def facturation_vente(request, facture_id=None):
    # Récupérer l'agence de l'utilisateur connecté
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    
    
    # Vérifier le statut de la caisse pour l'affichage mais ne pas bloquer l'accès

    caisse_ouverte = False

    caisse_actuelle = None

    try:

        aujourd_hui = timezone.now().date()

        session_caisse = SessionCaisse.objects.filter(

            agence=agence, 

            date_ouverture__date=aujourd_hui, 

            statut='ouverte'

        ).first()

        caisse_ouverte = session_caisse is not None

        if session_caisse:

            caisse_actuelle = session_caisse.caisse

    except:

        pass
    
    
    
    # Si aucune caisse ouverte, utiliser la première caisse disponible

    if not caisse_actuelle:

        caisse_actuelle = Caisse.objects.filter(agence=agence, statut='active').first()

        if not caisse_actuelle:

            # Chercher une caisse inactive et l'activer
            caisse_inactive = Caisse.objects.filter(agence=agence).first()
            if caisse_inactive:
                caisse_inactive.statut = 'active'
                caisse_inactive.save()
                caisse_actuelle = caisse_inactive
            else:
                # Créer une caisse par défaut si aucune n'existe
                caisse_actuelle, created = Caisse.objects.get_or_create(
                    numero_caisse='CAISSE001',
                    defaults={
                        'nom_caisse': 'Caisse Principale',
                        'agence': agence,
                        'solde_initial': 0,
                        'solde_actuel': 0,
                        'statut': 'active'
                    }
                )
    
    # Créer une session si aucune n'est ouverte
    if not caisse_ouverte and caisse_actuelle:
        try:
            # Trouver un employé de cette agence
            compte = Compte.objects.filter(agence=agence).first()
            if compte:
                employe = Employe.objects.filter(compte=compte).first()
                if employe:
                    # Fermer les sessions ouvertes existantes
                    SessionCaisse.objects.filter(agence=agence, statut='ouverte').update(statut='fermee')
                    
                    # Créer une nouvelle session
                    SessionCaisse.objects.create(
                        agence=agence,
                        caisse=caisse_actuelle,
                        employe=employe,
                        date_ouverture=timezone.now(),
                        solde_ouverture=0,
                        statut='ouverte'
                    )
        except:
            pass
    
    # Générer le numéro de ticket auto-incrémenté

    numero_ticket = generate_ticket_number(agence)

    
    
    # Récupérer la facture temporaire de la session

    facture_temp = request.session.get('facture_temporaire', {

        'lignes': [], 'type_vente': 'detail', 'remise': 0, 'montant_regler': 0, 'nette_a_payer': 0, 'rendu': 0

    })

    
    
    print(f"DEBUG facturation_vente: facture_temp récupérée = {facture_temp}")

    print(f"DEBUG facturation_vente: type de facture_temp = {type(facture_temp)}")

    print(f"DEBUG facturation_vente: lignes = {facture_temp.get('lignes', [])}")

    print(f"DEBUG facturation_vente: nombre de lignes = {len(facture_temp.get('lignes', []))}")

    
    
    # Si la session est vide, essayer de récupérer depuis les cookies ou créer une nouvelle

    if not facture_temp or not facture_temp.get('lignes'):

        print("DEBUG: Session vide, création d'une nouvelle facture temporaire")

        facture_temp = {

            'lignes': [], 

            'type_vente': 'detail', 

            'remise': 0, 

            'montant_regler': 0, 

            'nette_a_payer': 0, 

            'rendu': 0

        }

        request.session['facture_temporaire'] = facture_temp

    
    
    # Récupérer les clients pour le formulaire

    clients = Client.objects.filter(agence=agence).order_by('intitule')

    
    
    # Récupérer l'employé de l'utilisateur connecté
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
        employe = Employe.objects.filter(compte=compte).first()
        
        # Si pas d'employé, créer les infos à partir du compte
        if not employe:
            # L'utilisateur a un compte mais pas de fiche employé
            # On utilise les infos du compte directement
            employe = type('obj', (object,), {
                'compte': type('obj', (object,), {
                    'nom': compte.nom,
                    'prenom': compte.prenom,
                    'nom_complet': compte.nom_complet
                })()
            })()
    except Compte.DoesNotExist:
        # Fallback : utiliser le username
        employe = type('obj', (object,), {
            'compte': type('obj', (object,), {
                'nom': request.user.last_name or request.user.username,
                'prenom': request.user.first_name or '',
                'nom_complet': f"{request.user.first_name or ''} {request.user.last_name or request.user.username}".strip()
            })()
        })()

    
    
    return render(request, 'supermarket/caisse/facturation_vente.html', {

        'agence': agence, 

        'facture_temp': facture_temp, 

        'caisse_ouverte': caisse_ouverte,

        'caisse_actuelle': caisse_actuelle,

        'numero_ticket': numero_ticket,

        'clients': clients,

        'employe': employe,

        'caisses': Caisse.objects.filter(agence=agence, statut='active'),

        'today': aujourd_hui

    })



@login_required
def ouvrir_caisse(request):
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    
    
    try:

        caisse, created = Caisse.objects.get_or_create(

            numero_caisse='CAISSE001',

            defaults={'agence': agence, 'solde_actuel': 0, 'statut': 'active'}

        )

        
        
        employe = Employe.objects.filter(compte__agence=agence).first()

        if not employe:

            messages.error(request, 'Aucun employé trouvé.')
            return redirect('dashboard_caisse')
        
        
        
        aujourd_hui = timezone.now().date()

        sessions_ouvertes = SessionCaisse.objects.filter(agence=agence, date_ouverture=aujourd_hui, statut='ouverte').count()

        
        
        if sessions_ouvertes > 0:

            messages.info(request, 'Caisse déjà ouverte.')

            return redirect('dashboard_caisse')
        
        
        
        session = SessionCaisse.objects.create(

            caisse=caisse, utilisateur=None, employe=employe, agence=agence,

            solde_ouverture=caisse.solde_actuel, statut='ouverte'

        )

        
        # Vider les factures temporaires en attente pour la nouvelle session
        FactureTemporaire.objects.all().delete()
        
        
        messages.success(request, f'Caisse ouverte avec succès. Solde: {caisse.solde_actuel} FCFA')

        return redirect('dashboard_caisse')
        
        
        
    except Exception as e:

        messages.error(request, f'Erreur: {str(e)}')

        return redirect('dashboard_caisse')




@login_required
def search_window(request):

    search_term = request.GET.get('q', '')

    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    
    
    articles = []

    if search_term and len(search_term) >= 2:


        articles = Article.objects.filter(agence=agence, designation__icontains=search_term)[:50]
    
    
    
    return render(request, 'supermarket/caisse/search_window.html', {

        'articles': articles, 'search_term': search_term

    })



@login_required
def search_articles_api(request):

    search_term = request.GET.get('q', '')

    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'articles': []})
    
    
    
    articles = []

    if search_term:

        articles = Article.objects.filter(agence=agence, designation__icontains=search_term).order_by('designation')


    else:


        articles = Article.objects.filter(agence=agence).order_by('designation')
    
    
    
    results = []

    for article in articles:

        results.append({

            'id': article.id, 'reference': article.reference_article, 'designation': article.designation,

            'prix_achat': float(article.prix_achat or 0), 'prix_vente': float(article.prix_vente or 0),

            'stock': int(article.stock_actuel or 0)

        })
    
    
    
    return JsonResponse({'articles': results})



def get_prix_by_type(request):

    article_id = request.GET.get('article_id')

    type_vente = request.GET.get('type_vente', 'detail')

    
    
    try:

        article = Article.objects.get(id=article_id)

        type_vente_obj = TypeVente.objects.filter(article=article, intitule__iexact=type_vente).first()

        
        
        if type_vente_obj:

            prix = float(type_vente_obj.prix)

        else:

            prix = float(article.prix_vente or 0)
        
        
        
        return JsonResponse({'success': True, 'prix': prix})

    except Exception as e:

        return JsonResponse({'success': False, 'prix': 0, 'error': str(e)})



@login_required
def ajouter_article_facture(request):

    if request.method == 'POST':
        
        agence = get_user_agence(request)
        if not agence:

            return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
            
            
            
        try:

            article_id = request.POST.get('article_id')

            quantite = int(request.POST.get('quantite', 1))

            
            
            print(f"DEBUG: Article ID reçu: {article_id}")

            print(f"DEBUG: Quantité reçue: {quantite}")

            
            
            article = Article.objects.get(id=article_id, agence=agence)

            print(f"DEBUG: Article trouvé: {article.designation}")

            
            
            facture_temp = request.session.get('facture_temporaire', {

                'lignes': [], 'type_vente': 'detail', 'remise': 0, 'montant_regler': 0, 'nette_a_payer': 0, 'rendu': 0

            })

            
            
            print(f"DEBUG: Facture temp avant: {facture_temp}")

            
            
            article_existe = False

            for ligne in facture_temp['lignes']:

                # Comparer les IDs en tant que strings pour éviter les problèmes de type

                if str(ligne['article_id']) == str(article_id):

                    ligne['quantite'] += quantite

                    ligne['prix_total'] = safe_decimal_calculation(ligne['quantite'], ligne['prix_unitaire'], 'multiply')

                    article_existe = True

                    print(f"DEBUG: Article existant mis à jour")

                    break
            
            
            
            if not article_existe:

                prix_unitaire = normalize_decimal_input(article.prix_vente or 0)
                prix_total = safe_decimal_calculation(quantite, prix_unitaire, 'multiply')

                
                
                nouvelle_ligne = {

                    'article_id': article_id, 'reference': article.reference_article, 'designation': article.designation,

                    'quantite': quantite, 'prix_unitaire': prix_unitaire, 'prix_total': prix_total, 'type_vente': 'detail'

                }

                facture_temp['lignes'].append(nouvelle_ligne)

            
            
            
            request.session['facture_temporaire'] = facture_temp

            
            
            return JsonResponse({
                'success': True, 
                'message': f'Article "{article.designation}" ajouté avec succès!',
                'article': {

                    'id': article_id,

                    'designation': article.designation,

                    'reference': article.reference_article,

                    'prix_vente': float(article.prix_vente or 0)

                }

            })
            
            
            
        except Exception as e:

            print(f"DEBUG: Erreur: {str(e)}")

            return JsonResponse({'success': False, 'error': str(e)})



    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})



@csrf_exempt
@login_required
def enregistrer_facture(request):
    try:
        print("=" * 80)
        print("*** NOUVEAU CODE VERSION 2.0 CHARGE AVEC SUCCES ***")
        print("=" * 80)
        print("[ALERTE] ENREGISTRER_FACTURE APPELÉE")
        
        if request.method == 'POST':
            print("[ALERTE] MÉTHODE POST DÉTECTÉE")
        
        # Vérifier spécifiquement le champ facture_data
        facture_data = request.POST.get('facture_data', '')
        print(f"[ALERTE] facture_data: {facture_data[:100] if facture_data else 'VIDE'}")

        
        
        # Récupérer l'agence de l'utilisateur connecté
        agence = get_user_agence(request)
        if not agence:
            print(f"🔴 DEBUG: [ERREUR] Aucune agence trouvée pour l'utilisateur")
            return JsonResponse({
                'success': False,
                'error': 'Votre compte n\'est pas configuré correctement.'
            })
        
        print(f"🔴 DEBUG: [OK] Agence de l'utilisateur: {agence.nom_agence} (ID: {agence.id_agence})")

        # S'assurer qu'il y a une caisse active pour cette agence
        caisse_actuelle = Caisse.objects.filter(agence=agence, statut='active').first()
        if not caisse_actuelle:
            # Chercher une caisse inactive et l'activer
            caisse_inactive = Caisse.objects.filter(agence=agence).first()
            if caisse_inactive:
                caisse_inactive.statut = 'active'
                caisse_inactive.save()
                caisse_actuelle = caisse_inactive
                print(f"🔴 DEBUG: [OK] Caisse activée: {caisse_actuelle.numero_caisse}")
            else:
                # Créer une caisse par défaut si aucune n'existe
                caisse_actuelle, created = Caisse.objects.get_or_create(
                    numero_caisse='CAISSE001',
                    defaults={
                        'nom_caisse': 'Caisse Principale',
                        'agence': agence,
                        'solde_initial': 0,
                        'solde_actuel': 0,
                        'statut': 'active'
                    }
                )
                print(f"🔴 DEBUG: [OK] Caisse créée: {caisse_actuelle.numero_caisse}")

        # S'assurer qu'il y a une session ouverte
        session_ouverte = SessionCaisse.objects.filter(agence=agence, statut='ouverte').first()
        if not session_ouverte and caisse_actuelle:
            try:
                # Trouver un employé de cette agence
                compte = Compte.objects.filter(agence=agence).first()
                if compte:
                    employe = Employe.objects.filter(compte=compte).first()
                    if employe:
                        # Fermer les sessions ouvertes existantes
                        SessionCaisse.objects.filter(agence=agence, statut='ouverte').update(statut='fermee')
                        
                        # Créer une nouvelle session
                        session_ouverte = SessionCaisse.objects.create(
                            agence=agence,
                            caisse=caisse_actuelle,
                            employe=employe,
                            date_ouverture=timezone.now(),
                            solde_ouverture=0,
                            statut='ouverte'
                        )
                        print(f"🔴 DEBUG: [OK] Session créée: {session_ouverte.id}")
            except Exception as e:
                print(f"🔴 DEBUG: [ERREUR] Erreur création session: {e}")

        try:

            # Logs détaillés pour debug
            print("=" * 80)
            print("DEBUG ENREGISTRER_FACTURE: Début de la fonction")
            print(f"DEBUG: Méthode de requête: {request.method}")
            print(f"DEBUG: POST data: {dict(request.POST)}")
            print(f"DEBUG: Session key: {request.session.session_key}")
            print(f"DEBUG: Session data: {dict(request.session)}")
            
            # NOUVELLE LOGIQUE: Récupérer les articles du POST en priorité
            facture_temp = {'lignes': [], 'remise': 0, 'montant_regler': 0, 'rendu': 0}
            articles_from_post = []
            
            # Méthode 1: Chercher dans le champ facture_data (JSON)
            facture_data_str = request.POST.get('facture_data', '')
            if facture_data_str:
                try:
                    import json
                    facture_data_parsed = json.loads(facture_data_str)
                    lignes = facture_data_parsed.get('lignes', [])
                    if lignes:
                        print(f"🔴 DEBUG: {len(lignes)} articles récupérés depuis facture_data JSON")
                        facture_temp['lignes'] = lignes
                        facture_temp['remise'] = facture_data_parsed.get('remise', 0)
                        facture_temp['montant_regler'] = facture_data_parsed.get('montant_regler', 0)
                        facture_temp['rendu'] = facture_data_parsed.get('rendu', 0)
                except json.JSONDecodeError as e:
                    print(f"🔴 DEBUG: Erreur parsing JSON facture_data: {e}")
            
            # Méthode 2: Si pas de facture_data, chercher les champs article_X_id dans POST
            if not facture_temp.get('lignes'):
                print("🔴 DEBUG: Tentative de récupération des articles depuis les champs POST")
                for key, value in request.POST.items():
                    if key.startswith('article_') and key.endswith('_id') and value:
                        article_id = value
                        prefix = key.replace('_id', '')
                        quantite = request.POST.get(f'{prefix}_quantite', 1)
                        prix_unitaire = request.POST.get(f'{prefix}_prix_unitaire', 0)
                        prix_total = request.POST.get(f'{prefix}_prix_total', 0)
                        designation = request.POST.get(f'{prefix}_designation', '')
                        reference = request.POST.get(f'{prefix}_reference', '')
                        
                        try:
                            article = Article.objects.get(id=article_id)
                            articles_from_post.append({
                                'article_id': int(article_id),
                                'designation': designation or article.designation,
                                'quantite': int(quantite),
                                'prix_unitaire': float(prix_unitaire),
                                'prix_total': float(prix_total),
                                'reference': reference or article.reference_article
                            })
                            print(f"🔴 DEBUG: Article récupéré depuis POST: {designation} (ID: {article_id})")
                        except Article.DoesNotExist:
                            print(f"🔴 DEBUG: Article avec ID {article_id} non trouvé")
                        except Exception as e:
                            print(f"🔴 DEBUG: Erreur récupération article {article_id}: {e}")
                
                if articles_from_post:
                    facture_temp['lignes'] = articles_from_post
                    print(f"🔴 DEBUG: {len(articles_from_post)} articles récupérés depuis les champs POST")
            
            # Méthode 3: Si toujours pas d'articles, chercher dans la session
            if not facture_temp.get('lignes'):
                facture_temp_session = request.session.get('facture_temporaire', {'lignes': []})
                if facture_temp_session.get('lignes'):
                    facture_temp = facture_temp_session
                    print(f"🔴 DEBUG: {len(facture_temp['lignes'])} articles récupérés depuis la session")
            
            # Méthode 4: Si toujours pas d'articles, chercher dans FactureTemporaire (DB)
            if not facture_temp.get('lignes'):
                try:
                    session_key = request.session.session_key
                    if session_key:
                        facture_temp_db = FactureTemporaire.objects.filter(session_key=session_key).first()
                        if facture_temp_db and facture_temp_db.contenu:
                            import json
                            facture_temp = json.loads(facture_temp_db.contenu)
                            print(f"🔴 DEBUG: {len(facture_temp.get('lignes', []))} articles récupérés depuis FactureTemporaire DB")
                except Exception as e:
                    print(f"🔴 DEBUG: Erreur récupération depuis DB: {e}")
            
            print(f"🔴 DEBUG: [FINAL] Nombre total d'articles: {len(facture_temp.get('lignes', []))}")
            
            # Si toujours aucun article, retourner une erreur
            if not facture_temp.get('lignes'):
                print("🔴 DEBUG: [ERREUR FINALE] Aucun article trouvé après toutes les tentatives")
                return JsonResponse({
                    'success': False,
                    'error': 'Aucun article dans la facture. Veuillez ajouter des articles.'
                })
            
            print(f"🔴 DEBUG: [LIST] Articles à enregistrer:")
            for i, ligne in enumerate(facture_temp.get('lignes', [])):
                print(f"🔴 DEBUG: Ligne {i}: article_id={ligne.get('article_id', 'VIDE')}, designation={ligne.get('designation', 'VIDE')}, qté={ligne.get('quantite', 0)}")

            # Passer au traitement (la suite du code reste inchangée)
            # Utiliser la caisse actuelle (déjà vérifiée/créée plus haut)
            caisse = caisse_actuelle

            print(f"🔴 DEBUG: [OK] Caisse utilisée: {caisse.numero_caisse} (ID: {caisse.id})")

            
            
            # Récupérer le client

            client_id = request.POST.get('client_id')

            client_name = request.POST.get('client_name')

            print(f"DEBUG: Client ID: {client_id}, Client Name: {client_name}")

            client = None
            
            if client_id:
                try:
                    # Essayer de récupérer le client par ID (sans filtrer par agence)
                    client = Client.objects.get(id=client_id)
                    print(f"DEBUG: Client existant trouvé: {client}")
                    
                    # Vérifier si le client appartient à une autre agence
                    if client.agence != agence:
                        print(f"DEBUG: ATTENTION - Client de l'agence {client.agence.nom_agence}, utilisateur de l'agence {agence.nom_agence}")
                        # On peut soit utiliser ce client, soit utiliser le client par défaut de l'agence
                        # Pour l'instant, on utilise le client trouvé (vente inter-agence possible)
                except Client.DoesNotExist:
                    print(f"DEBUG: Client avec ID {client_id} non trouvé, utilisation du client par défaut")
                    client = None

            elif client_name:

                # Créer un nouveau client si le nom est fourni

                client = Client.objects.create(

                    intitule=client_name,

                    adresse='Adresse non spécifiée',

                    telephone='Non spécifié',

                    email='',

                    agence=agence

                )

                print(f"DEBUG: Nouveau client créé: {client}")

            else:
                # Utiliser un client par défaut
                client = Client.objects.filter(agence=agence).first()

            if not client:
                client = Client.objects.create(
                    intitule='Client Général',
                    adresse='Adresse non spécifiée',
                    telephone='Non spécifié',
                    email='',
                    agence=agence
                )
                print(f"DEBUG: Client par défaut créé: {client}")
            else:
                print(f"DEBUG: Client par défaut existant: {client}")
                
            # Générer un nouveau numéro de ticket
            numero_ticket = generate_ticket_number(agence)

            print(f"DEBUG: Numéro de ticket généré: {numero_ticket}")

            
            
            # Calculer les totaux avec conversion sécurisée et recalcul correct
            total_ht = Decimal('0')
            for ligne in facture_temp['lignes']:
                quantite = safe_quantity_conversion(ligne['quantite'])
                prix_unitaire = safe_price_conversion(ligne['prix_unitaire'])
                prix_total_correct = safe_decimal_calculation(quantite, prix_unitaire, 'multiply')
                total_ht = safe_decimal_calculation(total_ht, prix_total_correct, 'add')
            
            remise = safe_price_conversion(facture_temp.get('remise', 0))
            nette_a_payer = safe_decimal_calculation(total_ht, remise, 'subtract')
            montant_regler = safe_price_conversion(facture_temp.get('montant_regler', 0))
            rendu = safe_decimal_calculation(montant_regler, nette_a_payer, 'subtract')

            
            
            print(f"DEBUG: Totaux - Total HT: {total_ht}, Remise: {remise}, Nette: {nette_a_payer}")

            print(f"DEBUG: Montant réglé: {montant_regler}, Rendu: {rendu}")

            
            
            # Récupérer l'employé de l'utilisateur connecté
            try:
                compte = Compte.objects.get(user=request.user, actif=True)
                employe = Employe.objects.filter(compte=compte).first()
                
                # Si pas d'employé, utiliser les infos du compte
                if not employe:
                    employe = type('obj', (object,), {
                        'compte': compte
                    })()
                
                print(f"DEBUG: Employé vendeur: {employe.compte.nom_complet if employe and hasattr(employe, 'compte') else 'Aucun'}")
            except Compte.DoesNotExist:
                employe = None
                print(f"DEBUG: Aucun compte trouvé pour l'utilisateur: {request.user.username}")
            
            

            # Récupérer la session de caisse active
            session_caisse = None
            try:
                aujourd_hui = timezone.now().date()

                session_caisse = SessionCaisse.objects.filter(
                    agence=agence,
                    caisse=caisse,
                    date_ouverture__date=aujourd_hui,
                    statut='ouverte'
                ).first()
                print(f"DEBUG: Session de caisse active: {session_caisse}")
            except Exception as e:
                print(f"DEBUG: Erreur lors de la récupération de la session: {e}")
            
            # Récupérer la date de vente depuis le formulaire ou utiliser la date actuelle
            sale_date = request.POST.get('sale_date')
            if sale_date:
                try:
                    from datetime import datetime
                    aujourd_hui = datetime.strptime(sale_date, '%Y-%m-%d').date()
                    print(f"DEBUG: Date de vente personnalisée: {aujourd_hui}")
                except ValueError:
                    aujourd_hui = timezone.now().date()
                    print(f"DEBUG: Date invalide, utilisation de la date actuelle: {aujourd_hui}")
            else:
                aujourd_hui = timezone.now().date()
                print(f"DEBUG: Aucune date fournie, utilisation de la date actuelle: {aujourd_hui}")

            heure_actuelle = timezone.now().time()

            
            
            print("DEBUG: Création de la facture...")
            
            facture = FactureVente.objects.create(

                numero_ticket=numero_ticket, 

                client=client, 

                agence=agence, 

                caisse=caisse,

                vendeur=employe, 

                session_caisse=session_caisse,
                date=aujourd_hui, 

                heure=heure_actuelle,

                nette_a_payer=nette_a_payer, 

                remise=remise, 

                montant_regler=montant_regler,

                rendu=rendu, 

                en_attente=False,

                nom_vendeuse=employe.compte.nom_complet if employe else 'Vendeur'

            )

            print(f"DEBUG: Facture créée avec ID: {facture.id}")

            
            
            # Créer les lignes de facture

            print(f"DEBUG: Création de {len(facture_temp['lignes'])} lignes de facture...")

            for i, ligne_temp in enumerate(facture_temp['lignes']):

                print(f"DEBUG: Ligne {i+1}: {ligne_temp}")

                
                # Vérifier que l'article_id est valide
                article_id = ligne_temp.get('article_id')
                if not article_id:
                    print(f"DEBUG: ERREUR - article_id manquant pour la ligne {i+1}: {ligne_temp}")
                    continue
                
                try:
                    article = Article.objects.get(id=article_id)
                    print(f"DEBUG: Article trouvé: {article} (ID: {article.id})")
                except Article.DoesNotExist:
                    print(f"DEBUG: ERREUR - Article avec ID {article_id} non trouvé")
                    continue
                
                
                # Recalculer le prix total pour éviter les erreurs de multiplication
                quantite_correcte = safe_quantity_conversion(ligne_temp['quantite'])
                prix_unitaire_correct = safe_price_conversion(ligne_temp['prix_unitaire'])
                prix_total_correct = safe_decimal_calculation(quantite_correcte, prix_unitaire_correct, 'multiply')
                
                ligne_facture = LigneFactureVente.objects.create(
                    facture_vente=facture, 
                    article=article, 
                    designation=ligne_temp['designation'],
                    quantite=quantite_correcte, 
                    prix_unitaire=prix_unitaire_correct,
                    prix_total=prix_total_correct
                )
            
                print(f"DEBUG: Ligne de facture créée: {ligne_facture.id}")
                
                # [HOT] GESTION AUTOMATIQUE DU STOCK - RÉDUCTION LORS DE LA VENTE
                ancien_stock = article.stock_actuel
                quantite_vendue_decimal = safe_quantity_conversion(ligne_temp['quantite'])
                
                # Vérifier si le stock est suffisant
                if article.stock_actuel >= quantite_vendue_decimal:
                    # Réduire le stock
                    article.stock_actuel -= quantite_vendue_decimal
                    article.save()
                    print(f"🛒 STOCK VENTE - Article: {article.designation}")
                    print(f"🛒 STOCK VENTE - Quantité vendue: {quantite_vendue_decimal}")
                    print(f"🛒 STOCK VENTE - Stock mis à jour: {ancien_stock} → {article.stock_actuel}")
                    
                    # Créer un mouvement de stock pour traçabilité
                    try:
                        MouvementStock.objects.create(
                            article=article,
                            agence=agence,
                            type_mouvement='sortie',
                            date_mouvement=timezone.now(),
                            numero_piece=facture.numero_ticket,
                            quantite_stock=article.stock_actuel,
                            stock_initial=ancien_stock,
                            solde=article.stock_actuel,
                            quantite=quantite_vendue_decimal,
                            cout_moyen_pondere=float(article.prix_achat),
                            stock_permanent=float(article.stock_actuel * article.prix_achat),
                            facture_vente=facture,
                            commentaire=f"Vente - Facture {facture.numero_ticket}"
                        )
                        print(f"[NOTE] MOUVEMENT STOCK - Sortie enregistrée pour {article.designation}")
                    except Exception as e:
                        print(f"[WARNING] ERREUR MOUVEMENT STOCK: {e}")
                else:
                    print(f"[WARNING] STOCK INSUFFISANT - Article: {article.designation}")
                    print(f"[WARNING] STOCK INSUFFISANT - Stock disponible: {article.stock_actuel}")
                    print(f"[WARNING] STOCK INSUFFISANT - Quantité demandée: {quantite_vendue_decimal}")
                    # On peut choisir de continuer avec stock négatif ou d'arrêter la vente
                    # Pour l'instant, on continue mais on log l'alerte
                    article.stock_actuel -= quantite_vendue_decimal  # Peut devenir négatif
                    article.save()
                    print(f"[WARNING] STOCK NÉGATIF AUTORISÉ - Stock final: {article.stock_actuel}")
            
            
            
            # Vider la facture temporaire de la session ET de la base de données
            
            # 1. Vider la session
            request.session['facture_temporaire'] = {
                'lignes': [], 'type_vente': 'detail', 'remise': 0, 'montant_regler': 0, 'nette_a_payer': 0, 'rendu': 0
            }
            
            # 2. Supprimer la facture temporaire de la base de données
            session_key = request.session.session_key
            if session_key:
                factures_temp_supprimees = FactureTemporaire.objects.filter(session_key=session_key).delete()
                print(f"DEBUG: {factures_temp_supprimees[0]} facture(s) temporaire(s) supprimée(s) de la base de données")
            
            print("DEBUG: Facture temporaire vidée de la session ET de la base de données")

            
            
            # Retourner une réponse JSON pour AJAX

            print(f"DEBUG: Envoi de la réponse de succès pour la facture {numero_ticket}")

            return JsonResponse({
                'success': True,
                'message': f'Facture {numero_ticket} enregistrée avec succès!',
                'numero_ticket': numero_ticket
            })
            
            
            
        except Exception as e:

            print(f"DEBUG: Erreur lors de l'enregistrement: {str(e)}")

            import traceback

            print(f"DEBUG: Traceback: {traceback.format_exc()}")
            
            # Sauvegarder l'erreur dans un fichier pour debug
            try:
                with open('ERREUR_500_SERVEUR.txt', 'w', encoding='utf-8') as f:
                    f.write(f"ERREUR 500 - ENREGISTREMENT FACTURE\n")
                    f.write(f"=" * 50 + "\n\n")
                    f.write(f"Erreur: {str(e)}\n\n")
                    f.write(f"Traceback:\n{traceback.format_exc()}\n")
                print("📄 Erreur sauvegardée dans ERREUR_500_SERVEUR.txt")
            except:
                pass

            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    except Exception as global_error:
        # Capturer TOUTES les erreurs non gérées
        print(f"❌❌❌ ERREUR GLOBALE 500: {str(global_error)}")
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌❌❌ TRACEBACK:")
        print(error_trace)
        
        # Sauvegarder dans un fichier
        try:
            with open('ERREUR_500_SERVEUR.txt', 'w', encoding='utf-8') as f:
                f.write(f"ERREUR 500 - ENREGISTREMENT FACTURE\n")
                f.write(f"=" * 50 + "\n\n")
                f.write(f"Erreur: {str(global_error)}\n\n")
                f.write(f"Traceback complet:\n{error_trace}\n")
            print("📄 Erreur sauvegardée dans ERREUR_500_SERVEUR.txt")
        except:
            pass
        
        return JsonResponse({
            'success': False,
            'error': f'Erreur serveur 500: {str(global_error)}'
        }, status=500)



@login_required
def rapport_caisse(request):

    """Afficher le rapport de caisse avec les données de la journée"""

    agence = get_user_agence(request)
    if not agence:

        messages.error(request, 'Aucune agence trouvée.')

        return redirect('dashboard_caisse')
    
    
    
    # Récupérer la date d'aujourd'hui

    aujourd_hui = timezone.now().date()

    
    
    # Vérifier le statut de la caisse

    caisse_ouverte = False

    try:

        session_caisse = SessionCaisse.objects.filter(

            agence=agence, 

            date_ouverture__date=aujourd_hui, 

            statut='ouverte'

        ).first()

        caisse_ouverte = session_caisse is not None

    except:

        pass
    
    
    
    # Récupérer toutes les sessions de la journée

    sessions_jour = SessionCaisse.objects.filter(

        agence=agence, 

        date_ouverture__date=aujourd_hui

    ).order_by('date_ouverture').select_related('employe__compte')

    
    
    # Calculer les KPIs de la journée

    chiffre_affaires = 0

    nombre_ventes = 0

    tickets_attente = 0

    

    try:

        # Récupérer les ventes du jour liées aux sessions de caisse
        ventes_jour = FactureVente.objects.filter(
            agence=agence, 
            date=aujourd_hui,
            session_caisse__isnull=False
        ).select_related('session_caisse')
        
        chiffre_affaires = ventes_jour.aggregate(total=Sum('nette_a_payer'))['total'] or 0

        nombre_ventes = ventes_jour.count()

        # Compter les factures temporaires UNIQUEMENT pour la session courante ouverte
        if session_caisse:
            tickets_attente = FactureTemporaire.objects.filter(session_caisse=session_caisse).count()
        else:
            tickets_attente = 0

        
        print(f"DEBUG: Chiffre d'affaires calculé: {chiffre_affaires}")
        print(f"DEBUG: Nombre de ventes: {nombre_ventes}")
        print(f"DEBUG: Tickets en attente: {tickets_attente}")
    except Exception as e:
        print(f"DEBUG: Erreur lors du calcul des KPIs: {e}")
        pass

    

    # Récupérer les détails des ventes pour le rapport

    factures_jour = FactureVente.objects.filter(

        agence=agence, 

        date=aujourd_hui,
        session_caisse__isnull=False
    ).select_related('client', 'vendeur__compte', 'session_caisse').prefetch_related('lignes__article').order_by('-heure')
    
    
    # Calculer les statistiques par heure

    ventes_par_heure = {}

    for facture in factures_jour:

        heure = facture.heure.hour

        if heure not in ventes_par_heure:

            ventes_par_heure[heure] = {'nombre': 0, 'montant': 0}

        ventes_par_heure[heure]['nombre'] += 1

        ventes_par_heure[heure]['montant'] += float(facture.nette_a_payer or 0)
    
    
    
    # Récupérer les articles les plus vendus

    articles_vendus = LigneFactureVente.objects.filter(

        facture_vente__agence=agence,

        facture_vente__date=aujourd_hui

    ).values('article__designation').annotate(

        total_quantite=Sum('quantite'),

        total_montant=Sum('prix_total')

    ).order_by('-total_quantite')[:10]

    
    
    # Récupérer le nom du compte connecté pour l'affichage
    try:
        compte_connecte = Compte.objects.get(user=request.user, actif=True)
        vendeuse_nom = compte_connecte.nom_complet
    except Compte.DoesNotExist:
        vendeuse_nom = "-"
    
    context = {

        'agence': agence,

        'date_aujourd_hui': aujourd_hui,

        'caisse_ouverte': caisse_ouverte,

        'sessions_jour': sessions_jour,

        'chiffre_affaires': chiffre_affaires,

        'nombre_ventes': nombre_ventes,

        'tickets_attente': tickets_attente,

        'factures_jour': factures_jour,

        'ventes_par_heure': ventes_par_heure,
        
        'vendeuse_nom': vendeuse_nom,

        'articles_vendus': articles_vendus,

    }

    
    
    return render(request, 'supermarket/caisse/rapport_caisse.html', context)



@login_required
def detail_factures(request):
    print("DEBUG: Début de detail_factures")
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    print(f"DEBUG: Agence de l'utilisateur: {agence}")
    
    
    
    # Récupérer les paramètres de filtrage

    search_query = request.GET.get('search', '')

    date_filter = request.GET.get('date_filter', 'all')
    

    

    # Récupérer toutes les factures de l'agence liées aux sessions de caisse
    factures = FactureVente.objects.filter(agence=agence).select_related('client', 'vendeur__compte', 'session_caisse').prefetch_related('lignes__article').order_by('-date', '-heure')
    print(f"DEBUG: Nombre de factures trouvées: {factures.count()}")
    
    
    # Appliquer les filtres

    if search_query:

        factures = factures.filter(

            Q(numero_ticket__icontains=search_query) |

            Q(client__intitule__icontains=search_query) |

            Q(nom_vendeuse__icontains=search_query)

        )
    
    
    
    # Filtre par date

    from datetime import datetime, timedelta

    today = datetime.now().date()

    
    
    if date_filter == 'today':

        factures = factures.filter(date=today)

    elif date_filter == 'week':

        week_ago = today - timedelta(days=7)

        factures = factures.filter(date__gte=week_ago)

    elif date_filter == 'month':

        factures = factures.filter(date__year=today.year, date__month=today.month)
    

    

    # Calculer les statistiques (seulement les ventes liées aux sessions de caisse)
    total_factures = FactureVente.objects.filter(agence=agence, session_caisse__isnull=False).count()
    chiffre_affaires_total = FactureVente.objects.filter(agence=agence, session_caisse__isnull=False).aggregate(
        total=Sum('nette_a_payer')

    )['total'] or 0
    
    

    factures_aujourd_hui = FactureVente.objects.filter(agence=agence, date=today, session_caisse__isnull=False).count()
    ca_aujourd_hui = FactureVente.objects.filter(agence=agence, date=today, session_caisse__isnull=False).aggregate(
        total=Sum('nette_a_payer')

    )['total'] or 0

    
    
    # Récupérer l'employé connecté

    employe = None

    if request.user.is_authenticated:

        try:

            employe = Employe.objects.get(compte__user=request.user, compte__agence=agence)

        except Employe.DoesNotExist:

            pass

    

    print(f"DEBUG: Statistiques - Total factures: {total_factures}, CA total: {chiffre_affaires_total}")
    print(f"DEBUG: Aujourd'hui - Factures: {factures_aujourd_hui}, CA: {ca_aujourd_hui}")
    
    context = {

        'factures': factures,

        'total_factures': total_factures,

        'chiffre_affaires_total': chiffre_affaires_total,

        'factures_aujourd_hui': factures_aujourd_hui,

        'ca_aujourd_hui': ca_aujourd_hui,

        'search_query': search_query,

        'date_filter': date_filter,

        'agence': agence,

        'employe': employe,

    }
    
    

    print("DEBUG: Rendu du template detail_factures.html")
    return render(request, 'supermarket/caisse/detail_factures.html', context)



@login_required
def mouvement_vente(request):
    """Afficher les mouvements de vente du jour - même logique que le document HTML"""
    print(f"[ALERTE] MOUVEMENT_VENTE: {request.user.username}")
    
    agence = get_user_agence(request)
    if not agence:
        print(f"[ALERTE] MOUVEMENT_VENTE: Aucune agence pour {request.user.username}")
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    print(f"[ALERTE] MOUVEMENT_VENTE: Agence {agence.nom_agence}")
    
    # Récupérer la date d'aujourd'hui
    aujourd_hui = timezone.now().date()
    
    # Récupérer la session active du jour
    session_caisse = None
    
    # Récupérer le nom du compte connecté
    try:
        compte_connecte = Compte.objects.get(user=request.user, actif=True)
        vendeuse_nom = compte_connecte.nom_complet
    except Compte.DoesNotExist:
        vendeuse_nom = "-"
    
    try:
        # Récupérer seulement la session ACTIVE (ouverte)
        session_caisse = SessionCaisse.objects.filter(
            agence=agence, 
            date_ouverture__date=aujourd_hui,
            statut='ouverte'
        ).first()
    except:
        pass
    
    # Récupérer les factures de la session active
    if session_caisse:
        print(f"[ALERTE] MOUVEMENT_VENTE: Session trouvée {session_caisse.id}")
        factures_jour = FactureVente.objects.filter(
            agence=agence, 
            date=aujourd_hui,
            session_caisse=session_caisse
        ).select_related('client', 'vendeur__compte', 'session_caisse').prefetch_related('lignes__article').order_by('-heure')
        print(f"[ALERTE] MOUVEMENT_VENTE: {factures_jour.count()} factures trouvées")
        
        # Si aucune facture liée à la session, récupérer toutes les factures du jour et les lier
        if factures_jour.count() == 0:
            print(f"[ALERTE] MOUVEMENT_VENTE: Aucune facture liée à la session, recherche des factures orphelines")
            factures_orphelines = FactureVente.objects.filter(
                agence=agence,
                date=aujourd_hui,
                session_caisse__isnull=True
            ).select_related('client', 'vendeur__compte').prefetch_related('lignes__article')
            
            if factures_orphelines.count() > 0:
                print(f"[ALERTE] MOUVEMENT_VENTE: {factures_orphelines.count()} factures orphelines trouvées, liaison à la session")
                # Lier ces factures à la session actuelle
                factures_orphelines.update(session_caisse=session_caisse)
                
                # Récupérer maintenant toutes les factures de la session
                factures_jour = FactureVente.objects.filter(
                    agence=agence, 
                    date=aujourd_hui,
                    session_caisse=session_caisse
                ).select_related('client', 'vendeur__compte', 'session_caisse').prefetch_related('lignes__article').order_by('-heure')
                print(f"[ALERTE] MOUVEMENT_VENTE: {factures_jour.count()} factures après liaison")
    else:
        print(f"[ALERTE] MOUVEMENT_VENTE: Aucune session active")
        # Aucune session active, donc pas de factures à afficher
        factures_jour = FactureVente.objects.none()
    
    # Calculer les statistiques selon la même logique que le HTML
    total_tickets = factures_jour.count()
    total_articles = 0
    chiffre_affaires = 0
    derniere_vente_heure = None
    
    # Préparer les données pour le template (format similaire au localStorage)
    ventes_data = []
    tickets_uniques = set()
    
    for facture in factures_jour:
        tickets_uniques.add(facture.numero_ticket)
        chiffre_affaires += float(facture.nette_a_payer or 0)
        
        # Mettre à jour l'heure de la dernière vente
        if facture.heure:
            derniere_vente_heure = facture.heure.strftime('%H:%M')
        
        # Ajouter les lignes de la facture
        for ligne in facture.lignes.all():
            total_articles += ligne.quantite
            ventes_data.append({
                'ticket': facture.numero_ticket,
                'designation': ligne.designation,
                'reference': ligne.article.reference_article if ligne.article else '',
                'quantite': ligne.quantite,
                'prix_unitaire': float(ligne.prix_unitaire),
                'total': float(ligne.prix_total),
                'heure': facture.heure.strftime('%H:%M') if facture.heure else '-'
            })
    
    context = {
        'agence': agence,
        'date_aujourd_hui': aujourd_hui,
        'session_caisse': session_caisse,
        'vendeuse_nom': vendeuse_nom,
        'total_tickets': len(tickets_uniques),
        'total_articles': total_articles,
        'derniere_vente_heure': derniere_vente_heure or '-',
        'chiffre_affaires': chiffre_affaires,
        'ventes_data': ventes_data,
    }
    
    return render(request, 'supermarket/caisse/mouvement_vente.html', context)

@login_required
def documents_vente(request):
    """Afficher les documents de vente journaliers"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    # Récupérer les paramètres de filtrage
    search_query = request.GET.get('search', '')
    date_filter = request.GET.get('date_filter', 'all')
    
    # Récupérer tous les documents de vente
    documents = DocumentVente.objects.filter(agence=agence).order_by('-date', '-heure_fermeture')
    
    # Appliquer les filtres
    if search_query:
        documents = documents.filter(
            Q(numero_document__icontains=search_query) |
            Q(vendeuse_nom__icontains=search_query)
        )
    
    # Filtre par date
    from datetime import datetime, timedelta
    today = datetime.now().date()
    
    if date_filter == 'today':
        documents = documents.filter(date=today)
    elif date_filter == 'week':
        week_ago = today - timedelta(days=7)
        documents = documents.filter(date__gte=week_ago)
    elif date_filter == 'month':
        documents = documents.filter(date__year=today.year, date__month=today.month)
    elif date_filter == 'year':
        documents = documents.filter(date__year=today.year)
    
    # Calculer les statistiques
    total_documents = DocumentVente.objects.filter(agence=agence).count()
    chiffre_affaires_total = DocumentVente.objects.filter(agence=agence).aggregate(
        total=Sum('chiffre_affaires')
    )['total'] or 0
    total_factures = DocumentVente.objects.filter(agence=agence).aggregate(
        total=Sum('nombre_factures')
    )['total'] or 0
    
    # Dernier archivage
    dernier_archivage = None
    if documents.exists():
        dernier_doc = documents.first()
        dernier_archivage = dernier_doc.heure_fermeture
    
    context = {
        'agence': agence,
        'documents': documents,
        'total_documents': total_documents,
        'chiffre_affaires_total': chiffre_affaires_total,
        'total_factures': total_factures,
        'dernier_archivage': dernier_archivage,
        'search_query': search_query,
        'date_filter': date_filter,
    }
    
    return render(request, 'supermarket/caisse/documents_vente.html', context)

@login_required
def document_vente_details(request, document_id):
    """Afficher les détails d'un document de vente"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    try:
        document = DocumentVente.objects.get(id=document_id, agence=agence)
        
        context = {
            'document': document,
            'factures_data': document.factures_data,
        }
        
        return render(request, 'supermarket/caisse/document_vente_details.html', context)
        
    except DocumentVente.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Document non trouvé.'})

@login_required
def facture_details(request, facture_id):

    print(f"DEBUG: facture_details appelée pour ID: {facture_id}")
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    print(f"DEBUG: Agence de l'utilisateur: {agence}")
    
    
    
    try:

        # Récupérer la facture avec ses lignes

        facture = FactureVente.objects.filter(id=facture_id, agence=agence).select_related('client', 'vendeur__compte').prefetch_related('lignes__article').first()

        print(f"DEBUG: Facture trouvée: {facture}")
        
        
        if not facture:

            print(f"DEBUG: Facture ID {facture_id} non trouvée")
            return JsonResponse({'success': False, 'error': 'Facture non trouvée.'})
        
        
        
        # Préparer les données de la facture

        facture_data = {

            'id': facture.id,

            'numero_ticket': facture.numero_ticket,

            'date': facture.date.strftime('%d/%m/%Y'),

            'heure': facture.heure.strftime('%H:%M'),

            'nette_a_payer': float(facture.nette_a_payer),

            'montant_regler': float(facture.montant_regler),

            'rendu': float(facture.rendu),

            'remise': float(facture.remise),

            'en_attente': facture.en_attente,

            'nom_vendeuse': facture.nom_vendeuse,

            'client': {

                'intitule': facture.client.intitule if facture.client else 'Client anonyme'

            } if facture.client else None,

            'lignes': []

        }

        
        
        # Ajouter les lignes de facture

        for ligne in facture.lignes.all():

            ligne_data = {

                'designation': ligne.designation,

                'reference': ligne.article.reference_article,

                'quantite': ligne.quantite,

                'prix_unitaire': float(ligne.prix_unitaire),

                'prix_total': float(ligne.prix_total),

                'type_vente': 'Détail'  # Type de vente par défaut
            }

            facture_data['lignes'].append(ligne_data)
        
        

        print(f"DEBUG: Données de la facture préparées: {len(facture_data['lignes'])} lignes")
        return JsonResponse({'success': True, 'facture': facture_data})
        
        
        
    except Exception as e:

        print(f"DEBUG: Erreur lors de la récupération des détails de la facture: {str(e)}")

        return JsonResponse({'success': False, 'error': str(e)})



def fermer_caisse(request):


    """Fermer la caisse et créer un document de vente journalier"""
    if request.method == 'GET':
        # Afficher une page de confirmation pour la fermeture
        agence = get_user_agence(request)
        if not agence:
            messages.error(request, 'Votre compte n\'est pas configuré correctement.')
            return redirect('logout_caisse')

        # Vérifier s'il y a une session ouverte
        aujourd_hui = timezone.now().date()
        session_caisse = SessionCaisse.objects.filter(
            agence=agence,
            date_ouverture__date=aujourd_hui,
            statut='ouverte'
        ).first()
        
        if not session_caisse:
            messages.error(request, 'Aucune session de caisse ouverte.')
            return redirect('dashboard_caisse')
        
        # Récupérer le nom du compte connecté
        try:
            compte_connecte = Compte.objects.get(user=request.user, actif=True)
            vendeuse_nom = compte_connecte.nom_complet
        except Compte.DoesNotExist:
            vendeuse_nom = session_caisse.employe.compte.nom_complet if session_caisse.employe else 'Vendeur'
        
        context = {
            'session_caisse': session_caisse,
            'agence': agence,
            'vendeuse_nom': vendeuse_nom,
        }
        return render(request, 'supermarket/caisse/fermer_caisse_confirmation.html', context)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    # Récupérer la session de caisse ouverte
    aujourd_hui = timezone.now().date()
    session_caisse = SessionCaisse.objects.filter(
        agence=agence,
        date_ouverture__date=aujourd_hui,
        statut='ouverte'
    ).first()
    
    if not session_caisse:
        return JsonResponse({'success': False, 'error': 'Aucune session de caisse ouverte trouvée'})
    
    # Récupérer toutes les factures de la session de caisse ouverte
    factures_jour = FactureVente.objects.filter(
        agence=agence,
        date=aujourd_hui,
        session_caisse=session_caisse
    ).select_related('client', 'vendeur__compte').prefetch_related('lignes__article')
    
    # Si aucune facture liée à la session, récupérer toutes les factures du jour et les lier
    if factures_jour.count() == 0:
        # Récupérer toutes les factures du jour non liées à une session
        factures_orphelines = FactureVente.objects.filter(
            agence=agence,
            date=aujourd_hui,
            session_caisse__isnull=True
        ).select_related('client', 'vendeur__compte').prefetch_related('lignes__article')
        
        # Lier ces factures à la session actuelle
        factures_orphelines.update(session_caisse=session_caisse)
        
        # Récupérer maintenant toutes les factures de la session
        factures_jour = FactureVente.objects.filter(
            agence=agence,
            date=aujourd_hui,
            session_caisse=session_caisse
        ).select_related('client', 'vendeur__compte').prefetch_related('lignes__article')
    
    # Calculer les statistiques
    nombre_factures = factures_jour.count()
    total_articles = 0
    chiffre_affaires = 0
    
    # Préparer les données des factures pour le document
    factures_data = []
    
    for facture in factures_jour:
        chiffre_affaires += float(facture.nette_a_payer or 0)
        
        # Préparer les données de cette facture
        facture_data = {
            'numero_ticket': facture.numero_ticket,
            'date': facture.date.strftime('%Y-%m-%d'),
            'heure': facture.heure.strftime('%H:%M') if facture.heure else '',
            'client': facture.client.intitule if facture.client else 'Client anonyme',
            'nette_a_payer': float(facture.nette_a_payer or 0),
            'articles': []
        }
        
        # Ajouter les articles de cette facture
        for ligne in facture.lignes.all():
            total_articles += ligne.quantite
            facture_data['articles'].append({
                'designation': ligne.designation,
                'reference': ligne.article.reference_article if ligne.article else '',
                'quantite': ligne.quantite,
                'prix_unitaire': float(ligne.prix_unitaire),
                'total': float(ligne.prix_total)
            })
        
        factures_data.append(facture_data)
    
    # Générer un numéro de document unique
    numero_document = f"DOC{aujourd_hui.strftime('%Y%m%d')}{session_caisse.id:03d}"
    
    # Récupérer le nom de la vendeuse depuis le compte connecté
    try:
        compte_connecte = Compte.objects.get(user=request.user, actif=True)
        vendeuse_nom = compte_connecte.nom_complet
    except Compte.DoesNotExist:
        # Fallback sur le compte de la session si le compte connecté n'est pas trouvé
        vendeuse_nom = session_caisse.employe.compte.nom_complet if session_caisse.employe else 'Vendeur'
    
    # Créer le document de vente
    document_vente = DocumentVente.objects.create(
        numero_document=numero_document,
        date=aujourd_hui,
        heure_fermeture=timezone.now(),
        session_caisse=session_caisse,
        vendeuse_nom=vendeuse_nom,
        nombre_factures=nombre_factures,
        total_articles=total_articles,
        chiffre_affaires=chiffre_affaires,
        factures_data=factures_data,
        agence=agence
    )
    
    # Fermer la session de caisse
    session_caisse.statut = 'fermee'
    session_caisse.date_fermeture = timezone.now()
    session_caisse.save()
    
    # Fermer la caisse
    caisse = session_caisse.caisse
    caisse.statut = 'fermee'
    caisse.date_fermeture = timezone.now()
    caisse.save()
    
    messages.success(request, f'Caisse fermée avec succès! Document {numero_document} créé.')
    
    # Rediriger vers le dashboard pour voir le changement immédiatement
    return JsonResponse({
        'success': True,
        'message': f'Caisse fermée avec succès! Document {numero_document} créé.',
        'document_id': document_vente.id,
        'redirect_url': '/caisse/'
    })


def finaliser_facture(request, facture_id):

    messages.info(request, 'Fonction en développement')

    return redirect('facturation_vente')



@login_required
def mettre_en_attente(request):


    """Mettre une facture en attente"""
    print("=" * 80)
    print("🔵 DEBUG METTRE_EN_ATTENTE: Début de la fonction")
    print(f"🔵 DEBUG: Méthode de requête: {request.method}")
    print(f"🔵 DEBUG: Données POST: {dict(request.POST)}")
    print(f"🔵 DEBUG: Session key: {request.session.session_key}")
    print(f"🔵 DEBUG: Utilisateur: {request.user.username if request.user.is_authenticated else 'Non authentifié'}")
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        # Récupérer les données de la facture depuis la requête
        facture_data = request.POST.get('facture_data')
        print(f"🔵 DEBUG: facture_data reçu (longueur: {len(facture_data) if facture_data else 0}): {facture_data[:200] if facture_data else 'VIDE'}...")
        
        if not facture_data:
            print("🔵 DEBUG: [ERREUR] Aucune donnée facture_data trouvée dans POST")
            return JsonResponse({'success': False, 'error': 'Aucune donnée de facture fournie'})
        
        import json
        facture_content = json.loads(facture_data)
        print(f"🔵 DEBUG: [OK] facture_content parsé avec succès")
        print(f"🔵 DEBUG: Clés disponibles: {list(facture_content.keys())}")
        
        lignes = facture_content.get('lignes', [])
        print(f"🔵 DEBUG: Nombre de lignes: {len(lignes)}")
        
        for i, ligne in enumerate(lignes):
            print(f"🔵 DEBUG: Ligne {i}: article_id={ligne.get('article_id', 'VIDE')}, designation={ligne.get('designation', 'VIDE')}, quantite={ligne.get('quantite', 'VIDE')}")
        
        # Récupérer la session de caisse active (optionnel)
        agence = get_user_agence(request)
        session_caisse = None
        
        if agence:
            aujourd_hui = timezone.now().date()
            session_caisse = SessionCaisse.objects.filter(
                agence=agence,
                date_ouverture__date=aujourd_hui,
                statut='ouverte'
            ).first()
        
        # Utiliser la session key comme identifiant unique
        session_key = request.session.session_key
        if not session_key:
            request.session.create()
            session_key = request.session.session_key
        
        # Pour la mise en attente, on accepte les lignes même sans article_id
        # car elles peuvent être restaurées lors du rappel
        lignes_a_sauvegarder = []
        for ligne in facture_content.get('lignes', []):
            if ligne.get('article_id'):
                lignes_a_sauvegarder.append(ligne)
                print(f"DEBUG: Ligne avec article_id sauvegardée: {ligne.get('designation')} (ID: {ligne.get('article_id')})")
            else:
                # Inclure quand même la ligne pour la mise en attente
                lignes_a_sauvegarder.append(ligne)
                print(f"DEBUG: Ligne sans article_id incluse pour mise en attente: {ligne.get('designation', 'Sans désignation')}")
        
        # Mettre à jour le contenu avec toutes les lignes
        facture_content['lignes'] = lignes_a_sauvegarder
        print(f"DEBUG: Sauvegarde de {len(lignes_a_sauvegarder)} lignes (avec ou sans article_id)")
        
        # TOUJOURS créer un nouveau ticket en attente (ne pas utiliser get_or_create)
        # Chaque mise en attente = 1 nouveau ticket
        facture_temp = FactureTemporaire.objects.create(
            session_key=session_key,
            contenu=facture_content,
            session_caisse=session_caisse
        )
        print(f"🔵 DEBUG: Nouveau ticket en attente {facture_temp.id} créé pour session {session_caisse.id if session_caisse else 'None'}")
        
        # Mettre à jour la session Django avec les données de la facture
        request.session['facture_temporaire'] = facture_content
        print(f"DEBUG: Session mise à jour avec {len(facture_content.get('lignes', []))} lignes")
        
        # Compter UNIQUEMENT les factures en attente de la session courante
        if session_caisse:
            tickets_en_attente = FactureTemporaire.objects.filter(session_caisse=session_caisse).count()
        else:
            # Si pas de session, compter celles sans session
            tickets_en_attente = FactureTemporaire.objects.filter(session_caisse__isnull=True).count()
        
        return JsonResponse({
            'success': True,
            'message': 'Facture mise en attente avec succès',
            'tickets_en_attente': tickets_en_attente,
            'facture_id': facture_temp.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Format de données invalide'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erreur lors de la mise en attente: {str(e)}'})


@login_required
def rappeler_ticket(request):

    """Rappeler une facture en attente"""
    print("=" * 80)
    print("🟡 DEBUG RAPPELER_TICKET: Début de la fonction")
    print(f"🟡 DEBUG: Méthode: {request.method}")
    print(f"🟡 DEBUG: Utilisateur: {request.user.username if request.user.is_authenticated else 'Non authentifié'}")
    
    if request.method != 'POST':
        print("🟡 DEBUG: [ERREUR] Méthode non autorisée")
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        session_key = request.session.session_key
        print(f"🟡 DEBUG: Session key: {session_key}")
        
        if not session_key:
            print("🟡 DEBUG: [ERREUR] Aucune session active")
            return JsonResponse({'success': False, 'error': 'Aucune session active'})
        
        # Récupérer la facture temporaire la plus récente (toutes sessions confondues)
        try:
            # Rechercher dans toutes les sessions pour éviter les problèmes de session key
            factures_attente = FactureTemporaire.objects.all().order_by('-date_creation')
            print(f"🟡 DEBUG: Nombre total de factures temporaires: {factures_attente.count()}")
            
            for ft in factures_attente:
                print(f"🟡 DEBUG: Facture {ft.id} créée le {ft.date_creation} (session: {ft.session_key})")
            
            facture_temp = factures_attente.first()
            if not facture_temp:
                print("🟡 DEBUG: [ERREUR] Aucune facture en attente trouvée")
                return JsonResponse({'success': False, 'error': 'Aucune facture en attente trouvée'})
            
            print(f"🟡 DEBUG: [OK] Facture temporaire sélectionnée: {facture_temp.id}")
            print(f"🟡 DEBUG: Date de création: {facture_temp.date_creation}")
            print(f"🟡 DEBUG: Session de la facture: {facture_temp.session_key}")
            print(f"🟡 DEBUG: Session actuelle: {session_key}")
            
            contenu = facture_temp.contenu
            if isinstance(contenu, dict):
                print(f"🟡 DEBUG: [OK] Contenu est un dictionnaire")
                print(f"🟡 DEBUG: Clés du contenu: {list(contenu.keys())}")
                lignes = contenu.get('lignes', [])
                print(f"🟡 DEBUG: Nombre de lignes dans le contenu: {len(lignes)}")
                
                for i, ligne in enumerate(lignes):
                    print(f"🟡 DEBUG: Ligne {i}: article_id={ligne.get('article_id', 'VIDE')}, designation={ligne.get('designation', 'VIDE')}")
            else:
                print(f"🟡 DEBUG: [ERREUR] Contenu n'est pas un dictionnaire: {type(contenu)}")
                print(f"🟡 DEBUG: Contenu brut: {str(contenu)[:200]}...")
                
        except Exception as e:
            print(f"🟡 DEBUG: [ERREUR] Erreur lors de la récupération: {str(e)}")
            import traceback
            print(f"🟡 DEBUG: Traceback: {traceback.format_exc()}")
            return JsonResponse({'success': False, 'error': 'Erreur lors de la récupération de la facture'})
        
        # Retourner les données de la facture
        print("DEBUG: Envoi de la réponse de succès")
        import json
        # Convertir le contenu en JSON string si c'est un dictionnaire
        contenu_json = facture_temp.contenu if isinstance(facture_temp.contenu, str) else json.dumps(facture_temp.contenu)
        
        # Sauvegarder les infos avant suppression
        facture_id_supprime = facture_temp.id
        date_creation_str = facture_temp.date_creation.strftime('%d/%m/%Y %H:%M')
        
        # IMPORTANT: Supprimer le ticket après l'avoir rappelé
        facture_temp.delete()
        print(f"🟡 DEBUG: ✅ Ticket {facture_id_supprime} supprimé après rappel")
        
        return JsonResponse({
            'success': True,
            'message': 'Facture rappelée avec succès',
            'facture_data': contenu_json,
            'facture_id': facture_id_supprime,
            'date_creation': date_creation_str
        })
        
    except Exception as e:
        print(f"DEBUG: Erreur lors du rappel: {str(e)}")
        import traceback
        print(f"DEBUG: Traceback: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': f'Erreur lors du rappel: {str(e)}'})


@login_required
def lister_factures_attente(request):
    """Lister toutes les factures en attente pour la session courante"""
    try:
        session_key = request.session.session_key
        if not session_key:
            return JsonResponse({'success': False, 'error': 'Aucune session active'})
        
        # Récupérer toutes les factures en attente pour cette session
        factures_attente = FactureTemporaire.objects.filter(
            session_key=session_key
        ).order_by('-date_creation')
        
        factures_data = []
        for facture in factures_attente:
            import json
            try:
                # facture.contenu est déjà un dictionnaire Python (JSONField)
                contenu = facture.contenu if isinstance(facture.contenu, dict) else json.loads(facture.contenu)
                lignes = contenu.get('lignes', [])
                total = contenu.get('total', 0)
                rendu = contenu.get('rendu', 0)
                client = contenu.get('client', '')
                
                factures_data.append({
                    'id': facture.id,
                    'date_creation': facture.date_creation.strftime('%d/%m/%Y %H:%M'),
                    'nombre_articles': len(lignes),
                    'total': total,
                    'rendu': rendu,
                    'client': client,
                    'contenu': json.dumps(contenu)  # Convertir en JSON string pour le frontend
                })
            except Exception as e:
                print(f"DEBUG: Erreur lors du parsing de la facture {facture.id}: {e}")
                factures_data.append({
                    'id': facture.id,
                    'date_creation': facture.date_creation.strftime('%d/%m/%Y %H:%M'),
                    'nombre_articles': 0,
                    'total': 0,
                    'rendu': 0,
                    'client': '',
                    'contenu': '{}'
                })
        
        return JsonResponse({
            'success': True,
            'factures': factures_data,
            'nombre_factures': len(factures_data)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erreur lors du listage: {str(e)}'})


@login_required
def rappeler_facture_specifique(request):
    """Rappeler une facture spécifique par son ID"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        facture_id = request.POST.get('facture_id')
        if not facture_id:
            return JsonResponse({'success': False, 'error': 'ID de facture manquant'})
        
        session_key = request.session.session_key
        if not session_key:
            return JsonResponse({'success': False, 'error': 'Aucune session active'})
        
        # Récupérer la facture spécifique
        try:
            facture_temp = FactureTemporaire.objects.get(
                id=facture_id,
                session_key=session_key
            )
        except FactureTemporaire.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Facture non trouvée'})
        
        print(f"DEBUG: Facture spécifique trouvée: {facture_temp.id}")
        print(f"DEBUG: Contenu: {facture_temp.contenu}")
        
        import json
        # Convertir le contenu en JSON string si c'est un dictionnaire
        contenu_json = facture_temp.contenu if isinstance(facture_temp.contenu, str) else json.dumps(facture_temp.contenu)
        
        return JsonResponse({
            'success': True,
            'message': 'Facture rappelée avec succès',
            'facture_data': contenu_json,
            'facture_id': facture_temp.id,
            'date_creation': facture_temp.date_creation.strftime('%d/%m/%Y %H:%M')
        })
        
    except Exception as e:
        print(f"DEBUG: Erreur lors du rappel spécifique: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erreur lors du rappel: {str(e)}'})


def test_urls_page(request):
    """Page de test pour vérifier les URLs des tickets en attente"""
    return render(request, 'supermarket/caisse/test_urls.html')


@login_required
def supprimer_facture_attente(request):
    """Supprimer une facture en attente par son ID"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        facture_id = request.POST.get('facture_id')
        if not facture_id:
            return JsonResponse({'success': False, 'error': 'ID de facture manquant'})
        
        session_key = request.session.session_key
        if not session_key:
            return JsonResponse({'success': False, 'error': 'Aucune session active'})
        
        # Récupérer et supprimer la facture spécifique
        try:
            facture_temp = FactureTemporaire.objects.get(
                id=facture_id,
                session_key=session_key
            )
            
            # Supprimer la facture
            facture_temp.delete()
            
            print(f"DEBUG: Facture temporaire {facture_id} supprimée avec succès")
            
            # Compter le nombre restant de factures en attente
            tickets_restants = FactureTemporaire.objects.filter(session_key=session_key).count()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket en attente supprimé avec succès',
                'tickets_restants': tickets_restants
            })
            
        except FactureTemporaire.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Facture non trouvée'})
        
    except Exception as e:
        print(f"DEBUG: Erreur lors de la suppression: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erreur lors de la suppression: {str(e)}'})


def supprimer_ligne_facture(request):

    messages.info(request, 'Fonction en développement')

    return redirect('facturation_vente')



def modifier_quantite_ligne(request):

    messages.info(request, 'Fonction en développement')

    return redirect('facturation_vente')



def supprimer_vente(request):

    messages.info(request, 'Fonction en développement')

    return redirect('facturation_vente')



@login_required
def imprimer_facture(request, facture_id=None):
    """Afficher et imprimer une facture depuis le formulaire ou depuis la base"""
    print(f"[ALERTE] IMPRIMER_FACTURE: {request.user.username}")
    print(f"[ALERTE] IMPRIMER_FACTURE: Méthode: {request.method}")
    print(f"[ALERTE] IMPRIMER_FACTURE: GET params: {dict(request.GET)}")
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('facturation_vente')
    
    # Vérifier si des données sont passées en paramètres GET (depuis le JavaScript)
    if request.GET.get('numero_ticket'):
        print(f"[ALERTE] IMPRIMER_FACTURE: Données reçues depuis JavaScript")
        try:
            import json
            
            # Récupérer les paramètres depuis l'URL
            numero_ticket = request.GET.get('numero_ticket', '')
            date = request.GET.get('date', '')
            heure = request.GET.get('heure', '')
            caisse = request.GET.get('caisse', '')
            vendeur = request.GET.get('vendeur', '')
            client = request.GET.get('client', '')
            nette_a_payer = float(request.GET.get('nette_a_payer', 0))
            montant_regler = float(request.GET.get('montant_regler', 0))
            rendu = float(request.GET.get('rendu', 0))
            remise = float(request.GET.get('remise', 0))
            
            # Récupérer les articles
            articles_json = request.GET.get('articles', '[]')
            articles = json.loads(articles_json)
            
            print(f"[ALERTE] IMPRIMER_FACTURE: {len(articles)} articles reçus")
            
            # Créer les données de la facture avec structure optimisée
            # S'assurer que chaque article a les bonnes clés
            lignes_formatees = []
            for article in articles:
                ligne = {
                    'designation': article.get('designation', article.get('nom_article', 'Article')),
                    'quantite': article.get('quantite', 1),
                    'prix_unitaire': article.get('prix_unitaire', 0),
                    'prix_total': article.get('prix_total', 0),
                    'reference': article.get('reference', ''),
                    'article_id': article.get('article_id', '')
                }
                lignes_formatees.append(ligne)
            
            facture_data = {
                'numero_ticket': numero_ticket,
                'date': date,
                'heure': heure,
                'nette_a_payer': nette_a_payer,
                'montant_regler': montant_regler,
                'rendu': rendu,
                'remise': remise,
                'nom_vendeuse': vendeur or 'Vendeur',
                'client_nom': client or 'Client anonyme',
                'caisse_numero': caisse or 'CAISSE001',
                'lignes': lignes_formatees
            }
            
            print(f"[ALERTE] IMPRIMER_FACTURE: Données préparées: {facture_data}")
            
            # Rendre le template d'impression professionnel
            return render(request, 'supermarket/caisse/facture_impression.html', {
                'facture': facture_data,
                'agence': agence
            })
            
        except Exception as e:
            print(f"[ALERTE] IMPRIMER_FACTURE: Erreur traitement GET: {e}")
            messages.error(request, f'Erreur lors de la préparation de l\'impression: {e}')
            return redirect('facturation_vente')
    
    # Si pas de paramètres GET, essayer de récupérer depuis la session
    print(f"[ALERTE] IMPRIMER_FACTURE: Pas de paramètres GET, récupération depuis la session")
    
    try:
        # Récupérer les données de la facture temporaire depuis la session
        facture_temp = request.session.get('facture_temporaire', {})
        
        if facture_temp and facture_temp.get('lignes'):
            print(f"[ALERTE] IMPRIMER_FACTURE: Données de session trouvées")
            
            # Calculer les totaux
            total_ht = sum(ligne.get('prix_total', 0) for ligne in facture_temp.get('lignes', []))
            remise = facture_temp.get('remise', 0)
            nette_a_payer = total_ht - remise
            montant_regler = facture_temp.get('montant_regler', 0)
            rendu = montant_regler - nette_a_payer
            
            # Générer un numéro de ticket
            from datetime import datetime
            now = datetime.now()
            numero_ticket = f'TICKET_{now.strftime("%Y%m%d_%H%M%S")}'
            
            # Formater les lignes de la session
            lignes_session = facture_temp.get('lignes', [])
            lignes_formatees = []
            for ligne in lignes_session:
                ligne_formatee = {
                    'designation': ligne.get('designation', ligne.get('nom_article', 'Article')),
                    'quantite': ligne.get('quantite', 1),
                    'prix_unitaire': ligne.get('prix_unitaire', 0),
                    'prix_total': ligne.get('prix_total', 0),
                    'reference': ligne.get('reference', ''),
                    'article_id': ligne.get('article_id', '')
                }
                lignes_formatees.append(ligne_formatee)
            
            facture_data = {
                'numero_ticket': numero_ticket,
                'date': now.strftime('%d/%m/%Y'),
                'heure': now.strftime('%H:%M'),
                'nette_a_payer': nette_a_payer,
                'montant_regler': montant_regler,
                'rendu': rendu,
                'remise': remise,
                'nom_vendeuse': 'Vendeur',
                'client_nom': facture_temp.get('client', 'Client anonyme'),
                'caisse_numero': 'CAISSE001',
                'lignes': lignes_formatees
            }
            
            print(f"[ALERTE] IMPRIMER_FACTURE: Données de session préparées: {facture_data}")
            
            # Rendre le template d'impression professionnel
            return render(request, 'supermarket/caisse/facture_impression.html', {
                'facture': facture_data,
                'agence': agence
            })
        else:
            print(f"[ALERTE] IMPRIMER_FACTURE: Aucune donnée de session, création de données de test")
            
            # Créer des données de test pour l'impression
            from datetime import datetime
            now = datetime.now()
            
            facture_data = {
                'numero_ticket': f'TICKET_{now.strftime("%Y%m%d_%H%M%S")}',
                'date': now.strftime('%d/%m/%Y'),
                'heure': now.strftime('%H:%M'),
                'nette_a_payer': 0,
                'montant_regler': 0,
                'rendu': 0,
                'remise': 0,
                'nom_vendeuse': 'Vendeur',
                'client_nom': 'Client anonyme',
                'caisse_numero': 'CAISSE001',
                'lignes': []
            }
            
            print(f"[ALERTE] IMPRIMER_FACTURE: Données de test créées: {facture_data}")
            
            # Rendre le template d'impression avec les données de test
            return render(request, 'supermarket/caisse/facture_impression.html', {
                'facture': facture_data,
                'agence': agence
            })
            
    except Exception as e:
        print(f"[ALERTE] IMPRIMER_FACTURE: Erreur générale: {e}")
        messages.error(request, f'Erreur lors de l\'impression: {e}')
        return redirect('facturation_vente')



@login_required
def init_test_data(request):

    """Initialiser des données de test pour le développement"""

    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    
    
    if not agence:

        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    
    
    try:

        # Créer quelques articles de test

        articles_data = [

            {'reference': 'ART001', 'designation': 'Coca Cola 33cl', 'prix_vente': 500},

            {'reference': 'ART002', 'designation': 'Pain de mie', 'prix_vente': 1000},

            {'reference': 'ART003', 'designation': 'Savon de Marseille', 'prix_vente': 400},

            {'reference': 'ART004', 'designation': 'Riz 5kg', 'prix_vente': 3000},

            {'reference': 'ART005', 'designation': 'Eau minérale 1.5L', 'prix_vente': 300},

        ]

        
        
        for article_data in articles_data:

            article, created = Article.objects.get_or_create(

                reference_article=article_data['reference'],

                defaults={

                    'designation': article_data['designation'],

                    'prix_vente': article_data['prix_vente'],

                    'stock_disponible': 100,

                    'agence': agence

                }

            )

            if created:

                print(f"Article créé: {article.designation}")
        
        
        
        # Créer des factures de test pour aujourd'hui

        aujourd_hui = timezone.now().date()

        
        
        # Créer une session de caisse de test

        session_caisse, created = SessionCaisse.objects.get_or_create(

            agence=agence,

            date_ouverture__date=aujourd_hui,

            defaults={

                'date_ouverture': timezone.now(),

                'statut': 'ouverte',

                'montant_initial': 10000

            }

        )

        
        
        if created:

            print(f"Session de caisse créée: {session_caisse}")
        
        
        
        # Créer quelques factures de test

        factures_test = [

            {'numero_ticket': 'TKT20250923001', 'nette_a_payer': 1500, 'en_attente': False},

            {'numero_ticket': 'TKT20250923002', 'nette_a_payer': 2300, 'en_attente': False},

            {'numero_ticket': 'TKT20250923003', 'nette_a_payer': 800, 'en_attente': True},

        ]

        
        
        for facture_data in factures_test:

            facture, created = FactureVente.objects.get_or_create(

                numero_ticket=facture_data['numero_ticket'],

                defaults={

                    'agence': agence,

                    'date': aujourd_hui,

                    'heure': timezone.now().time(),

                    'nette_a_payer': facture_data['nette_a_payer'],

                    'montant_regler': facture_data['nette_a_payer'],

                    'rendu': 0,

                    'remise': 0,

                    'en_attente': facture_data['en_attente'],

                    'nom_vendeuse': 'Test Vendeuse'

                }

            )

            if created:

                print(f"Facture créée: {facture.numero_ticket}")
        
        
        
        return JsonResponse({'success': True, 'message': 'Données de test initialisées avec factures'})
        
        
        
    except Exception as e:

        return JsonResponse({'success': False, 'error': str(e)})



@login_required
def update_quantity_temp(request):

    return JsonResponse({'success': True})



@login_required
def remove_article_temp(request):
    """Supprimer un article de la facture temporaire dans la session"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        index = int(request.POST.get('index', -1))
        if index < 0:
            return JsonResponse({'success': False, 'error': 'Index invalide'})
        
        # Récupérer la facture temporaire de la session
        facture_temp = request.session.get('facture_temporaire', {'lignes': []})
        lignes = facture_temp.get('lignes', [])
        
        # Supprimer l'article à l'index spécifié
        if 0 <= index < len(lignes):
            lignes.pop(index)
            facture_temp['lignes'] = lignes
            
            # Recalculer les totaux
            total_ht = sum(ligne.get('prix_total', 0) for ligne in lignes)
            facture_temp['nette_a_payer'] = total_ht - facture_temp.get('remise', 0)
            
            # Mettre à jour la session
            request.session['facture_temporaire'] = facture_temp
            request.session.modified = True
            
            print(f"DEBUG: Article à l'index {index} supprimé. {len(lignes)} articles restants.")
            return JsonResponse({'success': True, 'message': 'Article supprimé avec succès'})
        else:
            return JsonResponse({'success': False, 'error': 'Index hors limites'})
            
    except Exception as e:
        print(f"DEBUG: Erreur lors de la suppression de l'article: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})



@login_required
def update_montant_regler(request):

    return JsonResponse({'success': True})



@login_required
def clear_facture_temp(request):

    request.session['facture_temporaire'] = {

        'lignes': [], 'type_vente': 'detail', 'remise': 0, 'montant_regler': 0, 'nette_a_payer': 0, 'rendu': 0

    }

    return JsonResponse({'success': True})



@login_required
def update_type_vente_temp(request):

    if request.method == 'POST':

        try:

            index = int(request.POST.get('index'))

            type_vente = request.POST.get('type_vente')

            
            
            facture_temp = request.session.get('facture_temporaire', {'lignes': []})

            
            
            if 0 <= index < len(facture_temp['lignes']):

                ligne = facture_temp['lignes'][index]

                ligne['type_vente'] = type_vente

                
                
                # Récupérer le nouveau prix selon le type de vente

                try:
                    agence = get_user_agence(request)
                    if not agence:
                        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
                    
                    article = Article.objects.get(id=ligne['article_id'], agence=agence)

                    type_vente_obj = TypeVente.objects.filter(article=article, intitule__iexact=type_vente).first()

                    
                    
                    if type_vente_obj:

                        nouveau_prix = float(type_vente_obj.prix)

                    else:

                        nouveau_prix = float(article.prix_vente or 0)
                    
                    
                    
                    ligne['prix_unitaire'] = nouveau_prix

                    ligne['prix_total'] = nouveau_prix * ligne['quantite']

                    
                    
                    request.session['facture_temporaire'] = facture_temp

                    
                    
                    return JsonResponse({'success': True})

                except Exception as e:

                    return JsonResponse({'success': False, 'error': str(e)})

            else:

                return JsonResponse({'success': False, 'error': 'Index invalide'})
                
                
                
        except Exception as e:

            return JsonResponse({'success': False, 'error': str(e)})
    
    
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})



@login_required
def generate_ticket_number_api(request):

    """API pour générer un nouveau numéro de ticket"""

    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    
    
    if not agence:

        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    
    
    numero_ticket = generate_ticket_number(agence)

    
    
    return JsonResponse({

        'success': True,

        'numero_ticket': numero_ticket

    })



@login_required
def get_article_types(request):

    """Récupérer les types de vente disponibles pour un article"""

    article_id = request.GET.get('article_id')

    
    
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    try:
        article = Article.objects.get(id=article_id, agence=agence)

        
        
        # Récupérer tous les types de vente pour cet article

        types_vente = TypeVente.objects.filter(article=article).order_by('intitule')

        
        
        # Ajouter aussi le prix de vente par défaut comme option "Détail"

        types_data = []

        
        
        # Prix par défaut de l'article comme "Détail"

        types_data.append({

            'intitule': 'Détail',

            'prix': float(article.prix_vente or 0)

        })

        
        
        # Types de vente spécifiques de la base de données

        for type_vente in types_vente:

            types_data.append({

                'intitule': type_vente.intitule,

                'prix': float(type_vente.prix)

            })
        
        
        
        return JsonResponse({

            'success': True,

            'types': types_data,

            'article': {

                'id': article.id,

                'designation': article.designation,


                'prix_vente': float(article.prix_vente or 0)


            }


        })
        
        
        
    except Exception as e:

        return JsonResponse({

            'success': False,

            'error': str(e),

            'types': []

        })



@login_required
def update_all_types_vente_temp(request):

    """Mettre à jour le type de vente global pour tous les articles de la facture temporaire"""

    if request.method == 'POST':

        try:

            type_vente_global = request.POST.get('type_vente_global')


            
            
            if not type_vente_global:


                return JsonResponse({'success': False, 'error': 'Type de vente non spécifié'})
            
            
            
            facture_temp = request.session.get('facture_temporaire', {'lignes': []})


            
            
            if not facture_temp.get('lignes'):

                return JsonResponse({'success': True, 'message': 'Aucun article dans la facture'})
            
            
            
            agence = get_user_agence(request)
            if not agence:
                return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
            
            try:

                
                
                # Mettre à jour chaque ligne de la facture temporaire

                for ligne in facture_temp['lignes']:

                    ligne['type_vente'] = type_vente_global

                    
                    
                    # Récupérer le nouveau prix selon le type de vente

                    article = Article.objects.get(id=ligne['article_id'], agence=agence)

                    type_vente_obj = TypeVente.objects.filter(article=article, intitule__iexact=type_vente_global).first()

                    
                    
                    if type_vente_obj:

                        nouveau_prix = float(type_vente_obj.prix)

                    else:

                        nouveau_prix = float(article.prix_vente or 0)
                    
                    
                    
                    ligne['prix_unitaire'] = nouveau_prix

                    ligne['prix_total'] = nouveau_prix * ligne['quantite']
                
                
                
                # Mettre à jour le type de vente global dans la facture temporaire

                facture_temp['type_vente'] = type_vente_global


                
                
                request.session['facture_temporaire'] = facture_temp

                
                
                return JsonResponse({
                    'success': True,
                    'message': f'Type de vente mis à jour vers {type_vente_global} pour tous les articles',
                    'lignes_updated': len(facture_temp['lignes'])
                })
                

                
                
            except Exception as e:

                return JsonResponse({'success': False, 'error': str(e)})
                

                
                
        except Exception as e:

            return JsonResponse({'success': False, 'error': str(e)})
    
    
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})



@login_required
def get_article_types_vente(request, article_id):
    """Récupérer les types de vente d'un article"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    try:
        article = Article.objects.get(id=article_id, agence=agence)
        types_vente = TypeVente.objects.filter(article=article).order_by('intitule')
        
        types_data = []
        for type_vente in types_vente:
            types_data.append({
                'intitule': type_vente.intitule,
                'prix': float(type_vente.prix)
            })
        
        return JsonResponse({
            'success': True,
            'types_vente': types_data
        })
        
    except Article.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Article non trouvé'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def update_type_vente(request):
    """Mettre à jour le type de vente d'un article"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    try:
        article_id = request.POST.get('article_id')
        type_vente = request.POST.get('type_vente')
        
        if not article_id or not type_vente:
            return JsonResponse({'success': False, 'error': 'Données manquantes'})
        
        article = Article.objects.get(id=article_id, agence=agence)
        type_vente_obj = TypeVente.objects.filter(
            article=article, 
            intitule__iexact=type_vente
        ).first()
        
        if type_vente_obj:
            nouveau_prix = float(type_vente_obj.prix)
        else:
            nouveau_prix = float(article.prix_vente or 0)
        
        return JsonResponse({
            'success': True,
            'nouveau_prix': nouveau_prix,
            'type_vente': type_vente
        })
        
    except Article.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Article non trouvé'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_article_id_by_designation(request):
    """Récupérer l'ID d'un article par sa désignation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    try:
        designation = request.POST.get('designation', '').strip()
        reference = request.POST.get('reference', '').strip()
        
        if not designation:
            return JsonResponse({'success': False, 'error': 'Désignation manquante'})
        
        # Chercher l'article par désignation exacte d'abord
        article = Article.objects.filter(
            agence=agence,
            designation__iexact=designation
        ).first()
        
        # Si pas trouvé par désignation exacte, chercher par référence
        if not article and reference:
            article = Article.objects.filter(
                agence=agence,
                reference_article__iexact=reference
            ).first()
        
        # Si toujours pas trouvé, chercher par désignation contenant
        if not article:
            article = Article.objects.filter(
                agence=agence,
                designation__icontains=designation
            ).first()
        
        if article:
            return JsonResponse({
                'success': True,
                'article_id': article.id,
                'designation': article.designation,
                'reference': article.reference_article,
                'prix_vente': float(article.prix_vente or 0)
            })
        else:
            return JsonResponse({
                'success': False,
                'error': f'Article non trouvé pour: "{designation}"'
            })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})




@login_required
def liste_articles(request):
    """Afficher la liste des articles avec recherche et statistiques"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    categorie_filter = request.GET.get('categorie', '')
    
    # Récupérer tous les articles de l'agence
    articles = Article.objects.filter(agence=agence).select_related('categorie').order_by('designation')
    
    # Appliquer les filtres
    if search_query:
        articles = articles.filter(
            Q(designation__icontains=search_query) |
            Q(reference_article__icontains=search_query)
        )
    
    if categorie_filter:
        articles = articles.filter(categorie__id=categorie_filter)
    
    # Récupérer les catégories pour le filtre
    categories = Article.objects.filter(agence=agence).values_list('categorie', flat=True).distinct()
    categories_list = [categorie for categorie in categories if categorie]
    
    # Calculer les statistiques
    total_articles = articles.count()
    stock_total = articles.aggregate(total=Sum('stock_actuel'))['total'] or 0
    valeur_stock = articles.aggregate(total=Sum(models.F('stock_actuel') * models.F('prix_vente')))['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'categories': categories_list,
        'search_query': search_query,
        'categorie_filter': categorie_filter,
        'total_articles': total_articles,
        'stock_total': stock_total,
        'valeur_stock': valeur_stock,
    }
    
    return render(request, 'supermarket/caisse/liste_articles.html', context)

@login_required
def liste_clients(request):
    """Afficher la liste des clients avec statistiques"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    
    # Récupérer tous les clients de l'agence
    clients = Client.objects.filter(agence=agence).order_by('intitule')
    
    # Appliquer le filtre de recherche
    if search_query:
        clients = clients.filter(
            Q(intitule__icontains=search_query) |
            Q(telephone__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Calculer les statistiques pour chaque client
    clients_avec_stats = []
    for client in clients:
        # Statistiques des ventes
        factures_client = FactureVente.objects.filter(client=client, agence=agence)
        nombre_achats = factures_client.count()
        montant_total = factures_client.aggregate(total=Sum('nette_a_payer'))['total'] or 0
        
        # Articles les plus achetés
        articles_achetes = LigneFactureVente.objects.filter(
            facture_vente__client=client,
            facture_vente__agence=agence
        ).values('article__designation').annotate(
            total_quantite=Sum('quantite'),
            total_montant=Sum('prix_total')
        ).order_by('-total_quantite')[:3]
        
        clients_avec_stats.append({
            'client': client,
            'nombre_achats': nombre_achats,
            'montant_total': montant_total,
            'articles_achetes': articles_achetes,
        })
    
    # Statistiques générales
    total_clients = clients.count()
    clients_actifs = len([c for c in clients_avec_stats if c['nombre_achats'] > 0])
    montant_total_ventes = sum([c['montant_total'] for c in clients_avec_stats])
    
    # Statistiques pour les cartes du template
    clients_avec_email = clients.filter(email__isnull=False).exclude(email='').count()
    clients_sans_email = clients.filter(Q(email__isnull=True) | Q(email='')).count()
    clients_avec_achats = clients_actifs
    
    context = {
        'agence': agence,
        'clients': clients,  # Passer directement les clients pour le template
        'clients_avec_stats': clients_avec_stats,
        'search_query': search_query,
        'total_clients': total_clients,
        'clients_actifs': clients_actifs,
        'montant_total_ventes': montant_total_ventes,
        'clients_avec_email': clients_avec_email,
        'clients_sans_email': clients_sans_email,
        'clients_avec_achats': clients_avec_achats,
    }
    
    return render(request, 'supermarket/caisse/liste_clients.html', context)

@login_required
def test_flux_complet(request):
    """Vue de test pour le flux complet"""
    return render(request, 'test_flux_complet.html')

@login_required
def creer_article(request):
    """Vue pour créer un nouvel article"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            designation = request.POST.get('designation')
            prix_achat = request.POST.get('prix_achat')
            dernier_prix_achat = request.POST.get('dernier_prix_achat', prix_achat)  # Par défaut, égal au prix d'achat
            prix_vente = request.POST.get('prix_vente')
            stock_actuel = request.POST.get('stock_actuel', 0)
            stock_minimum = request.POST.get('stock_minimum', 0)
            unite_vente = request.POST.get('unite_vente')
            conditionnement = request.POST.get('conditionnement')
            agence_id = request.POST.get('agence')
            famille_id = request.POST.get('famille')
            
            # Debug des champs reçus
            print(f"[SEARCH] DEBUG: Champs reçus:")
            print(f"[SEARCH] DEBUG: - designation: '{designation}'")
            print(f"[SEARCH] DEBUG: - prix_achat: '{prix_achat}'")
            print(f"[SEARCH] DEBUG: - prix_vente: '{prix_vente}'")
            print(f"[SEARCH] DEBUG: - agence_id: '{agence_id}'")
            print(f"[SEARCH] DEBUG: - famille_id: '{famille_id}'")
            print(f"[SEARCH] DEBUG: - unite_vente: '{unite_vente}'")
            print(f"[SEARCH] DEBUG: - conditionnement: '{conditionnement}'")
            
            # Validation des champs obligatoires
            if not designation or not prix_achat or not prix_vente or not agence_id or not famille_id:
                print(f"[SEARCH] DEBUG: [ERREUR] Validation échouée - Champs manquants:")
                if not designation: print("  - designation manquant")
                if not prix_achat: print("  - prix_achat manquant")
                if not prix_vente: print("  - prix_vente manquant")
                if not agence_id: print("  - agence_id manquant")
                if not famille_id: print("  - famille_id manquant")
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_article')
            
            # Récupérer l'agence et la famille
            try:
                agence = Agence.objects.get(id_agence=agence_id)
                famille = Famille.objects.get(id=famille_id)
            except (Agence.DoesNotExist, Famille.DoesNotExist):
                messages.error(request, 'Agence ou famille invalide.')
                return redirect('creer_article')
            
            # Générer automatiquement la référence article
            print(f"[SEARCH] DEBUG: Génération automatique de la référence...")
            
            # Récupérer la dernière référence
            last_article = Article.objects.order_by('-id').first()
            
            if last_article and last_article.reference_article:
                # Extraire le numéro de la dernière référence
                import re
                match = re.search(r'ART(\d+)', last_article.reference_article)
                if match:
                    last_number = int(match.group(1))
                    new_number = last_number + 1
                else:
                    new_number = 1
            else:
                new_number = 1
            
            # Générer la nouvelle référence
            reference_article = f"ART{new_number:03d}"
            
            # Vérifier que la référence n'existe pas déjà
            while Article.objects.filter(reference_article=reference_article).exists():
                new_number += 1
                reference_article = f"ART{new_number:03d}"
            
            print(f"[SEARCH] DEBUG: [OK] Référence générée: {reference_article}")
            
            # Créer l'article
            article = Article.objects.create(
                reference_article=reference_article,
                designation=designation,
                prix_achat=float(prix_achat),
                dernier_prix_achat=float(dernier_prix_achat) if dernier_prix_achat else float(prix_achat),
                prix_vente=float(prix_vente),
                stock_actuel=float(stock_actuel),
                stock_minimum=float(stock_minimum),
                unite_vente=unite_vente,
                conditionnement=conditionnement,
                agence=agence,
                categorie=famille
            )
            
            # Créer les types de vente avec leurs prix spécifiques
            # Récupérer les prix depuis le formulaire
            prix_gros = request.POST.get('prix_gros', '').strip()
            prix_demi_gros = request.POST.get('prix_demi_gros', '').strip()
            prix_detail = request.POST.get('prix_detail', '').strip()
            
            print(f"[DEBUG] Prix récupérés - Gros: '{prix_gros}', Demi-Gros: '{prix_demi_gros}', Détail: '{prix_detail}'")
            
            # Créer les types de vente avec les bons prix
            types_vente_data = [
                {'intitule': 'Gros', 'prix': float(prix_gros) if prix_gros else float(prix_vente)},
                {'intitule': 'Demi-Gros', 'prix': float(prix_demi_gros) if prix_demi_gros else float(prix_vente)},
                {'intitule': 'Détail', 'prix': float(prix_detail) if prix_detail else float(prix_vente)},
            ]
            
            for type_vente_info in types_vente_data:
                TypeVente.objects.create(
                    article=article,
                    intitule=type_vente_info['intitule'],
                    prix=type_vente_info['prix']
                )
                print(f"[DEBUG] Type de vente créé: {type_vente_info['intitule']} à {type_vente_info['prix']} FCFA")
            
            # Créer un mouvement de stock initial si stock > 0
            if float(stock_actuel) > 0:
                MouvementStock.objects.create(
                    article=article,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece=f'INIT-{article.id}',
                    quantite_stock=float(stock_actuel),
                    stock_initial=0,
                    solde=float(stock_actuel),
                    quantite=float(stock_actuel),
                    cout_moyen_pondere=float(prix_achat),
                    stock_permanent=float(stock_actuel) * float(prix_achat),
                    commentaire=f"Création article - Stock initial"
                )
            
            messages.success(request, f'Article "{designation}" créé avec succès!')
            return redirect('dashboard_stock')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de l\'article: {str(e)}')
            return redirect('creer_article')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    familles = Famille.objects.all()
    
    # Debug des familles disponibles
    print(f"[SEARCH] DEBUG: Familles disponibles: {familles.count()}")
    for famille in familles:
        print(f"[SEARCH] DEBUG: - Famille: {famille.intitule} (ID: {famille.id})")
    
    context = {
        'agences': agences,
        'familles': familles,
    }
    return render(request, 'supermarket/stock/creer_article.html', context)

@login_required
def generate_reference(request):
    """Vue pour générer une référence article automatique"""
    try:
        # Récupérer la dernière référence
        last_article = Article.objects.order_by('-id').first()
        
        if last_article and last_article.reference_article:
            # Extraire le numéro de la dernière référence
            import re
            match = re.search(r'ART(\d+)', last_article.reference_article)
            if match:
                last_number = int(match.group(1))
                new_number = last_number + 1
            else:
                new_number = 1
        else:
            new_number = 1
        
        # Générer la nouvelle référence
        new_reference = f"ART{new_number:03d}"
        
        # Vérifier que la référence n'existe pas déjà
        while Article.objects.filter(reference_article=new_reference).exists():
            new_number += 1
            new_reference = f"ART{new_number:03d}"
        
        return JsonResponse({
            'success': True,
            'reference': new_reference
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def consulter_articles(request):
    """Vue pour consulter la liste des articles"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    famille_id = request.GET.get('famille', '')
    stock_filter = request.GET.get('stock_filter', '')
    force_refresh = request.GET.get('refresh', '')
    
    # Forcer le rafraîchissement si demandé
    if force_refresh == '1':
        # Ne pas utiliser session.flush() car cela supprime l'authentification
        # Le simple rechargement de la page suffira
        pass
    
    # Construire la requête de base avec select_related pour éviter les requêtes multiples
    articles = Article.objects.filter(agence=agence).select_related('categorie')
    
    # Appliquer les filtres
    if search_query:
        articles = articles.filter(
            Q(designation__icontains=search_query) |
            Q(categorie__intitule__icontains=search_query)
        )
    
    if famille_id:
        articles = articles.filter(categorie_id=famille_id)
    
    if stock_filter == 'low':
        articles = articles.filter(stock_actuel__lte=F('stock_minimum'))
    elif stock_filter == 'zero':
        articles = articles.filter(stock_actuel__lte=0)
    
    # Trier par ID décroissant
    articles = articles.order_by('-id')
    
    # Calculer les statistiques
    total_articles = Article.objects.filter(agence=agence).count()
    articles_stock_faible = Article.objects.filter(
        agence=agence,
        stock_actuel__lte=F('stock_minimum')
    ).count()
    articles_rupture = Article.objects.filter(
        agence=agence,
        stock_actuel__lte=0
    ).count()
    
    # Valeur totale du stock
    articles_avec_prix = Article.objects.filter(
        agence=agence,
        prix_achat__isnull=False,
        stock_actuel__isnull=False
    ).exclude(prix_achat=0).exclude(stock_actuel=0)
    
    valeur_totale_stock = 0
    for article in articles_avec_prix:
        try:
            valeur_article = float(article.prix_achat) * float(article.stock_actuel)
            valeur_totale_stock += valeur_article
        except (ValueError, TypeError):
            continue
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Calculer les statistiques pour le template
    articles_stock_normal = total_articles - articles_stock_faible
    articles_stock_low = articles_stock_faible - articles_rupture
    articles_stock_zero = articles_rupture
    
    context = {
        'articles': articles,
        'agence': agence,
        'total_articles': total_articles,
        'articles_stock_normal': articles_stock_normal,
        'articles_stock_low': articles_stock_low,
        'articles_stock_zero': articles_stock_zero,
        'articles_stock_faible': articles_stock_faible,
        'articles_rupture': articles_rupture,
        'valeur_totale_stock': valeur_totale_stock,
        'familles': familles,
    }
    return render(request, 'supermarket/stock/consulter_articles.html', context)

@login_required
def consulter_fournisseurs(request):
    """Vue pour consulter la liste des fournisseurs"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    fournisseurs = Fournisseur.objects.filter(agence=agence).order_by('-id')
    context = {
        'fournisseurs': fournisseurs,
        'agence': agence,
    }
    return render(request, 'supermarket/stock/consulter_fournisseurs.html', context)

@login_required
def consulter_clients(request):
    """Vue pour consulter la liste des clients"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    agence_filter = request.GET.get('agence_filter', '')
    
    # Construire la requête de base - Limiter aux clients de l'agence de l'utilisateur
    clients = Client.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        clients = clients.filter(
            Q(intitule__icontains=search_query) |
            Q(telephone__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if agence_filter:
        clients = clients.filter(agence_id=agence_filter)
    
    # Trier par ID décroissant
    clients = clients.order_by('-id')
    
    # Calculer les statistiques
    total_clients = Client.objects.filter(agence=agence).count()
    clients_avec_telephone = Client.objects.filter(agence=agence).exclude(telephone='').count()
    clients_avec_email = Client.objects.filter(agence=agence).exclude(email='').count()
    clients_recents = Client.objects.filter(agence=agence).count()  # Peut être amélioré avec une date
    
    # Récupérer toutes les agences pour le filtre
    agences = Agence.objects.all()
    
    context = {
        'clients': clients,
        'agence': agence,
        'total_clients': total_clients,
        'clients_avec_telephone': clients_avec_telephone,
        'clients_avec_email': clients_avec_email,
        'clients_recents': clients_recents,
        'agences': agences,
    }
    return render(request, 'supermarket/stock/consulter_clients.html', context)

@login_required
def creer_client(request):
    """Vue pour créer un nouveau client"""
    if request.method == 'POST':
        try:
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email', '')
            agence_id = request.POST.get('agence')
            
            # Validation des champs obligatoires
            if not all([intitule, adresse, telephone, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_client')
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence sélectionnée invalide.')
                return redirect('creer_client')
            
            # Créer le client
            client = Client.objects.create(
                intitule=intitule,
                adresse=adresse,
                telephone=telephone,
                email=email,
                agence=agence
            )
            
            messages.success(request, f'Client "{client.intitule}" créé avec succès!')
            return redirect('dashboard_stock')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du client: {str(e)}')
            return redirect('creer_client')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    context = {
        'agences': agences,
    }
    return render(request, 'supermarket/stock/creer_client.html', context)

@login_required
def creer_fournisseur(request):
    """Vue pour créer un nouveau fournisseur"""
    if request.method == 'POST':
        try:
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email', '')
            agence_id = request.POST.get('agence')
            
            # Validation des champs obligatoires
            if not all([intitule, adresse, telephone, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_fournisseur')
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence sélectionnée invalide.')
                return redirect('creer_fournisseur')
            
            # Créer le fournisseur
            fournisseur = Fournisseur.objects.create(
                intitule=intitule,
                adresse=adresse,
                telephone=telephone,
                email=email,
                agence=agence
            )
            
            messages.success(request, f'Fournisseur "{fournisseur.intitule}" créé avec succès!')
            return redirect('dashboard_stock')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du fournisseur: {str(e)}')
            return redirect('creer_fournisseur')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    context = {
        'agences': agences,
    }
    return render(request, 'supermarket/stock/creer_fournisseur.html', context)

@login_required
def detail_fournisseur(request, fournisseur_id):
    """Vue pour afficher les détails d'un fournisseur"""
    try:
        fournisseur = Fournisseur.objects.get(id=fournisseur_id, agence=get_user_agence(request))
        
        # Récupérer les factures d'achat du fournisseur (si elles existent)
        factures_achat = FactureAchat.objects.filter(fournisseur=fournisseur).order_by('-date_achat', '-heure')[:10]
        
        context = {
            'fournisseur': fournisseur,
            'factures_achat': factures_achat,
        }
        return render(request, 'supermarket/stock/detail_fournisseur.html', context)
    except Fournisseur.DoesNotExist:
        messages.error(request, 'Fournisseur non trouvé.')
        return redirect('consulter_fournisseurs')

@login_required
def modifier_fournisseur(request, fournisseur_id):
    """Vue pour modifier un fournisseur existant"""
    try:
        fournisseur = Fournisseur.objects.get(id=fournisseur_id, agence=get_user_agence(request))
    except Fournisseur.DoesNotExist:
        messages.error(request, 'Fournisseur non trouvé.')
        return redirect('consulter_fournisseurs')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([intitule, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_fournisseur', fournisseur_id=fournisseur_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_fournisseur', fournisseur_id=fournisseur_id)
            
            # Mettre à jour le fournisseur
            fournisseur.intitule = intitule
            fournisseur.adresse = adresse
            fournisseur.telephone = telephone
            fournisseur.email = email
            fournisseur.agence = agence
            fournisseur.save()
            
            messages.success(request, f'Fournisseur "{intitule}" modifié avec succès!')
            return redirect('consulter_fournisseurs')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du fournisseur: {str(e)}')
            return redirect('modifier_fournisseur', fournisseur_id=fournisseur_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    
    context = {
        'fournisseur': fournisseur,
        'agences': agences,
    }
    return render(request, 'supermarket/stock/modifier_fournisseur.html', context)

@login_required
def supprimer_fournisseur(request, fournisseur_id):
    """Vue pour supprimer un fournisseur"""
    if request.method == 'POST':
        try:
            fournisseur = Fournisseur.objects.get(id=fournisseur_id, agence=get_user_agence(request))
            fournisseur_name = fournisseur.intitule
            fournisseur.delete()
            messages.success(request, f'Fournisseur "{fournisseur_name}" supprimé avec succès!')
        except Fournisseur.DoesNotExist:
            messages.error(request, 'Fournisseur non trouvé.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression du fournisseur: {str(e)}')
    
    return redirect('consulter_fournisseurs')

@login_required
def dashboard_stock(request):
    """Dashboard principal du module de gestion de stock"""
    try:
        # Récupérer l'agence de l'utilisateur
        agence = get_user_agence(request)
        if not agence:
            messages.error(request, 'Votre compte n\'est pas configuré correctement.')
            return redirect('logout_stock')

        # Calculer les KPIs
        total_articles = Article.objects.filter(agence=agence).count()
        articles_stock_faible = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=10
        ).count()
        articles_rupture = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=0
        ).count()
        
        # Valeur totale du stock
        articles_avec_prix = Article.objects.filter(
            agence=agence,
            prix_achat__isnull=False,
            stock_actuel__isnull=False
        ).exclude(prix_achat=0).exclude(stock_actuel=0)
        
        valeur_totale_stock = 0
        for article in articles_avec_prix:
            try:
                valeur_article = float(article.prix_achat) * float(article.stock_actuel)
                valeur_totale_stock += valeur_article
            except (ValueError, TypeError):
                continue
        
        # Mouvements récents
        mouvements_recents = MouvementStock.objects.filter(agence=agence).order_by('-date_mouvement')[:5]
        
        # Articles les plus vendus (simulation)
        articles_populaires = Article.objects.filter(agence=agence).order_by('-stock_actuel')[:5]
        
        # Alertes de stock
        alertes_stock = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=5
        ).order_by('stock_actuel')[:5]

        # Récupérer le nom de l'utilisateur
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except Compte.DoesNotExist:
            nom_utilisateur = request.user.username

        context = {
            'agence': agence,
            'nom_utilisateur': nom_utilisateur,
            'total_articles': total_articles,
            'articles_stock_faible': articles_stock_faible,
            'articles_rupture': articles_rupture,
            'valeur_stock': valeur_totale_stock,  # Corrigé le nom de la variable
            'mouvements_recents': mouvements_recents,
            'articles_populaires': articles_populaires,
            'alertes_stock': alertes_stock,
        }
        
        return render(request, 'supermarket/stock/dashboard_stock.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur lors du chargement du dashboard: {str(e)}')
        # Récupérer le nom de l'utilisateur même en cas d'erreur
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except:
            nom_utilisateur = request.user.username if request.user.is_authenticated else "Utilisateur"

        return render(request, 'supermarket/stock/dashboard_stock.html', {
            'agence': None,
            'nom_utilisateur': nom_utilisateur,
            'total_articles': 0,
            'articles_stock_faible': 0,
            'articles_rupture': 0,
            'valeur_stock': 0,
            'mouvements_recents': [],
            'articles_populaires': [],
            'alertes_stock': [],
        })

def login_stock(request):
    """Page de connexion pour la gestion de stock"""
    if request.user.is_authenticated:
        return redirect('dashboard_stock')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                # Vérifier que l'utilisateur a un compte lié à une agence
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        login(request, user)
                        # Stocker l'agence dans la session
                        request.session['agence_id'] = compte.agence.id_agence
                        request.session['agence_nom'] = compte.agence.nom_agence
                        messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                        return redirect('dashboard_stock')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Aucun compte actif trouvé pour cet utilisateur.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')
    
    return render(request, 'supermarket/stock/login.html')

@login_required
def logout_stock(request):
    """Vue de logout pour le module stock"""
    logout(request)
    return redirect('login_stock')

@login_required
def modifier_client(request, client_id):
    """Vue pour modifier un client existant"""
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([intitule, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_client', client_id=client_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_client', client_id=client_id)
            
            # Mettre à jour le client
            client.intitule = intitule
            client.adresse = adresse
            client.telephone = telephone
            client.email = email
            client.agence = agence
            client.save()
            
            messages.success(request, f'Client "{intitule}" modifié avec succès!')
            return redirect('consulter_clients')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du client: {str(e)}')
            return redirect('modifier_client', client_id=client_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    
    context = {
        'client': client,
        'agences': agences,
    }
    return render(request, 'supermarket/stock/modifier_client.html', context)

@login_required
def supprimer_client(request, client_id):
    """Vue pour supprimer un client"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_clients')
    
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
        client_name = client.intitule
        client.delete()
        messages.success(request, f'Client "{client_name}" supprimé avec succès!')
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_clients')

@login_required
def detail_client(request, client_id):
    """Vue pour afficher les détails d'un client"""
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
        
        # Récupérer les factures du client (si elles existent)
        factures = FactureVente.objects.filter(client=client).order_by('-date', '-heure')[:10]
        
        context = {
            'client': client,
            'factures': factures,
        }
        return render(request, 'supermarket/stock/detail_client.html', context)
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients')

# ==================== PLAN COMPTABLE ====================

@login_required
def consulter_plan_comptable(request):
    """Vue pour consulter le plan comptable"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    nature_filter = request.GET.get('nature_filter', '')
    
    # Construire la requête de base
    comptes = PlanComptable.objects.all()
    
    # Appliquer les filtres
    if search_query:
        comptes = comptes.filter(
            Q(intitule__icontains=search_query) |
            Q(compte__icontains=search_query) |
            Q(abrege__icontains=search_query)
        )
    
    if nature_filter:
        comptes = comptes.filter(nature_compte=nature_filter)
    
    # Trier par numéro
    comptes = comptes.order_by('numero')
    
    # Calculer les statistiques
    total_comptes = PlanComptable.objects.count()
    comptes_actifs = PlanComptable.objects.filter(actif=True).count()
    
    # Récupérer les natures de compte pour le filtre
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'comptes': comptes,
        'agence': agence,
        'total_comptes': total_comptes,
        'comptes_actifs': comptes_actifs,
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/consulter_plan_comptable.html', context)

@login_required
def creer_plan_comptable(request):
    """Vue pour créer un nouveau compte comptable"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero = request.POST.get('numero')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            abrege = request.POST.get('abrege')
            nature_compte = request.POST.get('nature_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([numero, intitule, compte]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_plan_comptable')
            
            # Créer le compte comptable
            PlanComptable.objects.create(
                numero=numero,
                intitule=intitule,
                compte=compte,
                abrege=abrege,
                nature_compte=nature_compte
            )
            
            messages.success(request, f'Compte comptable "{intitule}" créé avec succès!')
            return redirect('consulter_plan_comptable')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du compte: {str(e)}')
            return redirect('creer_plan_comptable')
    
    # GET - Afficher le formulaire
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/creer_plan_comptable.html', context)

@login_required
def modifier_plan_comptable(request, compte_id):
    """Vue pour modifier un compte comptable existant"""
    try:
        compte = PlanComptable.objects.get(id=compte_id)
    except PlanComptable.DoesNotExist:
        messages.error(request, 'Compte comptable non trouvé.')
        return redirect('consulter_plan_comptable')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero = request.POST.get('numero')
            intitule = request.POST.get('intitule')
            compte_field = request.POST.get('compte')
            abrege = request.POST.get('abrege')
            nature_compte = request.POST.get('nature_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([numero, intitule, compte_field]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_plan_comptable', compte_id=compte_id)
            
            # Mettre à jour le compte
            compte.numero = numero
            compte.intitule = intitule
            compte.compte = compte_field
            compte.abrege = abrege
            compte.nature_compte = nature_compte
            compte.save()
            
            messages.success(request, f'Compte comptable "{intitule}" modifié avec succès!')
            return redirect('consulter_plan_comptable')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du compte: {str(e)}')
            return redirect('modifier_plan_comptable', compte_id=compte_id)
    
    # GET - Afficher le formulaire pré-rempli
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'compte': compte,
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/modifier_plan_comptable.html', context)

@login_required
def supprimer_plan_comptable(request, compte_id):
    """Vue pour supprimer un compte comptable"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_plan_comptable')
    
    try:
        compte = PlanComptable.objects.get(id=compte_id)
        compte_name = compte.intitule
        compte.delete()
        messages.success(request, f'Compte comptable "{compte_name}" supprimé avec succès!')
    except PlanComptable.DoesNotExist:
        messages.error(request, 'Compte comptable non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_plan_comptable')

# ==================== PLAN TIERS ====================

@login_required
def consulter_plan_tiers(request):
    """Vue pour consulter le plan tiers"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    tiers = PlanTiers.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        tiers = tiers.filter(
            Q(intitule_compte__icontains=search_query) |
            Q(numero_compte__icontains=search_query)
        )
    
    if type_filter:
        tiers = tiers.filter(type=type_filter)
    
    # Trier par numéro de compte
    tiers = tiers.order_by('numero_compte')
    
    # Calculer les statistiques
    total_tiers = PlanTiers.objects.filter(agence=agence).count()
    clients_count = PlanTiers.objects.filter(agence=agence, type='client').count()
    fournisseurs_count = PlanTiers.objects.filter(agence=agence, type='fournisseur').count()
    
    # Récupérer les types pour le filtre
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'tiers': tiers,
        'agence': agence,
        'total_tiers': total_tiers,
        'clients_count': clients_count,
        'fournisseurs_count': fournisseurs_count,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_plan_tiers.html', context)

@login_required
def creer_plan_tiers(request):
    """Vue pour créer un nouveau tiers"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_tiers = request.POST.get('type')
            numero_compte = request.POST.get('numero_compte')
            intitule_compte = request.POST.get('intitule_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_tiers, numero_compte, intitule_compte, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_plan_tiers')
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('creer_plan_tiers')
            
            # Créer le tiers
            PlanTiers.objects.create(
                type=type_tiers,
                numero_compte=numero_compte,
                intitule_compte=intitule_compte,
                agence=agence
            )
            
            messages.success(request, f'Tiers "{intitule_compte}" créé avec succès!')
            return redirect('consulter_plan_tiers')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du tiers: {str(e)}')
            return redirect('creer_plan_tiers')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'agences': agences,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/creer_plan_tiers.html', context)

@login_required
def modifier_plan_tiers(request, tiers_id):
    """Vue pour modifier un tiers existant"""
    try:
        tiers = PlanTiers.objects.get(id=tiers_id, agence=get_user_agence(request))
    except PlanTiers.DoesNotExist:
        messages.error(request, 'Tiers non trouvé.')
        return redirect('consulter_plan_tiers')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_tiers = request.POST.get('type')
            numero_compte = request.POST.get('numero_compte')
            intitule_compte = request.POST.get('intitule_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_tiers, numero_compte, intitule_compte, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_plan_tiers', tiers_id=tiers_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_plan_tiers', tiers_id=tiers_id)
            
            # Mettre à jour le tiers
            tiers.type = type_tiers
            tiers.numero_compte = numero_compte
            tiers.intitule_compte = intitule_compte
            tiers.agence = agence
            tiers.save()
            
            messages.success(request, f'Tiers "{intitule_compte}" modifié avec succès!')
            return redirect('consulter_plan_tiers')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du tiers: {str(e)}')
            return redirect('modifier_plan_tiers', tiers_id=tiers_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'tiers': tiers,
        'agences': agences,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/modifier_plan_tiers.html', context)

@login_required
def supprimer_plan_tiers(request, tiers_id):
    """Vue pour supprimer un tiers"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_plan_tiers')
    
    try:
        tiers = PlanTiers.objects.get(id=tiers_id, agence=get_user_agence(request))
        tiers_name = tiers.intitule_compte
        tiers.delete()
        messages.success(request, f'Tiers "{tiers_name}" supprimé avec succès!')
    except PlanTiers.DoesNotExist:
        messages.error(request, 'Tiers non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_plan_tiers')

# ==================== CODE JOURNAUX ====================

@login_required
def consulter_code_journaux(request):
    """Vue pour consulter les codes journaux"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    journaux = CodeJournaux.objects.all()
    
    # Appliquer les filtres
    if search_query:
        journaux = journaux.filter(
            Q(intitule__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    if type_filter:
        journaux = journaux.filter(type_document=type_filter)
    
    # Trier par code
    journaux = journaux.order_by('code')
    
    # Calculer les statistiques
    total_journaux = CodeJournaux.objects.count()
    journaux_achat = CodeJournaux.objects.filter(type_document='document_achat').count()
    journaux_vente = CodeJournaux.objects.filter(type_document='caisse').count()
    
    # Récupérer les types pour le filtre
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'journaux': journaux,
        'agence': agence,
        'total_journaux': total_journaux,
        'journaux_achat': journaux_achat,
        'journaux_vente': journaux_vente,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_code_journaux.html', context)

@login_required
def creer_code_journaux(request):
    """Vue pour créer un nouveau code journal"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_document = request.POST.get('type_document')
            intitule = request.POST.get('intitule')
            code = request.POST.get('code')
            compte_contrepartie = request.POST.get('compte_contrepartie')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_document, intitule, code]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_code_journaux')
            
            # Créer le code journal
            CodeJournaux.objects.create(
                type_document=type_document,
                intitule=intitule,
                code=code,
                compte_contrepartie=compte_contrepartie
            )
            
            messages.success(request, f'Code journal "{intitule}" créé avec succès!')
            return redirect('consulter_code_journaux')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du code journal: {str(e)}')
            return redirect('creer_code_journaux')
    
    # GET - Afficher le formulaire
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/creer_code_journaux.html', context)

@login_required
def modifier_code_journaux(request, journal_id):
    """Vue pour modifier un code journal existant"""
    try:
        journal = CodeJournaux.objects.get(id=journal_id)
    except CodeJournaux.DoesNotExist:
        messages.error(request, 'Code journal non trouvé.')
        return redirect('consulter_code_journaux')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_document = request.POST.get('type_document')
            intitule = request.POST.get('intitule')
            code = request.POST.get('code')
            compte_contrepartie = request.POST.get('compte_contrepartie')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_document, intitule, code]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_code_journaux', journal_id=journal_id)
            
            # Mettre à jour le code journal
            journal.type_document = type_document
            journal.intitule = intitule
            journal.code = code
            journal.compte_contrepartie = compte_contrepartie
            journal.save()
            
            messages.success(request, f'Code journal "{intitule}" modifié avec succès!')
            return redirect('consulter_code_journaux')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du code journal: {str(e)}')
            return redirect('modifier_code_journaux', journal_id=journal_id)
    
    # GET - Afficher le formulaire pré-rempli
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'journal': journal,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/modifier_code_journaux.html', context)

@login_required
def supprimer_code_journaux(request, journal_id):
    """Vue pour supprimer un code journal"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_code_journaux')
    
    try:
        journal = CodeJournaux.objects.get(id=journal_id)
        journal_name = journal.intitule
        journal.delete()
        messages.success(request, f'Code journal "{journal_name}" supprimé avec succès!')
    except CodeJournaux.DoesNotExist:
        messages.error(request, 'Code journal non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_code_journaux')

# ==================== TAUX TAXE ====================

@login_required
def consulter_taux_taxe(request):
    """Vue pour consulter les taux de taxe"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    taux = TauxTaxe.objects.all()
    
    # Appliquer les filtres
    if search_query:
        taux = taux.filter(
            Q(intitule__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    if type_filter:
        taux = taux.filter(type=type_filter)
    
    # Trier par code
    taux = taux.order_by('code')
    
    # Calculer les statistiques
    total_taux = TauxTaxe.objects.count()
    taux_actifs = TauxTaxe.objects.filter(actif=True).count()
    
    # Récupérer les types pour le filtre
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    
    context = {
        'taux': taux,
        'agence': agence,
        'total_taux': total_taux,
        'taux_actifs': taux_actifs,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_taux_taxe.html', context)

@login_required
def creer_taux_taxe(request):
    """Vue pour créer un nouveau taux de taxe"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            code = request.POST.get('code')
            sens = request.POST.get('sens')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            taux = request.POST.get('taux')
            type_taxe = request.POST.get('type')
            assujettissement = request.POST.get('assujettissement')
            code_regroupement = request.POST.get('code_regroupement')
            provenance = request.POST.get('provenance')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([code, intitule, taux]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_taux_taxe')
            
            # Créer le taux de taxe
            TauxTaxe.objects.create(
                code=code,
                sens=sens,
                intitule=intitule,
                compte=compte,
                taux=float(taux),
                type=type_taxe,
                assujettissement=assujettissement,
                code_regroupement=code_regroupement,
                provenance=provenance
            )
            
            messages.success(request, f'Taux de taxe "{intitule}" créé avec succès!')
            return redirect('consulter_taux_taxe')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du taux de taxe: {str(e)}')
            return redirect('creer_taux_taxe')
    
    # GET - Afficher le formulaire
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    sens_choices = TauxTaxe.SENS_CHOICES
    assujettissement_choices = TauxTaxe.ASSUJETTISSEMENT_CHOICES
    
    context = {
        'type_choices': type_choices,
        'sens_choices': sens_choices,
        'assujettissement_choices': assujettissement_choices,
    }
    return render(request, 'supermarket/stock/creer_taux_taxe.html', context)

@login_required
def modifier_taux_taxe(request, taux_id):
    """Vue pour modifier un taux de taxe existant"""
    try:
        taux = TauxTaxe.objects.get(id=taux_id)
    except TauxTaxe.DoesNotExist:
        messages.error(request, 'Taux de taxe non trouvé.')
        return redirect('consulter_taux_taxe')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            code = request.POST.get('code')
            sens = request.POST.get('sens')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            taux_value = request.POST.get('taux')
            type_taxe = request.POST.get('type')
            assujettissement = request.POST.get('assujettissement')
            code_regroupement = request.POST.get('code_regroupement')
            provenance = request.POST.get('provenance')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([code, intitule, taux_value]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_taux_taxe', taux_id=taux_id)
            
            # Mettre à jour le taux de taxe
            taux.code = code
            taux.sens = sens
            taux.intitule = intitule
            taux.compte = compte
            taux.taux = float(taux_value)
            taux.type = type_taxe
            taux.assujettissement = assujettissement
            taux.code_regroupement = code_regroupement
            taux.provenance = provenance
            taux.save()
            
            messages.success(request, f'Taux de taxe "{intitule}" modifié avec succès!')
            return redirect('consulter_taux_taxe')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du taux de taxe: {str(e)}')
            return redirect('modifier_taux_taxe', taux_id=taux_id)
    
    # GET - Afficher le formulaire pré-rempli
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    sens_choices = TauxTaxe.SENS_CHOICES
    assujettissement_choices = TauxTaxe.ASSUJETTISSEMENT_CHOICES
    
    context = {
        'taux': taux,
        'type_choices': type_choices,
        'sens_choices': sens_choices,
        'assujettissement_choices': assujettissement_choices,
    }
    return render(request, 'supermarket/stock/modifier_taux_taxe.html', context)

@login_required
def supprimer_taux_taxe(request, taux_id):
    """Vue pour supprimer un taux de taxe"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_taux_taxe')
    
    try:
        taux = TauxTaxe.objects.get(id=taux_id)
        taux_name = taux.intitule
        taux.delete()
        messages.success(request, f'Taux de taxe "{taux_name}" supprimé avec succès!')
    except TauxTaxe.DoesNotExist:
        messages.error(request, 'Taux de taxe non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_taux_taxe')

@login_required
def modifier_article(request, article_id):
    """Vue pour modifier un article existant"""
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
        return redirect('consulter_articles')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            designation = request.POST.get('designation')
            agence_id = request.POST.get('agence')
            prix_achat = request.POST.get('prix_achat')
            prix_vente = request.POST.get('prix_vente')
            stock_actuel = request.POST.get('stock_actuel')
            stock_minimum = request.POST.get('stock_minimum', 0)
            unite_vente = request.POST.get('unite_vente')
            conditionnement = request.POST.get('conditionnement')
            famille_id = request.POST.get('famille')
            
            # Validation
            if not all([designation, agence_id, prix_achat, prix_vente, stock_actuel, unite_vente]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_article', article_id=article_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_article', article_id=article_id)
            
            # Récupérer la famille si spécifiée
            if famille_id:
                try:
                    categorie = Famille.objects.get(id=famille_id)
                except Famille.DoesNotExist:
                    messages.error(request, 'Famille non trouvée.')
                    return redirect('modifier_article', article_id=article_id)
            else:
                # Si aucune famille n'est fournie, garder l'ancienne
                categorie = article.categorie
            
            # Mettre à jour l'article
            article.designation = designation
            article.agence = agence
            article.prix_achat = float(prix_achat)
            article.prix_vente = float(prix_vente)
            article.stock_actuel = float(stock_actuel)
            article.stock_minimum = float(stock_minimum) if stock_minimum else 0
            article.unite_vente = unite_vente
            article.conditionnement = conditionnement
            article.categorie = categorie
            article.save()
            
            # Mettre à jour les types de vente
            prix_gros = request.POST.get('prix_gros')
            prix_demi_gros = request.POST.get('prix_demi_gros')
            prix_detail = request.POST.get('prix_detail')
            
            if prix_gros:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Gros',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_gros)}
                )
            
            if prix_demi_gros:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Demi-Gros',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_demi_gros)}
                )
            
            if prix_detail:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Détail',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_detail)}
                )
            
            messages.success(request, f'Article "{designation}" modifié avec succès!')
            return redirect('consulter_articles')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de l\'article: {str(e)}')
            return redirect('modifier_article', article_id=article_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    familles = Famille.objects.all()
    types_vente = TypeVente.objects.filter(article=article)
    
    # Créer un dictionnaire des types de vente avec des clés sans traits d'union
    types_vente_dict = {}
    for tv in types_vente:
        if tv.intitule == 'Demi-Gros':
            types_vente_dict['Demi_Gros'] = tv.prix
        elif tv.intitule == 'Détail':
            types_vente_dict['Détail'] = tv.prix
        else:
            types_vente_dict[tv.intitule] = tv.prix
    
    context = {
        'article': article,
        'agences': agences,
        'familles': familles,
        'types_vente': types_vente_dict
    }
    return render(request, 'supermarket/stock/modifier_article.html', context)

@login_required
def supprimer_article(request, article_id):
    """Vue pour supprimer un article"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_articles')
    
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
        article_name = article.designation
        article.delete()
        messages.success(request, f'Article "{article_name}" supprimé avec succès!')
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_articles')

@login_required
def detail_article(request, article_id):
    """Vue pour afficher les détails d'un article"""
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
        types_vente = TypeVente.objects.filter(article=article)
        mouvements = MouvementStock.objects.filter(article=article).order_by('-date_mouvement')[:10]
        
        # Calculer les marges
        marge_unitaire = float(article.prix_vente) - float(article.prix_achat) if article.prix_achat > 0 else 0
        marge_pourcentage = (marge_unitaire / float(article.prix_achat) * 100) if article.prix_achat > 0 else 0
        valeur_stock = float(article.prix_achat) * float(article.stock_actuel)
        
        # Calculer les marges pour chaque type de vente
        types_vente_with_marges = []
        for tv in types_vente:
            marge_tv = float(tv.prix) - float(article.prix_achat) if article.prix_achat > 0 else 0
            marge_tv_pourcentage = (marge_tv / float(article.prix_achat) * 100) if article.prix_achat > 0 else 0
            types_vente_with_marges.append({
                'type_vente': tv,
                'marge': marge_tv,
                'marge_pourcentage': marge_tv_pourcentage
            })
        
        # Debug: Vérifier la famille de l'article
        print(f"[ALERTE] DEBUG Article {article.id}:")
        print(f"   - Désignation: {article.designation}")
        print(f"   - Catégorie: {article.categorie}")
        print(f"   - Intitulé famille: {article.categorie.intitule if article.categorie else 'None'}")
        
        context = {
            'article': article,
            'types_vente': types_vente,
            'types_vente_with_marges': types_vente_with_marges,
            'mouvements': mouvements,
            'marge_unitaire': marge_unitaire,
            'marge_pourcentage': marge_pourcentage,
            'valeur_stock': valeur_stock
        }
        return render(request, 'supermarket/stock/detail_article.html', context)
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
        return redirect('consulter_articles')

# ==================== FACTURES D'ACHAT ====================

@login_required
def consulter_factures_achat(request):
    """Vue pour consulter les factures d'achat"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    statut_filter = request.GET.get('statut_filter', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Construire la requête de base
    factures = FactureAchat.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        factures = factures.filter(
            Q(numero_fournisseur__icontains=search_query) |
            Q(reference_achat__icontains=search_query) |
            Q(commentaire__icontains=search_query)
        )
    
    if statut_filter:
        factures = factures.filter(statut=statut_filter)
    
    if date_debut:
        factures = factures.filter(date_achat__gte=date_debut)
    
    if date_fin:
        factures = factures.filter(date_achat__lte=date_fin)
    
    # Trier par date décroissante
    factures = factures.order_by('-date_achat', '-heure')
    
    # Calculer les statistiques
    total_factures = FactureAchat.objects.filter(agence=agence).count()
    factures_validees = FactureAchat.objects.filter(agence=agence, statut='validee').count()
    factures_payees = FactureAchat.objects.filter(agence=agence, statut='payee').count()
    montant_total = FactureAchat.objects.filter(agence=agence).aggregate(
        total=Sum('prix_total_global')
    )['total'] or 0
    
    # Récupérer les statuts pour le filtre
    statut_choices = FactureAchat.STATUT_CHOICES
    
    context = {
        'factures': factures,
        'search_query': search_query,
        'statut_filter': statut_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statut_choices': statut_choices,
        'total_factures': total_factures,
        'factures_validees': factures_validees,
        'factures_payees': factures_payees,
        'montant_total': montant_total,
    }
    return render(request, 'supermarket/stock/consulter_factures_achat.html', context)

@login_required
def creer_facture_achat(request):
    """Vue pour créer une nouvelle facture d'achat"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    if request.method == 'POST':
        try:
            print("[START] DÉBUT CRÉATION FACTURE D'ACHAT")
            print(f"[LIST] Données POST reçues: {dict(request.POST)}")
            
            # Récupérer les données du formulaire
            numero_fournisseur = request.POST.get('numero_fournisseur')
            date_achat = request.POST.get('date_achat')
            heure = request.POST.get('heure')
            reference_achat = request.POST.get('reference_achat')
            prix_total_global = request.POST.get('prix_total_global')
            statut = request.POST.get('statut')
            commentaire = request.POST.get('commentaire')
            
            print(f"[NOTE] Données extraites:")
            print(f"  - Numéro fournisseur: {numero_fournisseur}")
            print(f"  - Date achat: {date_achat}")
            print(f"  - Heure: {heure}")
            print(f"  - Référence: {reference_achat}")
            print(f"  - Prix total: {prix_total_global}")
            print(f"  - Statut: {statut}")
            print(f"  - Commentaire: {commentaire}")
            
            print("[SEARCH] AVANT VALIDATION")
            
            # Validation
            if not all([numero_fournisseur, date_achat, heure, reference_achat, prix_total_global]):
                print("[ERREUR] VALIDATION ÉCHOUÉE")
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_facture_achat')
            
            print("[OK] VALIDATION RÉUSSIE")
            
            print("[SEARCH] AVANT CRÉATION FOURNISSEUR")
            
            # Récupérer ou créer le fournisseur
            fournisseur, created = Fournisseur.objects.get_or_create(
                intitule=numero_fournisseur,
                defaults={'agence': agence}
            )
            
            print(f"[OK] FOURNISSEUR: {fournisseur.intitule}")
            print("[SEARCH] AVANT CRÉATION FACTURE")
            
            try:
                # Vérifier si la référence existe déjà
                if FactureAchat.objects.filter(reference_achat=reference_achat).exists():
                    # Générer une référence unique
                    import time
                    reference_achat = f"{reference_achat}_{int(time.time())}"
                    print(f"[REFRESH] Référence modifiée pour éviter le doublon: {reference_achat}")
                
                # Créer la facture d'achat
                facture = FactureAchat.objects.create(
                    numero_fournisseur=numero_fournisseur,
                    date_achat=date_achat,
                    heure=heure,
                    reference_achat=reference_achat,
                    prix_total_global=float(prix_total_global),
                    statut=statut,
                    commentaire=commentaire,
                    fournisseur=fournisseur,
                    agence=agence
                )
                print(f"[OK] FACTURE CRÉÉE: {facture.id}")
            except Exception as e:
                print(f"[ERREUR] ERREUR CRÉATION FACTURE: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f'Erreur lors de la création de la facture: {str(e)}')
                return redirect('creer_facture_achat')
            
            print("[SEARCH] AVANT TRAITEMENT ARTICLES")
            
            # Traiter les articles sélectionnés
            articles_data = request.POST.get('articles_data', '')
            print(f"[SEARCH] Articles data reçus: {articles_data}")
            
            try:
                if articles_data:
                    import json
                    try:
                        articles = json.loads(articles_data)
                        print(f"[PACKAGE] Articles à traiter: {len(articles)}")
                        
                        for i, article_data in enumerate(articles):
                            print(f"[NOTE] Traitement article {i+1}: {article_data}")
                            
                            # Récupérer l'article
                            article = Article.objects.get(id=article_data['id'])
                            print(f"[OK] Article trouvé: {article.designation}")
                            
                            # Convertir les quantités et prix avec les fonctions sécurisées
                            quantite_decimale = safe_quantity_conversion(article_data['quantite'])
                            prix_achat_decimal = safe_price_conversion(article_data['prix_achat'])
                            prix_total_decimale = safe_decimal_calculation(quantite_decimale, prix_achat_decimal, 'multiply')
                            
                            # Créer la ligne de facture
                            ligne = LigneFactureAchat.objects.create(
                                facture_achat=facture,
                                article=article,
                                reference_article=article.reference_article,
                                designation=article.designation,
                                prix_unitaire=prix_achat_decimal,
                                quantite=quantite_decimale,
                                prix_total_article=prix_total_decimale
                            )
                            print(f"[OK] Ligne créée: {ligne.id}")
                            
                            # Mettre à jour le stock de l'article
                            ancien_stock = article.stock_actuel
                            quantite_ajoutee = quantite_decimale  # Déjà convertie en Decimal
                            article.stock_actuel += quantite_ajoutee
                            
                            # Mettre à jour le dernier prix d'achat avec le nouveau prix
                            ancien_dernier_prix = article.dernier_prix_achat
                            nouveau_prix_achat = prix_achat_decimal  # Déjà converti en Decimal
                            article.dernier_prix_achat = nouveau_prix_achat
                            
                            # Sauvegarder l'article avec validation
                            article.save()
                            
                            # Vérifier que la mise à jour a bien fonctionné
                            article.refresh_from_db()
                            if article.stock_actuel != ancien_stock + quantite_ajoutee:
                                raise Exception(f"Erreur de mise à jour du stock pour {article.designation}")
                            
                            print(f"[STATS] Stock mis à jour: {ancien_stock} → {article.stock_actuel}")
                            print(f"[MONEY] Dernier prix d'achat mis à jour: {ancien_dernier_prix} → {nouveau_prix_achat}")
                            
                            # [HOT] CRÉER UN MOUVEMENT DE STOCK POUR TRAÇABILITÉ
                            try:
                                MouvementStock.objects.create(
                                    article=article,
                                    agence=agence,
                                    type_mouvement='entree',
                                    date_mouvement=timezone.now(),
                                    numero_piece=facture.reference_achat,
                                    quantite_stock=article.stock_actuel,
                                    stock_initial=ancien_stock,
                                    solde=article.stock_actuel,
                                    quantite=quantite_decimale,  # Utiliser la quantité décimale convertie
                                    cout_moyen_pondere=safe_price_conversion(article.prix_achat),
                                    stock_permanent=safe_decimal_calculation(article.stock_actuel, safe_price_conversion(article.prix_achat), 'multiply'),
                                    facture_achat=facture,
                                    fournisseur=facture.fournisseur,
                                    commentaire=f"Achat - Facture {facture.reference_achat}"
                                )
                                print(f"[NOTE] MOUVEMENT STOCK - Entrée enregistrée pour {article.designation}")
                            except Exception as e:
                                print(f"[WARNING] ERREUR MOUVEMENT STOCK ACHAT: {e}")
                            
                    except (json.JSONDecodeError, Article.DoesNotExist, KeyError) as e:
                        print(f"[ERREUR] Erreur lors du traitement des articles: {e}")
                        messages.error(request, f'Erreur lors du traitement des articles: {str(e)}')
                else:
                    print("[WARNING] Aucun article sélectionné")
            except Exception as e:
                print(f"[ERREUR] ERREUR GÉNÉRALE: {e}")
                import traceback
                traceback.print_exc()
            
            messages.success(request, f'Facture d\'achat "{reference_achat}" créée avec succès! Stock mis à jour.')
            return redirect('creer_facture_achat')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de la facture: {str(e)}')
            return redirect('creer_facture_achat')
    
    # GET - Afficher le formulaire
    statut_choices = FactureAchat.STATUT_CHOICES
    
    context = {
        'statut_choices': statut_choices,
    }
    return render(request, 'supermarket/stock/creer_facture_achat.html', context)

@login_required
def detail_facture_achat(request, facture_id):
    """Vue pour afficher les détails d'une facture d'achat"""
    try:
        agence = get_user_agence(request)
        facture = FactureAchat.objects.get(id=facture_id, agence=agence)
        lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
        
        print(f"[SEARCH] DÉTAIL FACTURE D'ACHAT")
        print(f"[LIST] Facture ID: {facture_id}")
        print(f"[LIST] Facture: {facture.reference_achat}")
        print(f"[LIST] Agence: {agence.nom_agence}")
        print(f"[PACKAGE] Nombre de lignes trouvées: {lignes.count()}")
        
        # Debug: vérifier toutes les lignes de facture d'achat
        toutes_lignes = LigneFactureAchat.objects.all()
        print(f"[CHART] TOTAL LIGNES DANS LA BASE: {toutes_lignes.count()}")
        
        for ligne in toutes_lignes:
            print(f"  [NOTE] Ligne globale: {ligne.designation} - Facture: {ligne.facture_achat.reference_achat}")
        
        for i, ligne in enumerate(lignes):
            print(f"  [NOTE] Ligne {i+1}: {ligne.designation} - Qty: {ligne.quantite} - Prix: {ligne.prix_unitaire}")
        
        context = {
            'facture': facture,
            'lignes': lignes,
        }
        return render(request, 'supermarket/stock/detail_facture_achat.html', context)
    except FactureAchat.DoesNotExist:
        print(f"[ERREUR] Facture d'achat {facture_id} non trouvée")
        messages.error(request, 'Facture d\'achat non trouvée.')
        return redirect('consulter_factures_achat')

@login_required
def modifier_facture_achat(request, facture_id):
    """Vue pour modifier une facture d'achat existante"""
    try:
        agence = get_user_agence(request)
        facture = FactureAchat.objects.get(id=facture_id, agence=agence)
    except FactureAchat.DoesNotExist:
        messages.error(request, 'Facture d\'achat non trouvée.')
        return redirect('consulter_factures_achat')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_fournisseur = request.POST.get('numero_fournisseur')
            date_achat = request.POST.get('date_achat')
            heure = request.POST.get('heure')
            reference_achat = request.POST.get('reference_achat')
            prix_total_global = request.POST.get('prix_total_global')
            statut = request.POST.get('statut')
            commentaire = request.POST.get('commentaire')
            
            # Validation
            if not all([numero_fournisseur, date_achat, heure, reference_achat, prix_total_global]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_facture_achat', facture_id=facture_id)
            
            # Mettre à jour la facture d'achat
            facture.numero_fournisseur = numero_fournisseur
            facture.date_achat = date_achat
            facture.heure = heure
            facture.reference_achat = reference_achat
            facture.prix_total_global = float(prix_total_global)
            facture.statut = statut
            facture.commentaire = commentaire
            facture.save()
            
            messages.success(request, f'Facture d\'achat "{reference_achat}" modifiée avec succès!')
            return redirect('consulter_factures_achat')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de la facture: {str(e)}')
            return redirect('modifier_facture_achat', facture_id=facture_id)
    
    # GET - Afficher le formulaire pré-rempli
    statut_choices = FactureAchat.STATUT_CHOICES
    
    # Récupérer les lignes d'articles de la facture
    lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
    
    context = {
        'facture': facture,
        'statut_choices': statut_choices,
        'lignes': lignes,
    }
    return render(request, 'supermarket/stock/modifier_facture_achat.html', context)

@login_required
def supprimer_facture_achat(request, facture_id):
    """Vue pour supprimer une facture d'achat"""
    print(f"🗑️ SUPPRESSION FACTURE D'ACHAT")
    print(f"[LIST] Méthode: {request.method}")
    print(f"[LIST] Facture ID: {facture_id}")
    
    if request.method != 'POST':
        print("[ERREUR] Méthode non autorisée - redirection")
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_factures_achat')
    
    try:
        agence = get_user_agence(request)
        print(f"[BUILDING] Agence: {agence}")
        
        facture = FactureAchat.objects.get(id=facture_id, agence=agence)
        print(f"[LIST] Facture trouvée: {facture.reference_achat}")
        
        facture_name = facture.reference_achat
        facture.delete()
        print(f"[OK] Facture supprimée: {facture_name}")
        
        messages.success(request, f'Facture d\'achat "{facture_name}" supprimée avec succès!')
    except FactureAchat.DoesNotExist:
        print(f"[ERREUR] Facture d'achat {facture_id} non trouvée")
        messages.error(request, 'Facture d\'achat non trouvée.')
    except Exception as e:
        print(f"[ERREUR] Erreur lors de la suppression: {e}")
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_factures_achat')

# ==================== FACTURES DE TRANSFERT ====================

@login_required
def consulter_factures_transfert(request):
    """Vue pour consulter les factures de transfert"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    statut_filter = request.GET.get('statut_filter', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Construire la requête de base (factures où l'agence est source ou destination)
    factures = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    )
    
    # Appliquer les filtres
    if search_query:
        factures = factures.filter(
            Q(numero_compte__icontains=search_query) |
            Q(reference_transfert__icontains=search_query) |
            Q(lieu_depart__icontains=search_query) |
            Q(lieu_arrivee__icontains=search_query)
        )
    
    if statut_filter:
        factures = factures.filter(statut=statut_filter)
    
    if date_debut:
        factures = factures.filter(date_transfert__gte=date_debut)
    
    if date_fin:
        factures = factures.filter(date_transfert__lte=date_fin)
    
    # Trier par date décroissante
    factures = factures.order_by('-date_transfert')
    
    # Calculer les statistiques
    total_factures = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    ).count()
    factures_en_cours = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence),
        statut='en_cours'
    ).count()
    factures_terminees = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence),
        statut='termine'
    ).count()
    quantite_totale = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    ).aggregate(
        total=Sum('quantite')
    )['total'] or 0
    
    # Récupérer les statuts pour le filtre
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'factures': factures,
        'search_query': search_query,
        'statut_filter': statut_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statut_choices': statut_choices,
        'total_factures': total_factures,
        'factures_en_cours': factures_en_cours,
        'factures_terminees': factures_terminees,
        'quantite_totale': quantite_totale,
    }
    return render(request, 'supermarket/stock/consulter_factures_transfert.html', context)

@login_required
def creer_facture_transfert(request):
    """Vue pour créer une nouvelle facture de transfert"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_compte = request.POST.get('numero_compte')
            date_transfert = request.POST.get('date_transfert')
            reference_transfert = request.POST.get('reference_transfert')
            lieu_depart = request.POST.get('lieu_depart')
            lieu_arrivee = request.POST.get('lieu_arrivee')
            quantite = request.POST.get('quantite')
            statut = request.POST.get('statut')
            employe_expediteur = request.POST.get('employe_expediteur', '').strip()
            employe_destinataire = request.POST.get('employe_destinataire', '').strip()
            etat = request.POST.get('etat', 'sortir')
            
            # Validation
            if not all([numero_compte, date_transfert, reference_transfert, lieu_depart, lieu_arrivee, quantite, employe_expediteur]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_facture_transfert')
            
            # Rechercher les employés existants par nom
            expediteur_employe = None
            destinataire_employe = None
            
            # Rechercher l'employé expéditeur
            try:
                # Essayer de trouver par nom complet dans le compte
                expediteur_employe = Employe.objects.filter(
                    compte__agence=agence,
                    compte__user__first_name__icontains=employe_expediteur.split()[0] if employe_expediteur.split() else employe_expediteur
                ).first()
                
                # Si pas trouvé, essayer par nom de famille
                if not expediteur_employe and len(employe_expediteur.split()) > 1:
                    expediteur_employe = Employe.objects.filter(
                        compte__agence=agence,
                        compte__user__last_name__icontains=employe_expediteur.split()[-1]
                    ).first()
                
                # Si toujours pas trouvé, prendre le premier employé de l'agence
                if not expediteur_employe:
                    expediteur_employe = Employe.objects.filter(compte__agence=agence).first()
                    
            except Exception as e:
                print(f"Erreur lors de la recherche de l'employé expéditeur: {e}")
                expediteur_employe = Employe.objects.filter(compte__agence=agence).first()
            
            if not expediteur_employe:
                messages.error(request, 'Aucun employé trouvé dans cette agence.')
                return redirect('creer_facture_transfert')
            
            # Afficher un message informatif sur l'employé trouvé
            expediteur_nom = f"{expediteur_employe.compte.user.first_name} {expediteur_employe.compte.user.last_name}".strip()
            messages.info(request, f'Employé expéditeur trouvé: {expediteur_nom}')
            
            if destinataire_employe:
                destinataire_nom = f"{destinataire_employe.compte.user.first_name} {destinataire_employe.compte.user.last_name}".strip()
                messages.info(request, f'Employé destinataire trouvé: {destinataire_nom}')
            
            # Rechercher l'employé destinataire (si fourni)
            if employe_destinataire:
                try:
                    # Essayer de trouver par nom complet dans le compte
                    destinataire_employe = Employe.objects.filter(
                        compte__agence=agence,
                        compte__user__first_name__icontains=employe_destinataire.split()[0] if employe_destinataire.split() else employe_destinataire
                    ).first()
                    
                    # Si pas trouvé, essayer par nom de famille
                    if not destinataire_employe and len(employe_destinataire.split()) > 1:
                        destinataire_employe = Employe.objects.filter(
                            compte__agence=agence,
                            compte__user__last_name__icontains=employe_destinataire.split()[-1]
                        ).first()
                        
                except Exception as e:
                    print(f"Erreur lors de la recherche de l'employé destinataire: {e}")
                    destinataire_employe = None
            
            # Créer la facture de transfert
            facture = FactureTransfert.objects.create(
                numero_compte=numero_compte,
                date_transfert=date_transfert,
                reference_transfert=reference_transfert,
                lieu_depart=lieu_depart,
                lieu_arrivee=lieu_arrivee,
                quantite=int(quantite),
                statut=statut,
                agence_source=agence,
                agence_destination=agence,  # Pour l'instant, même agence (à modifier selon les besoins)
                employe_expediteur=expediteur_employe,
                employe_destinataire=destinataire_employe,
                etat=etat
            )
            
            # Traiter les articles sélectionnés
            articles_data = request.POST.get('articles_data', '')
            if articles_data:
                import json
                try:
                    articles = json.loads(articles_data)
                    for article_data in articles:
                        # Récupérer l'article
                        article = Article.objects.get(id=article_data['id'])
                        
                        # Créer la ligne de facture de transfert
                        LigneFactureTransfert.objects.create(
                            facture_transfert=facture,
                            article=article,
                            quantite=int(article_data['quantite']),
                            prix_unitaire=float(article_data['prix_achat']),
                            valeur_totale=float(article_data['prix_achat']) * int(article_data['quantite'])
                        )
                        
                        # Mettre à jour le stock de l'article (déduction pour transfert)
                        ancien_stock = article.stock_actuel
                        # Convertir en Decimal pour éviter les erreurs de type
                        from decimal import Decimal
                        quantite_decimal = Decimal(str(article_data['quantite']))
                        article.stock_actuel -= quantite_decimal
                        if article.stock_actuel < 0:
                            article.stock_actuel = 0
                        
                        # Mettre à jour le dernier prix d'achat avec le prix du transfert
                        ancien_dernier_prix = article.dernier_prix_achat
                        nouveau_prix_achat = float(article_data['prix_achat'])
                        article.dernier_prix_achat = nouveau_prix_achat
                        
                        article.save()
                        print(f"[PACKAGE] STOCK TRANSFERT - Article: {article.designation}")
                        print(f"[PACKAGE] STOCK TRANSFERT - Stock mis à jour: {ancien_stock} → {article.stock_actuel}")
                        print(f"[MONEY] Transfert - Dernier prix d'achat mis à jour: {ancien_dernier_prix} → {nouveau_prix_achat}")
                        
                        # [HOT] CRÉER UN MOUVEMENT DE STOCK POUR TRAÇABILITÉ
                        try:
                            MouvementStock.objects.create(
                                article=article,
                                agence=agence,
                                type_mouvement='sortie',
                                date_mouvement=timezone.now(),
                                numero_piece=facture.reference_transfert,
                                quantite_stock=article.stock_actuel,
                                stock_initial=ancien_stock,
                                solde=article.stock_actuel,
                                quantite=int(article_data['quantite']),
                                cout_moyen_pondere=float(article.prix_achat),
                                stock_permanent=float(article.stock_actuel * article.prix_achat),
                                commentaire=f"Transfert - Facture {facture.reference_transfert}"
                            )
                            print(f"[NOTE] MOUVEMENT STOCK - Sortie transfert enregistrée pour {article.designation}")
                        except Exception as e:
                            print(f"[WARNING] ERREUR MOUVEMENT STOCK TRANSFERT: {e}")
                        
                except (json.JSONDecodeError, Article.DoesNotExist, KeyError) as e:
                    print(f"Erreur lors du traitement des articles: {e}")
            
            messages.success(request, f'Facture de transfert "{reference_transfert}" créée avec succès!')
            return redirect('creer_facture_transfert')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de la facture: {str(e)}')
            return redirect('creer_facture_transfert')
    
    # GET - Afficher le formulaire
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'statut_choices': statut_choices,
    }
    return render(request, 'supermarket/stock/creer_facture_transfert_nouveau.html', context)

@login_required
def detail_facture_transfert(request, facture_id):
    """Vue pour afficher les détails d'une facture de transfert"""
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
        lignes = LigneFactureTransfert.objects.filter(facture_transfert=facture)
        
        context = {
            'facture': facture,
            'lignes': lignes,
        }
        return render(request, 'supermarket/stock/detail_facture_transfert.html', context)
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
        return redirect('consulter_factures_transfert')

@login_required
def modifier_facture_transfert(request, facture_id):
    """Vue pour modifier une facture de transfert existante"""
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
        return redirect('consulter_factures_transfert')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_compte = request.POST.get('numero_compte')
            date_transfert = request.POST.get('date_transfert')
            reference_transfert = request.POST.get('reference_transfert')
            lieu_depart = request.POST.get('lieu_depart')
            lieu_arrivee = request.POST.get('lieu_arrivee')
            quantite = request.POST.get('quantite')
            statut = request.POST.get('statut')
            employe_expediteur = request.POST.get('employe_expediteur', '').strip()
            employe_destinataire = request.POST.get('employe_destinataire', '').strip()
            etat = request.POST.get('etat', 'sortir')
            
            # Validation
            if not all([numero_compte, date_transfert, reference_transfert, lieu_depart, lieu_arrivee, quantite]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            # Mettre à jour la facture de transfert
            facture.numero_compte = numero_compte
            facture.date_transfert = date_transfert
            facture.reference_transfert = reference_transfert
            facture.lieu_depart = lieu_depart
            facture.lieu_arrivee = lieu_arrivee
            facture.quantite = int(quantite)
            facture.statut = statut
            facture.save()
            
            messages.success(request, f'Facture de transfert "{reference_transfert}" modifiée avec succès!')
            return redirect('consulter_factures_transfert')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de la facture: {str(e)}')
            return redirect('modifier_facture_transfert', facture_id=facture_id)
    
    # GET - Afficher le formulaire pré-rempli
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'facture': facture,
        'statut_choices': statut_choices,
    }
    return render(request, 'supermarket/stock/modifier_facture_transfert.html', context)

@login_required
def supprimer_facture_transfert(request, facture_id):
    """Vue pour supprimer une facture de transfert"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_factures_transfert')
    
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
        facture_name = facture.reference_transfert
        facture.delete()
        messages.success(request, f'Facture de transfert "{facture_name}" supprimée avec succès!')
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_factures_transfert')

# ===== RECHERCHE D'ARTICLES POUR STOCK =====

@login_required
def search_articles_stock(request):
    """Vue pour la recherche d'articles dans le module de stock"""
    search_term = request.GET.get('q', '')
    
    print(f"[SEARCH] search_articles_stock: recherche pour '{search_term}'")
    
    # Vérifier d'abord s'il y a des articles dans la base de données
    total_articles = Article.objects.count()
    print(f"[CHART] Total articles dans la base de données: {total_articles}")
    
    if total_articles == 0:
        print("[ERREUR] Aucun article dans la base de données!")
        return JsonResponse({'articles': []})
    
    agence = get_user_agence(request)
    print(f"[BUILDING] Agence trouvée: {agence}")
    print(f"[BUILDING] ID de l'agence: {agence.id_agence if agence else 'None'}")
    print(f"[BUILDING] Nom de l'agence: {agence.nom_agence if agence else 'None'}")
    
    if not agence:
        print("[ERREUR] Aucune agence trouvée")
        return JsonResponse({'articles': []})
    
    # Test: afficher tous les articles sans filtre d'agence
    articles_all = Article.objects.all()
    print(f"[PACKAGE] Tous les articles (toutes agences): {articles_all.count()}")
    for article in articles_all[:5]:  # Afficher les 5 premiers
        print(f"  - {article.designation} (Agence: {article.agence.nom_agence if article.agence else 'None'})")
    
    # Test: afficher les articles de cette agence spécifique
    articles_agence = Article.objects.filter(agence=agence)
    print(f"[PACKAGE] Articles de l'agence {agence.nom_agence}: {articles_agence.count()}")
    for article in articles_agence[:5]:  # Afficher les 5 premiers
        print(f"  - {article.designation} (ID: {article.id})")
    
    articles = []
    
    if search_term and len(search_term) >= 1:
        # Recherche avec filtre d'agence dès qu'on tape 1 lettre
        articles = Article.objects.filter(agence=agence, designation__icontains=search_term)[:50]
        print(f"[SEARCH] Articles trouvés avec recherche '{search_term}' (agence {agence.nom_agence}): {articles.count()}")
    else:
        # Afficher tous les articles de l'agence si pas de terme de recherche
        articles = Article.objects.filter(agence=agence)[:50]
        print(f"[PACKAGE] Tous les articles de l'agence {agence.nom_agence}: {articles.count()}")
    
    # Convertir les articles en format JSON
    articles_data = []
    for article in articles:
        articles_data.append({
            'id': article.id,
            'designation': article.designation,
            'prix_achat': float(article.prix_achat),
            'stock': article.stock_actuel,
            'reference_article': article.reference_article,
        })
        print(f"[NOTE] Article: {article.designation} (ID: {article.id})")
    
    print(f"[CHART] Total articles_data: {len(articles_data)}")
    return JsonResponse({'articles': articles_data})

def create_test_articles(request):
    """Vue temporaire pour créer des articles de test"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'error': 'Aucune agence trouvée'})
    
    # Créer quelques articles de test
    test_articles = [
        {'designation': 'Ordinateur Portable', 'prix_achat': 500000, 'prix_vente': 600000, 'stock_actuel': 10},
        {'designation': 'Souris USB', 'prix_achat': 5000, 'prix_vente': 7500, 'stock_actuel': 50},
        {'designation': 'Clavier Mécanique', 'prix_achat': 15000, 'prix_vente': 20000, 'stock_actuel': 25},
        {'designation': 'Écran 24 pouces', 'prix_achat': 80000, 'prix_vente': 100000, 'stock_actuel': 15},
        {'designation': 'Casque Audio', 'prix_achat': 25000, 'prix_vente': 35000, 'stock_actuel': 30},
    ]
    
    created_articles = []
    for article_data in test_articles:
        article, created = Article.objects.get_or_create(
            designation=article_data['designation'],
            agence=agence,
            defaults=article_data
        )
        if created:
            created_articles.append(article.designation)
    
    return JsonResponse({
        'message': f'Articles créés: {len(created_articles)}',
        'articles': created_articles
    })

# ==================== INVENTAIRE DE STOCK ====================

@login_required
def inventaire_stock(request):
    """Vue pour la page d'inventaire de stock"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques du stock
    total_articles = articles.count()
    total_quantite = articles.aggregate(total=Sum('stock_actuel'))['total'] or 0
    valeur_totale_stock = articles.aggregate(
        total=Sum(F('stock_actuel') * F('prix_achat'))
    )['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'total_quantite': total_quantite,
        'valeur_totale_stock': valeur_totale_stock,
    }
    
    return render(request, 'supermarket/stock/inventaire_stock.html', context)

@login_required
def generer_inventaire(request):
    """Vue pour générer l'inventaire selon les critères sélectionnés"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        famille_id = request.POST.get('famille')
        selection_articles = request.POST.get('selection_articles', 'tous')  # 'tous' ou 'selectionnes'
        articles_selectionnes = request.POST.getlist('articles_selectionnes[]')
        
        print(f"[SEARCH] PARAMÈTRES INVENTAIRE:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Famille: {famille_id}")
        print(f"  - Sélection: {selection_articles}")
        print(f"  - Articles sélectionnés: {articles_selectionnes}")
        
        # Filtrer les articles selon les critères
        articles_query = Article.objects.filter(agence=agence)
        
        # Filtre par famille
        if famille_id and famille_id != '':
            articles_query = articles_query.filter(categorie_id=famille_id)
        
        # Filtre par sélection d'articles
        if selection_articles == 'selectionnes' and articles_selectionnes:
            articles_query = articles_query.filter(id__in=articles_selectionnes)
        
        articles = articles_query.order_by('designation')
        
        print(f"[PACKAGE] ARTICLES FILTRÉS: {articles.count()}")
        
        # Créer l'inventaire
        numero_inventaire = f"INV-{timezone.now().strftime('%Y%m%d-%H%M%S')}"
        
        # Récupérer l'employé responsable
        employe = Employe.objects.filter(compte__agence=agence).first()
        
        inventaire = InventaireStock.objects.create(
            numero_inventaire=numero_inventaire,
            date_debut=timezone.now(),
            statut='en_cours',
            agence=agence,
            responsable=employe,
            commentaire=f"Inventaire généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        )
        
        # Créer les lignes d'inventaire
        total_quantite = 0
        total_valeur = 0
        
        for article in articles:
            valeur = float(article.stock_actuel) * float(article.prix_achat)
            
            LigneInventaireStock.objects.create(
                inventaire=inventaire,
                reference_article=article.reference_article,
                designation=article.designation,
                quantite_stock=article.stock_actuel,
                prix_unitaire=article.prix_achat,
                valeur=valeur,
                conditionnement=article.conditionnement,
                article=article
            )
            
            total_quantite += article.stock_actuel
            total_valeur += valeur
        
        # Marquer l'inventaire comme terminé
        inventaire.date_fin = timezone.now()
        inventaire.statut = 'termine'
        inventaire.save()
        
        print(f"[OK] INVENTAIRE CRÉÉ: {numero_inventaire}")
        print(f"[CHART] TOTAUX: {total_quantite} articles, {total_valeur} FCFA")
        
        return JsonResponse({
            'success': True,
            'inventaire_id': inventaire.id,
            'numero_inventaire': numero_inventaire,
            'total_articles': articles.count(),
            'total_quantite': total_quantite,
            'total_valeur': total_valeur
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR GÉNÉRATION INVENTAIRE: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_excel(request):
    """Vue pour exporter l'inventaire en format Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible, utilisation du format CSV")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Inventaire {inventaire.numero_inventaire}"
        
        # Style du titre
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire le titre "Inventaire" avec la date du jour
        date_du_jour = timezone.now().strftime('%d/%m/%Y')
        title_cell = ws.cell(row=1, column=1, value=f"Inventaire - {date_du_jour}")
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # Fusionner les cellules pour le titre (sur toutes les colonnes)
        from openpyxl.utils import get_column_letter
        ws.merge_cells(f'A1:{get_column_letter(6)}1')
        
        # Ligne vide
        ws.row_dimensions[2].height = 5
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité en Stock', 'Prix Unitaire (FCFA)', 'Valeur Totale (FCFA)']
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire les en-têtes (décalés à la ligne 3)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Écrire les données (décalées à partir de la ligne 4)
        total_quantite = 0
        total_valeur = 0
        
        for idx, ligne in enumerate(lignes):
            row_num = 4 + idx  # Commencer à la ligne 4 (après titre ligne 1, ligne vide ligne 2, en-têtes ligne 3)
            ws.cell(row=row_num, column=1, value=ligne.reference_article)
            ws.cell(row=row_num, column=2, value=ligne.designation)
            ws.cell(row=row_num, column=3, value=ligne.conditionnement)
            ws.cell(row=row_num, column=4, value=ligne.quantite_stock)
            ws.cell(row=row_num, column=5, value=float(ligne.prix_unitaire))
            ws.cell(row=row_num, column=6, value=float(ligne.valeur))
            
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux (décalée de 2 lignes supplémentaires)
        total_row = 4 + len(lignes) + 2  # Ligne de données + 1 ligne vide + ligne totaux
        ws.cell(row=total_row, column=3, value="TOTAL GÉNÉRAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_quantite).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=total_valeur).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        column_widths = [15, 40, 15, 15, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le classeur dans la réponse
        wb.save(response)
        
        print(f"[CHART] EXPORT EXCEL - Inventaire {inventaire.numero_inventaire}")
        print(f"[CHART] EXPORT EXCEL - {len(lignes)} articles exportés")
        print(f"[CHART] EXPORT EXCEL - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_pdf(request):
    """Vue pour exporter l'inventaire en format PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Centre
        )
        
        # Titre
        title = Paragraph(f"INVENTAIRE DE STOCK - {inventaire.numero_inventaire}", title_style)
        elements.append(title)
        
        # Informations de l'inventaire
        info_data = [
            ['Date de génération:', inventaire.date_debut.strftime('%d/%m/%Y à %H:%M')],
            ['Agence:', agence.nom_agence],
            ['Responsable:', inventaire.responsable.compte.nom_complet if inventaire.responsable else 'Non spécifié'],
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 20))
        
        # Tableau des articles
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité', 'Prix Unitaire', 'Valeur']
        
        # Données du tableau
        data = [headers]
        total_quantite = 0
        total_valeur = 0
        
        for ligne in lignes:
            row = [
                ligne.reference_article,
                ligne.designation,
                ligne.conditionnement,
                str(ligne.quantite_stock),
                f"{float(ligne.prix_unitaire):,.0f}",
                f"{float(ligne.valeur):,.0f}"
            ]
            data.append(row)
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux
        data.append(['', '', 'TOTAL GÉNÉRAL:', str(total_quantite), '', f"{total_valeur:,.0f}"])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 6*cm, 2.5*cm, 2*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            # En-têtes
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Données
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Ligne des totaux
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF - Inventaire {inventaire.numero_inventaire}")
        print(f"📄 EXPORT PDF - {len(lignes)} articles exportés")
        print(f"📄 EXPORT PDF - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_csv(request):
    """Vue pour exporter l'inventaire en format CSV (alternative si Excel n'est pas disponible)"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité en Stock', 'Prix Unitaire (FCFA)', 'Valeur Totale (FCFA)']
        writer.writerow(headers)
        
        # Données
        total_quantite = 0
        total_valeur = 0
        
        for ligne in lignes:
            row = [
                ligne.reference_article,
                ligne.designation,
                ligne.conditionnement,
                ligne.quantite_stock,
                float(ligne.prix_unitaire),
                float(ligne.valeur)
            ]
            writer.writerow(row)
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['', '', 'TOTAL GÉNÉRAL:', total_quantite, '', total_valeur])
        
        print(f"[CHART] EXPORT CSV - Inventaire {inventaire.numero_inventaire}")
        print(f"[CHART] EXPORT CSV - {len(lignes)} articles exportés")
        print(f"[CHART] EXPORT CSV - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== STATISTIQUES DE VENTE ====================

@login_required
def statistiques_vente(request):
    """Vue pour la page des statistiques de vente"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques générales
    total_articles = articles.count()
    
    # Calculer les statistiques de vente des 30 derniers jours
    date_debut = timezone.now().date() - timezone.timedelta(days=30)
    
    # Récupérer les ventes des 30 derniers jours
    ventes_recentes = LigneFactureVente.objects.filter(
        facture_vente__agence=agence,
        facture_vente__date__gte=date_debut
    ).select_related('article', 'facture_vente')
    
    # Calculer le chiffre d'affaires total
    chiffre_affaires_total = float(ventes_recentes.aggregate(
        total=Sum(F('quantite') * F('prix_unitaire'))
    )['total'] or 0)
    
    # Calculer la marge totale
    from decimal import Decimal
    marge_totale = Decimal('0')
    for vente in ventes_recentes:
        prix_achat = Decimal(str(vente.article.prix_achat))
        prix_vente = Decimal(str(vente.prix_unitaire))
        marge_unitaire = prix_vente - prix_achat
        marge_totale += marge_unitaire * Decimal(str(vente.quantite))
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'chiffre_affaires_total': chiffre_affaires_total,
        'marge_totale': float(marge_totale),
        'pourcentage_marge_global': (float(marge_totale) / float(chiffre_affaires_total) * 100) if chiffre_affaires_total > 0 else 0,
    }
    
    return render(request, 'supermarket/stock/statistiques_vente.html', context)

@login_required
def generer_statistiques_vente(request):
    """Vue pour générer les statistiques de vente selon les critères sélectionnés"""
    print("[START] DÉBUT GENERER_STATISTIQUES_VENTE")
    
    if request.method != 'POST':
        print("[ERREUR] Méthode non POST")
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
        print(f"[SEARCH] Agence récupérée: {agence}")
    except Exception as e:
        print(f"[ERREUR] Erreur get_user_agence: {e}")
        return JsonResponse({'success': False, 'error': f'Agence non trouvée: {str(e)}'})
    
    try:
        print("[SEARCH] Début du traitement des paramètres")
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        famille_id = request.POST.get('famille')
        selection_articles = request.POST.get('selection_articles', 'tous')
        articles_selectionnes = request.POST.getlist('articles_selectionnes[]')
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        print(f"[CHART] PARAMÈTRES STATISTIQUES:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Famille: {famille_id}")
        print(f"  - Sélection: {selection_articles}")
        print(f"  - Articles sélectionnés: {articles_selectionnes}")
        
        # Convertir les dates
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Format de date invalide: {str(e)}'})
        
        # Filtrer les articles selon les critères
        print(f"[SEARCH] Filtrage des articles pour agence: {agence}")
        articles_query = Article.objects.filter(agence=agence)
        print(f"[SEARCH] Articles de base: {articles_query.count()}")
        
        # Filtre par famille
        if famille_id and famille_id != '':
            print(f"[SEARCH] Filtrage par famille: {famille_id}")
            articles_query = articles_query.filter(categorie_id=famille_id)
            print(f"[SEARCH] Articles après filtre famille: {articles_query.count()}")
        
        # Filtre par sélection d'articles
        if selection_articles == 'selectionnes' and articles_selectionnes:
            print(f"[SEARCH] Filtrage par sélection: {articles_selectionnes}")
            articles_query = articles_query.filter(id__in=articles_selectionnes)
            print(f"[SEARCH] Articles après filtre sélection: {articles_query.count()}")
        
        articles = articles_query.order_by('designation')
        print(f"[PACKAGE] ARTICLES FILTRÉS: {articles.count()}")
        
        # Récupérer les ventes pour la période
        ventes = LigneFactureVente.objects.filter(
            facture_vente__agence=agence,
            facture_vente__date__gte=date_debut_obj,
            facture_vente__date__lte=date_fin_obj,
            article__in=articles
        ).select_related('article', 'facture_vente')
        
        # Calculer les statistiques par article
        statistiques_articles = []
        chiffre_affaires_total = 0.0
        marge_totale = 0.0
        quantite_totale_vendue = 0.0
        
        for article in articles:
            # Récupérer les ventes de cet article
            ventes_article = ventes.filter(article=article)
            
            # Calculer les totaux pour cet article
            quantite_vendue = float(ventes_article.aggregate(total=Sum('quantite'))['total'] or 0)
            chiffre_affaires_article = float(ventes_article.aggregate(
                total=Sum(F('quantite') * F('prix_unitaire'))
            )['total'] or 0)
            
            # Calculer la marge
            from decimal import Decimal
            prix_achat = Decimal(str(article.prix_achat))
            marge_unitaire = Decimal('0')
            marge_article = Decimal('0')
            
            if quantite_vendue > 0:
                prix_vente_moyen = Decimal(str(chiffre_affaires_article)) / Decimal(str(quantite_vendue))
                marge_unitaire = prix_vente_moyen - prix_achat
                marge_article = marge_unitaire * Decimal(str(quantite_vendue))
            
            # Calculer le pourcentage de marge
            if chiffre_affaires_article > 0:
                pourcentage_marge = (marge_article / Decimal(str(chiffre_affaires_article)) * Decimal('100'))
            else:
                pourcentage_marge = Decimal('0')
            
            if quantite_vendue > 0:  # Ne garder que les articles vendus
                statistiques_articles.append({
                    'reference_article': article.reference_article,
                    'designation': article.designation,
                    'quantite_vendue': float(quantite_vendue),
                    'chiffre_affaires': float(chiffre_affaires_article),
                    'marge_profit': float(marge_article),
                    'pourcentage_marge': float(pourcentage_marge),
                })
                
                chiffre_affaires_total += float(chiffre_affaires_article)
                marge_totale += float(marge_article)
                quantite_totale_vendue += float(quantite_vendue)
        
        # Calculer le pourcentage de marge global
        if chiffre_affaires_total > 0:
            pourcentage_marge_global = (marge_totale / float(chiffre_affaires_total) * 100)
        else:
            pourcentage_marge_global = 0
        
        print(f"[CHART] STATISTIQUES GÉNÉRÉES:")
        print(f"  - Articles vendus: {len(statistiques_articles)}")
        print(f"  - Quantité totale vendue: {quantite_totale_vendue}")
        print(f"  - Chiffre d'affaires total: {chiffre_affaires_total}")
        print(f"  - Marge totale: {marge_totale}")
        print(f"  - Pourcentage marge global: {pourcentage_marge_global:.2f}%")
        
        # Stocker les statistiques dans la session pour l'export (conversion en types sérialisables)
        request.session['statistiques_vente'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'statistiques_articles': statistiques_articles,  # Déjà converties en float
            'chiffre_affaires_total': float(chiffre_affaires_total),
            'marge_totale': float(marge_totale),
            'quantite_totale_vendue': int(quantite_totale_vendue),
            'pourcentage_marge_global': float(pourcentage_marge_global),
        }
        
        return JsonResponse({
            'success': True,
            'total_articles': len(statistiques_articles),
            'quantite_totale_vendue': int(quantite_totale_vendue),
            'chiffre_affaires_total': float(chiffre_affaires_total),
            'marge_totale': float(marge_totale),
            'pourcentage_marge_global': float(pourcentage_marge_global),
            'statistiques_articles': statistiques_articles  # Ajouter les données détaillées
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR GÉNÉRATION STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        error_details = f"Erreur: {str(e)}"
        if hasattr(e, '__class__'):
            error_details += f" (Type: {e.__class__.__name__})"
        return JsonResponse({'success': False, 'error': error_details})

@login_required
def export_statistiques_excel(request):
    """Vue pour exporter les statistiques de vente en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible, utilisation du format CSV")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Statistiques Vente {date_debut} - {date_fin}"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tête du document
        ws.merge_cells('A1:F1')
        ws['A1'] = f"STATISTIQUES DE VENTE - {agence.nom_agence}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Période: du {date_debut} au {date_fin}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes des colonnes
        headers = ['Référence', 'Désignation', 'Quantité Vendue', 'Chiffre d\'Affaires (FCFA)', 'Marge Profit (FCFA)', 'Pourcentage Marge (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données
        row = 5
        for stat in statistiques_articles:
            ws.cell(row=row, column=1, value=stat['reference_article'])
            ws.cell(row=row, column=2, value=stat['designation'])
            ws.cell(row=row, column=3, value=stat['quantite_vendue'])
            ws.cell(row=row, column=4, value=float(stat['chiffre_affaires']))
            ws.cell(row=row, column=5, value=float(stat['marge_profit']))
            ws.cell(row=row, column=6, value=float(stat['pourcentage_marge']))
            row += 1
        
        # Ligne des totaux
        row += 1
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value="").font = Font(bold=True)
        ws.cell(row=row, column=3, value=quantite_totale_vendue).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(chiffre_affaires_total)).font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(marge_totale)).font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(pourcentage_marge_global)).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 18
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le fichier dans la réponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_statistiques_pdf(request):
    """Vue pour exporter les statistiques de vente en PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            textColor=colors.darkblue
        )
        
        # Éléments du document
        elements = []
        
        # Titre
        title = Paragraph(f"STATISTIQUES DE VENTE - {agence.nom_agence}", title_style)
        elements.append(title)
        
        # Informations de la période
        period_text = f"<b>Période:</b> du {date_debut} au {date_fin}"
        period_para = Paragraph(period_text, styles['Normal'])
        elements.append(period_para)
        elements.append(Spacer(1, 20))
        
        # Données du tableau
        data = [['Référence', 'Désignation', 'Qté Vendue', 'Chiffre d\'Affaires', 'Marge Profit', 'Marge %']]
        
        for stat in statistiques_articles:
            data.append([
                stat['reference_article'],
                stat['designation'][:30] + '...' if len(stat['designation']) > 30 else stat['designation'],
                str(stat['quantite_vendue']),
                f"{float(stat['chiffre_affaires']):,.0f} FCFA",
                f"{float(stat['marge_profit']):,.0f} FCFA",
                f"{float(stat['pourcentage_marge']):.1f}%"
            ])
        
        # Ligne des totaux
        data.append([
            'TOTAL GÉNÉRAL',
            '',
            str(quantite_totale_vendue),
            f"{float(chiffre_affaires_total):,.0f} FCFA",
            f"{float(marge_totale):,.0f} FCFA",
            f"{float(pourcentage_marge_global):.1f}%"
        ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 6*cm, 2*cm, 3*cm, 3*cm, 2*cm])
        
        # Style du tableau
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"📄 EXPORT PDF STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"📄 EXPORT PDF STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_statistiques_csv(request):
    """Vue pour exporter les statistiques de vente en CSV"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Quantité Vendue', 'Chiffre d\'Affaires (FCFA)', 'Marge Profit (FCFA)', 'Pourcentage Marge (%)']
        writer.writerow(headers)
        
        # Données
        for stat in statistiques_articles:
            row = [
                stat['reference_article'],
                stat['designation'],
                stat['quantite_vendue'],
                float(stat['chiffre_affaires']),
                float(stat['marge_profit']),
                float(stat['pourcentage_marge'])
            ]
            writer.writerow(row)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['TOTAL GÉNÉRAL', '', quantite_totale_vendue, chiffre_affaires_total, marge_totale, pourcentage_marge_global])
        
        print(f"[CHART] EXPORT CSV STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"[CHART] EXPORT CSV STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"[CHART] EXPORT CSV STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_statistiques(request):
    """Vue de test pour diagnostiquer les problèmes"""
    try:
        print("🧪 TEST STATISTIQUES - Début")
        
        # Test 1: Récupération de l'agence
        agence = get_user_agence(request)
        print(f"🧪 TEST - Agence récupérée: {agence}")
        
        # Test 2: Vérification des imports
        from decimal import Decimal
        print("🧪 TEST - Import Decimal OK")
        
        # Test 3: Vérification des modèles
        articles_count = Article.objects.filter(agence=agence).count()
        print(f"🧪 TEST - Articles trouvés: {articles_count}")
        
        # Test 4: Vérification des ventes
        ventes_count = LigneFactureVente.objects.filter(facture_vente__agence=agence).count()
        print(f"🧪 TEST - Ventes trouvées: {ventes_count}")
        
        return JsonResponse({
            'success': True,
            'message': 'Tests réussis',
            'agence': str(agence),
            'articles_count': articles_count,
            'ventes_count': ventes_count
        })
        
    except Exception as e:
        print(f"🧪 TEST - Erreur: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== MOUVEMENTS DE STOCK ====================

@login_required
def mouvements_stock(request):
    """Vue pour la page des mouvements de stock"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques générales
    total_articles = articles.count()
    
    # Calculer les statistiques des mouvements des 30 derniers jours
    date_debut = timezone.now().date() - timezone.timedelta(days=30)
    
    # Récupérer les mouvements des 30 derniers jours
    mouvements_recentes = MouvementStock.objects.filter(
        agence=agence,
        date_mouvement__gte=date_debut
    ).select_related('article', 'fournisseur', 'facture_vente', 'facture_achat')
    
    # Statistiques des mouvements
    total_mouvements = mouvements_recentes.count()
    mouvements_entree = mouvements_recentes.filter(type_mouvement='entree').count()
    mouvements_sortie = mouvements_recentes.filter(type_mouvement='sortie').count()
    
    # Valeur totale du stock permanent
    valeur_stock_permanent = mouvements_recentes.aggregate(
        total=Sum('stock_permanent')
    )['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'total_mouvements': total_mouvements,
        'mouvements_entree': mouvements_entree,
        'mouvements_sortie': mouvements_sortie,
        'valeur_stock_permanent': valeur_stock_permanent,
    }
    
    return render(request, 'supermarket/stock/mouvements_stock.html', context)

@login_required
def consulter_mouvements_stock(request):
    """Vue pour consulter les mouvements de stock selon les critères sélectionnés"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        article_ids = request.POST.getlist('articles')  # Récupérer tous les articles sélectionnés
        type_mouvement = request.POST.get('type_mouvement', '')
        
        print(f"[CHART] PARAMÈTRES MOUVEMENTS:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Articles: {article_ids}")
        print(f"  - Type mouvement: {type_mouvement}")
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        if not article_ids:
            return JsonResponse({'success': False, 'error': 'Veuillez sélectionner au moins un article'})
        
        # Convertir les dates
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Format de date invalide: {str(e)}'})
        
        # Récupérer les articles
        try:
            articles = Article.objects.filter(id__in=article_ids, agence=agence)
            if not articles.exists():
                return JsonResponse({'success': False, 'error': 'Aucun article valide trouvé'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Erreur lors de la récupération des articles: {str(e)}'})
        
        # Filtrer les mouvements selon les critères
        # Utiliser __date pour comparer seulement les dates (ignorer l'heure)
        mouvements_query = MouvementStock.objects.filter(
            agence=agence,
            article__in=articles,  # Filtrer par plusieurs articles
            date_mouvement__date__gte=date_debut_obj,
            date_mouvement__date__lte=date_fin_obj
        ).select_related('article', 'fournisseur', 'facture_vente', 'facture_achat').order_by('article__reference_article', 'date_mouvement')
        
        # Filtre par type de mouvement
        if type_mouvement and type_mouvement != '':
            mouvements_query = mouvements_query.filter(type_mouvement=type_mouvement)
        
        mouvements = mouvements_query
        
        print(f"[PACKAGE] MOUVEMENTS FILTRÉS: {mouvements.count()}")
        
        # Debug: Afficher quelques mouvements pour vérifier
        if mouvements.count() > 0:
            print("[LIST] PREMIERS MOUVEMENTS TROUVÉS:")
            for i, mvt in enumerate(mouvements[:3]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement}")
        else:
            print("[ERREUR] AUCUN MOUVEMENT TROUVÉ - Vérifions les mouvements existants:")
            tous_mouvements = MouvementStock.objects.filter(agence=agence, article__in=articles)
            print(f"[CHART] Total mouvements pour ces articles: {tous_mouvements.count()}")
            for i, mvt in enumerate(tous_mouvements[:3]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement}")
        
        # Calculer les statistiques
        total_mouvements = mouvements.count()
        mouvements_entree = mouvements.filter(type_mouvement='entree').count()
        mouvements_sortie = mouvements.filter(type_mouvement='sortie').count()
        
        # Calculer la valeur totale du stock permanent
        valeur_stock_permanent = mouvements.aggregate(
            total=Sum('stock_permanent')
        )['total'] or 0
        
        # Stocker les données dans la session pour l'export
        mouvements_data = []
        for mouvement in mouvements:
            # Déterminer le tiers
            tiers = ""
            if mouvement.fournisseur:
                tiers = f"Fournisseur: {mouvement.fournisseur.intitule}"
            elif mouvement.facture_vente:
                tiers = f"Client: {mouvement.facture_vente.client.intitule if mouvement.facture_vente.client else 'N/A'}"
            elif mouvement.facture_achat:
                tiers = f"Fournisseur: {mouvement.facture_achat.fournisseur.intitule if mouvement.facture_achat.fournisseur else 'N/A'}"
            
            # Debug pour vérifier les données d'article
            print(f"[SEARCH] DEBUG ARTICLE: {mouvement.article.reference_article} - {mouvement.article.designation}")
            
            mouvements_data.append({
                'date_mouvement': mouvement.date_mouvement.strftime('%Y-%m-%d %H:%M'),
                'type_mouvement': mouvement.type_mouvement,
                'type_mouvement_display': mouvement.get_type_mouvement_display(),
                'reference_article': str(mouvement.article.reference_article) if mouvement.article.reference_article else 'N/A',
                'designation': str(mouvement.article.designation) if mouvement.article.designation else 'N/A',
                'tiers': tiers,
                'stock_initial': mouvement.stock_initial,
                'quantite': mouvement.quantite,
                'solde': mouvement.solde,
                'cout_moyen_pondere': float(mouvement.cout_moyen_pondere),
                'stock_permanent': float(mouvement.stock_permanent),
                'numero_piece': mouvement.numero_piece,
                'commentaire': mouvement.commentaire or '',
            })
        
        print(f"[CHART] MOUVEMENTS GÉNÉRÉS:")
        print(f"  - Total mouvements: {total_mouvements}")
        print(f"  - Entrées: {mouvements_entree}")
        print(f"  - Sorties: {mouvements_sortie}")
        print(f"  - Valeur stock permanent: {valeur_stock_permanent}")
        
        # Stocker les mouvements dans la session pour l'export
        request.session['mouvements_stock'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'articles_count': len(articles),
            'mouvements_data': mouvements_data,
            'total_mouvements': total_mouvements,
            'mouvements_entree': mouvements_entree,
            'mouvements_sortie': mouvements_sortie,
            'valeur_stock_permanent': float(valeur_stock_permanent),
            'article_reference': ', '.join([article.reference_article for article in articles]),
            'article_designation': ', '.join([article.designation for article in articles]),
        }
        
        # Préparer les informations des articles
        articles_info = []
        for article in articles:
            article_mouvements = mouvements.filter(article=article)
            articles_info.append({
                'id': article.id,
                'reference': article.reference_article,
                'designation': article.designation,
                'stock_actuel': article.stock_actuel,
                'mouvements_count': article_mouvements.count()
            })
        
        return JsonResponse({
            'success': True,
            'total_mouvements': total_mouvements,
            'mouvements_entree': mouvements_entree,
            'mouvements_sortie': mouvements_sortie,
            'valeur_stock_permanent': float(valeur_stock_permanent),
            'articles_info': articles_info,
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'mouvements': mouvements_data,
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR CONSULTATION MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        error_details = f"Erreur: {str(e)}"
        if hasattr(e, '__class__'):
            error_details += f" (Type: {e.__class__.__name__})"
        return JsonResponse({'success': False, 'error': error_details})

@login_required
def export_mouvements_excel(request):
    """Vue pour exporter les mouvements de stock en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Mouvements Stock {article_reference}"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tête du document
        ws.merge_cells('A1:P1')
        ws['A1'] = f"FICHE DE STOCK - {article_reference} - {article_designation}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:P2')
        ws['A2'] = f"Période: du {date_debut} au {date_fin}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes des colonnes
        headers = ['Date', 'Type', '', 'N°', '', '', 'Tiers', '', '', '', '', '+/-', 'Quantités en stock', 'Solde', 'C.M.U.P.', 'Stock permanent']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données
        row = 5
        for mouvement in mouvements_data_list:
            # Structure modifiée (16 colonnes - sans Référence/Désignation/unitaire, avec Tiers)
            ws.cell(row=row, column=1, value=mouvement['date_mouvement'])  # Date
            ws.cell(row=row, column=2, value=mouvement['type_mouvement'])  # Type
            ws.cell(row=row, column=3, value='')  # Colonne vide
            ws.cell(row=row, column=4, value=mouvement['numero_piece'])  # N°
            ws.cell(row=row, column=5, value='')  # Colonne vide
            ws.cell(row=row, column=6, value='')  # Colonne vide
            ws.cell(row=row, column=7, value=mouvement['tiers'])  # Tiers
            ws.cell(row=row, column=8, value='')  # Colonne vide
            ws.cell(row=row, column=9, value='')  # Colonne vide
            ws.cell(row=row, column=10, value='')  # Colonne vide
            ws.cell(row=row, column=11, value='')  # Colonne vide
            ws.cell(row=row, column=12, value=f"+{mouvement['quantite']}" if mouvement['quantite'] > 0 else mouvement['quantite'])  # +/-
            ws.cell(row=row, column=13, value=mouvement['stock_initial'])  # Quantités en stock
            ws.cell(row=row, column=14, value=mouvement['solde'])  # Solde
            ws.cell(row=row, column=15, value=mouvement['cout_moyen_pondere'])  # C.M.U.P.
            ws.cell(row=row, column=16, value=mouvement['stock_permanent'])  # Stock permanent
            row += 1
        
        # Ligne des totaux
        row += 1
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"Entrées: {mouvements_entree}, Sorties: {mouvements_sortie}").font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"Total: {total_mouvements} mouvements").font = Font(bold=True)
        ws.cell(row=row, column=4, value="").font = Font(bold=True)
        ws.cell(row=row, column=5, value="").font = Font(bold=True)
        ws.cell(row=row, column=6, value="").font = Font(bold=True)
        ws.cell(row=row, column=16, value=valeur_stock_permanent).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 15
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le fichier dans la réponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Article: {article_reference}")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - {total_mouvements} mouvements")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_mouvements_pdf(request):
    """Vue pour exporter les mouvements de stock en PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            textColor=colors.darkblue
        )
        
        # Éléments du document
        elements = []
        
        # Titre
        title = Paragraph(f"FICHE DE STOCK - {article_reference}", title_style)
        elements.append(title)
        
        # Informations de l'article et période
        info_text = f"<b>Article:</b> {article_designation}<br/><b>Période:</b> du {date_debut} au {date_fin}"
        info_para = Paragraph(info_text, styles['Normal'])
        elements.append(info_para)
        elements.append(Spacer(1, 20))
        
        # Données du tableau
        data = [['Date', 'Type', 'Tiers', 'Stock Init.', 'Solde', 'C.M.PU', 'Stock Perm.', 'N° Pièce']]
        
        for mouvement in mouvements_data_list:
            data.append([
                mouvement['date_mouvement'],
                mouvement['type_mouvement'],
                mouvement['tiers'][:20] + '...' if len(mouvement['tiers']) > 20 else mouvement['tiers'],
                str(mouvement['stock_initial']),
                str(mouvement['solde']),
                f"{mouvement['cout_moyen_pondere']:,.0f}",
                f"{mouvement['stock_permanent']:,.0f}",
                mouvement['numero_piece']
            ])
        
        # Ligne des totaux
        data.append([
            'TOTAL',
            f'E:{mouvements_entree} S:{mouvements_sortie}',
            f'{total_mouvements} mouvements',
            '',
            '',
            '',
            f"{valeur_stock_permanent:,.0f} FCFA",
            ''
        ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 1.5*cm, 3*cm, 1.5*cm, 1.5*cm, 2*cm, 2.5*cm, 2*cm])
        
        # Style du tableau
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF MOUVEMENTS - Article: {article_reference}")
        print(f"📄 EXPORT PDF MOUVEMENTS - {total_mouvements} mouvements")
        print(f"📄 EXPORT PDF MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_mouvements_csv(request):
    """Vue pour exporter les mouvements de stock en CSV"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Date', 'Type', '', 'N°', '', '', 'Tiers', '', '', '', '', '+/-', 'Quantités en stock', 'Solde', 'C.M.U.P.', 'Stock permanent']
        writer.writerow(headers)
        
        # Données
        for mouvement in mouvements_data_list:
            row = [
                mouvement['date_mouvement'],
                mouvement['type_mouvement'],
                mouvement['tiers'],
                mouvement['stock_initial'],
                mouvement['solde'],
                mouvement['cout_moyen_pondere'],
                mouvement['stock_permanent'],
                mouvement['numero_piece']
            ]
            writer.writerow(row)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['TOTAL GÉNÉRAL', f'Entrées: {mouvements_entree}, Sorties: {mouvements_sortie}', f'{total_mouvements} mouvements', '', '', '', valeur_stock_permanent, ''])
        
        print(f"[CHART] EXPORT CSV MOUVEMENTS - Article: {article_reference}")
        print(f"[CHART] EXPORT CSV MOUVEMENTS - {total_mouvements} mouvements")
        print(f"[CHART] EXPORT CSV MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def creer_mouvements_retroactifs(request):
    """Vue simplifiée pour créer des mouvements de stock rétroactifs"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[REFRESH] CRÉATION MOUVEMENTS RÉTROACTIFS (VERSION SIMPLIFIÉE)...")
        print(f"[TARGET] Agence utilisée: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # 1. Créer des mouvements pour les factures de vente
        factures_vente = FactureVente.objects.filter(agence=agence)
        print(f"[CHART] Factures de vente: {factures_vente.count()}")
        
        for facture in factures_vente:
            lignes = LigneFactureVente.objects.filter(facture_vente=facture)
            print(f"[SEARCH] Facture {facture.numero_ticket}: {lignes.count()} lignes")
            
            for ligne in lignes:
                print(f"  [PACKAGE] Article: {ligne.article.designation}, Quantité: {ligne.quantite}")
                
                # Vérifier si le mouvement existe déjà
                mouvement_existe = MouvementStock.objects.filter(facture_vente=facture, article=ligne.article).exists()
                print(f"  [SEARCH] Mouvement existe déjà: {mouvement_existe}")
                
                if not mouvement_existe:
                    try:
                        from datetime import datetime
                        from django.utils import timezone
                        
                        # Utiliser timezone.now() pour la date
                        date_mouvement = timezone.now()
                        
                        MouvementStock.objects.create(
                            article=ligne.article,
                            agence=agence,
                            type_mouvement='sortie',
                            date_mouvement=date_mouvement,
                            numero_piece=facture.numero_ticket,
                            quantite_stock=ligne.article.stock_actuel,
                            stock_initial=ligne.article.stock_actuel + ligne.quantite,
                            solde=ligne.article.stock_actuel,
                            quantite=ligne.quantite,
                            cout_moyen_pondere=float(ligne.prix_unitaire),
                            stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                            facture_vente=facture,
                            commentaire=f"Vente - {facture.numero_ticket}"
                        )
                        mouvements_crees += 1
                        print(f"[OK] Vente: {ligne.article.designation} - {facture.numero_ticket}")
                    except Exception as e:
                        print(f"[ERREUR] Erreur vente {facture.numero_ticket}: {e}")
        
        # 2. Créer des mouvements pour les factures d'achat
        factures_achat = FactureAchat.objects.filter(agence=agence)
        print(f"[CHART] Factures d'achat: {factures_achat.count()}")
        
        for facture in factures_achat:
            lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
            print(f"[SEARCH] Facture achat {facture.reference_achat}: {lignes.count()} lignes")
            
            for ligne in lignes:
                print(f"  [PACKAGE] Article: {ligne.article.designation}, Quantité: {ligne.quantite}")
                
                mouvement_existe = MouvementStock.objects.filter(facture_achat=facture, article=ligne.article).exists()
                print(f"  [SEARCH] Mouvement existe déjà: {mouvement_existe}")
                
                if not mouvement_existe:
                    try:
                        from django.utils import timezone
                        
                        MouvementStock.objects.create(
                            article=ligne.article,
                            agence=agence,
                            type_mouvement='entree',
                            date_mouvement=timezone.now(),
                            numero_piece=facture.reference_achat,
                            quantite_stock=ligne.article.stock_actuel,
                            stock_initial=ligne.article.stock_actuel - ligne.quantite,
                            solde=ligne.article.stock_actuel,
                            quantite=ligne.quantite,
                            cout_moyen_pondere=float(ligne.prix_unitaire),
                            stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                            facture_achat=facture,
                            fournisseur=facture.fournisseur,
                            commentaire=f"Achat - {facture.reference_achat}"
                        )
                        mouvements_crees += 1
                        print(f"[OK] Achat: {ligne.article.designation} - {facture.reference_achat}")
                    except Exception as e:
                        print(f"[ERREUR] Erreur achat {facture.reference_achat}: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'{mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def diagnostic_mouvements(request):
    """Vue de diagnostic pour les mouvements de stock"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Compter les données
        articles_count = Article.objects.filter(agence=agence).count()
        factures_vente_count = FactureVente.objects.filter(agence=agence).count()
        factures_achat_count = FactureAchat.objects.filter(agence=agence).count()
        factures_transfert_count = FactureTransfert.objects.filter(agence_source=agence).count()
        mouvements_count = MouvementStock.objects.filter(agence=agence).count()
        
        # Détails des factures
        factures_vente_details = []
        for facture in FactureVente.objects.filter(agence=agence)[:5]:
            lignes_count = LigneFactureVente.objects.filter(facture_vente=facture).count()
            factures_vente_details.append({
                'numero': facture.numero_ticket,
                'date': str(facture.date),
                'lignes': lignes_count
            })
        
        factures_achat_details = []
        for facture in FactureAchat.objects.filter(agence=agence)[:5]:
            lignes_count = LigneFactureAchat.objects.filter(facture_achat=facture).count()
            factures_achat_details.append({
                'numero': facture.reference_achat,
                'date': str(facture.date_achat),
                'lignes': lignes_count
            })
        
        return JsonResponse({
            'success': True,
            'agence': agence.nom_agence,
            'articles_count': articles_count,
            'factures_vente_count': factures_vente_count,
            'factures_achat_count': factures_achat_count,
            'factures_transfert_count': factures_transfert_count,
            'mouvements_count': mouvements_count,
            'factures_vente_details': factures_vente_details,
            'factures_achat_details': factures_achat_details
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def forcer_mouvements(request):
    """Vue pour forcer la création de mouvements même s'ils existent déjà"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[HOT] CRÉATION FORCÉE DE MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # Supprimer tous les mouvements existants d'abord
        anciens_mouvements = MouvementStock.objects.filter(agence=agence).count()
        MouvementStock.objects.filter(agence=agence).delete()
        print(f"🗑️ {anciens_mouvements} anciens mouvements supprimés")
        
        # Test de création d'un mouvement simple
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles trouvés: {articles.count()}")
        
        if articles.exists():
            article_test = articles.first()
            print(f"🧪 Test avec article: {article_test.designation}")
            
            try:
                from django.utils import timezone
                
                mouvement_test = MouvementStock.objects.create(
                    article=article_test,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece='TEST-001',
                    quantite_stock=article_test.stock_actuel,
                    stock_initial=0,
                    solde=article_test.stock_actuel,
                    quantite=1,
                    cout_moyen_pondere=float(article_test.prix_achat),
                    stock_permanent=float(article_test.stock_actuel * article_test.prix_achat),
                    commentaire='Test de création'
                )
                print(f"[OK] MOUVEMENT TEST CRÉÉ AVEC SUCCÈS: ID {mouvement_test.id}")
                mouvements_crees += 1
                
                # Supprimer le test
                mouvement_test.delete()
                print(f"🗑️ Mouvement test supprimé")
                
            except Exception as e:
                print(f"[ERREUR] ERREUR LORS DU TEST: {e}")
                import traceback
                traceback.print_exc()
                return JsonResponse({'success': False, 'error': f'Erreur lors du test de création: {str(e)}'})
        
        print(f"[OK] Test terminé, création des vrais mouvements...")
        
        # Créer des mouvements pour les factures de vente (version simplifiée)
        factures_vente = FactureVente.objects.filter(agence=agence)
        print(f"[CHART] Factures de vente: {factures_vente.count()}")
        
        for facture in factures_vente:
            lignes = LigneFactureVente.objects.filter(facture_vente=facture)
            print(f"  [SEARCH] {facture.numero_ticket}: {lignes.count()} lignes")
            
            for ligne in lignes:
                try:
                    from django.utils import timezone
                    
                    # Utiliser la même logique simple que le test qui fonctionne
                    MouvementStock.objects.create(
                        article=ligne.article,
                        agence=agence,
                        type_mouvement='sortie',
                        date_mouvement=timezone.now(),
                        numero_piece=facture.numero_ticket,
                        quantite_stock=ligne.article.stock_actuel,
                        stock_initial=0,  # Simplifié
                        solde=ligne.article.stock_actuel,
                        quantite=ligne.quantite,
                        cout_moyen_pondere=float(ligne.prix_unitaire),
                        stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                        facture_vente=facture,
                        commentaire=f"Vente - {facture.numero_ticket}"
                    )
                    mouvements_crees += 1
                    print(f"    [OK] {ligne.article.designation}")
                except Exception as e:
                    print(f"    [ERREUR] Erreur: {e}")
        
        # Créer des mouvements pour les factures d'achat (version simplifiée)
        factures_achat = FactureAchat.objects.filter(agence=agence)
        print(f"[CHART] Factures d'achat: {factures_achat.count()}")
        
        for facture in factures_achat:
            lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
            print(f"  [SEARCH] {facture.reference_achat}: {lignes.count()} lignes")
            
            for ligne in lignes:
                try:
                    from django.utils import timezone
                    
                    # Utiliser la même logique simple que le test qui fonctionne
                    MouvementStock.objects.create(
                        article=ligne.article,
                        agence=agence,
                        type_mouvement='entree',
                        date_mouvement=timezone.now(),
                        numero_piece=facture.reference_achat,
                        quantite_stock=ligne.article.stock_actuel,
                        stock_initial=0,  # Simplifié
                        solde=ligne.article.stock_actuel,
                        quantite=ligne.quantite,
                        cout_moyen_pondere=float(ligne.prix_unitaire),
                        stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                        facture_achat=facture,
                        fournisseur=facture.fournisseur,
                        commentaire=f"Achat - {facture.reference_achat}"
                    )
                    mouvements_crees += 1
                    print(f"    [OK] {ligne.article.designation}")
                except Exception as e:
                    print(f"    [ERREUR] Erreur: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'FORCÉ: {mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_mouvement_simple(request):
    """Test simple de création d'un mouvement"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("🧪 TEST SIMPLE DE CRÉATION DE MOUVEMENT...")
        
        # Vérifier les articles
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles: {articles.count()}")
        
        if not articles.exists():
            return JsonResponse({'success': False, 'error': 'Aucun article trouvé dans cette agence'})
        
        article = articles.first()
        print(f"[PACKAGE] Article test: {article.designation}")
        
        # Vérifier les champs obligatoires
        print(f"[PACKAGE] Stock actuel: {article.stock_actuel}")
        print(f"[PACKAGE] Prix achat: {article.prix_achat}")
        
        # Créer un mouvement simple
        from django.utils import timezone
        
        mouvement = MouvementStock.objects.create(
            article=article,
            agence=agence,
            type_mouvement='entree',
            date_mouvement=timezone.now(),
            numero_piece='TEST-SIMPLE',
            quantite_stock=article.stock_actuel,
            stock_initial=0,
            solde=article.stock_actuel,
            quantite=1,
            cout_moyen_pondere=float(article.prix_achat),
            stock_permanent=float(article.stock_actuel * article.prix_achat),
            commentaire='Test simple'
        )
        
        print(f"[OK] MOUVEMENT CRÉÉ: ID {mouvement.id}")
        
        # Vérifier qu'il existe
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        print(f"[CHART] Total mouvements: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'Test réussi! Mouvement ID {mouvement.id} créé. Total: {total_mouvements}',
            'mouvement_id': mouvement.id,
            'total_mouvements': total_mouvements
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR TEST: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def creer_mouvements_manuels(request):
    """Créer des mouvements manuels simples pour tester"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[TOOL] CRÉATION MANUELLE DE MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # Récupérer tous les articles
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles trouvés: {articles.count()}")
        
        if not articles.exists():
            return JsonResponse({'success': False, 'error': 'Aucun article trouvé'})
        
        # Créer un mouvement pour chaque article
        for article in articles:
            try:
                from django.utils import timezone
                
                MouvementStock.objects.create(
                    article=article,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece=f'MANUEL-{article.id}',
                    quantite_stock=article.stock_actuel,
                    stock_initial=0,
                    solde=article.stock_actuel,
                    quantite=article.stock_actuel,
                    cout_moyen_pondere=float(article.prix_achat),
                    stock_permanent=float(article.stock_actuel * article.prix_achat),
                    commentaire=f'Création manuelle - {article.designation}'
                )
                mouvements_crees += 1
                print(f"[OK] {article.designation}")
                
            except Exception as e:
                print(f"[ERREUR] Erreur pour {article.designation}: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements manuels créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'MANUEL: {mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_consultation_mouvements(request):
    """Test simple pour vérifier les mouvements existants"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[SEARCH] TEST CONSULTATION MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        # Vérifier tous les mouvements de l'agence
        tous_mouvements = MouvementStock.objects.filter(agence=agence)
        print(f"[CHART] Total mouvements dans l'agence: {tous_mouvements.count()}")
        
        if tous_mouvements.count() > 0:
            print("[LIST] PREMIERS MOUVEMENTS:")
            for i, mvt in enumerate(tous_mouvements[:5]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement} - {mvt.numero_piece}")
        
        # Vérifier les articles avec mouvements
        articles_avec_mouvements = Article.objects.filter(
            agence=agence,
            mouvementstock__isnull=False
        ).distinct()
        print(f"[PACKAGE] Articles avec mouvements: {articles_avec_mouvements.count()}")
        
        for article in articles_avec_mouvements[:3]:
            mouvements_article = MouvementStock.objects.filter(agence=agence, article=article)
            print(f"  - {article.designation}: {mouvements_article.count()} mouvements")
        
        return JsonResponse({
            'success': True,
            'message': f'Test terminé - {tous_mouvements.count()} mouvements trouvés',
            'total_mouvements': tous_mouvements.count(),
            'articles_avec_mouvements': articles_avec_mouvements.count()
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_mouvements_session(request):
    """Récupérer les données de mouvements depuis la session"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les données depuis la session
        mouvements_data = request.session.get('mouvements_data', [])
        article_info = request.session.get('article_info', {})
        
        print(f"[CHART] RÉCUPÉRATION SESSION:")
        print(f"  - Mouvements en session: {len(mouvements_data)}")
        print(f"  - Article info: {article_info}")
        
        return JsonResponse({
            'success': True,
            'mouvements': mouvements_data,
            'article_info': article_info
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def debug_session(request):
    """Debug simple pour voir le contenu de la session"""
    try:
        print("[SEARCH] DEBUG SESSION:")
        print(f"  - Clés de session: {list(request.session.keys())}")
        
        mouvements_stock = request.session.get('mouvements_stock', {})
        print(f"  - mouvements_stock: {list(mouvements_stock.keys()) if mouvements_stock else 'Aucun'}")
        
        if mouvements_stock:
            print(f"  - mouvements_data count: {len(mouvements_stock.get('mouvements_data', []))}")
            print(f"  - article_reference: {mouvements_stock.get('article_reference', 'N/A')}")
            print(f"  - article_designation: {mouvements_stock.get('article_designation', 'N/A')}")
        
        return JsonResponse({
            'success': True,
            'session_keys': list(request.session.keys()),
            'mouvements_stock_keys': list(mouvements_stock.keys()) if mouvements_stock else [],
            'mouvements_count': len(mouvements_stock.get('mouvements_data', [])) if mouvements_stock else 0
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def dashboard_stock(request):
    """Dashboard principal du module de gestion de stock"""
    try:
        # Récupérer l'agence de l'utilisateur
        agence = get_user_agence(request)
        if not agence:
            messages.error(request, 'Votre compte n\'est pas configuré correctement.')
            return redirect('logout_stock')

        # Calculer les KPIs
        total_articles = Article.objects.filter(agence=agence).count()
        articles_stock_faible = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=10
        ).count()
        articles_rupture = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=0
        ).count()
        
        # Valeur totale du stock
        articles_avec_prix = Article.objects.filter(
            agence=agence,
            prix_achat__isnull=False,
            stock_actuel__isnull=False
        ).exclude(prix_achat=0).exclude(stock_actuel=0)
        
        valeur_totale_stock = 0
        for article in articles_avec_prix:
            try:
                valeur_article = float(article.prix_achat) * float(article.stock_actuel)
                valeur_totale_stock += valeur_article
            except (ValueError, TypeError):
                continue
        
        # Mouvements récents
        mouvements_recents = MouvementStock.objects.filter(agence=agence).order_by('-date_mouvement')[:5]
        
        # Articles les plus vendus (simulation)
        articles_populaires = Article.objects.filter(agence=agence).order_by('-stock_actuel')[:5]
        
        # Alertes de stock
        alertes_stock = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=5
        ).order_by('stock_actuel')[:5]

        # Récupérer le nom de l'utilisateur
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except Compte.DoesNotExist:
            nom_utilisateur = request.user.username

        context = {
            'agence': agence,
            'nom_utilisateur': nom_utilisateur,
            'total_articles': total_articles,
            'articles_stock_faible': articles_stock_faible,
            'articles_rupture': articles_rupture,
            'valeur_stock': valeur_totale_stock,  # Corrigé le nom de la variable
            'mouvements_recents': mouvements_recents,
            'articles_populaires': articles_populaires,
            'alertes_stock': alertes_stock,
        }
        
        return render(request, 'supermarket/stock/dashboard_stock.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur lors du chargement du dashboard: {str(e)}')
        # Récupérer le nom de l'utilisateur même en cas d'erreur
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except:
            nom_utilisateur = request.user.username if request.user.is_authenticated else "Utilisateur"

        return render(request, 'supermarket/stock/dashboard_stock.html', {
            'agence': None,
            'nom_utilisateur': nom_utilisateur,
            'total_articles': 0,
            'articles_stock_faible': 0,
            'articles_rupture': 0,
            'valeur_stock': 0,
            'mouvements_recents': [],
            'articles_populaires': [],
            'alertes_stock': [],
        })

def login_stock(request):
    """Page de connexion pour la gestion de stock"""
    if request.user.is_authenticated:
        return redirect('dashboard_stock')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                # Vérifier que l'utilisateur a un compte lié à une agence
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        login(request, user)
                        # Stocker l'agence dans la session
                        request.session['agence_id'] = compte.agence.id_agence
                        request.session['agence_nom'] = compte.agence.nom_agence
                        messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                        return redirect('dashboard_stock')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Aucun compte actif trouvé pour cet utilisateur.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')
    
    return render(request, 'supermarket/stock/login.html')

@login_required
def logout_stock(request):
    """Vue de logout pour le module stock"""
    logout(request)
    return redirect('login_stock')

@login_required
def modifier_client(request, client_id):
    """Vue pour modifier un client existant"""
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([intitule, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_client', client_id=client_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_client', client_id=client_id)
            
            # Mettre à jour le client
            client.intitule = intitule
            client.adresse = adresse
            client.telephone = telephone
            client.email = email
            client.agence = agence
            client.save()
            
            messages.success(request, f'Client "{intitule}" modifié avec succès!')
            return redirect('consulter_clients')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du client: {str(e)}')
            return redirect('modifier_client', client_id=client_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    
    context = {
        'client': client,
        'agences': agences,
    }
    return render(request, 'supermarket/stock/modifier_client.html', context)

@login_required
def supprimer_client(request, client_id):
    """Vue pour supprimer un client"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_clients')
    
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
        client_name = client.intitule
        client.delete()
        messages.success(request, f'Client "{client_name}" supprimé avec succès!')
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_clients')

@login_required
def detail_client(request, client_id):
    """Vue pour afficher les détails d'un client"""
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
        
        # Récupérer les factures du client (si elles existent)
        factures = FactureVente.objects.filter(client=client).order_by('-date', '-heure')[:10]
        
        context = {
            'client': client,
            'factures': factures,
        }
        return render(request, 'supermarket/stock/detail_client.html', context)
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients')

# ==================== PLAN COMPTABLE ====================

@login_required
def consulter_plan_comptable(request):
    """Vue pour consulter le plan comptable"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    nature_filter = request.GET.get('nature_filter', '')
    
    # Construire la requête de base
    comptes = PlanComptable.objects.all()
    
    # Appliquer les filtres
    if search_query:
        comptes = comptes.filter(
            Q(intitule__icontains=search_query) |
            Q(compte__icontains=search_query) |
            Q(abrege__icontains=search_query)
        )
    
    if nature_filter:
        comptes = comptes.filter(nature_compte=nature_filter)
    
    # Trier par numéro
    comptes = comptes.order_by('numero')
    
    # Calculer les statistiques
    total_comptes = PlanComptable.objects.count()
    comptes_actifs = PlanComptable.objects.filter(actif=True).count()
    
    # Récupérer les natures de compte pour le filtre
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'comptes': comptes,
        'agence': agence,
        'total_comptes': total_comptes,
        'comptes_actifs': comptes_actifs,
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/consulter_plan_comptable.html', context)

@login_required
def creer_plan_comptable(request):
    """Vue pour créer un nouveau compte comptable"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero = request.POST.get('numero')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            abrege = request.POST.get('abrege')
            nature_compte = request.POST.get('nature_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([numero, intitule, compte]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_plan_comptable')
            
            # Créer le compte comptable
            PlanComptable.objects.create(
                numero=numero,
                intitule=intitule,
                compte=compte,
                abrege=abrege,
                nature_compte=nature_compte
            )
            
            messages.success(request, f'Compte comptable "{intitule}" créé avec succès!')
            return redirect('consulter_plan_comptable')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du compte: {str(e)}')
            return redirect('creer_plan_comptable')
    
    # GET - Afficher le formulaire
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/creer_plan_comptable.html', context)

@login_required
def modifier_plan_comptable(request, compte_id):
    """Vue pour modifier un compte comptable existant"""
    try:
        compte = PlanComptable.objects.get(id=compte_id)
    except PlanComptable.DoesNotExist:
        messages.error(request, 'Compte comptable non trouvé.')
        return redirect('consulter_plan_comptable')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero = request.POST.get('numero')
            intitule = request.POST.get('intitule')
            compte_field = request.POST.get('compte')
            abrege = request.POST.get('abrege')
            nature_compte = request.POST.get('nature_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([numero, intitule, compte_field]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_plan_comptable', compte_id=compte_id)
            
            # Mettre à jour le compte
            compte.numero = numero
            compte.intitule = intitule
            compte.compte = compte_field
            compte.abrege = abrege
            compte.nature_compte = nature_compte
            compte.save()
            
            messages.success(request, f'Compte comptable "{intitule}" modifié avec succès!')
            return redirect('consulter_plan_comptable')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du compte: {str(e)}')
            return redirect('modifier_plan_comptable', compte_id=compte_id)
    
    # GET - Afficher le formulaire pré-rempli
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'compte': compte,
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/modifier_plan_comptable.html', context)

@login_required
def supprimer_plan_comptable(request, compte_id):
    """Vue pour supprimer un compte comptable"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_plan_comptable')
    
    try:
        compte = PlanComptable.objects.get(id=compte_id)
        compte_name = compte.intitule
        compte.delete()
        messages.success(request, f'Compte comptable "{compte_name}" supprimé avec succès!')
    except PlanComptable.DoesNotExist:
        messages.error(request, 'Compte comptable non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_plan_comptable')

# ==================== PLAN TIERS ====================

@login_required
def consulter_plan_tiers(request):
    """Vue pour consulter le plan tiers"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    tiers = PlanTiers.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        tiers = tiers.filter(
            Q(intitule_compte__icontains=search_query) |
            Q(numero_compte__icontains=search_query)
        )
    
    if type_filter:
        tiers = tiers.filter(type=type_filter)
    
    # Trier par numéro de compte
    tiers = tiers.order_by('numero_compte')
    
    # Calculer les statistiques
    total_tiers = PlanTiers.objects.filter(agence=agence).count()
    clients_count = PlanTiers.objects.filter(agence=agence, type='client').count()
    fournisseurs_count = PlanTiers.objects.filter(agence=agence, type='fournisseur').count()
    
    # Récupérer les types pour le filtre
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'tiers': tiers,
        'agence': agence,
        'total_tiers': total_tiers,
        'clients_count': clients_count,
        'fournisseurs_count': fournisseurs_count,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_plan_tiers.html', context)

@login_required
def creer_plan_tiers(request):
    """Vue pour créer un nouveau tiers"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_tiers = request.POST.get('type')
            numero_compte = request.POST.get('numero_compte')
            intitule_compte = request.POST.get('intitule_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_tiers, numero_compte, intitule_compte, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_plan_tiers')
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('creer_plan_tiers')
            
            # Créer le tiers
            PlanTiers.objects.create(
                type=type_tiers,
                numero_compte=numero_compte,
                intitule_compte=intitule_compte,
                agence=agence
            )
            
            messages.success(request, f'Tiers "{intitule_compte}" créé avec succès!')
            return redirect('consulter_plan_tiers')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du tiers: {str(e)}')
            return redirect('creer_plan_tiers')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'agences': agences,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/creer_plan_tiers.html', context)

@login_required
def modifier_plan_tiers(request, tiers_id):
    """Vue pour modifier un tiers existant"""
    try:
        tiers = PlanTiers.objects.get(id=tiers_id, agence=get_user_agence(request))
    except PlanTiers.DoesNotExist:
        messages.error(request, 'Tiers non trouvé.')
        return redirect('consulter_plan_tiers')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_tiers = request.POST.get('type')
            numero_compte = request.POST.get('numero_compte')
            intitule_compte = request.POST.get('intitule_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_tiers, numero_compte, intitule_compte, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_plan_tiers', tiers_id=tiers_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_plan_tiers', tiers_id=tiers_id)
            
            # Mettre à jour le tiers
            tiers.type = type_tiers
            tiers.numero_compte = numero_compte
            tiers.intitule_compte = intitule_compte
            tiers.agence = agence
            tiers.save()
            
            messages.success(request, f'Tiers "{intitule_compte}" modifié avec succès!')
            return redirect('consulter_plan_tiers')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du tiers: {str(e)}')
            return redirect('modifier_plan_tiers', tiers_id=tiers_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'tiers': tiers,
        'agences': agences,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/modifier_plan_tiers.html', context)

@login_required
def supprimer_plan_tiers(request, tiers_id):
    """Vue pour supprimer un tiers"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_plan_tiers')
    
    try:
        tiers = PlanTiers.objects.get(id=tiers_id, agence=get_user_agence(request))
        tiers_name = tiers.intitule_compte
        tiers.delete()
        messages.success(request, f'Tiers "{tiers_name}" supprimé avec succès!')
    except PlanTiers.DoesNotExist:
        messages.error(request, 'Tiers non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_plan_tiers')

# ==================== CODE JOURNAUX ====================

@login_required
def consulter_code_journaux(request):
    """Vue pour consulter les codes journaux"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    journaux = CodeJournaux.objects.all()
    
    # Appliquer les filtres
    if search_query:
        journaux = journaux.filter(
            Q(intitule__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    if type_filter:
        journaux = journaux.filter(type_document=type_filter)
    
    # Trier par code
    journaux = journaux.order_by('code')
    
    # Calculer les statistiques
    total_journaux = CodeJournaux.objects.count()
    journaux_achat = CodeJournaux.objects.filter(type_document='document_achat').count()
    journaux_vente = CodeJournaux.objects.filter(type_document='caisse').count()
    
    # Récupérer les types pour le filtre
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'journaux': journaux,
        'agence': agence,
        'total_journaux': total_journaux,
        'journaux_achat': journaux_achat,
        'journaux_vente': journaux_vente,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_code_journaux.html', context)

@login_required
def creer_code_journaux(request):
    """Vue pour créer un nouveau code journal"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_document = request.POST.get('type_document')
            intitule = request.POST.get('intitule')
            code = request.POST.get('code')
            compte_contrepartie = request.POST.get('compte_contrepartie')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_document, intitule, code]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_code_journaux')
            
            # Créer le code journal
            CodeJournaux.objects.create(
                type_document=type_document,
                intitule=intitule,
                code=code,
                compte_contrepartie=compte_contrepartie
            )
            
            messages.success(request, f'Code journal "{intitule}" créé avec succès!')
            return redirect('consulter_code_journaux')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du code journal: {str(e)}')
            return redirect('creer_code_journaux')
    
    # GET - Afficher le formulaire
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/creer_code_journaux.html', context)

@login_required
def modifier_code_journaux(request, journal_id):
    """Vue pour modifier un code journal existant"""
    try:
        journal = CodeJournaux.objects.get(id=journal_id)
    except CodeJournaux.DoesNotExist:
        messages.error(request, 'Code journal non trouvé.')
        return redirect('consulter_code_journaux')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_document = request.POST.get('type_document')
            intitule = request.POST.get('intitule')
            code = request.POST.get('code')
            compte_contrepartie = request.POST.get('compte_contrepartie')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_document, intitule, code]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_code_journaux', journal_id=journal_id)
            
            # Mettre à jour le code journal
            journal.type_document = type_document
            journal.intitule = intitule
            journal.code = code
            journal.compte_contrepartie = compte_contrepartie
            journal.save()
            
            messages.success(request, f'Code journal "{intitule}" modifié avec succès!')
            return redirect('consulter_code_journaux')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du code journal: {str(e)}')
            return redirect('modifier_code_journaux', journal_id=journal_id)
    
    # GET - Afficher le formulaire pré-rempli
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'journal': journal,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/modifier_code_journaux.html', context)

@login_required
def supprimer_code_journaux(request, journal_id):
    """Vue pour supprimer un code journal"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_code_journaux')
    
    try:
        journal = CodeJournaux.objects.get(id=journal_id)
        journal_name = journal.intitule
        journal.delete()
        messages.success(request, f'Code journal "{journal_name}" supprimé avec succès!')
    except CodeJournaux.DoesNotExist:
        messages.error(request, 'Code journal non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_code_journaux')

# ==================== TAUX TAXE ====================

@login_required
def consulter_taux_taxe(request):
    """Vue pour consulter les taux de taxe"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    taux = TauxTaxe.objects.all()
    
    # Appliquer les filtres
    if search_query:
        taux = taux.filter(
            Q(intitule__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    if type_filter:
        taux = taux.filter(type=type_filter)
    
    # Trier par code
    taux = taux.order_by('code')
    
    # Calculer les statistiques
    total_taux = TauxTaxe.objects.count()
    taux_actifs = TauxTaxe.objects.filter(actif=True).count()
    
    # Récupérer les types pour le filtre
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    
    context = {
        'taux': taux,
        'agence': agence,
        'total_taux': total_taux,
        'taux_actifs': taux_actifs,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_taux_taxe.html', context)

@login_required
def creer_taux_taxe(request):
    """Vue pour créer un nouveau taux de taxe"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            code = request.POST.get('code')
            sens = request.POST.get('sens')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            taux = request.POST.get('taux')
            type_taxe = request.POST.get('type')
            assujettissement = request.POST.get('assujettissement')
            code_regroupement = request.POST.get('code_regroupement')
            provenance = request.POST.get('provenance')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([code, intitule, taux]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_taux_taxe')
            
            # Créer le taux de taxe
            TauxTaxe.objects.create(
                code=code,
                sens=sens,
                intitule=intitule,
                compte=compte,
                taux=float(taux),
                type=type_taxe,
                assujettissement=assujettissement,
                code_regroupement=code_regroupement,
                provenance=provenance
            )
            
            messages.success(request, f'Taux de taxe "{intitule}" créé avec succès!')
            return redirect('consulter_taux_taxe')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du taux de taxe: {str(e)}')
            return redirect('creer_taux_taxe')
    
    # GET - Afficher le formulaire
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    sens_choices = TauxTaxe.SENS_CHOICES
    assujettissement_choices = TauxTaxe.ASSUJETTISSEMENT_CHOICES
    
    context = {
        'type_choices': type_choices,
        'sens_choices': sens_choices,
        'assujettissement_choices': assujettissement_choices,
    }
    return render(request, 'supermarket/stock/creer_taux_taxe.html', context)

@login_required
def modifier_taux_taxe(request, taux_id):
    """Vue pour modifier un taux de taxe existant"""
    try:
        taux = TauxTaxe.objects.get(id=taux_id)
    except TauxTaxe.DoesNotExist:
        messages.error(request, 'Taux de taxe non trouvé.')
        return redirect('consulter_taux_taxe')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            code = request.POST.get('code')
            sens = request.POST.get('sens')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            taux_value = request.POST.get('taux')
            type_taxe = request.POST.get('type')
            assujettissement = request.POST.get('assujettissement')
            code_regroupement = request.POST.get('code_regroupement')
            provenance = request.POST.get('provenance')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([code, intitule, taux_value]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_taux_taxe', taux_id=taux_id)
            
            # Mettre à jour le taux de taxe
            taux.code = code
            taux.sens = sens
            taux.intitule = intitule
            taux.compte = compte
            taux.taux = float(taux_value)
            taux.type = type_taxe
            taux.assujettissement = assujettissement
            taux.code_regroupement = code_regroupement
            taux.provenance = provenance
            taux.save()
            
            messages.success(request, f'Taux de taxe "{intitule}" modifié avec succès!')
            return redirect('consulter_taux_taxe')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du taux de taxe: {str(e)}')
            return redirect('modifier_taux_taxe', taux_id=taux_id)
    
    # GET - Afficher le formulaire pré-rempli
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    sens_choices = TauxTaxe.SENS_CHOICES
    assujettissement_choices = TauxTaxe.ASSUJETTISSEMENT_CHOICES
    
    context = {
        'taux': taux,
        'type_choices': type_choices,
        'sens_choices': sens_choices,
        'assujettissement_choices': assujettissement_choices,
    }
    return render(request, 'supermarket/stock/modifier_taux_taxe.html', context)

@login_required
def supprimer_taux_taxe(request, taux_id):
    """Vue pour supprimer un taux de taxe"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_taux_taxe')
    
    try:
        taux = TauxTaxe.objects.get(id=taux_id)
        taux_name = taux.intitule
        taux.delete()
        messages.success(request, f'Taux de taxe "{taux_name}" supprimé avec succès!')
    except TauxTaxe.DoesNotExist:
        messages.error(request, 'Taux de taxe non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_taux_taxe')

@login_required
def modifier_article(request, article_id):
    """Vue pour modifier un article existant"""
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
        return redirect('consulter_articles')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            designation = request.POST.get('designation')
            agence_id = request.POST.get('agence')
            prix_achat = request.POST.get('prix_achat')
            prix_vente = request.POST.get('prix_vente')
            stock_actuel = request.POST.get('stock_actuel')
            stock_minimum = request.POST.get('stock_minimum', 0)
            unite_vente = request.POST.get('unite_vente')
            conditionnement = request.POST.get('conditionnement')
            famille_id = request.POST.get('famille')
            
            # Validation
            if not all([designation, agence_id, prix_achat, prix_vente, stock_actuel, unite_vente]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_article', article_id=article_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_article', article_id=article_id)
            
            # Récupérer la famille si spécifiée
            if famille_id:
                try:
                    categorie = Famille.objects.get(id=famille_id)
                except Famille.DoesNotExist:
                    messages.error(request, 'Famille non trouvée.')
                    return redirect('modifier_article', article_id=article_id)
            else:
                # Si aucune famille n'est fournie, garder l'ancienne
                categorie = article.categorie
            
            # Mettre à jour l'article
            article.designation = designation
            article.agence = agence
            article.prix_achat = float(prix_achat)
            article.prix_vente = float(prix_vente)
            article.stock_actuel = float(stock_actuel)
            article.stock_minimum = float(stock_minimum) if stock_minimum else 0
            article.unite_vente = unite_vente
            article.conditionnement = conditionnement
            article.categorie = categorie
            article.save()
            
            # Mettre à jour les types de vente
            prix_gros = request.POST.get('prix_gros')
            prix_demi_gros = request.POST.get('prix_demi_gros')
            prix_detail = request.POST.get('prix_detail')
            
            if prix_gros:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Gros',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_gros)}
                )
            
            if prix_demi_gros:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Demi-Gros',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_demi_gros)}
                )
            
            if prix_detail:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Détail',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_detail)}
                )
            
            messages.success(request, f'Article "{designation}" modifié avec succès!')
            return redirect('consulter_articles')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de l\'article: {str(e)}')
            return redirect('modifier_article', article_id=article_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    familles = Famille.objects.all()
    types_vente = TypeVente.objects.filter(article=article)
    
    # Créer un dictionnaire des types de vente avec des clés sans traits d'union
    types_vente_dict = {}
    for tv in types_vente:
        if tv.intitule == 'Demi-Gros':
            types_vente_dict['Demi_Gros'] = tv.prix
        elif tv.intitule == 'Détail':
            types_vente_dict['Détail'] = tv.prix
        else:
            types_vente_dict[tv.intitule] = tv.prix
    
    context = {
        'article': article,
        'agences': agences,
        'familles': familles,
        'types_vente': types_vente_dict
    }
    return render(request, 'supermarket/stock/modifier_article.html', context)

@login_required
def supprimer_article(request, article_id):
    """Vue pour supprimer un article"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_articles')
    
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
        article_name = article.designation
        article.delete()
        messages.success(request, f'Article "{article_name}" supprimé avec succès!')
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_articles')

@login_required
def detail_article(request, article_id):
    """Vue pour afficher les détails d'un article"""
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
        types_vente = TypeVente.objects.filter(article=article)
        mouvements = MouvementStock.objects.filter(article=article).order_by('-date_mouvement')[:10]
        
        # Calculer les marges
        marge_unitaire = float(article.prix_vente) - float(article.prix_achat) if article.prix_achat > 0 else 0
        marge_pourcentage = (marge_unitaire / float(article.prix_achat) * 100) if article.prix_achat > 0 else 0
        valeur_stock = float(article.prix_achat) * float(article.stock_actuel)
        
        # Calculer les marges pour chaque type de vente
        types_vente_with_marges = []
        for tv in types_vente:
            marge_tv = float(tv.prix) - float(article.prix_achat) if article.prix_achat > 0 else 0
            marge_tv_pourcentage = (marge_tv / float(article.prix_achat) * 100) if article.prix_achat > 0 else 0
            types_vente_with_marges.append({
                'type_vente': tv,
                'marge': marge_tv,
                'marge_pourcentage': marge_tv_pourcentage
            })
        
        # Debug: Vérifier la famille de l'article
        print(f"[ALERTE] DEBUG Article {article.id}:")
        print(f"   - Désignation: {article.designation}")
        print(f"   - Catégorie: {article.categorie}")
        print(f"   - Intitulé famille: {article.categorie.intitule if article.categorie else 'None'}")
        
        context = {
            'article': article,
            'types_vente': types_vente,
            'types_vente_with_marges': types_vente_with_marges,
            'mouvements': mouvements,
            'marge_unitaire': marge_unitaire,
            'marge_pourcentage': marge_pourcentage,
            'valeur_stock': valeur_stock
        }
        return render(request, 'supermarket/stock/detail_article.html', context)
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
        return redirect('consulter_articles')

# ==================== FACTURES D'ACHAT ====================

@login_required
def consulter_factures_achat(request):
    """Vue pour consulter les factures d'achat"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    statut_filter = request.GET.get('statut_filter', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Construire la requête de base
    factures = FactureAchat.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        factures = factures.filter(
            Q(numero_fournisseur__icontains=search_query) |
            Q(reference_achat__icontains=search_query) |
            Q(commentaire__icontains=search_query)
        )
    
    if statut_filter:
        factures = factures.filter(statut=statut_filter)
    
    if date_debut:
        factures = factures.filter(date_achat__gte=date_debut)
    
    if date_fin:
        factures = factures.filter(date_achat__lte=date_fin)
    
    # Trier par date décroissante
    factures = factures.order_by('-date_achat', '-heure')
    
    # Calculer les statistiques
    total_factures = FactureAchat.objects.filter(agence=agence).count()
    factures_validees = FactureAchat.objects.filter(agence=agence, statut='validee').count()
    factures_payees = FactureAchat.objects.filter(agence=agence, statut='payee').count()
    montant_total = FactureAchat.objects.filter(agence=agence).aggregate(
        total=Sum('prix_total_global')
    )['total'] or 0
    
    # Récupérer les statuts pour le filtre
    statut_choices = FactureAchat.STATUT_CHOICES
    
    context = {
        'factures': factures,
        'search_query': search_query,
        'statut_filter': statut_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statut_choices': statut_choices,
        'total_factures': total_factures,
        'factures_validees': factures_validees,
        'factures_payees': factures_payees,
        'montant_total': montant_total,
    }
    return render(request, 'supermarket/stock/consulter_factures_achat.html', context)


@login_required
def consulter_factures_transfert(request):
    """Vue pour consulter les factures de transfert"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    statut_filter = request.GET.get('statut_filter', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Construire la requête de base (factures où l'agence est source ou destination)
    factures = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    )
    
    # Appliquer les filtres
    if search_query:
        factures = factures.filter(
            Q(numero_compte__icontains=search_query) |
            Q(reference_transfert__icontains=search_query) |
            Q(lieu_depart__icontains=search_query) |
            Q(lieu_arrivee__icontains=search_query)
        )
    
    if statut_filter:
        factures = factures.filter(statut=statut_filter)
    
    if date_debut:
        factures = factures.filter(date_transfert__gte=date_debut)
    
    if date_fin:
        factures = factures.filter(date_transfert__lte=date_fin)
    
    # Trier par date décroissante
    factures = factures.order_by('-date_transfert')
    
    # Calculer les statistiques
    total_factures = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    ).count()
    factures_en_cours = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence),
        statut='en_cours'
    ).count()
    factures_terminees = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence),
        statut='termine'
    ).count()
    quantite_totale = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    ).aggregate(
        total=Sum('quantite')
    )['total'] or 0
    
    # Récupérer les statuts pour le filtre
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'factures': factures,
        'search_query': search_query,
        'statut_filter': statut_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statut_choices': statut_choices,
        'total_factures': total_factures,
        'factures_en_cours': factures_en_cours,
        'factures_terminees': factures_terminees,
        'quantite_totale': quantite_totale,
    }
    return render(request, 'supermarket/stock/consulter_factures_transfert.html', context)

@login_required
def creer_facture_transfert(request):
    """Vue pour créer une nouvelle facture de transfert"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_compte = request.POST.get('numero_compte')
            date_transfert = request.POST.get('date_transfert')
            reference_transfert = request.POST.get('reference_transfert')
            lieu_depart = request.POST.get('lieu_depart')
            lieu_arrivee = request.POST.get('lieu_arrivee')
            quantite = request.POST.get('quantite')
            statut = request.POST.get('statut')
            employe_expediteur = request.POST.get('employe_expediteur', '').strip()
            employe_destinataire = request.POST.get('employe_destinataire', '').strip()
            etat = request.POST.get('etat', 'sortir')
            
            # Validation
            if not all([numero_compte, date_transfert, reference_transfert, lieu_depart, lieu_arrivee, quantite, employe_expediteur]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_facture_transfert')
            
            # Rechercher les employés existants par nom
            expediteur_employe = None
            destinataire_employe = None
            
            # Rechercher l'employé expéditeur
            try:
                # Essayer de trouver par nom complet dans le compte
                expediteur_employe = Employe.objects.filter(
                    compte__agence=agence,
                    compte__user__first_name__icontains=employe_expediteur.split()[0] if employe_expediteur.split() else employe_expediteur
                ).first()
                
                # Si pas trouvé, essayer par nom de famille
                if not expediteur_employe and len(employe_expediteur.split()) > 1:
                    expediteur_employe = Employe.objects.filter(
                        compte__agence=agence,
                        compte__user__last_name__icontains=employe_expediteur.split()[-1]
                    ).first()
                
                # Si toujours pas trouvé, prendre le premier employé de l'agence
                if not expediteur_employe:
                    expediteur_employe = Employe.objects.filter(compte__agence=agence).first()
                    
            except Exception as e:
                print(f"Erreur lors de la recherche de l'employé expéditeur: {e}")
                expediteur_employe = Employe.objects.filter(compte__agence=agence).first()
            
            if not expediteur_employe:
                messages.error(request, 'Aucun employé trouvé dans cette agence.')
                return redirect('creer_facture_transfert')
            
            # Afficher un message informatif sur l'employé trouvé
            expediteur_nom = f"{expediteur_employe.compte.user.first_name} {expediteur_employe.compte.user.last_name}".strip()
            messages.info(request, f'Employé expéditeur trouvé: {expediteur_nom}')
            
            if destinataire_employe:
                destinataire_nom = f"{destinataire_employe.compte.user.first_name} {destinataire_employe.compte.user.last_name}".strip()
                messages.info(request, f'Employé destinataire trouvé: {destinataire_nom}')
            
            # Rechercher l'employé destinataire (si fourni)
            if employe_destinataire:
                try:
                    # Essayer de trouver par nom complet dans le compte
                    destinataire_employe = Employe.objects.filter(
                        compte__agence=agence,
                        compte__user__first_name__icontains=employe_destinataire.split()[0] if employe_destinataire.split() else employe_destinataire
                    ).first()
                    
                    # Si pas trouvé, essayer par nom de famille
                    if not destinataire_employe and len(employe_destinataire.split()) > 1:
                        destinataire_employe = Employe.objects.filter(
                            compte__agence=agence,
                            compte__user__last_name__icontains=employe_destinataire.split()[-1]
                        ).first()
                        
                except Exception as e:
                    print(f"Erreur lors de la recherche de l'employé destinataire: {e}")
                    destinataire_employe = None
            
            # Créer la facture de transfert
            facture = FactureTransfert.objects.create(
                numero_compte=numero_compte,
                date_transfert=date_transfert,
                reference_transfert=reference_transfert,
                lieu_depart=lieu_depart,
                lieu_arrivee=lieu_arrivee,
                quantite=int(quantite),
                statut=statut,
                agence_source=agence,
                agence_destination=agence,  # Pour l'instant, même agence (à modifier selon les besoins)
                employe_expediteur=expediteur_employe,
                employe_destinataire=destinataire_employe,
                etat=etat
            )
            
            # Traiter les articles sélectionnés
            articles_data = request.POST.get('articles_data', '')
            if articles_data:
                import json
                try:
                    articles = json.loads(articles_data)
                    for article_data in articles:
                        # Récupérer l'article
                        article = Article.objects.get(id=article_data['id'])
                        
                        # Créer la ligne de facture de transfert
                        LigneFactureTransfert.objects.create(
                            facture_transfert=facture,
                            article=article,
                            quantite=int(article_data['quantite']),
                            prix_unitaire=float(article_data['prix_achat']),
                            valeur_totale=float(article_data['prix_achat']) * int(article_data['quantite'])
                        )
                        
                        # Mettre à jour le stock de l'article (déduction pour transfert)
                        ancien_stock = article.stock_actuel
                        # Convertir en Decimal pour éviter les erreurs de type
                        from decimal import Decimal
                        quantite_decimal = Decimal(str(article_data['quantite']))
                        article.stock_actuel -= quantite_decimal
                        if article.stock_actuel < 0:
                            article.stock_actuel = 0
                        
                        # Mettre à jour le dernier prix d'achat avec le prix du transfert
                        ancien_dernier_prix = article.dernier_prix_achat
                        nouveau_prix_achat = float(article_data['prix_achat'])
                        article.dernier_prix_achat = nouveau_prix_achat
                        
                        article.save()
                        print(f"[PACKAGE] STOCK TRANSFERT - Article: {article.designation}")
                        print(f"[PACKAGE] STOCK TRANSFERT - Stock mis à jour: {ancien_stock} → {article.stock_actuel}")
                        print(f"[MONEY] Transfert - Dernier prix d'achat mis à jour: {ancien_dernier_prix} → {nouveau_prix_achat}")
                        
                        # [HOT] CRÉER UN MOUVEMENT DE STOCK POUR TRAÇABILITÉ
                        try:
                            MouvementStock.objects.create(
                                article=article,
                                agence=agence,
                                type_mouvement='sortie',
                                date_mouvement=timezone.now(),
                                numero_piece=facture.reference_transfert,
                                quantite_stock=article.stock_actuel,
                                stock_initial=ancien_stock,
                                solde=article.stock_actuel,
                                quantite=int(article_data['quantite']),
                                cout_moyen_pondere=float(article.prix_achat),
                                stock_permanent=float(article.stock_actuel * article.prix_achat),
                                commentaire=f"Transfert - Facture {facture.reference_transfert}"
                            )
                            print(f"[NOTE] MOUVEMENT STOCK - Sortie transfert enregistrée pour {article.designation}")
                        except Exception as e:
                            print(f"[WARNING] ERREUR MOUVEMENT STOCK TRANSFERT: {e}")
                        
                except (json.JSONDecodeError, Article.DoesNotExist, KeyError) as e:
                    print(f"Erreur lors du traitement des articles: {e}")
            
            messages.success(request, f'Facture de transfert "{reference_transfert}" créée avec succès!')
            return redirect('creer_facture_transfert')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de la facture: {str(e)}')
            return redirect('creer_facture_transfert')
    
    # GET - Afficher le formulaire
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'statut_choices': statut_choices,
    }
    return render(request, 'supermarket/stock/creer_facture_transfert_nouveau.html', context)

@login_required
def detail_facture_transfert(request, facture_id):
    """Vue pour afficher les détails d'une facture de transfert"""
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
        lignes = LigneFactureTransfert.objects.filter(facture_transfert=facture)
        
        context = {
            'facture': facture,
            'lignes': lignes,
        }
        return render(request, 'supermarket/stock/detail_facture_transfert.html', context)
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
        return redirect('consulter_factures_transfert')

@login_required
def modifier_facture_transfert(request, facture_id):
    """Vue pour modifier une facture de transfert existante"""
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
        return redirect('consulter_factures_transfert')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_compte = request.POST.get('numero_compte')
            date_transfert = request.POST.get('date_transfert')
            reference_transfert = request.POST.get('reference_transfert')
            lieu_depart = request.POST.get('lieu_depart')
            lieu_arrivee = request.POST.get('lieu_arrivee')
            quantite = request.POST.get('quantite')
            statut = request.POST.get('statut')
            employe_expediteur = request.POST.get('employe_expediteur', '').strip()
            employe_destinataire = request.POST.get('employe_destinataire', '').strip()
            etat = request.POST.get('etat', 'sortir')
            
            # Validation
            if not all([numero_compte, date_transfert, reference_transfert, lieu_depart, lieu_arrivee, quantite]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            # Mettre à jour la facture de transfert
            facture.numero_compte = numero_compte
            facture.date_transfert = date_transfert
            facture.reference_transfert = reference_transfert
            facture.lieu_depart = lieu_depart
            facture.lieu_arrivee = lieu_arrivee
            facture.quantite = int(quantite)
            facture.statut = statut
            facture.save()
            
            messages.success(request, f'Facture de transfert "{reference_transfert}" modifiée avec succès!')
            return redirect('consulter_factures_transfert')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de la facture: {str(e)}')
            return redirect('modifier_facture_transfert', facture_id=facture_id)
    
    # GET - Afficher le formulaire pré-rempli
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'facture': facture,
        'statut_choices': statut_choices,
    }
    return render(request, 'supermarket/stock/modifier_facture_transfert.html', context)

@login_required
def supprimer_facture_transfert(request, facture_id):
    """Vue pour supprimer une facture de transfert"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_factures_transfert')
    
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
        facture_name = facture.reference_transfert
        facture.delete()
        messages.success(request, f'Facture de transfert "{facture_name}" supprimée avec succès!')
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_factures_transfert')

# ===== RECHERCHE D'ARTICLES POUR STOCK =====

@login_required
def search_articles_stock(request):
    """Vue pour la recherche d'articles dans le module de stock"""
    search_term = request.GET.get('q', '')
    
    print(f"[SEARCH] search_articles_stock: recherche pour '{search_term}'")
    
    # Vérifier d'abord s'il y a des articles dans la base de données
    total_articles = Article.objects.count()
    print(f"[CHART] Total articles dans la base de données: {total_articles}")
    
    if total_articles == 0:
        print("[ERREUR] Aucun article dans la base de données!")
        return JsonResponse({'articles': []})
    
    agence = get_user_agence(request)
    print(f"[BUILDING] Agence trouvée: {agence}")
    print(f"[BUILDING] ID de l'agence: {agence.id_agence if agence else 'None'}")
    print(f"[BUILDING] Nom de l'agence: {agence.nom_agence if agence else 'None'}")
    
    if not agence:
        print("[ERREUR] Aucune agence trouvée")
        return JsonResponse({'articles': []})
    
    # Test: afficher tous les articles sans filtre d'agence
    articles_all = Article.objects.all()
    print(f"[PACKAGE] Tous les articles (toutes agences): {articles_all.count()}")
    for article in articles_all[:5]:  # Afficher les 5 premiers
        print(f"  - {article.designation} (Agence: {article.agence.nom_agence if article.agence else 'None'})")
    
    # Test: afficher les articles de cette agence spécifique
    articles_agence = Article.objects.filter(agence=agence)
    print(f"[PACKAGE] Articles de l'agence {agence.nom_agence}: {articles_agence.count()}")
    for article in articles_agence[:5]:  # Afficher les 5 premiers
        print(f"  - {article.designation} (ID: {article.id})")
    
    articles = []
    
    if search_term and len(search_term) >= 1:
        # Recherche avec filtre d'agence dès qu'on tape 1 lettre
        articles = Article.objects.filter(agence=agence, designation__icontains=search_term)[:50]
        print(f"[SEARCH] Articles trouvés avec recherche '{search_term}' (agence {agence.nom_agence}): {articles.count()}")
    else:
        # Afficher tous les articles de l'agence si pas de terme de recherche
        articles = Article.objects.filter(agence=agence)[:50]
        print(f"[PACKAGE] Tous les articles de l'agence {agence.nom_agence}: {articles.count()}")
    
    # Convertir les articles en format JSON
    articles_data = []
    for article in articles:
        articles_data.append({
            'id': article.id,
            'designation': article.designation,
            'prix_achat': float(article.prix_achat),
            'stock': article.stock_actuel,
            'reference_article': article.reference_article,
        })
        print(f"[NOTE] Article: {article.designation} (ID: {article.id})")
    
    print(f"[CHART] Total articles_data: {len(articles_data)}")
    return JsonResponse({'articles': articles_data})

def create_test_articles(request):
    """Vue temporaire pour créer des articles de test"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'error': 'Aucune agence trouvée'})
    
    # Créer quelques articles de test
    test_articles = [
        {'designation': 'Ordinateur Portable', 'prix_achat': 500000, 'prix_vente': 600000, 'stock_actuel': 10},
        {'designation': 'Souris USB', 'prix_achat': 5000, 'prix_vente': 7500, 'stock_actuel': 50},
        {'designation': 'Clavier Mécanique', 'prix_achat': 15000, 'prix_vente': 20000, 'stock_actuel': 25},
        {'designation': 'Écran 24 pouces', 'prix_achat': 80000, 'prix_vente': 100000, 'stock_actuel': 15},
        {'designation': 'Casque Audio', 'prix_achat': 25000, 'prix_vente': 35000, 'stock_actuel': 30},
    ]
    
    created_articles = []
    for article_data in test_articles:
        article, created = Article.objects.get_or_create(
            designation=article_data['designation'],
            agence=agence,
            defaults=article_data
        )
        if created:
            created_articles.append(article.designation)
    
    return JsonResponse({
        'message': f'Articles créés: {len(created_articles)}',
        'articles': created_articles
    })

# ==================== INVENTAIRE DE STOCK ====================

@login_required
def inventaire_stock(request):
    """Vue pour la page d'inventaire de stock"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques du stock
    total_articles = articles.count()
    total_quantite = articles.aggregate(total=Sum('stock_actuel'))['total'] or 0
    valeur_totale_stock = articles.aggregate(
        total=Sum(F('stock_actuel') * F('prix_achat'))
    )['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'total_quantite': total_quantite,
        'valeur_totale_stock': valeur_totale_stock,
    }
    
    return render(request, 'supermarket/stock/inventaire_stock.html', context)

@login_required
def generer_inventaire(request):
    """Vue pour générer l'inventaire selon les critères sélectionnés"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        famille_id = request.POST.get('famille')
        selection_articles = request.POST.get('selection_articles', 'tous')  # 'tous' ou 'selectionnes'
        articles_selectionnes = request.POST.getlist('articles_selectionnes[]')
        
        print(f"[SEARCH] PARAMÈTRES INVENTAIRE:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Famille: {famille_id}")
        print(f"  - Sélection: {selection_articles}")
        print(f"  - Articles sélectionnés: {articles_selectionnes}")
        
        # Filtrer les articles selon les critères
        articles_query = Article.objects.filter(agence=agence)
        
        # Filtre par famille
        if famille_id and famille_id != '':
            articles_query = articles_query.filter(categorie_id=famille_id)
        
        # Filtre par sélection d'articles
        if selection_articles == 'selectionnes' and articles_selectionnes:
            articles_query = articles_query.filter(id__in=articles_selectionnes)
        
        articles = articles_query.order_by('designation')
        
        print(f"[PACKAGE] ARTICLES FILTRÉS: {articles.count()}")
        
        # Créer l'inventaire
        numero_inventaire = f"INV-{timezone.now().strftime('%Y%m%d-%H%M%S')}"
        
        # Récupérer l'employé responsable
        employe = Employe.objects.filter(compte__agence=agence).first()
        
        inventaire = InventaireStock.objects.create(
            numero_inventaire=numero_inventaire,
            date_debut=timezone.now(),
            statut='en_cours',
            agence=agence,
            responsable=employe,
            commentaire=f"Inventaire généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        )
        
        # Créer les lignes d'inventaire
        total_quantite = 0
        total_valeur = 0
        
        for article in articles:
            valeur = float(article.stock_actuel) * float(article.prix_achat)
            
            LigneInventaireStock.objects.create(
                inventaire=inventaire,
                reference_article=article.reference_article,
                designation=article.designation,
                quantite_stock=article.stock_actuel,
                prix_unitaire=article.prix_achat,
                valeur=valeur,
                conditionnement=article.conditionnement,
                article=article
            )
            
            total_quantite += article.stock_actuel
            total_valeur += valeur
        
        # Marquer l'inventaire comme terminé
        inventaire.date_fin = timezone.now()
        inventaire.statut = 'termine'
        inventaire.save()
        
        print(f"[OK] INVENTAIRE CRÉÉ: {numero_inventaire}")
        print(f"[CHART] TOTAUX: {total_quantite} articles, {total_valeur} FCFA")
        
        return JsonResponse({
            'success': True,
            'inventaire_id': inventaire.id,
            'numero_inventaire': numero_inventaire,
            'total_articles': articles.count(),
            'total_quantite': total_quantite,
            'total_valeur': total_valeur
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR GÉNÉRATION INVENTAIRE: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_excel(request):
    """Vue pour exporter l'inventaire en format Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible, utilisation du format CSV")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Inventaire {inventaire.numero_inventaire}"
        
        # Style du titre
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire le titre "Inventaire" avec la date du jour
        date_du_jour = timezone.now().strftime('%d/%m/%Y')
        title_cell = ws.cell(row=1, column=1, value=f"Inventaire - {date_du_jour}")
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # Fusionner les cellules pour le titre (sur toutes les colonnes)
        from openpyxl.utils import get_column_letter
        ws.merge_cells(f'A1:{get_column_letter(6)}1')
        
        # Ligne vide
        ws.row_dimensions[2].height = 5
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité en Stock', 'Prix Unitaire (FCFA)', 'Valeur Totale (FCFA)']
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire les en-têtes (décalés à la ligne 3)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Écrire les données (décalées à partir de la ligne 4)
        total_quantite = 0
        total_valeur = 0
        
        for idx, ligne in enumerate(lignes):
            row_num = 4 + idx  # Commencer à la ligne 4 (après titre ligne 1, ligne vide ligne 2, en-têtes ligne 3)
            ws.cell(row=row_num, column=1, value=ligne.reference_article)
            ws.cell(row=row_num, column=2, value=ligne.designation)
            ws.cell(row=row_num, column=3, value=ligne.conditionnement)
            ws.cell(row=row_num, column=4, value=ligne.quantite_stock)
            ws.cell(row=row_num, column=5, value=float(ligne.prix_unitaire))
            ws.cell(row=row_num, column=6, value=float(ligne.valeur))
            
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux (décalée de 2 lignes supplémentaires)
        total_row = 4 + len(lignes) + 2  # Ligne de données + 1 ligne vide + ligne totaux
        ws.cell(row=total_row, column=3, value="TOTAL GÉNÉRAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_quantite).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=total_valeur).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        column_widths = [15, 40, 15, 15, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le classeur dans la réponse
        wb.save(response)
        
        print(f"[CHART] EXPORT EXCEL - Inventaire {inventaire.numero_inventaire}")
        print(f"[CHART] EXPORT EXCEL - {len(lignes)} articles exportés")
        print(f"[CHART] EXPORT EXCEL - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_pdf(request):
    """Vue pour exporter l'inventaire en format PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Centre
        )
        
        # Titre
        title = Paragraph(f"INVENTAIRE DE STOCK - {inventaire.numero_inventaire}", title_style)
        elements.append(title)
        
        # Informations de l'inventaire
        info_data = [
            ['Date de génération:', inventaire.date_debut.strftime('%d/%m/%Y à %H:%M')],
            ['Agence:', agence.nom_agence],
            ['Responsable:', inventaire.responsable.compte.nom_complet if inventaire.responsable else 'Non spécifié'],
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 20))
        
        # Tableau des articles
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité', 'Prix Unitaire', 'Valeur']
        
        # Données du tableau
        data = [headers]
        total_quantite = 0
        total_valeur = 0
        
        for ligne in lignes:
            row = [
                ligne.reference_article,
                ligne.designation,
                ligne.conditionnement,
                str(ligne.quantite_stock),
                f"{float(ligne.prix_unitaire):,.0f}",
                f"{float(ligne.valeur):,.0f}"
            ]
            data.append(row)
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux
        data.append(['', '', 'TOTAL GÉNÉRAL:', str(total_quantite), '', f"{total_valeur:,.0f}"])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 6*cm, 2.5*cm, 2*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            # En-têtes
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Données
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Ligne des totaux
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF - Inventaire {inventaire.numero_inventaire}")
        print(f"📄 EXPORT PDF - {len(lignes)} articles exportés")
        print(f"📄 EXPORT PDF - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_csv(request):
    """Vue pour exporter l'inventaire en format CSV (alternative si Excel n'est pas disponible)"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité en Stock', 'Prix Unitaire (FCFA)', 'Valeur Totale (FCFA)']
        writer.writerow(headers)
        
        # Données
        total_quantite = 0
        total_valeur = 0
        
        for ligne in lignes:
            row = [
                ligne.reference_article,
                ligne.designation,
                ligne.conditionnement,
                ligne.quantite_stock,
                float(ligne.prix_unitaire),
                float(ligne.valeur)
            ]
            writer.writerow(row)
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['', '', 'TOTAL GÉNÉRAL:', total_quantite, '', total_valeur])
        
        print(f"[CHART] EXPORT CSV - Inventaire {inventaire.numero_inventaire}")
        print(f"[CHART] EXPORT CSV - {len(lignes)} articles exportés")
        print(f"[CHART] EXPORT CSV - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== STATISTIQUES DE VENTE ====================

@login_required
def statistiques_vente(request):
    """Vue pour la page des statistiques de vente"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques générales
    total_articles = articles.count()
    
    # Calculer les statistiques de vente des 30 derniers jours
    date_debut = timezone.now().date() - timezone.timedelta(days=30)
    
    # Récupérer les ventes des 30 derniers jours
    ventes_recentes = LigneFactureVente.objects.filter(
        facture_vente__agence=agence,
        facture_vente__date__gte=date_debut
    ).select_related('article', 'facture_vente')
    
    # Calculer le chiffre d'affaires total
    chiffre_affaires_total = float(ventes_recentes.aggregate(
        total=Sum(F('quantite') * F('prix_unitaire'))
    )['total'] or 0)
    
    # Calculer la marge totale
    from decimal import Decimal
    marge_totale = Decimal('0')
    for vente in ventes_recentes:
        prix_achat = Decimal(str(vente.article.prix_achat))
        prix_vente = Decimal(str(vente.prix_unitaire))
        marge_unitaire = prix_vente - prix_achat
        marge_totale += marge_unitaire * Decimal(str(vente.quantite))
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'chiffre_affaires_total': chiffre_affaires_total,
        'marge_totale': float(marge_totale),
        'pourcentage_marge_global': (float(marge_totale) / float(chiffre_affaires_total) * 100) if chiffre_affaires_total > 0 else 0,
    }
    
    return render(request, 'supermarket/stock/statistiques_vente.html', context)

@login_required
def generer_statistiques_vente(request):
    """Vue pour générer les statistiques de vente selon les critères sélectionnés"""
    print("[START] DÉBUT GENERER_STATISTIQUES_VENTE")
    
    if request.method != 'POST':
        print("[ERREUR] Méthode non POST")
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
        print(f"[SEARCH] Agence récupérée: {agence}")
    except Exception as e:
        print(f"[ERREUR] Erreur get_user_agence: {e}")
        return JsonResponse({'success': False, 'error': f'Agence non trouvée: {str(e)}'})
    
    try:
        print("[SEARCH] Début du traitement des paramètres")
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        famille_id = request.POST.get('famille')
        selection_articles = request.POST.get('selection_articles', 'tous')
        articles_selectionnes = request.POST.getlist('articles_selectionnes[]')
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        print(f"[CHART] PARAMÈTRES STATISTIQUES:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Famille: {famille_id}")
        print(f"  - Sélection: {selection_articles}")
        print(f"  - Articles sélectionnés: {articles_selectionnes}")
        
        # Convertir les dates
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Format de date invalide: {str(e)}'})
        
        # Filtrer les articles selon les critères
        print(f"[SEARCH] Filtrage des articles pour agence: {agence}")
        articles_query = Article.objects.filter(agence=agence)
        print(f"[SEARCH] Articles de base: {articles_query.count()}")
        
        # Filtre par famille
        if famille_id and famille_id != '':
            print(f"[SEARCH] Filtrage par famille: {famille_id}")
            articles_query = articles_query.filter(categorie_id=famille_id)
            print(f"[SEARCH] Articles après filtre famille: {articles_query.count()}")
        
        # Filtre par sélection d'articles
        if selection_articles == 'selectionnes' and articles_selectionnes:
            print(f"[SEARCH] Filtrage par sélection: {articles_selectionnes}")
            articles_query = articles_query.filter(id__in=articles_selectionnes)
            print(f"[SEARCH] Articles après filtre sélection: {articles_query.count()}")
        
        articles = articles_query.order_by('designation')
        print(f"[PACKAGE] ARTICLES FILTRÉS: {articles.count()}")
        
        # Récupérer les ventes pour la période
        ventes = LigneFactureVente.objects.filter(
            facture_vente__agence=agence,
            facture_vente__date__gte=date_debut_obj,
            facture_vente__date__lte=date_fin_obj,
            article__in=articles
        ).select_related('article', 'facture_vente')
        
        # Calculer les statistiques par article
        statistiques_articles = []
        chiffre_affaires_total = 0.0
        marge_totale = 0.0
        quantite_totale_vendue = 0.0
        
        for article in articles:
            # Récupérer les ventes de cet article
            ventes_article = ventes.filter(article=article)
            
            # Calculer les totaux pour cet article
            quantite_vendue = float(ventes_article.aggregate(total=Sum('quantite'))['total'] or 0)
            chiffre_affaires_article = float(ventes_article.aggregate(
                total=Sum(F('quantite') * F('prix_unitaire'))
            )['total'] or 0)
            
            # Calculer la marge
            from decimal import Decimal
            prix_achat = Decimal(str(article.prix_achat))
            marge_unitaire = Decimal('0')
            marge_article = Decimal('0')
            
            if quantite_vendue > 0:
                prix_vente_moyen = Decimal(str(chiffre_affaires_article)) / Decimal(str(quantite_vendue))
                marge_unitaire = prix_vente_moyen - prix_achat
                marge_article = marge_unitaire * Decimal(str(quantite_vendue))
            
            # Calculer le pourcentage de marge
            if chiffre_affaires_article > 0:
                pourcentage_marge = (marge_article / Decimal(str(chiffre_affaires_article)) * Decimal('100'))
            else:
                pourcentage_marge = Decimal('0')
            
            if quantite_vendue > 0:  # Ne garder que les articles vendus
                statistiques_articles.append({
                    'reference_article': article.reference_article,
                    'designation': article.designation,
                    'quantite_vendue': float(quantite_vendue),
                    'chiffre_affaires': float(chiffre_affaires_article),
                    'marge_profit': float(marge_article),
                    'pourcentage_marge': float(pourcentage_marge),
                })
                
                chiffre_affaires_total += float(chiffre_affaires_article)
                marge_totale += float(marge_article)
                quantite_totale_vendue += float(quantite_vendue)
        
        # Calculer le pourcentage de marge global
        if chiffre_affaires_total > 0:
            pourcentage_marge_global = (marge_totale / float(chiffre_affaires_total) * 100)
        else:
            pourcentage_marge_global = 0
        
        print(f"[CHART] STATISTIQUES GÉNÉRÉES:")
        print(f"  - Articles vendus: {len(statistiques_articles)}")
        print(f"  - Quantité totale vendue: {quantite_totale_vendue}")
        print(f"  - Chiffre d'affaires total: {chiffre_affaires_total}")
        print(f"  - Marge totale: {marge_totale}")
        print(f"  - Pourcentage marge global: {pourcentage_marge_global:.2f}%")
        
        # Stocker les statistiques dans la session pour l'export (conversion en types sérialisables)
        request.session['statistiques_vente'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'statistiques_articles': statistiques_articles,  # Déjà converties en float
            'chiffre_affaires_total': float(chiffre_affaires_total),
            'marge_totale': float(marge_totale),
            'quantite_totale_vendue': int(quantite_totale_vendue),
            'pourcentage_marge_global': float(pourcentage_marge_global),
        }
        
        return JsonResponse({
            'success': True,
            'total_articles': len(statistiques_articles),
            'quantite_totale_vendue': int(quantite_totale_vendue),
            'chiffre_affaires_total': float(chiffre_affaires_total),
            'marge_totale': float(marge_totale),
            'pourcentage_marge_global': float(pourcentage_marge_global),
            'statistiques_articles': statistiques_articles  # Ajouter les données détaillées
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR GÉNÉRATION STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        error_details = f"Erreur: {str(e)}"
        if hasattr(e, '__class__'):
            error_details += f" (Type: {e.__class__.__name__})"
        return JsonResponse({'success': False, 'error': error_details})

@login_required
def export_statistiques_excel(request):
    """Vue pour exporter les statistiques de vente en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible, utilisation du format CSV")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Statistiques Vente {date_debut} - {date_fin}"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tête du document
        ws.merge_cells('A1:F1')
        ws['A1'] = f"STATISTIQUES DE VENTE - {agence.nom_agence}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Période: du {date_debut} au {date_fin}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes des colonnes
        headers = ['Référence', 'Désignation', 'Quantité Vendue', 'Chiffre d\'Affaires (FCFA)', 'Marge Profit (FCFA)', 'Pourcentage Marge (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données
        row = 5
        for stat in statistiques_articles:
            ws.cell(row=row, column=1, value=stat['reference_article'])
            ws.cell(row=row, column=2, value=stat['designation'])
            ws.cell(row=row, column=3, value=stat['quantite_vendue'])
            ws.cell(row=row, column=4, value=float(stat['chiffre_affaires']))
            ws.cell(row=row, column=5, value=float(stat['marge_profit']))
            ws.cell(row=row, column=6, value=float(stat['pourcentage_marge']))
            row += 1
        
        # Ligne des totaux
        row += 1
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value="").font = Font(bold=True)
        ws.cell(row=row, column=3, value=quantite_totale_vendue).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(chiffre_affaires_total)).font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(marge_totale)).font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(pourcentage_marge_global)).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 18
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le fichier dans la réponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_statistiques_pdf(request):
    """Vue pour exporter les statistiques de vente en PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            textColor=colors.darkblue
        )
        
        # Éléments du document
        elements = []
        
        # Titre
        title = Paragraph(f"STATISTIQUES DE VENTE - {agence.nom_agence}", title_style)
        elements.append(title)
        
        # Informations de la période
        period_text = f"<b>Période:</b> du {date_debut} au {date_fin}"
        period_para = Paragraph(period_text, styles['Normal'])
        elements.append(period_para)
        elements.append(Spacer(1, 20))
        
        # Données du tableau
        data = [['Référence', 'Désignation', 'Qté Vendue', 'Chiffre d\'Affaires', 'Marge Profit', 'Marge %']]
        
        for stat in statistiques_articles:
            data.append([
                stat['reference_article'],
                stat['designation'][:30] + '...' if len(stat['designation']) > 30 else stat['designation'],
                str(stat['quantite_vendue']),
                f"{float(stat['chiffre_affaires']):,.0f} FCFA",
                f"{float(stat['marge_profit']):,.0f} FCFA",
                f"{float(stat['pourcentage_marge']):.1f}%"
            ])
        
        # Ligne des totaux
        data.append([
            'TOTAL GÉNÉRAL',
            '',
            str(quantite_totale_vendue),
            f"{float(chiffre_affaires_total):,.0f} FCFA",
            f"{float(marge_totale):,.0f} FCFA",
            f"{float(pourcentage_marge_global):.1f}%"
        ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 6*cm, 2*cm, 3*cm, 3*cm, 2*cm])
        
        # Style du tableau
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"📄 EXPORT PDF STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"📄 EXPORT PDF STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_statistiques_csv(request):
    """Vue pour exporter les statistiques de vente en CSV"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Quantité Vendue', 'Chiffre d\'Affaires (FCFA)', 'Marge Profit (FCFA)', 'Pourcentage Marge (%)']
        writer.writerow(headers)
        
        # Données
        for stat in statistiques_articles:
            row = [
                stat['reference_article'],
                stat['designation'],
                stat['quantite_vendue'],
                float(stat['chiffre_affaires']),
                float(stat['marge_profit']),
                float(stat['pourcentage_marge'])
            ]
            writer.writerow(row)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['TOTAL GÉNÉRAL', '', quantite_totale_vendue, chiffre_affaires_total, marge_totale, pourcentage_marge_global])
        
        print(f"[CHART] EXPORT CSV STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"[CHART] EXPORT CSV STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"[CHART] EXPORT CSV STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_statistiques(request):
    """Vue de test pour diagnostiquer les problèmes"""
    try:
        print("🧪 TEST STATISTIQUES - Début")
        
        # Test 1: Récupération de l'agence
        agence = get_user_agence(request)
        print(f"🧪 TEST - Agence récupérée: {agence}")
        
        # Test 2: Vérification des imports
        from decimal import Decimal
        print("🧪 TEST - Import Decimal OK")
        
        # Test 3: Vérification des modèles
        articles_count = Article.objects.filter(agence=agence).count()
        print(f"🧪 TEST - Articles trouvés: {articles_count}")
        
        # Test 4: Vérification des ventes
        ventes_count = LigneFactureVente.objects.filter(facture_vente__agence=agence).count()
        print(f"🧪 TEST - Ventes trouvées: {ventes_count}")
        
        return JsonResponse({
            'success': True,
            'message': 'Tests réussis',
            'agence': str(agence),
            'articles_count': articles_count,
            'ventes_count': ventes_count
        })
        
    except Exception as e:
        print(f"🧪 TEST - Erreur: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== MOUVEMENTS DE STOCK ====================

@login_required
def mouvements_stock(request):
    """Vue pour la page des mouvements de stock"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques générales
    total_articles = articles.count()
    
    # Calculer les statistiques des mouvements des 30 derniers jours
    date_debut = timezone.now().date() - timezone.timedelta(days=30)
    
    # Récupérer les mouvements des 30 derniers jours
    mouvements_recentes = MouvementStock.objects.filter(
        agence=agence,
        date_mouvement__gte=date_debut
    ).select_related('article', 'fournisseur', 'facture_vente', 'facture_achat')
    
    # Statistiques des mouvements
    total_mouvements = mouvements_recentes.count()
    mouvements_entree = mouvements_recentes.filter(type_mouvement='entree').count()
    mouvements_sortie = mouvements_recentes.filter(type_mouvement='sortie').count()
    
    # Valeur totale du stock permanent
    valeur_stock_permanent = mouvements_recentes.aggregate(
        total=Sum('stock_permanent')
    )['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'total_mouvements': total_mouvements,
        'mouvements_entree': mouvements_entree,
        'mouvements_sortie': mouvements_sortie,
        'valeur_stock_permanent': valeur_stock_permanent,
    }
    
    return render(request, 'supermarket/stock/mouvements_stock.html', context)

@login_required
def consulter_mouvements_stock(request):
    """Vue pour consulter les mouvements de stock selon les critères sélectionnés"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        article_ids = request.POST.getlist('articles')  # Récupérer tous les articles sélectionnés
        type_mouvement = request.POST.get('type_mouvement', '')
        
        print(f"[CHART] PARAMÈTRES MOUVEMENTS:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Articles: {article_ids}")
        print(f"  - Type mouvement: {type_mouvement}")
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        if not article_ids:
            return JsonResponse({'success': False, 'error': 'Veuillez sélectionner au moins un article'})
        
        # Convertir les dates
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Format de date invalide: {str(e)}'})
        
        # Récupérer les articles
        try:
            articles = Article.objects.filter(id__in=article_ids, agence=agence)
            if not articles.exists():
                return JsonResponse({'success': False, 'error': 'Aucun article valide trouvé'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Erreur lors de la récupération des articles: {str(e)}'})
        
        # Filtrer les mouvements selon les critères
        # Utiliser __date pour comparer seulement les dates (ignorer l'heure)
        mouvements_query = MouvementStock.objects.filter(
            agence=agence,
            article__in=articles,  # Filtrer par plusieurs articles
            date_mouvement__date__gte=date_debut_obj,
            date_mouvement__date__lte=date_fin_obj
        ).select_related('article', 'fournisseur', 'facture_vente', 'facture_achat').order_by('article__reference_article', 'date_mouvement')
        
        # Filtre par type de mouvement
        if type_mouvement and type_mouvement != '':
            mouvements_query = mouvements_query.filter(type_mouvement=type_mouvement)
        
        mouvements = mouvements_query
        
        print(f"[PACKAGE] MOUVEMENTS FILTRÉS: {mouvements.count()}")
        
        # Debug: Afficher quelques mouvements pour vérifier
        if mouvements.count() > 0:
            print("[LIST] PREMIERS MOUVEMENTS TROUVÉS:")
            for i, mvt in enumerate(mouvements[:3]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement}")
        else:
            print("[ERREUR] AUCUN MOUVEMENT TROUVÉ - Vérifions les mouvements existants:")
            tous_mouvements = MouvementStock.objects.filter(agence=agence, article__in=articles)
            print(f"[CHART] Total mouvements pour ces articles: {tous_mouvements.count()}")
            for i, mvt in enumerate(tous_mouvements[:3]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement}")
        
        # Calculer les statistiques
        total_mouvements = mouvements.count()
        mouvements_entree = mouvements.filter(type_mouvement='entree').count()
        mouvements_sortie = mouvements.filter(type_mouvement='sortie').count()
        
        # Calculer la valeur totale du stock permanent
        valeur_stock_permanent = mouvements.aggregate(
            total=Sum('stock_permanent')
        )['total'] or 0
        
        # Stocker les données dans la session pour l'export
        mouvements_data = []
        for mouvement in mouvements:
            # Déterminer le tiers
            tiers = ""
            if mouvement.fournisseur:
                tiers = f"Fournisseur: {mouvement.fournisseur.intitule}"
            elif mouvement.facture_vente:
                tiers = f"Client: {mouvement.facture_vente.client.intitule if mouvement.facture_vente.client else 'N/A'}"
            elif mouvement.facture_achat:
                tiers = f"Fournisseur: {mouvement.facture_achat.fournisseur.intitule if mouvement.facture_achat.fournisseur else 'N/A'}"
            
            mouvements_data.append({
                'date_mouvement': mouvement.date_mouvement.strftime('%Y-%m-%d %H:%M'),
                'type_mouvement': mouvement.type_mouvement,  # Garder la valeur brute pour le JavaScript
                'type_mouvement_display': mouvement.get_type_mouvement_display(),  # Label d'affichage
                'tiers': tiers,
                'stock_initial': mouvement.stock_initial,
                'quantite': mouvement.quantite,
                'solde': mouvement.solde,
                'cout_moyen_pondere': float(mouvement.cout_moyen_pondere),
                'stock_permanent': float(mouvement.stock_permanent),
                'numero_piece': mouvement.numero_piece,
                'commentaire': mouvement.commentaire or '',
            })
        
        print(f"[CHART] MOUVEMENTS GÉNÉRÉS:")
        print(f"  - Total mouvements: {total_mouvements}")
        print(f"  - Entrées: {mouvements_entree}")
        print(f"  - Sorties: {mouvements_sortie}")
        print(f"  - Valeur stock permanent: {valeur_stock_permanent}")
        
        # Stocker les mouvements dans la session pour l'export
        request.session['mouvements_stock'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'articles_count': len(articles),
            'mouvements_data': mouvements_data,
            'total_mouvements': total_mouvements,
            'mouvements_entree': mouvements_entree,
            'mouvements_sortie': mouvements_sortie,
            'valeur_stock_permanent': float(valeur_stock_permanent),
            'article_reference': ', '.join([article.reference_article for article in articles]),
            'article_designation': ', '.join([article.designation for article in articles]),
        }
        
        # Préparer les informations des articles
        articles_info = []
        for article in articles:
            article_mouvements = mouvements.filter(article=article)
            articles_info.append({
                'id': article.id,
                'reference': article.reference_article,
                'designation': article.designation,
                'stock_actuel': article.stock_actuel,
                'mouvements_count': article_mouvements.count()
            })
        
        return JsonResponse({
            'success': True,
            'total_mouvements': total_mouvements,
            'mouvements_entree': mouvements_entree,
            'mouvements_sortie': mouvements_sortie,
            'valeur_stock_permanent': float(valeur_stock_permanent),
            'articles_info': articles_info,
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'mouvements': mouvements_data,
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR CONSULTATION MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        error_details = f"Erreur: {str(e)}"
        if hasattr(e, '__class__'):
            error_details += f" (Type: {e.__class__.__name__})"
        return JsonResponse({'success': False, 'error': error_details})

@login_required
def export_mouvements_excel(request):
    """Vue pour exporter les mouvements de stock en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Mouvements Stock {article_reference}"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tête du document
        ws.merge_cells('A1:P1')
        ws['A1'] = f"FICHE DE STOCK - {article_reference} - {article_designation}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:P2')
        ws['A2'] = f"Période: du {date_debut} au {date_fin}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes des colonnes
        headers = ['Date', 'Type', '', 'N°', '', '', 'Tiers', '', '', '', '', '+/-', 'Quantités en stock', 'Solde', 'C.M.U.P.', 'Stock permanent']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données
        row = 5
        for mouvement in mouvements_data_list:
            # Structure modifiée (16 colonnes - sans Référence/Désignation/unitaire, avec Tiers)
            ws.cell(row=row, column=1, value=mouvement['date_mouvement'])  # Date
            ws.cell(row=row, column=2, value=mouvement['type_mouvement'])  # Type
            ws.cell(row=row, column=3, value='')  # Colonne vide
            ws.cell(row=row, column=4, value=mouvement['numero_piece'])  # N°
            ws.cell(row=row, column=5, value='')  # Colonne vide
            ws.cell(row=row, column=6, value='')  # Colonne vide
            ws.cell(row=row, column=7, value=mouvement['tiers'])  # Tiers
            ws.cell(row=row, column=8, value='')  # Colonne vide
            ws.cell(row=row, column=9, value='')  # Colonne vide
            ws.cell(row=row, column=10, value='')  # Colonne vide
            ws.cell(row=row, column=11, value='')  # Colonne vide
            ws.cell(row=row, column=12, value=f"+{mouvement['quantite']}" if mouvement['quantite'] > 0 else mouvement['quantite'])  # +/-
            ws.cell(row=row, column=13, value=mouvement['stock_initial'])  # Quantités en stock
            ws.cell(row=row, column=14, value=mouvement['solde'])  # Solde
            ws.cell(row=row, column=15, value=mouvement['cout_moyen_pondere'])  # C.M.U.P.
            ws.cell(row=row, column=16, value=mouvement['stock_permanent'])  # Stock permanent
            row += 1
        
        # Ligne des totaux
        row += 1
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"Entrées: {mouvements_entree}, Sorties: {mouvements_sortie}").font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"Total: {total_mouvements} mouvements").font = Font(bold=True)
        ws.cell(row=row, column=4, value="").font = Font(bold=True)
        ws.cell(row=row, column=5, value="").font = Font(bold=True)
        ws.cell(row=row, column=6, value="").font = Font(bold=True)
        ws.cell(row=row, column=16, value=valeur_stock_permanent).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 15
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le fichier dans la réponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Article: {article_reference}")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - {total_mouvements} mouvements")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_mouvements_pdf(request):
    """Vue pour exporter les mouvements de stock en PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            textColor=colors.darkblue
        )
        
        # Éléments du document
        elements = []
        
        # Titre
        title = Paragraph(f"FICHE DE STOCK - {article_reference}", title_style)
        elements.append(title)
        
        # Informations de l'article et période
        info_text = f"<b>Article:</b> {article_designation}<br/><b>Période:</b> du {date_debut} au {date_fin}"
        info_para = Paragraph(info_text, styles['Normal'])
        elements.append(info_para)
        elements.append(Spacer(1, 20))
        
        # Données du tableau
        data = [['Date', 'Type', 'Tiers', 'Stock Init.', 'Solde', 'C.M.PU', 'Stock Perm.', 'N° Pièce']]
        
        for mouvement in mouvements_data_list:
            data.append([
                mouvement['date_mouvement'],
                mouvement['type_mouvement'],
                mouvement['tiers'][:20] + '...' if len(mouvement['tiers']) > 20 else mouvement['tiers'],
                str(mouvement['stock_initial']),
                str(mouvement['solde']),
                f"{mouvement['cout_moyen_pondere']:,.0f}",
                f"{mouvement['stock_permanent']:,.0f}",
                mouvement['numero_piece']
            ])
        
        # Ligne des totaux
        data.append([
            'TOTAL',
            f'E:{mouvements_entree} S:{mouvements_sortie}',
            f'{total_mouvements} mouvements',
            '',
            '',
            '',
            f"{valeur_stock_permanent:,.0f} FCFA",
            ''
        ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 1.5*cm, 3*cm, 1.5*cm, 1.5*cm, 2*cm, 2.5*cm, 2*cm])
        
        # Style du tableau
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF MOUVEMENTS - Article: {article_reference}")
        print(f"📄 EXPORT PDF MOUVEMENTS - {total_mouvements} mouvements")
        print(f"📄 EXPORT PDF MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_mouvements_csv(request):
    """Vue pour exporter les mouvements de stock en CSV"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Date', 'Type', '', 'N°', '', '', 'Tiers', '', '', '', '', '+/-', 'Quantités en stock', 'Solde', 'C.M.U.P.', 'Stock permanent']
        writer.writerow(headers)
        
        # Données
        for mouvement in mouvements_data_list:
            row = [
                mouvement['date_mouvement'],
                mouvement['type_mouvement'],
                mouvement['tiers'],
                mouvement['stock_initial'],
                mouvement['solde'],
                mouvement['cout_moyen_pondere'],
                mouvement['stock_permanent'],
                mouvement['numero_piece']
            ]
            writer.writerow(row)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['TOTAL GÉNÉRAL', f'Entrées: {mouvements_entree}, Sorties: {mouvements_sortie}', f'{total_mouvements} mouvements', '', '', '', valeur_stock_permanent, ''])
        
        print(f"[CHART] EXPORT CSV MOUVEMENTS - Article: {article_reference}")
        print(f"[CHART] EXPORT CSV MOUVEMENTS - {total_mouvements} mouvements")
        print(f"[CHART] EXPORT CSV MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def creer_mouvements_retroactifs(request):
    """Vue simplifiée pour créer des mouvements de stock rétroactifs"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[REFRESH] CRÉATION MOUVEMENTS RÉTROACTIFS (VERSION SIMPLIFIÉE)...")
        print(f"[TARGET] Agence utilisée: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # 1. Créer des mouvements pour les factures de vente
        factures_vente = FactureVente.objects.filter(agence=agence)
        print(f"[CHART] Factures de vente: {factures_vente.count()}")
        
        for facture in factures_vente:
            lignes = LigneFactureVente.objects.filter(facture_vente=facture)
            print(f"[SEARCH] Facture {facture.numero_ticket}: {lignes.count()} lignes")
            
            for ligne in lignes:
                print(f"  [PACKAGE] Article: {ligne.article.designation}, Quantité: {ligne.quantite}")
                
                # Vérifier si le mouvement existe déjà
                mouvement_existe = MouvementStock.objects.filter(facture_vente=facture, article=ligne.article).exists()
                print(f"  [SEARCH] Mouvement existe déjà: {mouvement_existe}")
                
                if not mouvement_existe:
                    try:
                        from datetime import datetime
                        from django.utils import timezone
                        
                        # Utiliser timezone.now() pour la date
                        date_mouvement = timezone.now()
                        
                        MouvementStock.objects.create(
                            article=ligne.article,
                            agence=agence,
                            type_mouvement='sortie',
                            date_mouvement=date_mouvement,
                            numero_piece=facture.numero_ticket,
                            quantite_stock=ligne.article.stock_actuel,
                            stock_initial=ligne.article.stock_actuel + ligne.quantite,
                            solde=ligne.article.stock_actuel,
                            quantite=ligne.quantite,
                            cout_moyen_pondere=float(ligne.prix_unitaire),
                            stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                            facture_vente=facture,
                            commentaire=f"Vente - {facture.numero_ticket}"
                        )
                        mouvements_crees += 1
                        print(f"[OK] Vente: {ligne.article.designation} - {facture.numero_ticket}")
                    except Exception as e:
                        print(f"[ERREUR] Erreur vente {facture.numero_ticket}: {e}")
        
        # 2. Créer des mouvements pour les factures d'achat
        factures_achat = FactureAchat.objects.filter(agence=agence)
        print(f"[CHART] Factures d'achat: {factures_achat.count()}")
        
        for facture in factures_achat:
            lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
            print(f"[SEARCH] Facture achat {facture.reference_achat}: {lignes.count()} lignes")
            
            for ligne in lignes:
                print(f"  [PACKAGE] Article: {ligne.article.designation}, Quantité: {ligne.quantite}")
                
                mouvement_existe = MouvementStock.objects.filter(facture_achat=facture, article=ligne.article).exists()
                print(f"  [SEARCH] Mouvement existe déjà: {mouvement_existe}")
                
                if not mouvement_existe:
                    try:
                        from django.utils import timezone
                        
                        MouvementStock.objects.create(
                            article=ligne.article,
                            agence=agence,
                            type_mouvement='entree',
                            date_mouvement=timezone.now(),
                            numero_piece=facture.reference_achat,
                            quantite_stock=ligne.article.stock_actuel,
                            stock_initial=ligne.article.stock_actuel - ligne.quantite,
                            solde=ligne.article.stock_actuel,
                            quantite=ligne.quantite,
                            cout_moyen_pondere=float(ligne.prix_unitaire),
                            stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                            facture_achat=facture,
                            fournisseur=facture.fournisseur,
                            commentaire=f"Achat - {facture.reference_achat}"
                        )
                        mouvements_crees += 1
                        print(f"[OK] Achat: {ligne.article.designation} - {facture.reference_achat}")
                    except Exception as e:
                        print(f"[ERREUR] Erreur achat {facture.reference_achat}: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'{mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def diagnostic_mouvements(request):
    """Vue de diagnostic pour les mouvements de stock"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Compter les données
        articles_count = Article.objects.filter(agence=agence).count()
        factures_vente_count = FactureVente.objects.filter(agence=agence).count()
        factures_achat_count = FactureAchat.objects.filter(agence=agence).count()
        factures_transfert_count = FactureTransfert.objects.filter(agence_source=agence).count()
        mouvements_count = MouvementStock.objects.filter(agence=agence).count()
        
        # Détails des factures
        factures_vente_details = []
        for facture in FactureVente.objects.filter(agence=agence)[:5]:
            lignes_count = LigneFactureVente.objects.filter(facture_vente=facture).count()
            factures_vente_details.append({
                'numero': facture.numero_ticket,
                'date': str(facture.date),
                'lignes': lignes_count
            })
        
        factures_achat_details = []
        for facture in FactureAchat.objects.filter(agence=agence)[:5]:
            lignes_count = LigneFactureAchat.objects.filter(facture_achat=facture).count()
            factures_achat_details.append({
                'numero': facture.reference_achat,
                'date': str(facture.date_achat),
                'lignes': lignes_count
            })
        
        return JsonResponse({
            'success': True,
            'agence': agence.nom_agence,
            'articles_count': articles_count,
            'factures_vente_count': factures_vente_count,
            'factures_achat_count': factures_achat_count,
            'factures_transfert_count': factures_transfert_count,
            'mouvements_count': mouvements_count,
            'factures_vente_details': factures_vente_details,
            'factures_achat_details': factures_achat_details
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def forcer_mouvements(request):
    """Vue pour forcer la création de mouvements même s'ils existent déjà"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[HOT] CRÉATION FORCÉE DE MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # Supprimer tous les mouvements existants d'abord
        anciens_mouvements = MouvementStock.objects.filter(agence=agence).count()
        MouvementStock.objects.filter(agence=agence).delete()
        print(f"🗑️ {anciens_mouvements} anciens mouvements supprimés")
        
        # Test de création d'un mouvement simple
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles trouvés: {articles.count()}")
        
        if articles.exists():
            article_test = articles.first()
            print(f"🧪 Test avec article: {article_test.designation}")
            
            try:
                from django.utils import timezone
                
                mouvement_test = MouvementStock.objects.create(
                    article=article_test,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece='TEST-001',
                    quantite_stock=article_test.stock_actuel,
                    stock_initial=0,
                    solde=article_test.stock_actuel,
                    quantite=1,
                    cout_moyen_pondere=float(article_test.prix_achat),
                    stock_permanent=float(article_test.stock_actuel * article_test.prix_achat),
                    commentaire='Test de création'
                )
                print(f"[OK] MOUVEMENT TEST CRÉÉ AVEC SUCCÈS: ID {mouvement_test.id}")
                mouvements_crees += 1
                
                # Supprimer le test
                mouvement_test.delete()
                print(f"🗑️ Mouvement test supprimé")
                
            except Exception as e:
                print(f"[ERREUR] ERREUR LORS DU TEST: {e}")
                import traceback
                traceback.print_exc()
                return JsonResponse({'success': False, 'error': f'Erreur lors du test de création: {str(e)}'})
        
        print(f"[OK] Test terminé, création des vrais mouvements...")
        
        # Créer des mouvements pour les factures de vente (version simplifiée)
        factures_vente = FactureVente.objects.filter(agence=agence)
        print(f"[CHART] Factures de vente: {factures_vente.count()}")
        
        for facture in factures_vente:
            lignes = LigneFactureVente.objects.filter(facture_vente=facture)
            print(f"  [SEARCH] {facture.numero_ticket}: {lignes.count()} lignes")
            
            for ligne in lignes:
                try:
                    from django.utils import timezone
                    
                    # Utiliser la même logique simple que le test qui fonctionne
                    MouvementStock.objects.create(
                        article=ligne.article,
                        agence=agence,
                        type_mouvement='sortie',
                        date_mouvement=timezone.now(),
                        numero_piece=facture.numero_ticket,
                        quantite_stock=ligne.article.stock_actuel,
                        stock_initial=0,  # Simplifié
                        solde=ligne.article.stock_actuel,
                        quantite=ligne.quantite,
                        cout_moyen_pondere=float(ligne.prix_unitaire),
                        stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                        facture_vente=facture,
                        commentaire=f"Vente - {facture.numero_ticket}"
                    )
                    mouvements_crees += 1
                    print(f"    [OK] {ligne.article.designation}")
                except Exception as e:
                    print(f"    [ERREUR] Erreur: {e}")
        
        # Créer des mouvements pour les factures d'achat (version simplifiée)
        factures_achat = FactureAchat.objects.filter(agence=agence)
        print(f"[CHART] Factures d'achat: {factures_achat.count()}")
        
        for facture in factures_achat:
            lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
            print(f"  [SEARCH] {facture.reference_achat}: {lignes.count()} lignes")
            
            for ligne in lignes:
                try:
                    from django.utils import timezone
                    
                    # Utiliser la même logique simple que le test qui fonctionne
                    MouvementStock.objects.create(
                        article=ligne.article,
                        agence=agence,
                        type_mouvement='entree',
                        date_mouvement=timezone.now(),
                        numero_piece=facture.reference_achat,
                        quantite_stock=ligne.article.stock_actuel,
                        stock_initial=0,  # Simplifié
                        solde=ligne.article.stock_actuel,
                        quantite=ligne.quantite,
                        cout_moyen_pondere=float(ligne.prix_unitaire),
                        stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                        facture_achat=facture,
                        fournisseur=facture.fournisseur,
                        commentaire=f"Achat - {facture.reference_achat}"
                    )
                    mouvements_crees += 1
                    print(f"    [OK] {ligne.article.designation}")
                except Exception as e:
                    print(f"    [ERREUR] Erreur: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'FORCÉ: {mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_mouvement_simple(request):
    """Test simple de création d'un mouvement"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("🧪 TEST SIMPLE DE CRÉATION DE MOUVEMENT...")
        
        # Vérifier les articles
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles: {articles.count()}")
        
        if not articles.exists():
            return JsonResponse({'success': False, 'error': 'Aucun article trouvé dans cette agence'})
        
        article = articles.first()
        print(f"[PACKAGE] Article test: {article.designation}")
        
        # Vérifier les champs obligatoires
        print(f"[PACKAGE] Stock actuel: {article.stock_actuel}")
        print(f"[PACKAGE] Prix achat: {article.prix_achat}")
        
        # Créer un mouvement simple
        from django.utils import timezone
        
        mouvement = MouvementStock.objects.create(
            article=article,
            agence=agence,
            type_mouvement='entree',
            date_mouvement=timezone.now(),
            numero_piece='TEST-SIMPLE',
            quantite_stock=article.stock_actuel,
            stock_initial=0,
            solde=article.stock_actuel,
            quantite=1,
            cout_moyen_pondere=float(article.prix_achat),
            stock_permanent=float(article.stock_actuel * article.prix_achat),
            commentaire='Test simple'
        )
        
        print(f"[OK] MOUVEMENT CRÉÉ: ID {mouvement.id}")
        
        # Vérifier qu'il existe
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        print(f"[CHART] Total mouvements: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'Test réussi! Mouvement ID {mouvement.id} créé. Total: {total_mouvements}',
            'mouvement_id': mouvement.id,
            'total_mouvements': total_mouvements
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR TEST: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def creer_mouvements_manuels(request):
    """Créer des mouvements manuels simples pour tester"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[TOOL] CRÉATION MANUELLE DE MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # Récupérer tous les articles
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles trouvés: {articles.count()}")
        
        if not articles.exists():
            return JsonResponse({'success': False, 'error': 'Aucun article trouvé'})
        
        # Créer un mouvement pour chaque article
        for article in articles:
            try:
                from django.utils import timezone
                
                MouvementStock.objects.create(
                    article=article,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece=f'MANUEL-{article.id}',
                    quantite_stock=article.stock_actuel,
                    stock_initial=0,
                    solde=article.stock_actuel,
                    quantite=article.stock_actuel,
                    cout_moyen_pondere=float(article.prix_achat),
                    stock_permanent=float(article.stock_actuel * article.prix_achat),
                    commentaire=f'Création manuelle - {article.designation}'
                )
                mouvements_crees += 1
                print(f"[OK] {article.designation}")
                
            except Exception as e:
                print(f"[ERREUR] Erreur pour {article.designation}: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements manuels créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'MANUEL: {mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_consultation_mouvements(request):
    """Test simple pour vérifier les mouvements existants"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[SEARCH] TEST CONSULTATION MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        # Vérifier tous les mouvements de l'agence
        tous_mouvements = MouvementStock.objects.filter(agence=agence)
        print(f"[CHART] Total mouvements dans l'agence: {tous_mouvements.count()}")
        
        if tous_mouvements.count() > 0:
            print("[LIST] PREMIERS MOUVEMENTS:")
            for i, mvt in enumerate(tous_mouvements[:5]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement} - {mvt.numero_piece}")
        
        # Vérifier les articles avec mouvements
        articles_avec_mouvements = Article.objects.filter(
            agence=agence,
            mouvementstock__isnull=False
        ).distinct()
        print(f"[PACKAGE] Articles avec mouvements: {articles_avec_mouvements.count()}")
        
        for article in articles_avec_mouvements[:3]:
            mouvements_article = MouvementStock.objects.filter(agence=agence, article=article)
            print(f"  - {article.designation}: {mouvements_article.count()} mouvements")
        
        return JsonResponse({
            'success': True,
            'message': f'Test terminé - {tous_mouvements.count()} mouvements trouvés',
            'total_mouvements': tous_mouvements.count(),
            'articles_avec_mouvements': articles_avec_mouvements.count()
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_mouvements_session(request):
    """Récupérer les données de mouvements depuis la session"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les données depuis la session
        mouvements_data = request.session.get('mouvements_data', [])
        article_info = request.session.get('article_info', {})
        
        print(f"[CHART] RÉCUPÉRATION SESSION:")
        print(f"  - Mouvements en session: {len(mouvements_data)}")
        print(f"  - Article info: {article_info}")
        
        return JsonResponse({
            'success': True,
            'mouvements': mouvements_data,
            'article_info': article_info
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})


# ==================== DÉFACTURATION ====================

@login_required
def defacturer_vente(request, facture_id):
    """Vue pour défacturer (annuler) une vente"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('detail_factures')
    
    try:
        # Récupérer la facture
        facture = FactureVente.objects.get(id=facture_id, agence=agence)
        
        # Utiliser une transaction pour s'assurer de la cohérence
        from django.db import transaction
        
        with transaction.atomic():
            print(f"[DÉFACTURATION] Début de la défacturation de la facture {facture.numero_ticket}")
            
            # 1. Récupérer toutes les lignes de la facture
            lignes = LigneFactureVente.objects.filter(facture_vente=facture)
            print(f"[DÉFACTURATION] {lignes.count()} lignes à traiter")
            
            # 2. Remettre les produits en stock et créer des mouvements inversés
            for ligne in lignes:
                article = ligne.article
                quantite_a_remettre = ligne.quantite
                
                print(f"[DÉFACTURATION] Traitement de {article.designation} - Quantité: {quantite_a_remettre}")
                
                # Sauvegarder l'ancien stock
                ancien_stock = article.stock_actuel
                
                # Remettre en stock
                article.stock_actuel += quantite_a_remettre
                article.save()
                
                print(f"[DÉFACTURATION] Stock remis: {ancien_stock} → {article.stock_actuel} (+{quantite_a_remettre})")
                
                # Créer un mouvement de stock inverse (entrée)
                MouvementStock.objects.create(
                    article=article,
                    agence=agence,
                    type_mouvement='retour',  # Type spécial pour les retours
                    date_mouvement=timezone.now(),
                    numero_piece=f"RETOUR-{facture.numero_ticket}",
                    quantite_stock=article.stock_actuel,
                    stock_initial=ancien_stock,
                    solde=article.stock_actuel,
                    quantite=quantite_a_remettre,
                    cout_moyen_pondere=float(article.prix_achat),
                    stock_permanent=float(article.stock_actuel * article.prix_achat),
                    facture_vente=facture,  # Référence à la facture annulée
                    commentaire=f"Défacturation - Retour stock pour facture {facture.numero_ticket}"
                )
                
                print(f"[DÉFACTURATION] Mouvement de retour créé pour {article.designation}")
            
            # 3. Supprimer la facture et ses lignes
            # Supprimer les lignes de facture
            lignes.delete()
            print(f"[DÉFACTURATION] Lignes de facture supprimées")
            
            # Supprimer la facture
            numero_ticket = facture.numero_ticket
            facture.delete()
            print(f"[DÉFACTURATION] Facture {numero_ticket} supprimée")
            
            messages.success(request, f'Facture {numero_ticket} défacturée avec succès. Les produits ont été remis en stock.')
            
    except FactureVente.DoesNotExist:
        messages.error(request, 'Facture non trouvée.')
        return redirect('detail_factures')
    except Exception as e:
        print(f"[ERREUR] Erreur lors de la défacturation: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, f'Erreur lors de la défacturation: {str(e)}')
        return redirect('detail_factures')
    
    return redirect('detail_factures')


@login_required
def defacturer_vente_confirmation(request, facture_id):
    """Vue pour confirmer la défacturation d'une vente"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    try:
        facture = FactureVente.objects.get(id=facture_id, agence=agence)
        lignes = LigneFactureVente.objects.filter(facture_vente=facture)
        
        context = {
            'facture': facture,
            'lignes': lignes,
            'agence': agence,
        }
        
        return render(request, 'supermarket/stock/defacturation_confirmation.html', context)
        
    except FactureVente.DoesNotExist:
        messages.error(request, 'Facture non trouvée.')
        return redirect('detail_factures')
