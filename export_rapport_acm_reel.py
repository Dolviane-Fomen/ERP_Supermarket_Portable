"""
Script pour exporter le rapport ACM avec les vraies données de la base
Usage: py export_rapport_acm_reel.py [--date-debut YYYY-MM-DD] [--date-fin YYYY-MM-DD] [--zone ZONE] [--agence AGENCE_ID]
"""

import os
import sys
import django
import argparse
from datetime import datetime, timedelta, time
from collections import defaultdict

# Configuration de Django
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'erp_project.settings')
django.setup()

from supermarket.models import Agence, Client, SuiviClientAction, Commande
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def get_action_color(action_text, fill_rouge_fonce, fill_rouge_clair, fill_violet, fill_orange, 
                     fill_jaune, fill_vert_clair, fill_vert_fonce, fill_bleu_clair, fill_rose, fill_blanc):
    """Retourne la couleur appropriée selon le contenu de l'action"""
    if not action_text:
        return fill_blanc
    
    action_lower = action_text.lower().strip()
    
    # Rouge foncé : Actions critiques - Ne plus contacter
    if any(mot in action_lower for mot in ['ne plus rappeler', 'ne pas rappeler', 'ne plus contacter', 
                                            'ne plus appeler', 'arrêter', 'arreter', 'stop', 'bloquer']):
        return fill_rouge_fonce
    
    # Rouge clair : Actions négatives importantes
    if any(mot in action_lower for mot in ['refusé', 'refuse', 'annulé', 'annule', 'pas intéressé', 
                                             'pas interesse', 'désabonné', 'desabonne', 'refuse']):
        return fill_rouge_clair
    
    # Violet : Pas de commande, n'appel plus
    if any(mot in action_lower for mot in ['pas de commande', 'n\'appel plus', 'n\'appelle plus', 
                                             'n\'appel plus le matin', 'ne commande plus', 'ne commandera pas',
                                             'n\'appel plus', 'nappel plus']):
        return fill_violet
    
    # Orange : Rappels programmés et appels non aboutis
    if any(mot in action_lower for mot in ['rappel plus tard', 'rappel demain', 'rappel dans', 
                                            'rappel lundi', 'rappel mardi', 'rappel mercredi', 
                                            'rappel jeudi', 'rappel vendredi', 'rappel samedi', 
                                            'rappel dimanche', 'rappel la semaine', 'rappel le mois',
                                            'ne decroche pas', 'ne décroche pas', 'ne decroche', 
                                            'pas de reponse', 'pas de réponse', 'pas reponse',
                                            'ne repond pas', 'ne répond pas', 'ne repond', 
                                            'sonne occupé', 'sonne occupe', 'ligne occupee',
                                            'ligne occupée', 'pas disponible', 'indisponible']):
        return fill_orange
    
    # Jaune : En attente, à confirmer
    if any(mot in action_lower for mot in ['en attente', 'à confirmer', 'a confirmer', 'en cours', 
                                            'en discussion', 'en négociation', 'en negociation', 
                                            'à vérifier', 'a verifier', 'en suspens',
                                            'en deplacement', 'en déplacement', 'deplacement',
                                            'déplacement', 'absent', 'pas la', 'pas present',
                                            'pas présent', 'en reunion', 'en réunion']):
        return fill_jaune
    
    # Vert clair : Commandes confirmées
    if any(mot in action_lower for mot in ['commande confirmée', 'commande confirmee', 'commande passée', 
                                            'commande passee', 'commande validée', 'commande validee', 
                                            'commande acceptée', 'commande acceptee', 'commande prise']):
        return fill_vert_clair
    
    # Vert foncé : Commandes livrées
    if any(mot in action_lower for mot in ['commande livrée', 'commande livree', 'livré', 'livre', 
                                            'terminé', 'termine', 'finalisé', 'finalise', 'complété', 
                                            'complete', 'fini', 'achevé', 'acheve']):
        return fill_vert_fonce
    
    # Bleu clair : Client intéressé
    if any(mot in action_lower for mot in ['client intéressé', 'client interesse', 'intéressé', 'interesse', 
                                            'à suivre', 'a suivre', 'suivi', 'prospect', 'potentiel', 
                                            'envisage', 'réfléchit', 'reflechit', 'considère', 'considere',
                                            'disponible', 'libre', 'ok', 'd\'accord', 'daccord']):
        return fill_bleu_clair
    
    # Rose : Informations, notes
    if any(mot in action_lower for mot in ['note', 'information', 'info', 'remarque', 'commentaire', 
                                             'observation', 'précision', 'precision', 'détail', 'detail']):
        return fill_rose
    
    # Par défaut : Si l'action contient des mots qui suggèrent une commande
    if any(mot in action_lower for mot in ['qté', 'qte', 'quantité', 'quantite', 'kg', 'unité', 'unite']):
        return fill_vert_clair
    
    return fill_blanc

def export_rapport_acm_reel(agence_id=None, date_debut=None, date_fin=None, zone=None):
    """Export réel du rapport ACM"""
    
    print("=" * 80)
    print("EXPORT RÉEL - RAPPORT ACM")
    print("=" * 80)
    print()
    
    # 1. Récupérer l'agence (par défaut: MARCHE HUITIEME qui utilise gestion commande)
    if agence_id:
        try:
            agence = Agence.objects.get(id_agence=agence_id)
        except Agence.DoesNotExist:
            print(f"✗ Erreur: Agence avec ID {agence_id} non trouvée.")
            return
    else:
        # Par défaut, utiliser MARCHE HUITIEME (agence qui utilise gestion commande)
        agence = Agence.objects.filter(nom_agence__icontains='huitieme').first()
        if not agence:
            # Si pas trouvée, prendre la première agence disponible
            agences = Agence.objects.all()
            if not agences.exists():
                print("✗ Erreur: Aucune agence trouvée dans la base de données.")
                return
            agence = agences.first()
        print(f"ℹ Utilisation de l'agence: {agence.nom_agence} (ID: {agence.id_agence})")
    
    # 2. Définir les dates
    if not date_debut:
        date_debut = (datetime.now() - timedelta(days=7)).date()
    else:
        try:
            date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        except ValueError:
            print(f"✗ Erreur: Format de date invalide pour date_debut. Utilisez YYYY-MM-DD")
            return
    
    if not date_fin:
        date_fin = datetime.now().date()
    else:
        try:
            date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            print(f"✗ Erreur: Format de date invalide pour date_fin. Utilisez YYYY-MM-DD")
            return
    
    print(f"ℹ Période: {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}")
    if zone:
        print(f"ℹ Zone: {zone}")
    print()
    
    # 3. Récupérer les clients
    clients_query = Client.objects.filter(agence=agence)
    if zone:
        clients_query = clients_query.filter(zone=zone)
    
    clients = clients_query.select_related('agence').order_by('ref', 'detail', 'intitule')
    print(f"✓ Nombre de clients: {clients.count()}")
    
    if clients.count() == 0:
        print("✗ Aucun client trouvé. Export annulé.")
        return
    
    # 4. Définir les plages horaires
    PLAGES_HORAIRES = [
        ('06h30-08h30', '06h30-08h30'),
        ('08h30-10h00', '08h30-10h00'),
        ('10h00-11h30', '10h00-11h30'),
        ('14h30-16h00', '14h30-16h00'),
        ('16h00-17h30', '16h00-17h30'),
    ]
    
    # 5. Récupérer les actions de suivi
    suivi_actions = SuiviClientAction.objects.filter(
        agence=agence,
        client__in=clients,
        date_action__date__range=[date_debut, date_fin]
    ).select_related('client').order_by('client', 'plage_horaire', '-date_action', '-heure_appel')
    
    print(f"✓ Nombre d'actions de suivi: {suivi_actions.count()}")
    
    # 6. Organiser les actions par client et par plage horaire
    actions_dict = {}
    for action in suivi_actions:
        client_id = action.client.id
        plage = action.plage_horaire or ''
        
        if client_id not in actions_dict:
            actions_dict[client_id] = {}
        
        if plage not in actions_dict[client_id]:
            actions_dict[client_id][plage] = []
        
        actions_dict[client_id][plage].append({
            'action': action.action or '',
            'heure_appel': action.heure_appel,
            'date_action': action.date_action
        })
    
    # 7. Récupérer les commandes
    commandes = Commande.objects.filter(
        agence=agence,
        client__in=clients,
        date__range=[date_debut, date_fin]
    ).select_related('client', 'article').order_by('client', 'date', 'heure')
    
    print(f"✓ Nombre de commandes: {commandes.count()}")
    
    # 8. Grouper les commandes par client, date et heure
    commandes_dict = defaultdict(lambda: defaultdict(list))
    
    for cmd in commandes:
        key = (cmd.client.id, cmd.date, cmd.heure)
        commandes_dict[cmd.client.id][key].append({
            'article': cmd.article.designation if cmd.article else 'N/A',
            'quantite': float(cmd.quantite),
            'prix_total': float(cmd.prix_total),
        })
    
    # 9. Convertir en format final
    commandes_grouped = {}
    for client_id, commandes_par_client in commandes_dict.items():
        commandes_grouped[client_id] = []
        for (client_id_key, date, heure), articles in commandes_par_client.items():
            total_commande = sum(art['prix_total'] for art in articles)
            articles_list = [f"{art['article']} (Qté: {art['quantite']})" for art in articles]
            commandes_grouped[client_id].append({
                'date': date.strftime('%d/%m/%Y'),
                'heure': heure.strftime('%H:%M') if heure else '',
                'articles': ', '.join(articles_list),
                'total': total_commande,
            })
    
    # 10. Calculer les totaux par client et organiser par plage horaire
    clients_data = []
    for client in clients:
        client_actions = actions_dict.get(client.id, {})
        commandes_client = commandes_grouped.get(client.id, [])
        total_commandes = sum(cmd['total'] for cmd in commandes_client)
        
        # Organiser les actions et commandes par plage horaire
        actions_par_plage = {}
        commandes_par_plage = {}
        totaux_par_plage = {}
        
        for plage_code, plage_label in PLAGES_HORAIRES:
            actions_list = client_actions.get(plage_code, [])
            actions_par_plage[plage_code] = actions_list
            commandes_par_plage[plage_code] = []
            total_plage = 0
            
            # Déterminer la plage horaire selon l'heure de la commande
            for cmd in commandes_client:
                if cmd.get('heure'):
                    try:
                        heure_str = cmd['heure']
                        if isinstance(heure_str, str):
                            heure_obj = datetime.strptime(heure_str, '%H:%M').time()
                        else:
                            heure_obj = heure_str
                        
                        # Déterminer la plage horaire de la commande
                        plage_commande = None
                        if time(6, 30) <= heure_obj < time(8, 30):
                            plage_commande = '06h30-08h30'
                        elif time(8, 30) <= heure_obj < time(10, 0):
                            plage_commande = '08h30-10h00'
                        elif time(10, 0) <= heure_obj < time(11, 30):
                            plage_commande = '10h00-11h30'
                        elif time(14, 30) <= heure_obj < time(16, 0):
                            plage_commande = '14h30-16h00'
                        elif time(16, 0) <= heure_obj < time(17, 30):
                            plage_commande = '16h00-17h30'
                        
                        # Si la commande appartient à cette plage horaire
                        if plage_commande == plage_code:
                            commandes_par_plage[plage_code].append(cmd)
                            total_plage += cmd['total']
                    except:
                        pass
            
            totaux_par_plage[plage_code] = total_plage
        
        # Créer une ligne par client avec toutes les plages horaires
        clients_data.append({
            'id': client.id,
            'detail': client.detail or '',
            'ref': client.ref or '',
            'potentiel': client.potentiel or '',
            'nom': client.intitule,
            'telephone': client.telephone,
            'actions_par_plage': actions_par_plage,
            'commandes_par_plage': commandes_par_plage,
            'totaux_par_plage': totaux_par_plage,
            'total': total_commandes,
        })
    
    print(f"✓ Nombre de clients avec données: {len(clients_data)}")
    print()
    
    # 11. Créer le fichier Excel (même logique que export_rapport_acm_excel)
    print("Création du fichier Excel...")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport ACM"
        
        # Styles
        header_fill = PatternFill(start_color="06beb6", end_color="48b1bf", fill_type="solid")
        header_font_bold = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-tête
        ws.merge_cells('A1:H1')
        ws['A1'] = f"RAPPORT ACM - {agence.nom_agence}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Période : {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Styles pour les couleurs
        fill_rouge_fonce = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        fill_rouge_clair = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        fill_violet = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
        fill_orange = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        fill_jaune = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
        fill_vert_clair = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        fill_vert_fonce = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        fill_bleu_clair = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        fill_rose = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
        fill_blanc = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        row_num = 3
        if zone:
            ws.merge_cells(f'A{row_num}:H{row_num}')
            ws[f'A{row_num}'] = f"Zone : {zone}"
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center')
            row_num += 1
        
        # Légende des couleurs
        ws.merge_cells(f'A{row_num}:H{row_num}')
        ws[f'A{row_num}'] = "LÉGENDE DES COULEURS :"
        ws[f'A{row_num}'].font = Font(bold=True, size=10, name='Calibri')
        ws[f'A{row_num}'].alignment = Alignment(horizontal='left', vertical='center')
        row_num += 1
        
        legend_items = [
            ('Rouge foncé', fill_rouge_fonce, 'Ne plus rappeler'),
            ('Rouge clair', fill_rouge_clair, 'Refusé/Annulé'),
            ('Violet', fill_violet, 'Pas de commande'),
            ('Orange', fill_orange, 'Rappel programmé'),
            ('Jaune', fill_jaune, 'En attente'),
            ('Vert clair', fill_vert_clair, 'Commande confirmée'),
            ('Vert foncé', fill_vert_fonce, 'Commande livrée'),
            ('Bleu clair', fill_bleu_clair, 'Client intéressé'),
        ]
        
        for line_idx in range(2):
            for i in range(4):
                idx = line_idx * 4 + i
                if idx < len(legend_items):
                    label, color_fill, description = legend_items[idx]
                    col_label = 1 + i * 2
                    ws.cell(row=row_num, column=col_label, value=label).fill = color_fill
                    ws.cell(row=row_num, column=col_label).font = Font(size=9, name='Calibri')
                    ws.cell(row=row_num, column=col_label).border = border
                    ws.cell(row=row_num, column=col_label + 1, value=description).font = Font(size=9, name='Calibri')
                    ws.cell(row=row_num, column=col_label + 1).border = border
            row_num += 1
        
        row_num += 1
        
        # En-têtes de colonnes
        start_row = row_num
        headers_fixes = ['DETAIL', 'REF', 'POTENTIEL /5', 'NOMS', 'TELEPHONES']
        
        col = 1
        for header in headers_fixes:
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            col += 1
        
        # En-têtes des plages horaires
        for plage_code, plage_label in PLAGES_HORAIRES:
            ws.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col+2)}{start_row}')
            cell = ws.cell(row=start_row, column=col)
            cell.value = plage_label
            cell.font = header_font_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            col += 3
        
        # Colonne Total
        cell = ws.cell(row=start_row, column=col)
        cell.value = 'TOTAL'
        cell.font = header_font_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        
        # Deuxième ligne : sous-en-têtes
        start_row2 = start_row + 1
        col = len(headers_fixes) + 1
        for _ in PLAGES_HORAIRES:
            ws.cell(row=start_row2, column=col, value='Heure').font = header_font_bold
            ws.cell(row=start_row2, column=col).fill = header_fill
            ws.cell(row=start_row2, column=col).border = border
            col += 1
            
            ws.cell(row=start_row2, column=col, value='Commandes').font = header_font_bold
            ws.cell(row=start_row2, column=col).fill = header_fill
            ws.cell(row=start_row2, column=col).border = border
            col += 1
            
            ws.cell(row=start_row2, column=col, value='Total').font = header_font_bold
            ws.cell(row=start_row2, column=col).fill = header_fill
            ws.cell(row=start_row2, column=col).border = border
            col += 1
        
        ws.row_dimensions[start_row].height = 25
        ws.row_dimensions[start_row2].height = 25
        
        # Calculer les statistiques par vague
        stats_par_vague = {}
        for plage_code, plage_label in PLAGES_HORAIRES:
            stats_par_vague[plage_code] = {
                'nb_appels': 0,
                'nb_reponses': 0,
                'nb_commandes': 0,
                'total_commandes': 0.0
            }
        
        # Police pour les cellules de données
        data_font = Font(size=10, name='Calibri')
        
        # Données des clients
        row = start_row2 + 1
        for idx, client_data in enumerate(clients_data):
            try:
                actions_par_plage = client_data.get('actions_par_plage', {})
                total_client = float(client_data.get('total', 0) or 0)
                
                col = 1
                # Colonnes fixes
                ws.cell(row=row, column=col, value=client_data.get('detail', '')).border = border
                ws.cell(row=row, column=col).font = data_font
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
                col += 1
                
                ws.cell(row=row, column=col, value=client_data.get('ref', '')).border = border
                ws.cell(row=row, column=col).font = data_font
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
                col += 1
                
                ws.cell(row=row, column=col, value=client_data.get('potentiel', '')).border = border
                ws.cell(row=row, column=col).font = data_font
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                col += 1
                
                ws.cell(row=row, column=col, value=client_data.get('nom', '')).border = border
                ws.cell(row=row, column=col).font = Font(bold=True, size=10, name='Calibri')
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
                col += 1
                
                ws.cell(row=row, column=col, value=client_data.get('telephone', '')).border = border
                ws.cell(row=row, column=col).font = data_font
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center')
                col += 1
                
                # Pour chaque plage horaire : 3 colonnes (Heure, Commandes, Total)
                for plage_code, plage_label in PLAGES_HORAIRES:
                    actions_list = actions_par_plage.get(plage_code, [])
                    commandes_plage = client_data.get('commandes_par_plage', {}).get(plage_code, [])
                    
                    # Récupérer les heures et actions
                    heures_appels = []
                    actions_texts = []
                    for action in actions_list:
                        if action.get('heure_appel'):
                            if hasattr(action['heure_appel'], 'strftime'):
                                heures_appels.append(action['heure_appel'].strftime('%H:%M'))
                            else:
                                heures_appels.append(str(action['heure_appel']))
                        if action.get('action'):
                            actions_texts.append(action['action'])
                            stats_par_vague[plage_code]['nb_reponses'] += 1
                    
                    stats_par_vague[plage_code]['nb_appels'] += len(actions_list)
                    
                    # Compter les commandes
                    for action_text in actions_texts:
                        if any(mot in action_text.lower() for mot in ['commande', 'qte', 'quantité', 'kg', 'unité']):
                            stats_par_vague[plage_code]['nb_commandes'] += 1
                    
                    if commandes_plage:
                        stats_par_vague[plage_code]['nb_commandes'] += len(commandes_plage)
                    
                    # Colonne 1: Heure
                    all_heures = []
                    if heures_appels:
                        all_heures.extend(heures_appels)
                    for cmd in commandes_plage:
                        if cmd.get('heure') and cmd['heure'] not in all_heures:
                            all_heures.append(cmd['heure'])
                    
                    heures_str = '\n'.join(all_heures) if all_heures else ''
                    cell_heure = ws.cell(row=row, column=col, value=heures_str)
                    cell_heure.border = border
                    cell_heure.font = Font(size=10, name='Calibri', bold=True, color='2e7d32')
                    cell_heure.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                    cell_heure.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    col += 1
                    
                    # Colonne 2: Commandes
                    all_texts = []
                    if actions_texts:
                        all_texts.extend(actions_texts)
                    for cmd in commandes_plage:
                        cmd_text = f"{cmd.get('articles', '')} ({cmd.get('date', '')} {cmd.get('heure', '')})"
                        if cmd_text.strip():
                            all_texts.append(cmd_text)
                    
                    actions_str = '\n'.join(all_texts) if all_texts else ''
                    cell_cmd = ws.cell(row=row, column=col, value=actions_str)
                    cell_cmd.border = border
                    cell_cmd.font = data_font
                    cell_cmd.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    if actions_texts:
                        action_fill = get_action_color(actions_texts[0], fill_rouge_fonce, fill_rouge_clair, 
                                                      fill_violet, fill_orange, fill_jaune, fill_vert_clair,
                                                      fill_vert_fonce, fill_bleu_clair, fill_rose, fill_blanc)
                        cell_cmd.fill = action_fill
                    elif commandes_plage:
                        cell_cmd.fill = fill_vert_clair
                    else:
                        cell_cmd.fill = fill_blanc
                    col += 1
                    
                    # Colonne 3: Total
                    totaux_par_plage = client_data.get('totaux_par_plage', {})
                    total_plage = float(totaux_par_plage.get(plage_code, 0) or 0)
                    total_plage_str = f"{total_plage:.0f} FCFA" if total_plage > 0 else ''
                    cell_total_plage = ws.cell(row=row, column=col, value=total_plage_str)
                    cell_total_plage.border = border
                    cell_total_plage.font = Font(bold=True, size=10, name='Calibri')
                    cell_total_plage.alignment = Alignment(horizontal='right', vertical='center')
                    cell_total_plage.fill = fill_blanc
                    col += 1
                
                # Colonne Total général
                total_str = f"{total_client:.0f} FCFA" if total_client > 0 else ''
                ws.cell(row=row, column=col, value=total_str).border = border
                ws.cell(row=row, column=col).font = Font(bold=True, size=11, name='Calibri', color='1b5e20')
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=row, column=col).fill = PatternFill(start_color="E8F5E9", end_color="C8E6C9", fill_type="solid")
                
                # Ajouter le total aux statistiques
                if total_client > 0:
                    vagues_avec_actions = [p[0] for p in PLAGES_HORAIRES if actions_par_plage.get(p[0])]
                    if vagues_avec_actions:
                        total_par_vague = total_client / len(vagues_avec_actions)
                        for vague_code in vagues_avec_actions:
                            stats_par_vague[vague_code]['total_commandes'] += total_par_vague
                
                row += 1
                
            except Exception as client_error:
                print(f"✗ Erreur client {idx + 1}: {str(client_error)}")
                import traceback
                traceback.print_exc()
                continue
        
        # Ajouter les statistiques par vague (même logique que dans views.py)
        stats_row = row + 2
        stats_start_col = 1
        
        # Titre du tableau de statistiques - Première ligne
        # Colonne A avec "RAPPORT JOURNALIER-DATE" qui span 2 lignes (rowspan=2)
        # Écrire AVANT de fusionner
        cell_title = ws.cell(row=stats_row, column=stats_start_col)
        cell_title.value = 'RAPPORT JOURNALIER-DATE'
        cell_title.font = Font(bold=True, size=12, name='Calibri')
        cell_title.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        cell_title.alignment = Alignment(horizontal='left', vertical='center')
        cell_title.border = border
        ws.merge_cells(f'{get_column_letter(stats_start_col)}{stats_row}:{get_column_letter(stats_start_col)}{stats_row + 1}')
        
        # En-têtes des vagues
        col = stats_start_col + 1
        vague_names = ['PREMIERE', 'DEUXIEME VAGUE', 'TROISIEME VAGUE', 'QUATRIEME VAGUE', 'CINQUIEME VAGUE']
        for idx, (plage_code, plage_label) in enumerate(PLAGES_HORAIRES):
            vague_name = vague_names[idx] if idx < len(vague_names) else plage_label
            # Écrire AVANT de fusionner
            cell = ws.cell(row=stats_row, column=col)
            cell.value = vague_name
            cell.font = header_font_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.merge_cells(f'{get_column_letter(col)}{stats_row}:{get_column_letter(col)}{stats_row + 1}')
            col += 1
        
        # Colonne Total
        cell = ws.cell(row=stats_row, column=col)
        cell.value = 'TOTAL COMMANDES JOURNALIERES'
        cell.font = header_font_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.merge_cells(f'{get_column_letter(col)}{stats_row}:{get_column_letter(col)}{stats_row + 1}')
        # Ne pas incrémenter stats_row ici, on l'utilisera pour les lignes de données
        stats_data_row = stats_row + 1  # Ligne pour les données (après l'en-tête)
        
        # Lignes de statistiques
        metrics = [
            ('Nombre d\'app', 'nb_appels', False),
            ('Nombre de rép', 'nb_reponses', False),
            ('Nombre de Comm', 'nb_commandes', True),
            ('Total comman', 'total_commandes', True),
        ]
        
        fill_rose = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        
        for metric_label, metric_key, use_rose_label in metrics:
            col = stats_start_col
            # Écrire dans la ligne de données (stats_data_row)
            # La colonne stats_start_col est fusionnée verticalement sur stats_row et stats_row + 1
            # stats_data_row = stats_row + 1, donc on est dans la deuxième ligne de la fusion
            # On ne peut pas écrire dans une MergedCell, donc on doit utiliser la cellule principale
            # La cellule principale est à stats_row (première ligne de la fusion)
            # Mais on veut écrire dans stats_data_row, donc on doit "unmerge" temporairement ou utiliser une autre approche
            # Solution: écrire dans la cellule principale (stats_row) pour toutes les lignes
            # car elles partagent la même cellule fusionnée
            # Mais non, chaque ligne de métrique doit avoir sa propre valeur
            # Solution: ne pas fusionner cette colonne pour les lignes de données
            # Ou utiliser une approche différente: écrire dans stats_row pour la première métrique,
            # puis dans stats_row + 1 pour la deuxième, etc.
            # Mais comme la colonne est fusionnée, on ne peut écrire que dans la cellule principale
            # La meilleure solution est de ne pas fusionner cette colonne pour les lignes de données
            # ou d'utiliser une colonne différente
            # Utilisons la même approche que dans views.py: écrire dans stats_row (première ligne de la fusion)
            # pour toutes les métriques, car elles partagent la même cellule fusionnée
            # Mais cela ne fonctionnera pas car chaque métrique doit avoir sa propre valeur
            # Solution finale: ne pas fusionner cette colonne pour les lignes de données
            # On va "unmerge" cette colonne pour les lignes de données
            # Ou mieux: utiliser une approche différente - écrire dans stats_data_row qui est stats_row + 1
            # et utiliser la cellule principale si c'est une MergedCell
            cell = ws.cell(row=stats_data_row, column=col)
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                # Utiliser la cellule principale (première ligne de la fusion)
                cell = ws.cell(row=stats_row, column=col)
            cell.value = metric_label
            cell.font = Font(bold=True, size=10, name='Calibri')
            if use_rose_label:
                cell.fill = fill_rose
            else:
                cell.fill = fill_blanc
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            col += 1
            
            for plage_code, plage_label in PLAGES_HORAIRES:
                value = stats_par_vague[plage_code][metric_key]
                if metric_key == 'total_commandes':
                    value_str = f"{value:,.0f}".replace(',', '.') if value > 0 else ""
                else:
                    value_str = str(int(value)) if value > 0 else ""
                
                cell_data = ws.cell(row=stats_data_row, column=col)
                from openpyxl.cell.cell import MergedCell
                if isinstance(cell_data, MergedCell):
                    # Utiliser la cellule principale (première ligne de la fusion)
                    cell_data = ws.cell(row=stats_row, column=col)
                cell_data.value = value_str
                cell_data.border = border
                cell_data.font = data_font
                if metric_key == 'total_commandes':
                    cell_data.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell_data.alignment = Alignment(horizontal='center', vertical='center')
                if metric_key == 'nb_commandes':
                    cell_data.fill = fill_rose
                else:
                    cell_data.fill = fill_blanc
                col += 1
            
            # Total journalier
            if metric_key == 'total_commandes':
                total_journalier = sum(stats_par_vague[p[0]][metric_key] for p in PLAGES_HORAIRES)
                value_str = f"{total_journalier:,.0f}".replace(',', '.') if total_journalier > 0 else ""
            else:
                total_journalier = sum(stats_par_vague[p[0]][metric_key] for p in PLAGES_HORAIRES)
                value_str = str(int(total_journalier)) if total_journalier > 0 else ""
            
            cell_total = ws.cell(row=stats_data_row, column=col)
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell_total, MergedCell):
                # Utiliser la cellule principale (première ligne de la fusion)
                cell_total = ws.cell(row=stats_row, column=col)
            cell_total.value = value_str
            cell_total.border = border
            cell_total.font = Font(bold=True, size=10, name='Calibri')
            if metric_key == 'total_commandes':
                cell_total.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell_total.alignment = Alignment(horizontal='center', vertical='center')
            if metric_key == 'nb_commandes':
                cell_total.fill = fill_rose
            else:
                cell_total.fill = fill_blanc
            
            stats_data_row += 1
        
        # Ajuster les largeurs de colonnes
        column_widths = [30, 18, 15, 30, 25]
        for _ in PLAGES_HORAIRES:
            column_widths.extend([15, 40, 18])
        column_widths.append(18)
        
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Ajuster la hauteur des lignes
        for r in range(start_row2 + 1, row + 1):
            ws.row_dimensions[r].height = 30
        
        # Sauvegarder le fichier
        filename = f"rapport_acm_{agence.nom_agence.replace(' ', '_')}_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}_{datetime.now().strftime('%H%M%S')}.xlsx"
        wb.save(filename)
        
        print(f"✓ Fichier Excel créé: {filename}")
        print(f"✓ Chemin complet: {os.path.abspath(filename)}")
        print()
        print("=" * 80)
        print("EXPORT TERMINÉ AVEC SUCCÈS")
        print("=" * 80)
        
    except Exception as e:
        print(f"✗ Erreur lors de la création du fichier Excel: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Export réel du rapport ACM')
    parser.add_argument('--date-debut', type=str, help='Date de début (YYYY-MM-DD)')
    parser.add_argument('--date-fin', type=str, help='Date de fin (YYYY-MM-DD)')
    parser.add_argument('--zone', type=str, help='Zone à filtrer')
    parser.add_argument('--agence', type=int, help='ID de l\'agence')
    
    args = parser.parse_args()
    
    export_rapport_acm_reel(
        agence_id=args.agence,
        date_debut=args.date_debut,
        date_fin=args.date_fin,
        zone=args.zone
    )

