from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Max, Q, F
from django.db import models, connection, IntegrityError, transaction
from django.utils import timezone
from django.conf import settings
from datetime import datetime, time, timedelta
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal
from .models import (
    Agence, Compte, Employe, Caisse, SessionCaisse, Client, InventaireStock, LigneInventaireStock,
    StatistiqueVente, MouvementStock, Article, TypeVente, FactureVente, LigneFactureVente, DocumentVente,
    FactureTemporaire, Famille, Fournisseur, MouvementStock, PlanComptable, PlanTiers, CodeJournaux,
    TauxTaxe, FactureAchat, LigneFactureAchat, FactureTransfert, LigneFactureTransfert,
    Commande, Livraison, FactureCommande, StatistiqueClient, DocumentCommande, LigneCommande, Notification,
    Livreur, SuiviClientAction,Depense,ChiffreAffaire,Tresorerie,Beneficiaire
)
from .decorators import (
    require_stock_modify_access,
    require_commandes_feature, require_module_access, require_compte_type,
    require_caisse_access, require_stock_access, require_comptes_access,
    get_user_compte, get_user_livreur, require_comptable_access,
)
from .permissions_utils import (
    filter_commandes_by_user, filter_suivi_client_by_user, filter_livraisons_by_user
)

def get_user_agence(request):
    """Récupérer l'agence de l'utilisateur connecté"""
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
        print(f"[ALERTE] get_user_agence: {request.user.username} -> {compte.agence.nom_agence}")
        return compte.agence
    except Compte.DoesNotExist:
        print(f"[ALERTE] get_user_agence: {request.user.username} -> AUCUN COMPTE")
        return None

def login_caisse(request):
    """Page de connexion pour la gestion de caisse"""
    # Toujours afficher la page de connexion, même si l'utilisateur est déjà connecté
    # Cela permet de forcer la reconnexion depuis la page d'accueil
    if request.user.is_authenticated and request.method == 'GET':
        # Déconnecter l'utilisateur pour forcer la reconnexion
        logout(request)
        request.session.flush()
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                # Vérifier que l'utilisateur a un compte lié à une agence et de type caissier
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        # Vérifier que le type de compte est caissier ou admin
                        if compte.type_compte not in ['caissier', 'admin']:
                            type_display = compte.get_type_compte_display()
                            messages.error(
                                request, 
                                f'Accès refusé. Seuls les caissiers et les administrateurs peuvent accéder au module Caisse. '
                                f'Votre type de compte ({type_display}) n\'est pas autorisé.'
                            )
                            return render(request, 'supermarket/caisse/login.html', {'agences': Agence.objects.all()})
                        
                        login(request, user)
                        # Stocker l'agence dans la session
                        request.session['agence_id'] = compte.agence.id_agence
                        request.session['agence_nom'] = compte.agence.nom_agence
                        messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                        return redirect('dashboard_caisse')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Compte non trouvé ou inactif.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')
    
    # Récupérer les agences disponibles pour affichage
    agences = Agence.objects.all()
    context = {
        'agences': agences
    }
    return render(request, 'supermarket/caisse/login.html', context)

def logout_caisse(request):
    """Déconnexion de la gestion de caisse"""
    if request.user.is_authenticated:
        logout(request)
        # Nettoyer la session
        request.session.flush()
        messages.info(request, 'Vous avez été déconnecté avec succès.')
    return redirect('index')

def index(request):
    return render(request, 'supermarket/index.html')


# ==================== EXPORT/IMPORT DES DONNÉES ====================

def export_data_view(request):
    """Vue pour exporter les données (sans authentification requise)"""
    from .export_import import export_data, DecimalEncoder
    from django.http import JsonResponse
    from datetime import datetime
    import json
    
    if request.method == 'GET':
        # Afficher la page de sélection d'agence pour l'export
        agence = None
        try:
            agence = get_user_agence(request)
        except:
            pass
        
        agences = Agence.objects.all().order_by('nom_agence')
        
        context = {
            'agence': agence,
            'agences': agences,
        }
        return render(request, 'supermarket/export_import/export_data.html', context)
    
    if request.method == 'POST':
        try:
            # Récupérer l'agence sélectionnée
            agence_id = request.POST.get('agence_id')
            if agence_id:
                try:
                    agence_id = int(agence_id)
                    agence_obj = Agence.objects.get(id_agence=agence_id)
                except (ValueError, Agence.DoesNotExist):
                    messages.error(request, 'Agence invalide.')
                    return redirect('export_import_page')
            else:
                # Si aucune agence sélectionnée, essayer de récupérer celle de l'utilisateur
                try:
                    agence = get_user_agence(request)
                    agence_id = agence.id_agence if agence else None
                except:
                    agence_id = None
            
            # Exporter les données
            data = export_data(agence_id=agence_id, include_users=True)
            
            # Créer la réponse JSON
            response = JsonResponse(data, encoder=DecimalEncoder, json_dumps_params={'indent': 2, 'ensure_ascii': False})
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            agence_nom = Agence.objects.get(id_agence=agence_id).nom_agence.replace(' ', '_') if agence_id else 'toutes'
            response['Content-Disposition'] = f'attachment; filename="export_erp_{agence_nom}_{timestamp}.json"'
            
            return response
        
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'export: {str(e)}')
            return redirect('export_import_page')


def import_data_view(request):
    """Vue pour importer les données (sans authentification requise)"""
    from .export_import import import_data
    from django.http import JsonResponse
    import json
    
    if request.method == 'GET':
        # Afficher la page d'import
        agence = None
        try:
            agence = get_user_agence(request)
        except:
            pass
        
        agences = Agence.objects.all().order_by('nom_agence')
        
        context = {
            'agence': agence,
            'agences': agences,
        }
        return render(request, 'supermarket/export_import/import_data.html', context)
    
    if request.method == 'POST':
        try:
            # Récupérer l'agence cible pour l'import
            agence_id = request.POST.get('agence_id')
            if agence_id:
                try:
                    agence_id = int(agence_id)
                    agence_obj = Agence.objects.get(id_agence=agence_id)
                    agence = agence_obj
                except (ValueError, Agence.DoesNotExist):
                    messages.error(request, 'Agence invalide.')
                    return redirect('import_data')
            else:
                # Si aucune agence sélectionnée, essayer de récupérer celle de l'utilisateur
                try:
                    agence = get_user_agence(request)
                    agence_id = agence.id_agence if agence else None
                except:
                    agence = None
                    agence_id = None
            
            # Récupérer le fichier
            if 'file' not in request.FILES:
                messages.error(request, 'Aucun fichier fourni.')
                return redirect('import_data')
            
            file = request.FILES['file']
            
            # Lire le contenu JSON
            try:
                data = json.loads(file.read().decode('utf-8'))
            except json.JSONDecodeError:
                messages.error(request, 'Le fichier n\'est pas un JSON valide.')
                return redirect('import_data')
            
            # Option pour effacer les données existantes
            clear_existing = request.POST.get('clear_existing', 'off') == 'on'
            
            if clear_existing:
                # Avertissement supplémentaire
                if request.POST.get('confirm_clear') != 'yes':
                    messages.warning(request, 'Vous devez confirmer la suppression des données existantes.')
                    agences = Agence.objects.all().order_by('nom_agence')
                    context = {
                        'agence': agence,
                        'agences': agences,
                        'file_content': data,
                        'confirm_needed': True,
                        'selected_agence_id': agence_id,
                    }
                    return render(request, 'supermarket/export_import/import_data.html', context)
            
            # Importer les données
            stats = import_data(data, agence_id=agence_id, clear_existing=clear_existing)
            
            # Afficher les résultats
            agence_nom = agence.nom_agence if agence else "toutes les agences"
            success_msg = f"Import pour l'agence '{agence_nom}' : "
            success_msg += f"Agences: {stats['agences']}, "
            success_msg += f"Familles: {stats['familles']}, "
            success_msg += f"Articles: {stats['articles']}, "
            success_msg += f"Clients: {stats['clients']}, "
            success_msg += f"Fournisseurs: {stats['fournisseurs']}, "
            success_msg += f"Factures Vente: {stats['factures_vente']}, "
            success_msg += f"Factures Achat: {stats['factures_achat']}, "
            success_msg += f"Mouvements Stock: {stats['mouvements_stock']}"
            
            if stats['errors']:
                error_msg = f" {len(stats['errors'])} erreur(s) rencontrée(s)."
                # Afficher les 10 premières erreurs pour diagnostic
                if len(stats['errors']) > 0:
                    error_details = " | ".join(stats['errors'][:10])
                    if len(stats['errors']) > 10:
                        error_details += f" ... et {len(stats['errors']) - 10} autres erreurs"
                    messages.error(request, success_msg + error_msg + f" Détails: {error_details}")
                else:
                    messages.warning(request, success_msg + error_msg)
            else:
                messages.success(request, success_msg)
            
            return redirect('import_data')
        
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'import: {str(e)}')
            return redirect('import_data')


def export_import_page(request):
    """Page principale pour l'export/import (sans authentification requise)"""
    agence = None
    try:
        agence = get_user_agence(request)
    except:
        pass
    
    # Récupérer toutes les agences pour la sélection
    agences = Agence.objects.all().order_by('nom_agence')
    
    # L'export/import est accessible à tous (même sans authentification)
    can_export_import = True
    
    context = {
        'agence': agence,
        'agences': agences,
        'can_export_import': can_export_import,
    }
    return render(request, 'supermarket/export_import/export_import.html', context)



@login_required
@require_caisse_access
def dashboard_caisse(request):
    # Récupérer l'agence de l'utilisateur connecté
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    # Récupérer le compte de l'utilisateur
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
        return redirect('logout_caisse')
    
    caisse_ouverte = False

    try:
        aujourd_hui = timezone.now().date()

        session_caisse = SessionCaisse.objects.filter(

            agence=agence, 

            date_ouverture__date=aujourd_hui, 

            statut='ouverte'

        ).first()

        caisse_ouverte = session_caisse is not None

    except:
        pass
    
    chiffre_affaires = 0
    nombre_ventes = 0
    tickets_attente = 0

    try:
        # Récupérer les ventes de la session active si elle existe (toutes dates confondues)
        if session_caisse:
            ventes_jour = FactureVente.objects.filter(
                agence=agence, 
                session_caisse=session_caisse
            )
        else:
            # Sinon, récupérer les ventes du jour liées aux sessions de caisse
            ventes_jour = FactureVente.objects.filter(
                agence=agence, 
                date=aujourd_hui,
                session_caisse__isnull=False
            )
        chiffre_affaires = ventes_jour.aggregate(total=Sum('nette_a_payer'))['total'] or 0

        nombre_ventes = ventes_jour.count()

        # Compter les factures temporaires UNIQUEMENT pour la session courante ouverte
        if session_caisse:
            tickets_attente = FactureTemporaire.objects.filter(session_caisse=session_caisse).count()
        else:
            tickets_attente = 0

    except:

        pass
    

    
    
    # Récupérer l'employé associé
    employe = None
    try:
        employe = Employe.objects.get(compte=compte)
    except Employe.DoesNotExist:
        pass
    
    return render(request, 'supermarket/caisse/dashboard.html', {
        'agence': agence, 
        'compte': compte,
        'employe': employe,
        'caisse_ouverte': caisse_ouverte, 
        'chiffre_affaires': chiffre_affaires,
        'nombre_ventes': nombre_ventes, 
        'tickets_attente': tickets_attente, 
        'session_caisse': session_caisse
    })

@login_required
@require_caisse_access
def dashboard_kpis_api(request):
    print(f"[ALERTE] DASHBOARD_KPIS_API: {request.user.username}")
    
    agence = get_user_agence(request)
    if not agence:
        print(f"[ALERTE] DASHBOARD_KPIS_API: Aucune agence pour {request.user.username}")

        return JsonResponse({

            'chiffre_affaires': 0, 

            'nombre_ventes': 0, 

            'tickets_attente': 0, 

            'caisse_ouverte': False,

            'sessions_info': [],

            'premiere_ouverture': None

        })
    
    print(f"[ALERTE] DASHBOARD_KPIS_API: Agence {agence.nom_agence}")
    
    
    
    aujourd_hui = timezone.now().date()

    
    
    # Calculer les KPIs de manière simple d'abord

    chiffre_affaires = 0

    nombre_ventes = 0

    tickets_attente = 0
    
    caisse_ouverte = False

    

    try:

        # Récupérer la session de caisse active
        session_ouverte = SessionCaisse.objects.filter(
            agence=agence, 
            date_ouverture__date=aujourd_hui, 
            statut='ouverte'
        ).first()
        
        if session_ouverte:
            print(f"[ALERTE] DASHBOARD_KPIS_API: Session trouvée {session_ouverte.id}")
            # Récupérer toutes les factures de la session active (toutes dates confondues)
            ventes_jour = FactureVente.objects.filter(
                agence=agence, 
                session_caisse=session_ouverte
            )
            print(f"[ALERTE] DASHBOARD_KPIS_API: {ventes_jour.count()} factures trouvées")
            
            # Si aucune facture liée à la session, récupérer toutes les factures du jour et les lier
            if ventes_jour.count() == 0:
                print(f"[ALERTE] DASHBOARD_KPIS_API: Aucune facture liée à la session, recherche des factures orphelines")
                factures_orphelines = FactureVente.objects.filter(
                    agence=agence,
                    date=aujourd_hui,
                    session_caisse__isnull=True
                )
                
                if factures_orphelines.count() > 0:
                    print(f"[ALERTE] DASHBOARD_KPIS_API: {factures_orphelines.count()} factures orphelines trouvées, liaison à la session")
                    # Lier ces factures à la session actuelle
                    factures_orphelines.update(session_caisse=session_ouverte)
                    
                    # Récupérer maintenant toutes les factures de la session
                    ventes_jour = FactureVente.objects.filter(
                        agence=agence, 
                        session_caisse=session_ouverte
                    )
                    print(f"[ALERTE] DASHBOARD_KPIS_API: {ventes_jour.count()} factures après liaison")
        else:
            print(f"[ALERTE] DASHBOARD_KPIS_API: Aucune session active")
            # Aucune session active, donc pas de ventes
            ventes_jour = FactureVente.objects.none()
        
        
        chiffre_affaires = ventes_jour.aggregate(total=Sum('nette_a_payer'))['total'] or 0

        nombre_ventes = ventes_jour.count()

        # Compter les factures temporaires en attente UNIQUEMENT pour la session courante
        if session_ouverte:
            tickets_attente = FactureTemporaire.objects.filter(session_caisse=session_ouverte).count()
        else:
            tickets_attente = 0
        
        
    except Exception as e:
    

        pass
    

    # Vérifier le statut de la caisse (session_ouverte déjà récupérée plus haut)
    try:

        caisse_ouverte = session_ouverte is not None
        
        # Récupérer toutes les sessions pour l'affichage
        sessions_caisse = SessionCaisse.objects.filter(

            agence=agence, 

            date_ouverture__date=aujourd_hui

        ).order_by('date_ouverture')
        
        
        
    except Exception as e:

        pass
    
    
    # Préparer les informations des sessions

    sessions_info = []

    try:

        for i, session in enumerate(sessions_caisse[:2]):  # Max 2 sessions


            session_data = {

                'numero': i + 1,

                'ouverture': session.date_ouverture.strftime('%H:%M') if session.date_ouverture else '',

                'fermeture': session.date_fermeture.strftime('%H:%M') if session.date_fermeture else '',

                'ouverte_par': session.employe.compte.nom_complet if session.employe and hasattr(session.employe, 'compte') else 'Non spécifié',

                'fermee_par': session.employe.compte.nom_complet if session.employe and session.date_fermeture and hasattr(session.employe, 'compte') else '',


                'statut': 'Ouverte' if session.statut == 'ouverte' else 'Fermée'

            }

            sessions_info.append(session_data)

    except Exception as e:

        print(f"DEBUG: Erreur sessions_info: {str(e)}")
    

    
    
    result = {

        'chiffre_affaires': float(chiffre_affaires),

        'nombre_ventes': nombre_ventes,

        'tickets_attente': tickets_attente,

        'caisse_ouverte': caisse_ouverte,

        'sessions_info': sessions_info,


        'premiere_ouverture': None,

        'nombre_sessions': len(sessions_info)

    }

    
    
    print(f"DEBUG: Résultat final: {result}")

    
    
    return JsonResponse(result)



def generate_ticket_number(agence):

    """Générer un numéro de ticket auto-incrémenté"""

    try:

        # Obtenir la date actuelle

        aujourd_hui = timezone.now().date()

        
        
        # Chercher le dernier numéro de ticket pour aujourd'hui

        dernier_ticket = FactureVente.objects.filter(

            agence=agence,

            date=aujourd_hui

        ).aggregate(

            max_numero=Max('numero_ticket')

        )['max_numero']

        
        
        if dernier_ticket:

            # Extraire le numéro du dernier ticket (format: TKT20250923001)

            try:

                # Le format est TKT + YYYYMMDD + 3 chiffres

                numero_sequence = int(dernier_ticket[-3:])

                nouveau_numero = numero_sequence + 1

            except (ValueError, IndexError):

                nouveau_numero = 1

        else:

            nouveau_numero = 1
        
        
        
        # Formater le nouveau numéro de ticket

        date_str = aujourd_hui.strftime('%Y%m%d')

        numero_ticket = f"TKT{date_str}{nouveau_numero:03d}"

        # Vérifier que le numéro n'existe pas déjà
        while FactureVente.objects.filter(numero_ticket=numero_ticket).exists():
            nouveau_numero += 1
            numero_ticket = f"TKT{date_str}{nouveau_numero:03d}"

        print(f"[ALERTE] NUMÉRO TICKET GÉNÉRÉ: {numero_ticket}")
        
        return numero_ticket
        
        
        
    except Exception as e:

        print(f"Erreur lors de la génération du numéro de ticket: {e}")

        # En cas d'erreur, utiliser un format simple avec timestamp

        timestamp = timezone.now().strftime('%Y%m%d%H%M%S')

        return f"TKT{timestamp}"

    

@login_required
@require_caisse_access
def facturation_vente(request, facture_id=None):
    # Récupérer l'agence de l'utilisateur connecté
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    
    
    # Vérifier le statut de la caisse pour l'affichage mais ne pas bloquer l'accès

    caisse_ouverte = False

    caisse_actuelle = None

    try:

        aujourd_hui = timezone.now().date()

        session_caisse = SessionCaisse.objects.filter(

            agence=agence, 

            date_ouverture__date=aujourd_hui, 

            statut='ouverte'

        ).first()

        caisse_ouverte = session_caisse is not None

        if session_caisse:

            caisse_actuelle = session_caisse.caisse

    except:

        pass
    
    
    
    # Si aucune caisse ouverte, utiliser la première caisse disponible

    if not caisse_actuelle:

        caisse_actuelle = Caisse.objects.filter(agence=agence, statut='active').first()

        if not caisse_actuelle:

            # Chercher une caisse inactive et l'activer
            caisse_inactive = Caisse.objects.filter(agence=agence).first()
            if caisse_inactive:
                caisse_inactive.statut = 'active'
                caisse_inactive.save()
                caisse_actuelle = caisse_inactive
            else:
                # Créer une caisse par défaut si aucune n'existe
                caisse_actuelle, created = Caisse.objects.get_or_create(
                    numero_caisse='CAISSE001',
                    defaults={
                        'nom_caisse': 'Caisse Principale',
                        'agence': agence,
                        'solde_initial': 0,
                        'solde_actuel': 0,
                        'statut': 'active'
                    }
                )
    
    # Créer une session si aucune n'est ouverte
    if not caisse_ouverte and caisse_actuelle:
        try:
            # Trouver un employé de cette agence
            compte = Compte.objects.filter(agence=agence).first()
            if compte:
                employe = Employe.objects.filter(compte=compte).first()
                if employe:
                    # Fermer les sessions ouvertes existantes
                    SessionCaisse.objects.filter(agence=agence, statut='ouverte').update(statut='fermee')
                    
                    # Créer une nouvelle session
                    SessionCaisse.objects.create(
                        agence=agence,
                        caisse=caisse_actuelle,
                        employe=employe,
                        date_ouverture=timezone.now(),
                        solde_ouverture=0,
                        statut='ouverte'
                    )
        except:
            pass
    
    # Générer le numéro de ticket auto-incrémenté

    numero_ticket = generate_ticket_number(agence)

    
    
    # Récupérer la facture temporaire de la session

    facture_temp = request.session.get('facture_temporaire', {

        'lignes': [], 'type_vente': 'detail', 'remise': 0, 'montant_regler': 0, 'nette_a_payer': 0, 'rendu': 0

    })

    
    
    print(f"DEBUG facturation_vente: facture_temp récupérée = {facture_temp}")

    print(f"DEBUG facturation_vente: type de facture_temp = {type(facture_temp)}")

    print(f"DEBUG facturation_vente: lignes = {facture_temp.get('lignes', [])}")

    print(f"DEBUG facturation_vente: nombre de lignes = {len(facture_temp.get('lignes', []))}")

    
    
    # Si la session est vide, essayer de récupérer depuis les cookies ou créer une nouvelle

    if not facture_temp or not facture_temp.get('lignes'):

        print("DEBUG: Session vide, création d'une nouvelle facture temporaire")

        facture_temp = {

            'lignes': [], 

            'type_vente': 'detail', 

            'remise': 0, 

            'montant_regler': 0, 

            'nette_a_payer': 0, 

            'rendu': 0

        }

        request.session['facture_temporaire'] = facture_temp

    
    
    # Récupérer les clients pour le formulaire

    clients = Client.objects.filter(agence=agence).order_by('intitule')

    
    
    # Récupérer l'employé de l'utilisateur connecté
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
        employe = Employe.objects.filter(compte=compte).first()
        
        # Si pas d'employé, créer les infos à partir du compte
        if not employe:
            # L'utilisateur a un compte mais pas de fiche employé
            # On utilise les infos du compte directement
            employe = type('obj', (object,), {
                'compte': type('obj', (object,), {
                    'nom': compte.nom,
                    'prenom': compte.prenom,
                    'nom_complet': compte.nom_complet
                })()
            })()
    except Compte.DoesNotExist:
        # Fallback : utiliser le username
        employe = type('obj', (object,), {
            'compte': type('obj', (object,), {
                'nom': request.user.last_name or request.user.username,
                'prenom': request.user.first_name or '',
                'nom_complet': f"{request.user.first_name or ''} {request.user.last_name or request.user.username}".strip()
            })()
        })()

    
    
    return render(request, 'supermarket/caisse/facturation_vente.html', {

        'agence': agence, 

        'facture_temp': facture_temp, 

        'caisse_ouverte': caisse_ouverte,

        'caisse_actuelle': caisse_actuelle,

        'numero_ticket': numero_ticket,

        'clients': clients,

        'employe': employe,

        'caisses': Caisse.objects.filter(agence=agence, statut='active')

    })



@login_required
@require_caisse_access
def ouvrir_caisse(request):
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    
    
    try:

        caisse, created = Caisse.objects.get_or_create(

            numero_caisse='CAISSE001',

            defaults={'agence': agence, 'solde_actuel': 0, 'statut': 'active'}

        )

        
        
        employe = Employe.objects.filter(compte__agence=agence).first()

        if not employe:

            messages.error(request, 'Aucun employé trouvé.')
            return redirect('dashboard_caisse')
        
        
        
        aujourd_hui = timezone.now().date()

        sessions_ouvertes = SessionCaisse.objects.filter(agence=agence, date_ouverture=aujourd_hui, statut='ouverte').count()

        
        
        if sessions_ouvertes > 0:

            messages.info(request, 'Caisse déjà ouverte.')

            return redirect('dashboard_caisse')
        
        
        
        session = SessionCaisse.objects.create(

            caisse=caisse, utilisateur=None, employe=employe, agence=agence,

            solde_ouverture=caisse.solde_actuel, statut='ouverte'

        )

        
        # Vider les factures temporaires en attente pour la nouvelle session
        FactureTemporaire.objects.all().delete()
        
        
        messages.success(request, f'Caisse ouverte avec succès. Solde: {caisse.solde_actuel} FCFA')

        return redirect('dashboard_caisse')
        
        
        
    except Exception as e:

        messages.error(request, f'Erreur: {str(e)}')

        return redirect('dashboard_caisse')




@login_required
@require_caisse_access
def search_window(request):

    search_term = request.GET.get('q', '')

    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    
    
    articles = []

    if search_term and len(search_term) >= 2:


        articles = Article.objects.filter(agence=agence, designation__icontains=search_term)[:50]
    
    
    
    return render(request, 'supermarket/caisse/search_window.html', {

        'articles': articles, 'search_term': search_term

    })



@login_required
@require_caisse_access
def search_articles_api(request):

    search_term = request.GET.get('q', '')

    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'articles': []})
    
    
    
    articles = []

    if search_term:

        articles = Article.objects.filter(agence=agence, designation__icontains=search_term).order_by('designation')


    else:


        articles = Article.objects.filter(agence=agence).order_by('designation')
    
    
    
    results = []

    for article in articles:

        results.append({

            'id': article.id, 'reference': article.reference_article, 'designation': article.designation,

            'prix_achat': float(article.prix_achat or 0), 'prix_vente': float(article.prix_vente or 0),

            'stock': int(article.stock_actuel or 0)

        })
    
    
    
    return JsonResponse({'articles': results})

@login_required
@require_caisse_access
def get_prix_by_type(request):

    article_id = request.GET.get('article_id')

    type_vente = request.GET.get('type_vente', 'detail')

    
    
    try:

        article = Article.objects.get(id=article_id)

        type_vente_obj = TypeVente.objects.filter(article=article, intitule__iexact=type_vente).first()

        
        
        if type_vente_obj:

            prix = float(type_vente_obj.prix)

        else:

            prix = float(article.prix_vente or 0)
        
        
        
        return JsonResponse({'success': True, 'prix': prix})

    except Exception as e:

        return JsonResponse({'success': False, 'prix': 0, 'error': str(e)})



@login_required
@require_caisse_access
def ajouter_article_facture(request):

    if request.method == 'POST':
        
        agence = get_user_agence(request)
        if not agence:

            return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
            
            
            
        try:

            article_id = request.POST.get('article_id')

            quantite = int(request.POST.get('quantite', 1))

            
            
            print(f"DEBUG: Article ID reçu: {article_id}")

            print(f"DEBUG: Quantité reçue: {quantite}")

            
            
            article = Article.objects.get(id=article_id, agence=agence)

            print(f"DEBUG: Article trouvé: {article.designation}")

            
            
            facture_temp = request.session.get('facture_temporaire', {

                'lignes': [], 'type_vente': 'detail', 'remise': 0, 'montant_regler': 0, 'nette_a_payer': 0, 'rendu': 0

            })

            
            
            print(f"DEBUG: Facture temp avant: {facture_temp}")

            
            
            article_existe = False

            for ligne in facture_temp['lignes']:

                # Comparer les IDs en tant que strings pour éviter les problèmes de type

                if str(ligne['article_id']) == str(article_id):

                    ligne['quantite'] += quantite

                    ligne['prix_total'] = ligne['quantite'] * ligne['prix_unitaire']

                    article_existe = True

                    print(f"DEBUG: Article existant mis à jour")

                    break
            
            
            
            if not article_existe:

                prix_unitaire = float(article.prix_vente or 0)

                prix_total = quantite * prix_unitaire

                
                
                nouvelle_ligne = {

                    'article_id': article_id, 'reference': article.reference_article, 'designation': article.designation,

                    'quantite': quantite, 'prix_unitaire': prix_unitaire, 'prix_total': prix_total, 'type_vente': 'detail'

                }

                facture_temp['lignes'].append(nouvelle_ligne)

            
            
            
            request.session['facture_temporaire'] = facture_temp

            
            
            return JsonResponse({
                'success': True, 
                'message': f'Article "{article.designation}" ajouté avec succès!',
                'article': {

                    'id': article_id,

                    'designation': article.designation,

                    'reference': article.reference_article,

                    'prix_vente': float(article.prix_vente or 0)

                }

            })
            
            
            
        except Exception as e:

            print(f"DEBUG: Erreur: {str(e)}")

            return JsonResponse({'success': False, 'error': str(e)})



    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})



@csrf_exempt
@login_required
@require_caisse_access
def enregistrer_facture(request):
    try:
        print("=" * 80)
        print("*** NOUVEAU CODE VERSION 2.0 CHARGE AVEC SUCCES ***")
        print("=" * 80)
        print("[ALERTE] ENREGISTRER_FACTURE APPELÉE")
        
        if request.method == 'POST':
            print("[ALERTE] MÉTHODE POST DÉTECTÉE")
        
        # Vérifier spécifiquement le champ facture_data
        facture_data = request.POST.get('facture_data', '')
        print(f"[ALERTE] facture_data: {facture_data[:100] if facture_data else 'VIDE'}")

        
        
        # Récupérer l'agence de l'utilisateur connecté
        agence = get_user_agence(request)
        if not agence:
            print(f"🔴 DEBUG: [ERREUR] Aucune agence trouvée pour l'utilisateur")
            return JsonResponse({
                'success': False,
                'error': 'Votre compte n\'est pas configuré correctement.'
            })
        
        print(f"🔴 DEBUG: [OK] Agence de l'utilisateur: {agence.nom_agence} (ID: {agence.id_agence})")

        # S'assurer qu'il y a une caisse active pour cette agence
        caisse_actuelle = Caisse.objects.filter(agence=agence, statut='active').first()
        if not caisse_actuelle:
            # Chercher une caisse inactive et l'activer
            caisse_inactive = Caisse.objects.filter(agence=agence).first()
            if caisse_inactive:
                caisse_inactive.statut = 'active'
                caisse_inactive.save()
                caisse_actuelle = caisse_inactive
                print(f"🔴 DEBUG: [OK] Caisse activée: {caisse_actuelle.numero_caisse}")
            else:
                # Créer une caisse par défaut si aucune n'existe
                caisse_actuelle, created = Caisse.objects.get_or_create(
                    numero_caisse='CAISSE001',
                    defaults={
                        'nom_caisse': 'Caisse Principale',
                        'agence': agence,
                        'solde_initial': 0,
                        'solde_actuel': 0,
                        'statut': 'active'
                    }
                )
                print(f"🔴 DEBUG: [OK] Caisse créée: {caisse_actuelle.numero_caisse}")

        # S'assurer qu'il y a une session ouverte
        session_ouverte = SessionCaisse.objects.filter(agence=agence, statut='ouverte').first()
        if not session_ouverte and caisse_actuelle:
            try:
                # Trouver un employé de cette agence
                compte = Compte.objects.filter(agence=agence).first()
                if compte:
                    employe = Employe.objects.filter(compte=compte).first()
                    if employe:
                        # Fermer les sessions ouvertes existantes
                        SessionCaisse.objects.filter(agence=agence, statut='ouverte').update(statut='fermee')
                        
                        # Créer une nouvelle session
                        session_ouverte = SessionCaisse.objects.create(
                            agence=agence,
                            caisse=caisse_actuelle,
                            employe=employe,
                            date_ouverture=timezone.now(),
                            solde_ouverture=0,
                            statut='ouverte'
                        )
                        print(f"🔴 DEBUG: [OK] Session créée: {session_ouverte.id}")
            except Exception as e:
                print(f"🔴 DEBUG: [ERREUR] Erreur création session: {e}")

        try:

            # Logs détaillés pour debug
            print("=" * 80)
            print("DEBUG ENREGISTRER_FACTURE: Début de la fonction")
            print(f"DEBUG: Méthode de requête: {request.method}")
            print(f"DEBUG: POST data: {dict(request.POST)}")
            print(f"DEBUG: Session key: {request.session.session_key}")
            print(f"DEBUG: Session data: {dict(request.session)}")
            
            # NOUVELLE LOGIQUE: Récupérer les articles du POST en priorité
            facture_temp = {'lignes': [], 'remise': 0, 'montant_regler': 0, 'rendu': 0}
            articles_from_post = []
            
            # Méthode 1: Chercher dans le champ facture_data (JSON)
            facture_data_str = request.POST.get('facture_data', '')
            if facture_data_str:
                try:
                    import json
                    facture_data_parsed = json.loads(facture_data_str)
                    lignes = facture_data_parsed.get('lignes', [])
                    if lignes:
                        print(f"🔴 DEBUG: {len(lignes)} articles récupérés depuis facture_data JSON")
                        facture_temp['lignes'] = lignes
                        facture_temp['remise'] = facture_data_parsed.get('remise', 0)
                        facture_temp['montant_regler'] = facture_data_parsed.get('montant_regler', 0)
                        facture_temp['rendu'] = facture_data_parsed.get('rendu', 0)
                except json.JSONDecodeError as e:
                    print(f"🔴 DEBUG: Erreur parsing JSON facture_data: {e}")
            
            # Méthode 2: Si pas de facture_data, chercher les champs article_X_id dans POST
            if not facture_temp.get('lignes'):
                print("🔴 DEBUG: Tentative de récupération des articles depuis les champs POST")
                for key, value in request.POST.items():
                    if key.startswith('article_') and key.endswith('_id') and value:
                        article_id = value
                        prefix = key.replace('_id', '')
                        quantite = request.POST.get(f'{prefix}_quantite', 1)
                        prix_unitaire = request.POST.get(f'{prefix}_prix_unitaire', 0)
                        prix_total = request.POST.get(f'{prefix}_prix_total', 0)
                        designation = request.POST.get(f'{prefix}_designation', '')
                        reference = request.POST.get(f'{prefix}_reference', '')
                        
                        try:
                            article = Article.objects.get(id=article_id)
                            articles_from_post.append({
                                'article_id': int(article_id),
                                'designation': designation or article.designation,
                                'quantite': float(quantite),
                                'prix_unitaire': float(prix_unitaire),
                                'prix_total': float(prix_total),
                                'reference': reference or article.reference_article
                            })
                            print(f"🔴 DEBUG: Article récupéré depuis POST: {designation} (ID: {article_id})")
                        except Article.DoesNotExist:
                            print(f"🔴 DEBUG: Article avec ID {article_id} non trouvé")
                        except Exception as e:
                            print(f"🔴 DEBUG: Erreur récupération article {article_id}: {e}")
                
                if articles_from_post:
                    facture_temp['lignes'] = articles_from_post
                    print(f"🔴 DEBUG: {len(articles_from_post)} articles récupérés depuis les champs POST")
            
            # Méthode 3: Si toujours pas d'articles, chercher dans la session
            if not facture_temp.get('lignes'):
                facture_temp_session = request.session.get('facture_temporaire', {'lignes': []})
                if facture_temp_session.get('lignes'):
                    facture_temp = facture_temp_session
                    print(f"🔴 DEBUG: {len(facture_temp['lignes'])} articles récupérés depuis la session")
            
            # Méthode 4: Si toujours pas d'articles, chercher dans FactureTemporaire (DB)
            if not facture_temp.get('lignes'):
                try:
                    session_key = request.session.session_key
                    if session_key:
                        facture_temp_db = FactureTemporaire.objects.filter(session_key=session_key).first()
                        if facture_temp_db and facture_temp_db.contenu:
                            import json
                            facture_temp = json.loads(facture_temp_db.contenu)
                            print(f"🔴 DEBUG: {len(facture_temp.get('lignes', []))} articles récupérés depuis FactureTemporaire DB")
                except Exception as e:
                    print(f"🔴 DEBUG: Erreur récupération depuis DB: {e}")
            
            print(f"🔴 DEBUG: [FINAL] Nombre total d'articles: {len(facture_temp.get('lignes', []))}")
            
            # Si toujours aucun article, retourner une erreur
            if not facture_temp.get('lignes'):
                print("🔴 DEBUG: [ERREUR FINALE] Aucun article trouvé après toutes les tentatives")
                return JsonResponse({
                    'success': False,
                    'error': 'Aucun article dans la facture. Veuillez ajouter des articles.'
                })
            
            print(f"🔴 DEBUG: [LIST] Articles à enregistrer:")
            for i, ligne in enumerate(facture_temp.get('lignes', [])):
                print(f"🔴 DEBUG: Ligne {i}: article_id={ligne.get('article_id', 'VIDE')}, designation={ligne.get('designation', 'VIDE')}, qté={ligne.get('quantite', 0)}")

            # Passer au traitement (la suite du code reste inchangée)
            # Utiliser la caisse actuelle (déjà vérifiée/créée plus haut)
            caisse = caisse_actuelle

            print(f"🔴 DEBUG: [OK] Caisse utilisée: {caisse.numero_caisse} (ID: {caisse.id})")

            
            
            # Récupérer le client

            client_id = request.POST.get('client_id')

            client_name = request.POST.get('client_name')

            print(f"DEBUG: Client ID: {client_id}, Client Name: {client_name}")

            client = None
            
            if client_id:
                try:
                    # Essayer de récupérer le client par ID (sans filtrer par agence)
                    client = Client.objects.get(id=client_id)
                    print(f"DEBUG: Client existant trouvé: {client}")
                    
                    # Vérifier si le client appartient à une autre agence
                    if client.agence != agence:
                        print(f"DEBUG: ATTENTION - Client de l'agence {client.agence.nom_agence}, utilisateur de l'agence {agence.nom_agence}")
                        # On peut soit utiliser ce client, soit utiliser le client par défaut de l'agence
                        # Pour l'instant, on utilise le client trouvé (vente inter-agence possible)
                except Client.DoesNotExist:
                    print(f"DEBUG: Client avec ID {client_id} non trouvé, utilisation du client par défaut")
                    client = None

            elif client_name:

                # Créer un nouveau client si le nom est fourni

                client = Client.objects.create(

                    intitule=client_name,

                    adresse='Adresse non spécifiée',

                    telephone='Non spécifié',

                    email='',

                    agence=agence

                )

                print(f"DEBUG: Nouveau client créé: {client}")

            else:
                # Utiliser un client par défaut
                client = Client.objects.filter(agence=agence).first()

            if not client:
                client = Client.objects.create(
                    intitule='Client Général',
                    adresse='Adresse non spécifiée',
                    telephone='Non spécifié',
                    email='',
                    agence=agence
                )
                print(f"DEBUG: Client par défaut créé: {client}")
            else:
                print(f"DEBUG: Client par défaut existant: {client}")
                
            # Générer un nouveau numéro de ticket
            numero_ticket = generate_ticket_number(agence)

            print(f"DEBUG: Numéro de ticket généré: {numero_ticket}")

            
            
            # Calculer les totaux

            total_ht = sum(ligne['prix_total'] for ligne in facture_temp['lignes'])

            remise = facture_temp.get('remise', 0)

            nette_a_payer = total_ht - remise

            montant_regler = facture_temp.get('montant_regler', 0)

            rendu = montant_regler - nette_a_payer

            
            
            print(f"DEBUG: Totaux - Total HT: {total_ht}, Remise: {remise}, Nette: {nette_a_payer}")

            print(f"DEBUG: Montant réglé: {montant_regler}, Rendu: {rendu}")

            
            
            # Récupérer l'employé de l'utilisateur connecté
            try:
                compte = Compte.objects.get(user=request.user, actif=True)
                employe = Employe.objects.filter(compte=compte).first()
                
                # Si pas d'employé, utiliser les infos du compte
                if not employe:
                    employe = type('obj', (object,), {
                        'compte': compte
                    })()
                
                print(f"DEBUG: Employé vendeur: {employe.compte.nom_complet if employe and hasattr(employe, 'compte') else 'Aucun'}")
            except Compte.DoesNotExist:
                employe = None
                print(f"DEBUG: Aucun compte trouvé pour l'utilisateur: {request.user.username}")
            
            

            # Récupérer la session de caisse active
            session_caisse = None
            try:
                aujourd_hui = timezone.now().date()

                # Rechercher la session active (même critères que l'API KPIs)
                session_caisse = SessionCaisse.objects.filter(
                    agence=agence,
                    date_ouverture__date=aujourd_hui,
                    statut='ouverte'
                ).first()
                
                print(f"DEBUG: Session trouvée: {session_caisse}")
                
                print(f"DEBUG: Session de caisse active: {session_caisse}")
            except Exception as e:
                print(f"DEBUG: Erreur lors de la récupération de la session: {e}")

            # Déterminer la date de la facture (priorité à la date saisie)
            sale_date_str = request.POST.get('sale_date')
            date_facture = timezone.now().date()
            if sale_date_str:
                try:
                    from datetime import datetime
                    date_facture = datetime.strptime(sale_date_str, '%Y-%m-%d').date()
                    print(f"DEBUG: Date de vente saisie utilisée: {date_facture}")
                except ValueError:
                    print(f"DEBUG: Date de vente invalide '{sale_date_str}', utilisation de la date du jour")
            heure_actuelle = timezone.now().time()

            
            
            print("DEBUG: Création de la facture...")
            
            # Vérifier si le numéro de ticket existe déjà
            if FactureVente.objects.filter(numero_ticket=numero_ticket).exists():
                print(f"DEBUG: Numéro de ticket déjà existant, génération d'un nouveau...")
                numero_ticket = generate_ticket_number(agence)
                print(f"DEBUG: Nouveau numéro de ticket: {numero_ticket}")
            
            try:
                facture = FactureVente.objects.create(

                    numero_ticket=numero_ticket, 

                    client=client, 

                    agence=agence, 

                    caisse=caisse,

                    vendeur=employe, 

                    session_caisse=session_caisse,
                    date=date_facture, 

                    heure=heure_actuelle,

                    nette_a_payer=nette_a_payer, 

                    remise=remise, 

                    montant_regler=montant_regler,

                    rendu=rendu, 

                    en_attente=False,

                    nom_vendeuse=employe.compte.nom_complet if employe else 'Vendeur'

                )
                print(f"DEBUG: Facture créée avec ID: {facture.id}")
            except IntegrityError as e:
                # Si l'erreur persiste, générer un nouveau numéro et réessayer
                print(f"DEBUG: Erreur d'intégrité, génération d'un nouveau numéro...")
                numero_ticket = generate_ticket_number(agence)
                facture = FactureVente.objects.create(

                    numero_ticket=numero_ticket, 

                    client=client, 

                    agence=agence, 

                    caisse=caisse,

                    vendeur=employe, 

                    session_caisse=session_caisse,
                    date=date_facture, 

                    heure=heure_actuelle,

                    nette_a_payer=nette_a_payer, 

                    remise=remise, 

                    montant_regler=montant_regler,

                    rendu=rendu, 

                    en_attente=False,

                    nom_vendeuse=employe.compte.nom_complet if employe else 'Vendeur'

                )
                print(f"DEBUG: Facture créée avec ID: {facture.id} (nouveau numéro: {numero_ticket})")

            
            
            # Créer les lignes de facture

            print(f"DEBUG: Création de {len(facture_temp['lignes'])} lignes de facture...")

            for i, ligne_temp in enumerate(facture_temp['lignes']):

                print(f"DEBUG: Ligne {i+1}: {ligne_temp}")

                
                # Vérifier que l'article_id est valide
                article_id = ligne_temp.get('article_id')
                if not article_id:
                    print(f"DEBUG: ERREUR - article_id manquant pour la ligne {i+1}: {ligne_temp}")
                    continue
                
                try:
                    article = Article.objects.get(id=article_id)
                    print(f"DEBUG: Article trouvé: {article} (ID: {article.id})")
                except Article.DoesNotExist:
                    print(f"DEBUG: ERREUR - Article avec ID {article_id} non trouvé")
                    continue
                
                
                # S'assurer que les valeurs sont converties correctement
                # Convertir en float pour les valeurs numériques
                quantite_value = ligne_temp.get('quantite', 0)
                prix_unitaire_value = ligne_temp.get('prix_unitaire', 0)
                prix_total_value = ligne_temp.get('prix_total', 0)
                
                # Si les valeurs sont des Decimal, les convertir
                if isinstance(quantite_value, Decimal):
                    quantite_value = float(quantite_value)
                elif not isinstance(quantite_value, (int, float)):
                    quantite_value = float(quantite_value)
                
                if isinstance(prix_unitaire_value, Decimal):
                    prix_unitaire_value = float(prix_unitaire_value)
                elif not isinstance(prix_unitaire_value, (int, float)):
                    prix_unitaire_value = float(prix_unitaire_value)
                
                if isinstance(prix_total_value, Decimal):
                    prix_total_value = float(prix_total_value)
                elif not isinstance(prix_total_value, (int, float)):
                    prix_total_value = float(prix_total_value)
                
                # Contrôles de stock AVANT création de la ligne
                quantite_vendue = Decimal(str(ligne_temp['quantite']))
                if article.stock_actuel <= 0:
                    print(f"[ERREUR] STOCK NUL/NEGATIF - Article: {article.designation} (stock={article.stock_actuel})")
                    # Nettoyer toute création partielle et retourner une erreur claire
                    try:
                        LigneFactureVente.objects.filter(facture_vente=facture).delete()
                        facture.delete()
                    except Exception:
                        pass
                    return JsonResponse({
                        'success': False,
                        'message': 'Pas de stock',
                        'message_type': 'error'
                    }, status=400)
                if quantite_vendue > Decimal(str(article.stock_actuel)):
                    print(f"[ERREUR] STOCK INSUFFISANT - Article: {article.designation} | demandé={quantite_vendue} > stock={article.stock_actuel}")
                    try:
                        LigneFactureVente.objects.filter(facture_vente=facture).delete()
                        facture.delete()
                    except Exception:
                        pass
                    return JsonResponse({
                        'success': False,
                        'message': 'Stock insuffisant',
                        'message_type': 'error'
                    }, status=400)

                # Création de la ligne après validations
                ligne_facture = LigneFactureVente.objects.create(

                    facture_vente=facture, 

                    article=article, 

                    designation=ligne_temp['designation'],

                    quantite=quantite_value, 

                    prix_unitaire=prix_unitaire_value,

                    prix_total=prix_total_value
                )
            
                print(f"DEBUG: Ligne de facture créée: {ligne_facture.id}")
                
                # [HOT] GESTION AUTOMATIQUE DU STOCK - RÉDUCTION LORS DE LA VENTE
                ancien_stock = article.stock_actuel
                
                # Réduire le stock (sécurisé par validations plus haut)
                article.stock_actuel -= quantite_vendue
                # Le suivi de stock est toujours activé automatiquement lors de toute modification du stock
                article.suivi_stock = True
                article.save()
                print(f"🛒 STOCK VENTE - Article: {article.designation}")
                print(f"🛒 STOCK VENTE - Quantité vendue: {quantite_vendue}")
                print(f"🛒 STOCK VENTE - Stock mis à jour: {ancien_stock} → {article.stock_actuel}")
                
                # Créer un mouvement de stock pour traçabilité
                try:
                    mouvement_datetime = datetime.combine(
                        facture.date or timezone.now().date(),
                        facture.heure or heure_actuelle or timezone.now().time()
                    )
                    if timezone.is_naive(mouvement_datetime):
                        mouvement_datetime = timezone.make_aware(
                            mouvement_datetime,
                            timezone.get_current_timezone()
                        )
                    
                    MouvementStock.objects.create(
                        article=article,
                        agence=agence,
                        type_mouvement='sortie',
                        date_mouvement=mouvement_datetime,
                        numero_piece=facture.numero_ticket,
                        quantite_stock=article.stock_actuel,
                        stock_initial=ancien_stock,
                        solde=article.stock_actuel,
                        quantite=quantite_vendue,
                        cout_moyen_pondere=float(article.prix_achat),
                        stock_permanent=float(article.stock_actuel * article.prix_achat),
                        facture_vente=facture,
                        commentaire=f"Vente - Facture {facture.numero_ticket}"
                    )
                    print(f"[NOTE] MOUVEMENT STOCK - Sortie enregistrée pour {article.designation}")
                except Exception as e:
                    print(f"[WARNING] ERREUR MOUVEMENT STOCK: {e}")
            
            
            
            # Vider la facture temporaire de la session ET de la base de données
            
            # 1. Vider la session
            request.session['facture_temporaire'] = {
                'lignes': [], 'type_vente': 'detail', 'remise': 0, 'montant_regler': 0, 'nette_a_payer': 0, 'rendu': 0
            }
            
            # 2. Supprimer la facture temporaire de la base de données
            session_key = request.session.session_key
            if session_key:
                factures_temp_supprimees = FactureTemporaire.objects.filter(session_key=session_key).delete()
                print(f"DEBUG: {factures_temp_supprimees[0]} facture(s) temporaire(s) supprimée(s) de la base de données")
            
            print("DEBUG: Facture temporaire vidée de la session ET de la base de données")

            
            
            # Retourner une réponse JSON pour AJAX

            print(f"DEBUG: Envoi de la réponse de succès pour la facture {numero_ticket}")

            return JsonResponse({
                'success': True,
                'message': f'Facture {numero_ticket} enregistrée avec succès!',
                'numero_ticket': numero_ticket
            })
            
            
            
        except Exception as e:

            print(f"DEBUG: Erreur lors de l'enregistrement: {str(e)}")

            import traceback

            print(f"DEBUG: Traceback: {traceback.format_exc()}")
            
            # Sauvegarder l'erreur dans un fichier pour debug
            try:
                with open('ERREUR_500_SERVEUR.txt', 'w', encoding='utf-8') as f:
                    f.write(f"ERREUR 500 - ENREGISTREMENT FACTURE\n")
                    f.write(f"=" * 50 + "\n\n")
                    f.write(f"Erreur: {str(e)}\n\n")
                    f.write(f"Traceback:\n{traceback.format_exc()}\n")
                print("📄 Erreur sauvegardée dans ERREUR_500_SERVEUR.txt")
            except:
                pass

            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    except Exception as global_error:
        # Capturer TOUTES les erreurs non gérées
        print(f"❌❌❌ ERREUR GLOBALE 500: {str(global_error)}")
        import traceback
        error_trace = traceback.format_exc()
        print(f"❌❌❌ TRACEBACK:")
        print(error_trace)
        
        # Sauvegarder dans un fichier
        try:
            with open('ERREUR_500_SERVEUR.txt', 'w', encoding='utf-8') as f:
                f.write(f"ERREUR 500 - ENREGISTREMENT FACTURE\n")
                f.write(f"=" * 50 + "\n\n")
                f.write(f"Erreur: {str(global_error)}\n\n")
                f.write(f"Traceback complet:\n{error_trace}\n")
            print("📄 Erreur sauvegardée dans ERREUR_500_SERVEUR.txt")
        except:
            pass
        
        return JsonResponse({
            'success': False,
            'error': f'Erreur serveur 500: {str(global_error)}'
        }, status=500)
@login_required
@require_caisse_access
def rapport_caisse(request):

    """Afficher le rapport de caisse avec les données de la journée"""

    agence = get_user_agence(request)
    if not agence:

        messages.error(request, 'Aucune agence trouvée.')

        return redirect('dashboard_caisse')
    
    
    
    # Récupérer la date d'aujourd'hui

    aujourd_hui = timezone.now().date()

    
    
    # Vérifier le statut de la caisse

    caisse_ouverte = False

    try:

        session_caisse = SessionCaisse.objects.filter(

            agence=agence, 

            date_ouverture__date=aujourd_hui, 

            statut='ouverte'

        ).first()

        caisse_ouverte = session_caisse is not None

    except:

        pass
    
    
    
    # Récupérer toutes les sessions de la journée

    sessions_jour = SessionCaisse.objects.filter(

        agence=agence, 

        date_ouverture__date=aujourd_hui

    ).order_by('date_ouverture').select_related('employe__compte')

    
    
    # Calculer les KPIs de la journée

    chiffre_affaires = 0

    nombre_ventes = 0

    tickets_attente = 0

    

    try:

        # Récupérer les ventes du jour liées aux sessions de caisse
        ventes_jour = FactureVente.objects.filter(
            agence=agence, 
            date=aujourd_hui,
            session_caisse__isnull=False
        ).select_related('session_caisse')
        
        chiffre_affaires = ventes_jour.aggregate(total=Sum('nette_a_payer'))['total'] or 0

        nombre_ventes = ventes_jour.count()

        # Compter les factures temporaires UNIQUEMENT pour la session courante ouverte
        if session_caisse:
            tickets_attente = FactureTemporaire.objects.filter(session_caisse=session_caisse).count()
        else:
            tickets_attente = 0

        
        print(f"DEBUG: Chiffre d'affaires calculé: {chiffre_affaires}")
        print(f"DEBUG: Nombre de ventes: {nombre_ventes}")
        print(f"DEBUG: Tickets en attente: {tickets_attente}")
    except Exception as e:
        print(f"DEBUG: Erreur lors du calcul des KPIs: {e}")
        pass

    

    # Récupérer les détails des ventes pour le rapport

    factures_jour = FactureVente.objects.filter(

        agence=agence, 

        date=aujourd_hui,
        session_caisse__isnull=False
    ).select_related('client', 'vendeur__compte', 'session_caisse').prefetch_related('lignes__article').order_by('-heure')
    
    
    # Calculer les statistiques par heure

    ventes_par_heure = {}

    for facture in factures_jour:

        heure = facture.heure.hour

        if heure not in ventes_par_heure:

            ventes_par_heure[heure] = {'nombre': 0, 'montant': 0}

        ventes_par_heure[heure]['nombre'] += 1

        ventes_par_heure[heure]['montant'] += float(facture.nette_a_payer or 0)
    
    
    
    # Récupérer les articles les plus vendus

    articles_vendus = LigneFactureVente.objects.filter(

        facture_vente__agence=agence,

        facture_vente__date=aujourd_hui

    ).values('article__designation').annotate(

        total_quantite=Sum('quantite'),

        total_montant=Sum('prix_total')

    ).order_by('-total_quantite')[:10]

    
    
    # Récupérer le nom du compte connecté pour l'affichage
    try:
        compte_connecte = Compte.objects.get(user=request.user, actif=True)
        vendeuse_nom = compte_connecte.nom_complet
    except Compte.DoesNotExist:
        vendeuse_nom = "-"
    
    context = {

        'agence': agence,

        'date_aujourd_hui': aujourd_hui,

        'caisse_ouverte': caisse_ouverte,

        'sessions_jour': sessions_jour,

        'chiffre_affaires': chiffre_affaires,

        'nombre_ventes': nombre_ventes,

        'tickets_attente': tickets_attente,

        'factures_jour': factures_jour,

        'ventes_par_heure': ventes_par_heure,
        
        'vendeuse_nom': vendeuse_nom,

        'articles_vendus': articles_vendus,

    }

    
    
    return render(request, 'supermarket/caisse/rapport_caisse.html', context)



@login_required
@require_caisse_access
def detail_factures(request):
    print("DEBUG: Début de detail_factures")
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    print(f"DEBUG: Agence de l'utilisateur: {agence}")
    
    
    
    # Récupérer les paramètres de filtrage

    search_query = request.GET.get('search', '')

    date_filter = request.GET.get('date_filter', 'all')
    

    

    # Récupérer toutes les factures de l'agence liées aux sessions de caisse
    factures = FactureVente.objects.filter(agence=agence).select_related('client', 'vendeur__compte', 'session_caisse').prefetch_related('lignes__article').order_by('-date', '-heure')
    print(f"DEBUG: Nombre de factures trouvées: {factures.count()}")
    
    
    # Appliquer les filtres

    if search_query:

        factures = factures.filter(

            Q(numero_ticket__icontains=search_query) |

            Q(client__intitule__icontains=search_query) |

            Q(nom_vendeuse__icontains=search_query)

        )
    
    
    
    # Filtre par date

    from datetime import datetime, timedelta

    today = datetime.now().date()

    
    
    if date_filter == 'today':

        factures = factures.filter(date=today)

    elif date_filter == 'week':

        week_ago = today - timedelta(days=7)

        factures = factures.filter(date__gte=week_ago)

    elif date_filter == 'month':

        factures = factures.filter(date__year=today.year, date__month=today.month)
    

    

    # Calculer les statistiques (seulement les ventes liées aux sessions de caisse)
    total_factures = FactureVente.objects.filter(agence=agence, session_caisse__isnull=False).count()
    chiffre_affaires_total = FactureVente.objects.filter(agence=agence, session_caisse__isnull=False).aggregate(
        total=Sum('nette_a_payer')

    )['total'] or 0
    
    

    factures_aujourd_hui = FactureVente.objects.filter(agence=agence, date=today, session_caisse__isnull=False).count()
    ca_aujourd_hui = FactureVente.objects.filter(agence=agence, date=today, session_caisse__isnull=False).aggregate(
        total=Sum('nette_a_payer')

    )['total'] or 0

    
    
    # Récupérer l'employé connecté

    employe = None

    if request.user.is_authenticated:

        try:

            employe = Employe.objects.get(compte__user=request.user, compte__agence=agence)

        except Employe.DoesNotExist:

            pass

    

    print(f"DEBUG: Statistiques - Total factures: {total_factures}, CA total: {chiffre_affaires_total}")
    print(f"DEBUG: Aujourd'hui - Factures: {factures_aujourd_hui}, CA: {ca_aujourd_hui}")
    
    context = {

        'factures': factures,

        'total_factures': total_factures,
        'user': request.user,

        'chiffre_affaires_total': chiffre_affaires_total,

        'factures_aujourd_hui': factures_aujourd_hui,

        'ca_aujourd_hui': ca_aujourd_hui,

        'search_query': search_query,

        'date_filter': date_filter,

        'agence': agence,

        'employe': employe,

    }
    
    

    print("DEBUG: Rendu du template detail_factures.html")
    return render(request, 'supermarket/caisse/detail_factures.html', context)


@login_required
def liste_utilisateurs_huitieme(request):
    """Vue temporaire pour lister tous les utilisateurs de l'agence MARCHE HUITIEME"""
    try:
        agence = Agence.objects.filter(nom_agence__icontains='huitieme').first()
        
        if not agence:
            messages.error(request, 'Agence MARCHE HUITIEME non trouvée.')
            return redirect('detail_factures')
        
        comptes = Compte.objects.filter(agence=agence, actif=True).select_related('user').order_by('user__username')
        
        print(f"\n{'='*60}")
        print(f"AGENCE: {agence.nom_agence}")
        print(f"Nombre d'utilisateurs actifs: {comptes.count()}")
        print(f"{'-'*60}")
        print("LISTE DES UTILISATEURS:")
        print(f"{'-'*60}")
        
        for compte in comptes:
            print(f"Username: {compte.user.username}")
            print(f"  Nom complet: {compte.nom_complet}")
            print(f"  Type: {compte.type_compte}")
            print(f"  Email: {compte.email}")
            print()
        
        print(f"{'='*60}\n")
        
        context = {
            'agence': agence,
            'comptes': comptes,
            'total': comptes.count(),
        }
        
        # Retourner un message simple
        messages.info(request, f'Liste des utilisateurs de {agence.nom_agence}: {comptes.count()} utilisateur(s) actif(s). Voir la console pour les détails.')
        return redirect('detail_factures')
        
    except Exception as e:
        print(f"[ERREUR] Erreur lors de la liste: {e}")
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('detail_factures')



@login_required
@require_caisse_access
def mouvement_vente(request):
    """Afficher les mouvements de vente du jour - même logique que le document HTML"""
    print(f"[ALERTE] MOUVEMENT_VENTE: {request.user.username}")
    
    agence = get_user_agence(request)
    if not agence:
        print(f"[ALERTE] MOUVEMENT_VENTE: Aucune agence pour {request.user.username}")
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    print(f"[ALERTE] MOUVEMENT_VENTE: Agence {agence.nom_agence}")
    
    # Date affichée (par défaut aujourd'hui, remplacée par la date des factures si disponible)
    date_reference = timezone.now().date()
    periode_affichage = date_reference.strftime('%d/%m/%Y')
    
    # Récupérer la session active (peu importe la date d'ouverture)
    session_caisse = None
    
    # Récupérer le nom du compte connecté
    try:
        compte_connecte = Compte.objects.get(user=request.user, actif=True)
        vendeuse_nom = compte_connecte.nom_complet
    except Compte.DoesNotExist:
        vendeuse_nom = "-"
    
    try:
        # Récupérer seulement la session ACTIVE (ouverte) - peu importe la date d'ouverture
        session_caisse = SessionCaisse.objects.filter(
            agence=agence, 
            statut='ouverte'
        ).order_by('-date_ouverture').first()
    except:
        pass
    
    # Récupérer toutes les factures de la session active (toutes dates confondues)
    # Les factures utilisent maintenant la date saisie par l'utilisateur lors de la facturation
    if session_caisse:
        print(f"[ALERTE] MOUVEMENT_VENTE: Session trouvée {session_caisse.id}")
        factures_jour = FactureVente.objects.filter(
            agence=agence, 
            session_caisse=session_caisse
        ).select_related('client', 'vendeur__compte', 'session_caisse').prefetch_related('lignes__article').order_by('-date', '-heure')
        print(f"[ALERTE] MOUVEMENT_VENTE: {factures_jour.count()} factures trouvées")
        
        # Si aucune facture liée à la session, récupérer toutes les factures orphelines (toutes dates confondues) et les lier
        if factures_jour.count() == 0:
            print(f"[ALERTE] MOUVEMENT_VENTE: Aucune facture liée à la session, recherche des factures orphelines")
            factures_orphelines = FactureVente.objects.filter(
                agence=agence,
                session_caisse__isnull=True
            ).select_related('client', 'vendeur__compte').prefetch_related('lignes__article')
            
            if factures_orphelines.count() > 0:
                print(f"[ALERTE] MOUVEMENT_VENTE: {factures_orphelines.count()} factures orphelines trouvées, liaison à la session")
                # Lier ces factures à la session actuelle
                factures_orphelines.update(session_caisse=session_caisse)
                
                # Récupérer maintenant toutes les factures de la session
                factures_jour = FactureVente.objects.filter(
                    agence=agence, 
                    session_caisse=session_caisse
                ).select_related('client', 'vendeur__compte', 'session_caisse').prefetch_related('lignes__article').order_by('-date', '-heure')
                print(f"[ALERTE] MOUVEMENT_VENTE: {factures_jour.count()} factures après liaison")
    else:
        print(f"[ALERTE] MOUVEMENT_VENTE: Aucune session active")
        # Aucune session active, donc pas de factures à afficher
        factures_jour = FactureVente.objects.none()
    
    # Mettre à jour la date affichée selon les factures récupérées
    if factures_jour.exists():
        dates_couvertes = list(
            factures_jour.order_by('date').values_list('date', flat=True).distinct()
        )
        if dates_couvertes:
            date_reference = dates_couvertes[-1]
            if len(dates_couvertes) == 1:
                periode_affichage = date_reference.strftime('%d/%m/%Y')
            else:
                periode_affichage = f"{dates_couvertes[0].strftime('%d/%m/%Y')} au {dates_couvertes[-1].strftime('%d/%m/%Y')}"
    
    # Calculer les statistiques selon la même logique que le HTML
    total_tickets = factures_jour.count()
    total_articles = 0
    chiffre_affaires = 0
    derniere_vente_heure = None
    
    # Préparer les données pour le template (format similaire au localStorage)
    ventes_data = []
    tickets_uniques = set()
    
    for facture in factures_jour:
        tickets_uniques.add(facture.numero_ticket)
        chiffre_affaires += float(facture.nette_a_payer or 0)
        
        # Mettre à jour l'heure de la dernière vente
        if facture.heure:
            derniere_vente_heure = facture.heure.strftime('%H:%M')
        
        # Ajouter les lignes de la facture
        for ligne in facture.lignes.all():
            total_articles += ligne.quantite
            ventes_data.append({
                'ticket': facture.numero_ticket,
                'designation': ligne.designation,
                'reference': ligne.article.reference_article if ligne.article else '',
                'quantite': ligne.quantite,
                'prix_unitaire': float(ligne.prix_unitaire),
                'total': float(ligne.prix_total),
                'heure': facture.heure.strftime('%H:%M') if facture.heure else '-'
            })
    
    context = {
        'agence': agence,
        'date_reference': date_reference,
        'periode_affichage': periode_affichage,
        'session_caisse': session_caisse,
        'vendeuse_nom': vendeuse_nom,
        'total_tickets': len(tickets_uniques),
        'total_articles': total_articles,
        'derniere_vente_heure': derniere_vente_heure or '-',
        'chiffre_affaires': chiffre_affaires,
        'ventes_data': ventes_data,
    }
    
    return render(request, 'supermarket/caisse/mouvement_vente.html', context)

@login_required
@require_caisse_access
def documents_vente(request):
    """Afficher les documents de vente journaliers"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    # Récupérer les paramètres de filtrage
    search_query = request.GET.get('search', '')
    date_filter = request.GET.get('date_filter', 'all')
    
    # Récupérer tous les documents de vente
    documents = DocumentVente.objects.filter(agence=agence).order_by('-date', '-heure_fermeture')
    
    # Appliquer les filtres
    if search_query:
        documents = documents.filter(
            Q(numero_document__icontains=search_query) |
            Q(vendeuse_nom__icontains=search_query)
        )
    
    # Filtre par date
    from datetime import datetime, timedelta
    today = datetime.now().date()
    
    if date_filter == 'today':
        documents = documents.filter(date=today)
    elif date_filter == 'week':
        week_ago = today - timedelta(days=7)
        documents = documents.filter(date__gte=week_ago)
    elif date_filter == 'month':
        documents = documents.filter(date__year=today.year, date__month=today.month)
    elif date_filter == 'year':
        documents = documents.filter(date__year=today.year)
    
    # Calculer les statistiques
    total_documents = DocumentVente.objects.filter(agence=agence).count()
    chiffre_affaires_total = DocumentVente.objects.filter(agence=agence).aggregate(
        total=Sum('chiffre_affaires')
    )['total'] or 0
    total_factures = DocumentVente.objects.filter(agence=agence).aggregate(
        total=Sum('nombre_factures')
    )['total'] or 0
    
    # Dernier archivage
    dernier_archivage = None
    if documents.exists():
        dernier_doc = documents.first()
        dernier_archivage = dernier_doc.heure_fermeture
    
    context = {
        'agence': agence,
        'documents': documents,
        'total_documents': total_documents,
        'chiffre_affaires_total': chiffre_affaires_total,
        'total_factures': total_factures,
        'dernier_archivage': dernier_archivage,
        'search_query': search_query,
        'date_filter': date_filter,
    }
    
    return render(request, 'supermarket/caisse/documents_vente.html', context)

@login_required
@require_caisse_access
def document_vente_details(request, document_id):
    """Afficher les détails d'un document de vente"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    try:
        document = DocumentVente.objects.get(id=document_id, agence=agence)
        
        context = {
            'document': document,
            'factures_data': document.factures_data,
        }
        
        return render(request, 'supermarket/caisse/document_vente_details.html', context)
        
    except DocumentVente.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Document non trouvé.'})

@login_required
def facture_details(request, facture_id):

    print(f"DEBUG: facture_details appelée pour ID: {facture_id}")
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    print(f"DEBUG: Agence de l'utilisateur: {agence}")
    
    
    
    try:

        # Récupérer la facture avec ses lignes

        facture = FactureVente.objects.filter(id=facture_id, agence=agence).select_related('client', 'vendeur__compte').prefetch_related('lignes__article').first()

        print(f"DEBUG: Facture trouvée: {facture}")
        
        
        if not facture:

            print(f"DEBUG: Facture ID {facture_id} non trouvée")
            return JsonResponse({'success': False, 'error': 'Facture non trouvée.'})
        
        
        
        # Préparer les données de la facture

        facture_data = {

            'id': facture.id,

            'numero_ticket': facture.numero_ticket,

            'date': facture.date.strftime('%d/%m/%Y'),

            'heure': facture.heure.strftime('%H:%M'),

            'nette_a_payer': float(facture.nette_a_payer),

            'montant_regler': float(facture.montant_regler),

            'rendu': float(facture.rendu),

            'remise': float(facture.remise),

            'en_attente': facture.en_attente,

            'nom_vendeuse': facture.nom_vendeuse,

            'client': {

                'intitule': facture.client.intitule if facture.client else 'Client anonyme'

            } if facture.client else None,

            'lignes': []

        }

        
        
        # Ajouter les lignes de facture

        for ligne in facture.lignes.all():

            ligne_data = {
                'ligne_id': ligne.id,
                'article_id': ligne.article.id if ligne.article else None,
                'designation': ligne.designation,
                'reference': ligne.article.reference_article if ligne.article else '',

                'quantite': ligne.quantite,

                'prix_unitaire': float(ligne.prix_unitaire),

                'prix_total': float(ligne.prix_total),

                'type_vente': 'Détail'  # Type de vente par défaut
            }

            facture_data['lignes'].append(ligne_data)
        
        

        print(f"DEBUG: Données de la facture préparées: {len(facture_data['lignes'])} lignes")
        return JsonResponse({'success': True, 'facture': facture_data})
        
        
        
    except Exception as e:

        print(f"DEBUG: Erreur lors de la récupération des détails de la facture: {str(e)}")

        return JsonResponse({'success': False, 'error': str(e)})
@login_required
@require_caisse_access
def fermer_caisse(request):


    """Fermer la caisse et créer un document de vente journalier"""
    if request.method == 'GET':
        # Afficher une page de confirmation pour la fermeture
        agence = get_user_agence(request)
        if not agence:
            messages.error(request, 'Votre compte n\'est pas configuré correctement.')
            return redirect('logout_caisse')

        # Vérifier s'il y a une session ouverte (peu importe la date d'ouverture)
        session_caisse = SessionCaisse.objects.filter(
            agence=agence,
            statut='ouverte'
        ).order_by('-date_ouverture').first()
        
        if not session_caisse:
            messages.error(request, 'Aucune session de caisse ouverte.')
            return redirect('dashboard_caisse')
        
        # Récupérer le nom du compte connecté
        try:
            compte_connecte = Compte.objects.get(user=request.user, actif=True)
            vendeuse_nom = compte_connecte.nom_complet
        except Compte.DoesNotExist:
            vendeuse_nom = session_caisse.employe.compte.nom_complet if session_caisse.employe else 'Vendeur'
        
        # Calculer la période couverte par les factures de la session
        periode_affichage = None
        if session_caisse:
            factures_session = FactureVente.objects.filter(
                agence=agence,
                session_caisse=session_caisse
            ).order_by('date')
            dates_couvertes = list(factures_session.values_list('date', flat=True).distinct())
            if dates_couvertes:
                if len(dates_couvertes) == 1:
                    periode_affichage = dates_couvertes[0].strftime('%d/%m/%Y')
                else:
                    periode_affichage = f"{dates_couvertes[0].strftime('%d/%m/%Y')} au {dates_couvertes[-1].strftime('%d/%m/%Y')}"
        
        context = {
            'session_caisse': session_caisse,
            'agence': agence,
            'vendeuse_nom': vendeuse_nom,
            'periode_affichage': periode_affichage,
        }
        return render(request, 'supermarket/caisse/fermer_caisse_confirmation.html', context)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    # Bloc try/except pour capturer toutes les erreurs
    try:
        agence = get_user_agence(request)
        if not agence:
            return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
        
        # Récupérer la session de caisse ouverte (peu importe la date d'ouverture)
        aujourd_hui = timezone.now().date()
        session_caisse = SessionCaisse.objects.filter(
            agence=agence,
            statut='ouverte'
        ).order_by('-date_ouverture').first()
        
        if not session_caisse:
            return JsonResponse({'success': False, 'error': 'Aucune session de caisse ouverte trouvée'})
        
        # Récupérer toutes les factures de la session de caisse ouverte (toutes dates confondues)
        # Les factures utilisent maintenant la date saisie par l'utilisateur lors de la facturation
        factures_jour = FactureVente.objects.filter(
            agence=agence,
            session_caisse=session_caisse
        ).select_related('client', 'vendeur__compte').prefetch_related('lignes__article').order_by('-date', '-heure')
        
        # Si aucune facture liée à la session, récupérer toutes les factures orphelines (toutes dates confondues) et les lier
        if factures_jour.count() == 0:
            # Récupérer toutes les factures orphelines non liées à une session
            factures_orphelines = FactureVente.objects.filter(
                agence=agence,
                session_caisse__isnull=True
            ).select_related('client', 'vendeur__compte').prefetch_related('lignes__article')
            
            # Lier ces factures à la session actuelle
            factures_orphelines.update(session_caisse=session_caisse)
            
            # Récupérer maintenant toutes les factures de la session
            factures_jour = FactureVente.objects.filter(
                agence=agence,
                session_caisse=session_caisse
            ).select_related('client', 'vendeur__compte').prefetch_related('lignes__article').order_by('-date', '-heure')
        
        # Calculer les statistiques
        nombre_factures = factures_jour.count()
        total_articles = 0
        chiffre_affaires = 0.0  # Initialisé comme float pour éviter les problèmes de sérialisation JSON
        
        # Déterminer la date du document : utiliser la date la plus récente des factures
        # ou la date d'aujourd'hui s'il n'y a pas de factures
        date_document = aujourd_hui
        if factures_jour.exists():
            # Récupérer la date la plus récente parmi les factures
            date_document = factures_jour.first().date
        
        # Préparer les données des factures pour le document
        factures_data = []
        
        for facture in factures_jour:
            chiffre_affaires += float(facture.nette_a_payer or 0)
            
            # Préparer les données de cette facture
            facture_data = {
                'numero_ticket': facture.numero_ticket,
                'date': facture.date.strftime('%Y-%m-%d'),
                'heure': facture.heure.strftime('%H:%M') if facture.heure else '',
                'client': facture.client.intitule if facture.client else 'Client anonyme',
                'nette_a_payer': float(facture.nette_a_payer or 0),
                'articles': []
            }
            
            # Ajouter les articles de cette facture
            for ligne in facture.lignes.all():
                total_articles += int(ligne.quantite)
                facture_data['articles'].append({
                    'designation': ligne.designation,
                    'reference': ligne.article.reference_article if ligne.article else '',
                    'quantite': int(ligne.quantite),
                    'prix_unitaire': float(ligne.prix_unitaire or 0),
                    'total': float(ligne.prix_total or 0)
                })
            
            factures_data.append(facture_data)
        
        # Vérifier si un document existe déjà pour cette session
        document_existant = DocumentVente.objects.filter(session_caisse=session_caisse).first()
        if document_existant:
            # La caisse a déjà été fermée pour cette session
            messages.info(request, f'Caisse déjà fermée. Document {document_existant.numero_document} existant.')
            return JsonResponse({
                'success': True,
                'message': f'Caisse déjà fermée. Document {document_existant.numero_document} existant.',
                'document_id': document_existant.id,
                'redirect_url': '/caisse/'
            })
        
        # Générer un numéro de document unique avec un compteur incrémental
        # Utiliser la date du document (date la plus récente des factures)
        base_numero = f"DOC{date_document.strftime('%Y%m%d')}{session_caisse.id:03d}"
        numero_document = base_numero
        compteur = 1
        
        # S'assurer que le numéro est unique en ajoutant un suffixe si nécessaire
        while DocumentVente.objects.filter(numero_document=numero_document).exists():
            numero_document = f"{base_numero}-{compteur:02d}"
            compteur += 1
        
        # Récupérer le nom de la vendeuse depuis le compte connecté
        try:
            compte_connecte = Compte.objects.get(user=request.user, actif=True)
            vendeuse_nom = compte_connecte.nom_complet
        except Compte.DoesNotExist:
            # Fallback sur le compte de la session si le compte connecté n'est pas trouvé
            vendeuse_nom = session_caisse.employe.compte.nom_complet if session_caisse.employe else 'Vendeur'
        
        # Créer le document de vente
        # Convertir chiffre_affaires en Decimal pour le modèle
        import json
        
        # S'assurer que toutes les valeurs dans factures_data sont sérialisables en JSON
        # Cela force la sérialisation et détecte tout problème avec des Decimal
        try:
            json.dumps(factures_data)
            print("DEBUG: factures_data est sérialisable en JSON")
        except TypeError as e:
            print(f"ERREUR: Données non sérialisables dans factures_data: {e}")
            print(f"factures_data = {factures_data}")
            raise
        
        fermeture_heure = timezone.now().time()
        if factures_jour.exists():
            reference_facture = factures_jour.first()
            if reference_facture.heure:
                fermeture_heure = reference_facture.heure
        
        fermeture_datetime = datetime.combine(date_document, fermeture_heure)
        if timezone.is_naive(fermeture_datetime):
            fermeture_datetime = timezone.make_aware(
                fermeture_datetime,
                timezone.get_current_timezone()
            )
        
        document_vente = DocumentVente.objects.create(
            numero_document=numero_document,
            date=date_document,  # Utiliser la date la plus récente des factures (date saisie par l'utilisateur)
            heure_fermeture=fermeture_datetime,
            session_caisse=session_caisse,
            vendeuse_nom=vendeuse_nom,
            nombre_factures=nombre_factures,
            total_articles=total_articles,
            chiffre_affaires=Decimal(str(chiffre_affaires)),
            factures_data=factures_data,
            agence=agence
        )
        
        # Fermer la session de caisse
        session_caisse.statut = 'fermee'
        session_caisse.date_fermeture = fermeture_datetime
        session_caisse.save()
        
        # Fermer la caisse
        caisse = session_caisse.caisse
        caisse.statut = 'fermee'
        caisse.date_fermeture = fermeture_datetime
        caisse.save()
        
        messages.success(request, f'Caisse fermée avec succès! Document {numero_document} créé.')
        
        # Rediriger vers le dashboard pour voir le changement immédiatement
        return JsonResponse({
            'success': True,
            'message': f'Caisse fermée avec succès! Document {numero_document} créé.',
            'document_id': document_vente.id,
            'redirect_url': '/caisse/'
        })
    
    except Exception as e:
        # Afficher l'erreur complète dans la console pour le debugging
        import traceback
        print("=" * 80)
        print("ERREUR LORS DE LA FERMETURE DE CAISSE:")
        print(str(e))
        print(traceback.format_exc())
        print("=" * 80)
        
        # Retourner une erreur JSON claire
        return JsonResponse({
            'success': False, 
            'error': f'Erreur lors de la fermeture de caisse: {str(e)}'
        })


@login_required
@require_caisse_access
def finaliser_facture(request, facture_id):

    messages.info(request, 'Fonction en développement')

    return redirect('facturation_vente')



@login_required
@require_caisse_access
def mettre_en_attente(request):


    """Mettre une facture en attente"""
    print("=" * 80)
    print("🔵 DEBUG METTRE_EN_ATTENTE: Début de la fonction")
    print(f"🔵 DEBUG: Méthode de requête: {request.method}")
    print(f"🔵 DEBUG: Données POST: {dict(request.POST)}")
    print(f"🔵 DEBUG: Session key: {request.session.session_key}")
    print(f"🔵 DEBUG: Utilisateur: {request.user.username if request.user.is_authenticated else 'Non authentifié'}")
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        # Récupérer les données de la facture depuis la requête
        facture_data = request.POST.get('facture_data')
        print(f"🔵 DEBUG: facture_data reçu (longueur: {len(facture_data) if facture_data else 0}): {facture_data[:200] if facture_data else 'VIDE'}...")
        
        if not facture_data:
            print("🔵 DEBUG: [ERREUR] Aucune donnée facture_data trouvée dans POST")
            return JsonResponse({'success': False, 'error': 'Aucune donnée de facture fournie'})
        
        import json
        facture_content = json.loads(facture_data)
        print(f"🔵 DEBUG: [OK] facture_content parsé avec succès")
        print(f"🔵 DEBUG: Clés disponibles: {list(facture_content.keys())}")
        
        lignes = facture_content.get('lignes', [])
        print(f"🔵 DEBUG: Nombre de lignes: {len(lignes)}")
        
        for i, ligne in enumerate(lignes):
            print(f"🔵 DEBUG: Ligne {i}: article_id={ligne.get('article_id', 'VIDE')}, designation={ligne.get('designation', 'VIDE')}, quantite={ligne.get('quantite', 'VIDE')}")
        
        # Récupérer la session de caisse active (optionnel)
        agence = get_user_agence(request)
        session_caisse = None
        
        if agence:
            aujourd_hui = timezone.now().date()
            session_caisse = SessionCaisse.objects.filter(
                agence=agence,
                date_ouverture__date=aujourd_hui,
                statut='ouverte'
            ).first()
        
        # Utiliser la session key comme identifiant unique
        session_key = request.session.session_key
        if not session_key:
            request.session.create()
            session_key = request.session.session_key
        
        # Pour la mise en attente, on accepte les lignes même sans article_id
        # car elles peuvent être restaurées lors du rappel
        lignes_a_sauvegarder = []
        for ligne in facture_content.get('lignes', []):
            if ligne.get('article_id'):
                lignes_a_sauvegarder.append(ligne)
                print(f"DEBUG: Ligne avec article_id sauvegardée: {ligne.get('designation')} (ID: {ligne.get('article_id')})")
            else:
                # Inclure quand même la ligne pour la mise en attente
                lignes_a_sauvegarder.append(ligne)
                print(f"DEBUG: Ligne sans article_id incluse pour mise en attente: {ligne.get('designation', 'Sans désignation')}")
        
        # Mettre à jour le contenu avec toutes les lignes
        facture_content['lignes'] = lignes_a_sauvegarder
        print(f"DEBUG: Sauvegarde de {len(lignes_a_sauvegarder)} lignes (avec ou sans article_id)")
        
        # TOUJOURS créer un nouveau ticket en attente (ne pas utiliser get_or_create)
        # Chaque mise en attente = 1 nouveau ticket
        facture_temp = FactureTemporaire.objects.create(
            session_key=session_key,
            contenu=facture_content,
            session_caisse=session_caisse
        )
        print(f"🔵 DEBUG: Nouveau ticket en attente {facture_temp.id} créé pour session {session_caisse.id if session_caisse else 'None'}")
        
        # Mettre à jour la session Django avec les données de la facture
        request.session['facture_temporaire'] = facture_content
        print(f"DEBUG: Session mise à jour avec {len(facture_content.get('lignes', []))} lignes")
        
        # Compter UNIQUEMENT les factures en attente de la session courante
        if session_caisse:
            tickets_en_attente = FactureTemporaire.objects.filter(session_caisse=session_caisse).count()
        else:
            # Si pas de session, compter celles sans session
            tickets_en_attente = FactureTemporaire.objects.filter(session_caisse__isnull=True).count()
        
        return JsonResponse({
            'success': True,
            'message': 'Facture mise en attente avec succès',
            'tickets_en_attente': tickets_en_attente,
            'facture_id': facture_temp.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Format de données invalide'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erreur lors de la mise en attente: {str(e)}'})


@login_required
@require_caisse_access
def rappeler_ticket(request):

    """Rappeler une facture en attente"""
    print("=" * 80)
    print("🟡 DEBUG RAPPELER_TICKET: Début de la fonction")
    print(f"🟡 DEBUG: Méthode: {request.method}")
    print(f"🟡 DEBUG: Utilisateur: {request.user.username if request.user.is_authenticated else 'Non authentifié'}")
    
    if request.method != 'POST':
        print("🟡 DEBUG: [ERREUR] Méthode non autorisée")
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        session_key = request.session.session_key
        print(f"🟡 DEBUG: Session key: {session_key}")
        
        if not session_key:
            print("🟡 DEBUG: [ERREUR] Aucune session active")
            return JsonResponse({'success': False, 'error': 'Aucune session active'})
        
        # Récupérer la facture temporaire la plus récente (toutes sessions confondues)
        try:
            # Rechercher dans toutes les sessions pour éviter les problèmes de session key
            factures_attente = FactureTemporaire.objects.all().order_by('-date_creation')
            print(f"🟡 DEBUG: Nombre total de factures temporaires: {factures_attente.count()}")
            
            for ft in factures_attente:
                print(f"🟡 DEBUG: Facture {ft.id} créée le {ft.date_creation} (session: {ft.session_key})")
            
            facture_temp = factures_attente.first()
            if not facture_temp:
                print("🟡 DEBUG: [ERREUR] Aucune facture en attente trouvée")
                return JsonResponse({'success': False, 'error': 'Aucune facture en attente trouvée'})
            
            print(f"🟡 DEBUG: [OK] Facture temporaire sélectionnée: {facture_temp.id}")
            print(f"🟡 DEBUG: Date de création: {facture_temp.date_creation}")
            print(f"🟡 DEBUG: Session de la facture: {facture_temp.session_key}")
            print(f"🟡 DEBUG: Session actuelle: {session_key}")
            
            contenu = facture_temp.contenu
            if isinstance(contenu, dict):
                print(f"🟡 DEBUG: [OK] Contenu est un dictionnaire")
                print(f"🟡 DEBUG: Clés du contenu: {list(contenu.keys())}")
                lignes = contenu.get('lignes', [])
                print(f"🟡 DEBUG: Nombre de lignes dans le contenu: {len(lignes)}")
                
                for i, ligne in enumerate(lignes):
                    print(f"🟡 DEBUG: Ligne {i}: article_id={ligne.get('article_id', 'VIDE')}, designation={ligne.get('designation', 'VIDE')}")
            else:
                print(f"🟡 DEBUG: [ERREUR] Contenu n'est pas un dictionnaire: {type(contenu)}")
                print(f"🟡 DEBUG: Contenu brut: {str(contenu)[:200]}...")
                
        except Exception as e:
            print(f"🟡 DEBUG: [ERREUR] Erreur lors de la récupération: {str(e)}")
            import traceback
            print(f"🟡 DEBUG: Traceback: {traceback.format_exc()}")
            return JsonResponse({'success': False, 'error': 'Erreur lors de la récupération de la facture'})
        
        # Retourner les données de la facture
        print("DEBUG: Envoi de la réponse de succès")
        import json
        # Convertir le contenu en JSON string si c'est un dictionnaire
        contenu_json = facture_temp.contenu if isinstance(facture_temp.contenu, str) else json.dumps(facture_temp.contenu)
        
        # Sauvegarder les infos avant suppression
        facture_id_supprime = facture_temp.id
        date_creation_str = facture_temp.date_creation.strftime('%d/%m/%Y %H:%M')
        
        # IMPORTANT: Supprimer le ticket après l'avoir rappelé
        facture_temp.delete()
        print(f"🟡 DEBUG: ✅ Ticket {facture_id_supprime} supprimé après rappel")
        
        return JsonResponse({
            'success': True,
            'message': 'Facture rappelée avec succès',
            'facture_data': contenu_json,
            'facture_id': facture_id_supprime,
            'date_creation': date_creation_str
        })
        
    except Exception as e:
        print(f"DEBUG: Erreur lors du rappel: {str(e)}")
        import traceback
        print(f"DEBUG: Traceback: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': f'Erreur lors du rappel: {str(e)}'})


@login_required
@require_caisse_access
def lister_factures_attente(request):
    """Lister toutes les factures en attente pour la session courante"""
    try:
        session_key = request.session.session_key
        if not session_key:
            return JsonResponse({'success': False, 'error': 'Aucune session active'})
        
        # Récupérer toutes les factures en attente pour cette session
        factures_attente = FactureTemporaire.objects.filter(
            session_key=session_key
        ).order_by('-date_creation')
        
        factures_data = []
        for facture in factures_attente:
            import json
            try:
                # facture.contenu est déjà un dictionnaire Python (JSONField)
                contenu = facture.contenu if isinstance(facture.contenu, dict) else json.loads(facture.contenu)
                lignes = contenu.get('lignes', [])
                total = contenu.get('total', 0)
                rendu = contenu.get('rendu', 0)
                client = contenu.get('client', '')
                
                factures_data.append({
                    'id': facture.id,
                    'date_creation': facture.date_creation.strftime('%d/%m/%Y %H:%M'),
                    'nombre_articles': len(lignes),
                    'total': total,
                    'rendu': rendu,
                    'client': client,
                    'contenu': json.dumps(contenu)  # Convertir en JSON string pour le frontend
                })
            except Exception as e:
                print(f"DEBUG: Erreur lors du parsing de la facture {facture.id}: {e}")
                factures_data.append({
                    'id': facture.id,
                    'date_creation': facture.date_creation.strftime('%d/%m/%Y %H:%M'),
                    'nombre_articles': 0,
                    'total': 0,
                    'rendu': 0,
                    'client': '',
                    'contenu': '{}'
                })
        
        return JsonResponse({
            'success': True,
            'factures': factures_data,
            'nombre_factures': len(factures_data)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erreur lors du listage: {str(e)}'})


@login_required
@require_caisse_access
def rappeler_facture_specifique(request):
    """Rappeler une facture spécifique par son ID"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        facture_id = request.POST.get('facture_id')
        if not facture_id:
            return JsonResponse({'success': False, 'error': 'ID de facture manquant'})
        
        session_key = request.session.session_key
        if not session_key:
            return JsonResponse({'success': False, 'error': 'Aucune session active'})
        
        # Récupérer la facture spécifique
        try:
            facture_temp = FactureTemporaire.objects.get(
                id=facture_id,
                session_key=session_key
            )
        except FactureTemporaire.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Facture non trouvée'})
        
        print(f"DEBUG: Facture spécifique trouvée: {facture_temp.id}")
        print(f"DEBUG: Contenu: {facture_temp.contenu}")
        
        import json
        # Convertir le contenu en JSON string si c'est un dictionnaire
        contenu_json = facture_temp.contenu if isinstance(facture_temp.contenu, str) else json.dumps(facture_temp.contenu)
        
        return JsonResponse({
            'success': True,
            'message': 'Facture rappelée avec succès',
            'facture_data': contenu_json,
            'facture_id': facture_temp.id,
            'date_creation': facture_temp.date_creation.strftime('%d/%m/%Y %H:%M')
        })
        
    except Exception as e:
        print(f"DEBUG: Erreur lors du rappel spécifique: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erreur lors du rappel: {str(e)}'})


def test_urls_page(request):
    """Page de test pour vérifier les URLs des tickets en attente"""
    return render(request, 'supermarket/caisse/test_urls.html')


@login_required
@require_caisse_access
def supprimer_facture_attente(request):
    """Supprimer une facture en attente par son ID"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        facture_id = request.POST.get('facture_id')
        if not facture_id:
            return JsonResponse({'success': False, 'error': 'ID de facture manquant'})
        
        session_key = request.session.session_key
        if not session_key:
            return JsonResponse({'success': False, 'error': 'Aucune session active'})
        
        # Récupérer et supprimer la facture spécifique
        try:
            facture_temp = FactureTemporaire.objects.get(
                id=facture_id,
                session_key=session_key
            )
            
            # Supprimer la facture
            facture_temp.delete()
            
            print(f"DEBUG: Facture temporaire {facture_id} supprimée avec succès")
            
            # Compter le nombre restant de factures en attente
            tickets_restants = FactureTemporaire.objects.filter(session_key=session_key).count()
            
            return JsonResponse({
                'success': True,
                'message': 'Ticket en attente supprimé avec succès',
                'tickets_restants': tickets_restants
            })
            
        except FactureTemporaire.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Facture non trouvée'})
        
    except Exception as e:
        print(f"DEBUG: Erreur lors de la suppression: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erreur lors de la suppression: {str(e)}'})


@login_required
@require_caisse_access
def supprimer_ligne_facture(request):

    messages.info(request, 'Fonction en développement')

    return redirect('facturation_vente')



@login_required
@require_caisse_access
def modifier_quantite_ligne(request):

    messages.info(request, 'Fonction en développement')

    return redirect('facturation_vente')



@login_required
@require_caisse_access
def supprimer_vente(request):

    messages.info(request, 'Fonction en développement')

    return redirect('facturation_vente')



@login_required
@require_caisse_access
def imprimer_facture(request, facture_id=None):
    """Afficher et imprimer une facture depuis le formulaire ou depuis la base"""
    print(f"[ALERTE] IMPRIMER_FACTURE: {request.user.username}")
    print(f"[ALERTE] IMPRIMER_FACTURE: Méthode: {request.method}")
    print(f"[ALERTE] IMPRIMER_FACTURE: GET params: {dict(request.GET)}")
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('facturation_vente')
    
    # Vérifier si des données sont passées en paramètres GET (depuis le JavaScript)
    if request.GET.get('numero_ticket'):
        print(f"[ALERTE] IMPRIMER_FACTURE: Données reçues depuis JavaScript")
        try:
            import json
            
            # Récupérer les paramètres depuis l'URL
            numero_ticket = request.GET.get('numero_ticket', '')
            date = request.GET.get('date', '')
            heure = request.GET.get('heure', '')
            caisse = request.GET.get('caisse', '')
            vendeur = request.GET.get('vendeur', '')
            client = request.GET.get('client', '')
            nette_a_payer = float(request.GET.get('nette_a_payer', 0))
            montant_regler = float(request.GET.get('montant_regler', 0))
            rendu = float(request.GET.get('rendu', 0))
            remise = float(request.GET.get('remise', 0))
            
            # Récupérer les articles
            articles_json = request.GET.get('articles', '[]')
            articles = json.loads(articles_json)
            
            print(f"[ALERTE] IMPRIMER_FACTURE: {len(articles)} articles reçus")
            
            # Créer les données de la facture avec structure optimisée
            # S'assurer que chaque article a les bonnes clés
            lignes_formatees = []
            for article in articles:
                ligne = {
                    'designation': article.get('designation', article.get('nom_article', 'Article')),
                    'quantite': article.get('quantite', 1),
                    'prix_unitaire': article.get('prix_unitaire', 0),
                    'prix_total': article.get('prix_total', 0),
                    'reference': article.get('reference', ''),
                    'article_id': article.get('article_id', '')
                }
                lignes_formatees.append(ligne)
            
            facture_data = {
                'numero_ticket': numero_ticket,
                'date': date,
                'heure': heure,
                'nette_a_payer': nette_a_payer,
                'montant_regler': montant_regler,
                'rendu': rendu,
                'remise': remise,
                'nom_vendeuse': vendeur or 'Vendeur',
                'client_nom': client or 'Client anonyme',
                'caisse_numero': caisse or 'CAISSE001',
                'lignes': lignes_formatees
            }
            
            print(f"[ALERTE] IMPRIMER_FACTURE: Données préparées: {facture_data}")
            
            # Rendre le template d'impression professionnel
            return render(request, 'supermarket/caisse/facture_impression.html', {
                'facture': facture_data,
                'agence': agence
            })
            
        except Exception as e:
            print(f"[ALERTE] IMPRIMER_FACTURE: Erreur traitement GET: {e}")
            messages.error(request, f'Erreur lors de la préparation de l\'impression: {e}')
            return redirect('facturation_vente')
    
    # Si pas de paramètres GET, essayer de récupérer depuis la session
    print(f"[ALERTE] IMPRIMER_FACTURE: Pas de paramètres GET, récupération depuis la session")
    
    try:
        # Récupérer les données de la facture temporaire depuis la session
        facture_temp = request.session.get('facture_temporaire', {})
        
        if facture_temp and facture_temp.get('lignes'):
            print(f"[ALERTE] IMPRIMER_FACTURE: Données de session trouvées")
            
            # Calculer les totaux
            total_ht = sum(ligne.get('prix_total', 0) for ligne in facture_temp.get('lignes', []))
            remise = facture_temp.get('remise', 0)
            nette_a_payer = total_ht - remise
            montant_regler = facture_temp.get('montant_regler', 0)
            rendu = montant_regler - nette_a_payer
            
            # Générer un numéro de ticket
            from datetime import datetime
            now = datetime.now()
            numero_ticket = f'TICKET_{now.strftime("%Y%m%d_%H%M%S")}'
            
            # Formater les lignes de la session
            lignes_session = facture_temp.get('lignes', [])
            lignes_formatees = []
            for ligne in lignes_session:
                ligne_formatee = {
                    'designation': ligne.get('designation', ligne.get('nom_article', 'Article')),
                    'quantite': ligne.get('quantite', 1),
                    'prix_unitaire': ligne.get('prix_unitaire', 0),
                    'prix_total': ligne.get('prix_total', 0),
                    'reference': ligne.get('reference', ''),
                    'article_id': ligne.get('article_id', '')
                }
                lignes_formatees.append(ligne_formatee)
            
            facture_data = {
                'numero_ticket': numero_ticket,
                'date': now.strftime('%d/%m/%Y'),
                'heure': now.strftime('%H:%M'),
                'nette_a_payer': nette_a_payer,
                'montant_regler': montant_regler,
                'rendu': rendu,
                'remise': remise,
                'nom_vendeuse': 'Vendeur',
                'client_nom': facture_temp.get('client', 'Client anonyme'),
                'caisse_numero': 'CAISSE001',
                'lignes': lignes_formatees
            }
            
            print(f"[ALERTE] IMPRIMER_FACTURE: Données de session préparées: {facture_data}")
            
            # Rendre le template d'impression professionnel
            return render(request, 'supermarket/caisse/facture_impression.html', {
                'facture': facture_data,
                'agence': agence
            })
        else:
            print(f"[ALERTE] IMPRIMER_FACTURE: Aucune donnée de session, création de données de test")
            
            # Créer des données de test pour l'impression
            from datetime import datetime
            now = datetime.now()
            
            facture_data = {
                'numero_ticket': f'TICKET_{now.strftime("%Y%m%d_%H%M%S")}',
                'date': now.strftime('%d/%m/%Y'),
                'heure': now.strftime('%H:%M'),
                'nette_a_payer': 0,
                'montant_regler': 0,
                'rendu': 0,
                'remise': 0,
                'nom_vendeuse': 'Vendeur',
                'client_nom': 'Client anonyme',
                'caisse_numero': 'CAISSE001',
                'lignes': []
            }
            
            print(f"[ALERTE] IMPRIMER_FACTURE: Données de test créées: {facture_data}")
            
            # Rendre le template d'impression avec les données de test
            return render(request, 'supermarket/caisse/facture_impression.html', {
                'facture': facture_data,
                'agence': agence
            })
            
    except Exception as e:
        print(f"[ALERTE] IMPRIMER_FACTURE: Erreur générale: {e}")
        messages.error(request, f'Erreur lors de l\'impression: {e}')
        return redirect('facturation_vente')
@login_required
def init_test_data(request):

    """Initialiser des données de test pour le développement"""

    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    
    
    if not agence:

        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    
    
    try:

        # Créer quelques articles de test

        articles_data = [

            {'reference': 'ART001', 'designation': 'Coca Cola 33cl', 'prix_vente': 500},

            {'reference': 'ART002', 'designation': 'Pain de mie', 'prix_vente': 1000},

            {'reference': 'ART003', 'designation': 'Savon de Marseille', 'prix_vente': 400},

            {'reference': 'ART004', 'designation': 'Riz 5kg', 'prix_vente': 3000},

            {'reference': 'ART005', 'designation': 'Eau minérale 1.5L', 'prix_vente': 300},

        ]

        
        
        for article_data in articles_data:

            article, created = Article.objects.get_or_create(

                reference_article=article_data['reference'],

                defaults={

                    'designation': article_data['designation'],

                    'prix_vente': article_data['prix_vente'],

                    'stock_disponible': 100,

                    'agence': agence

                }

            )

            if created:

                print(f"Article créé: {article.designation}")
        
        
        
        # Créer des factures de test pour aujourd'hui

        aujourd_hui = timezone.now().date()

        
        
        # Créer une session de caisse de test

        session_caisse, created = SessionCaisse.objects.get_or_create(

            agence=agence,

            date_ouverture__date=aujourd_hui,

            defaults={

                'date_ouverture': timezone.now(),

                'statut': 'ouverte',

                'montant_initial': 10000

            }

        )

        
        
        if created:

            print(f"Session de caisse créée: {session_caisse}")
        
        
        
        # Créer quelques factures de test

        factures_test = [

            {'numero_ticket': 'TKT20250923001', 'nette_a_payer': 1500, 'en_attente': False},

            {'numero_ticket': 'TKT20250923002', 'nette_a_payer': 2300, 'en_attente': False},

            {'numero_ticket': 'TKT20250923003', 'nette_a_payer': 800, 'en_attente': True},

        ]

        
        
        for facture_data in factures_test:

            facture, created = FactureVente.objects.get_or_create(

                numero_ticket=facture_data['numero_ticket'],

                defaults={

                    'agence': agence,

                    'date': aujourd_hui,

                    'heure': timezone.now().time(),

                    'nette_a_payer': facture_data['nette_a_payer'],

                    'montant_regler': facture_data['nette_a_payer'],

                    'rendu': 0,

                    'remise': 0,

                    'en_attente': facture_data['en_attente'],

                    'nom_vendeuse': 'Test Vendeuse'

                }

            )

            if created:

                print(f"Facture créée: {facture.numero_ticket}")
        
        
        
        return JsonResponse({'success': True, 'message': 'Données de test initialisées avec factures'})
        
        
        
    except Exception as e:

        return JsonResponse({'success': False, 'error': str(e)})



@login_required
def update_quantity_temp(request):
    if request.method == 'POST':
        try:
            index = int(request.POST.get('index', 0))
            quantite = float(request.POST.get('quantite', 1))
            
            # Récupérer la facture temporaire de la session
            facture_temp = request.session.get('facture_temporaire', {'lignes': []})
            
            # Vérifier que l'index est valide
            if 0 <= index < len(facture_temp['lignes']):
                ligne = facture_temp['lignes'][index]
                
                # Mettre à jour la quantité
                ligne['quantite'] = quantite
                
                # Recalculer le prix total
                ligne['prix_total'] = quantite * ligne['prix_unitaire']
                
                # Sauvegarder en session
                request.session['facture_temporaire'] = facture_temp
                
                return JsonResponse({
                    'success': True,
                    'prix_total': ligne['prix_total']
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Index invalide'
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': True})



@login_required
@require_caisse_access
def remove_article_temp(request):
    """Supprimer un article de la facture temporaire dans la session"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        index = int(request.POST.get('index', -1))
        if index < 0:
            return JsonResponse({'success': False, 'error': 'Index invalide'})
        
        # Récupérer la facture temporaire de la session
        facture_temp = request.session.get('facture_temporaire', {'lignes': []})
        lignes = facture_temp.get('lignes', [])
        
        # Supprimer l'article à l'index spécifié
        if 0 <= index < len(lignes):
            lignes.pop(index)
            facture_temp['lignes'] = lignes
            
            # Recalculer les totaux
            total_ht = sum(ligne.get('prix_total', 0) for ligne in lignes)
            facture_temp['nette_a_payer'] = total_ht - facture_temp.get('remise', 0)
            
            # Mettre à jour la session
            request.session['facture_temporaire'] = facture_temp
            request.session.modified = True
            
            print(f"DEBUG: Article à l'index {index} supprimé. {len(lignes)} articles restants.")
            return JsonResponse({'success': True, 'message': 'Article supprimé avec succès'})
        else:
            return JsonResponse({'success': False, 'error': 'Index hors limites'})
            
    except Exception as e:
        print(f"DEBUG: Erreur lors de la suppression de l'article: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})



@login_required
@require_caisse_access
def update_montant_regler(request):

    return JsonResponse({'success': True})



@login_required
@require_caisse_access
def clear_facture_temp(request):

    request.session['facture_temporaire'] = {

        'lignes': [], 'type_vente': 'detail', 'remise': 0, 'montant_regler': 0, 'nette_a_payer': 0, 'rendu': 0

    }

    return JsonResponse({'success': True})



@login_required
@require_caisse_access
def update_type_vente_temp(request):

    if request.method == 'POST':

        try:

            index = int(request.POST.get('index'))

            type_vente = request.POST.get('type_vente')

            
            
            facture_temp = request.session.get('facture_temporaire', {'lignes': []})

            
            
            if 0 <= index < len(facture_temp['lignes']):

                ligne = facture_temp['lignes'][index]

                ligne['type_vente'] = type_vente

                
                
                # Récupérer le nouveau prix selon le type de vente

                try:
                    agence = get_user_agence(request)
                    if not agence:
                        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
                    
                    article = Article.objects.get(id=ligne['article_id'], agence=agence)

                    type_vente_obj = TypeVente.objects.filter(article=article, intitule__iexact=type_vente).first()

                    
                    
                    if type_vente_obj:

                        nouveau_prix = float(type_vente_obj.prix)

                    else:

                        nouveau_prix = float(article.prix_vente or 0)
                    
                    
                    
                    ligne['prix_unitaire'] = nouveau_prix

                    ligne['prix_total'] = nouveau_prix * ligne['quantite']

                    
                    
                    request.session['facture_temporaire'] = facture_temp

                    
                    
                    return JsonResponse({'success': True})

                except Exception as e:

                    return JsonResponse({'success': False, 'error': str(e)})

            else:

                return JsonResponse({'success': False, 'error': 'Index invalide'})
                
                
                
        except Exception as e:

            return JsonResponse({'success': False, 'error': str(e)})
    
    
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})



@login_required
@require_caisse_access
def generate_ticket_number_api(request):

    """API pour générer un nouveau numéro de ticket"""

    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    
    
    if not agence:

        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    
    
    numero_ticket = generate_ticket_number(agence)

    
    
    return JsonResponse({

        'success': True,

        'numero_ticket': numero_ticket

    })



@login_required
@require_caisse_access
def get_article_types(request):

    """Récupérer les types de vente disponibles pour un article"""

    article_id = request.GET.get('article_id')

    
    
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    try:
        article = Article.objects.get(id=article_id, agence=agence)

        
        
        # Récupérer tous les types de vente pour cet article

        types_vente = TypeVente.objects.filter(article=article).order_by('intitule')

        
        
        # Ajouter aussi le prix de vente par défaut comme option "Détail"

        types_data = []

        
        
        # Prix par défaut de l'article comme "Détail"

        types_data.append({

            'intitule': 'Détail',

            'prix': float(article.prix_vente or 0)

        })

        
        
        # Types de vente spécifiques de la base de données

        for type_vente in types_vente:

            types_data.append({

                'intitule': type_vente.intitule,

                'prix': float(type_vente.prix)

            })
        
        
        
        return JsonResponse({

            'success': True,

            'types': types_data,

            'article': {

                'id': article.id,

                'designation': article.designation,


                'prix_vente': float(article.prix_vente or 0)


            }


        })
        
        
        
    except Exception as e:

        return JsonResponse({

            'success': False,

            'error': str(e),

            'types': []

        })



@login_required
@require_caisse_access
def update_all_types_vente_temp(request):

    """Mettre à jour le type de vente global pour tous les articles de la facture temporaire"""

    if request.method == 'POST':

        try:

            type_vente_global = request.POST.get('type_vente_global')


            
            
            if not type_vente_global:


                return JsonResponse({'success': False, 'error': 'Type de vente non spécifié'})
            
            
            
            facture_temp = request.session.get('facture_temporaire', {'lignes': []})


            
            
            if not facture_temp.get('lignes'):

                return JsonResponse({'success': True, 'message': 'Aucun article dans la facture'})
            
            
            
            agence = get_user_agence(request)
            if not agence:
                return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
            
            try:

                
                
                # Mettre à jour chaque ligne de la facture temporaire

                for ligne in facture_temp['lignes']:

                    ligne['type_vente'] = type_vente_global

                    
                    
                    # Récupérer le nouveau prix selon le type de vente

                    article = Article.objects.get(id=ligne['article_id'], agence=agence)

                    type_vente_obj = TypeVente.objects.filter(article=article, intitule__iexact=type_vente_global).first()

                    
                    
                    if type_vente_obj:

                        nouveau_prix = float(type_vente_obj.prix)

                    else:

                        nouveau_prix = float(article.prix_vente or 0)
                    
                    
                    
                    ligne['prix_unitaire'] = nouveau_prix

                    ligne['prix_total'] = nouveau_prix * ligne['quantite']
                
                
                
                # Mettre à jour le type de vente global dans la facture temporaire

                facture_temp['type_vente'] = type_vente_global


                
                
                request.session['facture_temporaire'] = facture_temp

                
                
                return JsonResponse({
                    'success': True,
                    'message': f'Type de vente mis à jour vers {type_vente_global} pour tous les articles',
                    'lignes_updated': len(facture_temp['lignes'])
                })
                

                
                
            except Exception as e:

                return JsonResponse({'success': False, 'error': str(e)})
                

                
                
        except Exception as e:

            return JsonResponse({'success': False, 'error': str(e)})
    
    
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})



@login_required
@require_caisse_access
def get_article_types_vente(request, article_id):
    """Récupérer les types de vente d'un article"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    try:
        article = Article.objects.get(id=article_id, agence=agence)
        types_vente = TypeVente.objects.filter(article=article).order_by('intitule')
        
        types_data = []
        for type_vente in types_vente:
            types_data.append({
                'intitule': type_vente.intitule,
                'prix': float(type_vente.prix)
            })
        
        return JsonResponse({
            'success': True,
            'types_vente': types_data
        })
        
    except Article.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Article non trouvé'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_caisse_access
def update_type_vente(request):
    """Mettre à jour le type de vente d'un article"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    try:
        article_id = request.POST.get('article_id')
        type_vente = request.POST.get('type_vente')
        
        if not article_id or not type_vente:
            return JsonResponse({'success': False, 'error': 'Données manquantes'})
        
        article = Article.objects.get(id=article_id, agence=agence)
        type_vente_obj = TypeVente.objects.filter(
            article=article, 
            intitule__iexact=type_vente
        ).first()
        
        if type_vente_obj:
            nouveau_prix = float(type_vente_obj.prix)
        else:
            nouveau_prix = float(article.prix_vente or 0)
        
        return JsonResponse({
            'success': True,
            'nouveau_prix': nouveau_prix,
            'type_vente': type_vente
        })
        
    except Article.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Article non trouvé'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_caisse_access
def get_article_id_by_designation(request):
    """Récupérer l'ID d'un article par sa désignation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'})
    
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée'})
    
    try:
        designation = request.POST.get('designation', '').strip()
        reference = request.POST.get('reference', '').strip()
        
        if not designation:
            return JsonResponse({'success': False, 'error': 'Désignation manquante'})
        
        # Chercher l'article par désignation exacte d'abord
        article = Article.objects.filter(
            agence=agence,
            designation__iexact=designation
        ).first()
        
        # Si pas trouvé par désignation exacte, chercher par référence
        if not article and reference:
            article = Article.objects.filter(
                agence=agence,
                reference_article__iexact=reference
            ).first()
        
        # Si toujours pas trouvé, chercher par désignation contenant
        if not article:
            article = Article.objects.filter(
                agence=agence,
                designation__icontains=designation
            ).first()
        
        if article:
            return JsonResponse({
                'success': True,
                'article_id': article.id,
                'designation': article.designation,
                'reference': article.reference_article,
                'prix_vente': float(article.prix_vente or 0)
            })
        else:
            return JsonResponse({
                'success': False,
                'error': f'Article non trouvé pour: "{designation}"'
            })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})




@login_required
@require_caisse_access
def liste_articles(request):
    """Afficher la liste des articles avec recherche et statistiques"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    categorie_filter = request.GET.get('categorie', '')
    
    # Récupérer tous les articles de l'agence
    articles = Article.objects.filter(agence=agence).select_related('categorie').order_by('designation')
    
    # Appliquer les filtres
    if search_query:
        articles = articles.filter(
            Q(designation__icontains=search_query) |
            Q(reference_article__icontains=search_query)
        )
    
    if categorie_filter:
        articles = articles.filter(categorie__id=categorie_filter)
    
    # Récupérer les catégories pour le filtre
    categories = Article.objects.filter(agence=agence).values_list('categorie', flat=True).distinct()
    categories_list = [categorie for categorie in categories if categorie]
    
    # Calculer les statistiques
    total_articles = articles.count()
    stock_total = articles.aggregate(total=Sum('stock_actuel'))['total'] or 0
    valeur_stock = articles.aggregate(total=Sum(models.F('stock_actuel') * models.F('prix_vente')))['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'categories': categories_list,
        'search_query': search_query,
        'categorie_filter': categorie_filter,
        'total_articles': total_articles,
        'stock_total': stock_total,
        'valeur_stock': valeur_stock,
    }
    
    return render(request, 'supermarket/caisse/liste_articles.html', context)
@login_required
@require_caisse_access
def liste_clients(request):
    """Afficher la liste des clients avec statistiques"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_caisse')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    
    # Récupérer tous les clients de l'agence
    clients = Client.objects.filter(agence=agence).order_by('intitule')
    
    # Appliquer le filtre de recherche
    if search_query:
        clients = clients.filter(
            Q(intitule__icontains=search_query) |
            Q(telephone__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Calculer les statistiques pour chaque client
    clients_avec_stats = []
    for client in clients:
        # Statistiques des ventes
        factures_client = FactureVente.objects.filter(client=client, agence=agence)
        nombre_achats = factures_client.count()
        montant_total = factures_client.aggregate(total=Sum('nette_a_payer'))['total'] or 0
        
        # Articles les plus achetés
        articles_achetes = LigneFactureVente.objects.filter(
            facture_vente__client=client,
            facture_vente__agence=agence
        ).values('article__designation').annotate(
            total_quantite=Sum('quantite'),
            total_montant=Sum('prix_total')
        ).order_by('-total_quantite')[:3]
        
        clients_avec_stats.append({
            'client': client,
            'nombre_achats': nombre_achats,
            'montant_total': montant_total,
            'articles_achetes': articles_achetes,
        })
    
    # Statistiques générales
    total_clients = clients.count()
    clients_actifs = len([c for c in clients_avec_stats if c['nombre_achats'] > 0])
    montant_total_ventes = sum([c['montant_total'] for c in clients_avec_stats])
    
    # Statistiques pour les cartes du template
    clients_avec_email = clients.filter(email__isnull=False).exclude(email='').count()
    clients_sans_email = clients.filter(Q(email__isnull=True) | Q(email='')).count()
    clients_avec_achats = clients_actifs
    
    context = {
        'agence': agence,
        'clients': clients,  # Passer directement les clients pour le template
        'clients_avec_stats': clients_avec_stats,
        'search_query': search_query,
        'total_clients': total_clients,
        'clients_actifs': clients_actifs,
        'montant_total_ventes': montant_total_ventes,
        'clients_avec_email': clients_avec_email,
        'clients_sans_email': clients_sans_email,
        'clients_avec_achats': clients_avec_achats,
    }
    
    return render(request, 'supermarket/caisse/liste_clients.html', context)

@login_required
def test_flux_complet(request):
    """Vue de test pour le flux complet"""
    return render(request, 'test_flux_complet.html')

@login_required
def creer_article(request):
    """Vue pour créer un nouvel article"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            designation = request.POST.get('designation')
            prix_achat = request.POST.get('prix_achat')
            dernier_prix_achat = request.POST.get('dernier_prix_achat', prix_achat)  # Par défaut, égal au prix d'achat
            prix_vente = request.POST.get('prix_vente')
            stock_actuel = request.POST.get('stock_actuel', 0)
            stock_minimum = request.POST.get('stock_minimum', 0)
            unite_vente = request.POST.get('unite_vente')
            conditionnement = request.POST.get('conditionnement')
            # Le suivi de stock est toujours activé automatiquement (non modifiable par l'utilisateur)
            agence_id = request.POST.get('agence')
            famille_id = request.POST.get('famille')
            # Debug des champs reçus
            print(f"[SEARCH] DEBUG: Champs reçus:")
            print(f"[SEARCH] DEBUG: - designation: '{designation}'")
            print(f"[SEARCH] DEBUG: - prix_achat: '{prix_achat}'")
            print(f"[SEARCH] DEBUG: - prix_vente: '{prix_vente}'")
            print(f"[SEARCH] DEBUG: - agence_id: '{agence_id}'")
            print(f"[SEARCH] DEBUG: - famille_id: '{famille_id}'")
            print(f"[SEARCH] DEBUG: - unite_vente: '{unite_vente}'")
            print(f"[SEARCH] DEBUG: - conditionnement: '{conditionnement}'")
            
            # Validation des champs obligatoires
            if not designation or not prix_achat or not prix_vente or not agence_id or not famille_id:
                print(f"[SEARCH] DEBUG: [ERREUR] Validation échouée - Champs manquants:")
                if not designation: print("  - designation manquant")
                if not prix_achat: print("  - prix_achat manquant")
                if not prix_vente: print("  - prix_vente manquant")
                if not agence_id: print("  - agence_id manquant")
                if not famille_id: print("  - famille_id manquant")
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_article')
            
            # Récupérer l'agence et la famille
            try:
                agence = Agence.objects.get(id_agence=agence_id)
                famille = Famille.objects.get(id=famille_id)
            except (Agence.DoesNotExist, Famille.DoesNotExist):
                messages.error(request, 'Agence ou famille invalide.')
                return redirect('creer_article')
            
            # Générer automatiquement la référence article
            print(f"[SEARCH] DEBUG: Génération automatique de la référence...")
            
            # Récupérer la dernière référence
            last_article = Article.objects.order_by('-id').first()
            
            if last_article and last_article.reference_article:
                # Extraire le numéro de la dernière référence
                import re
                match = re.search(r'ART(\d+)', last_article.reference_article)
                if match:
                    last_number = int(match.group(1))
                    new_number = last_number + 1
                else:
                    new_number = 1
            else:
                new_number = 1
            
            # Générer la nouvelle référence
            reference_article = f"ART{new_number:03d}"
            
            # Vérifier que la référence n'existe pas déjà
            while Article.objects.filter(reference_article=reference_article).exists():
                new_number += 1
                reference_article = f"ART{new_number:03d}"
            
            print(f"[SEARCH] DEBUG: [OK] Référence générée: {reference_article}")
            
            # Créer l'article
            # Le suivi de stock est toujours activé automatiquement
            article = Article.objects.create(
                reference_article=reference_article,
                designation=designation,
                prix_achat=float(prix_achat),
                dernier_prix_achat=float(dernier_prix_achat) if dernier_prix_achat else float(prix_achat),
                prix_vente=float(prix_vente),
                stock_actuel=float(stock_actuel),
                stock_minimum=float(stock_minimum),
                unite_vente=unite_vente,
                conditionnement=conditionnement,
                agence=agence,
                categorie=famille,
                suivi_stock=True  # Toujours activé automatiquement
            )
            
            # Créer les types de vente avec leurs prix spécifiques
            # Récupérer les prix depuis le formulaire
            prix_gros = request.POST.get('prix_gros', '').strip()
            prix_demi_gros = request.POST.get('prix_demi_gros', '').strip()
            prix_detail = request.POST.get('prix_detail', '').strip()
            
            print(f"[DEBUG] Prix récupérés - Gros: '{prix_gros}', Demi-Gros: '{prix_demi_gros}', Détail: '{prix_detail}'")
            
            # Créer les types de vente avec les bons prix
            types_vente_data = [
                {'intitule': 'Gros', 'prix': float(prix_gros) if prix_gros else float(prix_vente)},
                {'intitule': 'Demi-Gros', 'prix': float(prix_demi_gros) if prix_demi_gros else float(prix_vente)},
                {'intitule': 'Détail', 'prix': float(prix_detail) if prix_detail else float(prix_vente)},
            ]
            
            for type_vente_info in types_vente_data:
                TypeVente.objects.create(
                    article=article,
                    intitule=type_vente_info['intitule'],
                    prix=type_vente_info['prix']
                )
                print(f"[DEBUG] Type de vente créé: {type_vente_info['intitule']} à {type_vente_info['prix']} FCFA")
            
            # Créer un mouvement de stock initial si stock > 0
            if float(stock_actuel) > 0:
                MouvementStock.objects.create(
                    article=article,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece=f'INIT-{article.id}',
                    quantite_stock=float(stock_actuel),
                    stock_initial=0,
                    solde=float(stock_actuel),
                    quantite=float(stock_actuel),
                    cout_moyen_pondere=float(prix_achat),
                    stock_permanent=float(stock_actuel) * float(prix_achat),
                    commentaire=f"Création article - Stock initial"
                )
            
            messages.success(request, f'Article "{designation}" créé avec succès!')
            return redirect('dashboard_stock')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de l\'article: {str(e)}')
            return redirect('creer_article')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    familles = Famille.objects.all()
    
    # Debug des familles disponibles
    print(f"[SEARCH] DEBUG: Familles disponibles: {familles.count()}")
    for famille in familles:
        print(f"[SEARCH] DEBUG: - Famille: {famille.intitule} (ID: {famille.id})")
    
    context = {
        'agences': agences,
        'familles': familles,
    }
    return render(request, 'supermarket/stock/creer_article.html', context)

@login_required
def generate_reference(request):
    """Vue pour générer une référence article automatique"""
    try:
        # Récupérer la dernière référence
        last_article = Article.objects.order_by('-id').first()
        
        if last_article and last_article.reference_article:
            # Extraire le numéro de la dernière référence
            import re
            match = re.search(r'ART(\d+)', last_article.reference_article)
            if match:
                last_number = int(match.group(1))
                new_number = last_number + 1
            else:
                new_number = 1
        else:
            new_number = 1
        
        # Générer la nouvelle référence
        new_reference = f"ART{new_number:03d}"
        
        # Vérifier que la référence n'existe pas déjà
        while Article.objects.filter(reference_article=new_reference).exists():
            new_number += 1
            new_reference = f"ART{new_number:03d}"
        
        return JsonResponse({
            'success': True,
            'reference': new_reference
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def consulter_articles(request):
    """Vue pour consulter la liste des articles"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    famille_id = request.GET.get('famille', '')
    stock_filter = request.GET.get('stock_filter', '')
    force_refresh = request.GET.get('refresh', '')
    
    # Forcer le rafraîchissement si demandé
    if force_refresh == '1':
        # Vider le cache de session pour forcer le rechargement
        if hasattr(request, 'session'):
            request.session.flush()
    
    # Construire la requête de base avec select_related pour éviter les requêtes multiples
    articles = Article.objects.filter(agence=agence).select_related('categorie')
    
    # Appliquer les filtres
    if search_query:
        articles = articles.filter(
            Q(designation__icontains=search_query) |
            Q(categorie__intitule__icontains=search_query)
        )
    
    if famille_id:
        articles = articles.filter(categorie_id=famille_id)
    
    if stock_filter == 'low':
        articles = articles.filter(stock_actuel__lte=F('stock_minimum'))
    elif stock_filter == 'zero':
        articles = articles.filter(stock_actuel__lte=0)
    
    # Trier par ID décroissant
    articles = articles.order_by('-id')
    
    # Calculer les statistiques
    total_articles = Article.objects.filter(agence=agence).count()
    articles_stock_faible = Article.objects.filter(
        agence=agence,
        stock_actuel__lte=F('stock_minimum')
    ).count()
    articles_rupture = Article.objects.filter(
        agence=agence,
        stock_actuel__lte=0
    ).count()
    
    # Valeur totale du stock
    articles_avec_prix = Article.objects.filter(
        agence=agence,
        prix_achat__isnull=False,
        stock_actuel__isnull=False
    ).exclude(prix_achat=0).exclude(stock_actuel=0)
    
    valeur_totale_stock = 0
    for article in articles_avec_prix:
        try:
            valeur_article = float(article.prix_achat) * float(article.stock_actuel)
            valeur_totale_stock += valeur_article
        except (ValueError, TypeError):
            continue
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Calculer les statistiques pour le template
    articles_stock_normal = total_articles - articles_stock_faible
    articles_stock_low = articles_stock_faible - articles_rupture
    articles_stock_zero = articles_rupture
    
    context = {
        'articles': articles,
        'agence': agence,
        'total_articles': total_articles,
        'articles_stock_normal': articles_stock_normal,
        'articles_stock_low': articles_stock_low,
        'articles_stock_zero': articles_stock_zero,
        'articles_stock_faible': articles_stock_faible,
        'articles_rupture': articles_rupture,
        'valeur_totale_stock': valeur_totale_stock,
        'familles': familles,
    }
    return render(request, 'supermarket/stock/consulter_articles.html', context)

@login_required
def consulter_fournisseurs(request):
    """Vue pour consulter la liste des fournisseurs"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    fournisseurs = Fournisseur.objects.filter(agence=agence).order_by('-id')
    context = {
        'fournisseurs': fournisseurs,
        'agence': agence,
    }
    return render(request, 'supermarket/stock/consulter_fournisseurs.html', context)

@login_required
def consulter_clients(request):
    """Vue pour consulter la liste des clients"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    agence_filter = request.GET.get('agence_filter', '')
    
    # Construire la requête de base - Limiter aux clients de l'agence de l'utilisateur
    clients = Client.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        clients = clients.filter(
            Q(intitule__icontains=search_query) |
            Q(telephone__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    if agence_filter:
        clients = clients.filter(agence_id=agence_filter)
    
    # Trier par ID décroissant
    clients = clients.order_by('-id')
    
    # Calculer les statistiques
    total_clients = Client.objects.filter(agence=agence).count()
    clients_avec_telephone = Client.objects.filter(agence=agence).exclude(telephone='').count()
    clients_avec_email = Client.objects.filter(agence=agence).exclude(email='').count()
    clients_recents = Client.objects.filter(agence=agence).count()  # Peut être amélioré avec une date
    
    # Récupérer toutes les agences pour le filtre
    agences = Agence.objects.all()
    
    context = {
        'clients': clients,
        'agence': agence,
        'total_clients': total_clients,
        'clients_avec_telephone': clients_avec_telephone,
        'clients_avec_email': clients_avec_email,
        'clients_recents': clients_recents,
        'agences': agences,
    }
    return render(request, 'supermarket/stock/consulter_clients.html', context)

@login_required
def creer_client(request):
    """Vue pour créer un nouveau client"""
    if request.method == 'POST':
        try:
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email', '')
            agence_id = request.POST.get('agence')
            
            # Validation des champs obligatoires
            if not all([intitule, adresse, telephone, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_client')
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence sélectionnée invalide.')
                return redirect('creer_client')
            
            # Créer le client
            client = Client.objects.create(
                intitule=intitule,
                adresse=adresse,
                telephone=telephone,
                email=email,
                agence=agence
            )
            
            messages.success(request, f'Client "{client.intitule}" créé avec succès!')
            return redirect('dashboard_stock')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du client: {str(e)}')
            return redirect('creer_client')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    context = {
        'agences': agences,
    }
    return render(request, 'supermarket/stock/creer_client.html', context)

@login_required
def creer_fournisseur(request):
    """Vue pour créer un nouveau fournisseur"""
    if request.method == 'POST':
        try:
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email', '')
            agence_id = request.POST.get('agence')
            
            # Validation des champs obligatoires
            if not all([intitule, adresse, telephone, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_fournisseur')
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence sélectionnée invalide.')
                return redirect('creer_fournisseur')
            
            # Créer le fournisseur
            fournisseur = Fournisseur.objects.create(
                intitule=intitule,
                adresse=adresse,
                telephone=telephone,
                email=email,
                agence=agence
            )
            
            messages.success(request, f'Fournisseur "{fournisseur.intitule}" créé avec succès!')
            return redirect('dashboard_stock')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du fournisseur: {str(e)}')
            return redirect('creer_fournisseur')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    context = {
        'agences': agences,
    }
    return render(request, 'supermarket/stock/creer_fournisseur.html', context)

@login_required
def detail_fournisseur(request, fournisseur_id):
    """Vue pour afficher les détails d'un fournisseur"""
    try:
        fournisseur = Fournisseur.objects.get(id=fournisseur_id, agence=get_user_agence(request))
        
        # Récupérer les factures d'achat du fournisseur (si elles existent)
        factures_achat = FactureAchat.objects.filter(fournisseur=fournisseur).order_by('-date_achat', '-heure')[:10]
        
        context = {
            'fournisseur': fournisseur,
            'factures_achat': factures_achat,
        }
        return render(request, 'supermarket/stock/detail_fournisseur.html', context)
    except Fournisseur.DoesNotExist:
        messages.error(request, 'Fournisseur non trouvé.')
        return redirect('consulter_fournisseurs')

@login_required
def modifier_fournisseur(request, fournisseur_id):
    """Vue pour modifier un fournisseur existant"""
    try:
        fournisseur = Fournisseur.objects.get(id=fournisseur_id, agence=get_user_agence(request))
    except Fournisseur.DoesNotExist:
        messages.error(request, 'Fournisseur non trouvé.')
        return redirect('consulter_fournisseurs')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([intitule, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_fournisseur', fournisseur_id=fournisseur_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_fournisseur', fournisseur_id=fournisseur_id)
            
            # Mettre à jour le fournisseur
            fournisseur.intitule = intitule
            fournisseur.adresse = adresse
            fournisseur.telephone = telephone
            fournisseur.email = email
            fournisseur.agence = agence
            fournisseur.save()
            
            messages.success(request, f'Fournisseur "{intitule}" modifié avec succès!')
            return redirect('consulter_fournisseurs')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du fournisseur: {str(e)}')
            return redirect('modifier_fournisseur', fournisseur_id=fournisseur_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    
    context = {
        'fournisseur': fournisseur,
        'agences': agences,
    }
    return render(request, 'supermarket/stock/modifier_fournisseur.html', context)

@login_required
def supprimer_fournisseur(request, fournisseur_id):
    """Vue pour supprimer un fournisseur"""
    if request.method == 'POST':
        try:
            fournisseur = Fournisseur.objects.get(id=fournisseur_id, agence=get_user_agence(request))
            fournisseur_name = fournisseur.intitule
            fournisseur.delete()
            messages.success(request, f'Fournisseur "{fournisseur_name}" supprimé avec succès!')
        except Fournisseur.DoesNotExist:
            messages.error(request, 'Fournisseur non trouvé.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression du fournisseur: {str(e)}')
    
    return redirect('consulter_fournisseurs')

@login_required
@require_stock_access
def dashboard_stock(request):
    """Dashboard principal du module de gestion de stock"""
    try:
        # Récupérer l'agence de l'utilisateur
        agence = get_user_agence(request)
        if not agence:
            messages.error(request, 'Votre compte n\'est pas configuré correctement.')
            return redirect('logout_stock')

        # Calculer les KPIs
        total_articles = Article.objects.filter(agence=agence).count()
        articles_stock_faible = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=10
        ).count()
        articles_rupture = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=0
        ).count()
        
        # Valeur totale du stock
        articles_avec_prix = Article.objects.filter(
            agence=agence,
            prix_achat__isnull=False,
            stock_actuel__isnull=False
        ).exclude(prix_achat=0).exclude(stock_actuel=0)
        
        valeur_totale_stock = 0
        for article in articles_avec_prix:
            try:
                valeur_article = float(article.prix_achat) * float(article.stock_actuel)
                valeur_totale_stock += valeur_article
            except (ValueError, TypeError):
                continue
        
        # Mouvements récents
        mouvements_recents = MouvementStock.objects.filter(agence=agence).order_by('-date_mouvement')[:5]
        
        # Articles les plus vendus (simulation)
        articles_populaires = Article.objects.filter(agence=agence).order_by('-stock_actuel')[:5]
        
        # Alertes de stock
        alertes_stock = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=5
        ).order_by('stock_actuel')[:5]

        # Récupérer le nom de l'utilisateur
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except Compte.DoesNotExist:
            nom_utilisateur = request.user.username

        context = {
            'agence': agence,
            'nom_utilisateur': nom_utilisateur,
            'total_articles': total_articles,
            'articles_stock_faible': articles_stock_faible,
            'articles_rupture': articles_rupture,
            'valeur_stock': valeur_totale_stock,  # Corrigé le nom de la variable
            'mouvements_recents': mouvements_recents,
            'articles_populaires': articles_populaires,
            'alertes_stock': alertes_stock,
        }
        
        return render(request, 'supermarket/stock/dashboard_stock.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur lors du chargement du dashboard: {str(e)}')
        # Récupérer le nom de l'utilisateur même en cas d'erreur
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except:
            nom_utilisateur = request.user.username if request.user.is_authenticated else "Utilisateur"

        return render(request, 'supermarket/stock/dashboard_stock.html', {
            'agence': None,
            'nom_utilisateur': nom_utilisateur,
            'total_articles': 0,
            'articles_stock_faible': 0,
            'articles_rupture': 0,
            'valeur_stock': 0,
            'mouvements_recents': [],
            'articles_populaires': [],
            'alertes_stock': [],
        })

def login_stock(request):
    """Page de connexion pour la gestion de stock"""
    # Toujours afficher la page de connexion, même si l'utilisateur est déjà connecté
    # Cela permet de forcer la reconnexion depuis la page d'accueil
    if request.user.is_authenticated and request.method == 'GET':
        # Déconnecter l'utilisateur pour forcer la reconnexion
        logout(request)
        request.session.flush()
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                # Vérifier que l'utilisateur a un compte lié à une agence
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        # Vérifier que l'utilisateur est un comptable, assistant_comptable ou admin
                        if compte.type_compte not in ['comptable', 'assistant_comptable', 'admin']:
                            messages.error(request, 'Accès refusé. Ce module est réservé aux comptables, assistants comptables et aux administrateurs.')
                        else:
                            login(request, user)
                            # Stocker l'agence dans la session
                            request.session['agence_id'] = compte.agence.id_agence
                            request.session['agence_nom'] = compte.agence.nom_agence
                            messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                            return redirect('dashboard_stock')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Aucun compte actif trouvé pour cet utilisateur.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')
    
    return render(request, 'supermarket/stock/login.html')

@login_required
@require_commandes_feature('dashboard')
def dashboard_commandes(request):
    """Dashboard principal du module de gestion des commandes"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')

    compte = get_user_compte(request)
    if not compte:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('login_commandes')
    
    # Rediriger vers les pages de travail personnalisées selon le type de compte
    if compte.type_compte == 'acm':
        return redirect('travail_acm')
    elif compte.type_compte == 'livreur':
        return redirect('travail_livreur')
    
    # Pour admin, gérant, etc. : afficher le dashboard normal
    nom_utilisateur = compte.nom_complet

    # Filtrer les données selon le type de compte
    commandes_queryset = Commande.objects.filter(agence=agence)
    commandes_queryset = filter_commandes_by_user(commandes_queryset, compte)
    
    livraisons_queryset = Livraison.objects.filter(agence=agence)
    livraisons_queryset = filter_livraisons_by_user(livraisons_queryset, compte)

    total_commandes = commandes_queryset.count()
    commandes_en_attente = commandes_queryset.filter(etat_commande='en_attente').count()
    commandes_en_livraison = commandes_queryset.filter(etat_commande='en_livraison').count()
    commandes_livrees = commandes_queryset.filter(etat_commande='livree').count()
    commandes_annulees = commandes_queryset.filter(etat_commande='annulee').count()

    total_livraisons = livraisons_queryset.count()
    livraisons_en_cours = livraisons_queryset.filter(etat_livraison='en_cours').count()
    total_factures = FactureCommande.objects.filter(agence=agence).count()
    total_statistiques = StatistiqueClient.objects.filter(agence=agence).count()

    commandes_recents = (
        commandes_queryset.select_related('client', 'article')
        .order_by('-date', '-heure')[:10]
    )

    context = {
        'agence': agence,
        'compte': compte,  # Ajouter le compte pour les vérifications dans le template
        'nom_utilisateur': nom_utilisateur,
        'total_commandes': total_commandes,
        'commandes_en_attente': commandes_en_attente,
        'commandes_en_livraison': commandes_en_livraison,
        'commandes_livrees': commandes_livrees,
        'commandes_annulees': commandes_annulees,
        'total_livraisons': total_livraisons,
        'livraisons_en_cours': livraisons_en_cours,
        'total_factures': total_factures,
        'total_statistiques': total_statistiques,
        'commandes_recents': commandes_recents,
    }
    return render(request, 'supermarket/commandes/dashboard.html', context)

def login_commandes(request):
    """Page de connexion pour le module gestion des commandes"""
    if request.user.is_authenticated and request.method == 'GET':
        logout(request)
        request.session.flush()

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        login(request, user)
                        request.session['agence_id'] = compte.agence.id_agence
                        request.session['agence_nom'] = compte.agence.nom_agence
                        messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                        return redirect('dashboard_commandes')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Aucun compte actif trouvé pour cet utilisateur.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')

    return render(request, 'supermarket/commandes/login.html')

@login_required
def logout_commandes(request):
    """Déconnexion pour le module commandes"""
    logout(request)
    return redirect('index')

@login_required
@require_commandes_feature('gestion_commande')
def enregistrer_commande(request):
    """Vue pour afficher le formulaire d'enregistrement de commande"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    # Récupérer le client_id depuis les paramètres GET si présent
    client_id_from_get = request.GET.get('client_id')
    
    # Récupérer les clients de l'agence
    clients = Client.objects.filter(agence=agence).order_by('intitule')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Initialiser la session de commande temporaire
    if 'commande_temporaire' not in request.session:
        request.session['commande_temporaire'] = {
            'lignes': [],
            'client_id': None,
            'adresse': '',
            'ville': '',
            'telephone': ''
        }
    
    commande_temp = request.session.get('commande_temporaire', {
        'lignes': [],
        'client_id': None,
        'adresse': '',
        'ville': '',
        'telephone': ''
    })
    
    # Si un client_id est fourni en GET, pré-remplir le client dans la session
    if client_id_from_get:
        try:
            client = Client.objects.get(id=client_id_from_get, agence=agence)
            commande_temp['client_id'] = client.id
            commande_temp['adresse'] = client.adresse or ''
            commande_temp['telephone'] = client.telephone or ''
            request.session['commande_temporaire'] = commande_temp
        except Client.DoesNotExist:
            pass
    
    print(f"[ENREGISTRER COMMANDE] Session récupérée: {len(commande_temp.get('lignes', []))} lignes")
    if commande_temp.get('lignes'):
        print(f"[ENREGISTRER COMMANDE] Lignes: {commande_temp['lignes']}")
    
    # Préparer les articles en format JSON pour JavaScript
    import json
    articles_json = json.dumps([
        {
            'id': article.id,
            'designation': article.designation,
            'reference_article': article.reference_article or '',
            'prix_vente': float(article.prix_vente or 0)
        }
        for article in articles
    ])
    
    # Préparer les clients en format JSON pour JavaScript
    clients_json = json.dumps([
        {
            'id': client.id,
            'intitule': client.intitule,
            'adresse': client.adresse or '',
            'telephone': client.telephone or '',
            'email': client.email or ''
        }
        for client in clients
    ])
    
    from django.utils import timezone
    today = timezone.now().date()
    now = timezone.now()
    
    context = {
        'agence': agence,
        'clients': clients,
        'articles': articles,
        'articles_json': articles_json,
        'clients_json': clients_json,
        'commande_temporaire': commande_temp,
        'today': today,
        'now': now,
    }
    return render(request, 'supermarket/commandes/enregistrer_commande.html', context)

@csrf_exempt
@login_required
def ajouter_article_commande(request):
    """Vue pour ajouter un article à la commande temporaire"""
    if request.method == 'POST':
        agence = get_user_agence(request)
        if not agence:
            return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
        
        try:
            article_id = request.POST.get('article_id')
            quantite = Decimal(str(request.POST.get('quantite', 1)))
            
            article = Article.objects.get(id=article_id, agence=agence)
            
            commande_temp = request.session.get('commande_temporaire', {
                'lignes': [],
                'client_id': None,
                'adresse': '',
                'ville': '',
                'telephone': ''
            })
            
            # Vérifier si l'article existe déjà
            article_existe = False
            for ligne in commande_temp['lignes']:
                if str(ligne['article_id']) == str(article_id):
                    ligne['quantite'] = float(Decimal(str(ligne['quantite'])) + quantite)
                    ligne['prix_total'] = float(Decimal(str(ligne['quantite'])) * Decimal(str(ligne['prix_unitaire'])))
                    article_existe = True
                    break
            
            if not article_existe:
                prix_unitaire = article.prix_vente or Decimal('0')
                prix_total = quantite * prix_unitaire
                
                nouvelle_ligne = {
                    'article_id': int(article_id),
                    'reference': article.reference_article or '',
                    'designation': article.designation,
                    'quantite': float(quantite),
                    'prix_unitaire': float(prix_unitaire),
                    'prix_total': float(prix_total),
                }
                commande_temp['lignes'].append(nouvelle_ligne)
            
            # Forcer la sauvegarde de la session
            request.session['commande_temporaire'] = commande_temp
            request.session.modified = True
            
            print(f"[AJOUTER ARTICLE] Session mise à jour: {len(commande_temp['lignes'])} lignes")
            print(f"[AJOUTER ARTICLE] Dernière ligne: {commande_temp['lignes'][-1] if commande_temp['lignes'] else 'Aucune'}")
            
            return JsonResponse({
                'success': True,
                'message': f'Article "{article.designation}" ajouté avec succès!',
                'article': {
                    'id': article_id,
                    'designation': article.designation,
                    'reference': article.reference_article or '',
                    'prix_vente': float(article.prix_vente or 0)
                },
                'lignes_count': len(commande_temp['lignes'])
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

@csrf_exempt
@login_required
def mettre_a_jour_quantites_commande(request):
    """Vue pour mettre à jour les quantités dans la commande temporaire"""
    if request.method == 'POST':
        agence = get_user_agence(request)
        if not agence:
            return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
        
        try:
            article_id = request.POST.get('article_id')
            quantite = Decimal(str(request.POST.get('quantite', 1)))
            
            commande_temp = request.session.get('commande_temporaire', {
                'lignes': [],
                'client_id': None,
                'adresse': '',
                'ville': '',
                'telephone': ''
            })
            
            # Trouver et mettre à jour la ligne correspondante
            for ligne in commande_temp['lignes']:
                if str(ligne['article_id']) == str(article_id):
                    ligne['quantite'] = float(quantite)
                    ligne['prix_total'] = float(quantite * Decimal(str(ligne['prix_unitaire'])))
                    break
            
            # Forcer la sauvegarde de la session
            request.session['commande_temporaire'] = commande_temp
            request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'message': 'Quantité mise à jour avec succès!'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

@csrf_exempt
@login_required
def supprimer_article_commande(request):
    """Vue pour supprimer un article de la commande temporaire"""
    if request.method == 'POST':
        agence = get_user_agence(request)
        if not agence:
            return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
        
        try:
            article_id = request.POST.get('article_id')
            
            commande_temp = request.session.get('commande_temporaire', {
                'lignes': [],
                'client_id': None,
                'adresse': '',
                'ville': '',
                'telephone': ''
            })
            
            # Supprimer la ligne correspondante
            commande_temp['lignes'] = [
                ligne for ligne in commande_temp['lignes']
                if str(ligne['article_id']) != str(article_id)
            ]
            
            # Forcer la sauvegarde de la session
            request.session['commande_temporaire'] = commande_temp
            request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'message': 'Article supprimé avec succès!',
                'lignes_count': len(commande_temp['lignes'])
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

@login_required
def sauvegarder_commande(request):
    """Vue pour sauvegarder la commande dans la base de données"""
    if request.method == 'POST':
        agence = get_user_agence(request)
        if not agence:
            messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
            return redirect('enregistrer_commande')
        
        try:
            commande_temp = request.session.get('commande_temporaire', {})
            lignes = commande_temp.get('lignes', [])
            
            if not lignes:
                messages.error(request, 'Veuillez ajouter au moins un article à la commande.')
                return redirect('enregistrer_commande')
            
            client_id = request.POST.get('client_id')
            if not client_id:
                messages.error(request, 'Veuillez sélectionner un client.')
                return redirect('enregistrer_commande')
            
            client = Client.objects.get(id=client_id, agence=agence)
            date_commande = request.POST.get('date')
            heure_commande = request.POST.get('heure')
            etat_commande = request.POST.get('etat_commande', 'en_attente')
            
            # Calculer les totaux globaux
            quantite_totale = sum(Decimal(str(ligne['quantite'])) for ligne in lignes)
            prix_total_global = sum(Decimal(str(ligne['prix_total'])) for ligne in lignes)
            
            # Créer les commandes (une par article) - Utiliser uniquement le modèle Commande
            commandes_creees = []
            
            for ligne in lignes:
                article = Article.objects.get(id=ligne['article_id'], agence=agence)
                quantite = Decimal(str(ligne['quantite']))
                prix_unitaire = Decimal(str(ligne['prix_unitaire']))
                prix_total_ligne = Decimal(str(ligne['prix_total']))
                
                # Créer une Commande pour chaque article
                commande = Commande.objects.create(
                    date=date_commande,
                    heure=heure_commande,
                    client=client,
                    article=article,
                    agence=agence,
                    quantite=quantite,
                    quantite_totale=quantite_totale,
                    prix_total=prix_total_ligne,
                    etat_commande=etat_commande,
                    created_by=request.user  # Enregistrer qui a créé la commande
                )
                commandes_creees.append(commande)
            
            # Essayer de créer DocumentCommande si le modèle existe (optionnel)
            try:
                import time
                numero_commande = f"CMD-{int(time.time())}"
                document_commande = DocumentCommande.objects.create(
                    numero_commande=numero_commande,
                    date=date_commande,
                    heure=heure_commande,
                    client=client,
                    agence=agence,
                    quantite_totale=quantite_totale,
                    prix_total_global=prix_total_global,
                    etat_commande=etat_commande
                )
                
                # Créer les lignes de commande si DocumentCommande existe
                for ligne, commande in zip(lignes, commandes_creees):
                    try:
                        LigneCommande.objects.create(
                            document_commande=document_commande,
                            article=commande.article,
                            quantite=commande.quantite,
                            prix_unitaire=commande.article.prix_vente,
                            prix_total=commande.prix_total
                        )
                    except Exception:
                        pass
            except Exception:
                # Si DocumentCommande n'existe pas, on continue avec Commande uniquement
                pass
            
            # Générer automatiquement une facture pour la commande - UNIQUEMENT pour les ACM
            compte = get_user_compte(request)
            facture_generee = False
            if commandes_creees and compte and compte.type_compte == 'acm':
                # Générer un numéro de facture unique
                from datetime import datetime, date as date_class
                # Convertir date_commande en objet date si c'est une chaîne
                if isinstance(date_commande, str):
                    date_obj = datetime.strptime(date_commande, '%Y-%m-%d').date()
                elif isinstance(date_commande, datetime):
                    date_obj = date_commande.date()
                elif isinstance(date_commande, date_class):
                    date_obj = date_commande
                else:
                    date_obj = timezone.now().date()
                
                date_str = date_obj.strftime('%Y%m%d')
                numero_facture = f"FAC{date_str}{len(commandes_creees):03d}"
                
                # Vérifier l'unicité du numéro
                compteur = 1
                while FactureCommande.objects.filter(numero_facture=numero_facture).exists():
                    numero_facture = f"FAC{date_str}{len(commandes_creees):03d}-{compteur}"
                    compteur += 1
                
                # Créer une facture pour la première commande
                # La base de données utilise 'date' et 'heure' (pas 'date_facture') pour l'instant
                # Récupérer l'heure de la commande ou utiliser l'heure actuelle
                if heure_commande:
                    if isinstance(heure_commande, str):
                        from datetime import datetime
                        try:
                            heure_obj = datetime.strptime(heure_commande, '%H:%M').time()
                        except ValueError:
                            try:
                                heure_obj = datetime.strptime(heure_commande, '%H:%M:%S').time()
                            except ValueError:
                                heure_obj = timezone.now().time()
                    elif hasattr(heure_commande, 'time'):
                        from datetime import datetime
                        heure_obj = heure_commande.time() if isinstance(heure_commande, datetime) else heure_commande
                    else:
                        heure_obj = heure_commande
                else:
                    heure_obj = timezone.now().time()
                
                try:
                    print(f"[DEBUG] AVANT création facture: agence.id_agence={agence.id_agence}, agence.nom_agence={agence.nom_agence}")
                    print(f"[DEBUG] Commande associée: commande.id={commandes_creees[0].id}, commande.agence.id_agence={commandes_creees[0].agence.id_agence}")
                    
                    # Calculer le prix total de toutes les commandes du groupe
                    prix_total_facture = sum(Decimal(str(c.prix_total)) for c in commandes_creees)
                    
                    facture_commande = FactureCommande.objects.create(
                        commande=commandes_creees[0],
                        numero_facture=numero_facture,
                        date=date_obj,  # Champ pour compatibilité
                        date_facture=date_obj,  # Champ principal dans la DB
                        heure=heure_obj,  # Ajouter l'heure
                        agence=agence,
                        prix_total=prix_total_facture,
                        net_a_payer=prix_total_facture  # Par défaut, net à payer = prix total
                    )
                    
                    # Vérifier après création
                    facture_commande.refresh_from_db()
                    facture_generee = True
                    print(f"[DEBUG] Facture créée avec succès: ID={facture_commande.id}, Numéro={numero_facture}")
                    print(f"[DEBUG] Facture agence vérification: facture.agence.id_agence={facture_commande.agence.id_agence}, agence.id_agence={agence.id_agence}")
                    print(f"[DEBUG] Facture agence nom: facture.agence.nom_agence={facture_commande.agence.nom_agence}")
                except Exception as e:
                    print(f"[DEBUG] ERREUR lors de la création de la facture: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    raise
            
            # Créer une notification pour la commande enregistrée
            try:
                Notification.objects.create(
                    type_notification='commande_enregistree',
                    titre='Nouvelle commande enregistrée',
                    message=f'Une nouvelle commande a été enregistrée pour le client {client.intitule}. Montant total: {prix_total_global} FCFA',
                    agence=agence,
                    commande=commandes_creees[0] if commandes_creees else None
                )
            except Exception as e:
                # Si le modèle Notification n'existe pas encore, on continue sans notification
                pass
            
            # Vider la session
            request.session['commande_temporaire'] = {
                'lignes': [],
                'client_id': None,
                'adresse': '',
                'ville': '',
                'telephone': ''
            }
            
            if facture_generee:
                messages.success(request, 'Commande enregistrée avec succès et facture générée automatiquement!')
            else:
                messages.success(request, 'Commande enregistrée avec succès!')
            # Rediriger vers la page de suivi client
            return redirect('suivi_client')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'enregistrement: {str(e)}')
            return redirect('enregistrer_commande')
    
    return redirect('enregistrer_commande')

@login_required
def search_window_commande(request):
    """Fenêtre de recherche d'articles pour les commandes"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        context = {
            'articles': [],
            'search_term': '',
            'agence': None
        }
        return render(request, 'supermarket/commandes/search_window.html', context)
    
    search_term = request.GET.get('q', '').strip()
    
    # Récupérer les articles de l'agence (tous les articles pour le chargement initial)
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    print(f"[SEARCH WINDOW COMMANDE] Agence: {agence.nom_agence} (ID: {agence.id_agence}), Articles trouvés: {articles.count()}")
    
    # Afficher quelques exemples d'articles pour déboguer
    if articles.exists():
        print(f"[SEARCH WINDOW COMMANDE] Exemples d'articles:")
        for article in articles[:3]:
            print(f"  - {article.designation} (ID: {article.id}, Agence: {article.agence.nom_agence})")
    else:
        print(f"[SEARCH WINDOW COMMANDE] ⚠️ AUCUN ARTICLE TROUVÉ pour l'agence {agence.nom_agence}")
        # Vérifier s'il y a des articles dans d'autres agences
        total_articles = Article.objects.count()
        print(f"[SEARCH WINDOW COMMANDE] Total articles dans la base: {total_articles}")
    
    context = {
        'articles': articles,
        'search_term': search_term,
        'agence': agence
    }
    return render(request, 'supermarket/commandes/search_window.html', context)

@csrf_exempt
@login_required
def rechercher_articles_commande(request):
    """Vue pour rechercher des articles (AJAX)"""
    if request.method == 'GET':
        agence = get_user_agence(request)
        if not agence:
            return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
        
        search_term = request.GET.get('search_term', '').strip()
        
        # Si pas de terme de recherche, retourner tous les articles de l'agence
        if not search_term or len(search_term) < 1:
            articles = Article.objects.filter(agence=agence).order_by('designation')[:100]
        else:
            # Recherche dans les articles de l'agence
            articles = Article.objects.filter(
                agence=agence
            ).filter(
                Q(designation__icontains=search_term) | 
                Q(reference_article__icontains=search_term)
            ).order_by('designation')[:100]
        
        print(f"[RECHERCHE COMMANDE] Agence: {agence.nom_agence}, Terme: '{search_term}', Articles trouvés: {articles.count()}")
        
        articles_data = [
            {
                'id': article.id,
                'designation': article.designation,
                'reference': article.reference_article or '',
                'prix_vente': float(article.prix_vente or 0),
                'stock': float(article.stock_actuel or 0)
            }
            for article in articles
        ]
        
        return JsonResponse({'success': True, 'articles': articles_data})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

@login_required
@require_commandes_feature('facture_commande')
def generer_facture_commande(request):
    """Vue pour générer automatiquement la facture à partir de la commande"""
    if request.method == 'POST':
        agence = get_user_agence(request)
        if not agence:
            messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
            return redirect('enregistrer_commande')
        
        try:
            commande_temp = request.session.get('commande_temporaire', {})
            lignes = commande_temp.get('lignes', [])
            
            if not lignes:
                messages.error(request, 'Veuillez ajouter au moins un article à la commande.')
                return redirect('enregistrer_commande')
            
            client_id = request.POST.get('client_id')
            if not client_id:
                messages.error(request, 'Veuillez sélectionner un client.')
                return redirect('enregistrer_commande')
            
            client = Client.objects.get(id=client_id, agence=agence)
            date_commande = request.POST.get('date')
            heure_commande = request.POST.get('heure')
            # Par défaut, toutes les commandes sont en attente
            etat_commande = 'en_attente'
            
            # Calculer les totaux
            quantite_totale = sum(Decimal(str(ligne['quantite'])) for ligne in lignes)
            prix_total = sum(Decimal(str(ligne['prix_total'])) for ligne in lignes)
            net_a_payer = prix_total  # Pour l'instant, pas de remise
            
            # Créer une commande pour chaque article
            commandes_creees = []
            for ligne in lignes:
                article = Article.objects.get(id=ligne['article_id'], agence=agence)
                
                commande = Commande.objects.create(
                    date=date_commande,
                    heure=heure_commande,
                    client=client,
                    article=article,
                    agence=agence,
                    quantite=Decimal(str(ligne['quantite'])),
                    quantite_totale=quantite_totale,
                    prix_total=prix_total,
                    etat_commande=etat_commande
                )
                commandes_creees.append(commande)
            
            # Créer une facture pour la première commande (ou toutes si nécessaire)
            if commandes_creees:
                # Générer un numéro de facture unique
                from datetime import datetime, date as date_class
                # Convertir date_commande en objet date si c'est une chaîne
                if isinstance(date_commande, str):
                    date_obj = datetime.strptime(date_commande, '%Y-%m-%d').date()
                elif isinstance(date_commande, datetime):
                    date_obj = date_commande.date()
                elif isinstance(date_commande, date_class):
                    date_obj = date_commande
                else:
                    date_obj = timezone.now().date()
                
                date_str = date_obj.strftime('%Y%m%d')
                numero_facture = f"FAC{date_str}{len(commandes_creees):03d}"
                
                # Vérifier l'unicité du numéro
                compteur = 1
                while FactureCommande.objects.filter(numero_facture=numero_facture).exists():
                    numero_facture = f"FAC{date_str}{len(commandes_creees):03d}-{compteur}"
                    compteur += 1
                
                # Créer une facture pour la première commande
                # Récupérer l'heure de la commande ou utiliser l'heure actuelle
                if heure_commande:
                    if isinstance(heure_commande, str):
                        from datetime import datetime
                        try:
                            heure_obj = datetime.strptime(heure_commande, '%H:%M').time()
                        except ValueError:
                            try:
                                heure_obj = datetime.strptime(heure_commande, '%H:%M:%S').time()
                            except ValueError:
                                heure_obj = timezone.now().time()
                    elif hasattr(heure_commande, 'time'):
                        heure_obj = heure_commande.time() if isinstance(heure_commande, datetime) else heure_commande
                    else:
                        heure_obj = heure_commande
                else:
                    heure_obj = timezone.now().time()
                
                # Calculer le prix total de toutes les commandes du groupe
                prix_total_facture = sum(Decimal(str(c.prix_total)) for c in commandes_creees)
                
                facture_commande, created = FactureCommande.objects.get_or_create(
                    commande=commandes_creees[0],
                    defaults={
                        'numero_facture': numero_facture,
                        'date': date_obj,  # Champ pour compatibilité
                        'date_facture': date_obj,  # Champ principal dans la DB
                        'heure': heure_obj,  # Ajouter l'heure
                        'agence': agence,
                        'prix_total': prix_total_facture,
                        'net_a_payer': prix_total_facture  # Par défaut, net à payer = prix total
                    }
                )
            
            # Préparer les données pour l'impression
            lignes_facture = []
            for ligne in lignes:
                article = Article.objects.get(id=ligne['article_id'], agence=agence)
                lignes_facture.append({
                    'designation': article.designation,
                    'quantite': float(ligne['quantite']),
                    'prix_unitaire': float(ligne['prix_unitaire']),
                    'prix_total': float(ligne['prix_total']),
                })
            
            # Convertir date et heure en chaînes pour la sérialisation JSON
            date_str = date_obj.strftime('%Y-%m-%d') if date_obj else ''
            date_display = date_obj.strftime('%d/%m/%Y') if date_obj else '--/--/----'
            heure_str = heure_obj.strftime('%H:%M:%S') if heure_obj else ''
            heure_display = heure_obj.strftime('%H:%M') if heure_obj else '--:--'
            
            facture_data = {
                'numero_facture': numero_facture,
                'date': date_str,  # Chaîne au format YYYY-MM-DD pour compatibilité
                'date_display': date_display,  # Chaîne formatée pour affichage DD/MM/YYYY
                'date_facture': date_str,  # Chaîne au format YYYY-MM-DD pour compatibilité
                'heure': heure_str,  # Chaîne au format HH:MM:SS pour compatibilité
                'heure_display': heure_display,  # Chaîne formatée pour affichage HH:MM
                'prix_total': float(prix_total_facture),
                'net_a_payer': float(prix_total_facture),
                'client_nom': client.intitule,
                'client_adresse': client.adresse if hasattr(client, 'adresse') else '',
                'client_telephone': client.telephone if hasattr(client, 'telephone') else '',
                'lignes': lignes_facture,
            }
            
            # Vider la session
            request.session['commande_temporaire'] = {
                'lignes': [],
                'client_id': None,
                'adresse': '',
                'ville': '',
                'telephone': ''
            }
            
            # Rediriger vers la page d'impression
            request.session['facture_commande_impression'] = facture_data
            return redirect('imprimer_facture_commande_session')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la génération: {str(e)}')
            return redirect('enregistrer_commande')
    
    return redirect('enregistrer_commande')

@login_required
@require_commandes_feature('consulter_commande')
def consulter_commandes(request):
    """Vue pour consulter la liste des commandes (documents de commande)"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    compte = get_user_compte(request)
    if not compte:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('login_commandes')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    etat_filter = request.GET.get('etat', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Utiliser Commande comme source principale et regrouper par (client, date, heure)
    from django.db.models import Q, Sum, Count
    from collections import defaultdict
    from decimal import Decimal
    
    # Récupérer les commandes de l'agence et filtrer selon l'utilisateur
    commandes = Commande.objects.filter(agence=agence).select_related('client', 'article', 'agence').order_by('-date', '-heure')
    commandes = filter_commandes_by_user(commandes, compte)
    
    # Appliquer les filtres
    if search_query:
        commandes = commandes.filter(
            Q(client__intitule__icontains=search_query) |
            Q(article__designation__icontains=search_query) |
            Q(article__reference_article__icontains=search_query)
        )
    
    if etat_filter:
        commandes = commandes.filter(etat_commande=etat_filter)
    
    if date_debut:
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            commandes = commandes.filter(date__gte=date_debut_obj)
        except ValueError:
            pass
    
    if date_fin:
        try:
            from datetime import datetime
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
            commandes = commandes.filter(date__lte=date_fin_obj)
        except ValueError:
            pass
    
    # Regrouper les commandes par (client, date, heure) pour créer des groupes de commandes
    commandes_list = list(commandes)
    
    def create_default_group():
        return {
        'commandes': [],
        'premiere_commande': None,
        'articles_uniques': set(),  # Set pour compter les articles uniques
        'quantite_totale': Decimal('0'),
        'prix_total': Decimal('0'),
        }
    
    groupes_dict = defaultdict(create_default_group)
    
    for cmd in commandes_list:
        heure_str = str(cmd.heure) if cmd.heure else ''
        cle_groupe = (cmd.client.id, str(cmd.date), heure_str)
        
        groupes_dict[cle_groupe]['commandes'].append(cmd)
        if groupes_dict[cle_groupe]['premiere_commande'] is None:
            groupes_dict[cle_groupe]['premiere_commande'] = cmd
        
        # Ajouter l'ID de l'article au set pour compter les articles uniques
        if cmd.article:
            groupes_dict[cle_groupe]['articles_uniques'].add(cmd.article.id)
        
        groupes_dict[cle_groupe]['quantite_totale'] += cmd.quantite
        groupes_dict[cle_groupe]['prix_total'] += cmd.prix_total
    
    # Créer la liste finale des groupes (sans doublons)
    commandes_list_final = []
    groupes_vus = set()  # Pour éviter les doublons
    
    for cle_groupe, groupe_data in groupes_dict.items():
        premiere_commande = groupe_data['premiere_commande']
        if premiere_commande:
            # Créer une clé unique pour éviter les doublons
            cle_unique = (premiere_commande.client.id, premiere_commande.date, premiere_commande.heure)
            
            if cle_unique not in groupes_vus:
                groupes_vus.add(cle_unique)
                commandes_list_final.append({
                    'id': premiere_commande.id,  # ID de la première commande pour les liens
                    'client': premiere_commande.client,
                    'date': premiere_commande.date,
                    'heure': premiere_commande.heure,
                    'etat_commande': premiere_commande.etat_commande,
                    'nombre_articles': len(groupe_data['articles_uniques']),  # Nombre d'articles uniques
                    'quantite_totale': groupe_data['quantite_totale'],
                    'prix_total': groupe_data['prix_total'],
                })
    
    # Trier par date et heure décroissantes
    commandes_list_final.sort(key=lambda x: (x['date'], x['heure']), reverse=True)
    
    # Calculer les statistiques sur les groupes de commandes
    total_commandes = len(commandes_list_final)
    commandes_en_attente = len([c for c in commandes_list_final if c['etat_commande'] == 'en_attente'])
    commandes_validees = len([c for c in commandes_list_final if c['etat_commande'] == 'validee'])
    commandes_en_livraison = len([c for c in commandes_list_final if c['etat_commande'] == 'en_livraison'])
    commandes_livrees = len([c for c in commandes_list_final if c['etat_commande'] == 'livree'])
    commandes_annulees = len([c for c in commandes_list_final if c['etat_commande'] == 'annulee'])
    
    # Calculer le montant total
    from decimal import Decimal
    montant_total = sum(Decimal(str(c['prix_total'])) for c in commandes_list_final)
    
    context = {
        'commandes': commandes_list_final,  # Liste des groupes de commandes
        'agence': agence,
        'total_commandes': total_commandes,
        'commandes_en_attente': commandes_en_attente,
        'commandes_validees': commandes_validees,
        'commandes_en_livraison': commandes_en_livraison,
        'commandes_livrees': commandes_livrees,
        'commandes_annulees': commandes_annulees,
        'montant_total': float(montant_total),
        'search_query': search_query,
        'etat_filter': etat_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'supermarket/commandes/consulter_commandes.html', context)

@login_required
@require_commandes_feature('consulter_commande')
def detail_commande(request, commande_id):
    """Vue pour afficher les détails d'une commande (avec tous les articles du groupe)"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    # Vérifier si on vient de consulter_livraisons (mode lecture seule)
    readonly = request.GET.get('readonly', 'false').lower() == 'true'
    from_livraison = request.GET.get('from_livraison', 'false').lower() == 'true'
    
    try:
        commande = Commande.objects.select_related('client', 'article').get(id=commande_id, client__agence=agence)
        
        # Récupérer toutes les commandes du même groupe (même client, date, heure)
        commandes_groupe = Commande.objects.filter(
            client=commande.client,
            date=commande.date,
            heure=commande.heure,
            client__agence=agence
        ).select_related('article').order_by('id')
        
        # Regrouper les commandes par article pour éviter les doublons
        from collections import defaultdict
        from decimal import Decimal
        commandes_grouped = defaultdict(lambda: {
            'article': None,
            'quantite': Decimal('0'),
            'prix_total': Decimal('0'),
            'prix_unitaire': Decimal('0')
        })
        
        for cmd in commandes_groupe:
            article_id = cmd.article.id
            if commandes_grouped[article_id]['article'] is None:
                commandes_grouped[article_id]['article'] = cmd.article
                commandes_grouped[article_id]['prix_unitaire'] = Decimal(str(cmd.article.prix_vente or 0))
            
            commandes_grouped[article_id]['quantite'] += cmd.quantite
            commandes_grouped[article_id]['prix_total'] += cmd.prix_total
        
        # Créer une classe simple pour permettre l'accès par point dans le template
        class CommandeGroupee:
            def __init__(self, article, quantite, prix_total, prix_unitaire):
                self.article = article
                self.quantite = quantite
                self.prix_total = prix_total
                self.prix_unitaire = prix_unitaire
        
        # Convertir en liste d'objets pour le template
        commandes_groupe_final = []
        for article_id, data in commandes_grouped.items():
            commandes_groupe_final.append(CommandeGroupee(
                article=data['article'],
                quantite=data['quantite'],
                prix_total=data['prix_total'],
                prix_unitaire=data['prix_unitaire']
            ))
        
        # Trier par ID d'article pour un affichage cohérent
        commandes_groupe_final.sort(key=lambda x: x.article.id)
        
        # Calculer les totaux du groupe
        quantite_totale = sum(cmd['quantite'] for cmd in commandes_groupe_final)
        prix_total = sum(cmd['prix_total'] for cmd in commandes_groupe_final)
        
        # Vérifier s'il y a une facture associée (prendre la première commande du groupe)
        facture = None
        try:
            facture = FactureCommande.objects.filter(commande__in=commandes_groupe).first()
        except:
            pass
        
        # Vérifier s'il y a une livraison associée (prendre la première commande du groupe)
        livraison = None
        try:
            livraison = Livraison.objects.filter(commande__in=commandes_groupe).first()
        except:
            pass
        
        context = {
            'commande': commande,  # Commande principale pour les infos générales
            'commandes_groupe': commandes_groupe,  # Tous les articles de la commande
            'facture': facture,
            'livraison': livraison,
            'agence': agence,
            'quantite_totale': totaux['quantite_totale'] or 0,
            'prix_total': totaux['prix_total'] or 0,
            'readonly': readonly or from_livraison,  # Mode lecture seule si depuis consulter_livraisons
        }
        return render(request, 'supermarket/commandes/detail_commande.html', context)
    except Commande.DoesNotExist:
        messages.error(request, 'Commande non trouvée.')
        return redirect('consulter_commandes')

@login_required
@require_commandes_feature('consulter_commande')
def supprimer_commande(request, commande_id):
    """Vue pour supprimer une commande (et toutes les commandes du même groupe)"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_commandes')
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        commande = Commande.objects.select_related('client').get(id=commande_id, client__agence=agence)
        
        # Récupérer toutes les commandes du même groupe (même client, date, heure)
        commandes_groupe = Commande.objects.filter(
            client=commande.client,
            date=commande.date,
            heure=commande.heure,
            client__agence=agence
        )
        
        # Informations pour le message
        client_nom = commande.client.intitule
        date_commande = commande.date
        nombre_articles = commandes_groupe.count()
        
        # Formater la date pour l'affichage
        date_formatee = date_commande.strftime('%d/%m/%Y') if date_commande else 'Date inconnue'
        
        # Supprimer toutes les commandes du groupe
        commandes_groupe.delete()
        
        messages.success(request, f'Commande du client "{client_nom}" du {date_formatee} ({nombre_articles} article{"s" if nombre_articles > 1 else ""}) supprimée avec succès!')
        
    except Commande.DoesNotExist:
        messages.error(request, 'Commande non trouvée.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_commandes')

@login_required
@require_commandes_feature('consulter_commande')
def modifier_commande(request, commande_id):
    """Vue pour modifier une commande existante"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        commande = Commande.objects.select_related('client', 'article').get(id=commande_id, client__agence=agence)
    except Commande.DoesNotExist:
        messages.error(request, 'Commande non trouvée.')
        return redirect('consulter_commandes')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            date_commande = request.POST.get('date')
            heure_commande = request.POST.get('heure')
            etat_commande = request.POST.get('etat_commande', 'en_attente')
            
            # Validation
            if not all([date_commande, heure_commande, etat_commande]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_commande', commande_id=commande_id)
            
            # Récupérer toutes les commandes du même groupe (même client, date, heure)
            commandes_groupe = Commande.objects.filter(
                client=commande.client,
                date=commande.date,
                heure=commande.heure,
                client__agence=agence
            )
            
            # Supprimer les commandes marquées comme supprimées
            commandes_supprimees = request.POST.getlist('commande_supprimee')
            for cmd_id in commandes_supprimees:
                try:
                    cmd_to_delete = Commande.objects.get(id=cmd_id, client__agence=agence)
                    cmd_to_delete.delete()
                except Commande.DoesNotExist:
                    pass
            
            # Mettre à jour les quantités et prix pour chaque ligne existante (non supprimée)
            quantite_totale = Decimal('0')
            prix_total_global = Decimal('0')
            
            # Récupérer les commandes restantes après suppression
            commandes_restantes = Commande.objects.filter(
                client=commande.client,
                date=commande.date,
                heure=commande.heure,
                client__agence=agence
            ).exclude(id__in=commandes_supprimees)
            
            for cmd in commandes_restantes:
                # Récupérer la quantité depuis le formulaire
                quantite_key = f'quantite_{cmd.id}'
                quantite_str = request.POST.get(quantite_key)
                
                if quantite_str:
                    quantite_decimal = Decimal(str(quantite_str))
                    if quantite_decimal > 0:
                        # Recalculer le prix total pour cette ligne
                        prix_unitaire = cmd.article.prix_vente or Decimal('0')
                        prix_total_ligne = quantite_decimal * prix_unitaire
                        
                        # Mettre à jour la commande
                        cmd.quantite = quantite_decimal
                        cmd.prix_total = prix_total_ligne
                        cmd.date = date_commande
                        cmd.heure = heure_commande
                        cmd.etat_commande = etat_commande
                        cmd.save()
                        
                        quantite_totale += quantite_decimal
                        prix_total_global += prix_total_ligne
                    else:
                        # Si quantité = 0, supprimer la commande
                        cmd.delete()
            
            # Après avoir traité les commandes existantes, mettre à jour la date/heure/état pour toutes les commandes du groupe final
            commandes_finales = Commande.objects.filter(
                client=commande.client,
                agence=agence
            ).filter(
                Q(date=date_commande, heure=heure_commande) | 
                Q(date=commande.date, heure=commande.heure)
            )
            commandes_finales.update(
                date=date_commande,
                heure=heure_commande,
                etat_commande=etat_commande
            )
            
            # Ajouter les nouveaux articles (via JSON comme dans modifier_facture_achat)
            articles_data = request.POST.get('articles_data')
            if articles_data:
                import json
                try:
                    articles = json.loads(articles_data)
                    for a in articles:
                        # a: {id, quantite, prix_vente}
                        try:
                            article = Article.objects.get(id=a['id'], agence=agence)
                            quantite_decimal = Decimal(str(a['quantite']))
                            
                            if quantite_decimal > 0:
                                prix_unitaire = Decimal(str(a.get('prix_vente', article.prix_vente or 0)))
                                prix_total_ligne = quantite_decimal * prix_unitaire
                                
                                # Créer une nouvelle commande pour cet article
                                nouvelle_commande = Commande.objects.create(
                                    date=date_commande,
                                    heure=heure_commande,
                                    client=commande.client,
                                    article=article,
                                    agence=agence,
                                    quantite=quantite_decimal,
                                    quantite_totale=Decimal('0'),  # Sera mis à jour après
                                    prix_total=prix_total_ligne,
                                    etat_commande=etat_commande
                                )
                                
                                quantite_totale += quantite_decimal
                                prix_total_global += prix_total_ligne
                        except (Article.DoesNotExist, KeyError, ValueError) as e:
                            pass
                except (json.JSONDecodeError, Exception) as e:
                            pass
            
            # Mettre à jour quantite_totale pour toutes les commandes du groupe
            commandes_groupe = Commande.objects.filter(
                client=commande.client,
                date=date_commande,
                heure=heure_commande,
                client__agence=agence
            )
            commandes_groupe.update(quantite_totale=quantite_totale)
            
            messages.success(request, 'Commande modifiée avec succès!')
            return redirect('detail_commande', commande_id=commande_id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
            return redirect('modifier_commande', commande_id=commande_id)
    
    # Récupérer toutes les commandes du même groupe (même client, date, heure)
    commandes_groupe = Commande.objects.filter(
        client=commande.client,
        date=commande.date,
        heure=commande.heure,
        client__agence=agence
    ).select_related('article').order_by('id')
    
    # Calculer les totaux
    from django.db.models import Sum
    totaux = commandes_groupe.aggregate(
        quantite_totale=Sum('quantite'),
        prix_total=Sum('prix_total')
    )
    quantite_totale = totaux['quantite_totale'] or Decimal('0')
    prix_total = totaux['prix_total'] or Decimal('0')
    
    # Afficher le formulaire de modification
    clients = Client.objects.filter(agence=agence).order_by('intitule')
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    context = {
        'commande': commande,
        'commandes_groupe': commandes_groupe,
        'clients': clients,
        'articles': articles,
        'quantite_totale': quantite_totale,
        'prix_total': prix_total,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/modifier_commande.html', context)

@login_required
@require_commandes_feature('definir_etat_livraison')
def definir_etat_livraison(request):
    """Vue pour définir l'état d'une livraison - affiche uniquement les livraisons confirmées avec tous les détails"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    compte = get_user_compte(request)
    if not compte:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('login_commandes')
    
    # Récupérer uniquement les livraisons confirmées (qui attendent la définition de l'état final)
    livraisons = Livraison.objects.filter(
        agence=agence,
        etat_livraison='confirmee'
    ).select_related('commande', 'commande__client', 'livreur').order_by('-date_livraison', '-heure_livraison')
    
    # Filtrer selon l'utilisateur (livreur voit seulement ses livraisons)
    livraisons = filter_livraisons_by_user(livraisons, compte)
    
    # Récupérer les commandes groupées et les détails pour chaque livraison
    livraisons_avec_details = []
    for livraison in livraisons:
        if livraison.commande:
            # Récupérer toutes les commandes du même groupe
            commandes_groupe = Commande.objects.filter(
                client=livraison.commande.client,
                date=livraison.commande.date,
                heure=livraison.commande.heure,
                agence=agence
            ).select_related('article').order_by('article__designation')
            
            # Calculer les totaux
            from django.db.models import Sum
            totaux = commandes_groupe.aggregate(
                total_quantite=Sum('quantite'),
                total_prix=Sum('prix_total')
            )
            
            livraisons_avec_details.append({
                'livraison': livraison,
                'commandes_groupe': commandes_groupe,
                'nombre_articles': commandes_groupe.count(),
                'quantite_totale': totaux['total_quantite'] or 0,
                'prix_total': totaux['total_prix'] or 0,
            })
    
    context = {
        'livraisons_avec_details': livraisons_avec_details,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/definir_etat_livraison.html', context)

@login_required
@require_commandes_feature('voir_itineraire')
def voir_itineraire(request):
    """Vue pour voir l'itinéraire de livraison avec les livraisons planifiées"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    compte = get_user_compte(request)
    if not compte:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('login_commandes')
    
    # Récupérer les livraisons planifiées (non confirmées et non livrées)
    livraisons = Livraison.objects.filter(
        agence=agence,
        etat_livraison__in=['planifiee', 'en_preparation', 'en_cours']
    ).select_related('commande', 'commande__client', 'livreur', 'agence').order_by('date_livraison', 'heure_depart', 'ordre_livraison')
    
    # Filtrer selon l'utilisateur (livreur voit seulement ses livraisons)
    livraisons = filter_livraisons_by_user(livraisons, compte)
    
    # Récupérer les commandes groupées pour chaque livraison
    livraisons_avec_details = []
    for livraison in livraisons:
        if livraison.commande:
            # Récupérer toutes les commandes du même groupe
            commandes_groupe = Commande.objects.filter(
                client=livraison.commande.client,
                date=livraison.commande.date,
                heure=livraison.commande.heure,
                agence=agence
            ).select_related('article')
            
            # Calculer les totaux
            from django.db.models import Sum
            totaux = commandes_groupe.aggregate(
                total_quantite=Sum('quantite'),
                total_prix=Sum('prix_total')
            )
            
            livraisons_avec_details.append({
                'livraison': livraison,
                'commandes_groupe': commandes_groupe,
                'nombre_articles': commandes_groupe.count(),
                'quantite_totale': totaux['total_quantite'] or 0,
                'prix_total': totaux['total_prix'] or 0,
            })
    
    context = {
        'agence': agence,
        'livraisons_avec_details': livraisons_avec_details,
    }
    return render(request, 'supermarket/commandes/voir_itineraire.html', context)

@login_required
def rapport_livraison(request):
    """Vue pour afficher le rapport de livraison - Redirige vers rapport_livreur"""
    return rapport_livreur(request)

@login_required
def consulter_factures_commande(request):
    """Vue pour consulter les factures de commande"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Debug: vérifier toutes les factures dans la base AVANT filtrage
    toutes_factures = FactureCommande.objects.all()
    print(f"[DEBUG] Total factures dans la DB: {toutes_factures.count()}")
    for f in toutes_factures[:5]:  # Afficher les 5 premières
        print(f"[DEBUG] Facture ID={f.id}, Numéro={f.numero_facture}, Agence={f.agence.id_agence if f.agence else 'None'}")
    
    # Récupérer toutes les factures de l'agence
    factures = FactureCommande.objects.filter(agence=agence).select_related('commande', 'commande__client', 'agence').order_by('-date')
    print(f"[DEBUG] Factures filtrées pour agence {agence.id_agence} ({agence.nom_agence}): {factures.count()}")
    
    # Appliquer les filtres
    if search_query:
        factures = factures.filter(
            Q(numero_facture__icontains=search_query) |
            Q(commande__client__intitule__icontains=search_query)
        )
    
    if date_debut:
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            factures = factures.filter(date__gte=date_debut_obj)
        except ValueError:
            pass
    
    if date_fin:
        try:
            from datetime import datetime
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
            factures = factures.filter(date__lte=date_fin_obj)
        except ValueError:
            pass
    
    # Calculer les statistiques
    from django.db.models import Sum
    total_factures = FactureCommande.objects.filter(agence=agence).count()
    
    # Calculer le montant total et préparer les données pour chaque facture
    factures_list = []
    montant_total = 0
    
    # Debug: vérifier combien de factures sont trouvées
    print(f"[DEBUG] consulter_factures_commande: {factures.count()} factures trouvées pour l'agence {agence.id_agence}")
    
    for facture in factures:
        montant_facture = 0
        if facture.commande:
            # Récupérer toutes les commandes du même groupe
            commandes_groupe = Commande.objects.filter(
                client=facture.commande.client,
                date=facture.commande.date,
                heure=facture.commande.heure,
                agence=agence
            )
            totaux_groupe = commandes_groupe.aggregate(total=Sum('prix_total'))
            montant_facture = totaux_groupe['total'] or 0
            montant_total += montant_facture
        else:
            # Si pas de commande associée, essayer de calculer depuis les commandes liées
            # ou utiliser 0
            print(f"[DEBUG] Facture {facture.id} n'a pas de commande associée")
        
        factures_list.append({
            'facture': facture,
            'montant': montant_facture
        })
    
    print(f"[DEBUG] consulter_factures_commande: {len(factures_list)} factures dans la liste finale")
    
    context = {
        'factures': factures_list,
        'agence': agence,
        'search_query': search_query,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'total_factures': total_factures,
        'montant_total': montant_total,
    }
    return render(request, 'supermarket/commandes/consulter_factures.html', context)

@login_required
def detail_facture_commande(request, facture_id):
    """Vue pour afficher les détails d'une facture de commande"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, "Votre compte n'est pas correctement lié à une agence.")
        return redirect('login_commandes')
    
    try:
        facture = FactureCommande.objects.select_related('commande', 'commande__client', 'agence').get(id=facture_id, agence=agence)
        
        # Récupérer toutes les commandes du même groupe (même client, date, heure)
        commandes_groupe = []
        quantite_totale = 0
        prix_total = 0
        
        if facture.commande:
            commandes_groupe = Commande.objects.filter(
                client=facture.commande.client,
                date=facture.commande.date,
                heure=facture.commande.heure,
                agence=agence
            ).select_related('article').order_by('id')
            
            # Calculer les totaux
            from django.db.models import Sum
            totaux = commandes_groupe.aggregate(
                quantite_totale=Sum('quantite'),
                prix_total=Sum('prix_total')
            )
            quantite_totale = totaux['quantite_totale'] or 0
            prix_total = totaux['prix_total'] or 0
        
    except FactureCommande.DoesNotExist:
        messages.error(request, 'Facture non trouvée.')
        return redirect('consulter_factures_commande')
    
    context = {
        'facture': facture,
        'commandes': commandes_groupe,
        'quantite_totale': quantite_totale,
        'prix_total': prix_total,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/detail_facture.html', context)

@login_required
def modifier_facture_commande(request, facture_id):
    """Vue pour modifier une facture de commande"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, "Votre compte n'est pas correctement lié à une agence.")
        return redirect('login_commandes')
    
    try:
        facture = FactureCommande.objects.select_related('commande', 'commande__client').get(id=facture_id, agence=agence)
        
        # Récupérer toutes les commandes du même groupe (même client, date, heure)
        commandes_groupe = []
        quantite_totale = 0
        prix_total = 0
        
        if facture.commande:
            commandes_groupe = Commande.objects.filter(
                client=facture.commande.client,
                date=facture.commande.date,
                heure=facture.commande.heure,
                agence=agence
            ).select_related('article').order_by('id')
            
            # Calculer les totaux
            from django.db.models import Sum
            totaux = commandes_groupe.aggregate(
                quantite_totale=Sum('quantite'),
                prix_total=Sum('prix_total')
            )
            quantite_totale = totaux['quantite_totale'] or 0
            prix_total = totaux['prix_total'] or 0
        
    except FactureCommande.DoesNotExist:
        messages.error(request, 'Facture non trouvée.')
        return redirect('consulter_factures_commande')
    
    if request.method == 'POST':
        try:
            numero_facture = request.POST.get('numero_facture')
            date_facture = request.POST.get('date_facture')
            
            if not all([numero_facture, date_facture]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_facture_commande', facture_id=facture_id)
            
            # Mettre à jour les informations générales de la facture
            facture.numero_facture = numero_facture
            facture.date = date_facture  # Champ pour compatibilité
            facture.date_facture = date_facture  # Champ principal dans la DB
            
            # Traiter les modifications des commandes (quantités et suppressions)
            commandes_supprimees = request.POST.getlist('commande_supprimee')
            for cmd_id in commandes_supprimees:
                try:
                    cmd_to_delete = Commande.objects.get(id=cmd_id, agence=agence)
                    cmd_to_delete.delete()
                except Commande.DoesNotExist:
                    pass
            
            # Mettre à jour les quantités des commandes restantes
            from decimal import Decimal
            for cmd in commandes_groupe:
                if str(cmd.id) not in commandes_supprimees:
                    quantite_nouvelle = request.POST.get(f'quantite_{cmd.id}')
                    if quantite_nouvelle:
                        try:
                            quantite_nouvelle = Decimal(quantite_nouvelle)
                            if quantite_nouvelle > 0:
                                # Recalculer le prix total
                                prix_unitaire = cmd.article.prix_vente if hasattr(cmd.article, 'prix_vente') else (cmd.prix_total / cmd.quantite if cmd.quantite > 0 else 0)
                                prix_total_nouveau = quantite_nouvelle * Decimal(str(prix_unitaire))
                                
                                cmd.quantite = quantite_nouvelle
                                cmd.prix_total = prix_total_nouveau
                                cmd.save()
                            else:
                                # Si quantité = 0, supprimer la commande
                                cmd.delete()
                        except (ValueError, TypeError):
                            pass
            
            # Recalculer les totaux de la facture
            commandes_restantes = Commande.objects.filter(
                client=facture.commande.client,
                date=facture.commande.date,
                heure=facture.commande.heure,
                agence=agence
            )
            
            from django.db.models import Sum
            totaux = commandes_restantes.aggregate(
                quantite_totale=Sum('quantite'),
                prix_total=Sum('prix_total')
            )
            
            facture.prix_total = totaux['prix_total'] or Decimal('0')
            facture.net_a_payer = facture.prix_total  # Mettre à jour le net à payer
            facture.save()
            
            # Recharger les commandes après modification
            commandes_groupe = Commande.objects.filter(
                client=facture.commande.client,
                date=facture.commande.date,
                heure=facture.commande.heure,
                agence=agence
            ).select_related('article').order_by('id')
            
            # Recalculer les totaux
            totaux = commandes_groupe.aggregate(
                quantite_totale=Sum('quantite'),
                prix_total=Sum('prix_total')
            )
            quantite_totale = totaux['quantite_totale'] or 0
            prix_total = totaux['prix_total'] or 0
            
            messages.success(request, 'Facture modifiée avec succès!')
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de la facture: {str(e)}')
    
    # Recharger les commandes pour l'affichage (au cas où elles auraient été modifiées)
    if facture.commande:
        commandes_groupe = Commande.objects.filter(
            client=facture.commande.client,
            date=facture.commande.date,
            heure=facture.commande.heure,
            agence=agence
        ).select_related('article').order_by('id')
        
        # Calculer les totaux
        from django.db.models import Sum
        totaux = commandes_groupe.aggregate(
            quantite_totale=Sum('quantite'),
            prix_total=Sum('prix_total')
        )
        quantite_totale = totaux['quantite_totale'] or 0
        prix_total = totaux['prix_total'] or 0
    
    context = {
        'facture': facture,
        'commandes': commandes_groupe,
        'quantite_totale': quantite_totale,
        'prix_total': prix_total,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/modifier_facture.html', context)

@login_required
@require_commandes_feature('consulter_commande')
def generer_facture_commande_existante(request, commande_id):
    """Vue pour générer une facture à partir d'une commande existante"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        commande = Commande.objects.get(id=commande_id, agence=agence)
        
        # Générer un numéro de facture unique
        from datetime import datetime
        date_obj = commande.date if isinstance(commande.date, datetime) else datetime.strptime(str(commande.date), '%Y-%m-%d').date()
        date_str = date_obj.strftime('%Y%m%d')
        numero_facture = f"FAC{date_str}{commande.id:03d}"
        
        # Vérifier l'unicité du numéro
        compteur = 1
        while FactureCommande.objects.filter(numero_facture=numero_facture).exists():
            numero_facture = f"FAC{date_str}{commande.id:03d}-{compteur}"
            compteur += 1
        
        # Créer la facture
        # Récupérer l'heure de la commande ou utiliser l'heure actuelle
        if commande.heure:
            heure_obj = commande.heure
        else:
            heure_obj = timezone.now().time()
        
        # Calculer le prix total pour cette commande
        prix_total_facture = Decimal(str(commande.prix_total))
        
        facture, created = FactureCommande.objects.get_or_create(
            commande=commande,
            defaults={
                'numero_facture': numero_facture,
                'date': date_obj,  # Champ pour compatibilité
                'date_facture': date_obj,  # Champ principal dans la DB
                'heure': heure_obj,  # Ajouter l'heure
                'agence': agence,
                'prix_total': prix_total_facture,
                'net_a_payer': prix_total_facture  # Par défaut, net à payer = prix total
            }
        )
        
        if created:
            messages.success(request, 'Facture générée avec succès!')
        else:
            messages.info(request, 'Facture déjà existante.')
        
        return redirect('detail_facture_commande', facture_id=facture.id)
    except Commande.DoesNotExist:
        messages.error(request, 'Commande non trouvée.')
        return redirect('consulter_commandes')

@login_required
def imprimer_facture_commande(request, facture_id=None):
    """Vue pour imprimer une facture de commande"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    if facture_id:
        try:
            facture_db = FactureCommande.objects.select_related('commande', 'commande__client', 'agence').get(id=facture_id, agence=agence)
            
            # Récupérer la première commande pour obtenir le client
            premiere_commande = facture_db.commande
            
            # Récupérer toutes les commandes du même groupe (utiliser la date/heure actuelles des commandes)
            # Après modification, les commandes peuvent avoir une nouvelle date/heure
            commandes_groupe = Commande.objects.filter(
                client=premiere_commande.client,
                date=premiere_commande.date,
                heure=premiere_commande.heure,
                agence=agence
            ).select_related('article').order_by('id')
            
            # Si aucune commande trouvée avec cette date/heure, essayer avec la date/heure de la facture
            if not commandes_groupe.exists():
                commandes_groupe = Commande.objects.filter(
                    client=premiere_commande.client,
                    agence=agence
                ).filter(
                    Q(date=premiere_commande.date, heure=premiere_commande.heure) |
                    Q(date=facture_db.date_facture if hasattr(facture_db, 'date_facture') else facture_db.date)
                ).select_related('article').order_by('id')
            
            # Préparer les lignes de facture avec les valeurs actuelles
            lignes_facture = []
            prix_total_calcule = Decimal('0')
            
            for cmd in commandes_groupe:
                prix_unitaire = cmd.article.prix_vente if hasattr(cmd.article, 'prix_vente') else (cmd.prix_total / cmd.quantite if cmd.quantite > 0 else 0)
                prix_total_ligne = cmd.prix_total  # Utiliser le prix_total de la commande (déjà calculé)
                
                lignes_facture.append({
                    'designation': cmd.article.designation,
                    'quantite': float(cmd.quantite),
                    'prix_unitaire': float(prix_unitaire),
                    'prix_total': float(prix_total_ligne),
                })
                prix_total_calcule += Decimal(str(prix_total_ligne))
            
            # Utiliser les valeurs de la facture mise à jour (date_facture, heure)
            date_facture_obj = facture_db.date_facture if hasattr(facture_db, 'date_facture') and facture_db.date_facture else facture_db.date
            date_str = date_facture_obj.strftime('%Y-%m-%d') if date_facture_obj else ''
            date_display = date_facture_obj.strftime('%d/%m/%Y') if date_facture_obj else '--/--/----'
            
            # Utiliser l'heure de la facture ou de la première commande
            heure_facture = facture_db.heure if facture_db.heure else premiere_commande.heure
            heure_str = heure_facture.strftime('%H:%M:%S') if heure_facture else ''
            heure_display = heure_facture.strftime('%H:%M') if heure_facture else '--:--'
            
            # Utiliser le prix_total calculé à partir des commandes actuelles ou celui de la facture
            prix_total_final = float(prix_total_calcule) if prix_total_calcule > 0 else float(facture_db.prix_total)
            net_a_payer_final = float(facture_db.net_a_payer) if facture_db.net_a_payer else prix_total_final
            
            facture_data = {
                'numero_facture': facture_db.numero_facture,
                'date': date_str,  # Chaîne pour compatibilité avec session
                'date_display': date_display,  # Chaîne formatée pour affichage
                'date_facture': date_str,  # Chaîne pour compatibilité avec session
                'heure': heure_str,  # Chaîne pour compatibilité avec session
                'heure_display': heure_display,  # Chaîne formatée pour affichage
                'prix_total': prix_total_final,  # Utiliser le prix calculé à partir des commandes actuelles
                'net_a_payer': net_a_payer_final,
                'client_nom': premiere_commande.client.intitule,
                'client_adresse': premiere_commande.client.adresse if hasattr(premiere_commande.client, 'adresse') else '',
                'client_telephone': premiere_commande.client.telephone if hasattr(premiere_commande.client, 'telephone') else '',
                'lignes': lignes_facture,
            }
            
        except FactureCommande.DoesNotExist:
            messages.error(request, 'Facture non trouvée.')
            return redirect('consulter_factures_commande')
    else:
        # Vérifier si on a une facture dans la session (depuis génération)
        facture_data = request.session.get('facture_commande_impression')
        if not facture_data:
            messages.error(request, 'Aucune facture spécifiée.')
            return redirect('consulter_factures_commande')
        # Supprimer de la session après utilisation
        del request.session['facture_commande_impression']
    
    context = {
        'facture': facture_data,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/facture_impression_commande.html', context)

@login_required
def imprimer_facture_commande_xprinter(request, facture_id):
    """Vue pour imprimer une facture de commande avec XPrinter (format 80mm)"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        facture_db = FactureCommande.objects.select_related('commande', 'commande__client', 'agence').get(id=facture_id, agence=agence)
        
        # Récupérer la première commande pour obtenir le client
        premiere_commande = facture_db.commande
        
        # Récupérer toutes les commandes du même groupe (utiliser la date/heure actuelles des commandes)
        # Après modification, les commandes peuvent avoir une nouvelle date/heure
        # On récupère toutes les commandes du même client qui ont la même date/heure que la première commande
        commandes_groupe = Commande.objects.filter(
            client=premiere_commande.client,
            date=premiere_commande.date,
            heure=premiere_commande.heure,
            agence=agence
        ).select_related('article').order_by('id')
        
        # Si aucune commande trouvée avec cette date/heure, essayer avec la date/heure de la facture
        if not commandes_groupe.exists():
            commandes_groupe = Commande.objects.filter(
                client=premiere_commande.client,
                agence=agence
            ).filter(
                Q(date=premiere_commande.date, heure=premiere_commande.heure) |
                Q(date=facture_db.date_facture if hasattr(facture_db, 'date_facture') else facture_db.date)
            ).select_related('article').order_by('id')
        
        # Préparer les lignes de facture avec les valeurs actuelles
        lignes_facture = []
        prix_total_calcule = Decimal('0')
        
        for cmd in commandes_groupe:
            prix_unitaire = cmd.article.prix_vente if hasattr(cmd.article, 'prix_vente') else (cmd.prix_total / cmd.quantite if cmd.quantite > 0 else 0)
            prix_total_ligne = cmd.prix_total  # Utiliser le prix_total de la commande (déjà calculé)
            
            lignes_facture.append({
                'designation': cmd.article.designation,
                'quantite': float(cmd.quantite),
                'prix_unitaire': float(prix_unitaire),
                'prix_total': float(prix_total_ligne),
            })
            prix_total_calcule += Decimal(str(prix_total_ligne))
        
        # Utiliser les valeurs de la facture mise à jour (date_facture, heure)
        date_facture_obj = facture_db.date_facture if hasattr(facture_db, 'date_facture') and facture_db.date_facture else facture_db.date
        date_str = date_facture_obj.strftime('%Y-%m-%d') if date_facture_obj else ''
        date_display = date_facture_obj.strftime('%d/%m/%Y') if date_facture_obj else '--/--/----'
        
        # Utiliser l'heure de la facture ou de la première commande
        heure_facture = facture_db.heure if facture_db.heure else premiere_commande.heure
        heure_str = heure_facture.strftime('%H:%M:%S') if heure_facture else ''
        heure_display = heure_facture.strftime('%H:%M') if heure_facture else '--:--'
        
        # Utiliser le prix_total calculé à partir des commandes actuelles ou celui de la facture
        prix_total_final = float(prix_total_calcule) if prix_total_calcule > 0 else float(facture_db.prix_total)
        net_a_payer_final = float(facture_db.net_a_payer) if facture_db.net_a_payer else prix_total_final
        
        facture_data = {
            'numero_facture': facture_db.numero_facture,
            'date': date_str,
            'date_display': date_display,
            'date_facture': date_str,
            'heure': heure_str,
            'heure_display': heure_display,
            'prix_total': prix_total_final,  # Utiliser le prix calculé à partir des commandes actuelles
            'net_a_payer': net_a_payer_final,
            'client_nom': premiere_commande.client.intitule,
            'client_adresse': premiere_commande.client.adresse if hasattr(premiere_commande.client, 'adresse') else '',
            'client_telephone': premiere_commande.client.telephone if hasattr(premiere_commande.client, 'telephone') else '',
            'lignes': lignes_facture,
        }
        
    except FactureCommande.DoesNotExist:
        messages.error(request, 'Facture non trouvée.')
        return redirect('consulter_factures_commande')
    
    context = {
        'facture': facture_data,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/facture_impression_commande_xprinter.html', context)

@login_required
def planification_livraison(request):
    """Vue pour afficher la planification des livraisons avec vérification du stock"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    # Récupérer les livraisons déjà planifiées
    livraisons = Livraison.objects.filter(agence=agence).select_related('commande', 'commande__client', 'agence').order_by('-date_livraison')
    
    # Récupérer les commandes enregistrées qui ne sont pas encore planifiées
    # (commandes en attente ou validées qui n'ont pas de livraison)
    from django.db.models import Q, Sum, Count
    from collections import defaultdict
    from decimal import Decimal
    
    # Récupérer les IDs des commandes qui ont déjà une livraison (convertir en set pour vérification rapide)
    commandes_avec_livraison_set = set(Livraison.objects.filter(
        agence=agence,
        commande__isnull=False
    ).values_list('commande_id', flat=True).distinct())
    
    # Récupérer TOUTES les commandes (planifiées et non planifiées) pour les afficher
    commandes_toutes = Commande.objects.filter(
        agence=agence,
        etat_commande__in=['en_attente', 'validee', 'en_livraison']
    ).select_related('client', 'article', 'agence').order_by('-date', '-heure')
    
    # Regrouper les commandes par (client, date, heure) comme dans consulter_commandes
    commandes_list = list(commandes_toutes)
    
    def create_default_group():
        return {
            'commandes': [],
            'premiere_commande': None,
            'nombre_articles': 0,
            'quantite_totale': Decimal('0'),
            'prix_total': Decimal('0'),
        }
    
    groupes_dict = defaultdict(create_default_group)
    
    for cmd in commandes_list:
        heure_str = str(cmd.heure) if cmd.heure else ''
        cle_groupe = (cmd.client.id, str(cmd.date), heure_str)
        
        groupes_dict[cle_groupe]['commandes'].append(cmd)
        if groupes_dict[cle_groupe]['premiere_commande'] is None:
            groupes_dict[cle_groupe]['premiere_commande'] = cmd
        
        groupes_dict[cle_groupe]['nombre_articles'] += 1
        groupes_dict[cle_groupe]['quantite_totale'] += cmd.quantite
        groupes_dict[cle_groupe]['prix_total'] += cmd.prix_total
    
    # Créer la liste finale des groupes avec vérification du stock
    commandes_a_planifier = []
    for cle_groupe, groupe_data in groupes_dict.items():
        premiere_commande = groupe_data['premiere_commande']
        if premiere_commande:
            # Vérifier le stock pour toutes les commandes du groupe
            stock_suffisant = True
            articles_insuffisants = []
            
            for cmd in groupe_data['commandes']:
                stock_disponible = cmd.article.stock_actuel or Decimal('0')
                quantite_requise = cmd.quantite
                
                if stock_disponible < quantite_requise:
                    stock_suffisant = False
                    articles_insuffisants.append({
                        'article': cmd.article.designation,
                        'reference': cmd.article.reference_article,
                        'stock_disponible': float(stock_disponible),
                        'quantite_commande': float(quantite_requise),
                        'manquant': float(quantite_requise - stock_disponible),
                    })
            
            # Vérifier si cette commande est déjà planifiée (vérifier si au moins une commande du groupe a une livraison)
            deja_planifiee = any(cmd.id in commandes_avec_livraison_set for cmd in groupe_data['commandes'])
            
            commandes_a_planifier.append({
                'id': premiere_commande.id,
                'client': premiere_commande.client,
                'client_intitule': premiere_commande.client.intitule,
                'date': premiere_commande.date,
                'heure': premiere_commande.heure,
                'etat_commande': premiere_commande.etat_commande,
                'nombre_articles': groupe_data['nombre_articles'],
                'quantite_totale': groupe_data['quantite_totale'],
                'prix_total': groupe_data['prix_total'],
                'stock_suffisant': stock_suffisant,
                'articles_insuffisants': articles_insuffisants,
                'commandes': groupe_data['commandes'],  # Pour afficher les détails
                'deja_planifiee': deja_planifiee,  # Indicateur si déjà planifiée
            })
    
    # Trier par date et heure décroissantes
    commandes_a_planifier.sort(key=lambda x: (x['date'], x['heure']), reverse=True)
    
    context = {
        'livraisons': livraisons,
        'commandes_a_planifier': commandes_a_planifier,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/planification_livraison.html', context)

@login_required
@require_commandes_feature('consulter_livraison')
def consulter_livraisons(request):
    """Vue pour consulter toutes les livraisons"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    compte = get_user_compte(request)
    if not compte:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('login_commandes')
    
    livraisons = Livraison.objects.filter(agence=agence).select_related('commande', 'commande__client', 'livreur', 'agence').order_by('-date_livraison', '-heure_depart')
    
    # Filtrer selon l'utilisateur (livreur voit seulement ses livraisons)
    livraisons = filter_livraisons_by_user(livraisons, compte)
    
    context = {
        'livraisons': livraisons,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/consulter_livraisons.html', context)

@login_required
def creer_planification_livraison(request):
    """Vue pour créer une nouvelle planification de livraison avec vérification du stock"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    if request.method == 'POST':
        try:
            commande_id = request.POST.get('commande_id')
            date_livraison = request.POST.get('date_livraison')
            heure_depart = request.POST.get('heure_depart')
            heure_arrivee = request.POST.get('heure_arrivee')
            livreur_id = request.POST.get('livreur_id')
            zone = request.POST.get('zone')
            notes = request.POST.get('notes', '')
            
            if not commande_id or not date_livraison or not heure_depart or not heure_arrivee or not livreur_id or not zone:
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_planification_livraison')
            
            premiere_commande = Commande.objects.get(id=commande_id, agence=agence)
            
            # Récupérer toutes les commandes du même groupe
            commandes_groupe = Commande.objects.filter(
                client=premiere_commande.client,
                date=premiere_commande.date,
                heure=premiere_commande.heure,
                agence=agence
            ).select_related('article')
            
            # Vérifier le stock pour toutes les commandes du groupe
            articles_insuffisants = []
            for cmd in commandes_groupe:
                stock_disponible = cmd.article.stock_actuel or Decimal('0')
                quantite_requise = cmd.quantite
                
                if stock_disponible < quantite_requise:
                    articles_insuffisants.append({
                        'article': cmd.article.designation,
                        'reference': cmd.article.reference_article,
                        'stock_disponible': float(stock_disponible),
                        'quantite_commande': float(quantite_requise),
                        'manquant': float(quantite_requise - stock_disponible),
                    })
            
            # Si stock insuffisant, ne pas créer la livraison
            if articles_insuffisants:
                message_erreur = 'Stock insuffisant pour planifier cette livraison. Articles en rupture : '
                details = []
                for art in articles_insuffisants:
                    details.append(f"{art['article']} (Stock: {art['stock_disponible']:.2f}, Commandé: {art['quantite_commande']:.2f}, Manquant: {art['manquant']:.2f})")
                messages.error(request, message_erreur + ' | '.join(details))
                return redirect('planification_livraison')
            
            # Récupérer le livreur
            try:
                livreur = Livreur.objects.get(id=livreur_id, agence=agence, actif=True)
            except Livreur.DoesNotExist:
                messages.error(request, 'Livreur non trouvé ou inactif.')
                return redirect('planification_livraison')
            
            # Convertir les dates et heures
            from datetime import datetime
            date_livraison_obj = datetime.strptime(date_livraison, '%Y-%m-%d').date()
            heure_depart_obj = datetime.strptime(heure_depart, '%H:%M').time()
            heure_arrivee_obj = datetime.strptime(heure_arrivee, '%H:%M').time()
            
            # Créer la livraison pour la première commande du groupe
            livraison = Livraison.objects.create(
                commande=premiere_commande,
                date_livraison=date_livraison_obj,
                heure_depart=heure_depart_obj,
                heure_arrivee=heure_arrivee_obj,
                livreur=livreur,
                zone=zone,
                notes=notes if notes else None,
                agence=agence,
                etat_livraison='planifiee'
            )
            
            # Mettre à jour l'état des commandes du groupe
            commandes_groupe.update(etat_commande='en_livraison')
            
            # Créer une notification pour la livraison planifiée
            try:
                Notification.objects.create(
                    type_notification='livraison_planifiee',
                    titre='Livraison planifiée',
                    message=f'Une livraison a été planifiée pour le client {premiere_commande.client.intitule} le {date_livraison_obj.strftime("%d/%m/%Y")} à {heure_depart_obj.strftime("%H:%M")}. Livreur: {livreur.nom} {livreur.prenom}',
                    agence=agence,
                    livraison=livraison,
                    commande=premiere_commande
                )
            except Exception as e:
                pass  # Continuer même si la notification échoue
            
            messages.success(request, 'Livraison planifiée avec succès!')
            return redirect('planification_livraison')
        except Commande.DoesNotExist:
            messages.error(request, 'Commande non trouvée.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {str(e)}')
    
    # Récupérer la commande depuis le paramètre GET si présent
    commande_id = request.GET.get('commande_id')
    commande = None
    commandes_groupe = None
    if commande_id:
        try:
            commande = Commande.objects.get(id=commande_id, agence=agence)
            # Récupérer toutes les commandes du même groupe
            commandes_groupe = Commande.objects.filter(
                client=commande.client,
                date=commande.date,
                heure=commande.heure,
                agence=agence
            ).select_related('article', 'client').order_by('article__designation')
        except Commande.DoesNotExist:
            pass
    
    # Récupérer les livreurs actifs de l'agence
    livreurs = Livreur.objects.filter(agence=agence, actif=True).order_by('nom', 'prenom')
    
    # Récupérer les IDs des commandes qui ont déjà une livraison
    commandes_avec_livraison = Livraison.objects.filter(
        agence=agence,
        commande__isnull=False
    ).values_list('commande_id', flat=True).distinct()
    
    # Récupérer les commandes disponibles pour la planification (toutes les commandes)
    commandes = Commande.objects.filter(
        agence=agence,
        etat_commande__in=['en_attente', 'validee']
    ).exclude(
        id__in=commandes_avec_livraison
    ).select_related('client', 'article', 'agence').order_by('-date', '-heure')
    
    context = {
        'commandes': commandes,
        'commande_selectionnee': commande,
        'commandes_groupe': commandes_groupe,
        'livreurs': livreurs,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/creer_planification_livraison.html', context)

@login_required
@require_commandes_feature('gestion_client')
def enregistrer_client(request):
    """Vue pour enregistrer un nouveau client dans le module commandes"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    if request.method == 'POST':
        try:
            intitule = request.POST.get('intitule', '').strip()
            adresse = request.POST.get('adresse', '').strip()
            telephone = request.POST.get('telephone', '').strip()
            email = request.POST.get('email', '').strip()
            zone = request.POST.get('zone', '').strip()
            
            # Validation de tous les champs obligatoires
            if not intitule:
                messages.error(request, 'Le nom du client est obligatoire.')
                return redirect('enregistrer_client_commandes')
            if not adresse:
                messages.error(request, 'L\'adresse est obligatoire.')
                return redirect('enregistrer_client_commandes')
            if not telephone:
                messages.error(request, 'Le téléphone est obligatoire.')
                return redirect('enregistrer_client_commandes')
            if not email:
                messages.error(request, 'L\'e-mail est obligatoire.')
                return redirect('enregistrer_client_commandes')
            if not zone:
                messages.error(request, 'La zone est obligatoire.')
                return redirect('enregistrer_client_commandes')
            
            client = Client.objects.create(
                intitule=intitule,
                adresse=adresse,
                telephone=telephone,
                email=email,
                zone=zone,
                agence=agence
            )
            
            messages.success(request, 'Client enregistré avec succès!')
            return redirect('consulter_clients_commandes')
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'enregistrement: {str(e)}')
    
    context = {
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/enregistrer_client.html', context)

@login_required
@require_commandes_feature('gestion_client')
def consulter_clients_commandes(request):
    """Vue pour consulter les clients dans le module commandes"""
    from django.core.paginator import Paginator
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    compte = get_user_compte(request)
    if not compte:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('login_commandes')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    
    # Construire la requête de base - Tous les clients de l'agence (plus de filtre ACM restrictif)
    clients = Client.objects.filter(agence=agence).order_by('intitule')
    
    # Appliquer le filtre de recherche
    if search_query:
        clients = clients.filter(
            Q(intitule__icontains=search_query) |
            Q(telephone__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(adresse__icontains=search_query)
        )
    
    # Pagination - 20 clients par page
    paginator = Paginator(clients, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'clients': page_obj,
        'agence': agence,
        'search_query': search_query,
        'page_obj': page_obj,
    }
    return render(request, 'supermarket/commandes/consulter_clients.html', context)

@login_required
def detail_client_commandes(request, client_id):
    """Vue pour afficher les détails d'un client dans le module commandes"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        client = Client.objects.get(id=client_id, agence=agence)
        commandes = Commande.objects.filter(client=client).select_related('article').order_by('-date', '-heure')
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients_commandes')
    
    context = {
        'client': client,
        'commandes': commandes,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/detail_client.html', context)

@login_required
def modifier_client_commandes(request, client_id):
    """Vue pour modifier un client dans le module commandes"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        client = Client.objects.get(id=client_id, agence=agence)
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients_commandes')
    
    if request.method == 'POST':
        try:
            client.intitule = request.POST.get('intitule', client.intitule)
            client.adresse = request.POST.get('adresse', client.adresse)
            client.telephone = request.POST.get('telephone', client.telephone)
            client.save()
            
            messages.success(request, 'Client modifié avec succès!')
            return redirect('detail_client_commandes', client_id=client.id)
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    context = {
        'client': client,
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/modifier_client.html', context)

@login_required
@require_commandes_feature('gestion_client')
def supprimer_client_commandes(request, client_id):
    """Vue pour supprimer un client dans le module commandes"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_clients_commandes')
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        client = Client.objects.get(id=client_id, agence=agence)
        client_nom = client.intitule
        client.delete()
        messages.success(request, f'Client "{client_nom}" supprimé avec succès!')
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_clients_commandes')

@login_required
def verifier_stock_livraison(request):
    """Vue pour vérifier le stock avant livraison"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
    
    if request.method == 'POST':
        try:
            commande_id = request.POST.get('commande_id')
            commande = Commande.objects.get(id=commande_id, agence=agence)
            
            stock_disponible = commande.article.stock_actuel or 0
            quantite_requise = float(commande.quantite)
            
            if stock_disponible >= quantite_requise:
                return JsonResponse({'success': True, 'stock_suffisant': True})
            else:
                return JsonResponse({
                    'success': True,
                    'stock_suffisant': False,
                    'stock_disponible': stock_disponible,
                    'quantite_requise': quantite_requise
                })
        except Commande.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Commande non trouvée.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})

@login_required
def reporter_livraison(request, livraison_id):
    """Vue pour reporter une livraison"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        livraison = Livraison.objects.get(id=livraison_id, agence=agence)
        if request.method == 'POST':
            nouvelle_date = request.POST.get('nouvelle_date')
            if nouvelle_date:
                livraison.date_livraison = nouvelle_date
                livraison.save()
                messages.success(request, 'Livraison reportée avec succès!')
            else:
                messages.error(request, 'Veuillez spécifier une nouvelle date.')
        return redirect('planification_livraison')
    except Livraison.DoesNotExist:
        messages.error(request, 'Livraison non trouvée.')
        return redirect('planification_livraison')

@login_required
def annuler_livraison(request, livraison_id):
    """Vue pour annuler une livraison"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        livraison = Livraison.objects.get(id=livraison_id, agence=agence)
        livraison.etat_livraison = 'annulee'
        livraison.save()
        messages.success(request, 'Livraison annulée avec succès!')
    except Livraison.DoesNotExist:
        messages.error(request, 'Livraison non trouvée.')
    
    return redirect('planification_livraison')

@login_required
def modifier_ordre_livraison(request):
    """Vue pour modifier l'ordre des livraisons"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    if request.method == 'POST':
        try:
            ordre = request.POST.getlist('ordre[]')
            for index, livraison_id in enumerate(ordre):
                Livraison.objects.filter(id=livraison_id, agence=agence).update(date_creation=timezone.now() - timezone.timedelta(seconds=index))
            messages.success(request, 'Ordre des livraisons modifié avec succès!')
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    return redirect('planification_livraison')

@login_required
def confirmer_livraison(request, livraison_id):
    """Vue pour confirmer une livraison - change l'état à 'confirmee'"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        livraison = Livraison.objects.get(id=livraison_id, agence=agence)
        livraison.etat_livraison = 'confirmee'
        
        # Enregistrer l'heure de livraison si elle n'est pas déjà définie
        if not livraison.heure_livraison:
            from datetime import datetime
            livraison.heure_livraison = datetime.now().time()
        
        livraison.save()
        
        # Créer une notification pour la livraison confirmée
        try:
            Notification.objects.create(
                type_notification='livraison_confirmee',
                titre='Livraison confirmée',
                message=f'La livraison pour le client {livraison.commande.client.intitule if livraison.commande else "N/A"} a été confirmée par le livreur {livraison.livreur.nom if livraison.livreur else "N/A"} {livraison.livreur.prenom if livraison.livreur else ""}',
                agence=agence,
                livraison=livraison,
                commande=livraison.commande
            )
        except Exception as e:
            pass  # Continuer même si la notification échoue
        
        messages.success(request, 'Livraison confirmée avec succès! Vous pouvez maintenant définir l\'état final dans "Définir État de Livraison".')
        return redirect('voir_itineraire')
    except Livraison.DoesNotExist:
        messages.error(request, 'Livraison non trouvée.')
        return redirect('voir_itineraire')

@login_required
def definir_etat_final_livraison(request, livraison_id):
    """Vue pour définir l'état final d'une livraison (Livré, Pas livré, Partiel)"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    try:
        livraison = Livraison.objects.get(id=livraison_id, agence=agence)
        
        # Récupérer toutes les commandes du groupe
        commandes_groupe = None
        if livraison.commande:
            commandes_groupe = Commande.objects.filter(
                client=livraison.commande.client,
                date=livraison.commande.date,
                heure=livraison.commande.heure,
                agence=agence
            ).select_related('article').order_by('article__designation')
        
        if request.method == 'POST':
            nouvel_etat = request.POST.get('etat_livraison')
            
            if nouvel_etat in ['livree', 'pas_livree', 'livree_partiellement']:
                livraison.etat_livraison = nouvel_etat
                
                # Enregistrer l'heure de livraison si elle n'est pas déjà définie
                if not livraison.heure_livraison:
                    from datetime import datetime
                    livraison.heure_livraison = datetime.now().time()
                
                # Si état partiel, traiter les quantités livrées par article
                if nouvel_etat == 'livree_partiellement' and commandes_groupe:
                    # Stocker les quantités livrées dans les notes (temporairement, en JSON)
                    quantites_livrees = {}
                    for cmd in commandes_groupe:
                        quantite_livree = request.POST.get(f'quantite_livree_{cmd.id}', '0')
                        try:
                            quantites_livrees[str(cmd.id)] = {
                                'article_id': cmd.article.id,
                                'article_designation': cmd.article.designation,
                                'quantite_commandee': float(cmd.quantite),
                                'quantite_livree': float(quantite_livree),
                                'quantite_non_livree': float(cmd.quantite) - float(quantite_livree)
                            }
                        except (ValueError, TypeError):
                            pass
                    
                    import json
                    livraison.notes = json.dumps(quantites_livrees, ensure_ascii=False)
                    
                    # Mettre à jour les états des commandes partiellement livrées
                    for cmd in commandes_groupe:
                        quantite_livree = request.POST.get(f'quantite_livree_{cmd.id}', '0')
                        try:
                            quantite_livree_decimal = Decimal(quantite_livree)
                            if quantite_livree_decimal > 0 and quantite_livree_decimal < cmd.quantite:
                                cmd.etat_commande = 'livree_partiellement'
                                cmd.save()
                        except (ValueError, TypeError):
                            pass
                
                livraison.save()
                
                # Créer une notification pour la livraison terminée
                try:
                    etat_display = {
                        'livree': 'Livrée',
                        'livree_partiellement': 'Livrée partiellement',
                        'pas_livree': 'Pas livrée'
                    }.get(nouvel_etat, 'Terminée')
                    
                    Notification.objects.create(
                        type_notification='livraison_terminee',
                        titre=f'Livraison {etat_display.lower()}',
                        message=f'La livraison pour le client {livraison.commande.client.intitule if livraison.commande else "N/A"} a été marquée comme {etat_display.lower()}.',
                        agence=agence,
                        livraison=livraison,
                        commande=livraison.commande
                    )
                except Exception as e:
                    pass  # Continuer même si la notification échoue
                
                # Mettre à jour l'état des commandes selon l'état de livraison
                if commandes_groupe:
                    if nouvel_etat == 'livree':
                        commandes_groupe.update(etat_commande='livree')
                    elif nouvel_etat == 'pas_livree':
                        # Ne pas changer l'état des commandes si pas livrée
                        pass
                    # Pour partiel, déjà traité ci-dessus
                
                messages.success(request, f'État de la livraison défini comme "{livraison.get_etat_livraison_display()}" avec succès!')
                return redirect('definir_etat_livraison')
            else:
                messages.error(request, 'État invalide.')
        
        # Charger les quantités livrées depuis les notes si état partiel
        quantites_livrees = {}
        if livraison.etat_livraison == 'livree_partiellement' and livraison.notes:
            try:
                import json
                quantites_livrees = json.loads(livraison.notes)
            except (json.JSONDecodeError, TypeError):
                pass
        
        # Préparer les quantités livrées pour le template (format simplifié avec clés numériques)
        quantites_livrees_dict = {}
        if quantites_livrees and commandes_groupe:
            for cmd in commandes_groupe:
                cmd_id_str = str(cmd.id)
                if cmd_id_str in quantites_livrees:
                    quantites_livrees_dict[cmd.id] = Decimal(str(quantites_livrees[cmd_id_str].get('quantite_livree', cmd.quantite)))
                else:
                    quantites_livrees_dict[cmd.id] = cmd.quantite
        elif commandes_groupe:
            # Si pas de quantités sauvegardées, initialiser avec les quantités commandées
            for cmd in commandes_groupe:
                quantites_livrees_dict[cmd.id] = cmd.quantite
        
        context = {
            'livraison': livraison,
            'commandes_groupe': commandes_groupe,
            'quantites_livrees': quantites_livrees_dict,
            'agence': agence,
        }
        return render(request, 'supermarket/commandes/definir_etat_final_livraison.html', context)
    except Livraison.DoesNotExist:
        messages.error(request, 'Livraison non trouvée.')
        return redirect('voir_itineraire')

@login_required
def statistiques_clients(request):
    """Vue pour afficher les statistiques clients"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    # Récupérer tous les clients de l'agence
    clients = Client.objects.filter(agence=agence).order_by('intitule')
    
    # Récupérer les familles d'articles pour le filtre
    familles = Famille.objects.all().order_by('intitule')
    
    # Statistiques générales
    total_clients = clients.count()
    
    # Calculer les statistiques des 30 derniers jours
    date_debut = timezone.now().date() - timezone.timedelta(days=30)
    
    # Récupérer les factures des 30 derniers jours
    factures_recentes = FactureCommande.objects.filter(
        agence=agence,
        date_facture__gte=date_debut
    ).select_related('commande', 'commande__client')
    
    # Calculer le montant total
    montant_total = float(factures_recentes.aggregate(
        total=Sum('net_a_payer')
    )['total'] or 0)
    
    # Nombre de factures
    nombre_factures = factures_recentes.count()
    
    context = {
        'agence': agence,
        'clients': clients,
        'familles': familles,
        'total_clients': total_clients,
        'montant_total': montant_total,
        'nombre_factures': nombre_factures,
    }
    return render(request, 'supermarket/commandes/statistiques_clients.html', context)

@login_required
def generer_statistiques_clients(request):
    """Vue pour générer les statistiques clients selon les critères sélectionnés"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Agence non trouvée: {str(e)}'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        client_id = request.POST.get('client')
        famille_id = request.POST.get('famille', '')
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        if not client_id:
            return JsonResponse({'success': False, 'error': 'Le client est obligatoire'})
        
        # Convertir les dates
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Format de date invalide: {str(e)}'})
        
        # Récupérer le client
        try:
            client = Client.objects.get(id=client_id, agence=agence)
        except Client.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Client non trouvé'})
        
        # Récupérer toutes les commandes du client pour la période
        commandes = Commande.objects.filter(
            client=client,
            agence=agence,
            date__gte=date_debut_obj,
            date__lte=date_fin_obj
        ).select_related('article', 'article__categorie').order_by('-date', '-heure')
        
        # Filtrer par famille si sélectionnée
        famille_nom = None
        if famille_id and famille_id != '':
            try:
                famille = Famille.objects.get(id=famille_id)
                famille_nom = famille.intitule
                commandes = commandes.filter(article__categorie=famille)
            except Famille.DoesNotExist:
                pass
        
        # Grouper les commandes par (date, heure) pour obtenir les groupes de commandes
        from collections import defaultdict
        from django.db.models import Sum
        
        groupes_commandes = defaultdict(lambda: {
            'commandes': [],
            'date': None,
            'heure': None,
            'articles': [],
            'quantite_totale': 0,
            'prix_total': 0,
            'etat_livraison': None,
            'livraison_id': None,
        })
        
        for commande in commandes:
            heure_str = commande.heure.strftime('%H:%M:%S') if commande.heure else ''
            groupe_key = f"{commande.date}_{heure_str}"
            
            groupes_commandes[groupe_key]['commandes'].append(commande)
            groupes_commandes[groupe_key]['date'] = commande.date
            groupes_commandes[groupe_key]['heure'] = commande.heure
            
            # Ajouter l'article à la liste des articles du groupe
            article_info = {
                'reference': commande.article.reference_article,
                'designation': commande.article.designation,
                'quantite': float(commande.quantite),
                'prix_unitaire': float(commande.prix_total / commande.quantite) if commande.quantite > 0 else 0,
                'prix_total': float(commande.prix_total),
            }
            groupes_commandes[groupe_key]['articles'].append(article_info)
            groupes_commandes[groupe_key]['quantite_totale'] += float(commande.quantite)
            groupes_commandes[groupe_key]['prix_total'] += float(commande.prix_total)
        
        # Récupérer les états de livraison pour chaque groupe
        liste_commandes_detaillees = []
        nombre_commandes = len(groupes_commandes)
        
        for groupe_key, groupe_data in groupes_commandes.items():
            premiere_commande = groupe_data['commandes'][0]
            
            # Chercher une livraison associée à ce groupe
            livraison = Livraison.objects.filter(
                commande__client=client,
                commande__date=premiere_commande.date,
                commande__heure=premiere_commande.heure,
                agence=agence
            ).first()
            
            etat_livraison = None
            etat_livraison_display = None
            livraison_id = None
            
            if livraison:
                etat_livraison = livraison.etat_livraison
                etat_livraison_display = livraison.get_etat_livraison_display()
                livraison_id = livraison.id
            
            # Convertir date et heure en chaînes pour JSON
            date_str = None
            if groupe_data['date']:
                if isinstance(groupe_data['date'], str):
                    date_str = groupe_data['date']
                else:
                    date_str = groupe_data['date'].strftime('%Y-%m-%d')
            
            heure_str = None
            if groupe_data['heure']:
                if isinstance(groupe_data['heure'], str):
                    heure_str = groupe_data['heure']
                else:
                    heure_str = groupe_data['heure'].strftime('%H:%M:%S')
            
            liste_commandes_detaillees.append({
                'date': date_str,
                'heure': heure_str,
                'articles': groupe_data['articles'],
                'quantite_totale': float(groupe_data['quantite_totale']),
                'prix_total': float(groupe_data['prix_total']),
                'nombre_articles': len(groupe_data['articles']),
                'etat_livraison': etat_livraison,
                'etat_livraison_display': etat_livraison_display or 'Non planifiée',
                'livraison_id': livraison_id,
            })
        
        # Trier par date décroissante
        liste_commandes_detaillees.sort(key=lambda x: (x['date'], x['heure'] or ''), reverse=True)
        
        # Calculer les statistiques par article (pour le tableau récapitulatif)
        statistiques_articles = []
        montant_total = 0.0
        quantite_totale_commandee = 0.0
        
        articles_dict = defaultdict(lambda: {'quantite': 0, 'montant': 0})
        
        for commande in commandes:
            article = commande.article
            articles_dict[article.id]['quantite'] += float(commande.quantite)
            articles_dict[article.id]['montant'] += float(commande.prix_total)
            articles_dict[article.id]['article'] = article
        
        # Construire la liste des statistiques
        for article_id, data in articles_dict.items():
            article = data['article']
            quantite = data['quantite']
            montant = data['montant']
            
            # Calculer le prix unitaire moyen
            prix_unitaire_moyen = float(montant) / float(quantite) if quantite > 0 else 0
            
            statistiques_articles.append({
                'reference_article': article.reference_article,
                'designation': article.designation,
                'quantite_commandee': float(quantite),
                'montant_total': float(montant),
                'prix_unitaire_moyen': prix_unitaire_moyen,
            })
            
            montant_total += float(montant)
            quantite_totale_commandee += float(quantite)
        
        # Trier par quantité commandée décroissante (articles les plus commandés)
        statistiques_articles.sort(key=lambda x: x['quantite_commandee'], reverse=True)
        
        # Stocker les statistiques dans la session pour l'export
        request.session['statistiques_clients'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'client_id': client_id,
            'client_nom': client.intitule,
            'famille_id': famille_id if famille_id else None,
            'famille_nom': famille_nom,
            'statistiques_articles': statistiques_articles,
            'montant_total': float(montant_total),
            'quantite_totale_commandee': float(quantite_totale_commandee),
            'nombre_commandes': nombre_commandes,
            'commandes_detaillees': liste_commandes_detaillees,
        }
        
        return JsonResponse({
            'success': True,
            'total_articles': len(statistiques_articles),
            'quantite_totale_commandee': float(quantite_totale_commandee),
            'montant_total': float(montant_total),
            'nombre_commandes': nombre_commandes,
            'client_nom': client.intitule,
            'famille_nom': famille_nom,
            'statistiques_articles': statistiques_articles,
            'commandes_detaillees': liste_commandes_detaillees
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        error_details = f"Erreur: {str(e)}"
        if hasattr(e, '__class__'):
            error_details += f" (Type: {e.__class__.__name__})"
        return JsonResponse({'success': False, 'error': error_details})

@login_required
def export_statistiques_clients_excel(request):
    """Vue pour exporter les statistiques clients en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_clients')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        client_nom = statistiques_data['client_nom']
        famille_nom = statistiques_data.get('famille_nom', None)
        statistiques_articles = statistiques_data['statistiques_articles']
        montant_total = statistiques_data['montant_total']
        quantite_totale_commandee = statistiques_data['quantite_totale_commandee']
        nombre_commandes = statistiques_data.get('nombre_commandes', 0)
        commandes_detaillees = statistiques_data.get('commandes_detaillees', [])
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Statistiques Client {date_debut} - {date_fin}"
        
        # Style des en-têtes (turquoise)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="06beb6", end_color="06beb6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Définir les bordures (nécessaire pour les cartes et le tableau)
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-tête du document
        ws.merge_cells('A1:F1')
        ws['A1'] = f"STATISTIQUES CLIENT - {agence.nom_agence}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Client: {client_nom}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A3:F3')
        ws['A3'] = f"Période: du {date_debut} au {date_fin}"
        ws['A3'].font = Font(size=12)
        ws['A3'].alignment = Alignment(horizontal="center")
        
        # Ajouter l'information sur la famille si filtrée
        current_row = 4
        if famille_nom:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = f"Famille d'articles: {famille_nom}"
            ws[f'A{current_row}'].font = Font(size=11, italic=True, color="06beb6")
            ws[f'A{current_row}'].alignment = Alignment(horizontal="center")
            ws[f'A{current_row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            current_row += 1
        
        # Informations résumées
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"Total: {len(statistiques_articles)} articles | {quantite_totale_commandee} unités | {montant_total:,.0f} FCFA | {nombre_commandes} commande(s)"
        ws[f'A{current_row}'].font = Font(size=11, italic=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal="center")
        ws[f'A{current_row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        
        # Ajuster les numéros de lignes suivantes
        current_row += 1
        
        # Section des articles les plus commandés (Top 6)
        if statistiques_articles and len(statistiques_articles) > 0:
            top_articles_row = current_row
            ws.merge_cells(f'A{top_articles_row}:F{top_articles_row}')
            ws.cell(row=top_articles_row, column=1, value="🏆 ARTICLES LES PLUS COMMANDÉS")
            top_title_cell = ws.cell(row=top_articles_row, column=1)
            top_title_cell.font = Font(bold=True, size=13, color="FFFFFF")
            top_title_cell.fill = PatternFill(start_color="06beb6", end_color="48b1bf", fill_type="solid")
            top_title_cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            
            # Prendre les 6 premiers articles
            top_articles = statistiques_articles[:6]
            max_quantite = max([a['quantite_commandee'] for a in top_articles]) if top_articles else 1
            
            # Créer une grille de cartes (2 colonnes, 3 lignes)
            for idx, article in enumerate(top_articles):
                rank = idx + 1
                row_start = current_row + (idx // 2) * 8  # 8 lignes par carte
                col_start = (idx % 2) * 3 + 1  # Colonnes A-C pour première carte, D-F pour deuxième
                
                # Couleur selon le rang
                if rank == 1:
                    card_color = "FFD700"  # Or
                    card_bg = "FFF9C4"
                elif rank == 2:
                    card_color = "C0C0C0"  # Argent
                    card_bg = "F5F5F5"
                elif rank == 3:
                    card_color = "CD7F32"  # Bronze
                    card_bg = "FFE0B2"
                else:
                    card_color = "06beb6"  # Turquoise
                    card_bg = "E3F2FD"
                
                # En-tête de la carte avec rang
                ws.merge_cells(f'{chr(64+col_start)}{row_start}:{chr(64+col_start+2)}{row_start}')
                rank_cell = ws.cell(row=row_start, column=col_start, value=f"Rang #{rank}")
                rank_cell.font = Font(bold=True, size=10, color="FFFFFF")
                rank_cell.fill = PatternFill(start_color=card_color, end_color=card_color, fill_type="solid")
                rank_cell.alignment = Alignment(horizontal="center", vertical="center")
                rank_cell.border = thin_border
                
                # Référence
                ws.merge_cells(f'{chr(64+col_start)}{row_start+1}:{chr(64+col_start+2)}{row_start+1}')
                ref_cell = ws.cell(row=row_start+1, column=col_start, value=article['reference_article'])
                ref_cell.font = Font(bold=True, size=9, color="06beb6")
                ref_cell.fill = PatternFill(start_color=card_bg, end_color=card_bg, fill_type="solid")
                ref_cell.alignment = Alignment(horizontal="left", vertical="center")
                ref_cell.border = thin_border
                
                # Désignation (peut être longue, donc sur plusieurs lignes)
                ws.merge_cells(f'{chr(64+col_start)}{row_start+2}:{chr(64+col_start+2)}{row_start+3}')
                desig_cell = ws.cell(row=row_start+2, column=col_start, value=article['designation'])
                desig_cell.font = Font(bold=True, size=9)
                desig_cell.fill = PatternFill(start_color=card_bg, end_color=card_bg, fill_type="solid")
                desig_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                desig_cell.border = thin_border
                
                # Quantité
                qty_label_cell = ws.cell(row=row_start+4, column=col_start, value="Quantité:")
                qty_label_cell.font = Font(bold=True, size=9)
                qty_label_cell.fill = PatternFill(start_color=card_bg, end_color=card_bg, fill_type="solid")
                qty_label_cell.border = thin_border
                quantite_val_cell = ws.cell(row=row_start+4, column=col_start+1, value=float(article['quantite_commandee']))
                quantite_val_cell.number_format = '#,##0.00'
                quantite_val_cell.font = Font(bold=True, size=10, color="28a745")
                quantite_val_cell.fill = PatternFill(start_color=card_bg, end_color=card_bg, fill_type="solid")
                quantite_val_cell.alignment = Alignment(horizontal="right")
                quantite_val_cell.border = thin_border
                
                # Barre de progression (représentée par un pourcentage)
                progress_pct = (article['quantite_commandee'] / max_quantite) * 100
                progress_cell = ws.cell(row=row_start+4, column=col_start+2, value=f"{progress_pct:.0f}%")
                progress_cell.font = Font(size=8, color="6c757d")
                progress_cell.fill = PatternFill(start_color=card_bg, end_color=card_bg, fill_type="solid")
                progress_cell.border = thin_border
                
                # Montant total
                amount_label_cell = ws.cell(row=row_start+5, column=col_start, value="Montant:")
                amount_label_cell.font = Font(bold=True, size=9)
                amount_label_cell.fill = PatternFill(start_color=card_bg, end_color=card_bg, fill_type="solid")
                amount_label_cell.border = thin_border
                montant_val_cell = ws.cell(row=row_start+5, column=col_start+1, value=float(article['montant_total']))
                montant_val_cell.number_format = '#,##0.00'
                montant_val_cell.font = Font(bold=True, size=10, color="06beb6")
                montant_val_cell.fill = PatternFill(start_color=card_bg, end_color=card_bg, fill_type="solid")
                montant_val_cell.alignment = Alignment(horizontal="right")
                montant_val_cell.border = thin_border
                
                # Prix unitaire moyen
                prix_unitaire = article.get('prix_unitaire_moyen', 0)
                if prix_unitaire == 0 and article['quantite_commandee'] > 0:
                    prix_unitaire = article['montant_total'] / article['quantite_commandee']
                prix_unit_cell = ws.cell(row=row_start+5, column=col_start+2, value=f"{prix_unitaire:.0f} FCFA/u")
                prix_unit_cell.font = Font(size=8, color="6c757d")
                prix_unit_cell.fill = PatternFill(start_color=card_bg, end_color=card_bg, fill_type="solid")
                prix_unit_cell.border = thin_border
                
                # Bordures pour toutes les cellules de la carte
                for r in range(row_start, row_start+6):
                    for c in range(col_start, col_start+3):
                        cell = ws.cell(row=r, column=c)
                        if not cell.border:
                            cell.border = thin_border
            
            # Ajuster la hauteur des lignes pour les cartes
            for idx in range(len(top_articles)):
                row_start = current_row + (idx // 2) * 8
                for r in range(row_start, row_start+6):
                    ws.row_dimensions[r].height = 20
            
            # Calculer le nombre de lignes nécessaires pour les cartes
            num_rows_for_cards = (len(top_articles) // 2 + len(top_articles) % 2) * 8
            current_row += num_rows_for_cards + 2
            
            # Titre pour le tableau complet
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws.cell(row=current_row, column=1, value="📊 LISTE COMPLÈTE DES ARTICLES")
            list_title_cell = ws.cell(row=current_row, column=1)
            list_title_cell.font = Font(bold=True, size=13, color="FFFFFF")
            list_title_cell.fill = PatternFill(start_color="06beb6", end_color="48b1bf", fill_type="solid")
            list_title_cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
        
        # En-têtes des colonnes pour le tableau complet
        headers = ['Rang', 'Référence', 'Désignation', 'Quantité Commandée', 'Prix Unitaire Moyen (FCFA)', 'Montant Total (FCFA)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        header_row = current_row
        data_start_row = header_row + 1
        current_row = data_start_row
        
        # Données avec classement (tableau complet)
        row = current_row
        for index, stat in enumerate(statistiques_articles, 1):
            # Rang
            rank_cell = ws.cell(row=row, column=1, value=index)
            rank_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Couleur selon le rang
            if index == 1:
                rank_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Or
                rank_cell.font = Font(bold=True, color="000000")
            elif index == 2:
                rank_cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Argent
                rank_cell.font = Font(bold=True, color="000000")
            elif index == 3:
                rank_cell.fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")  # Bronze
                rank_cell.font = Font(bold=True, color="FFFFFF")
            else:
                rank_cell.fill = PatternFill(start_color="06beb6", end_color="48b1bf", fill_type="solid")  # Turquoise
                rank_cell.font = Font(bold=True, color="FFFFFF")
            
            # Référence
            ws.cell(row=row, column=2, value=stat['reference_article'])
            
            # Désignation
            ws.cell(row=row, column=3, value=stat['designation'])
            
            # Quantité commandée
            quantite_cell = ws.cell(row=row, column=4, value=float(stat['quantite_commandee']))
            quantite_cell.number_format = '#,##0.00'
            quantite_cell.alignment = Alignment(horizontal="right")
            
            # Prix unitaire moyen
            prix_unitaire = stat.get('prix_unitaire_moyen', 0)
            if prix_unitaire == 0 and stat['quantite_commandee'] > 0:
                prix_unitaire = stat['montant_total'] / stat['quantite_commandee']
            prix_cell = ws.cell(row=row, column=5, value=float(prix_unitaire))
            prix_cell.number_format = '#,##0.00'
            prix_cell.alignment = Alignment(horizontal="right")
            
            # Montant total
            montant_cell = ws.cell(row=row, column=6, value=float(stat['montant_total']))
            montant_cell.number_format = '#,##0.00'
            montant_cell.alignment = Alignment(horizontal="right")
            
            # Mise en évidence des 3 premiers
            if index <= 3:
                for col in range(2, 7):
                    cell = ws.cell(row=row, column=col)
                    if index == 1:
                        cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                    elif index == 2:
                        cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                    elif index == 3:
                        cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            
            row += 1
        
        # Ligne des totaux (turquoise)
        row += 1
        total_fill = PatternFill(start_color="06beb6", end_color="48b1bf", fill_type="solid")
        total_font = Font(bold=True, color="FFFFFF", size=11)
        
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=2, value="").font = total_font
        ws.cell(row=row, column=2).fill = total_fill
        ws.cell(row=row, column=3, value="").font = total_font
        ws.cell(row=row, column=3).fill = total_fill
        
        quantite_total_cell = ws.cell(row=row, column=4, value=float(quantite_totale_commandee))
        quantite_total_cell.font = total_font
        quantite_total_cell.fill = total_fill
        quantite_total_cell.number_format = '#,##0.00'
        quantite_total_cell.alignment = Alignment(horizontal="right")
        
        ws.cell(row=row, column=5, value="").font = total_font
        ws.cell(row=row, column=5).fill = total_fill
        
        montant_total_cell = ws.cell(row=row, column=6, value=float(montant_total))
        montant_total_cell.font = total_font
        montant_total_cell.fill = total_fill
        montant_total_cell.number_format = '#,##0.00'
        montant_total_cell.alignment = Alignment(horizontal="right")
        
        # Informations supplémentaires
        row += 2
        ws.cell(row=row, column=1, value="Nombre de commandes:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=nombre_commandes)
        
        row += 1
        ws.cell(row=row, column=1, value="Nombre d'articles différents:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=len(statistiques_articles))
        
        row += 1
        if quantite_totale_commandee > 0:
            prix_moyen_global = montant_total / quantite_totale_commandee
            ws.cell(row=row, column=1, value="Prix unitaire moyen global:").font = Font(bold=True)
            prix_moyen_cell = ws.cell(row=row, column=2, value=float(prix_moyen_global))
            prix_moyen_cell.number_format = '#,##0.00'
            prix_moyen_cell.font = Font(bold=True)
        
        # Ajuster la largeur des colonnes pour le tableau complet (après les cartes)
        # Les colonnes ont déjà été ajustées pour les cartes, on les réajuste pour le tableau
        ws.column_dimensions['A'].width = 10  # Rang
        ws.column_dimensions['B'].width = 15  # Référence
        ws.column_dimensions['C'].width = 40  # Désignation
        ws.column_dimensions['D'].width = 18  # Quantité
        ws.column_dimensions['E'].width = 22  # Prix unitaire
        ws.column_dimensions['F'].width = 20  # Montant total
        
        # Appliquer les bordures aux en-têtes
        for col in range(1, 7):
            cell = ws.cell(row=header_row, column=col)
            cell.border = thin_border
        
        # Appliquer les bordures aux données
        for row_num in range(data_start_row, row - 2):
            for col in range(1, 7):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
        
        # Appliquer les bordures à la ligne des totaux
        for col in range(1, 7):
            cell = ws.cell(row=row - 2, column=col)
            cell.border = thin_border
        
        # Ajouter les détails des commandes sur la même feuille
        if commandes_detaillees:
            row_cmd = row + 3  # Espacement après les informations supplémentaires
            
            # Titre de la section avec style amélioré
            ws.merge_cells(f'A{row_cmd}:F{row_cmd}')
            ws.cell(row=row_cmd, column=1, value=f"📋 DÉTAILS DES COMMANDES ({nombre_commandes} commande(s))")
            title_cell = ws.cell(row=row_cmd, column=1)
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="06beb6", end_color="48b1bf", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            row_cmd += 2  # Espacement supplémentaire
            
            for index, commande in enumerate(commandes_detaillees, 1):
                # Déterminer la couleur selon l'état de livraison
                etat = commande.get('etat_livraison', '')
                if etat == 'livree':
                    etat_color = "28a745"  # Vert
                    etat_bg = "E8F5E9"
                elif etat == 'livree_partiellement':
                    etat_color = "ffc107"  # Jaune/Orange
                    etat_bg = "FFF9C4"
                elif etat == 'pas_livree':
                    etat_color = "dc3545"  # Rouge
                    etat_bg = "FFEBEE"
                elif etat in ['planifiee', 'en_preparation', 'en_cours']:
                    etat_color = "17a2b8"  # Bleu clair
                    etat_bg = "E3F2FD"
                elif etat == 'confirmee':
                    etat_color = "fd7e14"  # Orange
                    etat_bg = "FFF3E0"
                else:
                    etat_color = "6c757d"  # Gris
                    etat_bg = "F5F5F5"
                
                # En-tête de la commande avec style amélioré
                ws.merge_cells(f'A{row_cmd}:F{row_cmd}')
                date_str = commande['date'] if commande['date'] else 'Date non disponible'
                heure_str = commande['heure'] if commande['heure'] else 'N/A'
                if heure_str and ':' in heure_str:
                    heure_str = heure_str.split(':')[0] + ':' + heure_str.split(':')[1]
                
                # Formatage de la date en français
                try:
                    from datetime import datetime
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    date_formatted = date_obj.strftime('%d/%m/%Y')
                except:
                    date_formatted = date_str
                
                ws.cell(row=row_cmd, column=1, value=f"COMMANDE #{index}")
                cmd_header_cell = ws.cell(row=row_cmd, column=1)
                cmd_header_cell.font = Font(bold=True, size=12, color="FFFFFF")
                cmd_header_cell.fill = PatternFill(start_color=etat_color, end_color=etat_color, fill_type="solid")
                cmd_header_cell.alignment = Alignment(horizontal="left", vertical="center")
                row_cmd += 1
                
                # Informations de la commande sur plusieurs lignes avec style
                info_row = row_cmd
                ws.cell(row=info_row, column=1, value="📅 Date:").font = Font(bold=True, size=10)
                ws.cell(row=info_row, column=2, value=date_formatted).font = Font(size=10)
                ws.cell(row=info_row, column=3, value="🕐 Heure:").font = Font(bold=True, size=10)
                ws.cell(row=info_row, column=4, value=heure_str).font = Font(size=10)
                ws.cell(row=info_row, column=5, value="📦 État:").font = Font(bold=True, size=10)
                etat_cell = ws.cell(row=info_row, column=6, value=commande['etat_livraison_display'])
                etat_cell.font = Font(bold=True, size=10, color=etat_color)
                etat_cell.fill = PatternFill(start_color=etat_bg, end_color=etat_bg, fill_type="solid")
                row_cmd += 1
                
                # Statistiques de la commande avec style amélioré
                stats_row = row_cmd
                ws.merge_cells(f'A{stats_row}:B{stats_row}')
                ws.cell(row=stats_row, column=1, value="📊 Statistiques de la commande:").font = Font(bold=True, size=10, color="2c3e50")
                stats_header_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                ws.cell(row=stats_row, column=1).fill = stats_header_fill
                
                ws.cell(row=stats_row, column=3, value="Articles:").font = Font(bold=True, size=10)
                ws.cell(row=stats_row, column=4, value=commande['nombre_articles']).font = Font(size=10, color="28a745")
                ws.cell(row=stats_row, column=5, value="Quantité totale:").font = Font(bold=True, size=10)
                quantite_tot_cell = ws.cell(row=stats_row, column=6, value=float(commande['quantite_totale']))
                quantite_tot_cell.number_format = '#,##0.00'
                quantite_tot_cell.alignment = Alignment(horizontal="right")
                quantite_tot_cell.font = Font(size=10, color="28a745")
                row_cmd += 1
                
                ws.cell(row=row_cmd, column=5, value="Montant total:").font = Font(bold=True, size=10)
                montant_tot_cell = ws.cell(row=row_cmd, column=6, value=float(commande['prix_total']))
                montant_tot_cell.number_format = '#,##0.00'
                montant_tot_cell.alignment = Alignment(horizontal="right")
                montant_tot_cell.font = Font(bold=True, size=11, color="06beb6")
                row_cmd += 1
                
                # En-têtes des colonnes pour les articles avec style amélioré
                article_headers = ['Référence', 'Désignation', 'Quantité', 'Prix Unitaire (FCFA)', 'Prix Total (FCFA)']
                header_fill_articles = PatternFill(start_color="06beb6", end_color="48b1bf", fill_type="solid")
                header_font_articles = Font(bold=True, color="FFFFFF", size=10)
                
                for col, header in enumerate(article_headers, 1):
                    cell = ws.cell(row=row_cmd, column=col, value=header)
                    cell.font = header_font_articles
                    cell.fill = header_fill_articles
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                row_cmd += 1
                
                # Articles de la commande avec alternance de couleurs
                for art_index, article in enumerate(commande['articles']):
                    # Couleur alternée pour les lignes
                    row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    if art_index % 2 == 1:
                        row_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                    
                    ws.cell(row=row_cmd, column=1, value=article['reference'])
                    ref_cell = ws.cell(row=row_cmd, column=1)
                    ref_cell.font = Font(bold=True, size=9, color="06beb6")
                    ref_cell.fill = row_fill
                    ref_cell.border = thin_border
                    
                    ws.cell(row=row_cmd, column=2, value=article['designation'])
                    desig_cell = ws.cell(row=row_cmd, column=2)
                    desig_cell.font = Font(size=9)
                    desig_cell.fill = row_fill
                    desig_cell.border = thin_border
                    
                    quantite_cell = ws.cell(row=row_cmd, column=3, value=float(article['quantite']))
                    quantite_cell.number_format = '#,##0.00'
                    quantite_cell.alignment = Alignment(horizontal="right")
                    quantite_cell.font = Font(size=9, color="28a745")
                    quantite_cell.fill = row_fill
                    quantite_cell.border = thin_border
                    
                    prix_unit_cell = ws.cell(row=row_cmd, column=4, value=float(article['prix_unitaire']))
                    prix_unit_cell.number_format = '#,##0.00'
                    prix_unit_cell.alignment = Alignment(horizontal="right")
                    prix_unit_cell.font = Font(size=9)
                    prix_unit_cell.fill = row_fill
                    prix_unit_cell.border = thin_border
                    
                    prix_total_cell = ws.cell(row=row_cmd, column=5, value=float(article['prix_total']))
                    prix_total_cell.number_format = '#,##0.00'
                    prix_total_cell.alignment = Alignment(horizontal="right")
                    prix_total_cell.font = Font(bold=True, size=9, color="2c3e50")
                    prix_total_cell.fill = row_fill
                    prix_total_cell.border = thin_border
                    
                    row_cmd += 1
                
                # Ligne de séparation entre les commandes avec style
                ws.merge_cells(f'A{row_cmd}:F{row_cmd}')
                separator_cell = ws.cell(row=row_cmd, column=1)
                separator_cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
                row_cmd += 2  # Espacement entre les commandes
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Client_{client_nom}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le fichier dans la réponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        buffer.close()
        
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erreur lors de l\'export: {str(e)}'})

@login_required
def generer_rapport_livraison(request):
    """Vue pour générer un rapport de livraison - Redirige vers generer_rapport_livreur"""
    return generer_rapport_livreur(request)

@login_required
def export_rapport_livraison_excel(request):
    """Vue pour exporter le rapport de livraison en Excel - Redirige vers export_rapport_livreur_excel"""
    return export_rapport_livreur_excel(request)

@login_required
def consulter_livreurs(request):
    """Vue pour consulter les livreurs"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    # Pour l'instant, retourner une liste vide car le modèle Livreur n'existe peut-être pas encore
    context = {
        'livreurs': [],
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/consulter_livreurs.html', context)

@login_required
def enregistrer_livreur(request):
    """Vue pour enregistrer un nouveau livreur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    if request.method == 'POST':
        messages.info(request, 'Enregistrement de livreur en cours de développement.')
        return redirect('consulter_livreurs')
    
    context = {
        'agence': agence,
    }
    return render(request, 'supermarket/commandes/enregistrer_livreur.html', context)

@login_required
def detail_livreur(request, livreur_id):
    """Vue pour afficher les détails d'un livreur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    messages.info(request, 'Détails livreur en cours de développement.')
    return redirect('consulter_livreurs')

@login_required
def modifier_livreur(request, livreur_id):
    """Vue pour modifier un livreur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    messages.info(request, 'Modification livreur en cours de développement.')
    return redirect('consulter_livreurs')

@login_required
def supprimer_livreur(request, livreur_id):
    """Vue pour supprimer un livreur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    messages.info(request, 'Suppression livreur en cours de développement.')
    return redirect('consulter_livreurs')

@login_required
def get_notifications(request):
    """Vue pour récupérer les notifications de l'agence"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
    
    try:
        # Récupérer les notifications non lues et les dernières notifications lues
        notifications_non_lues = Notification.objects.filter(
            agence=agence,
            lu=False
        ).select_related('commande', 'livraison', 'commande__client', 'livraison__commande__client').order_by('-date_creation')[:50]
        
        notifications_lues = Notification.objects.filter(
            agence=agence,
            lu=True
        ).select_related('commande', 'livraison', 'commande__client', 'livraison__commande__client').order_by('-date_creation')[:20]
        
        # Compter le nombre total de notifications non lues
        nombre_non_lues = Notification.objects.filter(agence=agence, lu=False).count()
        
        # Préparer les données des notifications
        notifications_data = []
        
        for notif in notifications_non_lues:
            notifications_data.append({
                'id': notif.id,
                'type': notif.type_notification,
                'type_display': notif.get_type_notification_display(),
                'titre': notif.titre,
                'message': notif.message,
                'lu': notif.lu,
                'date_creation': notif.date_creation.strftime('%d/%m/%Y %H:%M'),
                'date_creation_iso': notif.date_creation.isoformat(),
                'commande_id': notif.commande.id if notif.commande else None,
                'livraison_id': notif.livraison.id if notif.livraison else None,
            })
        
        for notif in notifications_lues:
            notifications_data.append({
                'id': notif.id,
                'type': notif.type_notification,
                'type_display': notif.get_type_notification_display(),
                'titre': notif.titre,
                'message': notif.message,
                'lu': notif.lu,
                'date_creation': notif.date_creation.strftime('%d/%m/%Y %H:%M'),
                'date_creation_iso': notif.date_creation.isoformat(),
                'commande_id': notif.commande.id if notif.commande else None,
                'livraison_id': notif.livraison.id if notif.livraison else None,
            })
        
        return JsonResponse({
            'success': True,
            'notifications': notifications_data,
            'nombre_non_lues': nombre_non_lues,
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la récupération des notifications: {str(e)}',
            'traceback': traceback.format_exc()
        })

@login_required
def marquer_notification_lue(request, notification_id):
    """Vue pour marquer une notification comme lue"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
    
    try:
        notification = Notification.objects.get(id=notification_id, agence=agence)
        notification.lu = True
        notification.save()
        
        # Compter le nombre restant de notifications non lues
        nombre_non_lues = Notification.objects.filter(agence=agence, lu=False).count()
        
        return JsonResponse({
            'success': True,
            'message': 'Notification marquée comme lue.',
            'nombre_non_lues': nombre_non_lues
        })
    except Notification.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Notification non trouvée.'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erreur: {str(e)}'})

@login_required
def marquer_toutes_notifications_lues(request):
    """Vue pour marquer toutes les notifications comme lues"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
    
    try:
        Notification.objects.filter(agence=agence, lu=False).update(lu=True)
        return JsonResponse({
            'success': True,
            'message': 'Toutes les notifications ont été marquées comme lues.',
            'nombre_non_lues': 0
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erreur: {str(e)}'})

@login_required
@require_commandes_feature('suivi_client')
def suivi_client(request):
    """Vue pour le suivi des clients par zone"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    compte = get_user_compte(request)
    if not compte:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('login_commandes')
    
    # Récupérer toutes les zones distinctes des clients de l'agence
    zones = Client.objects.filter(agence=agence).values_list('zone', flat=True).distinct().exclude(zone__isnull=True).exclude(zone='').order_by('zone')
    
    # Zone sélectionnée (depuis GET)
    zone_selectionnee = request.GET.get('zone', '')
    
    # Date sélectionnée (depuis GET) - Par défaut, date du jour
    from django.utils import timezone
    date_selectionnee_str = request.GET.get('date_selectionnee', '')
    
    if date_selectionnee_str:
        try:
            date_selectionnee = datetime.strptime(date_selectionnee_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            date_selectionnee = timezone.now().date()
    else:
        date_selectionnee = timezone.now().date()
    
    # Récupérer les clients de la zone sélectionnée
    clients = Client.objects.filter(agence=agence)
    if zone_selectionnee:
        clients = clients.filter(zone=zone_selectionnee)
    
    clients = clients.order_by('intitule')
    
    # Récupérer les actions des clients pour la date sélectionnée uniquement
    # Organiser par client puis par plage horaire
    actions_dict = {}
    suivi_actions = SuiviClientAction.objects.filter(
        agence=agence,
        client__in=clients,
        date_action__date=date_selectionnee  # Filtrer par date uniquement
    ).select_related('client').order_by('client', 'plage_horaire', '-date_action', '-heure_appel')
    
    # Filtrer selon l'utilisateur (ACM voit seulement ses actions)
    suivi_actions = filter_suivi_client_by_user(suivi_actions, compte)
    
    for action in suivi_actions:
        client_id = action.client.id
        plage = action.plage_horaire or ''
        
        if client_id not in actions_dict:
            actions_dict[client_id] = {}
        
        if plage not in actions_dict[client_id]:
            actions_dict[client_id][plage] = []
        
        actions_dict[client_id][plage].append({
            'action': action.action or '',
            'heure_appel': action.heure_appel,
            'date_action': action.date_action,
            'id': action.id
        })
    
    # Définir les plages horaires
    plages_horaires = SuiviClientAction.PLAGE_HORAIRE_CHOICES
    
    # Fonction pour déterminer la plage horaire à partir d'une heure
    def get_plage_from_heure(heure_obj):
        if not heure_obj:
            return ''
        from datetime import time
        for plage_code, plage_label in plages_horaires:
            debut_str, fin_str = plage_code.split('-')
            try:
                debut_h, debut_m = map(int, debut_str.replace('h', ':').split(':'))
                fin_h, fin_m = map(int, fin_str.replace('h', ':').split(':'))
                debut = time(debut_h, debut_m)
                fin = time(fin_h, fin_m)
                if debut <= heure_obj <= fin:
                    return plage_code
            except:
                continue
        return ''
    
    # Récupérer les commandes avec leurs détails (articles, quantités, heures)
    commandes_dict = {}
    commandes_totaux = {}
    # Utiliser distinct() pour éviter les doublons et regrouper par client, date, heure, article
    commandes = Commande.objects.filter(
        agence=agence,
        client__in=clients,
        date=date_selectionnee
    ).select_related('client', 'article').order_by('client', 'heure', 'article').distinct()
    
    # Utiliser un set pour éviter de compter deux fois la même commande
    commandes_deja_comptees = set()
    
    for cmd in commandes:
        client_id = cmd.client.id
        if client_id not in commandes_dict:
            commandes_dict[client_id] = {}
            commandes_totaux[client_id] = 0
        
        # Déterminer la plage horaire de la commande
        plage_code = get_plage_from_heure(cmd.heure) if cmd.heure else ''
        
        if plage_code not in commandes_dict[client_id]:
            commandes_dict[client_id][plage_code] = []
        
        # Créer une clé unique pour éviter les doublons : (client_id, date, heure, article_id)
        cmd_key = (client_id, str(cmd.date), str(cmd.heure) if cmd.heure else '', cmd.article.id if cmd.article else None)
        
        # Ne compter qu'une seule fois chaque commande unique
        if cmd_key not in commandes_deja_comptees:
            commandes_totaux[client_id] += float(cmd.prix_total)
            commandes_deja_comptees.add(cmd_key)
        
        # Formater les informations de la commande
        quantite_int = int(cmd.quantite) if cmd.quantite else 0
        article_str = f"{cmd.article.designation if cmd.article else 'N/A'}*{quantite_int}PCS (Qté: {float(cmd.quantite)}) ({cmd.date.strftime('%d/%m/%Y')} {cmd.heure.strftime('%H:%M') if cmd.heure else ''})"
        
        commandes_dict[client_id][plage_code].append({
            'article': article_str,
            'heure': cmd.heure.strftime('%H:%M') if cmd.heure else '',
            'quantite': float(cmd.quantite),
            'prix_total': float(cmd.prix_total),
        })
    
    context = {
        'agence': agence,
        'zones': zones,
        'zone_selectionnee': zone_selectionnee,
        'date_selectionnee': date_selectionnee.strftime('%Y-%m-%d'),
        'clients': clients,
        'actions_dict': actions_dict,
        'plages_horaires': plages_horaires,
        'commandes_dict': commandes_dict,
        'commandes_totaux': commandes_totaux,
    }
    return render(request, 'supermarket/commandes/suivi_client.html', context)

@login_required
def sauvegarder_action_client(request):
    """Vue pour sauvegarder l'action d'un client"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
    
    if request.method == 'POST':
        try:
            client_id = request.POST.get('client_id')
            action = request.POST.get('action', '').strip()
            heure_appel_str = request.POST.get('heure_appel', '').strip()
            
            if not client_id:
                return JsonResponse({'success': False, 'error': 'Client ID manquant.'})
            
            client = Client.objects.get(id=client_id, agence=agence)
            
            from django.utils import timezone
            from datetime import datetime
            
            # Récupérer la date depuis le POST, sinon utiliser la date/heure actuelle
            date_action_str = request.POST.get('date_action', '')
            if date_action_str:
                try:
                    # Format attendu: YYYY-MM-DD
                    date_action = datetime.strptime(date_action_str, '%Y-%m-%d').date()
                    # Combiner avec l'heure actuelle pour avoir un datetime complet
                    date_action = timezone.make_aware(datetime.combine(date_action, timezone.now().time()))
                except (ValueError, TypeError):
                    date_action = timezone.now()
            else:
                date_action = timezone.now()
            
            # Remplir automatiquement l'heure si un commentaire est saisi
            heure_appel = None
            if action:  # Si un commentaire/action est saisi
                # Utiliser l'heure fournie par le navigateur (heure locale)
                if heure_appel_str:
                    try:
                        # Format attendu: HH:MM (ex: 07:30) - heure locale du navigateur
                        heure_appel = datetime.strptime(heure_appel_str, '%H:%M').time()
                    except ValueError:
                        # Si erreur, utiliser l'heure actuelle du serveur
                        heure_appel = timezone.now().time()
                else:
                    # Si pas d'heure fournie, utiliser l'heure actuelle du serveur
                    heure_appel = timezone.now().time()
            elif heure_appel_str:  # Sinon, utiliser l'heure fournie manuellement (si nécessaire)
                try:
                    # Format attendu: HH:MM (ex: 07:30)
                    heure_appel = datetime.strptime(heure_appel_str, '%H:%M').time()
                except ValueError:
                    pass
            
            # Récupérer la plage horaire depuis le POST
            plage_horaire = request.POST.get('plage_horaire', '').strip()
            
            # Créer une nouvelle action (plusieurs appels possibles par client)
            suivi_action = SuiviClientAction.objects.create(
                client=client,
                agence=agence,
                action=action,
                heure_appel=heure_appel,
                plage_horaire=plage_horaire if plage_horaire else None,
                date_action=date_action,
                created_by=request.user  # Enregistrer qui a créé l'action
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Action enregistrée avec succès.',
                'date_action': suivi_action.date_action.strftime('%d/%m/%Y %H:%M'),
                'heure_appel': suivi_action.heure_appel.strftime('%H:%M') if suivi_action.heure_appel else ''
            })
            
        except Client.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Client non trouvé.'})
        except Exception as e:
            import traceback
            return JsonResponse({
                'success': False,
                'error': f'Erreur: {str(e)}',
                'traceback': traceback.format_exc()
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'})

@login_required
def modifier_action_client(request):
    """Vue pour modifier une action existante"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
    
    if request.method == 'POST':
        try:
            action_id = request.POST.get('action_id')
            client_id = request.POST.get('client_id')
            action = request.POST.get('action', '').strip()
            heure_appel_str = request.POST.get('heure_appel', '').strip()
            
            if not action_id or not client_id:
                return JsonResponse({'success': False, 'error': 'ID manquant.'})
            
            client = Client.objects.get(id=client_id, agence=agence)
            suivi_action = SuiviClientAction.objects.get(id=action_id, client=client, agence=agence)
            
            from django.utils import timezone
            from datetime import datetime
            
            # Remplir automatiquement l'heure si un commentaire est modifié
            heure_appel = None
            if action:  # Si un commentaire/action est saisi
                # Utiliser l'heure fournie par le navigateur (heure locale)
                if heure_appel_str:
                    try:
                        # Format attendu: HH:MM (ex: 07:30) - heure locale du navigateur
                        heure_appel = datetime.strptime(heure_appel_str, '%H:%M').time()
                    except ValueError:
                        # Si erreur, utiliser l'heure actuelle du serveur
                        heure_appel = timezone.now().time()
                else:
                    # Si pas d'heure fournie, utiliser l'heure actuelle du serveur
                    heure_appel = timezone.now().time()
            elif heure_appel_str:  # Sinon, utiliser l'heure fournie manuellement (si nécessaire)
                try:
                    heure_appel = datetime.strptime(heure_appel_str, '%H:%M').time()
                except ValueError:
                    pass
            
            suivi_action.action = action
            if heure_appel:
                suivi_action.heure_appel = heure_appel
            suivi_action.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Action modifiée avec succès.',
            })
            
        except (Client.DoesNotExist, SuiviClientAction.DoesNotExist):
            return JsonResponse({'success': False, 'error': 'Action ou client non trouvé.'})
        except Exception as e:
            import traceback
            return JsonResponse({
                'success': False,
                'error': f'Erreur: {str(e)}',
                'traceback': traceback.format_exc()
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'})

@login_required
@require_commandes_feature('rapport_acm')
def rapport_acm(request):
    """Vue pour afficher la page de génération de rapport ACM"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    compte = get_user_compte(request)
    if not compte:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('login_commandes')
    
    # Récupérer toutes les zones distinctes
    zones = Client.objects.filter(agence=agence).values_list('zone', flat=True).distinct().exclude(zone__isnull=True).exclude(zone='').order_by('zone')
    
    context = {
        'agence': agence,
        'zones': zones,
        'compte': compte,
    }
    return render(request, 'supermarket/commandes/rapport_acm.html', context)

@login_required
def generer_rapport_acm(request):
    """Vue pour générer le rapport ACM"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'success': False, 'error': 'Aucune agence trouvée.'})
    
    if request.method == 'POST':
        try:
            date_debut = request.POST.get('date_debut')
            date_fin = request.POST.get('date_fin')
            zone = request.POST.get('zone', '')
            
            if not date_debut or not date_fin:
                return JsonResponse({'success': False, 'error': 'Veuillez sélectionner une période.'})
            
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
            
            # Récupérer les clients avec leurs actions de suivi
            clients_query = Client.objects.filter(agence=agence)
            if zone:
                clients_query = clients_query.filter(zone=zone)
            
            clients = clients_query.select_related('agence').order_by('ref', 'detail', 'intitule')
            
            # Définir les plages horaires
            PLAGES_HORAIRES = [
                ('06h30-08h30', '06h30-08h30'),
                ('08h30-10h00', '08h30-10h00'),
                ('10h00-11h30', '10h00-11h30'),
                ('14h30-16h00', '14h30-16h00'),
                ('16h00-17h30', '16h00-17h30'),
            ]
            
            # Récupérer toutes les actions de suivi dans la période (plusieurs par client)
            suivi_actions = SuiviClientAction.objects.filter(
                agence=agence,
                client__in=clients,
                date_action__date__range=[date_debut_obj, date_fin_obj]
            ).select_related('client').order_by('client', 'plage_horaire', '-date_action', '-heure_appel')
            
            # Filtrer selon l'utilisateur (ACM voit seulement ses actions)
            compte = get_user_compte(request)
            if compte:
                suivi_actions = filter_suivi_client_by_user(suivi_actions, compte)
            
            # Organiser les actions par client et par plage horaire
            actions_dict = {}
            for action in suivi_actions:
                client_id = action.client.id
                plage = action.plage_horaire or ''
                
                if client_id not in actions_dict:
                    actions_dict[client_id] = {}
                
                if plage not in actions_dict[client_id]:
                    actions_dict[client_id][plage] = []
                
                actions_dict[client_id][plage].append({
                    'action': action.action or '',
                    'heure_appel': action.heure_appel.strftime('%H:%M') if action.heure_appel else '',
                    'date_action': action.date_action
                })
            
            # Récupérer les commandes pour chaque client dans la période
            # Grouper les commandes par client, date et heure (une commande = plusieurs articles)
            from collections import defaultdict
            commandes_dict = defaultdict(lambda: defaultdict(list))
            
            commandes = Commande.objects.filter(
                agence=agence,
                client__in=clients,
                date__range=[date_debut_obj, date_fin_obj]
            ).select_related('client', 'article').order_by('client', 'date', 'heure')
            
            for cmd in commandes:
                # Clé pour regrouper par client, date et heure
                key = (cmd.client.id, cmd.date, cmd.heure)
                commandes_dict[cmd.client.id][key].append({
                    'article': cmd.article.designation if cmd.article else 'N/A',
                    'quantite': float(cmd.quantite),
                    'prix_total': float(cmd.prix_total),
                })
            
            # Convertir en format final : liste de commandes groupées par client
            commandes_grouped = {}
            for client_id, commandes_par_client in commandes_dict.items():
                commandes_grouped[client_id] = []
                for (client_id_key, date, heure), articles in commandes_par_client.items():
                    total_commande = sum(art['prix_total'] for art in articles)
                    articles_list = [f"{art['article']} (Qté: {art['quantite']})" for art in articles]
                    commandes_grouped[client_id].append({
                        'date': date.strftime('%d/%m/%Y'),
                        'heure': heure.strftime('%H:%M') if heure else '',
                        'articles': ', '.join(articles_list),
                        'total': total_commande,
                    })
            
            # Fonction pour déterminer la plage horaire à partir d'une heure
            def get_plage_from_heure(heure_str):
                if not heure_str:
                    return ''
                try:
                    from datetime import time
                    heure_obj = datetime.strptime(heure_str, '%H:%M').time()
                    for plage_code, plage_label in PLAGES_HORAIRES:
                        debut_str, fin_str = plage_code.split('-')
                        debut = datetime.strptime(debut_str, '%Hh%M').time()
                        fin = datetime.strptime(fin_str, '%Hh%M').time()
                        if debut <= heure_obj <= fin:
                            return plage_code
                except:
                    pass
                return ''
            
            # Calculer les totaux par client et organiser par plage horaire
            clients_data = []
            for client in clients:
                actions_par_plage = actions_dict.get(client.id, {})
                commandes_client = commandes_grouped.get(client.id, [])
                
                # Organiser les commandes par plage horaire
                commandes_par_plage = {}
                for cmd in commandes_client:
                    plage = get_plage_from_heure(cmd['heure'])
                    if plage:
                        if plage not in commandes_par_plage:
                            commandes_par_plage[plage] = []
                        commandes_par_plage[plage].append(cmd)
                
                # Calculer les totaux par plage horaire
                totaux_par_plage = {}
                for plage_code, plage_label in PLAGES_HORAIRES:
                    total_plage = sum(cmd['total'] for cmd in commandes_par_plage.get(plage_code, []))
                    totaux_par_plage[plage_code] = total_plage
                
                total_commandes = sum(cmd['total'] for cmd in commandes_client)
                
                # Créer UNE SEULE ligne par client avec toutes les données organisées par plage horaire
                client_json = {
                        'id': client.id,
                        'detail': client.detail or '',
                        'ref': client.ref or '',
                        'potentiel': client.potentiel or '',
                        'nom': client.intitule,
                        'telephone': client.telephone,
                    'actions_par_plage': actions_par_plage,
                    'commandes_par_plage': commandes_par_plage,
                    'totaux_par_plage': totaux_par_plage,
                    'total': float(total_commandes),
                }
                clients_data.append(client_json)
            
            # Convertir les données pour la session (sérialisation JSON)
            clients_data_json = []
            for client_data in clients_data:
                # Sérialiser les actions par plage
                actions_par_plage_json = {}
                for plage, actions_list in client_data['actions_par_plage'].items():
                    actions_par_plage_json[plage] = []
                    for action in actions_list:
                        actions_par_plage_json[plage].append({
                            'action': action['action'],
                            'heure_appel': action['heure_appel'],
                            'date_action': action['date_action'].strftime('%Y-%m-%d %H:%M:%S') if hasattr(action['date_action'], 'strftime') else str(action['date_action'])
                        })
                
                # Sérialiser les commandes par plage
                commandes_par_plage_json = {}
                for plage, commandes_list in client_data['commandes_par_plage'].items():
                    commandes_par_plage_json[plage] = []
                    for cmd in commandes_list:
                        commandes_par_plage_json[plage].append({
                        'date': cmd['date'],
                        'heure': cmd['heure'],
                        'articles': cmd['articles'],
                        'total': float(cmd['total']),
                    })
                
                client_json = {
                    'id': client_data['id'],
                    'detail': client_data['detail'],
                    'ref': client_data['ref'],
                    'potentiel': client_data['potentiel'],
                    'nom': client_data['nom'],
                    'telephone': client_data['telephone'],
                    'actions_par_plage': actions_par_plage_json,
                    'commandes_par_plage': commandes_par_plage_json,
                    'totaux_par_plage': {k: float(v) for k, v in client_data['totaux_par_plage'].items()},
                    'total': float(client_data['total']),
                }
                clients_data_json.append(client_json)
            
            # Stocker dans la session pour l'export Excel
            request.session['rapport_acm'] = {
                'date_debut': date_debut,
                'date_fin': date_fin,
                'zone': zone,
                'clients_data': clients_data_json,
            }
            
            return JsonResponse({
                'success': True,
                'clients': clients_data_json,
            })
            
        except Exception as e:
            import traceback
            return JsonResponse({
                'success': False,
                'error': f'Erreur: {str(e)}',
                'traceback': traceback.format_exc()
            })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'})

@login_required
def export_rapport_acm_excel(request):
    """Vue pour exporter le rapport ACM en Excel"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Aucune agence trouvée.')
        return redirect('rapport_acm')
    
    try:
        rapport_data = request.session.get('rapport_acm')
        if not rapport_data:
            messages.error(request, 'Aucun rapport à exporter. Veuillez générer un rapport d\'abord.')
            return redirect('rapport_acm')
        
        print(f"[DEBUG EXPORT] Données de session récupérées: {len(rapport_data.get('clients_data', []))} clients")
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Rapport ACM"
        
        # Styles
        header_fill = PatternFill(start_color="06beb6", end_color="48b1bf", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_font_bold = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-tête
        ws.merge_cells('A1:H1')
        ws['A1'] = f"RAPPORT ACM - {agence.nom_agence}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Période : {datetime.strptime(rapport_data['date_debut'], '%Y-%m-%d').strftime('%d/%m/%Y')} au {datetime.strptime(rapport_data['date_fin'], '%Y-%m-%d').strftime('%d/%m/%Y')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Styles pour les différentes actions - Palette de couleurs complète (définir AVANT la légende)
        # Utiliser des couleurs plus visibles avec fill_type="solid"
        fill_rouge_fonce = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")  # Rouge foncé : Ne plus rappeler
        fill_rouge_clair = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Rouge clair : Actions critiques
        fill_violet = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")  # Violet : Pas de commande, n'appel plus
        fill_orange = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # Orange : Rappel plus tard, rappel demain
        fill_jaune = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")  # Jaune : En attente, à confirmer
        fill_vert_clair = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Vert clair : Commande confirmée, commande passée
        fill_bleu_clair = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Bleu clair : Client intéressé, à suivre
        fill_vert_fonce = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Vert foncé : Commande livrée, terminé
        fill_rose = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")  # Rose : Informations, notes
        fill_blanc = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanc : Par défaut
        
        row_num = 3
        if rapport_data.get('zone'):
            ws.merge_cells(f'A{row_num}:H{row_num}')
            ws[f'A{row_num}'] = f"Zone : {rapport_data['zone']}"
            ws[f'A{row_num}'].alignment = Alignment(horizontal='center')
            row_num += 1
        
        # Ajouter une légende des couleurs
        ws.merge_cells(f'A{row_num}:H{row_num}')
        ws[f'A{row_num}'] = "LÉGENDE DES COULEURS :"
        ws[f'A{row_num}'].font = Font(bold=True, size=10, name='Calibri')
        ws[f'A{row_num}'].alignment = Alignment(horizontal='left', vertical='center')
        row_num += 1
        
        # Ligne de légende avec les couleurs (4 couleurs par ligne sur 2 lignes)
        legend_items = [
            ('Rouge foncé', fill_rouge_fonce, 'Ne plus rappeler'),
            ('Rouge clair', fill_rouge_clair, 'Refusé/Annulé'),
            ('Violet', fill_violet, 'Pas de commande'),
            ('Orange', fill_orange, 'Rappel programmé'),
            ('Jaune', fill_jaune, 'En attente'),
            ('Vert clair', fill_vert_clair, 'Commande confirmée'),
            ('Vert foncé', fill_vert_fonce, 'Commande livrée'),
            ('Bleu clair', fill_bleu_clair, 'Client intéressé'),
        ]
        
        # Créer deux lignes pour la légende (4 couleurs par ligne)
        for line_idx in range(2):
            for i in range(4):
                idx = line_idx * 4 + i
                if idx < len(legend_items):
                    label, color_fill, description = legend_items[idx]
                    # Utiliser les colonnes A-B, C-D, E-F, G-H pour chaque ligne
                    col_label = 1 + i * 2  # A, C, E, G
                    col_desc = 2 + i * 2   # B, D, F, H
                    
                    # Cellule avec couleur et label
                    cell = ws.cell(row=row_num, column=col_label)
                    cell.fill = color_fill
                    cell.border = border
                    cell.value = label
                    cell.font = Font(size=9, name='Calibri', bold=True)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Cellule avec description
                    cell_desc = ws.cell(row=row_num, column=col_desc)
                    cell_desc.value = f": {description}"
                    cell_desc.font = Font(size=9, name='Calibri', italic=True)
                    cell_desc.alignment = Alignment(horizontal='left', vertical='center')
            row_num += 1
        
        start_row = row_num + 1
        
        # Définir les plages horaires
        PLAGES_HORAIRES = [
            ('06h30-08h30', '06h30-08h30'),
            ('08h30-10h00', '08h30-10h00'),
            ('10h00-11h30', '10h00-11h30'),
            ('14h30-16h00', '14h30-16h00'),
            ('16h00-17h30', '16h00-17h30'),
        ]
        
        # En-têtes de colonnes avec meilleur style
        # Première ligne : colonnes fixes + plages horaires fusionnées
        headers_fixed = ['DETAIL', 'REF', 'POTENTIEL /5', 'NOMS', 'TELEPHONES']
        col = 1
        for header in headers_fixed:
            ws.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col)}{start_row + 1}')
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = header_font_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            col += 1
        
        # Pour chaque plage horaire : 3 colonnes (Heure, Commandes, Total)
        for plage_code, plage_label in PLAGES_HORAIRES:
            # Fusionner les 3 colonnes pour le titre de la plage
            ws.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col + 2)}{start_row}')
            cell = ws.cell(row=start_row, column=col)
            cell.value = plage_label
            cell.font = header_font_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # Sous-colonnes : Heure, Commandes, Total
            sub_headers = ['Heure', 'Commandes', 'Total']
            for sub_header in sub_headers:
                cell = ws.cell(row=start_row + 1, column=col)
                cell.value = sub_header
                cell.font = header_font_bold
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                col += 1
        
        # Colonne Total général
        ws.merge_cells(f'{get_column_letter(col)}{start_row}:{get_column_letter(col)}{start_row + 1}')
        cell = ws.cell(row=start_row, column=col)
        cell.value = 'TOTAL'
        cell.font = header_font_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        
        # Hauteur de ligne pour les en-têtes
        ws.row_dimensions[start_row].height = 25
        ws.row_dimensions[start_row + 1].height = 25
        
        # Ajuster start_row pour les données (après les 2 lignes d'en-têtes)
        start_row = start_row + 2
        
        # Récupérer les données des clients avec la nouvelle structure
        clients_data = rapport_data.get('clients_data', [])
        print(f"[DEBUG EXPORT] Nombre de clients à exporter: {len(clients_data)}")
        
        # Créer une structure pour l'export avec les plages horaires
        clients_export = []
        for client_data in clients_data:
            clients_export.append({
                'id': client_data['id'],
                'detail': client_data.get('detail', ''),
                'ref': client_data.get('ref', ''),
                'potentiel': client_data.get('potentiel', ''),
                'nom': client_data.get('nom', ''),
                'telephone': client_data.get('telephone', ''),
                'actions_par_plage': client_data.get('actions_par_plage', {}),
                'commandes_par_plage': client_data.get('commandes_par_plage', {}),
                'totaux_par_plage': client_data.get('totaux_par_plage', {}),
                'total': client_data.get('total', 0),
            })
        
        print(f"[DEBUG EXPORT] Clients export préparés: {len(clients_export)}")
        
        # Police pour les cellules de données
        data_font = Font(size=10, name='Calibri')
        
        # Fonction pour déterminer la couleur selon l'action
        def get_action_color(action_text):
            """Retourne la couleur appropriée selon le contenu de l'action"""
            if not action_text:
                return fill_blanc
            
            action_lower = action_text.lower().strip()
            
            # Rouge foncé : Actions critiques - Ne plus contacter
            if any(mot in action_lower for mot in ['ne plus rappeler', 'ne pas rappeler', 'ne plus contacter', 
                                                    'ne plus appeler', 'arrêter', 'arreter', 'stop', 'bloquer']):
                return fill_rouge_fonce
            
            # Rouge clair : Actions négatives importantes
            if any(mot in action_lower for mot in ['refusé', 'refuse', 'annulé', 'annule', 'pas intéressé', 
                                                     'pas interesse', 'désabonné', 'desabonne', 'refuse']):
                return fill_rouge_clair
            
            # Violet : Pas de commande, n'appel plus
            if any(mot in action_lower for mot in ['pas de commande', 'n\'appel plus', 'n\'appelle plus', 
                                                     'n\'appel plus le matin', 'ne commande plus', 'ne commandera pas',
                                                     'n\'appel plus', 'nappel plus']):
                return fill_violet
            
            # Orange : Rappels programmés et appels non aboutis (à rappeler)
            if any(mot in action_lower for mot in ['rappel plus tard', 'rappel demain', 'rappel dans', 
                                                    'rappel lundi', 'rappel mardi', 'rappel mercredi', 
                                                    'rappel jeudi', 'rappel vendredi', 'rappel samedi', 
                                                    'rappel dimanche', 'rappel la semaine', 'rappel le mois',
                                                    'ne decroche pas', 'ne décroche pas', 'ne decroche', 
                                                    'pas de reponse', 'pas de réponse', 'pas reponse',
                                                    'ne repond pas', 'ne répond pas', 'ne repond', 
                                                    'sonne occupé', 'sonne occupe', 'ligne occupee',
                                                    'ligne occupée', 'pas disponible', 'indisponible']):
                return fill_orange
            
            # Jaune : En attente, à confirmer, en déplacement
            if any(mot in action_lower for mot in ['en attente', 'à confirmer', 'a confirmer', 'en cours', 
                                                    'en discussion', 'en négociation', 'en negociation', 
                                                    'à vérifier', 'a verifier', 'en suspens',
                                                    'en deplacement', 'en déplacement', 'deplacement',
                                                    'déplacement', 'absent', 'pas la', 'pas present',
                                                    'pas présent', 'en reunion', 'en réunion']):
                return fill_jaune
            
            # Vert clair : Commandes confirmées
            if any(mot in action_lower for mot in ['commande confirmée', 'commande confirmee', 'commande passée', 
                                                    'commande passee', 'commande validée', 'commande validee', 
                                                    'commande acceptée', 'commande acceptee', 'commande prise']):
                return fill_vert_clair
            
            # Vert foncé : Commandes livrées, terminées
            if any(mot in action_lower for mot in ['commande livrée', 'commande livree', 'livré', 'livre', 
                                                    'terminé', 'termine', 'finalisé', 'finalise', 'complété', 
                                                    'complete', 'fini', 'achevé', 'acheve']):
                return fill_vert_fonce
            
            # Bleu clair : Client intéressé, à suivre, disponible
            if any(mot in action_lower for mot in ['client intéressé', 'client interesse', 'intéressé', 'interesse', 
                                                    'à suivre', 'a suivre', 'suivi', 'prospect', 'potentiel', 
                                                    'envisage', 'réfléchit', 'reflechit', 'considère', 'considere',
                                                    'disponible', 'libre', 'ok', 'd\'accord', 'daccord']):
                return fill_bleu_clair
            
            # Rose : Informations, notes générales
            if any(mot in action_lower for mot in ['note', 'information', 'info', 'remarque', 'commentaire', 
                                                     'observation', 'précision', 'precision', 'détail', 'detail']):
                return fill_rose
            
            # Par défaut : Si l'action contient des mots qui suggèrent une commande (articles, quantités)
            # On met en vert clair car c'est probablement une commande
            if any(mot in action_lower for mot in ['qté', 'qte', 'quantité', 'quantite', 'kg', 'unité', 'unite']):
                return fill_vert_clair
            
            # Blanc par défaut pour les autres cas
            return fill_blanc
        
        # Données - UNE SEULE LIGNE PAR CLIENT avec toutes les plages horaires
        row = start_row + 1
        for client in clients_export:
            try:
                # Colonnes fixes
                ws.cell(row=row, column=1, value=client['detail']).border = border
                ws.cell(row=row, column=1).fill = fill_blanc
                ws.cell(row=row, column=1).font = data_font
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                
                ws.cell(row=row, column=2, value=client['ref']).border = border
                ws.cell(row=row, column=2).fill = fill_blanc
                ws.cell(row=row, column=2).font = data_font
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')
                
                ws.cell(row=row, column=3, value=client['potentiel']).border = border
                ws.cell(row=row, column=3).fill = fill_blanc
                ws.cell(row=row, column=3).font = data_font
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='center')
                
                ws.cell(row=row, column=4, value=client['nom']).border = border
                ws.cell(row=row, column=4).fill = fill_blanc
                ws.cell(row=row, column=4).font = Font(bold=True, size=10, name='Calibri')
                ws.cell(row=row, column=4).alignment = Alignment(horizontal='left', vertical='center')
                
                ws.cell(row=row, column=5, value=client['telephone']).border = border
                ws.cell(row=row, column=5).fill = fill_blanc
                ws.cell(row=row, column=5).font = data_font
                ws.cell(row=row, column=5).alignment = Alignment(horizontal='left', vertical='center')
                
                # Pour chaque plage horaire, créer 3 colonnes : Heure, Commandes, Total
                col = 6
                actions_par_plage = client.get('actions_par_plage', {})
                commandes_par_plage = client.get('commandes_par_plage', {})
                totaux_par_plage = client.get('totaux_par_plage', {})
                
                for plage_code, plage_label in PLAGES_HORAIRES:
                    # Récupérer les actions et commandes pour cette plage
                    actions_list = actions_par_plage.get(plage_code, [])
                    commandes_list = commandes_par_plage.get(plage_code, [])
                    
                    # Colonne Heure : concaténer toutes les heures (actions + commandes)
                    heures_set = set()
                    for action in actions_list:
                        if action.get('heure_appel'):
                            heures_set.add(action['heure_appel'])
                    for cmd in commandes_list:
                        if cmd.get('heure'):
                            heures_set.add(cmd['heure'])
                    heures_str = ', '.join(sorted(heures_set)) if heures_set else ''
                    
                    # Colonne Commandes : concaténer toutes les actions et commandes
                    commandes_str_parts = []
                    for action in actions_list:
                        if action.get('action'):
                            commandes_str_parts.append(action['action'])
                    for cmd in commandes_list:
                        if cmd.get('articles'):
                            commandes_str_parts.append(cmd['articles'])
                    commandes_str = ' | '.join(commandes_str_parts) if commandes_str_parts else ''
                    
                    # Déterminer la couleur selon les actions
                    row_fill = fill_blanc
                    for action in actions_list:
                        if action.get('action'):
                            action_fill = get_action_color(action['action'])
                            if action_fill != fill_blanc:
                                row_fill = action_fill
                                break
                    
                    # Colonne Heure
                    ws.cell(row=row, column=col, value=heures_str).border = border
                    ws.cell(row=row, column=col).fill = row_fill
                    ws.cell(row=row, column=col).font = data_font
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    col += 1
                    
                    # Colonne Commandes
                    ws.cell(row=row, column=col, value=commandes_str).border = border
                    ws.cell(row=row, column=col).fill = row_fill
                    ws.cell(row=row, column=col).font = data_font
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    col += 1
                    
                    # Colonne Total pour cette plage
                    total_plage = totaux_par_plage.get(plage_code, 0)
                    ws.cell(row=row, column=col, value=float(total_plage) if total_plage > 0 else '').border = border
                    ws.cell(row=row, column=col).fill = row_fill
                    ws.cell(row=row, column=col).font = data_font
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')
                    col += 1
                
                # Colonne Total général
                total_client = client.get('total', 0)
                ws.cell(row=row, column=col, value=float(total_client) if total_client > 0 else '').border = border
                ws.cell(row=row, column=col).fill = fill_blanc
                ws.cell(row=row, column=col).font = Font(bold=True, size=10, name='Calibri')
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')
                
                row += 1
                
            except Exception as client_error:
                print(f"[DEBUG EXPORT] Erreur client {client.get('id', 'N/A')}: {str(client_error)}")
                import traceback
                traceback.print_exc()
                continue
        
        # Ajouter les statistiques par vague
        # Utiliser les mêmes plages horaires que dans les données
        # Calculer les statistiques par vague
        stats_par_vague = {}
        for plage_code, plage_label in PLAGES_HORAIRES:
            stats_par_vague[plage_code] = {
                'nb_appels': 0,
                'nb_reponses': 0,
                'nb_commandes': 0,
                'total_commandes': 0.0
            }
        
        # Parcourir les clients pour calculer les stats
        for client in clients_export:
            actions_par_plage = client.get('actions_par_plage', {})
            commandes_par_plage = client.get('commandes_par_plage', {})
            totaux_par_plage = client.get('totaux_par_plage', {})
            
            # Compter les appels et réponses par plage horaire
            for plage_code, plage_label in PLAGES_HORAIRES:
                actions_list = actions_par_plage.get(plage_code, [])
                commandes_list = commandes_par_plage.get(plage_code, [])
                
                # Compter les appels
                # 1. Compter toutes les actions enregistrées (chaque action = un appel)
                nb_appels_actions = len(actions_list)
                
                # 2. Compter toutes les commandes (une commande = un appel, car il faut appeler pour passer une commande)
                nb_appels_commandes = len(commandes_list)
                
                # 3. Total des appels = actions + commandes
                # MAIS : si une action et une commande sont liées (même client, même plage), on ne compte qu'un seul appel
                # Pour simplifier, on prend le maximum entre actions et commandes, ou la somme si elles sont indépendantes
                # Ici, on va compter : max(actions, commandes) pour éviter de compter deux fois le même appel
                # Mais si on a des actions ET des commandes séparées, on les additionne
                # Solution : compter toutes les actions + les commandes qui n'ont pas d'action associée
                nb_appels = nb_appels_actions
                if commandes_list:
                    # Si on a des commandes mais pas d'actions, on compte les commandes comme appels
                    if not actions_list:
                        nb_appels = nb_appels_commandes
                    # Si on a des actions ET des commandes, on compte les deux (car ce sont des appels différents)
                    else:
                        nb_appels = nb_appels_actions + nb_appels_commandes
                
                stats_par_vague[plage_code]['nb_appels'] += nb_appels
                
                # Calculer le nombre de réponses selon la logique :
                # Nombre de réponses = (Actions avec réponse + Commandes) - (Actions "pas de réponse")
                
                # 1. Compter les actions qui sont des réponses (pas "pas de réponse")
                nb_reponses_actions = 0
                nb_pas_de_reponse = 0
                for action in actions_list:
                    action_text = (action.get('action') or '').lower().strip()
                    # Vérifier si c'est "pas de réponse"
                    if any(mot in action_text for mot in [
                        'pas de reponse', 'pas de réponse', 'pas reponse', 'pas réponse',
                        'ne decroche pas', 'ne décroche pas', 'ne decroche', 'ne décroche',
                        'ne repond pas', 'ne répond pas', 'ne repond', 'ne répond',
                        'sonne occupé', 'sonne occupe', 'ligne occupee', 'ligne occupée',
                        'pas disponible', 'indisponible'
                    ]):
                        nb_pas_de_reponse += 1
                    elif action_text:  # Si l'action a du texte et n'est pas "pas de réponse", c'est une réponse
                        nb_reponses_actions += 1
                    # Si l'action n'a pas de texte mais a une heure, on ne la compte pas comme réponse
                
                # 2. Compter toutes les commandes (une commande = une réponse)
                nb_reponses_commandes = len(commandes_list)
                
                # 3. Calculer le nombre total de réponses
                # Nombre de réponses = Actions avec réponse + Commandes
                nb_reponses = nb_reponses_actions + nb_reponses_commandes
                
                # IMPORTANT : Le nombre de réponses ne peut pas dépasser le nombre d'appels
                if nb_reponses > nb_appels:
                    # Si on a plus de réponses que d'appels, on limite au nombre d'appels
                    nb_reponses = nb_appels
                
                if nb_reponses > 0:
                    stats_par_vague[plage_code]['nb_reponses'] += nb_reponses
                
                # Compter les commandes par plage horaire
                if commandes_list:
                    stats_par_vague[plage_code]['nb_commandes'] += len(commandes_list)
                    # Ajouter le total des commandes de cette plage pour ce client
                    total_plage_client = float(totaux_par_plage.get(plage_code, 0))
                    stats_par_vague[plage_code]['total_commandes'] += total_plage_client
        
        # Ajouter le tableau des statistiques
        stats_row = row + 2
        stats_start_col = 1
        
        # Titre du tableau de statistiques - Première ligne
        # Colonne A avec "RAPPORT JOURNALIER-DATE" qui span 2 lignes (rowspan=2)
        # Écrire AVANT de fusionner pour que le texte soit visible
        cell_title = ws.cell(row=stats_row, column=stats_start_col)
        cell_title.value = 'RAPPORT JOURNALIER-DATE'
        cell_title.font = Font(bold=True, size=12, name='Calibri')
        cell_title.fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        cell_title.alignment = Alignment(horizontal='left', vertical='center')
        cell_title.border = border
        ws.merge_cells(f'{get_column_letter(stats_start_col)}{stats_row}:{get_column_letter(stats_start_col)}{stats_row + 1}')
        
        # En-têtes des vagues - Première ligne (fusionnés verticalement sur 2 lignes)
        col = stats_start_col + 1
        vague_names = ['PREMIERE', 'DEUXIEME VAGUE', 'TROISIEME VAGUE', 'QUATRIEME VAGUE', 'CINQUIEME VAGUE']
        for idx, (plage_code, plage_label) in enumerate(PLAGES_HORAIRES):
            vague_name = vague_names[idx] if idx < len(vague_names) else plage_label
            # Écrire AVANT de fusionner pour que le texte soit visible
            cell = ws.cell(row=stats_row, column=col)
            cell.value = vague_name
            cell.font = header_font_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.merge_cells(f'{get_column_letter(col)}{stats_row}:{get_column_letter(col)}{stats_row + 1}')
            col += 1
        
        # Colonne Total (fusionnée verticalement sur 2 lignes)
        # Écrire AVANT de fusionner pour que le texte soit visible
        cell = ws.cell(row=stats_row, column=col)
        cell.value = 'TOTAL COMMANDES JOURNALIERES'
        cell.font = header_font_bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.merge_cells(f'{get_column_letter(col)}{stats_row}:{get_column_letter(col)}{stats_row + 1}')
        
        # Lignes de statistiques - Style exact comme l'image
        # "Nombre de Comm" : colonne A violet clair ET toutes les cellules de données violet clair
        # "Total comman" : colonne A rose clair ET toutes les cellules de données rose clair
        metrics = [
            ('Nombre d\'app', 'nb_appels', 'blanc', 'blanc'),  # (label, key, couleur_col_A, couleur_cells)
            ('Nombre de rép', 'nb_reponses', 'blanc', 'blanc'),
            ('Nombre de Comm', 'nb_commandes', 'violet', 'violet'),  # Violet clair pour colonne A ET toutes les cellules
            ('Total comman', 'total_commandes', 'rose', 'rose'),  # Rose clair pour colonne A ET toutes les cellules
        ]
        
        fill_rose = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Rose clair
        fill_violet = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")  # Violet clair
        fill_blanc = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        stats_data_row = stats_row + 2  # Ligne pour les données (après l'en-tête fusionné)
        
        for metric_label, metric_key, couleur_col_a, couleur_cells in metrics:
            col = stats_start_col
            # Colonne A avec le label - gérer les cellules fusionnées
            cell = ws.cell(row=stats_data_row, column=col)
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                # Utiliser la cellule principale (première ligne de la fusion)
                cell = ws.cell(row=stats_row, column=col)
            cell.value = metric_label
            cell.font = Font(bold=True, size=10, name='Calibri')
            if couleur_col_a == 'rose':
                cell.fill = fill_rose
            elif couleur_col_a == 'violet':
                cell.fill = fill_violet
            else:
                cell.fill = fill_blanc
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            col += 1
            
            # Une colonne par vague
            for plage_code, plage_label in PLAGES_HORAIRES:
                value = stats_par_vague[plage_code][metric_key]
                if metric_key == 'total_commandes':
                    # Format avec point comme séparateur de milliers (ex: 32.800)
                    if value > 0:
                        value_str = f"{value:,.0f}".replace(',', '.')
                    else:
                        value_str = ""
                else:
                    value_str = str(int(value)) if value > 0 else ""
                
                # Cellule de données pour cette vague - gérer les cellules fusionnées
                cell_data = ws.cell(row=stats_data_row, column=col)
                from openpyxl.cell.cell import MergedCell
                if isinstance(cell_data, MergedCell):
                    # Utiliser la cellule principale (première ligne de la fusion)
                    cell_data = ws.cell(row=stats_row, column=col)
                cell_data.value = value_str
                cell_data.border = border
                cell_data.font = data_font
                if metric_key == 'total_commandes':
                    cell_data.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell_data.alignment = Alignment(horizontal='center', vertical='center')
                # Appliquer la couleur selon la métrique
                if couleur_cells == 'rose':
                    cell_data.fill = fill_rose
                elif couleur_cells == 'violet':
                    cell_data.fill = fill_violet
                else:
                    cell_data.fill = fill_blanc
                col += 1
            
            # Total journalier (dans la colonne TOTAL COMMANDES JOURNALIERES)
            # Cette colonne affiche UNIQUEMENT le chiffre d'affaire total (total_commandes)
            # Pour les autres métriques (nb_appels, nb_reponses, nb_commandes), cette colonne reste vide
            if metric_key == 'total_commandes':
                # Somme de tous les total_commandes de toutes les vagues = chiffre d'affaire total
                total_journalier = sum(stats_par_vague[p[0]]['total_commandes'] for p in PLAGES_HORAIRES)
                if total_journalier > 0:
                    # Format avec point comme séparateur de milliers (ex: 1.154.150)
                    value_str = f"{total_journalier:,.0f}".replace(',', '.')
                else:
                    value_str = ""
            else:
                # Pour les autres métriques, la colonne TOTAL COMMANDES JOURNALIERES reste vide
                value_str = ""
            
            cell_total = ws.cell(row=stats_data_row, column=col)
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell_total, MergedCell):
                # Utiliser la cellule principale (première ligne de la fusion)
                cell_total = ws.cell(row=stats_row, column=col)
            cell_total.value = value_str
            cell_total.border = border
            cell_total.font = Font(bold=True, size=10, name='Calibri')
            if metric_key == 'total_commandes':
                cell_total.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell_total.alignment = Alignment(horizontal='center', vertical='center')
            # Appliquer la couleur selon la métrique
            if couleur_cells == 'rose':
                cell_total.fill = fill_rose
            elif couleur_cells == 'violet':
                cell_total.fill = fill_violet
            else:
                cell_total.fill = fill_blanc
            
            stats_data_row += 1
        
        # Ajuster les largeurs de colonnes pour un meilleur affichage
        # Colonnes fixes : DETAIL, REF, POTENTIEL, NOMS, TELEPHONES
        column_widths = [30, 18, 15, 25, 18]
        # Pour chaque plage horaire : Heure, Commandes, Total
        for _ in PLAGES_HORAIRES:
            column_widths.extend([15, 40, 15])
        # Colonne Total général
        column_widths.append(18)
        
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Ajuster la hauteur des lignes de données
        for r in range(start_row + 1, row + 1):
            ws.row_dimensions[r].height = 20
        
        # Réponse HTTP avec le fichier Excel
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="rapport_acm_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        print(f"[ERREUR EXPORT ACM] {str(e)}")
        print(f"[TRACEBACK] {error_traceback}")
        messages.error(request, f'Erreur lors de l\'export: {str(e)}')
        return redirect('rapport_acm')

@login_required
def logout_stock(request):
    """Vue de logout pour le module stock"""
    logout(request)
    return redirect('index')

@login_required
def modifier_client(request, client_id):
    """Vue pour modifier un client existant"""
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([intitule, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_client', client_id=client_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_client', client_id=client_id)
            
            # Mettre à jour le client
            client.intitule = intitule
            client.adresse = adresse
            client.telephone = telephone
            client.email = email
            client.agence = agence
            client.save()
            
            messages.success(request, f'Client "{intitule}" modifié avec succès!')
            return redirect('consulter_clients')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du client: {str(e)}')
            return redirect('modifier_client', client_id=client_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    
    context = {
        'client': client,
        'agences': agences,
    }
    return render(request, 'supermarket/stock/modifier_client.html', context)

@login_required
def supprimer_client(request, client_id):
    """Vue pour supprimer un client"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_clients')
    
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
        client_name = client.intitule
        client.delete()
        messages.success(request, f'Client "{client_name}" supprimé avec succès!')
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_clients')
@login_required
def detail_client(request, client_id):
    """Vue pour afficher les détails d'un client"""
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
        
        # Récupérer les factures du client (si elles existent)
        factures = FactureVente.objects.filter(client=client).order_by('-date', '-heure')[:10]
        
        context = {
            'client': client,
            'factures': factures,
        }
        return render(request, 'supermarket/stock/detail_client.html', context)
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients')

# ==================== PLAN COMPTABLE ====================

@login_required
def consulter_plan_comptable(request):
    """Vue pour consulter le plan comptable"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    nature_filter = request.GET.get('nature_filter', '')
    
    # Construire la requête de base
    comptes = PlanComptable.objects.all()
    
    # Appliquer les filtres
    if search_query:
        comptes = comptes.filter(
            Q(intitule__icontains=search_query) |
            Q(compte__icontains=search_query) |
            Q(abrege__icontains=search_query)
        )
    
    if nature_filter:
        comptes = comptes.filter(nature_compte=nature_filter)
    
    # Trier par numéro
    comptes = comptes.order_by('numero')
    
    # Calculer les statistiques
    total_comptes = PlanComptable.objects.count()
    comptes_actifs = PlanComptable.objects.filter(actif=True).count()
    
    # Récupérer les natures de compte pour le filtre
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'comptes': comptes,
        'agence': agence,
        'total_comptes': total_comptes,
        'comptes_actifs': comptes_actifs,
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/consulter_plan_comptable.html', context)

@login_required
def creer_plan_comptable(request):
    """Vue pour créer un nouveau compte comptable"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero = request.POST.get('numero')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            abrege = request.POST.get('abrege')
            nature_compte = request.POST.get('nature_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([numero, intitule, compte]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_plan_comptable')
            
            # Créer le compte comptable
            PlanComptable.objects.create(
                numero=numero,
                intitule=intitule,
                compte=compte,
                abrege=abrege,
                nature_compte=nature_compte
            )
            
            messages.success(request, f'Compte comptable "{intitule}" créé avec succès!')
            return redirect('consulter_plan_comptable')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du compte: {str(e)}')
            return redirect('creer_plan_comptable')
    
    # GET - Afficher le formulaire
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/creer_plan_comptable.html', context)

@login_required
def modifier_plan_comptable(request, compte_id):
    """Vue pour modifier un compte comptable existant"""
    try:
        compte = PlanComptable.objects.get(id=compte_id)
    except PlanComptable.DoesNotExist:
        messages.error(request, 'Compte comptable non trouvé.')
        return redirect('consulter_plan_comptable')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero = request.POST.get('numero')
            intitule = request.POST.get('intitule')
            compte_field = request.POST.get('compte')
            abrege = request.POST.get('abrege')
            nature_compte = request.POST.get('nature_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([numero, intitule, compte_field]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_plan_comptable', compte_id=compte_id)
            
            # Mettre à jour le compte
            compte.numero = numero
            compte.intitule = intitule
            compte.compte = compte_field
            compte.abrege = abrege
            compte.nature_compte = nature_compte
            compte.save()
            
            messages.success(request, f'Compte comptable "{intitule}" modifié avec succès!')
            return redirect('consulter_plan_comptable')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du compte: {str(e)}')
            return redirect('modifier_plan_comptable', compte_id=compte_id)
    
    # GET - Afficher le formulaire pré-rempli
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'compte': compte,
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/modifier_plan_comptable.html', context)

@login_required
def supprimer_plan_comptable(request, compte_id):
    """Vue pour supprimer un compte comptable"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_plan_comptable')
    
    try:
        compte = PlanComptable.objects.get(id=compte_id)
        compte_name = compte.intitule
        compte.delete()
        messages.success(request, f'Compte comptable "{compte_name}" supprimé avec succès!')
    except PlanComptable.DoesNotExist:
        messages.error(request, 'Compte comptable non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_plan_comptable')

# ==================== PLAN TIERS ====================

@login_required
def consulter_plan_tiers(request):
    """Vue pour consulter le plan tiers"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    tiers = PlanTiers.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        tiers = tiers.filter(
            Q(intitule_compte__icontains=search_query) |
            Q(numero_compte__icontains=search_query)
        )
    
    if type_filter:
        tiers = tiers.filter(type=type_filter)
    
    # Trier par numéro de compte
    tiers = tiers.order_by('numero_compte')
    
    # Calculer les statistiques
    total_tiers = PlanTiers.objects.filter(agence=agence).count()
    clients_count = PlanTiers.objects.filter(agence=agence, type='client').count()
    fournisseurs_count = PlanTiers.objects.filter(agence=agence, type='fournisseur').count()
    
    # Récupérer les types pour le filtre
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'tiers': tiers,
        'agence': agence,
        'total_tiers': total_tiers,
        'clients_count': clients_count,
        'fournisseurs_count': fournisseurs_count,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_plan_tiers.html', context)

@login_required
def creer_plan_tiers(request):
    """Vue pour créer un nouveau tiers"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_tiers = request.POST.get('type')
            numero_compte = request.POST.get('numero_compte')
            intitule_compte = request.POST.get('intitule_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_tiers, numero_compte, intitule_compte, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_plan_tiers')
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('creer_plan_tiers')
            
            # Créer le tiers
            PlanTiers.objects.create(
                type=type_tiers,
                numero_compte=numero_compte,
                intitule_compte=intitule_compte,
                agence=agence
            )
            
            messages.success(request, f'Tiers "{intitule_compte}" créé avec succès!')
            return redirect('consulter_plan_tiers')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du tiers: {str(e)}')
            return redirect('creer_plan_tiers')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'agences': agences,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/creer_plan_tiers.html', context)

@login_required
def modifier_plan_tiers(request, tiers_id):
    """Vue pour modifier un tiers existant"""
    try:
        tiers = PlanTiers.objects.get(id=tiers_id, agence=get_user_agence(request))
    except PlanTiers.DoesNotExist:
        messages.error(request, 'Tiers non trouvé.')
        return redirect('consulter_plan_tiers')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_tiers = request.POST.get('type')
            numero_compte = request.POST.get('numero_compte')
            intitule_compte = request.POST.get('intitule_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_tiers, numero_compte, intitule_compte, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_plan_tiers', tiers_id=tiers_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_plan_tiers', tiers_id=tiers_id)
            
            # Mettre à jour le tiers
            tiers.type = type_tiers
            tiers.numero_compte = numero_compte
            tiers.intitule_compte = intitule_compte
            tiers.agence = agence
            tiers.save()
            
            messages.success(request, f'Tiers "{intitule_compte}" modifié avec succès!')
            return redirect('consulter_plan_tiers')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du tiers: {str(e)}')
            return redirect('modifier_plan_tiers', tiers_id=tiers_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'tiers': tiers,
        'agences': agences,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/modifier_plan_tiers.html', context)

@login_required
def supprimer_plan_tiers(request, tiers_id):
    """Vue pour supprimer un tiers"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_plan_tiers')
    
    try:
        tiers = PlanTiers.objects.get(id=tiers_id, agence=get_user_agence(request))
        tiers_name = tiers.intitule_compte
        tiers.delete()
        messages.success(request, f'Tiers "{tiers_name}" supprimé avec succès!')
    except PlanTiers.DoesNotExist:
        messages.error(request, 'Tiers non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_plan_tiers')

# ==================== CODE JOURNAUX ====================

@login_required
def consulter_code_journaux(request):
    """Vue pour consulter les codes journaux"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    journaux = CodeJournaux.objects.all()
    
    # Appliquer les filtres
    if search_query:
        journaux = journaux.filter(
            Q(intitule__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    if type_filter:
        journaux = journaux.filter(type_document=type_filter)
    
    # Trier par code
    journaux = journaux.order_by('code')
    
    # Calculer les statistiques
    total_journaux = CodeJournaux.objects.count()
    journaux_achat = CodeJournaux.objects.filter(type_document='document_achat').count()
    journaux_vente = CodeJournaux.objects.filter(type_document='caisse').count()
    
    # Récupérer les types pour le filtre
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'journaux': journaux,
        'agence': agence,
        'total_journaux': total_journaux,
        'journaux_achat': journaux_achat,
        'journaux_vente': journaux_vente,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_code_journaux.html', context)

@login_required
def creer_code_journaux(request):
    """Vue pour créer un nouveau code journal"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_document = request.POST.get('type_document')
            intitule = request.POST.get('intitule')
            code = request.POST.get('code')
            compte_contrepartie = request.POST.get('compte_contrepartie')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_document, intitule, code]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_code_journaux')
            
            # Créer le code journal
            CodeJournaux.objects.create(
                type_document=type_document,
                intitule=intitule,
                code=code,
                compte_contrepartie=compte_contrepartie
            )
            
            messages.success(request, f'Code journal "{intitule}" créé avec succès!')
            return redirect('consulter_code_journaux')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du code journal: {str(e)}')
            return redirect('creer_code_journaux')
    
    # GET - Afficher le formulaire
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/creer_code_journaux.html', context)

@login_required
def modifier_code_journaux(request, journal_id):
    """Vue pour modifier un code journal existant"""
    try:
        journal = CodeJournaux.objects.get(id=journal_id)
    except CodeJournaux.DoesNotExist:
        messages.error(request, 'Code journal non trouvé.')
        return redirect('consulter_code_journaux')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_document = request.POST.get('type_document')
            intitule = request.POST.get('intitule')
            code = request.POST.get('code')
            compte_contrepartie = request.POST.get('compte_contrepartie')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_document, intitule, code]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_code_journaux', journal_id=journal_id)
            
            # Mettre à jour le code journal
            journal.type_document = type_document
            journal.intitule = intitule
            journal.code = code
            journal.compte_contrepartie = compte_contrepartie
            journal.save()
            
            messages.success(request, f'Code journal "{intitule}" modifié avec succès!')
            return redirect('consulter_code_journaux')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du code journal: {str(e)}')
            return redirect('modifier_code_journaux', journal_id=journal_id)
    
    # GET - Afficher le formulaire pré-rempli
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'journal': journal,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/modifier_code_journaux.html', context)

@login_required
def supprimer_code_journaux(request, journal_id):
    """Vue pour supprimer un code journal"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_code_journaux')
    
    try:
        journal = CodeJournaux.objects.get(id=journal_id)
        journal_name = journal.intitule
        journal.delete()
        messages.success(request, f'Code journal "{journal_name}" supprimé avec succès!')
    except CodeJournaux.DoesNotExist:
        messages.error(request, 'Code journal non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_code_journaux')

# ==================== TAUX TAXE ====================

@login_required
def consulter_taux_taxe(request):
    """Vue pour consulter les taux de taxe"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    taux = TauxTaxe.objects.all()
    
    # Appliquer les filtres
    if search_query:
        taux = taux.filter(
            Q(intitule__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    if type_filter:
        taux = taux.filter(type=type_filter)
    
    # Trier par code
    taux = taux.order_by('code')
    
    # Calculer les statistiques
    total_taux = TauxTaxe.objects.count()
    taux_actifs = TauxTaxe.objects.filter(actif=True).count()
    
    # Récupérer les types pour le filtre
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    
    context = {
        'taux': taux,
        'agence': agence,
        'total_taux': total_taux,
        'taux_actifs': taux_actifs,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_taux_taxe.html', context)

@login_required
def creer_taux_taxe(request):
    """Vue pour créer un nouveau taux de taxe"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            code = request.POST.get('code')
            sens = request.POST.get('sens')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            taux = request.POST.get('taux')
            type_taxe = request.POST.get('type')
            assujettissement = request.POST.get('assujettissement')
            code_regroupement = request.POST.get('code_regroupement')
            provenance = request.POST.get('provenance')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([code, intitule, taux]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_taux_taxe')
            
            # Créer le taux de taxe
            TauxTaxe.objects.create(
                code=code,
                sens=sens,
                intitule=intitule,
                compte=compte,
                taux=float(taux),
                type=type_taxe,
                assujettissement=assujettissement,
                code_regroupement=code_regroupement,
                provenance=provenance
            )
            
            messages.success(request, f'Taux de taxe "{intitule}" créé avec succès!')
            return redirect('consulter_taux_taxe')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du taux de taxe: {str(e)}')
            return redirect('creer_taux_taxe')
    
    # GET - Afficher le formulaire
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    sens_choices = TauxTaxe.SENS_CHOICES
    assujettissement_choices = TauxTaxe.ASSUJETTISSEMENT_CHOICES
    
    context = {
        'type_choices': type_choices,
        'sens_choices': sens_choices,
        'assujettissement_choices': assujettissement_choices,
    }
    return render(request, 'supermarket/stock/creer_taux_taxe.html', context)

@login_required
def modifier_taux_taxe(request, taux_id):
    """Vue pour modifier un taux de taxe existant"""
    try:
        taux = TauxTaxe.objects.get(id=taux_id)
    except TauxTaxe.DoesNotExist:
        messages.error(request, 'Taux de taxe non trouvé.')
        return redirect('consulter_taux_taxe')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            code = request.POST.get('code')
            sens = request.POST.get('sens')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            taux_value = request.POST.get('taux')
            type_taxe = request.POST.get('type')
            assujettissement = request.POST.get('assujettissement')
            code_regroupement = request.POST.get('code_regroupement')
            provenance = request.POST.get('provenance')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([code, intitule, taux_value]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_taux_taxe', taux_id=taux_id)
            
            # Mettre à jour le taux de taxe
            taux.code = code
            taux.sens = sens
            taux.intitule = intitule
            taux.compte = compte
            taux.taux = float(taux_value)
            taux.type = type_taxe
            taux.assujettissement = assujettissement
            taux.code_regroupement = code_regroupement
            taux.provenance = provenance
            taux.save()
            
            messages.success(request, f'Taux de taxe "{intitule}" modifié avec succès!')
            return redirect('consulter_taux_taxe')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du taux de taxe: {str(e)}')
            return redirect('modifier_taux_taxe', taux_id=taux_id)
    
    # GET - Afficher le formulaire pré-rempli
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    sens_choices = TauxTaxe.SENS_CHOICES
    assujettissement_choices = TauxTaxe.ASSUJETTISSEMENT_CHOICES
    
    context = {
        'taux': taux,
        'type_choices': type_choices,
        'sens_choices': sens_choices,
        'assujettissement_choices': assujettissement_choices,
    }
    return render(request, 'supermarket/stock/modifier_taux_taxe.html', context)

@login_required
def supprimer_taux_taxe(request, taux_id):
    """Vue pour supprimer un taux de taxe"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_taux_taxe')
    
    try:
        taux = TauxTaxe.objects.get(id=taux_id)
        taux_name = taux.intitule
        taux.delete()
        messages.success(request, f'Taux de taxe "{taux_name}" supprimé avec succès!')
    except TauxTaxe.DoesNotExist:
        messages.error(request, 'Taux de taxe non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_taux_taxe')

@login_required
@require_stock_modify_access
def modifier_article(request, article_id):
    """Vue pour modifier un article existant"""
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
        return redirect('consulter_articles')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            designation = request.POST.get('designation')
            agence_id = request.POST.get('agence')
            prix_achat = request.POST.get('prix_achat')
            prix_vente = request.POST.get('prix_vente')
            stock_actuel = request.POST.get('stock_actuel')
            stock_minimum = request.POST.get('stock_minimum', 0)
            unite_vente = request.POST.get('unite_vente')
            conditionnement = request.POST.get('conditionnement')
            famille_id = request.POST.get('famille')
            # Le suivi de stock est toujours activé automatiquement (non modifiable par l'utilisateur)
            
            ancien_stock = article.stock_actuel
            
            # Validation (stock_actuel est optionnel)
            if not all([designation, agence_id, prix_achat, prix_vente, unite_vente]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_article', article_id=article_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_article', article_id=article_id)
            
            # Récupérer la famille si spécifiée
            if famille_id:
                try:
                    categorie = Famille.objects.get(id=famille_id)
                except Famille.DoesNotExist:
                    messages.error(request, 'Famille non trouvée.')
                    return redirect('modifier_article', article_id=article_id)
            else:
                # Si aucune famille n'est fournie, garder l'ancienne
                categorie = article.categorie
            
            # Mettre à jour l'article
            article.designation = designation
            article.agence = agence
            article.prix_achat = Decimal(str(prix_achat)).quantize(Decimal('0.01'))
            article.prix_vente = Decimal(str(prix_vente)).quantize(Decimal('0.01'))
            
            if stock_actuel and stock_actuel.strip():
                stock_value = stock_actuel.replace(',', '.')
                article.stock_actuel = Decimal(stock_value)
            
            article.stock_minimum = Decimal(str(stock_minimum)).quantize(Decimal('0.01')) if stock_minimum and stock_minimum.strip() else Decimal('0')
            article.unite_vente = unite_vente
            article.conditionnement = conditionnement
            article.categorie = categorie
            # Le suivi de stock est toujours activé automatiquement
            article.suivi_stock = True
            article.save()
            
            nouveau_stock = article.stock_actuel
            difference_stock = nouveau_stock - ancien_stock
            
            # Créer un mouvement pour TOUS les changements de stock (même si différence = 0 pour traçabilité complète)
            if article.suivi_stock:
                try:
                    numero_piece = f"AJUST-{article.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}"
                    MouvementStock.objects.create(
                        article=article,
                        agence=agence,
                        type_mouvement='ajustement',
                        date_mouvement=timezone.now(),
                        numero_piece=numero_piece,
                        quantite_stock=nouveau_stock,
                        stock_initial=ancien_stock,
                        solde=nouveau_stock,
                        quantite=abs(difference_stock) if difference_stock != Decimal('0') else Decimal('0'),
                        cout_moyen_pondere=float(article.prix_achat),
                        stock_permanent=float(nouveau_stock * article.prix_achat),
                        commentaire=f"Ajustement manuel via modification d'article par {request.user.username}" + (f" (différence: {difference_stock})" if difference_stock != Decimal('0') else " (vérification)")
                    )
                    print(f"[STOCK] Ajustement manuel enregistré: {article.designation} {ancien_stock} -> {nouveau_stock}")
                except Exception as movement_error:
                    print(f"[WARNING] Impossible d'enregistrer le mouvement d'ajustement: {movement_error}")
            
            # Mettre à jour les types de vente
            prix_gros = request.POST.get('prix_gros')
            prix_demi_gros = request.POST.get('prix_demi_gros')
            prix_detail = request.POST.get('prix_detail')
            
            if prix_gros:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Gros',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_gros)}
                )
            
            if prix_demi_gros:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Demi-Gros',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_demi_gros)}
                )
            
            if prix_detail:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Détail',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_detail)}
                )
            
            messages.success(request, f'Article "{designation}" modifié avec succès!')
            return redirect('consulter_articles')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de l\'article: {str(e)}')
            return redirect('modifier_article', article_id=article_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    familles = Famille.objects.all()
    types_vente = TypeVente.objects.filter(article=article)
    
    # Créer un dictionnaire des types de vente avec des clés sans traits d'union
    types_vente_dict = {}
    for tv in types_vente:
        if tv.intitule == 'Demi-Gros':
            types_vente_dict['Demi_Gros'] = tv.prix
        elif tv.intitule == 'Détail':
            types_vente_dict['Détail'] = tv.prix
        else:
            types_vente_dict[tv.intitule] = tv.prix
    
    context = {
        'article': article,
        'agences': agences,
        'familles': familles,
        'types_vente': types_vente_dict
    }
    return render(request, 'supermarket/stock/modifier_article.html', context)
@login_required
def supprimer_article(request, article_id):
    """Vue pour supprimer un article"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_articles')
    
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
        article_name = article.designation
        article.delete()
        messages.success(request, f'Article "{article_name}" supprimé avec succès!')
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_articles')

@login_required
def detail_article(request, article_id):
    """Vue pour afficher les détails d'un article"""
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
        types_vente = TypeVente.objects.filter(article=article)
        # Le suivi de stock est toujours activé, donc on affiche tous les mouvements
        mouvements = MouvementStock.objects.filter(article=article).order_by('-date_mouvement')[:10]
        
        # Calculer les marges
        marge_unitaire = float(article.prix_vente) - float(article.prix_achat) if article.prix_achat > 0 else 0
        marge_pourcentage = (marge_unitaire / float(article.prix_achat) * 100) if article.prix_achat > 0 else 0
        valeur_stock = float(article.prix_achat) * float(article.stock_actuel)
        
        # Calculer les marges pour chaque type de vente
        types_vente_with_marges = []
        for tv in types_vente:
            marge_tv = float(tv.prix) - float(article.prix_achat) if article.prix_achat > 0 else 0
            marge_tv_pourcentage = (marge_tv / float(article.prix_achat) * 100) if article.prix_achat > 0 else 0
            types_vente_with_marges.append({
                'type_vente': tv,
                'marge': marge_tv,
                'marge_pourcentage': marge_tv_pourcentage
            })
        
        # Debug: Vérifier la famille de l'article
        print(f"[ALERTE] DEBUG Article {article.id}:")
        print(f"   - Désignation: {article.designation}")
        print(f"   - Catégorie: {article.categorie}")
        print(f"   - Intitulé famille: {article.categorie.intitule if article.categorie else 'None'}")
        
        context = {
            'article': article,
            'types_vente': types_vente,
            'types_vente_with_marges': types_vente_with_marges,
            'mouvements': mouvements,
            'marge_unitaire': marge_unitaire,
            'marge_pourcentage': marge_pourcentage,
            'valeur_stock': valeur_stock
        }
        return render(request, 'supermarket/stock/detail_article.html', context)
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
        return redirect('consulter_articles')

# ==================== FACTURES D'ACHAT ====================

@login_required
def consulter_factures_achat(request):
    """Vue pour consulter les factures d'achat"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    statut_filter = request.GET.get('statut_filter', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Construire la requête de base
    factures = FactureAchat.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        factures = factures.filter(
            Q(numero_fournisseur__icontains=search_query) |
            Q(reference_achat__icontains=search_query) |
            Q(commentaire__icontains=search_query)
        )
    
    if statut_filter:
        factures = factures.filter(statut=statut_filter)
    
    if date_debut:
        factures = factures.filter(date_achat__gte=date_debut)
    
    if date_fin:
        factures = factures.filter(date_achat__lte=date_fin)
    
    # Trier par date décroissante
    factures = factures.order_by('-date_achat', '-heure')
    
    # Calculer les statistiques
    total_factures = FactureAchat.objects.filter(agence=agence).count()
    factures_validees = FactureAchat.objects.filter(agence=agence, statut='validee').count()
    factures_payees = FactureAchat.objects.filter(agence=agence, statut='payee').count()
    montant_total = FactureAchat.objects.filter(agence=agence).aggregate(
        total=Sum('prix_total_global')
    )['total'] or 0
    
    # Récupérer les statuts pour le filtre
    statut_choices = FactureAchat.STATUT_CHOICES
    
    context = {
        'factures': factures,
        'search_query': search_query,
        'statut_filter': statut_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statut_choices': statut_choices,
        'total_factures': total_factures,
        'factures_validees': factures_validees,
        'factures_payees': factures_payees,
        'montant_total': montant_total,
    }
    return render(request, 'supermarket/stock/consulter_factures_achat.html', context)

@login_required
def creer_facture_achat(request):
    """Vue pour créer une nouvelle facture d'achat"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    if request.method == 'POST':
        try:
            print("[START] DÉBUT CRÉATION FACTURE D'ACHAT")
            print(f"[LIST] Données POST reçues: {dict(request.POST)}")
            
            # Récupérer les données du formulaire
            numero_fournisseur = request.POST.get('numero_fournisseur')
            date_achat = request.POST.get('date_achat')
            heure = request.POST.get('heure')
            reference_achat = request.POST.get('reference_achat')
            prix_total_global = request.POST.get('prix_total_global')
            statut = request.POST.get('statut')
            commentaire = request.POST.get('commentaire')
            
            print(f"[NOTE] Données extraites:")
            print(f"  - Numéro fournisseur: {numero_fournisseur}")
            print(f"  - Date achat: {date_achat}")
            print(f"  - Heure: {heure}")
            print(f"  - Référence: {reference_achat}")
            print(f"  - Prix total: {prix_total_global}")
            print(f"  - Statut: {statut}")
            print(f"  - Commentaire: {commentaire}")
            
            print("[SEARCH] AVANT VALIDATION")
            
            # Validation
            if not all([numero_fournisseur, date_achat, heure, reference_achat, prix_total_global]):
                print("[ERREUR] VALIDATION ÉCHOUÉE")
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_facture_achat')
            
            print("[OK] VALIDATION RÉUSSIE")
            
            print("[SEARCH] AVANT CRÉATION FOURNISSEUR")
            
            # Récupérer ou créer le fournisseur
            fournisseur, created = Fournisseur.objects.get_or_create(
                intitule=numero_fournisseur,
                defaults={'agence': agence}
            )
            
            print(f"[OK] FOURNISSEUR: {fournisseur.intitule}")
            print("[SEARCH] AVANT CRÉATION FACTURE")
            
            try:
                # Vérifier si la référence existe déjà
                if FactureAchat.objects.filter(reference_achat=reference_achat).exists():
                    # Générer une référence unique
                    import time
                    reference_achat = f"{reference_achat}_{int(time.time())}"
                    print(f"[REFRESH] Référence modifiée pour éviter le doublon: {reference_achat}")
                
                # Créer la facture d'achat avec statut temporaire (sera validée après création des lignes)
                facture = FactureAchat.objects.create(
                    numero_fournisseur=numero_fournisseur,
                    date_achat=date_achat,
                    heure=heure,
                    reference_achat=reference_achat,
                    prix_total_global=float(prix_total_global),
                    statut='brouillon',  # Statut temporaire
                    commentaire=commentaire,
                    fournisseur=fournisseur,
                    agence=agence
                )
                print(f"[OK] FACTURE CRÉÉE: {facture.id}")
            except Exception as e:
                print(f"[ERREUR] ERREUR CRÉATION FACTURE: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f'Erreur lors de la création de la facture: {str(e)}')
                return redirect('creer_facture_achat')
            
            print("[SEARCH] AVANT TRAITEMENT ARTICLES")
            
            # Traiter les articles sélectionnés
            articles_data = request.POST.get('articles_data', '')
            print(f"[SEARCH] Articles data reçus: {articles_data}")
            
            try:
                if articles_data:
                    import json
                    try:
                        articles = json.loads(articles_data)
                        print(f"[PACKAGE] Articles à traiter: {len(articles)}")
                        
                        for i, article_data in enumerate(articles):
                            print(f"[NOTE] Traitement article {i+1}: {article_data}")
                            
                            # Récupérer l'article
                            article = Article.objects.get(id=article_data['id'])
                            print(f"[OK] Article trouvé: {article.designation}")
                            
                            # Convertir les quantités et prix en Decimal pour gérer les décimales correctement
                            quantite_decimale = Decimal(str(article_data['quantite']))
                            prix_achat_decimal = Decimal(str(article_data['prix_achat']))
                            prix_total_decimale = quantite_decimale * prix_achat_decimal
                            
                            # Créer la ligne de facture
                            ligne = LigneFactureAchat.objects.create(
                                facture_achat=facture,
                                article=article,
                                reference_article=article.reference_article,
                                designation=article.designation,
                                prix_unitaire=prix_achat_decimal,
                                quantite=quantite_decimale,
                                prix_total_article=prix_total_decimale
                            )
                            print(f"[OK] Ligne créée: {ligne.id}")
                            
                            # Le stock sera mis à jour automatiquement par le modèle FactureAchat
                            # car la facture est créée avec statut='validee'
                            print(f"🛒 STOCK ACHAT - Article: {article.designation}")
                            print(f"🛒 STOCK ACHAT - Quantité: {article_data['quantite']}")
                            print(f"🛒 STOCK ACHAT - Stock sera mis à jour automatiquement par le modèle")
                            
                    except (json.JSONDecodeError, Article.DoesNotExist, KeyError) as e:
                        print(f"[ERREUR] Erreur lors du traitement des articles: {e}")
                        messages.error(request, f'Erreur lors du traitement des articles: {str(e)}')
                else:
                    print("[WARNING] Aucun article sélectionné")
            except Exception as e:
                print(f"[ERREUR] ERREUR GÉNÉRALE: {e}")
                import traceback
                traceback.print_exc()
            
            # Valider la facture pour déclencher la mise à jour automatique du stock
            print("[INFO] Validation de la facture pour déclencher la mise à jour du stock...")
            facture.statut = 'validee'
            facture.save()  # Cela déclenchera automatiquement la mise à jour du stock
            
            print("[INFO] Stock mis à jour automatiquement par le modèle FactureAchat")
            
            messages.success(request, f'Facture d\'achat "{reference_achat}" créée avec succès! Stock mis à jour automatiquement.')
            return redirect('creer_facture_achat')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de la facture: {str(e)}')
            return redirect('creer_facture_achat')
    
    # GET - Afficher le formulaire
    statut_choices = FactureAchat.STATUT_CHOICES
    
    context = {
        'statut_choices': statut_choices,
    }
    return render(request, 'supermarket/stock/creer_facture_achat.html', context)

@login_required
def detail_facture_achat(request, facture_id):
    """Vue pour afficher les détails d'une facture d'achat"""
    try:
        agence = get_user_agence(request)
        facture = FactureAchat.objects.get(id=facture_id, agence=agence)
        lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
        
        print(f"[SEARCH] DÉTAIL FACTURE D'ACHAT")
        print(f"[LIST] Facture ID: {facture_id}")
        print(f"[LIST] Facture: {facture.reference_achat}")
        print(f"[LIST] Agence: {agence.nom_agence}")
        print(f"[PACKAGE] Nombre de lignes trouvées: {lignes.count()}")
        
        # Debug: vérifier toutes les lignes de facture d'achat
        toutes_lignes = LigneFactureAchat.objects.all()
        print(f"[CHART] TOTAL LIGNES DANS LA BASE: {toutes_lignes.count()}")
        
        for ligne in toutes_lignes:
            print(f"  [NOTE] Ligne globale: {ligne.designation} - Facture: {ligne.facture_achat.reference_achat}")
        
        for i, ligne in enumerate(lignes):
            print(f"  [NOTE] Ligne {i+1}: {ligne.designation} - Qty: {ligne.quantite} - Prix: {ligne.prix_unitaire}")
        
        context = {
            'facture': facture,
            'lignes': lignes,
        }
        return render(request, 'supermarket/stock/detail_facture_achat.html', context)
    except FactureAchat.DoesNotExist:
        print(f"[ERREUR] Facture d'achat {facture_id} non trouvée")
        messages.error(request, 'Facture d\'achat non trouvée.')
        return redirect('consulter_factures_achat')

@login_required
def modifier_facture_achat(request, facture_id):
    """Vue pour modifier une facture d'achat existante"""
    try:
        agence = get_user_agence(request)
        facture = FactureAchat.objects.get(id=facture_id, agence=agence)
    except FactureAchat.DoesNotExist:
        messages.error(request, 'Facture d\'achat non trouvée.')
        return redirect('consulter_factures_achat')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_fournisseur = request.POST.get('numero_fournisseur')
            date_achat = request.POST.get('date_achat')
            heure = request.POST.get('heure')
            reference_achat = request.POST.get('reference_achat')
            prix_total_global = request.POST.get('prix_total_global')
            statut = request.POST.get('statut')
            commentaire = request.POST.get('commentaire')
            articles_data = request.POST.get('articles_data')  # JSON des nouveaux articles ajoutés
            ligne_supprimee_id = request.POST.get('ligne_supprimee')  # id d'une ligne à supprimer
            
            # Validation
            if not all([numero_fournisseur, date_achat, heure, reference_achat, prix_total_global]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_facture_achat', facture_id=facture_id)
            
            # Mettre à jour la facture d'achat
            # IMPORTANT : Utiliser update() pour éviter que la méthode save() du modèle
            # n'appelle mettre_a_jour_stock() automatiquement, car on gère manuellement
            # la logique de modification (annulation + nouvelle quantité avec différence)
            # 
            # ATTENTION : Ne pas changer le statut si la facture est déjà validée,
            # car cela pourrait déclencher mettre_a_jour_stock() qui réappliquerait la quantité totale
            statut_avant = facture.statut
            if statut != statut_avant and statut_avant == 'validee' and statut == 'validee':
                # Si le statut reste 'validee', on ne change pas le statut pour éviter de déclencher mettre_a_jour_stock()
                print(f"[DEBUG] Statut inchangé (validee), on ne modifie pas le statut pour éviter mettre_a_jour_stock()")
                statut_a_sauvegarder = statut_avant
            else:
                statut_a_sauvegarder = statut
            
            FactureAchat.objects.filter(id=facture.id).update(
                numero_fournisseur=numero_fournisseur,
                date_achat=date_achat,
                heure=heure,
                reference_achat=reference_achat,
                prix_total_global=float(prix_total_global),
                statut=statut_a_sauvegarder,
                commentaire=commentaire
            )
            # Recharger la facture pour avoir les valeurs à jour
            facture.refresh_from_db()

            # Nouvelle logique : Annuler d'abord l'effet de l'ancienne facture, puis appliquer la nouvelle
            # 1. Soustraire les anciennes quantités du stock
            # 2. Ajouter les nouvelles quantités au stock
            # 3. Créer un seul mouvement de stock qui reflète le changement net
            from datetime import timedelta
            with transaction.atomic():
                # 1) Récupérer les anciennes lignes AVANT de les supprimer
                lignes_existantes = LigneFactureAchat.objects.select_related('article').filter(facture_achat=facture)
                
                # Créer un dictionnaire des anciennes quantités par article
                anciennes_quantites = {}
                
                print(f"[DEBUG] Récupération des anciennes lignes: {lignes_existantes.count()} lignes trouvées")
                for ligne in lignes_existantes:
                    if ligne.article:
                        article_id = ligne.article.id
                        anciennes_quantites[article_id] = Decimal(str(ligne.quantite))
                        print(f"  [DEBUG] Ancienne ligne - Article ID: {article_id}, Quantité: {ligne.quantite}")
                
                # 2) Supprimer toutes les anciennes lignes
                LigneFactureAchat.objects.filter(facture_achat=facture).delete()
                print(f"[DEBUG] Anciennes lignes supprimées")
                
                # 3) Traiter les nouvelles lignes : annuler l'ancienne puis appliquer la nouvelle
                if articles_data:
                    import json
                    try:
                        articles = json.loads(articles_data)
                        print(f"[DEBUG] Articles reçus du formulaire: {articles}")
                        for a in articles:
                            # Convertir l'ID en entier (peut être string ou int selon le JSON)
                            article_id = int(a.get('id', 0)) if a.get('id') else None
                            # Normaliser les valeurs numériques (gérer les virgules, etc.)
                            quantite_str = str(a.get('quantite', 0)).replace(',', '.')
                            prix_str = str(a.get('prix_achat', 0)).replace(',', '.')
                            quantite_nouvelle = Decimal(quantite_str)
                            prix_achat_nouveau = Decimal(prix_str)
                            
                            print(f"[DEBUG] Article ID: {article_id}, Quantité: {quantite_nouvelle}, Prix: {prix_achat_nouveau}")
                            
                            if not article_id or quantite_nouvelle <= 0:
                                print(f"[DEBUG] Article ignoré: ID={article_id}, Quantité={quantite_nouvelle}")
                                continue
                            
                            try:
                                article = Article.objects.get(id=article_id)
                                article.refresh_from_db()
                            except Article.DoesNotExist:
                                messages.error(request, f"Article introuvable (ID: {article_id}).")
                                continue
                            
                            # Stock AVANT modification
                            # IMPORTANT : Le stock actuel contient déjà l'ancienne quantité de la facture
                            stock_avant_modification = article.stock_actuel
                            
                            # Quantité de l'ancienne facture
                            quantite_ancienne = anciennes_quantites.get(article_id, Decimal('0'))
                            
                            # Calcul de la différence entre nouvelle et ancienne quantité
                            difference = quantite_nouvelle - quantite_ancienne
                            
                            # DEBUG : Afficher les valeurs pour vérification
                            print(f"[DEBUG MODIFICATION FACTURE ACHAT]")
                            print(f"  Article: {article.designation} (ID: {article_id})")
                            print(f"  Stock avant modification: {stock_avant_modification}")
                            print(f"  Quantité ancienne (dans facture): {quantite_ancienne}")
                            print(f"  Quantité nouvelle (dans facture): {quantite_nouvelle}")
                            print(f"  Différence calculée: {difference} (nouvelle {quantite_nouvelle} - ancienne {quantite_ancienne})")
                            
                            # LOGIQUE CORRIGÉE : Calculer le stock final avec la différence
                            # Le stock actuel contient déjà l'effet de l'ancienne facture
                            # Pour modifier : stock_final = stock_actuel - ancienne_quantité + nouvelle_quantité
                            # Ce qui équivaut à : stock_final = stock_actuel + différence
                            # 
                            # Exemples :
                            # - Stock = 30, ancienne = 10, nouvelle = 20 → différence = +10 → stock final = 30 + 10 = 40 ✓
                            # - Stock = 30, ancienne = 20, nouvelle = 10 → différence = -10 → stock final = 30 - 10 = 20 ✓
                            # - Stock = 30, ancienne = 10, nouvelle = 10 → différence = 0 → stock final = 30 + 0 = 30 ✓
                            
                            # Calculer le stock final
                            stock_final = stock_avant_modification + difference
                            
                            # S'assurer que le stock ne devient pas négatif
                            if stock_final < 0:
                                messages.error(request, f'Stock insuffisant pour l\'article {article.designation}. Stock actuel: {stock_avant_modification}, différence: {difference}')
                                raise ValueError(f"Stock insuffisant: {stock_avant_modification} + {difference} = {stock_final} < 0")
                            
                            # Stock après annulation (pour information, mais on utilise directement la différence)
                            stock_apres_annulation = stock_avant_modification - quantite_ancienne
                            
                            print(f"  Stock avant modification: {stock_avant_modification}")
                            print(f"  Différence: {difference} (nouvelle {quantite_nouvelle} - ancienne {quantite_ancienne})")
                            print(f"  Stock final: {stock_final} = {stock_avant_modification} + {difference}")
                            
                            # Créer la nouvelle ligne
                            # IMPORTANT : La création de la ligne ne doit PAS déclencher de mise à jour du stock
                            # car on gère manuellement la logique avec la différence
                            ligne = LigneFactureAchat.objects.create(
                                facture_achat=facture,
                                article=article,
                                reference_article=article.reference_article,
                                designation=article.designation,
                                prix_unitaire=float(prix_achat_nouveau),
                                quantite=int(quantite_nouvelle),
                                prix_total_article=float(prix_achat_nouveau * quantite_nouvelle)
                            )
                            print(f"  ✅ Ligne créée avec quantité: {quantite_nouvelle} (mais le stock sera mis à jour avec différence: {difference})")
                            
                            # Mettre à jour le stock avec le nouveau stock calculé
                            # Le stock final = stock_initial + différence (augmentation ou diminution selon le signe)
                            # IMPORTANT : On utilise stock_final qui est calculé avec la différence, JAMAIS quantite_nouvelle
                            ancien_stock_pour_verif = article.stock_actuel
                            
                            # VÉRIFICATION CRITIQUE : S'assurer qu'on utilise bien la différence et non quantite_nouvelle
                            if stock_final != stock_avant_modification + difference:
                                print(f"[ERREUR] Calcul incorrect du stock final!")
                                raise ValueError(f"Stock final incorrect: {stock_final} != {stock_avant_modification} + {difference}")
                            
                            # Mettre à jour le stock avec le stock final calculé
                            article.stock_actuel = stock_final
                            
                            # Vérification : le stock doit être égal à stock_avant_modification + difference
                            if abs(article.stock_actuel - (stock_avant_modification + difference)) > Decimal('0.01'):
                                raise ValueError(f"Stock incorrect: {article.stock_actuel} != {stock_avant_modification} + {difference}")
                            
                            print(f"  ✅ Stock mis à jour: {ancien_stock_pour_verif} → {article.stock_actuel}")
                            print(f"  ✅ Différence appliquée au stock: {article.stock_actuel - ancien_stock_pour_verif} (doit être {difference}, PAS {quantite_nouvelle})")
                            print(f"  ✅ Vérification: stock_initial ({stock_avant_modification}) + différence ({difference}) = solde ({stock_final})")
                            
                            # Vérification finale
                            if abs((article.stock_actuel - ancien_stock_pour_verif) - difference) > Decimal('0.01'):
                                print(f"[ERREUR CRITIQUE] La différence appliquée ({article.stock_actuel - ancien_stock_pour_verif}) ne correspond pas à la différence calculée ({difference})!")
                                raise ValueError(f"Différence incorrecte appliquée au stock!")
                            
                            # La différence a déjà été calculée plus haut pour clarifier la logique
                            
                            # Déterminer le type de mouvement et le commentaire selon la différence
                            if difference > 0:
                                # Augmentation : nouvelle quantité > ancienne quantité
                                # Exemple : ancienne = 10, nouvelle = 20, différence = +10, stock augmente de 10
                                type_mouvement = 'entree'
                                quantite_mouvement = difference  # Utiliser la différence réelle (positive)
                                commentaire = f"Achat - Facture {facture.reference_achat} (modification - augmentation: {quantite_ancienne} → {quantite_nouvelle}, +{difference})"
                            elif difference < 0:
                                # Diminution : nouvelle quantité < ancienne quantité
                                # Exemple : ancienne = 20, nouvelle = 10, différence = -10, stock diminue de 10
                                type_mouvement = 'sortie'  # Changé de 'ajustement' à 'sortie' pour refléter la diminution
                                quantite_mouvement = abs(difference)  # Utiliser la valeur absolue de la différence
                                commentaire = f"Achat - Facture {facture.reference_achat} (modification - diminution: {quantite_ancienne} → {quantite_nouvelle}, -{abs(difference)})"
                            else:
                                # Quantité inchangée : nouvelle quantité = ancienne quantité, différence = 0
                                # Pas de mouvement de stock nécessaire car le stock ne change pas
                                type_mouvement = None
                                quantite_mouvement = Decimal('0')
                                commentaire = f"Achat - Facture {facture.reference_achat} (modification - quantité inchangée: {quantite_nouvelle})"
                            
                            article.dernier_prix_achat = float(prix_achat_nouveau)
                            # Le suivi de stock est toujours activé automatiquement lors de toute modification du stock
                            article.suivi_stock = True
                            article.save()
                            
                            # IMPORTANT : Ne pas supprimer les anciens mouvements pour garder la traçabilité complète
                            # Tous les mouvements effectués doivent être conservés pour l'audit
                            
                            # Créer un mouvement de stock pour refléter le changement net
                            # IMPORTANT : quantite_mouvement = différence (ex: 50 → 70, différence = 20, mouvement = +20)
                            # Le stock_initial est le stock AVANT la modification
                            # Le solde est le stock APRÈS la modification = stock_initial + différence
                            # La quantité du mouvement = différence (pas la quantité totale)
                            # Exemple : stock_initial = 100, différence = +20, solde = 120 (pas 170)
                            # On crée un mouvement seulement si il y a un changement (quantite_mouvement > 0)
                            if quantite_mouvement > 0 and type_mouvement:
                                try:
                                    # IMPORTANT : Le mouvement doit refléter la différence, pas la quantité totale
                                    # Exemple : stock avant modification = 21, ancienne quantité = 20, nouvelle quantité = 30
                                    # différence = 30 - 20 = 10
                                    # stock final = 21 + 10 = 31
                                    # Pour le mouvement : stock_initial = 21, quantite = 10 (différence), solde = 31
                                    # Vérification : 21 + 10 = 31 ✓
                                    # VÉRIFICATION CRITIQUE : quantite_mouvement doit être égal à difference (ou abs(difference))
                                    if difference > 0 and quantite_mouvement != difference:
                                        raise ValueError(f"ERREUR: quantite_mouvement ({quantite_mouvement}) != difference ({difference})")
                                    elif difference < 0 and quantite_mouvement != abs(difference):
                                        raise ValueError(f"ERREUR: quantite_mouvement ({quantite_mouvement}) != abs(difference) ({abs(difference)})")
                                    
                                    print(f"  [DEBUG] Création mouvement - quantite_mouvement: {quantite_mouvement}, difference: {difference}, quantite_nouvelle: {quantite_nouvelle}")
                                    
                                    MouvementStock.objects.create(
                                        article=article,
                                        agence=agence,
                                        type_mouvement=type_mouvement,
                                        date_mouvement=timezone.now(),
                                        numero_piece=facture.reference_achat,
                                        quantite_stock=article.stock_actuel,
                                        stock_initial=stock_avant_modification,  # Stock AVANT modification (effet net)
                                        solde=stock_final,  # Stock APRÈS modification = stock_initial + différence
                                        quantite=quantite_mouvement,  # La DIFFÉRENCE (ex: 10 si 20→30, pas 30)
                                        cout_moyen_pondere=float(article.prix_achat),
                                        stock_permanent=float(article.stock_actuel * article.prix_achat),
                                        facture_achat=facture,
                                        fournisseur=facture.fournisseur,
                                        commentaire=commentaire
                                    )
                                    print(f"  ✅ Mouvement de stock créé - stock_initial: {stock_avant_modification}, quantite: {quantite_mouvement} (DIFFÉRENCE), solde: {stock_final}")
                                    print(f"  ✅ Vérification: {stock_avant_modification} + {quantite_mouvement} = {stock_avant_modification + quantite_mouvement} (doit être {stock_final})")
                                except Exception as e:
                                    print(f"[WARNING] Erreur création mouvement stock: {e}")
                                    import traceback
                                    traceback.print_exc()
                    except json.JSONDecodeError:
                        messages.error(request, 'Format des articles invalide.')
                    except Exception as e:
                        messages.error(request, f'Erreur lors de la modification des lignes: {str(e)}')
                else:
                    # Si aucun article n'est fourni, la facture est vide (toutes les lignes ont été supprimées)
                    messages.info(request, "Toutes les lignes ont été supprimées. La facture est maintenant vide.")
            
            # IMPORTANT: Ne JAMAIS appeler facture.save() après une modification
            # car cela déclencherait mettre_a_jour_stock() qui réappliquerait la quantité totale
            # au lieu de respecter la différence que nous avons calculée manuellement
            # La facture a déjà été mise à jour avec update() plus haut, donc pas besoin de save()
            
            print(f"[DEBUG] Modification terminée - Stock mis à jour avec la DIFFÉRENCE, pas la quantité totale")
            print(f"[DEBUG] Ne pas appeler facture.save() pour éviter mettre_a_jour_stock()")
            
            messages.success(request, f'Facture d\'achat "{reference_achat}" modifiée avec succès!')
            # Rediriger vers le détail de la facture pour voir les modifications
            return redirect('detail_facture_achat', facture_id=facture.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de la facture: {str(e)}')
            return redirect('modifier_facture_achat', facture_id=facture_id)
    
    # GET - Afficher le formulaire pré-rempli
    import json
    statut_choices = FactureAchat.STATUT_CHOICES
    
    # Récupérer les lignes d'articles de la facture
    lignes = LigneFactureAchat.objects.select_related('article').filter(facture_achat=facture)
    
    # Préparer les données initiales pour le template (comme dans modifier_facture_transfert)
    initial_articles = []
    for ligne in lignes:
        if ligne.article:
            initial_articles.append({
                'id': ligne.article.id,
                'designation': ligne.article.designation,
                'reference_article': ligne.article.reference_article,
                'stock_actuel': float(ligne.article.stock_actuel),
                'prix_achat': float(ligne.prix_unitaire),
                'quantite': float(ligne.quantite)  # Convertir Decimal en float pour JSON
            })
    
    context = {
        'facture': facture,
        'statut_choices': statut_choices,
        'lignes': lignes,
        'initial_articles_json': json.dumps(initial_articles),
    }
    return render(request, 'supermarket/stock/modifier_facture_achat.html', context)

@login_required
def supprimer_facture_achat(request, facture_id):
    """Vue pour supprimer une facture d'achat"""
    print(f"🗑️ SUPPRESSION FACTURE D'ACHAT")
    print(f"[LIST] Méthode: {request.method}")
    print(f"[LIST] Facture ID: {facture_id}")
    
    if request.method != 'POST':
        print("[ERREUR] Méthode non autorisée - redirection")
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_factures_achat')
    
    try:
        agence = get_user_agence(request)
        print(f"[BUILDING] Agence: {agence}")
        
        facture = FactureAchat.objects.get(id=facture_id, agence=agence)
        print(f"[LIST] Facture trouvée: {facture.reference_achat}")
        
        # Restocker: retirer les quantités ajoutées par cette facture et créer des mouvements de correction
        try:
            lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
            for ligne in lignes:
                article = ligne.article
                if not article:
                    continue
                ancien_stock = getattr(article, 'stock_actuel', 0)
                # Empêcher les stocks négatifs
                article.stock_actuel = max(0, ancien_stock - int(ligne.quantite))
                # Le suivi de stock est toujours activé automatiquement lors de toute modification du stock
                article.suivi_stock = True
                article.save()
                print(f"[STOCK] Reversion achat: {article.designation} {ancien_stock} -> {article.stock_actuel} (-{ligne.quantite})")
                
                # Créer un mouvement de correction au lieu de supprimer les mouvements existants
                try:
                    MouvementStock.objects.create(
                        article=article,
                        agence=agence,
                        type_mouvement='ajustement',
                        date_mouvement=timezone.now(),
                        numero_piece=f"SUPP-{facture.reference_achat}",
                        quantite_stock=article.stock_actuel,
                        stock_initial=ancien_stock,
                        solde=article.stock_actuel,
                        quantite=int(ligne.quantite),
                        cout_moyen_pondere=float(article.prix_achat),
                        stock_permanent=float(article.stock_actuel * article.prix_achat),
                        facture_achat=facture,
                        commentaire=f"Correction - Suppression facture achat {facture.reference_achat}"
                    )
                except Exception as e:
                    print(f"[WARNING] Erreur création mouvement correction suppression: {e}")
        except Exception as e:
            print(f"[WARNING] Echec reversion stock facture achat: {e}")

        facture_name = facture.reference_achat
        facture.delete()
        print(f"[OK] Facture supprimée: {facture_name}")
        
        messages.success(request, f'Facture d\'achat "{facture_name}" supprimée avec succès!')
    except FactureAchat.DoesNotExist:
        print(f"[ERREUR] Facture d'achat {facture_id} non trouvée")
        messages.error(request, 'Facture d\'achat non trouvée.')
    except Exception as e:
        print(f"[ERREUR] Erreur lors de la suppression: {e}")
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_factures_achat')

# ==================== FACTURES DE TRANSFERT ====================

@login_required
def consulter_factures_transfert(request):
    """Vue pour consulter les factures de transfert"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    statut_filter = request.GET.get('statut_filter', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Construire la requête de base (factures où l'agence est source ou destination)
    factures = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    )
    
    # Appliquer les filtres
    if search_query:
        factures = factures.filter(
            Q(numero_compte__icontains=search_query) |
            Q(reference_transfert__icontains=search_query) |
            Q(lieu_depart__icontains=search_query) |
            Q(lieu_arrivee__icontains=search_query)
        )
    
    if statut_filter:
        factures = factures.filter(statut=statut_filter)
    
    if date_debut:
        factures = factures.filter(date_transfert__gte=date_debut)
    
    if date_fin:
        factures = factures.filter(date_transfert__lte=date_fin)
    
    # Trier par date décroissante
    factures = factures.order_by('-date_transfert')
    
    # Calculer les statistiques
    total_factures = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    ).count()
    factures_en_cours = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence),
        statut='en_cours'
    ).count()
    factures_terminees = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence),
        statut='termine'
    ).count()
    quantite_totale = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    ).aggregate(
        total=Sum('quantite')
    )['total'] or 0
    
    # Récupérer les statuts pour le filtre
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'factures': factures,
        'search_query': search_query,
        'statut_filter': statut_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statut_choices': statut_choices,
        'total_factures': total_factures,
        'factures_en_cours': factures_en_cours,
        'factures_terminees': factures_terminees,
        'quantite_totale': quantite_totale,
    }
    return render(request, 'supermarket/stock/consulter_factures_transfert.html', context)

@login_required
def creer_facture_transfert(request):
    """Vue pour créer une nouvelle facture de transfert"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_compte = request.POST.get('numero_compte')
            date_transfert = request.POST.get('date_transfert')
            reference_transfert = request.POST.get('reference_transfert')
            lieu_depart = request.POST.get('lieu_depart')
            lieu_arrivee = request.POST.get('lieu_arrivee')
            quantite = request.POST.get('quantite')
            statut = request.POST.get('statut')
            employe_expediteur = request.POST.get('employe_expediteur', '').strip()
            employe_destinataire = request.POST.get('employe_destinataire', '').strip()
            etat = request.POST.get('etat', 'sortir')
            
            # Validation
            if not all([numero_compte, date_transfert, reference_transfert, lieu_depart, lieu_arrivee, quantite, employe_expediteur]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_facture_transfert')
            
            # Rechercher les employés existants par nom
            expediteur_employe = None
            destinataire_employe = None
            
            # Rechercher l'employé expéditeur
            try:
                # Essayer de trouver par nom complet dans le compte
                expediteur_employe = Employe.objects.filter(
                    compte__agence=agence,
                    compte__user__first_name__icontains=employe_expediteur.split()[0] if employe_expediteur.split() else employe_expediteur
                ).first()
                
                # Si pas trouvé, essayer par nom de famille
                if not expediteur_employe and len(employe_expediteur.split()) > 1:
                    expediteur_employe = Employe.objects.filter(
                        compte__agence=agence,
                        compte__user__last_name__icontains=employe_expediteur.split()[-1]
                    ).first()
                
                # Si toujours pas trouvé, prendre le premier employé de l'agence
                if not expediteur_employe:
                    expediteur_employe = Employe.objects.filter(compte__agence=agence).first()
                    
            except Exception as e:
                print(f"Erreur lors de la recherche de l'employé expéditeur: {e}")
                expediteur_employe = Employe.objects.filter(compte__agence=agence).first()
            
            if not expediteur_employe:
                messages.error(request, 'Aucun employé trouvé dans cette agence.')
                return redirect('creer_facture_transfert')
            
            # Afficher un message informatif sur l'employé trouvé
            expediteur_nom = f"{expediteur_employe.compte.user.first_name} {expediteur_employe.compte.user.last_name}".strip()
            messages.info(request, f'Employé expéditeur trouvé: {expediteur_nom}')
            
            if destinataire_employe:
                destinataire_nom = f"{destinataire_employe.compte.user.first_name} {destinataire_employe.compte.user.last_name}".strip()
                messages.info(request, f'Employé destinataire trouvé: {destinataire_nom}')
            
            # Rechercher l'employé destinataire (si fourni)
            if employe_destinataire:
                try:
                    # Essayer de trouver par nom complet dans le compte
                    destinataire_employe = Employe.objects.filter(
                        compte__agence=agence,
                        compte__user__first_name__icontains=employe_destinataire.split()[0] if employe_destinataire.split() else employe_destinataire
                    ).first()
                    
                    # Si pas trouvé, essayer par nom de famille
                    if not destinataire_employe and len(employe_destinataire.split()) > 1:
                        destinataire_employe = Employe.objects.filter(
                            compte__agence=agence,
                            compte__user__last_name__icontains=employe_destinataire.split()[-1]
                        ).first()
                        
                except Exception as e:
                    print(f"Erreur lors de la recherche de l'employé destinataire: {e}")
                    destinataire_employe = None
            
            # Vérifier l'unicité de la référence et générer une référence unique si nécessaire
            reference_finale = reference_transfert
            if FactureTransfert.objects.filter(reference_transfert=reference_transfert).exists():
                # Générer une référence unique avec timestamp
                import time
                reference_finale = f"{reference_transfert}_{int(time.time())}"
                print(f"[INFO] Référence modifiée pour éviter le doublon: {reference_finale}")
            
            # Créer la facture de transfert
            facture = FactureTransfert.objects.create(
                numero_compte=numero_compte,
                date_transfert=date_transfert,
                reference_transfert=reference_finale,
                lieu_depart=lieu_depart,
                lieu_arrivee=lieu_arrivee,
                quantite=int(quantite),
                statut=statut,
                agence_source=agence,
                agence_destination=agence,  # Pour l'instant, même agence (à modifier selon les besoins)
                employe_expediteur=expediteur_employe,
                employe_destinataire=destinataire_employe,
                etat=etat
            )
            
            # Traiter les articles sélectionnés
            articles_data = request.POST.get('articles_data', '')
            if articles_data:
                import json
                try:
                    articles = json.loads(articles_data)
                    for article_data in articles:
                        # Récupérer l'article
                        article = Article.objects.get(id=article_data['id'])
                        
                        # Créer la ligne de facture de transfert
                        LigneFactureTransfert.objects.create(
                            facture_transfert=facture,
                            article=article,
                            quantite=int(article_data['quantite']),
                            prix_unitaire=float(article_data['prix_achat']),
                            valeur_totale=float(article_data['prix_achat']) * int(article_data['quantite'])
                        )
                        
                        # Mettre à jour le stock de l'article selon l'état du transfert
                        ancien_stock = article.stock_actuel
                        quantite_transfert = Decimal(str(article_data['quantite']))
                        
                        # Normaliser l'état pour éviter les problèmes de casse ou d'espaces
                        etat_normalise = str(etat).strip().lower() if etat else 'sortir'
                        
                        # Si l'état est "entrer", augmenter le stock; si "sortir", diminuer le stock
                        # IMPORTANT: "entrer" doit TOUJOURS augmenter le stock
                        if etat_normalise == 'entrer':
                            article.stock_actuel += quantite_transfert
                            type_mouvement = 'entree'
                            action_stock = "augmentation"
                        else:  # sortir
                            article.stock_actuel -= quantite_transfert
                            if article.stock_actuel < 0:
                                article.stock_actuel = 0
                            type_mouvement = 'sortie'
                            action_stock = "diminution"
                        
                        # Mettre à jour le dernier prix d'achat avec le prix du transfert
                        ancien_dernier_prix = article.dernier_prix_achat
                        nouveau_prix_achat = float(article_data['prix_achat'])
                        article.dernier_prix_achat = nouveau_prix_achat
                        # Le suivi de stock est toujours activé automatiquement lors de toute modification du stock
                        article.suivi_stock = True
                        article.save()
                        print(f"[PACKAGE] STOCK TRANSFERT - Article: {article.designation}")
                        print(f"[PACKAGE] STOCK TRANSFERT - État: {etat} (normalisé: {etat_normalise}), {action_stock} du stock")
                        print(f"[PACKAGE] STOCK TRANSFERT - Stock mis à jour: {ancien_stock} → {article.stock_actuel}")
                        print(f"[MONEY] Transfert - Dernier prix d'achat mis à jour: {ancien_dernier_prix} → {nouveau_prix_achat}")
                        
                        # [HOT] CRÉER UN MOUVEMENT DE STOCK POUR TRAÇABILITÉ
                        try:
                            MouvementStock.objects.create(
                                article=article,
                                agence=agence,
                                type_mouvement=type_mouvement,
                                date_mouvement=timezone.now(),
                                numero_piece=facture.reference_transfert,
                                quantite_stock=article.stock_actuel,
                                stock_initial=ancien_stock,
                                solde=article.stock_actuel,
                                quantite=int(article_data['quantite']),
                                cout_moyen_pondere=float(article.prix_achat),
                                stock_permanent=float(article.stock_actuel * article.prix_achat),
                                commentaire=f"Transfert ({etat}) - Facture {facture.reference_transfert}"
                            )
                            print(f"[NOTE] MOUVEMENT STOCK - Transfert ({etat}) enregistré pour {article.designation}")
                        except Exception as e:
                            print(f"[WARNING] ERREUR MOUVEMENT STOCK TRANSFERT: {e}")
                        
                except (json.JSONDecodeError, Article.DoesNotExist, KeyError) as e:
                    print(f"Erreur lors du traitement des articles: {e}")
            
            messages.success(request, f'Facture de transfert "{reference_transfert}" créée avec succès!')
            return redirect('creer_facture_transfert')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de la facture: {str(e)}')
            return redirect('creer_facture_transfert')
    
    # GET - Afficher le formulaire
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'statut_choices': statut_choices,
    }
    return render(request, 'supermarket/stock/creer_facture_transfert_nouveau.html', context)

@login_required
def detail_facture_transfert(request, facture_id):
    """Vue pour afficher les détails d'une facture de transfert"""
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
        lignes = LigneFactureTransfert.objects.filter(facture_transfert=facture)
        
        context = {
            'facture': facture,
            'lignes': lignes,
        }
        return render(request, 'supermarket/stock/detail_facture_transfert.html', context)
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
        return redirect('consulter_factures_transfert')

@login_required
def modifier_facture_transfert(request, facture_id):
    """Vue pour modifier une facture de transfert existante"""
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
        return redirect('consulter_factures_transfert')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_compte = request.POST.get('numero_compte')
            date_transfert = request.POST.get('date_transfert')
            reference_transfert = request.POST.get('reference_transfert')
            lieu_depart = request.POST.get('lieu_depart')
            lieu_arrivee = request.POST.get('lieu_arrivee')
            quantite = request.POST.get('quantite')
            statut = request.POST.get('statut')
            employe_expediteur = request.POST.get('employe_expediteur', '').strip()
            employe_destinataire = request.POST.get('employe_destinataire', '').strip()
            etat = request.POST.get('etat', 'sortir')
            
            articles_data_raw = request.POST.get('articles_data')
            
            # Validation
            if not all([numero_compte, date_transfert, reference_transfert, lieu_depart, lieu_arrivee, quantite]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            if not articles_data_raw:
                messages.error(request, 'Veuillez sélectionner au moins un article.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            try:
                articles_payload = json.loads(articles_data_raw)
            except json.JSONDecodeError:
                messages.error(request, 'Format des articles invalide.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            if not articles_payload:
                messages.error(request, 'Veuillez sélectionner au moins un article.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            # Récupérer l'ancien état pour gérer le stock
            ancien_etat = facture.etat
            ancien_etat_normalise = str(ancien_etat).strip().lower() if ancien_etat else 'sortir'
            etat_normalise = str(etat).strip().lower() if etat else 'sortir'
            if ancien_etat_normalise == 'sortie':
                ancien_etat_normalise = 'sortir'
            if etat_normalise == 'sortie':
                etat_normalise = 'sortir'
            if ancien_etat_normalise == 'sortie':
                ancien_etat_normalise = 'sortir'
            if etat_normalise == 'sortie':
                etat_normalise = 'sortir'
            
            # Nouvelle logique : Calculer la différence et l'appliquer directement
            # Prendre en compte l'état (entrer/sortir) et la différence de quantité
            # Ajouter un flag pour empêcher la mise à jour automatique du stock
            request.session['modification_en_cours'] = True
            with transaction.atomic():
                # 1) Récupérer les anciennes lignes et créer un dictionnaire
                lignes_existantes = LigneFactureTransfert.objects.select_related('article').filter(facture_transfert=facture)
                anciennes_quantites = {}
                for ligne in lignes_existantes:
                    if ligne.article:
                        anciennes_quantites[ligne.article.id] = Decimal(str(ligne.quantite))
                
                # 2) Supprimer toutes les anciennes lignes
                LigneFactureTransfert.objects.filter(facture_transfert=facture).delete()
                
                # 3) Mettre à jour la facture de transfert
                facture.numero_compte = numero_compte
                facture.date_transfert = date_transfert
                facture.reference_transfert = reference_transfert
                facture.lieu_depart = lieu_depart
                facture.lieu_arrivee = lieu_arrivee
                facture.statut = statut
                facture.etat = etat
                
                total_quantite = Decimal('0')
                
                # 4) Traiter les nouvelles lignes avec la logique de différence
                for article_data in articles_payload:
                    article_id = article_data.get('id')
                    quantite_nouvelle = Decimal(str(article_data.get('quantite', 0)))
                    prix_achat_ligne = Decimal(str(article_data.get('prix_achat', 0)))
                    
                    if not article_id or quantite_nouvelle <= 0:
                        continue
                    
                    try:
                        article = Article.objects.get(id=article_id)
                        article.refresh_from_db()
                    except Article.DoesNotExist:
                        messages.error(request, f"Article introuvable (ID: {article_id}).")
                        raise
                    
                    # Vérifier le stock si c'est une sortie
                    if etat_normalise == 'sortir' and article.stock_actuel < quantite_nouvelle:
                        messages.error(request, f"Stock insuffisant pour l'article {article.designation}.")
                        raise ValueError("Stock insuffisant")
                    
                    # Créer la nouvelle ligne
                    LigneFactureTransfert.objects.create(
                        facture_transfert=facture,
                        article=article,
                        quantite=int(quantite_nouvelle),
                        prix_unitaire=prix_achat_ligne,
                        valeur_totale=quantite_nouvelle * prix_achat_ligne
                    )
                    
                    # Stock AVANT modification
                    # IMPORTANT : Le stock actuel contient déjà l'ancienne quantité de la facture
                    stock_avant_modification = article.stock_actuel
                    
                    # Quantité de l'ancienne facture
                    quantite_ancienne = anciennes_quantites.get(article_id, Decimal('0'))
                    
                    # Calcul de la différence entre nouvelle et ancienne quantité
                    difference = quantite_nouvelle - quantite_ancienne
                    
                    # LOGIQUE SIMPLIFIÉE (comme facture achat) : Calculer le stock final avec la différence
                    # Le stock actuel contient déjà l'effet de l'ancienne facture
                    # Pour modifier selon l'état :
                    # - Si état = "entrer" : stock_final = stock_actuel + différence (comme facture achat)
                    # - Si état = "sortir" : stock_final = stock_actuel - différence
                    # 
                    # Si l'état a changé, il faut d'abord annuler l'ancien effet puis appliquer le nouveau
                    etat_a_change = (ancien_etat_normalise != etat_normalise)
                    
                    if etat_a_change:
                        # L'état a changé : annuler l'ancien effet puis appliquer le nouveau
                        if ancien_etat_normalise == 'entrer':
                            # Annuler l'entrée : soustraire l'ancienne quantité
                            stock_apres_annulation = stock_avant_modification - quantite_ancienne
                        else:  # ancien = 'sortir'
                            # Annuler la sortie : ajouter l'ancienne quantité
                            stock_apres_annulation = stock_avant_modification + quantite_ancienne
                        
                        # Appliquer le nouvel effet
                        if etat_normalise == 'entrer':
                            stock_final = stock_apres_annulation + quantite_nouvelle
                        else:  # nouveau = 'sortir'
                            stock_final = max(Decimal('0'), stock_apres_annulation - quantite_nouvelle)
                    else:
                        # L'état n'a pas changé : utiliser directement la différence
                        if etat_normalise == 'entrer':
                            # État "entrer" : augmente le stock (comme facture achat)
                            stock_final = stock_avant_modification + difference
                        else:  # 'sortir'
                            # État "sortir" : diminue le stock
                            stock_final = max(Decimal('0'), stock_avant_modification - difference)
                    
                    # S'assurer que le stock ne devient pas négatif
                    if stock_final < 0:
                        messages.error(request, f'Stock insuffisant pour l\'article {article.designation}. Stock actuel: {stock_avant_modification}, différence: {difference}')
                        raise ValueError(f"Stock insuffisant: {stock_avant_modification} + {difference} = {stock_final} < 0")
                    
                    # Mettre à jour le stock avec le stock final calculé
                    article.stock_actuel = stock_final
                    article.save()
                    
                    # Déterminer le type de mouvement et le commentaire selon la différence
                    if etat_normalise == 'entrer':
                        # État "entrer" : augmente le stock
                        if difference > 0:
                            type_mouvement = 'entree'
                            quantite_mouvement = difference
                            commentaire = f"Transfert entrer - Facture {facture.reference_transfert} (modification - augmentation: {quantite_ancienne} → {quantite_nouvelle}, +{difference})"
                        elif difference < 0:
                            type_mouvement = 'sortie'
                            quantite_mouvement = abs(difference)
                            commentaire = f"Transfert entrer - Facture {facture.reference_transfert} (modification - diminution: {quantite_ancienne} → {quantite_nouvelle}, -{abs(difference)})"
                        else:
                            type_mouvement = None
                            quantite_mouvement = Decimal('0')
                            commentaire = f"Transfert entrer - Facture {facture.reference_transfert} (modification - quantité inchangée: {quantite_nouvelle})"
                    else:
                        # État "sortir" : diminue le stock
                        if difference > 0:
                            type_mouvement = 'sortie'
                            quantite_mouvement = difference
                            commentaire = f"Transfert sortir - Facture {facture.reference_transfert} (modification - augmentation: {quantite_ancienne} → {quantite_nouvelle}, -{difference})"
                        elif difference < 0:
                            type_mouvement = 'entree'
                            quantite_mouvement = abs(difference)
                            commentaire = f"Transfert sortir - Facture {facture.reference_transfert} (modification - diminution: {quantite_ancienne} → {quantite_nouvelle}, +{abs(difference)})"
                        else:
                            type_mouvement = None
                            quantite_mouvement = Decimal('0')
                            commentaire = f"Transfert sortir - Facture {facture.reference_transfert} (modification - quantité inchangée: {quantite_nouvelle})"
                    
                    article.dernier_prix_achat = prix_achat_ligne
                    # Le suivi de stock est toujours activé automatiquement lors de toute modification du stock
                    article.suivi_stock = True
                    
                    total_quantite += quantite_nouvelle
                    
                    # Créer un mouvement de stock pour refléter le changement
                    if quantite_mouvement > 0 and type_mouvement:
                        try:
                            MouvementStock.objects.create(
                                article=article,
                                agence=agence,
                                type_mouvement=type_mouvement,
                                date_mouvement=timezone.now(),
                                numero_piece=facture.reference_transfert,
                                quantite_stock=article.stock_actuel,
                                stock_initial=stock_avant_modification,
                                solde=article.stock_actuel,
                                quantite=quantite_mouvement,
                                cout_moyen_pondere=float(article.prix_achat),
                                stock_permanent=float(article.stock_actuel * article.prix_achat),
                                facture_transfert=facture,
                                commentaire=commentaire
                            )
                        except Exception as e:
                            print(f"[WARNING] ERREUR MOUVEMENT STOCK TRANSFERT MODIFICATION: {e}")
                
                facture.quantite = int(total_quantite)
                facture.save()
            
            # Retirer le flag de modification en cours
            request.session.pop('modification_en_cours', None)
            
            messages.success(request, f'Facture de transfert "{reference_transfert}" modifiée avec succès!')
            return redirect('consulter_factures_transfert')
            
        except Exception as e:
            # Retirer le flag en cas d'erreur
            request.session.pop('modification_en_cours', None)
            messages.error(request, f'Erreur lors de la modification de la facture: {str(e)}')
            return redirect('modifier_facture_transfert', facture_id=facture_id)
    
    # GET - Afficher le formulaire pré-rempli
    statut_choices = FactureTransfert.STATUT_CHOICES
    lignes = LigneFactureTransfert.objects.select_related('article').filter(facture_transfert=facture)
    
    initial_articles = []
    for ligne in lignes:
        if ligne.article:
            initial_articles.append({
                'id': ligne.article.id,
                'designation': ligne.article.designation,
                'reference_article': ligne.article.reference_article,
                'stock_actuel': float(ligne.article.stock_actuel),
                'prix_achat': float(ligne.prix_unitaire),
                'quantite': ligne.quantite
            })
    
    etat_actuel = (facture.etat or 'sortir').strip().lower()
    if etat_actuel == 'sortie':
        etat_actuel = 'sortir'
    
    context = {
        'facture': facture,
        'statut_choices': statut_choices,
        'initial_articles_json': json.dumps(initial_articles),
        'etat_actuel': etat_actuel,
    }
    return render(request, 'supermarket/stock/modifier_facture_transfert.html', context)

@login_required
def supprimer_facture_transfert(request, facture_id):
    """Vue pour supprimer une facture de transfert"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_factures_transfert')
    
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.select_related('agence_source').get(id=facture_id, agence_source=agence)
        facture_name = facture.reference_transfert
        etat_normalise = (facture.etat or 'sortir').strip().lower()
        
        with transaction.atomic():
            # NE PLUS SUPPRIMER LES MOUVEMENTS - Créer des mouvements de correction à la place
            lignes = LigneFactureTransfert.objects.select_related('article').filter(facture_transfert=facture)
            
            for ligne in lignes:
                article = ligne.article
                if not article:
                    continue
                
                quantite = Decimal(str(ligne.quantite))
                ancien_stock = article.stock_actuel
                
                if etat_normalise == 'entrer':
                    # La facture avait augmenté le stock, on revient en arrière
                    article.stock_actuel = max(Decimal('0'), article.stock_actuel - quantite)
                    commentaire = f"Correction - Suppression facture transfert {facture.reference_transfert} - retrait stock"
                else:
                    # La facture avait diminué le stock, on le restaure
                    article.stock_actuel += quantite
                    commentaire = f"Correction - Suppression facture transfert {facture.reference_transfert} - restauration stock"
                
                # Le suivi de stock est toujours activé automatiquement lors de toute modification du stock
                article.suivi_stock = True
                article.save()
                
                # Créer un mouvement de correction (les mouvements originaux sont conservés pour traçabilité)
                try:
                    MouvementStock.objects.create(
                        article=article,
                        agence=agence,
                        type_mouvement='ajustement',
                        date_mouvement=timezone.now(),
                        numero_piece=f"SUPP-{facture.reference_transfert}",
                        quantite_stock=article.stock_actuel,
                        stock_initial=ancien_stock,
                        solde=article.stock_actuel,
                        quantite=quantite,
                        cout_moyen_pondere=float(article.prix_achat),
                        stock_permanent=float(article.stock_actuel * article.prix_achat),
                        commentaire=commentaire,
                        facture_transfert=facture
                    )
                except Exception as mouvement_error:
                    print(f"[WARNING] Erreur enregistrement mouvement correction suppression transfert: {mouvement_error}")
            
            facture.delete()
        
        messages.success(request, f'Facture de transfert "{facture_name}" supprimée avec succès!')
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_factures_transfert')
# ===== RECHERCHE D'ARTICLES POUR STOCK =====

@login_required
@require_stock_access
def search_articles_stock(request):
    """Vue pour la recherche d'articles dans le module de stock"""
    search_term = request.GET.get('q', '')
    
    print(f"[SEARCH] search_articles_stock: recherche pour '{search_term}'")
    
    # Vérifier d'abord s'il y a des articles dans la base de données
    total_articles = Article.objects.count()
    print(f"[CHART] Total articles dans la base de données: {total_articles}")
    
    if total_articles == 0:
        print("[ERREUR] Aucun article dans la base de données!")
        return JsonResponse({'articles': []})
    
    agence = get_user_agence(request)
    print(f"[BUILDING] Agence trouvée: {agence}")
    print(f"[BUILDING] ID de l'agence: {agence.id_agence if agence else 'None'}")
    print(f"[BUILDING] Nom de l'agence: {agence.nom_agence if agence else 'None'}")
    
    if not agence:
        print("[ERREUR] Aucune agence trouvée")
        return JsonResponse({'articles': []})
    
    # Test: afficher tous les articles sans filtre d'agence
    articles_all = Article.objects.all()
    print(f"[PACKAGE] Tous les articles (toutes agences): {articles_all.count()}")
    for article in articles_all[:5]:  # Afficher les 5 premiers
        print(f"  - {article.designation} (Agence: {article.agence.nom_agence if article.agence else 'None'})")
    
    # Test: afficher les articles de cette agence spécifique
    articles_agence = Article.objects.filter(agence=agence)
    print(f"[PACKAGE] Articles de l'agence {agence.nom_agence}: {articles_agence.count()}")
    for article in articles_agence[:5]:  # Afficher les 5 premiers
        print(f"  - {article.reference_article} - {article.designation} (ID: {article.id})")
    
    articles = []
    
    if search_term and len(search_term) >= 1:
        # Recherche avec filtre d'agence : recherche dans designation ET reference_article
        from django.db.models import Q
        articles = Article.objects.filter(
            agence=agence
        ).filter(
            Q(designation__icontains=search_term) | Q(reference_article__icontains=search_term)
        ).order_by('designation')[:50]
        print(f"[SEARCH] Articles trouvés avec recherche '{search_term}' (agence {agence.nom_agence}): {articles.count()}")
        for article in articles[:10]:  # Afficher les 10 premiers résultats
            print(f"  - {article.reference_article} - {article.designation} (ID: {article.id})")
    else:
        # Afficher tous les articles de l'agence si pas de terme de recherche
        articles = Article.objects.filter(agence=agence).order_by('designation')[:50]
        print(f"[PACKAGE] Tous les articles de l'agence {agence.nom_agence}: {articles.count()}")
    
    # Convertir les articles en format JSON
    articles_data = []
    for article in articles:
        articles_data.append({
            'id': article.id,
            'designation': article.designation,
            'prix_achat': float(article.prix_achat) if article.prix_achat else 0.0,
            'stock': float(article.stock_actuel) if article.stock_actuel else 0.0,
            'reference_article': article.reference_article or '',
        })
        print(f"[NOTE] Article ajouté: {article.reference_article} - {article.designation} (ID: {article.id}, Agence: {article.agence.nom_agence})")
    
    print(f"[CHART] Total articles_data: {len(articles_data)}")
    return JsonResponse({'articles': articles_data})

def create_test_articles(request):
    """Vue temporaire pour créer des articles de test"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'error': 'Aucune agence trouvée'})
    
    # Créer quelques articles de test
    test_articles = [
        {'designation': 'Ordinateur Portable', 'prix_achat': 500000, 'prix_vente': 600000, 'stock_actuel': 10},
        {'designation': 'Souris USB', 'prix_achat': 5000, 'prix_vente': 7500, 'stock_actuel': 50},
        {'designation': 'Clavier Mécanique', 'prix_achat': 15000, 'prix_vente': 20000, 'stock_actuel': 25},
        {'designation': 'Écran 24 pouces', 'prix_achat': 80000, 'prix_vente': 100000, 'stock_actuel': 15},
        {'designation': 'Casque Audio', 'prix_achat': 25000, 'prix_vente': 35000, 'stock_actuel': 30},
    ]
    
    created_articles = []
    for article_data in test_articles:
        article, created = Article.objects.get_or_create(
            designation=article_data['designation'],
            agence=agence,
            defaults=article_data
        )
        if created:
            created_articles.append(article.designation)
    
    return JsonResponse({
        'message': f'Articles créés: {len(created_articles)}',
        'articles': created_articles
    })

# ==================== INVENTAIRE DE STOCK ====================

@login_required
@require_stock_access
def inventaire_stock(request):
    """Vue pour la page d'inventaire de stock"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques du stock
    total_articles = articles.count()
    total_quantite = articles.aggregate(total=Sum('stock_actuel'))['total'] or 0
    valeur_totale_stock = articles.aggregate(
        total=Sum(F('stock_actuel') * F('prix_achat'))
    )['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'total_quantite': total_quantite,
        'valeur_totale_stock': valeur_totale_stock,
    }
    
    return render(request, 'supermarket/stock/inventaire_stock.html', context)

@login_required
def generer_inventaire(request):
    """Vue pour générer l'inventaire selon les critères sélectionnés"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        famille_id = request.POST.get('famille')
        selection_articles = request.POST.get('selection_articles', 'tous')  # 'tous' ou 'selectionnes'
        articles_selectionnes = request.POST.getlist('articles_selectionnes[]')
        
        print(f"[SEARCH] PARAMÈTRES INVENTAIRE:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Famille: {famille_id}")
        print(f"  - Sélection: {selection_articles}")
        print(f"  - Articles sélectionnés: {articles_selectionnes}")
        
        # Filtrer les articles selon les critères
        articles_query = Article.objects.filter(agence=agence)
        
        # Filtre par famille
        if famille_id and famille_id != '':
            articles_query = articles_query.filter(categorie_id=famille_id)
        
        # Filtre par sélection d'articles
        if selection_articles == 'selectionnes' and articles_selectionnes:
            articles_query = articles_query.filter(id__in=articles_selectionnes)
        
        articles = articles_query.order_by('designation')
        
        print(f"[PACKAGE] ARTICLES FILTRÉS: {articles.count()}")
        
        # Créer l'inventaire
        numero_inventaire = f"INV-{timezone.now().strftime('%Y%m%d-%H%M%S')}"
        
        # Récupérer l'employé responsable
        employe = Employe.objects.filter(compte__agence=agence).first()
        
        inventaire = InventaireStock.objects.create(
            numero_inventaire=numero_inventaire,
            date_debut=timezone.now(),
            statut='en_cours',
            agence=agence,
            responsable=employe,
            commentaire=f"Inventaire généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        )
        
        # Créer les lignes d'inventaire
        total_quantite = 0
        total_valeur = 0
        
        for article in articles:
            valeur = float(article.stock_actuel) * float(article.prix_achat)
            
            LigneInventaireStock.objects.create(
                inventaire=inventaire,
                reference_article=article.reference_article,
                designation=article.designation,
                quantite_stock=article.stock_actuel,
                prix_unitaire=article.prix_achat,
                valeur=valeur,
                conditionnement=article.conditionnement,
                article=article
            )
            
            total_quantite += article.stock_actuel
            total_valeur += valeur
        
        # Marquer l'inventaire comme terminé
        inventaire.date_fin = timezone.now()
        inventaire.statut = 'termine'
        inventaire.save()
        
        print(f"[OK] INVENTAIRE CRÉÉ: {numero_inventaire}")
        print(f"[CHART] TOTAUX: {total_quantite} articles, {total_valeur} FCFA")
        
        return JsonResponse({
            'success': True,
            'inventaire_id': inventaire.id,
            'numero_inventaire': numero_inventaire,
            'total_articles': articles.count(),
            'total_quantite': total_quantite,
            'total_valeur': total_valeur
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR GÉNÉRATION INVENTAIRE: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_excel(request):
    """Vue pour exporter l'inventaire en format Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible, utilisation du format CSV")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Inventaire {inventaire.numero_inventaire}"
        
        # Style du titre
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire le titre "Inventaire" avec la date du jour
        date_du_jour = timezone.now().strftime('%d/%m/%Y')
        title_cell = ws.cell(row=1, column=1, value=f"Inventaire - {date_du_jour}")
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # Fusionner les cellules pour le titre (sur toutes les colonnes)
        from openpyxl.utils import get_column_letter
        ws.merge_cells(f'A1:{get_column_letter(6)}1')
        
        # Ligne vide
        ws.row_dimensions[2].height = 5
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité en Stock', 'Prix Unitaire (FCFA)', 'Valeur Totale (FCFA)']
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire les en-têtes (décalés à la ligne 3)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Écrire les données (décalées à partir de la ligne 4)
        total_quantite = 0
        total_valeur = 0
        
        for idx, ligne in enumerate(lignes):
            row_num = 4 + idx  # Commencer à la ligne 4 (après titre ligne 1, ligne vide ligne 2, en-têtes ligne 3)
            ws.cell(row=row_num, column=1, value=ligne.reference_article)
            ws.cell(row=row_num, column=2, value=ligne.designation)
            ws.cell(row=row_num, column=3, value=ligne.conditionnement)
            ws.cell(row=row_num, column=4, value=ligne.quantite_stock)
            ws.cell(row=row_num, column=5, value=float(ligne.prix_unitaire))
            ws.cell(row=row_num, column=6, value=float(ligne.valeur))
            
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux (décalée de 2 lignes supplémentaires)
        total_row = 4 + len(lignes) + 2  # Ligne de données + 1 ligne vide + ligne totaux
        ws.cell(row=total_row, column=3, value="TOTAL GÉNÉRAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_quantite).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=total_valeur).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        column_widths = [15, 40, 15, 15, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le classeur dans la réponse
        wb.save(response)
        
        print(f"[CHART] EXPORT EXCEL - Inventaire {inventaire.numero_inventaire}")
        print(f"[CHART] EXPORT EXCEL - {len(lignes)} articles exportés")
        print(f"[CHART] EXPORT EXCEL - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_pdf(request):
    """Vue pour exporter l'inventaire en format PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Centre
        )
        
        # Titre
        title = Paragraph(f"INVENTAIRE DE STOCK - {inventaire.numero_inventaire}", title_style)
        elements.append(title)
        
        # Informations de l'inventaire
        info_data = [
            ['Date de génération:', inventaire.date_debut.strftime('%d/%m/%Y à %H:%M')],
            ['Agence:', agence.nom_agence],
            ['Responsable:', inventaire.responsable.compte.nom_complet if inventaire.responsable else 'Non spécifié'],
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 20))
        
        # Tableau des articles
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité', 'Prix Unitaire', 'Valeur']
        
        # Données du tableau
        data = [headers]
        total_quantite = 0
        total_valeur = 0
        
        for ligne in lignes:
            row = [
                ligne.reference_article,
                ligne.designation,
                ligne.conditionnement,
                str(ligne.quantite_stock),
                f"{float(ligne.prix_unitaire):,.0f}",
                f"{float(ligne.valeur):,.0f}"
            ]
            data.append(row)
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux
        data.append(['', '', 'TOTAL GÉNÉRAL:', str(total_quantite), '', f"{total_valeur:,.0f}"])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 6*cm, 2.5*cm, 2*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            # En-têtes
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Données
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Ligne des totaux
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF - Inventaire {inventaire.numero_inventaire}")
        print(f"📄 EXPORT PDF - {len(lignes)} articles exportés")
        print(f"📄 EXPORT PDF - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_csv(request):
    """Vue pour exporter l'inventaire en format CSV (alternative si Excel n'est pas disponible)"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité en Stock', 'Prix Unitaire (FCFA)', 'Valeur Totale (FCFA)']
        writer.writerow(headers)
        
        # Données
        total_quantite = 0
        total_valeur = 0
        
        for ligne in lignes:
            row = [
                ligne.reference_article,
                ligne.designation,
                ligne.conditionnement,
                ligne.quantite_stock,
                float(ligne.prix_unitaire),
                float(ligne.valeur)
            ]
            writer.writerow(row)
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['', '', 'TOTAL GÉNÉRAL:', total_quantite, '', total_valeur])
        
        print(f"[CHART] EXPORT CSV - Inventaire {inventaire.numero_inventaire}")
        print(f"[CHART] EXPORT CSV - {len(lignes)} articles exportés")
        print(f"[CHART] EXPORT CSV - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== STATISTIQUES DE VENTE ====================

@login_required
def statistiques_vente(request):
    """Vue pour la page des statistiques de vente"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques générales
    total_articles = articles.count()
    
    # Calculer les statistiques de vente des 30 derniers jours
    date_debut = timezone.now().date() - timezone.timedelta(days=30)
    
    # Récupérer les ventes des 30 derniers jours
    ventes_recentes = LigneFactureVente.objects.filter(
        facture_vente__agence=agence,
        facture_vente__date__gte=date_debut
    ).select_related('article', 'facture_vente')
    
    # Calculer le chiffre d'affaires total
    chiffre_affaires_total = float(ventes_recentes.aggregate(
        total=Sum(F('quantite') * F('prix_unitaire'))
    )['total'] or 0)
    
    # Calculer la marge totale
    marge_totale = Decimal('0')
    for vente in ventes_recentes:
        prix_achat = Decimal(str(vente.article.prix_achat))
        prix_vente = Decimal(str(vente.prix_unitaire))
        marge_unitaire = prix_vente - prix_achat
        marge_totale += marge_unitaire * Decimal(str(vente.quantite))
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'chiffre_affaires_total': chiffre_affaires_total,
        'marge_totale': float(marge_totale),
        'pourcentage_marge_global': (float(marge_totale) / float(chiffre_affaires_total) * 100) if chiffre_affaires_total > 0 else 0,
    }
    
    return render(request, 'supermarket/stock/statistiques_vente.html', context)

@login_required
def generer_statistiques_vente(request):
    """Vue pour générer les statistiques de vente selon les critères sélectionnés"""
    print("[START] DÉBUT GENERER_STATISTIQUES_VENTE")
    
    if request.method != 'POST':
        print("[ERREUR] Méthode non POST")
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
        print(f"[SEARCH] Agence récupérée: {agence}")
    except Exception as e:
        print(f"[ERREUR] Erreur get_user_agence: {e}")
        return JsonResponse({'success': False, 'error': f'Agence non trouvée: {str(e)}'})
    
    try:
        print("[SEARCH] Début du traitement des paramètres")
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        famille_id = request.POST.get('famille')
        selection_articles = request.POST.get('selection_articles', 'tous')
        articles_selectionnes = request.POST.getlist('articles_selectionnes[]')
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        print(f"[CHART] PARAMÈTRES STATISTIQUES:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Famille: {famille_id}")
        print(f"  - Sélection: {selection_articles}")
        print(f"  - Articles sélectionnés: {articles_selectionnes}")
        
        # Convertir les dates
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Format de date invalide: {str(e)}'})
        
        # Filtrer les articles selon les critères
        print(f"[SEARCH] Filtrage des articles pour agence: {agence}")
        articles_query = Article.objects.filter(agence=agence)
        print(f"[SEARCH] Articles de base: {articles_query.count()}")
        
        # Filtre par famille
        if famille_id and famille_id != '':
            print(f"[SEARCH] Filtrage par famille: {famille_id}")
            articles_query = articles_query.filter(categorie_id=famille_id)
            print(f"[SEARCH] Articles après filtre famille: {articles_query.count()}")
        
        # Filtre par sélection d'articles
        if selection_articles == 'selectionnes' and articles_selectionnes:
            print(f"[SEARCH] Filtrage par sélection: {articles_selectionnes}")
            articles_query = articles_query.filter(id__in=articles_selectionnes)
            print(f"[SEARCH] Articles après filtre sélection: {articles_query.count()}")
        
        articles = articles_query.order_by('designation')
        print(f"[PACKAGE] ARTICLES FILTRÉS: {articles.count()}")
        
        # Récupérer les ventes pour la période
        ventes = LigneFactureVente.objects.filter(
            facture_vente__agence=agence,
            facture_vente__date__gte=date_debut_obj,
            facture_vente__date__lte=date_fin_obj,
            article__in=articles
        ).select_related('article', 'facture_vente')
        
        # Calculer les statistiques par article
        statistiques_articles = []
        chiffre_affaires_total = 0.0
        marge_totale = 0.0
        quantite_totale_vendue = 0.0
        
        for article in articles:
            # Récupérer les ventes de cet article
            ventes_article = ventes.filter(article=article)
            
            # Calculer les totaux pour cet article
            quantite_vendue = float(ventes_article.aggregate(total=Sum('quantite'))['total'] or 0)
            chiffre_affaires_article = float(ventes_article.aggregate(
                total=Sum(F('quantite') * F('prix_unitaire'))
            )['total'] or 0)
            
            # Calculer la marge
            prix_achat = Decimal(str(article.prix_achat))
            marge_unitaire = Decimal('0')
            marge_article = Decimal('0')
            
            if quantite_vendue > 0:
                prix_vente_moyen = Decimal(str(chiffre_affaires_article)) / Decimal(str(quantite_vendue))
                marge_unitaire = prix_vente_moyen - prix_achat
                marge_article = marge_unitaire * Decimal(str(quantite_vendue))
            
            # Calculer le pourcentage de marge
            if chiffre_affaires_article > 0:
                pourcentage_marge = (marge_article / Decimal(str(chiffre_affaires_article)) * Decimal('100'))
            else:
                pourcentage_marge = Decimal('0')
            
            if quantite_vendue > 0:  # Ne garder que les articles vendus
                statistiques_articles.append({
                    'reference_article': article.reference_article,
                    'designation': article.designation,
                    'quantite_vendue': float(quantite_vendue),
                    'chiffre_affaires': float(chiffre_affaires_article),
                    'marge_profit': float(marge_article),
                    'pourcentage_marge': float(pourcentage_marge),
                })
                
                chiffre_affaires_total += float(chiffre_affaires_article)
                marge_totale += float(marge_article)
                quantite_totale_vendue += float(quantite_vendue)
        
        # Calculer le pourcentage de marge global
        if chiffre_affaires_total > 0:
            pourcentage_marge_global = (marge_totale / float(chiffre_affaires_total) * 100)
        else:
            pourcentage_marge_global = 0
        
        print(f"[CHART] STATISTIQUES GÉNÉRÉES:")
        print(f"  - Articles vendus: {len(statistiques_articles)}")
        print(f"  - Quantité totale vendue: {quantite_totale_vendue}")
        print(f"  - Chiffre d'affaires total: {chiffre_affaires_total}")
        print(f"  - Marge totale: {marge_totale}")
        print(f"  - Pourcentage marge global: {pourcentage_marge_global:.2f}%")
        
        # Stocker les statistiques dans la session pour l'export (conversion en types sérialisables)
        request.session['statistiques_vente'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'statistiques_articles': statistiques_articles,  # Déjà converties en float
            'chiffre_affaires_total': float(chiffre_affaires_total),
            'marge_totale': float(marge_totale),
            'quantite_totale_vendue': int(quantite_totale_vendue),
            'pourcentage_marge_global': float(pourcentage_marge_global),
        }
        
        return JsonResponse({
            'success': True,
            'total_articles': len(statistiques_articles),
            'quantite_totale_vendue': int(quantite_totale_vendue),
            'chiffre_affaires_total': float(chiffre_affaires_total),
            'marge_totale': float(marge_totale),
            'pourcentage_marge_global': float(pourcentage_marge_global),
            'statistiques_articles': statistiques_articles  # Ajouter les données détaillées
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR GÉNÉRATION STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        error_details = f"Erreur: {str(e)}"
        if hasattr(e, '__class__'):
            error_details += f" (Type: {e.__class__.__name__})"
        return JsonResponse({'success': False, 'error': error_details})
@login_required
def export_statistiques_excel(request):
    """Vue pour exporter les statistiques de vente en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible, utilisation du format CSV")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Statistiques Vente {date_debut} - {date_fin}"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tête du document
        ws.merge_cells('A1:F1')
        ws['A1'] = f"STATISTIQUES DE VENTE - {agence.nom_agence}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Période: du {date_debut} au {date_fin}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes des colonnes
        headers = ['Référence', 'Désignation', 'Quantité Vendue', 'Chiffre d\'Affaires (FCFA)', 'Marge Profit (FCFA)', 'Pourcentage Marge (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données
        row = 5
        for stat in statistiques_articles:
            ws.cell(row=row, column=1, value=stat['reference_article'])
            ws.cell(row=row, column=2, value=stat['designation'])
            ws.cell(row=row, column=3, value=stat['quantite_vendue'])
            ws.cell(row=row, column=4, value=float(stat['chiffre_affaires']))
            ws.cell(row=row, column=5, value=float(stat['marge_profit']))
            ws.cell(row=row, column=6, value=float(stat['pourcentage_marge']))
            row += 1
        
        # Ligne des totaux
        row += 1
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value="").font = Font(bold=True)
        ws.cell(row=row, column=3, value=quantite_totale_vendue).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(chiffre_affaires_total)).font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(marge_totale)).font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(pourcentage_marge_global)).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 18
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le fichier dans la réponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_statistiques_pdf(request):
    """Vue pour exporter les statistiques de vente en PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            textColor=colors.darkblue
        )
        
        # Éléments du document
        elements = []
        
        # Titre
        title = Paragraph(f"STATISTIQUES DE VENTE - {agence.nom_agence}", title_style)
        elements.append(title)
        
        # Informations de la période
        period_text = f"<b>Période:</b> du {date_debut} au {date_fin}"
        period_para = Paragraph(period_text, styles['Normal'])
        elements.append(period_para)
        elements.append(Spacer(1, 20))
        
        # Données du tableau
        data = [['Référence', 'Désignation', 'Qté Vendue', 'Chiffre d\'Affaires', 'Marge Profit', 'Marge %']]
        
        for stat in statistiques_articles:
            data.append([
                stat['reference_article'],
                stat['designation'][:30] + '...' if len(stat['designation']) > 30 else stat['designation'],
                str(stat['quantite_vendue']),
                f"{float(stat['chiffre_affaires']):,.0f} FCFA",
                f"{float(stat['marge_profit']):,.0f} FCFA",
                f"{float(stat['pourcentage_marge']):.1f}%"
            ])
        
        # Ligne des totaux
        data.append([
            'TOTAL GÉNÉRAL',
            '',
            str(quantite_totale_vendue),
            f"{float(chiffre_affaires_total):,.0f} FCFA",
            f"{float(marge_totale):,.0f} FCFA",
            f"{float(pourcentage_marge_global):.1f}%"
        ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 6*cm, 2*cm, 3*cm, 3*cm, 2*cm])
        
        # Style du tableau
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"📄 EXPORT PDF STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"📄 EXPORT PDF STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_statistiques_csv(request):
    """Vue pour exporter les statistiques de vente en CSV"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Quantité Vendue', 'Chiffre d\'Affaires (FCFA)', 'Marge Profit (FCFA)', 'Pourcentage Marge (%)']
        writer.writerow(headers)
        
        # Données
        for stat in statistiques_articles:
            row = [
                stat['reference_article'],
                stat['designation'],
                stat['quantite_vendue'],
                float(stat['chiffre_affaires']),
                float(stat['marge_profit']),
                float(stat['pourcentage_marge'])
            ]
            writer.writerow(row)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['TOTAL GÉNÉRAL', '', quantite_totale_vendue, chiffre_affaires_total, marge_totale, pourcentage_marge_global])
        
        print(f"[CHART] EXPORT CSV STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"[CHART] EXPORT CSV STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"[CHART] EXPORT CSV STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_statistiques(request):
    """Vue de test pour diagnostiquer les problèmes"""
    try:
        print("🧪 TEST STATISTIQUES - Début")
        
        # Test 1: Récupération de l'agence
        agence = get_user_agence(request)
        print(f"🧪 TEST - Agence récupérée: {agence}")
        
        # Test 2: Vérification des imports
        print("🧪 TEST - Import Decimal OK")
        
        # Test 3: Vérification des modèles
        articles_count = Article.objects.filter(agence=agence).count()
        print(f"🧪 TEST - Articles trouvés: {articles_count}")
        
        # Test 4: Vérification des ventes
        ventes_count = LigneFactureVente.objects.filter(facture_vente__agence=agence).count()
        print(f"🧪 TEST - Ventes trouvées: {ventes_count}")
        
        return JsonResponse({
            'success': True,
            'message': 'Tests réussis',
            'agence': str(agence),
            'articles_count': articles_count,
            'ventes_count': ventes_count
        })
        
    except Exception as e:
        print(f"🧪 TEST - Erreur: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== MOUVEMENTS DE STOCK ====================

@login_required
@require_stock_access
def mouvements_stock(request):
    """Vue pour la page des mouvements de stock"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques générales
    total_articles = articles.count()
    
    # Calculer les statistiques des mouvements des 30 derniers jours
    date_debut = timezone.now().date() - timezone.timedelta(days=30)
    
    # Récupérer les mouvements des 30 derniers jours
    mouvements_recentes = MouvementStock.objects.filter(
        agence=agence,
        date_mouvement__gte=date_debut
    ).select_related('article', 'fournisseur', 'facture_vente', 'facture_achat')
    
    # Statistiques des mouvements
    total_mouvements = mouvements_recentes.count()
    mouvements_entree = mouvements_recentes.filter(type_mouvement='entree').count()
    mouvements_sortie = mouvements_recentes.filter(type_mouvement='sortie').count()
    
    # Valeur totale du stock permanent
    valeur_stock_permanent = mouvements_recentes.aggregate(
        total=Sum('stock_permanent')
    )['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'total_mouvements': total_mouvements,
        'mouvements_entree': mouvements_entree,
        'mouvements_sortie': mouvements_sortie,
        'valeur_stock_permanent': valeur_stock_permanent,
    }
    
    return render(request, 'supermarket/stock/mouvements_stock.html', context)

@login_required
@require_stock_access
def consulter_mouvements_stock(request):
    """Vue pour consulter les mouvements de stock selon les critères sélectionnés"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        article_ids = request.POST.getlist('articles')  # Récupérer tous les articles sélectionnés
        type_mouvement = request.POST.get('type_mouvement', '')
        
        print(f"[CHART] PARAMÈTRES MOUVEMENTS:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Articles: {article_ids}")
        print(f"  - Type mouvement: {type_mouvement}")
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        if not article_ids:
            return JsonResponse({'success': False, 'error': 'Veuillez sélectionner au moins un article'})
        
        # Convertir les dates
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Format de date invalide: {str(e)}'})
        
        # Récupérer les articles
        try:
            articles = Article.objects.filter(id__in=article_ids, agence=agence)
            if not articles.exists():
                return JsonResponse({'success': False, 'error': 'Aucun article valide trouvé'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Erreur lors de la récupération des articles: {str(e)}'})
        
        # Filtrer les mouvements selon les critères
        # Utiliser __date pour comparer seulement les dates (ignorer l'heure)
        mouvements_query = MouvementStock.objects.filter(
            agence=agence,
            article__in=articles,  # Filtrer par plusieurs articles
            date_mouvement__date__gte=date_debut_obj,
            date_mouvement__date__lte=date_fin_obj
        ).select_related('article', 'fournisseur', 'facture_vente', 'facture_achat').order_by('article__reference_article', 'date_mouvement')
        
        # Filtre par type de mouvement
        if type_mouvement and type_mouvement != '':
            mouvements_query = mouvements_query.filter(type_mouvement=type_mouvement)
        
        mouvements = mouvements_query
        
        print(f"[PACKAGE] MOUVEMENTS FILTRÉS: {mouvements.count()}")
        
        # Debug: Afficher quelques mouvements pour vérifier
        if mouvements.count() > 0:
            print("[LIST] PREMIERS MOUVEMENTS TROUVÉS:")
            for i, mvt in enumerate(mouvements[:3]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement}")
        else:
            print("[ERREUR] AUCUN MOUVEMENT TROUVÉ - Vérifions les mouvements existants:")
            tous_mouvements = MouvementStock.objects.filter(agence=agence, article__in=articles)
            print(f"[CHART] Total mouvements pour ces articles: {tous_mouvements.count()}")
            for i, mvt in enumerate(tous_mouvements[:3]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement}")
        
        # Calculer les statistiques
        total_mouvements = mouvements.count()
        mouvements_entree = mouvements.filter(type_mouvement='entree').count()
        mouvements_sortie = mouvements.filter(type_mouvement='sortie').count()
        
        # Calculer la valeur totale du stock permanent
        valeur_stock_permanent = mouvements.aggregate(
            total=Sum('stock_permanent')
        )['total'] or 0
        
        # Stocker les données dans la session pour l'export
        mouvements_data = []
        for mouvement in mouvements:
            # Déterminer le tiers
            tiers = ""
            if mouvement.fournisseur:
                tiers = f"Fournisseur: {mouvement.fournisseur.intitule}"
            elif mouvement.facture_vente:
                tiers = f"Client: {mouvement.facture_vente.client.intitule if mouvement.facture_vente.client else 'N/A'}"
            elif mouvement.facture_achat:
                tiers = f"Fournisseur: {mouvement.facture_achat.fournisseur.intitule if mouvement.facture_achat.fournisseur else 'N/A'}"
            
            # Debug pour vérifier les données d'article
            print(f"[SEARCH] DEBUG ARTICLE: {mouvement.article.reference_article} - {mouvement.article.designation}")
            
            mouvements_data.append({
                'date_mouvement': mouvement.date_mouvement.strftime('%Y-%m-%d %H:%M'),
                'type_mouvement': mouvement.type_mouvement,
                'type_mouvement_display': mouvement.get_type_mouvement_display(),
                'reference_article': str(mouvement.article.reference_article) if mouvement.article.reference_article else 'N/A',
                'designation': str(mouvement.article.designation) if mouvement.article.designation else 'N/A',
                'tiers': tiers,
                'stock_initial': float(mouvement.stock_initial) if mouvement.stock_initial is not None else 0.0,
                'quantite': float(mouvement.quantite) if mouvement.quantite is not None else 0.0,
                'solde': float(mouvement.solde) if mouvement.solde is not None else 0.0,
                'quantite_stock': float(mouvement.quantite_stock) if mouvement.quantite_stock is not None else 0.0,
                'cout_moyen_pondere': float(mouvement.cout_moyen_pondere) if mouvement.cout_moyen_pondere is not None else 0.0,
                'stock_permanent': float(mouvement.stock_permanent) if mouvement.stock_permanent is not None else 0.0,
                'numero_piece': mouvement.numero_piece,
                'commentaire': mouvement.commentaire or '',
            })
        
        print(f"[CHART] MOUVEMENTS GÉNÉRÉS:")
        print(f"  - Total mouvements: {total_mouvements}")
        print(f"  - Entrées: {mouvements_entree}")
        print(f"  - Sorties: {mouvements_sortie}")
        print(f"  - Valeur stock permanent: {valeur_stock_permanent}")
        
        # Stocker les mouvements dans la session pour l'export (sans Decimal)
        session_payload = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'articles_count': len(articles),
            'mouvements_data': mouvements_data,
            'total_mouvements': total_mouvements,
            'mouvements_entree': mouvements_entree,
            'mouvements_sortie': mouvements_sortie,
            'valeur_stock_permanent': float(valeur_stock_permanent),
            'article_reference': ', '.join([str(article.reference_article or 'N/A') for article in articles]),
            'article_designation': ', '.join([str(article.designation or 'N/A') for article in articles]),
        }
        from django.core.serializers.json import DjangoJSONEncoder
        import json
        clean_payload = json.loads(json.dumps(session_payload, cls=DjangoJSONEncoder))
        try:
            json.dumps(clean_payload)
        except TypeError as json_error:
            print(f"[ERROR] JSON serialization issue dans mouvements_stock: {json_error}")
            for idx, mouvement in enumerate(clean_payload.get('mouvements_data', [])):
                for key, value in mouvement.items():
                    print(f"    mouvement[{idx}].{key} -> {type(value)} = {value}")
            raise
        request.session['mouvements_stock'] = clean_payload
        
        # Préparer les informations des articles
        articles_info = []
        for article in articles:
            article_mouvements = mouvements.filter(article=article)
            articles_info.append({
                'id': article.id,
                'reference': article.reference_article,
                'designation': article.designation,
                'stock_actuel': article.stock_actuel,
                'mouvements_count': article_mouvements.count()
            })
        
        return JsonResponse({
            'success': True,
            'total_mouvements': total_mouvements,
            'mouvements_entree': mouvements_entree,
            'mouvements_sortie': mouvements_sortie,
            'valeur_stock_permanent': float(valeur_stock_permanent),
            'articles_info': articles_info,
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'mouvements': mouvements_data,
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR CONSULTATION MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        error_details = f"Erreur: {str(e)}"
        if hasattr(e, '__class__'):
            error_details += f" (Type: {e.__class__.__name__})"
        return JsonResponse({'success': False, 'error': error_details})

@login_required
def export_mouvements_excel(request):
    """Vue pour exporter les mouvements de stock en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Mouvements Stock {article_reference}"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tête du document (17 colonnes maintenant avec Référence/Désignation)
        ws.merge_cells('A1:Q1')
        ws['A1'] = f"FICHE DE STOCK - {article_reference} - {article_designation}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:Q2')
        ws['A2'] = f"Période: du {date_debut} au {date_fin}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes des colonnes - Ajout d'une colonne Référence/Désignation
        headers = ['Référence/Désignation', 'Date', 'Type', '', 'N°', '', '', 'Tiers', '', '', '', '', '+/-', 'Quantités en stock', 'Solde', 'C.M.U.P.', 'Stock permanent']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Grouper les mouvements par article (par référence_article)
        from collections import defaultdict
        mouvements_par_article = defaultdict(list)
        for mouvement in mouvements_data_list:
            # Utiliser référence_article comme clé de regroupement
            ref_article = mouvement.get('reference_article', 'N/A')
            mouvements_par_article[ref_article].append(mouvement)
        
        # Trier les articles par référence pour un affichage ordonné
        articles_tries = sorted(mouvements_par_article.keys())
        
        # Style pour les en-têtes d'article
        article_header_font = Font(bold=True, size=12, color="FFFFFF")
        article_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        article_header_alignment = Alignment(horizontal="left", vertical="center")
        
        # Données groupées par article
        row = 5
        total_stock_permanent_global = 0
        
        for ref_article in articles_tries:
            mouvements_article = mouvements_par_article[ref_article]
            
            # Récupérer la désignation du premier mouvement (tous les mouvements d'un article ont la même désignation)
            designation_article = mouvements_article[0].get('designation', 'N/A') if mouvements_article else 'N/A'
            
            # En-tête de l'article - Ligne de séparation avec référence et désignation
            ws.merge_cells(f'A{row}:Q{row}')
            cell_article = ws.cell(row=row, column=1, value=f"ARTICLE: {ref_article} - {designation_article}")
            cell_article.font = article_header_font
            cell_article.fill = article_header_fill
            cell_article.alignment = article_header_alignment
            row += 1
            
            # Calculer les totaux pour cet article
            total_entrees_article = sum(1 for m in mouvements_article if m['type_mouvement'] == 'entree')
            total_sorties_article = sum(1 for m in mouvements_article if m['type_mouvement'] == 'sortie')
            total_stock_permanent_article = sum(float(m.get('stock_permanent', 0)) for m in mouvements_article)
            total_stock_permanent_global += total_stock_permanent_article
            
            # Afficher tous les mouvements de cet article
            for mouvement in mouvements_article:
                # Déterminer le type de mouvement avec "Correction" si c'est une modification
                type_mouvement_display = mouvement.get('type_mouvement_display', mouvement['type_mouvement'])
                commentaire = (mouvement.get('commentaire', '') or '').lower()
                if 'modification' in commentaire or 'modifier' in commentaire or '(modification' in commentaire:
                    type_mouvement_display = 'Correction'
                
                # Structure avec colonne Référence/Désignation (17 colonnes maintenant)
                ws.cell(row=row, column=1, value=f"{ref_article} - {designation_article}")  # Référence/Désignation
                ws.cell(row=row, column=2, value=mouvement['date_mouvement'])  # Date
                ws.cell(row=row, column=3, value=type_mouvement_display)  # Type (avec Correction si modification)
                ws.cell(row=row, column=4, value='')  # Colonne vide
                ws.cell(row=row, column=5, value=mouvement['numero_piece'])  # N°
                ws.cell(row=row, column=6, value='')  # Colonne vide
                ws.cell(row=row, column=7, value='')  # Colonne vide
                ws.cell(row=row, column=8, value=mouvement['tiers'])  # Tiers
                ws.cell(row=row, column=9, value='')  # Colonne vide
                ws.cell(row=row, column=10, value='')  # Colonne vide
                ws.cell(row=row, column=11, value='')  # Colonne vide
                ws.cell(row=row, column=12, value='')  # Colonne vide
                # Calculer le signe +/- selon le type de mouvement
                quantite = float(mouvement.get('quantite', 0))
                if mouvement['type_mouvement'] in ['sortie', 'perte']:
                    quantite_display = f"-{abs(quantite)}"
                elif mouvement['type_mouvement'] in ['entree', 'retour']:
                    quantite_display = f"+{quantite}" if quantite > 0 else str(quantite)
                elif mouvement['type_mouvement'] == 'ajustement':
                    # Pour ajustement, déterminer selon solde vs stock_initial
                    stock_initial = float(mouvement.get('stock_initial', 0))
                    solde = float(mouvement.get('solde', 0))
                    if solde < stock_initial:
                        quantite_display = f"-{abs(quantite)}"
                    elif solde > stock_initial:
                        quantite_display = f"+{quantite}"
                    else:
                        quantite_display = str(quantite)
                else:
                    quantite_display = f"+{quantite}" if quantite > 0 else str(quantite)
                ws.cell(row=row, column=13, value=quantite_display)  # +/-
                ws.cell(row=row, column=14, value=mouvement['stock_initial'])  # Quantités en stock
                ws.cell(row=row, column=15, value=mouvement['solde'])  # Solde
                ws.cell(row=row, column=16, value=mouvement['cout_moyen_pondere'])  # C.M.U.P.
                ws.cell(row=row, column=17, value=mouvement['stock_permanent'])  # Stock permanent
                row += 1
            
            # Ligne de sous-total pour cet article
            ws.merge_cells(f'A{row}:B{row}')
            ws.cell(row=row, column=1, value=f"SOUS-TOTAL {ref_article}").font = Font(bold=True)
            ws.cell(row=row, column=3, value=f"Entrées: {total_entrees_article}, Sorties: {total_sorties_article}").font = Font(bold=True)
            ws.cell(row=row, column=4, value=f"Total: {len(mouvements_article)} mouvements").font = Font(bold=True)
            ws.cell(row=row, column=17, value=total_stock_permanent_article).font = Font(bold=True)
            # Remplir les colonnes vides
            for col in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
                ws.cell(row=row, column=col, value='').font = Font(bold=True)
            row += 1
            
            # Ligne vide pour séparer les articles
            row += 1
        
        # Ligne des totaux généraux
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True, size=14)
        ws.cell(row=row, column=3, value=f"Entrées: {mouvements_entree}, Sorties: {mouvements_sortie}").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"Total: {total_mouvements} mouvements").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"Articles: {len(articles_tries)}").font = Font(bold=True)
        # Remplir les colonnes vides
        for col in [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
            ws.cell(row=row, column=col, value="").font = Font(bold=True)
        ws.cell(row=row, column=17, value=total_stock_permanent_global).font = Font(bold=True, size=12)
        
        # Ajuster la largeur des colonnes (17 colonnes maintenant avec Référence/Désignation)
        ws.column_dimensions['A'].width = 35  # Référence/Désignation (plus large)
        ws.column_dimensions['B'].width = 20  # Date
        ws.column_dimensions['C'].width = 15  # Type
        ws.column_dimensions['D'].width = 5   # Colonne vide
        ws.column_dimensions['E'].width = 15  # N°
        ws.column_dimensions['F'].width = 5   # Colonne vide
        ws.column_dimensions['G'].width = 5  # Colonne vide
        ws.column_dimensions['H'].width = 25  # Tiers
        ws.column_dimensions['I'].width = 5   # Colonne vide
        ws.column_dimensions['J'].width = 5   # Colonne vide
        ws.column_dimensions['K'].width = 5   # Colonne vide
        ws.column_dimensions['L'].width = 5  # Colonne vide
        ws.column_dimensions['M'].width = 12  # +/-
        ws.column_dimensions['N'].width = 18  # Quantités en stock
        ws.column_dimensions['O'].width = 12  # Solde
        ws.column_dimensions['P'].width = 15  # C.M.U.P.
        ws.column_dimensions['Q'].width = 18  # Stock permanent
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le fichier dans la réponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Article: {article_reference}")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - {total_mouvements} mouvements")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Articles groupés: {len(articles_tries)}")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})
@login_required
def export_mouvements_pdf(request):
    """Vue pour exporter les mouvements de stock en PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            textColor=colors.darkblue
        )
        
        # Éléments du document
        elements = []
        
        # Titre
        title = Paragraph(f"FICHE DE STOCK - {article_reference}", title_style)
        elements.append(title)
        
        # Informations de l'article et période
        info_text = f"<b>Article:</b> {article_designation}<br/><b>Période:</b> du {date_debut} au {date_fin}"
        info_para = Paragraph(info_text, styles['Normal'])
        elements.append(info_para)
        elements.append(Spacer(1, 20))
        
        # Données du tableau
        data = [['Date', 'Type', 'Tiers', 'Stock Init.', 'Solde', 'C.M.PU', 'Stock Perm.', 'N° Pièce']]
        
        for mouvement in mouvements_data_list:
            data.append([
                mouvement['date_mouvement'],
                mouvement['type_mouvement'],
                mouvement['tiers'][:20] + '...' if len(mouvement['tiers']) > 20 else mouvement['tiers'],
                str(mouvement['stock_initial']),
                str(mouvement['solde']),
                f"{mouvement['cout_moyen_pondere']:,.0f}",
                f"{mouvement['stock_permanent']:,.0f}",
                mouvement['numero_piece']
            ])
        
        # Ligne des totaux
        data.append([
            'TOTAL',
            f'E:{mouvements_entree} S:{mouvements_sortie}',
            f'{total_mouvements} mouvements',
            '',
            '',
            '',
            f"{valeur_stock_permanent:,.0f} FCFA",
            ''
        ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 1.5*cm, 3*cm, 1.5*cm, 1.5*cm, 2*cm, 2.5*cm, 2*cm])
        
        # Style du tableau
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF MOUVEMENTS - Article: {article_reference}")
        print(f"📄 EXPORT PDF MOUVEMENTS - {total_mouvements} mouvements")
        print(f"📄 EXPORT PDF MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_mouvements_csv(request):
    """Vue pour exporter les mouvements de stock en CSV"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Date', 'Type', '', 'N°', '', '', 'Tiers', '', '', '', '', '+/-', 'Quantités en stock', 'Solde', 'C.M.U.P.', 'Stock permanent']
        writer.writerow(headers)
        
        # Données
        for mouvement in mouvements_data_list:
            row = [
                mouvement['date_mouvement'],
                mouvement['type_mouvement'],
                mouvement['tiers'],
                mouvement['stock_initial'],
                mouvement['solde'],
                mouvement['cout_moyen_pondere'],
                mouvement['stock_permanent'],
                mouvement['numero_piece']
            ]
            writer.writerow(row)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['TOTAL GÉNÉRAL', f'Entrées: {mouvements_entree}, Sorties: {mouvements_sortie}', f'{total_mouvements} mouvements', '', '', '', valeur_stock_permanent, ''])
        
        print(f"[CHART] EXPORT CSV MOUVEMENTS - Article: {article_reference}")
        print(f"[CHART] EXPORT CSV MOUVEMENTS - {total_mouvements} mouvements")
        print(f"[CHART] EXPORT CSV MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def creer_mouvements_retroactifs(request):
    """Vue simplifiée pour créer des mouvements de stock rétroactifs"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[REFRESH] CRÉATION MOUVEMENTS RÉTROACTIFS (VERSION SIMPLIFIÉE)...")
        print(f"[TARGET] Agence utilisée: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # 1. Créer des mouvements pour les factures de vente
        factures_vente = FactureVente.objects.filter(agence=agence)
        print(f"[CHART] Factures de vente: {factures_vente.count()}")
        
        for facture in factures_vente:
            lignes = LigneFactureVente.objects.filter(facture_vente=facture)
            print(f"[SEARCH] Facture {facture.numero_ticket}: {lignes.count()} lignes")
            
            for ligne in lignes:
                print(f"  [PACKAGE] Article: {ligne.article.designation}, Quantité: {ligne.quantite}")
                
                # Vérifier si le mouvement existe déjà
                mouvement_existe = MouvementStock.objects.filter(facture_vente=facture, article=ligne.article).exists()
                print(f"  [SEARCH] Mouvement existe déjà: {mouvement_existe}")
                
                if not mouvement_existe:
                    try:
                        mouvement_datetime = datetime.combine(
                            facture.date or timezone.now().date(),
                            facture.heure or time(12, 0)
                        )
                        if timezone.is_naive(mouvement_datetime):
                            mouvement_datetime = timezone.make_aware(
                                mouvement_datetime,
                                timezone.get_current_timezone()
                            )
                        
                        MouvementStock.objects.create(
                            article=ligne.article,
                            agence=agence,
                            type_mouvement='sortie',
                            date_mouvement=mouvement_datetime,
                            numero_piece=facture.numero_ticket,
                            quantite_stock=ligne.article.stock_actuel,
                            stock_initial=ligne.article.stock_actuel + ligne.quantite,
                            solde=ligne.article.stock_actuel,
                            quantite=ligne.quantite,
                            cout_moyen_pondere=float(ligne.prix_unitaire),
                            stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                            facture_vente=facture,
                            commentaire=f"Vente - {facture.numero_ticket}"
                        )
                        mouvements_crees += 1
                        print(f"[OK] Vente: {ligne.article.designation} - {facture.numero_ticket}")
                    except Exception as e:
                        print(f"[ERREUR] Erreur vente {facture.numero_ticket}: {e}")
        
        # 2. Créer des mouvements pour les factures d'achat
        factures_achat = FactureAchat.objects.filter(agence=agence)
        print(f"[CHART] Factures d'achat: {factures_achat.count()}")
        
        for facture in factures_achat:
            lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
            print(f"[SEARCH] Facture achat {facture.reference_achat}: {lignes.count()} lignes")
            
            for ligne in lignes:
                print(f"  [PACKAGE] Article: {ligne.article.designation}, Quantité: {ligne.quantite}")
                
                mouvement_existe = MouvementStock.objects.filter(facture_achat=facture, article=ligne.article).exists()
                print(f"  [SEARCH] Mouvement existe déjà: {mouvement_existe}")
                
                if not mouvement_existe:
                    try:
                        from django.utils import timezone
                        
                        MouvementStock.objects.create(
                            article=ligne.article,
                            agence=agence,
                            type_mouvement='entree',
                            date_mouvement=timezone.now(),
                            numero_piece=facture.reference_achat,
                            quantite_stock=ligne.article.stock_actuel,
                            stock_initial=ligne.article.stock_actuel - ligne.quantite,
                            solde=ligne.article.stock_actuel,
                            quantite=ligne.quantite,
                            cout_moyen_pondere=float(ligne.prix_unitaire),
                            stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                            facture_achat=facture,
                            fournisseur=facture.fournisseur,
                            commentaire=f"Achat - {facture.reference_achat}"
                        )
                        mouvements_crees += 1
                        print(f"[OK] Achat: {ligne.article.designation} - {facture.reference_achat}")
                    except Exception as e:
                        print(f"[ERREUR] Erreur achat {facture.reference_achat}: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'{mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def diagnostic_mouvements(request):
    """Vue de diagnostic pour les mouvements de stock"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Compter les données
        articles_count = Article.objects.filter(agence=agence).count()
        factures_vente_count = FactureVente.objects.filter(agence=agence).count()
        factures_achat_count = FactureAchat.objects.filter(agence=agence).count()
        factures_transfert_count = FactureTransfert.objects.filter(agence_source=agence).count()
        mouvements_count = MouvementStock.objects.filter(agence=agence).count()
        
        # Détails des factures
        factures_vente_details = []
        for facture in FactureVente.objects.filter(agence=agence)[:5]:
            lignes_count = LigneFactureVente.objects.filter(facture_vente=facture).count()
            factures_vente_details.append({
                'numero': facture.numero_ticket,
                'date': str(facture.date),
                'lignes': lignes_count
            })
        
        factures_achat_details = []
        for facture in FactureAchat.objects.filter(agence=agence)[:5]:
            lignes_count = LigneFactureAchat.objects.filter(facture_achat=facture).count()
            factures_achat_details.append({
                'numero': facture.reference_achat,
                'date': str(facture.date_achat),
                'lignes': lignes_count
            })
        
        return JsonResponse({
            'success': True,
            'agence': agence.nom_agence,
            'articles_count': articles_count,
            'factures_vente_count': factures_vente_count,
            'factures_achat_count': factures_achat_count,
            'factures_transfert_count': factures_transfert_count,
            'mouvements_count': mouvements_count,
            'factures_vente_details': factures_vente_details,
            'factures_achat_details': factures_achat_details
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def forcer_mouvements(request):
    """Vue pour forcer la création de mouvements même s'ils existent déjà"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[HOT] CRÉATION FORCÉE DE MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # Supprimer tous les mouvements existants d'abord
        anciens_mouvements = MouvementStock.objects.filter(agence=agence).count()
        MouvementStock.objects.filter(agence=agence).delete()
        print(f"🗑️ {anciens_mouvements} anciens mouvements supprimés")
        
        # Test de création d'un mouvement simple
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles trouvés: {articles.count()}")
        
        if articles.exists():
            article_test = articles.first()
            print(f"🧪 Test avec article: {article_test.designation}")
            
            try:
                from django.utils import timezone
                
                mouvement_test = MouvementStock.objects.create(
                    article=article_test,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece='TEST-001',
                    quantite_stock=article_test.stock_actuel,
                    stock_initial=0,
                    solde=article_test.stock_actuel,
                    quantite=1,
                    cout_moyen_pondere=float(article_test.prix_achat),
                    stock_permanent=float(article_test.stock_actuel * article_test.prix_achat),
                    commentaire='Test de création'
                )
                print(f"[OK] MOUVEMENT TEST CRÉÉ AVEC SUCCÈS: ID {mouvement_test.id}")
                mouvements_crees += 1
                
                # Supprimer le test
                mouvement_test.delete()
                print(f"🗑️ Mouvement test supprimé")
                
            except Exception as e:
                print(f"[ERREUR] ERREUR LORS DU TEST: {e}")
                import traceback
                traceback.print_exc()
                return JsonResponse({'success': False, 'error': f'Erreur lors du test de création: {str(e)}'})
        
        print(f"[OK] Test terminé, création des vrais mouvements...")
        
        # Créer des mouvements pour les factures de vente (version simplifiée)
        factures_vente = FactureVente.objects.filter(agence=agence)
        print(f"[CHART] Factures de vente: {factures_vente.count()}")
        
        for facture in factures_vente:
            lignes = LigneFactureVente.objects.filter(facture_vente=facture)
            print(f"  [SEARCH] {facture.numero_ticket}: {lignes.count()} lignes")
            
            for ligne in lignes:
                try:
                    from django.utils import timezone
                    
                    # Utiliser la même logique simple que le test qui fonctionne
                    MouvementStock.objects.create(
                        article=ligne.article,
                        agence=agence,
                        type_mouvement='sortie',
                        date_mouvement=timezone.now(),
                        numero_piece=facture.numero_ticket,
                        quantite_stock=ligne.article.stock_actuel,
                        stock_initial=0,  # Simplifié
                        solde=ligne.article.stock_actuel,
                        quantite=ligne.quantite,
                        cout_moyen_pondere=float(ligne.prix_unitaire),
                        stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                        facture_vente=facture,
                        commentaire=f"Vente - {facture.numero_ticket}"
                    )
                    mouvements_crees += 1
                    print(f"    [OK] {ligne.article.designation}")
                except Exception as e:
                    print(f"    [ERREUR] Erreur: {e}")
        
        # Créer des mouvements pour les factures d'achat (version simplifiée)
        factures_achat = FactureAchat.objects.filter(agence=agence)
        print(f"[CHART] Factures d'achat: {factures_achat.count()}")
        
        for facture in factures_achat:
            lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
            print(f"  [SEARCH] {facture.reference_achat}: {lignes.count()} lignes")
            
            for ligne in lignes:
                try:
                    from django.utils import timezone
                    
                    # Utiliser la même logique simple que le test qui fonctionne
                    MouvementStock.objects.create(
                        article=ligne.article,
                        agence=agence,
                        type_mouvement='entree',
                        date_mouvement=timezone.now(),
                        numero_piece=facture.reference_achat,
                        quantite_stock=ligne.article.stock_actuel,
                        stock_initial=0,  # Simplifié
                        solde=ligne.article.stock_actuel,
                        quantite=ligne.quantite,
                        cout_moyen_pondere=float(ligne.prix_unitaire),
                        stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                        facture_achat=facture,
                        fournisseur=facture.fournisseur,
                        commentaire=f"Achat - {facture.reference_achat}"
                    )
                    mouvements_crees += 1
                    print(f"    [OK] {ligne.article.designation}")
                except Exception as e:
                    print(f"    [ERREUR] Erreur: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'FORCÉ: {mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_mouvement_simple(request):
    """Test simple de création d'un mouvement"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("🧪 TEST SIMPLE DE CRÉATION DE MOUVEMENT...")
        
        # Vérifier les articles
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles: {articles.count()}")
        
        if not articles.exists():
            return JsonResponse({'success': False, 'error': 'Aucun article trouvé dans cette agence'})
        
        article = articles.first()
        print(f"[PACKAGE] Article test: {article.designation}")
        
        # Vérifier les champs obligatoires
        print(f"[PACKAGE] Stock actuel: {article.stock_actuel}")
        print(f"[PACKAGE] Prix achat: {article.prix_achat}")
        
        # Créer un mouvement simple
        from django.utils import timezone
        
        mouvement = MouvementStock.objects.create(
            article=article,
            agence=agence,
            type_mouvement='entree',
            date_mouvement=timezone.now(),
            numero_piece='TEST-SIMPLE',
            quantite_stock=article.stock_actuel,
            stock_initial=0,
            solde=article.stock_actuel,
            quantite=1,
            cout_moyen_pondere=float(article.prix_achat),
            stock_permanent=float(article.stock_actuel * article.prix_achat),
            commentaire='Test simple'
        )
        
        print(f"[OK] MOUVEMENT CRÉÉ: ID {mouvement.id}")
        
        # Vérifier qu'il existe
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        print(f"[CHART] Total mouvements: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'Test réussi! Mouvement ID {mouvement.id} créé. Total: {total_mouvements}',
            'mouvement_id': mouvement.id,
            'total_mouvements': total_mouvements
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR TEST: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def creer_mouvements_manuels(request):
    """Créer des mouvements manuels simples pour tester"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[TOOL] CRÉATION MANUELLE DE MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # Récupérer tous les articles
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles trouvés: {articles.count()}")
        
        if not articles.exists():
            return JsonResponse({'success': False, 'error': 'Aucun article trouvé'})
        
        # Créer un mouvement pour chaque article
        for article in articles:
            try:
                from django.utils import timezone
                
                MouvementStock.objects.create(
                    article=article,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece=f'MANUEL-{article.id}',
                    quantite_stock=article.stock_actuel,
                    stock_initial=0,
                    solde=article.stock_actuel,
                    quantite=article.stock_actuel,
                    cout_moyen_pondere=float(article.prix_achat),
                    stock_permanent=float(article.stock_actuel * article.prix_achat),
                    commentaire=f'Création manuelle - {article.designation}'
                )
                mouvements_crees += 1
                print(f"[OK] {article.designation}")
                
            except Exception as e:
                print(f"[ERREUR] Erreur pour {article.designation}: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements manuels créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'MANUEL: {mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_consultation_mouvements(request):
    """Test simple pour vérifier les mouvements existants"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[SEARCH] TEST CONSULTATION MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        # Vérifier tous les mouvements de l'agence
        tous_mouvements = MouvementStock.objects.filter(agence=agence)
        print(f"[CHART] Total mouvements dans l'agence: {tous_mouvements.count()}")
        
        if tous_mouvements.count() > 0:
            print("[LIST] PREMIERS MOUVEMENTS:")
            for i, mvt in enumerate(tous_mouvements[:5]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement} - {mvt.numero_piece}")
        
        # Vérifier les articles avec mouvements
        articles_avec_mouvements = Article.objects.filter(
            agence=agence,
            mouvementstock__isnull=False
        ).distinct()
        print(f"[PACKAGE] Articles avec mouvements: {articles_avec_mouvements.count()}")
        
        for article in articles_avec_mouvements[:3]:
            mouvements_article = MouvementStock.objects.filter(agence=agence, article=article)
            print(f"  - {article.designation}: {mouvements_article.count()} mouvements")
        
        return JsonResponse({
            'success': True,
            'message': f'Test terminé - {tous_mouvements.count()} mouvements trouvés',
            'total_mouvements': tous_mouvements.count(),
            'articles_avec_mouvements': articles_avec_mouvements.count()
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_mouvements_session(request):
    """Récupérer les données de mouvements depuis la session"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les données depuis la session
        mouvements_data = request.session.get('mouvements_data', [])
        article_info = request.session.get('article_info', {})
        
        print(f"[CHART] RÉCUPÉRATION SESSION:")
        print(f"  - Mouvements en session: {len(mouvements_data)}")
        print(f"  - Article info: {article_info}")
        
        return JsonResponse({
            'success': True,
            'mouvements': mouvements_data,
            'article_info': article_info
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ===== GESTION DES COMPTES UTILISATEURS =====

def login_comptes(request):
    """Page de connexion pour le module gestion des comptes"""
    if request.user.is_authenticated and request.method == 'GET':
        logout(request)
        request.session.flush()

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        login(request, user)
                        request.session['agence_id'] = compte.agence.id_agence
                        request.session['agence_nom'] = compte.agence.nom_agence
                        messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                        return redirect('gestion_comptes')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Aucun compte actif trouvé pour cet utilisateur.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')

    return render(request, 'supermarket/comptes/login.html')

@login_required
def logout_comptes(request):
    """Déconnexion pour le module comptes"""
    logout(request)
    return redirect('index')

@login_required
@require_comptes_access
def gestion_comptes(request):
    """Vue pour lister tous les comptes utilisateurs"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    comptes = Compte.objects.filter(agence=agence).order_by('nom', 'prenom')
    
    context = {
        'comptes': comptes,
        'agence': agence,
    }
    return render(request, 'supermarket/comptes/gestion_comptes.html', context)

@login_required
@require_comptes_access
def creer_compte(request):
    """Vue pour créer un nouveau compte utilisateur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        numero_compte = request.POST.get('numero_compte')
        type_compte = request.POST.get('type_compte')
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        telephone = request.POST.get('telephone')
        email = request.POST.get('email')
        agence_id = request.POST.get('agence')
        
        try:
            # Récupérer l'agence sélectionnée
            if agence_id:
                agence_selectionnee = Agence.objects.get(id_agence=agence_id)
            else:
                agence_selectionnee = agence  # Utiliser l'agence de l'utilisateur connecté par défaut
            
            # Créer l'utilisateur Django
            user = User.objects.create_user(username=username, password=password, email=email)
            
            # Créer le compte
            compte = Compte.objects.create(
                user=user,
                numero_compte=numero_compte,
                type_compte=type_compte,
                nom=nom,
                prenom=prenom,
                telephone=telephone,
                email=email,
                agence=agence_selectionnee
            )
            
            messages.success(request, f'Compte créé avec succès pour {compte.nom_complet} - Agence: {agence_selectionnee.nom_agence}')
            return redirect('gestion_comptes')
        except Agence.DoesNotExist:
            messages.error(request, 'Agence sélectionnée non trouvée.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du compte: {str(e)}')
    
    agences = Agence.objects.all().order_by('nom_agence')
    type_compte_choices = Compte._meta.get_field('type_compte').choices
    context = {
        'agences': agences,
        'agence': agence,
        'type_compte_choices': type_compte_choices,
    }
    return render(request, 'supermarket/comptes/creer_compte.html', context)

@login_required
@require_comptes_access
def detail_compte(request, compte_id):
    """Vue pour afficher les détails d'un compte"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
        return redirect('gestion_comptes')
    
    context = {
        'compte': compte,
        'agence': agence,
    }
    return render(request, 'supermarket/comptes/detail_compte.html', context)

@login_required
@require_comptes_access
def modifier_compte(request, compte_id):
    """Vue pour modifier un compte utilisateur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
        return redirect('gestion_comptes')
    
    if request.method == 'POST':
        compte.numero_compte = request.POST.get('numero_compte', compte.numero_compte)
        compte.type_compte = request.POST.get('type_compte', compte.type_compte)
        compte.nom = request.POST.get('nom', compte.nom)
        compte.prenom = request.POST.get('prenom', compte.prenom)
        compte.telephone = request.POST.get('telephone', compte.telephone)
        compte.email = request.POST.get('email', compte.email)
        compte.actif = request.POST.get('actif') == 'on'
        
        # Mettre à jour l'utilisateur Django si nécessaire
        if request.POST.get('username'):
            compte.user.username = request.POST.get('username')
            compte.user.save()
        
        if request.POST.get('password'):
            compte.user.set_password(request.POST.get('password'))
            compte.user.save()
        
        compte.save()
        messages.success(request, f'Compte modifié avec succès pour {compte.nom_complet}')
        return redirect('detail_compte', compte_id=compte.id)
    
    agences = Agence.objects.all()
    type_compte_choices = Compte._meta.get_field('type_compte').choices
    context = {
        'compte': compte,
        'agences': agences,
        'agence': agence,
        'type_compte_choices': type_compte_choices,
    }
    return render(request, 'supermarket/comptes/modifier_compte.html', context)

@login_required
@require_comptes_access
def supprimer_compte(request, compte_id):
    """Vue pour supprimer un compte utilisateur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
        nom_complet = compte.nom_complet
        compte.user.delete()  # Cela supprimera aussi le compte grâce à CASCADE
        messages.success(request, f'Compte {nom_complet} supprimé avec succès')
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('gestion_comptes')

@login_required
@require_comptes_access
def activer_desactiver_compte(request, compte_id):
    """Vue pour activer ou désactiver un compte"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('gestion_comptes')
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
        
        # Inverser le statut actif/inactif
        compte.actif = not compte.actif
        compte.save()
        
        statut = "activé" if compte.actif else "désactivé"
        messages.success(request, f'Compte {compte.nom_complet} {statut} avec succès.')
        
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    return redirect('gestion_comptes')

@login_required
@require_comptes_access
def reinitialiser_mot_de_passe(request, compte_id):
    """Vue pour réinitialiser le mot de passe d'un compte"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
        return redirect('gestion_comptes')
    
    if request.method == 'POST':
        nouveau_mot_de_passe = request.POST.get('nouveau_mot_de_passe')
        confirmer_mot_de_passe = request.POST.get('confirmer_mot_de_passe')
        
        if nouveau_mot_de_passe and nouveau_mot_de_passe == confirmer_mot_de_passe:
            compte.user.set_password(nouveau_mot_de_passe)
            compte.user.save()
            messages.success(request, f'Mot de passe réinitialisé avec succès pour {compte.nom_complet}')
            return redirect('detail_compte', compte_id=compte.id)
        else:
            messages.error(request, 'Les mots de passe ne correspondent pas.')
    
    context = {
        'compte': compte,
        'agence': agence,
    }
    return render(request, 'supermarket/comptes/reinitialiser_mot_de_passe.html', context)

@login_required
def debug_session(request):
    """Debug simple pour voir le contenu de la session"""
    try:
        print("[SEARCH] DEBUG SESSION:")
        print(f"  - Clés de session: {list(request.session.keys())}")
        
        mouvements_stock = request.session.get('mouvements_stock', {})
        print(f"  - mouvements_stock: {list(mouvements_stock.keys()) if mouvements_stock else 'Aucun'}")
        
        if mouvements_stock:
            print(f"  - mouvements_data count: {len(mouvements_stock.get('mouvements_data', []))}")
            print(f"  - article_reference: {mouvements_stock.get('article_reference', 'N/A')}")
            print(f"  - article_designation: {mouvements_stock.get('article_designation', 'N/A')}")
        
        return JsonResponse({
            'success': True,
            'session_keys': list(request.session.keys()),
            'mouvements_stock_keys': list(mouvements_stock.keys()) if mouvements_stock else [],
            'mouvements_count': len(mouvements_stock.get('mouvements_data', [])) if mouvements_stock else 0
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        return JsonResponse({'success': False, 'error': str(e)})
@login_required
@require_stock_access
def dashboard_stock(request):
    """Dashboard principal du module de gestion de stock"""
    try:
        # Récupérer l'agence de l'utilisateur
        agence = get_user_agence(request)
        if not agence:
            messages.error(request, 'Votre compte n\'est pas configuré correctement.')
            return redirect('logout_stock')

        # Calculer les KPIs
        total_articles = Article.objects.filter(agence=agence).count()
        articles_stock_faible = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=10
        ).count()
        articles_rupture = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=0
        ).count()
        
        # Valeur totale du stock
        articles_avec_prix = Article.objects.filter(
            agence=agence,
            prix_achat__isnull=False,
            stock_actuel__isnull=False
        ).exclude(prix_achat=0).exclude(stock_actuel=0)
        
        valeur_totale_stock = 0
        for article in articles_avec_prix:
            try:
                valeur_article = float(article.prix_achat) * float(article.stock_actuel)
                valeur_totale_stock += valeur_article
            except (ValueError, TypeError):
                continue
        
        # Mouvements récents
        mouvements_recents = MouvementStock.objects.filter(agence=agence).order_by('-date_mouvement')[:5]
        
        # Articles les plus vendus (simulation)
        articles_populaires = Article.objects.filter(agence=agence).order_by('-stock_actuel')[:5]
        
        # Alertes de stock
        alertes_stock = Article.objects.filter(
            agence=agence,
            stock_actuel__lte=5
        ).order_by('stock_actuel')[:5]

        # Récupérer le nom de l'utilisateur
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except Compte.DoesNotExist:
            nom_utilisateur = request.user.username

        context = {
            'agence': agence,
            'nom_utilisateur': nom_utilisateur,
            'total_articles': total_articles,
            'articles_stock_faible': articles_stock_faible,
            'articles_rupture': articles_rupture,
            'valeur_stock': valeur_totale_stock,  # Corrigé le nom de la variable
            'mouvements_recents': mouvements_recents,
            'articles_populaires': articles_populaires,
            'alertes_stock': alertes_stock,
        }
        
        return render(request, 'supermarket/stock/dashboard_stock.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur lors du chargement du dashboard: {str(e)}')
        # Récupérer le nom de l'utilisateur même en cas d'erreur
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except:
            nom_utilisateur = request.user.username if request.user.is_authenticated else "Utilisateur"

        return render(request, 'supermarket/stock/dashboard_stock.html', {
            'agence': None,
            'nom_utilisateur': nom_utilisateur,
            'total_articles': 0,
            'articles_stock_faible': 0,
            'articles_rupture': 0,
            'valeur_stock': 0,
            'mouvements_recents': [],
            'articles_populaires': [],
            'alertes_stock': [],
        })

def login_stock(request):
    """Page de connexion pour la gestion de stock"""
    # Toujours afficher la page de connexion, même si l'utilisateur est déjà connecté
    # Cela permet de forcer la reconnexion depuis la page d'accueil
    if request.user.is_authenticated and request.method == 'GET':
        # Déconnecter l'utilisateur pour forcer la reconnexion
        logout(request)
        request.session.flush()
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                # Vérifier que l'utilisateur a un compte lié à une agence
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        # Vérifier que l'utilisateur est un comptable, assistant_comptable ou admin
                        if compte.type_compte not in ['comptable', 'assistant_comptable', 'admin']:
                            messages.error(request, 'Accès refusé. Ce module est réservé aux comptables, assistants comptables et aux administrateurs.')
                        else:
                            login(request, user)
                            # Stocker l'agence dans la session
                            request.session['agence_id'] = compte.agence.id_agence
                            request.session['agence_nom'] = compte.agence.nom_agence
                            messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                            return redirect('dashboard_stock')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Aucun compte actif trouvé pour cet utilisateur.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')
    
    return render(request, 'supermarket/stock/login.html')

@login_required
def logout_stock(request):
    """Vue de logout pour le module stock"""
    logout(request)
    return redirect('index')

@login_required
def modifier_client(request, client_id):
    """Vue pour modifier un client existant"""
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            intitule = request.POST.get('intitule')
            adresse = request.POST.get('adresse')
            telephone = request.POST.get('telephone')
            email = request.POST.get('email')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([intitule, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_client', client_id=client_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_client', client_id=client_id)
            
            # Mettre à jour le client
            client.intitule = intitule
            client.adresse = adresse
            client.telephone = telephone
            client.email = email
            client.agence = agence
            client.save()
            
            messages.success(request, f'Client "{intitule}" modifié avec succès!')
            return redirect('consulter_clients')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du client: {str(e)}')
            return redirect('modifier_client', client_id=client_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    
    context = {
        'client': client,
        'agences': agences,
    }
    return render(request, 'supermarket/stock/modifier_client.html', context)

@login_required
def supprimer_client(request, client_id):
    """Vue pour supprimer un client"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_clients')
    
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
        client_name = client.intitule
        client.delete()
        messages.success(request, f'Client "{client_name}" supprimé avec succès!')
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_clients')

@login_required
def detail_client(request, client_id):
    """Vue pour afficher les détails d'un client"""
    try:
        client = Client.objects.get(id=client_id, agence=get_user_agence(request))
        
        # Récupérer les factures du client (si elles existent)
        factures = FactureVente.objects.filter(client=client).order_by('-date', '-heure')[:10]
        
        context = {
            'client': client,
            'factures': factures,
        }
        return render(request, 'supermarket/stock/detail_client.html', context)
    except Client.DoesNotExist:
        messages.error(request, 'Client non trouvé.')
        return redirect('consulter_clients')

# ==================== PLAN COMPTABLE ====================

@login_required
def consulter_plan_comptable(request):
    """Vue pour consulter le plan comptable"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    nature_filter = request.GET.get('nature_filter', '')
    
    # Construire la requête de base
    comptes = PlanComptable.objects.all()
    
    # Appliquer les filtres
    if search_query:
        comptes = comptes.filter(
            Q(intitule__icontains=search_query) |
            Q(compte__icontains=search_query) |
            Q(abrege__icontains=search_query)
        )
    
    if nature_filter:
        comptes = comptes.filter(nature_compte=nature_filter)
    
    # Trier par numéro
    comptes = comptes.order_by('numero')
    
    # Calculer les statistiques
    total_comptes = PlanComptable.objects.count()
    comptes_actifs = PlanComptable.objects.filter(actif=True).count()
    
    # Récupérer les natures de compte pour le filtre
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'comptes': comptes,
        'agence': agence,
        'total_comptes': total_comptes,
        'comptes_actifs': comptes_actifs,
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/consulter_plan_comptable.html', context)

@login_required
def creer_plan_comptable(request):
    """Vue pour créer un nouveau compte comptable"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero = request.POST.get('numero')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            abrege = request.POST.get('abrege')
            nature_compte = request.POST.get('nature_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([numero, intitule, compte]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_plan_comptable')
            
            # Créer le compte comptable
            PlanComptable.objects.create(
                numero=numero,
                intitule=intitule,
                compte=compte,
                abrege=abrege,
                nature_compte=nature_compte
            )
            
            messages.success(request, f'Compte comptable "{intitule}" créé avec succès!')
            return redirect('consulter_plan_comptable')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du compte: {str(e)}')
            return redirect('creer_plan_comptable')
    
    # GET - Afficher le formulaire
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/creer_plan_comptable.html', context)

@login_required
def modifier_plan_comptable(request, compte_id):
    """Vue pour modifier un compte comptable existant"""
    try:
        compte = PlanComptable.objects.get(id=compte_id)
    except PlanComptable.DoesNotExist:
        messages.error(request, 'Compte comptable non trouvé.')
        return redirect('consulter_plan_comptable')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero = request.POST.get('numero')
            intitule = request.POST.get('intitule')
            compte_field = request.POST.get('compte')
            abrege = request.POST.get('abrege')
            nature_compte = request.POST.get('nature_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([numero, intitule, compte_field]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_plan_comptable', compte_id=compte_id)
            
            # Mettre à jour le compte
            compte.numero = numero
            compte.intitule = intitule
            compte.compte = compte_field
            compte.abrege = abrege
            compte.nature_compte = nature_compte
            compte.save()
            
            messages.success(request, f'Compte comptable "{intitule}" modifié avec succès!')
            return redirect('consulter_plan_comptable')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du compte: {str(e)}')
            return redirect('modifier_plan_comptable', compte_id=compte_id)
    
    # GET - Afficher le formulaire pré-rempli
    nature_choices = PlanComptable.NATURE_COMPTE_CHOICES
    
    context = {
        'compte': compte,
        'nature_choices': nature_choices,
    }
    return render(request, 'supermarket/stock/modifier_plan_comptable.html', context)

@login_required
def supprimer_plan_comptable(request, compte_id):
    """Vue pour supprimer un compte comptable"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_plan_comptable')
    
    try:
        compte = PlanComptable.objects.get(id=compte_id)
        compte_name = compte.intitule
        compte.delete()
        messages.success(request, f'Compte comptable "{compte_name}" supprimé avec succès!')
    except PlanComptable.DoesNotExist:
        messages.error(request, 'Compte comptable non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_plan_comptable')

# ==================== PLAN TIERS ====================

@login_required
def consulter_plan_tiers(request):
    """Vue pour consulter le plan tiers"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    tiers = PlanTiers.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        tiers = tiers.filter(
            Q(intitule_compte__icontains=search_query) |
            Q(numero_compte__icontains=search_query)
        )
    
    if type_filter:
        tiers = tiers.filter(type=type_filter)
    
    # Trier par numéro de compte
    tiers = tiers.order_by('numero_compte')
    
    # Calculer les statistiques
    total_tiers = PlanTiers.objects.filter(agence=agence).count()
    clients_count = PlanTiers.objects.filter(agence=agence, type='client').count()
    fournisseurs_count = PlanTiers.objects.filter(agence=agence, type='fournisseur').count()
    
    # Récupérer les types pour le filtre
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'tiers': tiers,
        'agence': agence,
        'total_tiers': total_tiers,
        'clients_count': clients_count,
        'fournisseurs_count': fournisseurs_count,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_plan_tiers.html', context)

@login_required
def creer_plan_tiers(request):
    """Vue pour créer un nouveau tiers"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_tiers = request.POST.get('type')
            numero_compte = request.POST.get('numero_compte')
            intitule_compte = request.POST.get('intitule_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_tiers, numero_compte, intitule_compte, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_plan_tiers')
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('creer_plan_tiers')
            
            # Créer le tiers
            PlanTiers.objects.create(
                type=type_tiers,
                numero_compte=numero_compte,
                intitule_compte=intitule_compte,
                agence=agence
            )
            
            messages.success(request, f'Tiers "{intitule_compte}" créé avec succès!')
            return redirect('consulter_plan_tiers')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du tiers: {str(e)}')
            return redirect('creer_plan_tiers')
    
    # GET - Afficher le formulaire
    agences = Agence.objects.all()
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'agences': agences,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/creer_plan_tiers.html', context)

@login_required
def modifier_plan_tiers(request, tiers_id):
    """Vue pour modifier un tiers existant"""
    try:
        tiers = PlanTiers.objects.get(id=tiers_id, agence=get_user_agence(request))
    except PlanTiers.DoesNotExist:
        messages.error(request, 'Tiers non trouvé.')
        return redirect('consulter_plan_tiers')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_tiers = request.POST.get('type')
            numero_compte = request.POST.get('numero_compte')
            intitule_compte = request.POST.get('intitule_compte')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_tiers, numero_compte, intitule_compte, agence_id]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_plan_tiers', tiers_id=tiers_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_plan_tiers', tiers_id=tiers_id)
            
            # Mettre à jour le tiers
            tiers.type = type_tiers
            tiers.numero_compte = numero_compte
            tiers.intitule_compte = intitule_compte
            tiers.agence = agence
            tiers.save()
            
            messages.success(request, f'Tiers "{intitule_compte}" modifié avec succès!')
            return redirect('consulter_plan_tiers')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du tiers: {str(e)}')
            return redirect('modifier_plan_tiers', tiers_id=tiers_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    type_choices = PlanTiers.TYPE_TIERS_CHOICES
    
    context = {
        'tiers': tiers,
        'agences': agences,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/modifier_plan_tiers.html', context)

@login_required
def supprimer_plan_tiers(request, tiers_id):
    """Vue pour supprimer un tiers"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_plan_tiers')
    
    try:
        tiers = PlanTiers.objects.get(id=tiers_id, agence=get_user_agence(request))
        tiers_name = tiers.intitule_compte
        tiers.delete()
        messages.success(request, f'Tiers "{tiers_name}" supprimé avec succès!')
    except PlanTiers.DoesNotExist:
        messages.error(request, 'Tiers non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_plan_tiers')

# ==================== CODE JOURNAUX ====================

@login_required
def consulter_code_journaux(request):
    """Vue pour consulter les codes journaux"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    journaux = CodeJournaux.objects.all()
    
    # Appliquer les filtres
    if search_query:
        journaux = journaux.filter(
            Q(intitule__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    if type_filter:
        journaux = journaux.filter(type_document=type_filter)
    
    # Trier par code
    journaux = journaux.order_by('code')
    
    # Calculer les statistiques
    total_journaux = CodeJournaux.objects.count()
    journaux_achat = CodeJournaux.objects.filter(type_document='document_achat').count()
    journaux_vente = CodeJournaux.objects.filter(type_document='caisse').count()
    
    # Récupérer les types pour le filtre
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'journaux': journaux,
        'agence': agence,
        'total_journaux': total_journaux,
        'journaux_achat': journaux_achat,
        'journaux_vente': journaux_vente,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_code_journaux.html', context)

@login_required
def creer_code_journaux(request):
    """Vue pour créer un nouveau code journal"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_document = request.POST.get('type_document')
            intitule = request.POST.get('intitule')
            code = request.POST.get('code')
            compte_contrepartie = request.POST.get('compte_contrepartie')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_document, intitule, code]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_code_journaux')
            
            # Créer le code journal
            CodeJournaux.objects.create(
                type_document=type_document,
                intitule=intitule,
                code=code,
                compte_contrepartie=compte_contrepartie
            )
            
            messages.success(request, f'Code journal "{intitule}" créé avec succès!')
            return redirect('consulter_code_journaux')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du code journal: {str(e)}')
            return redirect('creer_code_journaux')
    
    # GET - Afficher le formulaire
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/creer_code_journaux.html', context)

@login_required
def modifier_code_journaux(request, journal_id):
    """Vue pour modifier un code journal existant"""
    try:
        journal = CodeJournaux.objects.get(id=journal_id)
    except CodeJournaux.DoesNotExist:
        messages.error(request, 'Code journal non trouvé.')
        return redirect('consulter_code_journaux')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_document = request.POST.get('type_document')
            intitule = request.POST.get('intitule')
            code = request.POST.get('code')
            compte_contrepartie = request.POST.get('compte_contrepartie')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([type_document, intitule, code]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_code_journaux', journal_id=journal_id)
            
            # Mettre à jour le code journal
            journal.type_document = type_document
            journal.intitule = intitule
            journal.code = code
            journal.compte_contrepartie = compte_contrepartie
            journal.save()
            
            messages.success(request, f'Code journal "{intitule}" modifié avec succès!')
            return redirect('consulter_code_journaux')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du code journal: {str(e)}')
            return redirect('modifier_code_journaux', journal_id=journal_id)
    
    # GET - Afficher le formulaire pré-rempli
    type_choices = CodeJournaux.TYPE_DOCUMENT_CHOICES
    
    context = {
        'journal': journal,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/modifier_code_journaux.html', context)

@login_required
def supprimer_code_journaux(request, journal_id):
    """Vue pour supprimer un code journal"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_code_journaux')
    
    try:
        journal = CodeJournaux.objects.get(id=journal_id)
        journal_name = journal.intitule
        journal.delete()
        messages.success(request, f'Code journal "{journal_name}" supprimé avec succès!')
    except CodeJournaux.DoesNotExist:
        messages.error(request, 'Code journal non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_code_journaux')

# ==================== TAUX TAXE ====================

@login_required
def consulter_taux_taxe(request):
    """Vue pour consulter les taux de taxe"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('logout_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    type_filter = request.GET.get('type_filter', '')
    
    # Construire la requête de base
    taux = TauxTaxe.objects.all()
    
    # Appliquer les filtres
    if search_query:
        taux = taux.filter(
            Q(intitule__icontains=search_query) |
            Q(code__icontains=search_query)
        )
    
    if type_filter:
        taux = taux.filter(type=type_filter)
    
    # Trier par code
    taux = taux.order_by('code')
    
    # Calculer les statistiques
    total_taux = TauxTaxe.objects.count()
    taux_actifs = TauxTaxe.objects.filter(actif=True).count()
    
    # Récupérer les types pour le filtre
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    
    context = {
        'taux': taux,
        'agence': agence,
        'total_taux': total_taux,
        'taux_actifs': taux_actifs,
        'type_choices': type_choices,
    }
    return render(request, 'supermarket/stock/consulter_taux_taxe.html', context)

@login_required
def creer_taux_taxe(request):
    """Vue pour créer un nouveau taux de taxe"""
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            code = request.POST.get('code')
            sens = request.POST.get('sens')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            taux = request.POST.get('taux')
            type_taxe = request.POST.get('type')
            assujettissement = request.POST.get('assujettissement')
            code_regroupement = request.POST.get('code_regroupement')
            provenance = request.POST.get('provenance')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([code, intitule, taux]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_taux_taxe')
            
            # Créer le taux de taxe
            TauxTaxe.objects.create(
                code=code,
                sens=sens,
                intitule=intitule,
                compte=compte,
                taux=float(taux),
                type=type_taxe,
                assujettissement=assujettissement,
                code_regroupement=code_regroupement,
                provenance=provenance
            )
            
            messages.success(request, f'Taux de taxe "{intitule}" créé avec succès!')
            return redirect('consulter_taux_taxe')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du taux de taxe: {str(e)}')
            return redirect('creer_taux_taxe')
    
    # GET - Afficher le formulaire
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    sens_choices = TauxTaxe.SENS_CHOICES
    assujettissement_choices = TauxTaxe.ASSUJETTISSEMENT_CHOICES
    
    context = {
        'type_choices': type_choices,
        'sens_choices': sens_choices,
        'assujettissement_choices': assujettissement_choices,
    }
    return render(request, 'supermarket/stock/creer_taux_taxe.html', context)
@login_required
def modifier_taux_taxe(request, taux_id):
    """Vue pour modifier un taux de taxe existant"""
    try:
        taux = TauxTaxe.objects.get(id=taux_id)
    except TauxTaxe.DoesNotExist:
        messages.error(request, 'Taux de taxe non trouvé.')
        return redirect('consulter_taux_taxe')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            code = request.POST.get('code')
            sens = request.POST.get('sens')
            intitule = request.POST.get('intitule')
            compte = request.POST.get('compte')
            taux_value = request.POST.get('taux')
            type_taxe = request.POST.get('type')
            assujettissement = request.POST.get('assujettissement')
            code_regroupement = request.POST.get('code_regroupement')
            provenance = request.POST.get('provenance')
            agence_id = request.POST.get('agence')
            
            # Validation
            if not all([code, intitule, taux_value]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_taux_taxe', taux_id=taux_id)
            
            # Mettre à jour le taux de taxe
            taux.code = code
            taux.sens = sens
            taux.intitule = intitule
            taux.compte = compte
            taux.taux = float(taux_value)
            taux.type = type_taxe
            taux.assujettissement = assujettissement
            taux.code_regroupement = code_regroupement
            taux.provenance = provenance
            taux.save()
            
            messages.success(request, f'Taux de taxe "{intitule}" modifié avec succès!')
            return redirect('consulter_taux_taxe')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification du taux de taxe: {str(e)}')
            return redirect('modifier_taux_taxe', taux_id=taux_id)
    
    # GET - Afficher le formulaire pré-rempli
    type_choices = TauxTaxe.TYPE_TAXE_CHOICES
    sens_choices = TauxTaxe.SENS_CHOICES
    assujettissement_choices = TauxTaxe.ASSUJETTISSEMENT_CHOICES
    
    context = {
        'taux': taux,
        'type_choices': type_choices,
        'sens_choices': sens_choices,
        'assujettissement_choices': assujettissement_choices,
    }
    return render(request, 'supermarket/stock/modifier_taux_taxe.html', context)

@login_required
def supprimer_taux_taxe(request, taux_id):
    """Vue pour supprimer un taux de taxe"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_taux_taxe')
    
    try:
        taux = TauxTaxe.objects.get(id=taux_id)
        taux_name = taux.intitule
        taux.delete()
        messages.success(request, f'Taux de taxe "{taux_name}" supprimé avec succès!')
    except TauxTaxe.DoesNotExist:
        messages.error(request, 'Taux de taxe non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_taux_taxe')

@login_required
@require_stock_modify_access
def modifier_article(request, article_id):
    """Vue pour modifier un article existant"""
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
        return redirect('consulter_articles')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            designation = request.POST.get('designation')
            agence_id = request.POST.get('agence')
            prix_achat = request.POST.get('prix_achat')
            prix_vente = request.POST.get('prix_vente')
            stock_actuel = request.POST.get('stock_actuel')
            stock_minimum = request.POST.get('stock_minimum', 0)
            unite_vente = request.POST.get('unite_vente')
            conditionnement = request.POST.get('conditionnement')
            famille_id = request.POST.get('famille')
            # Le suivi de stock est toujours activé automatiquement (non modifiable par l'utilisateur)
            
            ancien_stock = article.stock_actuel
            
            # Validation (stock_actuel est optionnel)
            if not all([designation, agence_id, prix_achat, prix_vente, unite_vente]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_article', article_id=article_id)
            
            # Récupérer l'agence
            try:
                agence = Agence.objects.get(id_agence=agence_id)
            except Agence.DoesNotExist:
                messages.error(request, 'Agence non trouvée.')
                return redirect('modifier_article', article_id=article_id)
            
            # Récupérer la famille si spécifiée
            if famille_id:
                try:
                    categorie = Famille.objects.get(id=famille_id)
                except Famille.DoesNotExist:
                    messages.error(request, 'Famille non trouvée.')
                    return redirect('modifier_article', article_id=article_id)
            else:
                # Si aucune famille n'est fournie, garder l'ancienne
                categorie = article.categorie
            
            # Mettre à jour l'article
            article.designation = designation
            article.agence = agence
            article.prix_achat = Decimal(str(prix_achat)).quantize(Decimal('0.01'))
            article.prix_vente = Decimal(str(prix_vente)).quantize(Decimal('0.01'))
            if stock_actuel and stock_actuel.strip():
                stock_value = stock_actuel.replace(',', '.')
                article.stock_actuel = Decimal(stock_value)
            article.stock_minimum = Decimal(str(stock_minimum)).quantize(Decimal('0.01')) if stock_minimum and stock_minimum.strip() else Decimal('0')
            article.unite_vente = unite_vente
            article.conditionnement = conditionnement
            article.categorie = categorie
            # Le suivi de stock est toujours activé automatiquement
            article.suivi_stock = True
            article.save()
            
            nouveau_stock = article.stock_actuel
            difference_stock = nouveau_stock - ancien_stock
            
            # Créer un mouvement pour TOUS les changements de stock (même si différence = 0 pour traçabilité complète)
            if article.suivi_stock:
                try:
                    numero_piece = f"AJUST-{article.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}"
                    MouvementStock.objects.create(
                        article=article,
                        agence=agence,
                        type_mouvement='ajustement',
                        date_mouvement=timezone.now(),
                        numero_piece=numero_piece,
                        quantite_stock=nouveau_stock,
                        stock_initial=ancien_stock,
                        solde=nouveau_stock,
                        quantite=abs(difference_stock) if difference_stock != Decimal('0') else Decimal('0'),
                        cout_moyen_pondere=float(article.prix_achat),
                        stock_permanent=float(nouveau_stock * article.prix_achat),
                        commentaire=f"Ajustement manuel via modification d'article par {request.user.username}" + (f" (différence: {difference_stock})" if difference_stock != Decimal('0') else " (vérification)")
                    )
                    print(f"[STOCK] Ajustement manuel enregistré: {article.designation} {ancien_stock} -> {nouveau_stock}")
                except Exception as movement_error:
                    print(f"[WARNING] Impossible d'enregistrer le mouvement d'ajustement: {movement_error}")
            
            # Mettre à jour les types de vente
            prix_gros = request.POST.get('prix_gros')
            prix_demi_gros = request.POST.get('prix_demi_gros')
            prix_detail = request.POST.get('prix_detail')
            
            if prix_gros:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Gros',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_gros)}
                )
            
            if prix_demi_gros:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Demi-Gros',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_demi_gros)}
                )
            
            if prix_detail:
                TypeVente.objects.update_or_create(
                    article=article,
                    intitule='Détail',  # CORRECTION: Utiliser 'intitule' au lieu de 'type_vente'
                    defaults={'prix': float(prix_detail)}
                )
            
            messages.success(request, f'Article "{designation}" modifié avec succès!')
            return redirect('consulter_articles')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification de l\'article: {str(e)}')
            return redirect('modifier_article', article_id=article_id)
    
    # GET - Afficher le formulaire pré-rempli
    agences = Agence.objects.all()
    familles = Famille.objects.all()
    types_vente = TypeVente.objects.filter(article=article)
    
    # Créer un dictionnaire des types de vente avec des clés sans traits d'union
    types_vente_dict = {}
    for tv in types_vente:
        if tv.intitule == 'Demi-Gros':
            types_vente_dict['Demi_Gros'] = tv.prix
        elif tv.intitule == 'Détail':
            types_vente_dict['Détail'] = tv.prix
        else:
            types_vente_dict[tv.intitule] = tv.prix
    
    context = {
        'article': article,
        'agences': agences,
        'familles': familles,
        'types_vente': types_vente_dict
    }
    return render(request, 'supermarket/stock/modifier_article.html', context)

@login_required
def supprimer_article(request, article_id):
    """Vue pour supprimer un article"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_articles')
    
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
        article_name = article.designation
        article.delete()
        messages.success(request, f'Article "{article_name}" supprimé avec succès!')
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_articles')

@login_required
def detail_article(request, article_id):
    """Vue pour afficher les détails d'un article"""
    try:
        article = Article.objects.get(id=article_id, agence=get_user_agence(request))
        types_vente = TypeVente.objects.filter(article=article)
        # Le suivi de stock est toujours activé, donc on affiche tous les mouvements
        mouvements = MouvementStock.objects.filter(article=article).order_by('-date_mouvement')[:10]
        
        # Calculer les marges
        marge_unitaire = float(article.prix_vente) - float(article.prix_achat) if article.prix_achat > 0 else 0
        marge_pourcentage = (marge_unitaire / float(article.prix_achat) * 100) if article.prix_achat > 0 else 0
        valeur_stock = float(article.prix_achat) * float(article.stock_actuel)
        
        # Calculer les marges pour chaque type de vente
        types_vente_with_marges = []
        for tv in types_vente:
            marge_tv = float(tv.prix) - float(article.prix_achat) if article.prix_achat > 0 else 0
            marge_tv_pourcentage = (marge_tv / float(article.prix_achat) * 100) if article.prix_achat > 0 else 0
            types_vente_with_marges.append({
                'type_vente': tv,
                'marge': marge_tv,
                'marge_pourcentage': marge_tv_pourcentage
            })
        
        # Debug: Vérifier la famille de l'article
        print(f"[ALERTE] DEBUG Article {article.id}:")
        print(f"   - Désignation: {article.designation}")
        print(f"   - Catégorie: {article.categorie}")
        print(f"   - Intitulé famille: {article.categorie.intitule if article.categorie else 'None'}")
        
        context = {
            'article': article,
            'types_vente': types_vente,
            'types_vente_with_marges': types_vente_with_marges,
            'mouvements': mouvements,
            'marge_unitaire': marge_unitaire,
            'marge_pourcentage': marge_pourcentage,
            'valeur_stock': valeur_stock
        }
        return render(request, 'supermarket/stock/detail_article.html', context)
    except Article.DoesNotExist:
        messages.error(request, 'Article non trouvé.')
        return redirect('consulter_articles')

# ==================== FACTURES D'ACHAT ====================

@login_required
def consulter_factures_achat(request):
    """Vue pour consulter les factures d'achat"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    statut_filter = request.GET.get('statut_filter', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Construire la requête de base
    factures = FactureAchat.objects.filter(agence=agence)
    
    # Appliquer les filtres
    if search_query:
        factures = factures.filter(
            Q(numero_fournisseur__icontains=search_query) |
            Q(reference_achat__icontains=search_query) |
            Q(commentaire__icontains=search_query)
        )
    
    if statut_filter:
        factures = factures.filter(statut=statut_filter)
    
    if date_debut:
        factures = factures.filter(date_achat__gte=date_debut)
    
    if date_fin:
        factures = factures.filter(date_achat__lte=date_fin)
    
    # Trier par date décroissante
    factures = factures.order_by('-date_achat', '-heure')
    
    # Calculer les statistiques
    total_factures = FactureAchat.objects.filter(agence=agence).count()
    factures_validees = FactureAchat.objects.filter(agence=agence, statut='validee').count()
    factures_payees = FactureAchat.objects.filter(agence=agence, statut='payee').count()
    montant_total = FactureAchat.objects.filter(agence=agence).aggregate(
        total=Sum('prix_total_global')
    )['total'] or 0
    
    # Récupérer les statuts pour le filtre
    statut_choices = FactureAchat.STATUT_CHOICES
    
    context = {
        'factures': factures,
        'search_query': search_query,
        'statut_filter': statut_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statut_choices': statut_choices,
        'total_factures': total_factures,
        'factures_validees': factures_validees,
        'factures_payees': factures_payees,
        'montant_total': montant_total,
    }
    return render(request, 'supermarket/stock/consulter_factures_achat.html', context)


@login_required
def consulter_factures_transfert(request):
    """Vue pour consulter les factures de transfert"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les paramètres de recherche
    search_query = request.GET.get('search', '')
    statut_filter = request.GET.get('statut_filter', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Construire la requête de base (factures où l'agence est source ou destination)
    factures = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    )
    
    # Appliquer les filtres
    if search_query:
        factures = factures.filter(
            Q(numero_compte__icontains=search_query) |
            Q(reference_transfert__icontains=search_query) |
            Q(lieu_depart__icontains=search_query) |
            Q(lieu_arrivee__icontains=search_query)
        )
    
    if statut_filter:
        factures = factures.filter(statut=statut_filter)
    
    if date_debut:
        factures = factures.filter(date_transfert__gte=date_debut)
    
    if date_fin:
        factures = factures.filter(date_transfert__lte=date_fin)
    
    # Trier par date décroissante
    factures = factures.order_by('-date_transfert')
    
    # Calculer les statistiques
    total_factures = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    ).count()
    factures_en_cours = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence),
        statut='en_cours'
    ).count()
    factures_terminees = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence),
        statut='termine'
    ).count()
    quantite_totale = FactureTransfert.objects.filter(
        Q(agence_source=agence) | Q(agence_destination=agence)
    ).aggregate(
        total=Sum('quantite')
    )['total'] or 0
    
    # Récupérer les statuts pour le filtre
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'factures': factures,
        'search_query': search_query,
        'statut_filter': statut_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'statut_choices': statut_choices,
        'total_factures': total_factures,
        'factures_en_cours': factures_en_cours,
        'factures_terminees': factures_terminees,
        'quantite_totale': quantite_totale,
    }
    return render(request, 'supermarket/stock/consulter_factures_transfert.html', context)

@login_required
def creer_facture_transfert(request):
    """Vue pour créer une nouvelle facture de transfert"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_compte = request.POST.get('numero_compte')
            date_transfert = request.POST.get('date_transfert')
            reference_transfert = request.POST.get('reference_transfert')
            lieu_depart = request.POST.get('lieu_depart')
            lieu_arrivee = request.POST.get('lieu_arrivee')
            quantite = request.POST.get('quantite')
            statut = request.POST.get('statut')
            employe_expediteur = request.POST.get('employe_expediteur', '').strip()
            employe_destinataire = request.POST.get('employe_destinataire', '').strip()
            etat = request.POST.get('etat', 'sortir')
            
            # Validation
            if not all([numero_compte, date_transfert, reference_transfert, lieu_depart, lieu_arrivee, quantite, employe_expediteur]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_facture_transfert')
            
            # Rechercher les employés existants par nom
            expediteur_employe = None
            destinataire_employe = None
            
            # Rechercher l'employé expéditeur
            try:
                # Essayer de trouver par nom complet dans le compte
                expediteur_employe = Employe.objects.filter(
                    compte__agence=agence,
                    compte__user__first_name__icontains=employe_expediteur.split()[0] if employe_expediteur.split() else employe_expediteur
                ).first()
                
                # Si pas trouvé, essayer par nom de famille
                if not expediteur_employe and len(employe_expediteur.split()) > 1:
                    expediteur_employe = Employe.objects.filter(
                        compte__agence=agence,
                        compte__user__last_name__icontains=employe_expediteur.split()[-1]
                    ).first()
                
                # Si toujours pas trouvé, prendre le premier employé de l'agence
                if not expediteur_employe:
                    expediteur_employe = Employe.objects.filter(compte__agence=agence).first()
                    
            except Exception as e:
                print(f"Erreur lors de la recherche de l'employé expéditeur: {e}")
                expediteur_employe = Employe.objects.filter(compte__agence=agence).first()
            
            if not expediteur_employe:
                messages.error(request, 'Aucun employé trouvé dans cette agence.')
                return redirect('creer_facture_transfert')
            
            # Afficher un message informatif sur l'employé trouvé
            expediteur_nom = f"{expediteur_employe.compte.user.first_name} {expediteur_employe.compte.user.last_name}".strip()
            messages.info(request, f'Employé expéditeur trouvé: {expediteur_nom}')
            
            if destinataire_employe:
                destinataire_nom = f"{destinataire_employe.compte.user.first_name} {destinataire_employe.compte.user.last_name}".strip()
                messages.info(request, f'Employé destinataire trouvé: {destinataire_nom}')
            
            # Rechercher l'employé destinataire (si fourni)
            if employe_destinataire:
                try:
                    # Essayer de trouver par nom complet dans le compte
                    destinataire_employe = Employe.objects.filter(
                        compte__agence=agence,
                        compte__user__first_name__icontains=employe_destinataire.split()[0] if employe_destinataire.split() else employe_destinataire
                    ).first()
                    
                    # Si pas trouvé, essayer par nom de famille
                    if not destinataire_employe and len(employe_destinataire.split()) > 1:
                        destinataire_employe = Employe.objects.filter(
                            compte__agence=agence,
                            compte__user__last_name__icontains=employe_destinataire.split()[-1]
                        ).first()
                        
                except Exception as e:
                    print(f"Erreur lors de la recherche de l'employé destinataire: {e}")
                    destinataire_employe = None
            
            # Vérifier l'unicité de la référence et générer une référence unique si nécessaire
            reference_finale = reference_transfert
            if FactureTransfert.objects.filter(reference_transfert=reference_transfert).exists():
                # Générer une référence unique avec timestamp
                import time
                reference_finale = f"{reference_transfert}_{int(time.time())}"
                print(f"[INFO] Référence modifiée pour éviter le doublon: {reference_finale}")
            
            # Créer la facture de transfert
            facture = FactureTransfert.objects.create(
                numero_compte=numero_compte,
                date_transfert=date_transfert,
                reference_transfert=reference_finale,
                lieu_depart=lieu_depart,
                lieu_arrivee=lieu_arrivee,
                quantite=int(quantite),
                statut=statut,
                agence_source=agence,
                agence_destination=agence,  # Pour l'instant, même agence (à modifier selon les besoins)
                employe_expediteur=expediteur_employe,
                employe_destinataire=destinataire_employe,
                etat=etat
            )
            
            # Traiter les articles sélectionnés
            articles_data = request.POST.get('articles_data', '')
            if articles_data:
                import json
                try:
                    articles = json.loads(articles_data)
                    for article_data in articles:
                        # Récupérer l'article
                        article = Article.objects.get(id=article_data['id'])
                        
                        # Créer la ligne de facture de transfert
                        LigneFactureTransfert.objects.create(
                            facture_transfert=facture,
                            article=article,
                            quantite=int(article_data['quantite']),
                            prix_unitaire=float(article_data['prix_achat']),
                            valeur_totale=float(article_data['prix_achat']) * int(article_data['quantite'])
                        )
                        
                        # Mettre à jour le stock de l'article selon l'état du transfert
                        ancien_stock = article.stock_actuel
                        quantite_transfert = Decimal(str(article_data['quantite']))
                        
                        # Normaliser l'état pour éviter les problèmes de casse ou d'espaces
                        etat_normalise = str(etat).strip().lower() if etat else 'sortir'
                        
                        # Si l'état est "entrer", augmenter le stock; si "sortir", diminuer le stock
                        # IMPORTANT: "entrer" doit TOUJOURS augmenter le stock
                        if etat_normalise == 'entrer':
                            article.stock_actuel += quantite_transfert
                            type_mouvement = 'entree'
                            action_stock = "augmentation"
                        else:  # sortir
                            article.stock_actuel -= quantite_transfert
                            if article.stock_actuel < 0:
                                article.stock_actuel = 0
                            type_mouvement = 'sortie'
                            action_stock = "diminution"
                        
                        # Mettre à jour le dernier prix d'achat avec le prix du transfert
                        ancien_dernier_prix = article.dernier_prix_achat
                        nouveau_prix_achat = float(article_data['prix_achat'])
                        article.dernier_prix_achat = nouveau_prix_achat
                        # Le suivi de stock est toujours activé automatiquement lors de toute modification du stock
                        article.suivi_stock = True
                        article.save()
                        print(f"[PACKAGE] STOCK TRANSFERT - Article: {article.designation}")
                        print(f"[PACKAGE] STOCK TRANSFERT - État: {etat} (normalisé: {etat_normalise}), {action_stock} du stock")
                        print(f"[PACKAGE] STOCK TRANSFERT - Stock mis à jour: {ancien_stock} → {article.stock_actuel}")
                        print(f"[MONEY] Transfert - Dernier prix d'achat mis à jour: {ancien_dernier_prix} → {nouveau_prix_achat}")
                        
                        # [HOT] CRÉER UN MOUVEMENT DE STOCK POUR TRAÇABILITÉ
                        try:
                            MouvementStock.objects.create(
                                article=article,
                                agence=agence,
                                type_mouvement=type_mouvement,
                                date_mouvement=timezone.now(),
                                numero_piece=facture.reference_transfert,
                                quantite_stock=article.stock_actuel,
                                stock_initial=ancien_stock,
                                solde=article.stock_actuel,
                                quantite=int(article_data['quantite']),
                                cout_moyen_pondere=float(article.prix_achat),
                                stock_permanent=float(article.stock_actuel * article.prix_achat),
                                commentaire=f"Transfert ({etat}) - Facture {facture.reference_transfert}"
                            )
                            print(f"[NOTE] MOUVEMENT STOCK - Transfert ({etat}) enregistré pour {article.designation}")
                        except Exception as e:
                            print(f"[WARNING] ERREUR MOUVEMENT STOCK TRANSFERT: {e}")
                        
                except (json.JSONDecodeError, Article.DoesNotExist, KeyError) as e:
                    print(f"Erreur lors du traitement des articles: {e}")
            
            messages.success(request, f'Facture de transfert "{reference_transfert}" créée avec succès!')
            return redirect('creer_facture_transfert')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de la facture: {str(e)}')
            return redirect('creer_facture_transfert')
    
    # GET - Afficher le formulaire
    statut_choices = FactureTransfert.STATUT_CHOICES
    
    context = {
        'statut_choices': statut_choices,
    }
    return render(request, 'supermarket/stock/creer_facture_transfert_nouveau.html', context)

@login_required
def detail_facture_transfert(request, facture_id):
    """Vue pour afficher les détails d'une facture de transfert"""
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
        lignes = LigneFactureTransfert.objects.filter(facture_transfert=facture)
        
        context = {
            'facture': facture,
            'lignes': lignes,
        }
        return render(request, 'supermarket/stock/detail_facture_transfert.html', context)
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
        return redirect('consulter_factures_transfert')

@login_required
def modifier_facture_transfert(request, facture_id):
    """Vue pour modifier une facture de transfert existante"""
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.get(id=facture_id, agence_source=agence)
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
        return redirect('consulter_factures_transfert')
    
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            numero_compte = request.POST.get('numero_compte')
            date_transfert = request.POST.get('date_transfert')
            reference_transfert = request.POST.get('reference_transfert')
            lieu_depart = request.POST.get('lieu_depart')
            lieu_arrivee = request.POST.get('lieu_arrivee')
            quantite = request.POST.get('quantite')
            statut = request.POST.get('statut')
            employe_expediteur = request.POST.get('employe_expediteur', '').strip()
            employe_destinataire = request.POST.get('employe_destinataire', '').strip()
            etat = request.POST.get('etat', 'sortir')
            
            articles_data_raw = request.POST.get('articles_data')
            
            # Validation
            if not all([numero_compte, date_transfert, reference_transfert, lieu_depart, lieu_arrivee, quantite]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            if not articles_data_raw:
                messages.error(request, 'Veuillez sélectionner au moins un article.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            try:
                articles_payload = json.loads(articles_data_raw)
            except json.JSONDecodeError:
                messages.error(request, 'Format des articles invalide.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            if not articles_payload:
                messages.error(request, 'Veuillez sélectionner au moins un article.')
                return redirect('modifier_facture_transfert', facture_id=facture_id)
            
            # Récupérer l'ancien état pour gérer le stock
            ancien_etat = facture.etat
            ancien_etat_normalise = str(ancien_etat).strip().lower() if ancien_etat else 'sortir'
            etat_normalise = str(etat).strip().lower() if etat else 'sortir'
            
            # Nouvelle logique : Calculer la différence et l'appliquer directement
            # Prendre en compte l'état (entrer/sortir) et la différence de quantité
            # Ajouter un flag pour empêcher la mise à jour automatique du stock
            request.session['modification_en_cours'] = True
            with transaction.atomic():
                    # 1) Récupérer les anciennes lignes et créer un dictionnaire
                    lignes_existantes = LigneFactureTransfert.objects.select_related('article').filter(facture_transfert=facture)
                    anciennes_quantites = {}
                    for ligne in lignes_existantes:
                        if ligne.article:
                            anciennes_quantites[ligne.article.id] = Decimal(str(ligne.quantite))
                    
                    # 2) Supprimer toutes les anciennes lignes
                    LigneFactureTransfert.objects.filter(facture_transfert=facture).delete()
                    
                    # 3) Mettre à jour la facture de transfert
                    facture.numero_compte = numero_compte
                    facture.date_transfert = date_transfert
                    facture.reference_transfert = reference_transfert
                    facture.lieu_depart = lieu_depart
                    facture.lieu_arrivee = lieu_arrivee
                    facture.statut = statut
                    facture.etat = etat
                    
                    total_quantite = Decimal('0')
                    
                    # 4) Traiter les nouvelles lignes avec la logique de différence
                    for article_data in articles_payload:
                        article_id = article_data.get('id')
                        quantite_nouvelle = Decimal(str(article_data.get('quantite', 0)))
                        prix_achat_ligne = Decimal(str(article_data.get('prix_achat', 0)))
                        
                        if not article_id or quantite_nouvelle <= 0:
                            continue
                        
                        try:
                            article = Article.objects.get(id=article_id)
                            article.refresh_from_db()
                        except Article.DoesNotExist:
                            messages.error(request, f"Article introuvable (ID: {article_id}).")
                            raise
                        
                        # Vérifier le stock si c'est une sortie
                        if etat_normalise == 'sortir' and article.stock_actuel < quantite_nouvelle:
                            messages.error(request, f"Stock insuffisant pour l'article {article.designation}.")
                            raise ValueError("Stock insuffisant")
                        
                        # Créer la nouvelle ligne
                        LigneFactureTransfert.objects.create(
                            facture_transfert=facture,
                            article=article,
                            quantite=int(quantite_nouvelle),
                            prix_unitaire=prix_achat_ligne,
                            valeur_totale=quantite_nouvelle * prix_achat_ligne
                        )
                        
                        # Calculer la différence et l'effet sur le stock
                        quantite_ancienne = anciennes_quantites.get(article_id, Decimal('0'))
                        difference = quantite_nouvelle - quantite_ancienne
                        
                        # Stock avant modification (ce stock contient déjà l'effet de l'ancienne facture)
                        stock_avant_modification = article.stock_actuel
                        
                        # LOGIQUE CORRIGÉE : Le stock actuel contient déjà l'effet de l'ancienne facture
                        # Pour modifier : il faut annuler l'effet de l'ancienne facture, puis appliquer le nouvel effet
                        # Ce qui équivaut à : stock_final = stock_actuel - effet_ancien + effet_nouveau
                        
                        # Vérifier si l'état a changé
                        etat_a_change = (ancien_etat_normalise != etat_normalise)
                        
                        # Calculer le stock final selon l'état
                        if etat_normalise == 'entrer':
                            # État "entrer" : augmente le stock
                            # Le stock actuel contient déjà l'effet de l'ancienne facture
                            # Si ancien état était "entrer" : stock_actuel = stock_initial + ancienne_quantité
                            #   -> stock_final = stock_actuel - ancienne_quantité + nouvelle_quantité = stock_actuel + différence
                            # Si ancien état était "sortir" : stock_actuel = stock_initial - ancienne_quantité
                            #   -> stock_final = stock_actuel + ancienne_quantité + nouvelle_quantité = stock_actuel + ancienne_quantité + nouvelle_quantité
                            if etat_a_change and ancien_etat_normalise == 'sortir':
                                # L'état a changé de "sortir" à "entrer"
                                # Il faut remettre l'ancienne quantité (qui avait été soustraite) et ajouter la nouvelle
                                stock_final = stock_avant_modification + quantite_ancienne + quantite_nouvelle
                            else:
                                # L'état n'a pas changé ou était déjà "entrer"
                                # stock_final = stock_actuel + différence
                                stock_final = stock_avant_modification + difference
                            
                            # S'assurer que le stock ne devient pas négatif
                            if stock_final < 0:
                                messages.error(request, f'Stock insuffisant pour l\'article {article.designation}. Stock actuel: {stock_avant_modification}, différence: {difference}')
                                raise ValueError(f"Stock insuffisant: {stock_avant_modification} + {difference} = {stock_final} < 0")
                            
                            article.stock_actuel = stock_final
                            
                            # Déterminer le type de mouvement et le commentaire selon la différence
                            if difference > 0:
                                type_mouvement = 'entree'
                                quantite_mouvement = difference if not etat_a_change else quantite_nouvelle
                                commentaire = f"Transfert entrer - Facture {facture.reference_transfert} (modification - augmentation: {quantite_ancienne} → {quantite_nouvelle}, +{difference})"
                            elif difference < 0:
                                type_mouvement = 'sortie'
                                quantite_mouvement = abs(difference) if not etat_a_change else quantite_ancienne
                                commentaire = f"Transfert entrer - Facture {facture.reference_transfert} (modification - diminution: {quantite_ancienne} → {quantite_nouvelle}, -{abs(difference)})"
                            else:
                                type_mouvement = None
                                quantite_mouvement = Decimal('0')
                                commentaire = f"Transfert entrer - Facture {facture.reference_transfert} (modification - quantité inchangée: {quantite_nouvelle})"
                        else:
                            # État "sortir" : diminue le stock
                            # Le stock actuel contient déjà l'effet de l'ancienne facture
                            # Si ancien état était "sortir" : stock_actuel = stock_initial - ancienne_quantité
                            #   -> stock_final = stock_actuel + ancienne_quantité - nouvelle_quantité = stock_actuel - différence
                            # Si ancien état était "entrer" : stock_actuel = stock_initial + ancienne_quantité
                            #   -> stock_final = stock_actuel - ancienne_quantité - nouvelle_quantité = stock_actuel - ancienne_quantité - nouvelle_quantité
                            if etat_a_change and ancien_etat_normalise == 'entrer':
                                # L'état a changé de "entrer" à "sortir"
                                # Il faut retirer l'ancienne quantité (qui avait été ajoutée) et soustraire la nouvelle
                                stock_final = max(Decimal('0'), stock_avant_modification - quantite_ancienne - quantite_nouvelle)
                            else:
                                # L'état n'a pas changé ou était déjà "sortir"
                                # stock_final = stock_actuel - différence
                                stock_final = max(Decimal('0'), stock_avant_modification - difference)
                            
                            # Vérifier le stock avant d'appliquer
                            if stock_final < 0:
                                messages.error(request, f'Stock insuffisant pour l\'article {article.designation}. Stock actuel: {stock_avant_modification}, différence: {difference}')
                                raise ValueError(f"Stock insuffisant: {stock_avant_modification} - {difference} = {stock_final} < 0")
                            
                            article.stock_actuel = stock_final
                            
                            # Déterminer le type de mouvement et le commentaire selon la différence
                            if difference > 0:
                                type_mouvement = 'sortie'
                                quantite_mouvement = difference if not etat_a_change else quantite_nouvelle
                                commentaire = f"Transfert sortir - Facture {facture.reference_transfert} (modification - augmentation: {quantite_ancienne} → {quantite_nouvelle}, -{difference})"
                            elif difference < 0:
                                type_mouvement = 'entree'
                                quantite_mouvement = abs(difference) if not etat_a_change else quantite_ancienne
                                commentaire = f"Transfert sortir - Facture {facture.reference_transfert} (modification - diminution: {quantite_ancienne} → {quantite_nouvelle}, +{abs(difference)})"
                            else:
                                type_mouvement = None
                                quantite_mouvement = Decimal('0')
                                commentaire = f"Transfert sortir - Facture {facture.reference_transfert} (modification - quantité inchangée: {quantite_nouvelle})"
                        
                        article.dernier_prix_achat = prix_achat_ligne
                        # Le suivi de stock est toujours activé automatiquement lors de toute modification du stock
                        article.suivi_stock = True
                        article.save()
                        
                        total_quantite += quantite_nouvelle
                        
                        # Créer un mouvement de stock pour refléter le changement
                        if quantite_mouvement > 0 and type_mouvement:
                            try:
                                MouvementStock.objects.create(
                                    article=article,
                                    agence=agence,
                                    type_mouvement=type_mouvement,
                                    date_mouvement=timezone.now(),
                                    numero_piece=facture.reference_transfert,
                                    quantite_stock=article.stock_actuel,
                                    stock_initial=stock_avant_modification,
                                    solde=article.stock_actuel,
                                    quantite=quantite_mouvement,
                                    cout_moyen_pondere=float(article.prix_achat),
                                    stock_permanent=float(article.stock_actuel * article.prix_achat),
                                    facture_transfert=facture,
                                    commentaire=commentaire
                                )
                            except Exception as e:
                                print(f"[WARNING] ERREUR MOUVEMENT STOCK TRANSFERT MODIFICATION: {e}")
                    
                    facture.quantite = int(total_quantite)
                    facture.save()
            
            # Retirer le flag de modification en cours
            request.session.pop('modification_en_cours', None)
            
            messages.success(request, f'Facture de transfert "{reference_transfert}" modifiée avec succès!')
            return redirect('consulter_factures_transfert')
            
        except Exception as e:
            # Retirer le flag en cas d'erreur
            request.session.pop('modification_en_cours', None)
            messages.error(request, f'Erreur lors de la modification de la facture: {str(e)}')
            return redirect('modifier_facture_transfert', facture_id=facture_id)
    
    # GET - Afficher le formulaire pré-rempli
    statut_choices = FactureTransfert.STATUT_CHOICES
    lignes = LigneFactureTransfert.objects.select_related('article').filter(facture_transfert=facture)
    
    initial_articles = []
    for ligne in lignes:
        if ligne.article:
            initial_articles.append({
                'id': ligne.article.id,
                'designation': ligne.article.designation,
                'reference_article': ligne.article.reference_article,
                'stock_actuel': float(ligne.article.stock_actuel),
                'prix_achat': float(ligne.prix_unitaire),
                'quantite': ligne.quantite
            })
    
    etat_actuel = (facture.etat or 'sortir').strip().lower()
    if etat_actuel == 'sortie':
        etat_actuel = 'sortir'
    
    context = {
        'facture': facture,
        'statut_choices': statut_choices,
        'initial_articles_json': json.dumps(initial_articles),
        'etat_actuel': etat_actuel,
    }
    return render(request, 'supermarket/stock/modifier_facture_transfert.html', context)

@login_required
def supprimer_facture_transfert(request, facture_id):
    """Vue pour supprimer une facture de transfert"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('consulter_factures_transfert')
    
    try:
        agence = get_user_agence(request)
        facture = FactureTransfert.objects.select_related('agence_source').get(id=facture_id, agence_source=agence)
        facture_name = facture.reference_transfert
        etat_normalise = (facture.etat or 'sortir').strip().lower()
        
        with transaction.atomic():
            # NE PLUS SUPPRIMER LES MOUVEMENTS - Créer des mouvements de correction à la place
            lignes = LigneFactureTransfert.objects.select_related('article').filter(facture_transfert=facture)
            
            for ligne in lignes:
                article = ligne.article
                if not article:
                    continue
                
                quantite = Decimal(str(ligne.quantite))
                ancien_stock = article.stock_actuel
                
                if etat_normalise == 'entrer':
                    article.stock_actuel = max(Decimal('0'), article.stock_actuel - quantite)
                    commentaire = f"Correction - Suppression facture transfert {facture.reference_transfert} - retrait stock"
                else:
                    article.stock_actuel += quantite
                    commentaire = f"Correction - Suppression facture transfert {facture.reference_transfert} - restauration stock"
                
                article.save()
                
                # Créer un mouvement de correction (les mouvements originaux sont conservés pour traçabilité)
                try:
                    MouvementStock.objects.create(
                        article=article,
                        agence=agence,
                        type_mouvement='ajustement',
                        date_mouvement=timezone.now(),
                        numero_piece=f"SUPP-{facture.reference_transfert}",
                        quantite_stock=article.stock_actuel,
                        stock_initial=ancien_stock,
                        solde=article.stock_actuel,
                        quantite=quantite,
                        cout_moyen_pondere=float(article.prix_achat),
                        stock_permanent=float(article.stock_actuel * article.prix_achat),
                        commentaire=commentaire,
                        facture_transfert=facture
                    )
                except Exception as mouvement_error:
                    print(f"[WARNING] Erreur enregistrement mouvement correction suppression transfert: {mouvement_error}")
            
            facture.delete()
        
        messages.success(request, f'Facture de transfert "{facture_name}" supprimée avec succès!')
    except FactureTransfert.DoesNotExist:
        messages.error(request, 'Facture de transfert non trouvée.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('consulter_factures_transfert')

# ===== RECHERCHE D'ARTICLES POUR STOCK =====

@login_required
@require_stock_access
def search_articles_stock(request):
    """Vue pour la recherche d'articles dans le module de stock"""
    search_term = request.GET.get('q', '')
    
    print(f"[SEARCH] search_articles_stock: recherche pour '{search_term}'")
    
    # Vérifier d'abord s'il y a des articles dans la base de données
    total_articles = Article.objects.count()
    print(f"[CHART] Total articles dans la base de données: {total_articles}")
    
    if total_articles == 0:
        print("[ERREUR] Aucun article dans la base de données!")
        return JsonResponse({'articles': []})
    
    agence = get_user_agence(request)
    print(f"[BUILDING] Agence trouvée: {agence}")
    print(f"[BUILDING] ID de l'agence: {agence.id_agence if agence else 'None'}")
    print(f"[BUILDING] Nom de l'agence: {agence.nom_agence if agence else 'None'}")
    
    if not agence:
        print("[ERREUR] Aucune agence trouvée")
        return JsonResponse({'articles': []})
    
    # Test: afficher tous les articles sans filtre d'agence
    articles_all = Article.objects.all()
    print(f"[PACKAGE] Tous les articles (toutes agences): {articles_all.count()}")
    for article in articles_all[:5]:  # Afficher les 5 premiers
        print(f"  - {article.designation} (Agence: {article.agence.nom_agence if article.agence else 'None'})")
    
    # Test: afficher les articles de cette agence spécifique
    articles_agence = Article.objects.filter(agence=agence)
    print(f"[PACKAGE] Articles de l'agence {agence.nom_agence}: {articles_agence.count()}")
    for article in articles_agence[:5]:  # Afficher les 5 premiers
        print(f"  - {article.reference_article} - {article.designation} (ID: {article.id})")
    
    articles = []
    
    if search_term and len(search_term) >= 1:
        # Recherche avec filtre d'agence : recherche dans designation ET reference_article
        from django.db.models import Q
        articles = Article.objects.filter(
            agence=agence
        ).filter(
            Q(designation__icontains=search_term) | Q(reference_article__icontains=search_term)
        ).order_by('designation')[:50]
        print(f"[SEARCH] Articles trouvés avec recherche '{search_term}' (agence {agence.nom_agence}): {articles.count()}")
        for article in articles[:10]:  # Afficher les 10 premiers résultats
            print(f"  - {article.reference_article} - {article.designation} (ID: {article.id})")
    else:
        # Afficher tous les articles de l'agence si pas de terme de recherche
        articles = Article.objects.filter(agence=agence).order_by('designation')[:50]
        print(f"[PACKAGE] Tous les articles de l'agence {agence.nom_agence}: {articles.count()}")
    
    # Convertir les articles en format JSON
    articles_data = []
    for article in articles:
        articles_data.append({
            'id': article.id,
            'designation': article.designation,
            'prix_achat': float(article.prix_achat) if article.prix_achat else 0.0,
            'stock': float(article.stock_actuel) if article.stock_actuel else 0.0,
            'reference_article': article.reference_article or '',
        })
        print(f"[NOTE] Article ajouté: {article.reference_article} - {article.designation} (ID: {article.id}, Agence: {article.agence.nom_agence})")
    
    print(f"[CHART] Total articles_data: {len(articles_data)}")
    return JsonResponse({'articles': articles_data})

def create_test_articles(request):
    """Vue temporaire pour créer des articles de test"""
    agence = get_user_agence(request)
    if not agence:
        return JsonResponse({'error': 'Aucune agence trouvée'})
    
    # Créer quelques articles de test
    test_articles = [
        {'designation': 'Ordinateur Portable', 'prix_achat': 500000, 'prix_vente': 600000, 'stock_actuel': 10},
        {'designation': 'Souris USB', 'prix_achat': 5000, 'prix_vente': 7500, 'stock_actuel': 50},
        {'designation': 'Clavier Mécanique', 'prix_achat': 15000, 'prix_vente': 20000, 'stock_actuel': 25},
        {'designation': 'Écran 24 pouces', 'prix_achat': 80000, 'prix_vente': 100000, 'stock_actuel': 15},
        {'designation': 'Casque Audio', 'prix_achat': 25000, 'prix_vente': 35000, 'stock_actuel': 30},
    ]
    
    created_articles = []
    for article_data in test_articles:
        article, created = Article.objects.get_or_create(
            designation=article_data['designation'],
            agence=agence,
            defaults=article_data
        )
        if created:
            created_articles.append(article.designation)
    
    return JsonResponse({
        'message': f'Articles créés: {len(created_articles)}',
        'articles': created_articles
    })
# ==================== INVENTAIRE DE STOCK ====================

@login_required
@require_stock_access
def inventaire_stock(request):
    """Vue pour la page d'inventaire de stock"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques du stock
    total_articles = articles.count()
    total_quantite = articles.aggregate(total=Sum('stock_actuel'))['total'] or 0
    valeur_totale_stock = articles.aggregate(
        total=Sum(F('stock_actuel') * F('prix_achat'))
    )['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'total_quantite': total_quantite,
        'valeur_totale_stock': valeur_totale_stock,
    }
    
    return render(request, 'supermarket/stock/inventaire_stock.html', context)

@login_required
def generer_inventaire(request):
    """Vue pour générer l'inventaire selon les critères sélectionnés"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        famille_id = request.POST.get('famille')
        selection_articles = request.POST.get('selection_articles', 'tous')  # 'tous' ou 'selectionnes'
        articles_selectionnes = request.POST.getlist('articles_selectionnes[]')
        
        print(f"[SEARCH] PARAMÈTRES INVENTAIRE:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Famille: {famille_id}")
        print(f"  - Sélection: {selection_articles}")
        print(f"  - Articles sélectionnés: {articles_selectionnes}")
        
        # Filtrer les articles selon les critères
        articles_query = Article.objects.filter(agence=agence)
        
        # Filtre par famille
        if famille_id and famille_id != '':
            articles_query = articles_query.filter(categorie_id=famille_id)
        
        # Filtre par sélection d'articles
        if selection_articles == 'selectionnes' and articles_selectionnes:
            articles_query = articles_query.filter(id__in=articles_selectionnes)
        
        articles = articles_query.order_by('designation')
        
        print(f"[PACKAGE] ARTICLES FILTRÉS: {articles.count()}")
        
        # Créer l'inventaire
        numero_inventaire = f"INV-{timezone.now().strftime('%Y%m%d-%H%M%S')}"
        
        # Récupérer l'employé responsable
        employe = Employe.objects.filter(compte__agence=agence).first()
        
        inventaire = InventaireStock.objects.create(
            numero_inventaire=numero_inventaire,
            date_debut=timezone.now(),
            statut='en_cours',
            agence=agence,
            responsable=employe,
            commentaire=f"Inventaire généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
        )
        
        # Créer les lignes d'inventaire
        total_quantite = 0
        total_valeur = 0
        
        for article in articles:
            valeur = float(article.stock_actuel) * float(article.prix_achat)
            
            LigneInventaireStock.objects.create(
                inventaire=inventaire,
                reference_article=article.reference_article,
                designation=article.designation,
                quantite_stock=article.stock_actuel,
                prix_unitaire=article.prix_achat,
                valeur=valeur,
                conditionnement=article.conditionnement,
                article=article
            )
            
            total_quantite += article.stock_actuel
            total_valeur += valeur
        
        # Marquer l'inventaire comme terminé
        inventaire.date_fin = timezone.now()
        inventaire.statut = 'termine'
        inventaire.save()
        
        print(f"[OK] INVENTAIRE CRÉÉ: {numero_inventaire}")
        print(f"[CHART] TOTAUX: {total_quantite} articles, {total_valeur} FCFA")
        
        return JsonResponse({
            'success': True,
            'inventaire_id': inventaire.id,
            'numero_inventaire': numero_inventaire,
            'total_articles': articles.count(),
            'total_quantite': total_quantite,
            'total_valeur': total_valeur
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR GÉNÉRATION INVENTAIRE: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_excel(request):
    """Vue pour exporter l'inventaire en format Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible, utilisation du format CSV")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Inventaire {inventaire.numero_inventaire}"
        
        # Style du titre
        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire le titre "Inventaire" avec la date du jour
        date_du_jour = timezone.now().strftime('%d/%m/%Y')
        title_cell = ws.cell(row=1, column=1, value=f"Inventaire - {date_du_jour}")
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # Fusionner les cellules pour le titre (sur toutes les colonnes)
        from openpyxl.utils import get_column_letter
        ws.merge_cells(f'A1:{get_column_letter(6)}1')
        
        # Ligne vide
        ws.row_dimensions[2].height = 5
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité en Stock', 'Prix Unitaire (FCFA)', 'Valeur Totale (FCFA)']
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Écrire les en-têtes (décalés à la ligne 3)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Écrire les données (décalées à partir de la ligne 4)
        total_quantite = 0
        total_valeur = 0
        
        for idx, ligne in enumerate(lignes):
            row_num = 4 + idx  # Commencer à la ligne 4 (après titre ligne 1, ligne vide ligne 2, en-têtes ligne 3)
            ws.cell(row=row_num, column=1, value=ligne.reference_article)
            ws.cell(row=row_num, column=2, value=ligne.designation)
            ws.cell(row=row_num, column=3, value=ligne.conditionnement)
            ws.cell(row=row_num, column=4, value=ligne.quantite_stock)
            ws.cell(row=row_num, column=5, value=float(ligne.prix_unitaire))
            ws.cell(row=row_num, column=6, value=float(ligne.valeur))
            
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux (décalée de 2 lignes supplémentaires)
        total_row = 4 + len(lignes) + 2  # Ligne de données + 1 ligne vide + ligne totaux
        ws.cell(row=total_row, column=3, value="TOTAL GÉNÉRAL:").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_quantite).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=total_valeur).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        column_widths = [15, 40, 15, 15, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le classeur dans la réponse
        wb.save(response)
        
        print(f"[CHART] EXPORT EXCEL - Inventaire {inventaire.numero_inventaire}")
        print(f"[CHART] EXPORT EXCEL - {len(lignes)} articles exportés")
        print(f"[CHART] EXPORT EXCEL - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_pdf(request):
    """Vue pour exporter l'inventaire en format PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Centre
        )
        
        # Titre
        title = Paragraph(f"INVENTAIRE DE STOCK - {inventaire.numero_inventaire}", title_style)
        elements.append(title)
        
        # Informations de l'inventaire
        info_data = [
            ['Date de génération:', inventaire.date_debut.strftime('%d/%m/%Y à %H:%M')],
            ['Agence:', agence.nom_agence],
            ['Responsable:', inventaire.responsable.compte.nom_complet if inventaire.responsable else 'Non spécifié'],
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 8*cm])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 20))
        
        # Tableau des articles
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité', 'Prix Unitaire', 'Valeur']
        
        # Données du tableau
        data = [headers]
        total_quantite = 0
        total_valeur = 0
        
        for ligne in lignes:
            row = [
                ligne.reference_article,
                ligne.designation,
                ligne.conditionnement,
                str(ligne.quantite_stock),
                f"{float(ligne.prix_unitaire):,.0f}",
                f"{float(ligne.valeur):,.0f}"
            ]
            data.append(row)
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux
        data.append(['', '', 'TOTAL GÉNÉRAL:', str(total_quantite), '', f"{total_valeur:,.0f}"])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 6*cm, 2.5*cm, 2*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            # En-têtes
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Données
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Ligne des totaux
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF - Inventaire {inventaire.numero_inventaire}")
        print(f"📄 EXPORT PDF - {len(lignes)} articles exportés")
        print(f"📄 EXPORT PDF - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_inventaire_csv(request):
    """Vue pour exporter l'inventaire en format CSV (alternative si Excel n'est pas disponible)"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer l'ID de l'inventaire depuis la session ou les paramètres
        inventaire_id = request.GET.get('inventaire_id') or request.session.get('last_inventaire_id')
        
        if not inventaire_id:
            return JsonResponse({'success': False, 'error': 'Aucun inventaire spécifié'})
        
        # Récupérer l'inventaire
        try:
            inventaire = InventaireStock.objects.get(id=inventaire_id, agence=agence)
        except InventaireStock.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Inventaire non trouvé'})
        
        # Récupérer les lignes d'inventaire
        lignes = LigneInventaireStock.objects.filter(inventaire=inventaire).order_by('designation')
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Inventaire_{inventaire.numero_inventaire}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Conditionnement', 'Quantité en Stock', 'Prix Unitaire (FCFA)', 'Valeur Totale (FCFA)']
        writer.writerow(headers)
        
        # Données
        total_quantite = 0
        total_valeur = 0
        
        for ligne in lignes:
            row = [
                ligne.reference_article,
                ligne.designation,
                ligne.conditionnement,
                ligne.quantite_stock,
                float(ligne.prix_unitaire),
                float(ligne.valeur)
            ]
            writer.writerow(row)
            total_quantite += ligne.quantite_stock
            total_valeur += float(ligne.valeur)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['', '', 'TOTAL GÉNÉRAL:', total_quantite, '', total_valeur])
        
        print(f"[CHART] EXPORT CSV - Inventaire {inventaire.numero_inventaire}")
        print(f"[CHART] EXPORT CSV - {len(lignes)} articles exportés")
        print(f"[CHART] EXPORT CSV - Total: {total_quantite} articles, {total_valeur} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== STATISTIQUES DE VENTE ====================

@login_required
def statistiques_vente(request):
    """Vue pour la page des statistiques de vente"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques générales
    total_articles = articles.count()
    
    # Calculer les statistiques de vente des 30 derniers jours
    date_debut = timezone.now().date() - timezone.timedelta(days=30)
    
    # Récupérer les ventes des 30 derniers jours
    ventes_recentes = LigneFactureVente.objects.filter(
        facture_vente__agence=agence,
        facture_vente__date__gte=date_debut
    ).select_related('article', 'facture_vente')
    
    # Calculer le chiffre d'affaires total
    chiffre_affaires_total = float(ventes_recentes.aggregate(
        total=Sum(F('quantite') * F('prix_unitaire'))
    )['total'] or 0)
    
    # Calculer la marge totale
    marge_totale = Decimal('0')
    for vente in ventes_recentes:
        prix_achat = Decimal(str(vente.article.prix_achat))
        prix_vente = Decimal(str(vente.prix_unitaire))
        marge_unitaire = prix_vente - prix_achat
        marge_totale += marge_unitaire * Decimal(str(vente.quantite))
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'chiffre_affaires_total': chiffre_affaires_total,
        'marge_totale': float(marge_totale),
        'pourcentage_marge_global': (float(marge_totale) / float(chiffre_affaires_total) * 100) if chiffre_affaires_total > 0 else 0,
    }
    
    return render(request, 'supermarket/stock/statistiques_vente.html', context)

@login_required
def generer_statistiques_vente(request):
    """Vue pour générer les statistiques de vente selon les critères sélectionnés"""
    print("[START] DÉBUT GENERER_STATISTIQUES_VENTE")
    
    if request.method != 'POST':
        print("[ERREUR] Méthode non POST")
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
        print(f"[SEARCH] Agence récupérée: {agence}")
    except Exception as e:
        print(f"[ERREUR] Erreur get_user_agence: {e}")
        return JsonResponse({'success': False, 'error': f'Agence non trouvée: {str(e)}'})
    
    try:
        print("[SEARCH] Début du traitement des paramètres")
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        famille_id = request.POST.get('famille')
        selection_articles = request.POST.get('selection_articles', 'tous')
        articles_selectionnes = request.POST.getlist('articles_selectionnes[]')
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        print(f"[CHART] PARAMÈTRES STATISTIQUES:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Famille: {famille_id}")
        print(f"  - Sélection: {selection_articles}")
        print(f"  - Articles sélectionnés: {articles_selectionnes}")
        
        # Convertir les dates
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Format de date invalide: {str(e)}'})
        
        # Filtrer les articles selon les critères
        print(f"[SEARCH] Filtrage des articles pour agence: {agence}")
        articles_query = Article.objects.filter(agence=agence)
        print(f"[SEARCH] Articles de base: {articles_query.count()}")
        
        # Filtre par famille
        if famille_id and famille_id != '':
            print(f"[SEARCH] Filtrage par famille: {famille_id}")
            articles_query = articles_query.filter(categorie_id=famille_id)
            print(f"[SEARCH] Articles après filtre famille: {articles_query.count()}")
        
        # Filtre par sélection d'articles
        if selection_articles == 'selectionnes' and articles_selectionnes:
            print(f"[SEARCH] Filtrage par sélection: {articles_selectionnes}")
            articles_query = articles_query.filter(id__in=articles_selectionnes)
            print(f"[SEARCH] Articles après filtre sélection: {articles_query.count()}")
        
        articles = articles_query.order_by('designation')
        print(f"[PACKAGE] ARTICLES FILTRÉS: {articles.count()}")
        
        # Récupérer les ventes pour la période
        ventes = LigneFactureVente.objects.filter(
            facture_vente__agence=agence,
            facture_vente__date__gte=date_debut_obj,
            facture_vente__date__lte=date_fin_obj,
            article__in=articles
        ).select_related('article', 'facture_vente')
        
        # Calculer les statistiques par article
        statistiques_articles = []
        chiffre_affaires_total = 0.0
        marge_totale = 0.0
        quantite_totale_vendue = 0.0
        
        for article in articles:
            # Récupérer les ventes de cet article
            ventes_article = ventes.filter(article=article)
            
            # Calculer les totaux pour cet article
            quantite_vendue = float(ventes_article.aggregate(total=Sum('quantite'))['total'] or 0)
            chiffre_affaires_article = float(ventes_article.aggregate(
                total=Sum(F('quantite') * F('prix_unitaire'))
            )['total'] or 0)
            
            # Calculer la marge
            prix_achat = Decimal(str(article.prix_achat))
            marge_unitaire = Decimal('0')
            marge_article = Decimal('0')
            
            if quantite_vendue > 0:
                prix_vente_moyen = Decimal(str(chiffre_affaires_article)) / Decimal(str(quantite_vendue))
                marge_unitaire = prix_vente_moyen - prix_achat
                marge_article = marge_unitaire * Decimal(str(quantite_vendue))
            
            # Calculer le pourcentage de marge
            if chiffre_affaires_article > 0:
                pourcentage_marge = (marge_article / Decimal(str(chiffre_affaires_article)) * Decimal('100'))
            else:
                pourcentage_marge = Decimal('0')
            
            if quantite_vendue > 0:  # Ne garder que les articles vendus
                statistiques_articles.append({
                    'reference_article': article.reference_article,
                    'designation': article.designation,
                    'quantite_vendue': float(quantite_vendue),
                    'chiffre_affaires': float(chiffre_affaires_article),
                    'marge_profit': float(marge_article),
                    'pourcentage_marge': float(pourcentage_marge),
                })
                
                chiffre_affaires_total += float(chiffre_affaires_article)
                marge_totale += float(marge_article)
                quantite_totale_vendue += float(quantite_vendue)
        
        # Calculer le pourcentage de marge global
        if chiffre_affaires_total > 0:
            pourcentage_marge_global = (marge_totale / float(chiffre_affaires_total) * 100)
        else:
            pourcentage_marge_global = 0
        
        print(f"[CHART] STATISTIQUES GÉNÉRÉES:")
        print(f"  - Articles vendus: {len(statistiques_articles)}")
        print(f"  - Quantité totale vendue: {quantite_totale_vendue}")
        print(f"  - Chiffre d'affaires total: {chiffre_affaires_total}")
        print(f"  - Marge totale: {marge_totale}")
        print(f"  - Pourcentage marge global: {pourcentage_marge_global:.2f}%")
        
        # Stocker les statistiques dans la session pour l'export (conversion en types sérialisables)
        request.session['statistiques_vente'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'statistiques_articles': statistiques_articles,  # Déjà converties en float
            'chiffre_affaires_total': float(chiffre_affaires_total),
            'marge_totale': float(marge_totale),
            'quantite_totale_vendue': int(quantite_totale_vendue),
            'pourcentage_marge_global': float(pourcentage_marge_global),
        }
        
        return JsonResponse({
            'success': True,
            'total_articles': len(statistiques_articles),
            'quantite_totale_vendue': int(quantite_totale_vendue),
            'chiffre_affaires_total': float(chiffre_affaires_total),
            'marge_totale': float(marge_totale),
            'pourcentage_marge_global': float(pourcentage_marge_global),
            'statistiques_articles': statistiques_articles  # Ajouter les données détaillées
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR GÉNÉRATION STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        error_details = f"Erreur: {str(e)}"
        if hasattr(e, '__class__'):
            error_details += f" (Type: {e.__class__.__name__})"
        return JsonResponse({'success': False, 'error': error_details})
@login_required
def export_statistiques_excel(request):
    """Vue pour exporter les statistiques de vente en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible, utilisation du format CSV")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Statistiques Vente {date_debut} - {date_fin}"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tête du document
        ws.merge_cells('A1:F1')
        ws['A1'] = f"STATISTIQUES DE VENTE - {agence.nom_agence}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Période: du {date_debut} au {date_fin}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes des colonnes
        headers = ['Référence', 'Désignation', 'Quantité Vendue', 'Chiffre d\'Affaires (FCFA)', 'Marge Profit (FCFA)', 'Pourcentage Marge (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données
        row = 5
        for stat in statistiques_articles:
            ws.cell(row=row, column=1, value=stat['reference_article'])
            ws.cell(row=row, column=2, value=stat['designation'])
            ws.cell(row=row, column=3, value=stat['quantite_vendue'])
            ws.cell(row=row, column=4, value=float(stat['chiffre_affaires']))
            ws.cell(row=row, column=5, value=float(stat['marge_profit']))
            ws.cell(row=row, column=6, value=float(stat['pourcentage_marge']))
            row += 1
        
        # Ligne des totaux
        row += 1
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value="").font = Font(bold=True)
        ws.cell(row=row, column=3, value=quantite_totale_vendue).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(chiffre_affaires_total)).font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(marge_totale)).font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(pourcentage_marge_global)).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 18
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le fichier dans la réponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"[CHART] EXPORT EXCEL STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_statistiques_pdf(request):
    """Vue pour exporter les statistiques de vente en PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            textColor=colors.darkblue
        )
        
        # Éléments du document
        elements = []
        
        # Titre
        title = Paragraph(f"STATISTIQUES DE VENTE - {agence.nom_agence}", title_style)
        elements.append(title)
        
        # Informations de la période
        period_text = f"<b>Période:</b> du {date_debut} au {date_fin}"
        period_para = Paragraph(period_text, styles['Normal'])
        elements.append(period_para)
        elements.append(Spacer(1, 20))
        
        # Données du tableau
        data = [['Référence', 'Désignation', 'Qté Vendue', 'Chiffre d\'Affaires', 'Marge Profit', 'Marge %']]
        
        for stat in statistiques_articles:
            data.append([
                stat['reference_article'],
                stat['designation'][:30] + '...' if len(stat['designation']) > 30 else stat['designation'],
                str(stat['quantite_vendue']),
                f"{float(stat['chiffre_affaires']):,.0f} FCFA",
                f"{float(stat['marge_profit']):,.0f} FCFA",
                f"{float(stat['pourcentage_marge']):.1f}%"
            ])
        
        # Ligne des totaux
        data.append([
            'TOTAL GÉNÉRAL',
            '',
            str(quantite_totale_vendue),
            f"{float(chiffre_affaires_total):,.0f} FCFA",
            f"{float(marge_totale):,.0f} FCFA",
            f"{float(pourcentage_marge_global):.1f}%"
        ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 6*cm, 2*cm, 3*cm, 3*cm, 2*cm])
        
        # Style du tableau
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"📄 EXPORT PDF STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"📄 EXPORT PDF STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_statistiques_csv(request):
    """Vue pour exporter les statistiques de vente en CSV"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les statistiques depuis la session
        statistiques_data = request.session.get('statistiques_vente')
        
        if not statistiques_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        # Récupérer les données
        date_debut = statistiques_data['date_debut']
        date_fin = statistiques_data['date_fin']
        statistiques_articles = statistiques_data['statistiques_articles']
        chiffre_affaires_total = statistiques_data['chiffre_affaires_total']
        marge_totale = statistiques_data['marge_totale']
        quantite_totale_vendue = statistiques_data['quantite_totale_vendue']
        pourcentage_marge_global = statistiques_data['pourcentage_marge_global']
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Statistiques_Vente_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Référence', 'Désignation', 'Quantité Vendue', 'Chiffre d\'Affaires (FCFA)', 'Marge Profit (FCFA)', 'Pourcentage Marge (%)']
        writer.writerow(headers)
        
        # Données
        for stat in statistiques_articles:
            row = [
                stat['reference_article'],
                stat['designation'],
                stat['quantite_vendue'],
                float(stat['chiffre_affaires']),
                float(stat['marge_profit']),
                float(stat['pourcentage_marge'])
            ]
            writer.writerow(row)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['TOTAL GÉNÉRAL', '', quantite_totale_vendue, chiffre_affaires_total, marge_totale, pourcentage_marge_global])
        
        print(f"[CHART] EXPORT CSV STATISTIQUES - Période: {date_debut} à {date_fin}")
        print(f"[CHART] EXPORT CSV STATISTIQUES - {len(statistiques_articles)} articles")
        print(f"[CHART] EXPORT CSV STATISTIQUES - Total: {quantite_totale_vendue} vendus, {chiffre_affaires_total} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV STATISTIQUES: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_statistiques(request):
    """Vue de test pour diagnostiquer les problèmes"""
    try:
        print("🧪 TEST STATISTIQUES - Début")
        
        # Test 1: Récupération de l'agence
        agence = get_user_agence(request)
        print(f"🧪 TEST - Agence récupérée: {agence}")
        
        # Test 2: Vérification des imports
        print("🧪 TEST - Import Decimal OK")
        
        # Test 3: Vérification des modèles
        articles_count = Article.objects.filter(agence=agence).count()
        print(f"🧪 TEST - Articles trouvés: {articles_count}")
        
        # Test 4: Vérification des ventes
        ventes_count = LigneFactureVente.objects.filter(facture_vente__agence=agence).count()
        print(f"🧪 TEST - Ventes trouvées: {ventes_count}")
        
        return JsonResponse({
            'success': True,
            'message': 'Tests réussis',
            'agence': str(agence),
            'articles_count': articles_count,
            'ventes_count': ventes_count
        })
        
    except Exception as e:
        print(f"🧪 TEST - Erreur: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ==================== MOUVEMENTS DE STOCK ====================

@login_required
@require_stock_access
def mouvements_stock(request):
    """Vue pour la page des mouvements de stock"""
    try:
        agence = get_user_agence(request)
    except:
        messages.error(request, 'Agence non trouvée.')
        return redirect('login_stock')
    
    # Récupérer les articles de l'agence
    articles = Article.objects.filter(agence=agence).order_by('designation')
    
    # Récupérer les familles pour le filtre
    familles = Famille.objects.all()
    
    # Statistiques générales
    total_articles = articles.count()
    
    # Calculer les statistiques des mouvements des 30 derniers jours
    date_debut = timezone.now().date() - timezone.timedelta(days=30)
    
    # Récupérer les mouvements des 30 derniers jours
    mouvements_recentes = MouvementStock.objects.filter(
        agence=agence,
        date_mouvement__gte=date_debut
    ).select_related('article', 'fournisseur', 'facture_vente', 'facture_achat')
    
    # Statistiques des mouvements
    total_mouvements = mouvements_recentes.count()
    mouvements_entree = mouvements_recentes.filter(type_mouvement='entree').count()
    mouvements_sortie = mouvements_recentes.filter(type_mouvement='sortie').count()
    
    # Valeur totale du stock permanent
    valeur_stock_permanent = mouvements_recentes.aggregate(
        total=Sum('stock_permanent')
    )['total'] or 0
    
    context = {
        'agence': agence,
        'articles': articles,
        'familles': familles,
        'total_articles': total_articles,
        'total_mouvements': total_mouvements,
        'mouvements_entree': mouvements_entree,
        'mouvements_sortie': mouvements_sortie,
        'valeur_stock_permanent': valeur_stock_permanent,
    }
    
    return render(request, 'supermarket/stock/mouvements_stock.html', context)

@login_required
@require_stock_access
def consulter_mouvements_stock(request):
    """Vue pour consulter les mouvements de stock selon les critères sélectionnés"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        article_ids = request.POST.getlist('articles')  # Récupérer tous les articles sélectionnés
        type_mouvement = request.POST.get('type_mouvement', '')
        
        print(f"[CHART] PARAMÈTRES MOUVEMENTS:")
        print(f"  - Date début: {date_debut}")
        print(f"  - Date fin: {date_fin}")
        print(f"  - Articles: {article_ids}")
        print(f"  - Type mouvement: {type_mouvement}")
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        if not article_ids:
            return JsonResponse({'success': False, 'error': 'Veuillez sélectionner au moins un article'})
        
        # Convertir les dates
        try:
            from datetime import datetime
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Format de date invalide: {str(e)}'})
        
        # Récupérer les articles
        try:
            articles = Article.objects.filter(id__in=article_ids, agence=agence)
            if not articles.exists():
                return JsonResponse({'success': False, 'error': 'Aucun article valide trouvé'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Erreur lors de la récupération des articles: {str(e)}'})
        
        # Filtrer les mouvements selon les critères
        # Utiliser __date pour comparer seulement les dates (ignorer l'heure)
        mouvements_query = MouvementStock.objects.filter(
            agence=agence,
            article__in=articles,  # Filtrer par plusieurs articles
            date_mouvement__date__gte=date_debut_obj,
            date_mouvement__date__lte=date_fin_obj
        ).select_related('article', 'fournisseur', 'facture_vente', 'facture_achat').order_by('article__reference_article', 'date_mouvement')
        
        # Filtre par type de mouvement
        if type_mouvement and type_mouvement != '':
            mouvements_query = mouvements_query.filter(type_mouvement=type_mouvement)
        
        mouvements = mouvements_query
        
        print(f"[PACKAGE] MOUVEMENTS FILTRÉS: {mouvements.count()}")
        
        # Debug: Afficher quelques mouvements pour vérifier
        if mouvements.count() > 0:
            print("[LIST] PREMIERS MOUVEMENTS TROUVÉS:")
            for i, mvt in enumerate(mouvements[:3]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement}")
        else:
            print("[ERREUR] AUCUN MOUVEMENT TROUVÉ - Vérifions les mouvements existants:")
            tous_mouvements = MouvementStock.objects.filter(agence=agence, article__in=articles)
            print(f"[CHART] Total mouvements pour ces articles: {tous_mouvements.count()}")
            for i, mvt in enumerate(tous_mouvements[:3]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement}")
        
        # Calculer les statistiques
        total_mouvements = mouvements.count()
        mouvements_entree = mouvements.filter(type_mouvement='entree').count()
        mouvements_sortie = mouvements.filter(type_mouvement='sortie').count()
        
        # Calculer la valeur totale du stock permanent
        valeur_stock_permanent = mouvements.aggregate(
            total=Sum('stock_permanent')
        )['total'] or 0
        
        # Stocker les données dans la session pour l'export
        mouvements_data = []
        for mouvement in mouvements:
            # Déterminer le tiers
            tiers = ""
            if mouvement.fournisseur:
                tiers = f"Fournisseur: {mouvement.fournisseur.intitule}"
            elif mouvement.facture_vente:
                tiers = f"Client: {mouvement.facture_vente.client.intitule if mouvement.facture_vente.client else 'N/A'}"
            elif mouvement.facture_achat:
                tiers = f"Fournisseur: {mouvement.facture_achat.fournisseur.intitule if mouvement.facture_achat.fournisseur else 'N/A'}"

            mouvements_data.append({
                'date_mouvement': mouvement.date_mouvement.strftime('%Y-%m-%d %H:%M'),
                'type_mouvement': mouvement.type_mouvement,
                'type_mouvement_display': mouvement.get_type_mouvement_display(),
                'reference_article': str(mouvement.article.reference_article) if getattr(mouvement.article, 'reference_article', None) else 'N/A',
                'designation': str(mouvement.article.designation) if getattr(mouvement.article, 'designation', None) else 'N/A',
                'tiers': tiers,
                'stock_initial': float(mouvement.stock_initial) if mouvement.stock_initial is not None else 0.0,
                'quantite': float(mouvement.quantite) if mouvement.quantite is not None else 0.0,
                'solde': float(mouvement.solde) if mouvement.solde is not None else 0.0,
                'quantite_stock': float(mouvement.quantite_stock) if mouvement.quantite_stock is not None else 0.0,
                'cout_moyen_pondere': float(mouvement.cout_moyen_pondere) if mouvement.cout_moyen_pondere is not None else 0.0,
                'stock_permanent': float(mouvement.stock_permanent) if mouvement.stock_permanent is not None else 0.0,
                'numero_piece': mouvement.numero_piece,
                'commentaire': mouvement.commentaire or '',
            })
        
        print(f"[CHART] MOUVEMENTS GÉNÉRÉS:")
        print(f"  - Total mouvements: {total_mouvements}")
        print(f"  - Entrées: {mouvements_entree}")
        print(f"  - Sorties: {mouvements_sortie}")
        print(f"  - Valeur stock permanent: {valeur_stock_permanent}")
        
        # Stocker les mouvements dans la session pour l'export (sans Decimal)
        session_payload = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'articles_count': len(articles),
            'mouvements_data': mouvements_data,
            'total_mouvements': total_mouvements,
            'mouvements_entree': mouvements_entree,
            'mouvements_sortie': mouvements_sortie,
            'valeur_stock_permanent': float(valeur_stock_permanent) if valeur_stock_permanent is not None else 0.0,
            'article_reference': ', '.join([str(article.reference_article or 'N/A') for article in articles]),
            'article_designation': ', '.join([str(article.designation or 'N/A') for article in articles]),
        }

        from django.core.serializers.json import DjangoJSONEncoder
        import json

        clean_payload = json.loads(json.dumps(session_payload, cls=DjangoJSONEncoder))

        try:
            json.dumps(clean_payload)
        except TypeError as json_error:
            print(f"[ERROR] JSON serialization issue dans mouvements_stock: {json_error}")
            for idx, mouvement in enumerate(clean_payload.get('mouvements_data', [])):
                for key, value in mouvement.items():
                    print(f"    mouvement[{idx}].{key} -> {type(value)} = {value}")
            raise

        request.session['mouvements_stock'] = clean_payload
        
        # Préparer les informations des articles
        articles_info = []
        for article in articles:
            article_mouvements = mouvements.filter(article=article)
            articles_info.append({
                'id': article.id,
                'reference': article.reference_article,
                'designation': article.designation,
                'stock_actuel': article.stock_actuel,
                'mouvements_count': article_mouvements.count()
            })
        
        return JsonResponse({
            'success': True,
            'total_mouvements': total_mouvements,
            'mouvements_entree': mouvements_entree,
            'mouvements_sortie': mouvements_sortie,
            'valeur_stock_permanent': float(valeur_stock_permanent) if valeur_stock_permanent is not None else 0.0,
            'articles_info': articles_info,
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'mouvements': mouvements_data,
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR CONSULTATION MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        error_details = f"Erreur: {str(e)}"
        if hasattr(e, '__class__'):
            error_details += f" (Type: {e.__class__.__name__})"
        return JsonResponse({'success': False, 'error': error_details})

@login_required
def export_mouvements_excel(request):
    """Vue pour exporter les mouvements de stock en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le fichier Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            print("[ERREUR] Module openpyxl non disponible")
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé. Veuillez installer openpyxl pour l\'export Excel.'})
        
        from django.http import HttpResponse
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Mouvements Stock {article_reference}"
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tête du document (17 colonnes maintenant avec Référence/Désignation)
        ws.merge_cells('A1:Q1')
        ws['A1'] = f"FICHE DE STOCK - {article_reference} - {article_designation}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:Q2')
        ws['A2'] = f"Période: du {date_debut} au {date_fin}"
        ws['A2'].font = Font(size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        # En-têtes des colonnes - Ajout d'une colonne Référence/Désignation
        headers = ['Référence/Désignation', 'Date', 'Type', '', 'N°', '', '', 'Tiers', '', '', '', '', '+/-', 'Quantités en stock', 'Solde', 'C.M.U.P.', 'Stock permanent']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Grouper les mouvements par article (par référence_article)
        from collections import defaultdict
        mouvements_par_article = defaultdict(list)
        for mouvement in mouvements_data_list:
            # Utiliser référence_article comme clé de regroupement
            ref_article = mouvement.get('reference_article', 'N/A')
            mouvements_par_article[ref_article].append(mouvement)
        
        # Trier les articles par référence pour un affichage ordonné
        articles_tries = sorted(mouvements_par_article.keys())
        
        # Style pour les en-têtes d'article
        article_header_font = Font(bold=True, size=12, color="FFFFFF")
        article_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        article_header_alignment = Alignment(horizontal="left", vertical="center")
        
        # Données groupées par article
        row = 5
        total_stock_permanent_global = 0
        
        for ref_article in articles_tries:
            mouvements_article = mouvements_par_article[ref_article]
            
            # Récupérer la désignation du premier mouvement (tous les mouvements d'un article ont la même désignation)
            designation_article = mouvements_article[0].get('designation', 'N/A') if mouvements_article else 'N/A'
            
            # En-tête de l'article - Ligne de séparation avec référence et désignation
            ws.merge_cells(f'A{row}:Q{row}')
            cell_article = ws.cell(row=row, column=1, value=f"ARTICLE: {ref_article} - {designation_article}")
            cell_article.font = article_header_font
            cell_article.fill = article_header_fill
            cell_article.alignment = article_header_alignment
            row += 1
            
            # Calculer les totaux pour cet article
            total_entrees_article = sum(1 for m in mouvements_article if m['type_mouvement'] == 'entree')
            total_sorties_article = sum(1 for m in mouvements_article if m['type_mouvement'] == 'sortie')
            total_stock_permanent_article = sum(float(m.get('stock_permanent', 0)) for m in mouvements_article)
            total_stock_permanent_global += total_stock_permanent_article
            
            # Afficher tous les mouvements de cet article
            for mouvement in mouvements_article:
                # Déterminer le type de mouvement avec "Correction" si c'est une modification
                type_mouvement_display = mouvement.get('type_mouvement_display', mouvement['type_mouvement'])
                commentaire = (mouvement.get('commentaire', '') or '').lower()
                if 'modification' in commentaire or 'modifier' in commentaire or '(modification' in commentaire:
                    type_mouvement_display = 'Correction'
                
                # Structure avec colonne Référence/Désignation (17 colonnes maintenant)
                ws.cell(row=row, column=1, value=f"{ref_article} - {designation_article}")  # Référence/Désignation
                ws.cell(row=row, column=2, value=mouvement['date_mouvement'])  # Date
                ws.cell(row=row, column=3, value=type_mouvement_display)  # Type (avec Correction si modification)
                ws.cell(row=row, column=4, value='')  # Colonne vide
                ws.cell(row=row, column=5, value=mouvement['numero_piece'])  # N°
                ws.cell(row=row, column=6, value='')  # Colonne vide
                ws.cell(row=row, column=7, value='')  # Colonne vide
                ws.cell(row=row, column=8, value=mouvement['tiers'])  # Tiers
                ws.cell(row=row, column=9, value='')  # Colonne vide
                ws.cell(row=row, column=10, value='')  # Colonne vide
                ws.cell(row=row, column=11, value='')  # Colonne vide
                ws.cell(row=row, column=12, value='')  # Colonne vide
                # Calculer le signe +/- selon le type de mouvement
                quantite = float(mouvement.get('quantite', 0))
                if mouvement['type_mouvement'] in ['sortie', 'perte']:
                    quantite_display = f"-{abs(quantite)}"
                elif mouvement['type_mouvement'] in ['entree', 'retour']:
                    quantite_display = f"+{quantite}" if quantite > 0 else str(quantite)
                elif mouvement['type_mouvement'] == 'ajustement':
                    # Pour ajustement, déterminer selon solde vs stock_initial
                    stock_initial = float(mouvement.get('stock_initial', 0))
                    solde = float(mouvement.get('solde', 0))
                    if solde < stock_initial:
                        quantite_display = f"-{abs(quantite)}"
                    elif solde > stock_initial:
                        quantite_display = f"+{quantite}"
                    else:
                        quantite_display = str(quantite)
                else:
                    quantite_display = f"+{quantite}" if quantite > 0 else str(quantite)
                ws.cell(row=row, column=13, value=quantite_display)  # +/-
                ws.cell(row=row, column=14, value=mouvement['stock_initial'])  # Quantités en stock
                ws.cell(row=row, column=15, value=mouvement['solde'])  # Solde
                ws.cell(row=row, column=16, value=mouvement['cout_moyen_pondere'])  # C.M.U.P.
                ws.cell(row=row, column=17, value=mouvement['stock_permanent'])  # Stock permanent
                row += 1
            
            # Ligne de sous-total pour cet article
            ws.merge_cells(f'A{row}:B{row}')
            ws.cell(row=row, column=1, value=f"SOUS-TOTAL {ref_article}").font = Font(bold=True)
            ws.cell(row=row, column=3, value=f"Entrées: {total_entrees_article}, Sorties: {total_sorties_article}").font = Font(bold=True)
            ws.cell(row=row, column=4, value=f"Total: {len(mouvements_article)} mouvements").font = Font(bold=True)
            ws.cell(row=row, column=17, value=total_stock_permanent_article).font = Font(bold=True)
            # Remplir les colonnes vides
            for col in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
                ws.cell(row=row, column=col, value='').font = Font(bold=True)
            row += 1
            
            # Ligne vide pour séparer les articles
            row += 1
        
        # Ligne des totaux généraux
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = Font(bold=True, size=14)
        ws.cell(row=row, column=3, value=f"Entrées: {mouvements_entree}, Sorties: {mouvements_sortie}").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"Total: {total_mouvements} mouvements").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"Articles: {len(articles_tries)}").font = Font(bold=True)
        # Remplir les colonnes vides
        for col in [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
            ws.cell(row=row, column=col, value="").font = Font(bold=True)
        ws.cell(row=row, column=17, value=total_stock_permanent_global).font = Font(bold=True, size=12)
        
        # Ajuster la largeur des colonnes (17 colonnes maintenant avec Référence/Désignation)
        ws.column_dimensions['A'].width = 35  # Référence/Désignation (plus large)
        ws.column_dimensions['B'].width = 20  # Date
        ws.column_dimensions['C'].width = 15  # Type
        ws.column_dimensions['D'].width = 5   # Colonne vide
        ws.column_dimensions['E'].width = 15  # N°
        ws.column_dimensions['F'].width = 5   # Colonne vide
        ws.column_dimensions['G'].width = 5  # Colonne vide
        ws.column_dimensions['H'].width = 25  # Tiers
        ws.column_dimensions['I'].width = 5   # Colonne vide
        ws.column_dimensions['J'].width = 5   # Colonne vide
        ws.column_dimensions['K'].width = 5   # Colonne vide
        ws.column_dimensions['L'].width = 5  # Colonne vide
        ws.column_dimensions['M'].width = 12  # +/-
        ws.column_dimensions['N'].width = 18  # Quantités en stock
        ws.column_dimensions['O'].width = 12  # Solde
        ws.column_dimensions['P'].width = 15  # C.M.U.P.
        ws.column_dimensions['Q'].width = 18  # Stock permanent
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le fichier dans la réponse
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Article: {article_reference}")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - {total_mouvements} mouvements")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        print(f"[CHART] EXPORT EXCEL MOUVEMENTS - Articles groupés: {len(articles_tries)}")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT EXCEL MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})
@login_required
def export_mouvements_pdf(request):
    """Vue pour exporter les mouvements de stock en PDF"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le PDF
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import cm
        except ImportError:
            print("[ERREUR] Module reportlab non disponible")
            return JsonResponse({'success': False, 'error': 'Module reportlab non installé. Veuillez installer reportlab pour l\'export PDF.'})
        
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,
            textColor=colors.darkblue
        )
        
        # Éléments du document
        elements = []
        
        # Titre
        title = Paragraph(f"FICHE DE STOCK - {article_reference}", title_style)
        elements.append(title)
        
        # Informations de l'article et période
        info_text = f"<b>Article:</b> {article_designation}<br/><b>Période:</b> du {date_debut} au {date_fin}"
        info_para = Paragraph(info_text, styles['Normal'])
        elements.append(info_para)
        elements.append(Spacer(1, 20))
        
        # Données du tableau
        data = [['Date', 'Type', 'Tiers', 'Stock Init.', 'Solde', 'C.M.PU', 'Stock Perm.', 'N° Pièce']]
        
        for mouvement in mouvements_data_list:
            data.append([
                mouvement['date_mouvement'],
                mouvement['type_mouvement'],
                mouvement['tiers'][:20] + '...' if len(mouvement['tiers']) > 20 else mouvement['tiers'],
                str(mouvement['stock_initial']),
                str(mouvement['solde']),
                f"{mouvement['cout_moyen_pondere']:,.0f}",
                f"{mouvement['stock_permanent']:,.0f}",
                mouvement['numero_piece']
            ])
        
        # Ligne des totaux
        data.append([
            'TOTAL',
            f'E:{mouvements_entree} S:{mouvements_sortie}',
            f'{total_mouvements} mouvements',
            '',
            '',
            '',
            f"{valeur_stock_permanent:,.0f} FCFA",
            ''
        ])
        
        # Créer le tableau
        table = Table(data, colWidths=[2*cm, 1.5*cm, 3*cm, 1.5*cm, 1.5*cm, 2*cm, 2.5*cm, 2*cm])
        
        # Style du tableau
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(table)
        
        # Construire le PDF
        doc.build(elements)
        
        print(f"📄 EXPORT PDF MOUVEMENTS - Article: {article_reference}")
        print(f"📄 EXPORT PDF MOUVEMENTS - {total_mouvements} mouvements")
        print(f"📄 EXPORT PDF MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT PDF MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def export_mouvements_csv(request):
    """Vue pour exporter les mouvements de stock en CSV"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les mouvements depuis la session
        mouvements_data = request.session.get('mouvements_stock')
        
        if not mouvements_data:
            return JsonResponse({'success': False, 'error': 'Aucun mouvement consulté'})
        
        # Récupérer les données
        date_debut = mouvements_data['date_debut']
        date_fin = mouvements_data['date_fin']
        article_reference = mouvements_data.get('article_reference', 'Articles multiples')
        article_designation = mouvements_data.get('article_designation', 'Articles multiples')
        mouvements_data_list = mouvements_data['mouvements_data']
        total_mouvements = mouvements_data['total_mouvements']
        mouvements_entree = mouvements_data['mouvements_entree']
        mouvements_sortie = mouvements_data['mouvements_sortie']
        valeur_stock_permanent = mouvements_data['valeur_stock_permanent']
        
        # Créer le fichier CSV
        import csv
        from django.http import HttpResponse
        
        # Créer la réponse HTTP
        filename = f"Mouvements_Stock_{article_reference}_{date_debut}_{date_fin}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Ajouter BOM pour Excel (UTF-8)
        response.write('\ufeff')
        
        # Créer le writer CSV
        writer = csv.writer(response, delimiter=';')
        
        # En-têtes
        headers = ['Date', 'Type', '', 'N°', '', '', 'Tiers', '', '', '', '', '+/-', 'Quantités en stock', 'Solde', 'C.M.U.P.', 'Stock permanent']
        writer.writerow(headers)
        
        # Données
        for mouvement in mouvements_data_list:
            row = [
                mouvement['date_mouvement'],
                mouvement['type_mouvement'],
                mouvement['tiers'],
                mouvement['stock_initial'],
                mouvement['solde'],
                mouvement['cout_moyen_pondere'],
                mouvement['stock_permanent'],
                mouvement['numero_piece']
            ]
            writer.writerow(row)
        
        # Ligne des totaux
        writer.writerow([])  # Ligne vide
        writer.writerow(['TOTAL GÉNÉRAL', f'Entrées: {mouvements_entree}, Sorties: {mouvements_sortie}', f'{total_mouvements} mouvements', '', '', '', valeur_stock_permanent, ''])
        
        print(f"[CHART] EXPORT CSV MOUVEMENTS - Article: {article_reference}")
        print(f"[CHART] EXPORT CSV MOUVEMENTS - {total_mouvements} mouvements")
        print(f"[CHART] EXPORT CSV MOUVEMENTS - Valeur stock permanent: {valeur_stock_permanent} FCFA")
        
        return response
        
    except Exception as e:
        print(f"[ERREUR] ERREUR EXPORT CSV MOUVEMENTS: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def creer_mouvements_retroactifs(request):
    """Vue simplifiée pour créer des mouvements de stock rétroactifs"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[REFRESH] CRÉATION MOUVEMENTS RÉTROACTIFS (VERSION SIMPLIFIÉE)...")
        print(f"[TARGET] Agence utilisée: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # 1. Créer des mouvements pour les factures de vente
        factures_vente = FactureVente.objects.filter(agence=agence)
        print(f"[CHART] Factures de vente: {factures_vente.count()}")
        
        for facture in factures_vente:
            lignes = LigneFactureVente.objects.filter(facture_vente=facture)
            print(f"[SEARCH] Facture {facture.numero_ticket}: {lignes.count()} lignes")
            
            for ligne in lignes:
                print(f"  [PACKAGE] Article: {ligne.article.designation}, Quantité: {ligne.quantite}")
                
                # Vérifier si le mouvement existe déjà
                mouvement_existe = MouvementStock.objects.filter(facture_vente=facture, article=ligne.article).exists()
                print(f"  [SEARCH] Mouvement existe déjà: {mouvement_existe}")
                
                if not mouvement_existe:
                    try:
                        from datetime import datetime
                        from django.utils import timezone
                        
                        # Utiliser timezone.now() pour la date
                        date_mouvement = timezone.now()
                        
                        MouvementStock.objects.create(
                            article=ligne.article,
                            agence=agence,
                            type_mouvement='sortie',
                            date_mouvement=date_mouvement,
                            numero_piece=facture.numero_ticket,
                            quantite_stock=ligne.article.stock_actuel,
                            stock_initial=ligne.article.stock_actuel + ligne.quantite,
                            solde=ligne.article.stock_actuel,
                            quantite=ligne.quantite,
                            cout_moyen_pondere=float(ligne.prix_unitaire),
                            stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                            facture_vente=facture,
                            commentaire=f"Vente - {facture.numero_ticket}"
                        )
                        mouvements_crees += 1
                        print(f"[OK] Vente: {ligne.article.designation} - {facture.numero_ticket}")
                    except Exception as e:
                        print(f"[ERREUR] Erreur vente {facture.numero_ticket}: {e}")
        
        # 2. Créer des mouvements pour les factures d'achat
        factures_achat = FactureAchat.objects.filter(agence=agence)
        print(f"[CHART] Factures d'achat: {factures_achat.count()}")
        
        for facture in factures_achat:
            lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
            print(f"[SEARCH] Facture achat {facture.reference_achat}: {lignes.count()} lignes")
            
            for ligne in lignes:
                print(f"  [PACKAGE] Article: {ligne.article.designation}, Quantité: {ligne.quantite}")
                
                mouvement_existe = MouvementStock.objects.filter(facture_achat=facture, article=ligne.article).exists()
                print(f"  [SEARCH] Mouvement existe déjà: {mouvement_existe}")
                
                if not mouvement_existe:
                    try:
                        from django.utils import timezone
                        
                        MouvementStock.objects.create(
                            article=ligne.article,
                            agence=agence,
                            type_mouvement='entree',
                            date_mouvement=timezone.now(),
                            numero_piece=facture.reference_achat,
                            quantite_stock=ligne.article.stock_actuel,
                            stock_initial=ligne.article.stock_actuel - ligne.quantite,
                            solde=ligne.article.stock_actuel,
                            quantite=ligne.quantite,
                            cout_moyen_pondere=float(ligne.prix_unitaire),
                            stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                            facture_achat=facture,
                            fournisseur=facture.fournisseur,
                            commentaire=f"Achat - {facture.reference_achat}"
                        )
                        mouvements_crees += 1
                        print(f"[OK] Achat: {ligne.article.designation} - {facture.reference_achat}")
                    except Exception as e:
                        print(f"[ERREUR] Erreur achat {facture.reference_achat}: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'{mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def diagnostic_mouvements(request):
    """Vue de diagnostic pour les mouvements de stock"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Compter les données
        articles_count = Article.objects.filter(agence=agence).count()
        factures_vente_count = FactureVente.objects.filter(agence=agence).count()
        factures_achat_count = FactureAchat.objects.filter(agence=agence).count()
        factures_transfert_count = FactureTransfert.objects.filter(agence_source=agence).count()
        mouvements_count = MouvementStock.objects.filter(agence=agence).count()
        
        # Détails des factures
        factures_vente_details = []
        for facture in FactureVente.objects.filter(agence=agence)[:5]:
            lignes_count = LigneFactureVente.objects.filter(facture_vente=facture).count()
            factures_vente_details.append({
                'numero': facture.numero_ticket,
                'date': str(facture.date),
                'lignes': lignes_count
            })
        
        factures_achat_details = []
        for facture in FactureAchat.objects.filter(agence=agence)[:5]:
            lignes_count = LigneFactureAchat.objects.filter(facture_achat=facture).count()
            factures_achat_details.append({
                'numero': facture.reference_achat,
                'date': str(facture.date_achat),
                'lignes': lignes_count
            })
        
        return JsonResponse({
            'success': True,
            'agence': agence.nom_agence,
            'articles_count': articles_count,
            'factures_vente_count': factures_vente_count,
            'factures_achat_count': factures_achat_count,
            'factures_transfert_count': factures_transfert_count,
            'mouvements_count': mouvements_count,
            'factures_vente_details': factures_vente_details,
            'factures_achat_details': factures_achat_details
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def forcer_mouvements(request):
    """Vue pour forcer la création de mouvements même s'ils existent déjà"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[HOT] CRÉATION FORCÉE DE MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # Supprimer tous les mouvements existants d'abord
        anciens_mouvements = MouvementStock.objects.filter(agence=agence).count()
        MouvementStock.objects.filter(agence=agence).delete()
        print(f"🗑️ {anciens_mouvements} anciens mouvements supprimés")
        
        # Test de création d'un mouvement simple
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles trouvés: {articles.count()}")
        
        if articles.exists():
            article_test = articles.first()
            print(f"🧪 Test avec article: {article_test.designation}")
            
            try:
                from django.utils import timezone
                
                mouvement_test = MouvementStock.objects.create(
                    article=article_test,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece='TEST-001',
                    quantite_stock=article_test.stock_actuel,
                    stock_initial=0,
                    solde=article_test.stock_actuel,
                    quantite=1,
                    cout_moyen_pondere=float(article_test.prix_achat),
                    stock_permanent=float(article_test.stock_actuel * article_test.prix_achat),
                    commentaire='Test de création'
                )
                print(f"[OK] MOUVEMENT TEST CRÉÉ AVEC SUCCÈS: ID {mouvement_test.id}")
                mouvements_crees += 1
                
                # Supprimer le test
                mouvement_test.delete()
                print(f"🗑️ Mouvement test supprimé")
                
            except Exception as e:
                print(f"[ERREUR] ERREUR LORS DU TEST: {e}")
                import traceback
                traceback.print_exc()
                return JsonResponse({'success': False, 'error': f'Erreur lors du test de création: {str(e)}'})
        
        print(f"[OK] Test terminé, création des vrais mouvements...")
        
        # Créer des mouvements pour les factures de vente (version simplifiée)
        factures_vente = FactureVente.objects.filter(agence=agence)
        print(f"[CHART] Factures de vente: {factures_vente.count()}")
        
        for facture in factures_vente:
            lignes = LigneFactureVente.objects.filter(facture_vente=facture)
            print(f"  [SEARCH] {facture.numero_ticket}: {lignes.count()} lignes")
            
            for ligne in lignes:
                try:
                    from django.utils import timezone
                    
                    # Utiliser la même logique simple que le test qui fonctionne
                    MouvementStock.objects.create(
                        article=ligne.article,
                        agence=agence,
                        type_mouvement='sortie',
                        date_mouvement=timezone.now(),
                        numero_piece=facture.numero_ticket,
                        quantite_stock=ligne.article.stock_actuel,
                        stock_initial=0,  # Simplifié
                        solde=ligne.article.stock_actuel,
                        quantite=ligne.quantite,
                        cout_moyen_pondere=float(ligne.prix_unitaire),
                        stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                        facture_vente=facture,
                        commentaire=f"Vente - {facture.numero_ticket}"
                    )
                    mouvements_crees += 1
                    print(f"    [OK] {ligne.article.designation}")
                except Exception as e:
                    print(f"    [ERREUR] Erreur: {e}")
        
        # Créer des mouvements pour les factures d'achat (version simplifiée)
        factures_achat = FactureAchat.objects.filter(agence=agence)
        print(f"[CHART] Factures d'achat: {factures_achat.count()}")
        
        for facture in factures_achat:
            lignes = LigneFactureAchat.objects.filter(facture_achat=facture)
            print(f"  [SEARCH] {facture.reference_achat}: {lignes.count()} lignes")
            
            for ligne in lignes:
                try:
                    from django.utils import timezone
                    
                    # Utiliser la même logique simple que le test qui fonctionne
                    MouvementStock.objects.create(
                        article=ligne.article,
                        agence=agence,
                        type_mouvement='entree',
                        date_mouvement=timezone.now(),
                        numero_piece=facture.reference_achat,
                        quantite_stock=ligne.article.stock_actuel,
                        stock_initial=0,  # Simplifié
                        solde=ligne.article.stock_actuel,
                        quantite=ligne.quantite,
                        cout_moyen_pondere=float(ligne.prix_unitaire),
                        stock_permanent=float(ligne.article.stock_actuel * ligne.prix_unitaire),
                        facture_achat=facture,
                        fournisseur=facture.fournisseur,
                        commentaire=f"Achat - {facture.reference_achat}"
                    )
                    mouvements_crees += 1
                    print(f"    [OK] {ligne.article.designation}")
                except Exception as e:
                    print(f"    [ERREUR] Erreur: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'FORCÉ: {mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_mouvement_simple(request):
    """Test simple de création d'un mouvement"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("🧪 TEST SIMPLE DE CRÉATION DE MOUVEMENT...")
        
        # Vérifier les articles
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles: {articles.count()}")
        
        if not articles.exists():
            return JsonResponse({'success': False, 'error': 'Aucun article trouvé dans cette agence'})
        
        article = articles.first()
        print(f"[PACKAGE] Article test: {article.designation}")
        
        # Vérifier les champs obligatoires
        print(f"[PACKAGE] Stock actuel: {article.stock_actuel}")
        print(f"[PACKAGE] Prix achat: {article.prix_achat}")
        
        # Créer un mouvement simple
        from django.utils import timezone
        
        mouvement = MouvementStock.objects.create(
            article=article,
            agence=agence,
            type_mouvement='entree',
            date_mouvement=timezone.now(),
            numero_piece='TEST-SIMPLE',
            quantite_stock=article.stock_actuel,
            stock_initial=0,
            solde=article.stock_actuel,
            quantite=1,
            cout_moyen_pondere=float(article.prix_achat),
            stock_permanent=float(article.stock_actuel * article.prix_achat),
            commentaire='Test simple'
        )
        
        print(f"[OK] MOUVEMENT CRÉÉ: ID {mouvement.id}")
        
        # Vérifier qu'il existe
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        print(f"[CHART] Total mouvements: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'Test réussi! Mouvement ID {mouvement.id} créé. Total: {total_mouvements}',
            'mouvement_id': mouvement.id,
            'total_mouvements': total_mouvements
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR TEST: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def creer_mouvements_manuels(request):
    """Créer des mouvements manuels simples pour tester"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[TOOL] CRÉATION MANUELLE DE MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        mouvements_crees = 0
        
        # Récupérer tous les articles
        articles = Article.objects.filter(agence=agence)
        print(f"[PACKAGE] Articles trouvés: {articles.count()}")
        
        if not articles.exists():
            return JsonResponse({'success': False, 'error': 'Aucun article trouvé'})
        
        # Créer un mouvement pour chaque article
        for article in articles:
            try:
                from django.utils import timezone
                
                MouvementStock.objects.create(
                    article=article,
                    agence=agence,
                    type_mouvement='entree',
                    date_mouvement=timezone.now(),
                    numero_piece=f'MANUEL-{article.id}',
                    quantite_stock=article.stock_actuel,
                    stock_initial=0,
                    solde=article.stock_actuel,
                    quantite=article.stock_actuel,
                    cout_moyen_pondere=float(article.prix_achat),
                    stock_permanent=float(article.stock_actuel * article.prix_achat),
                    commentaire=f'Création manuelle - {article.designation}'
                )
                mouvements_crees += 1
                print(f"[OK] {article.designation}")
                
            except Exception as e:
                print(f"[ERREUR] Erreur pour {article.designation}: {e}")
        
        total_mouvements = MouvementStock.objects.filter(agence=agence).count()
        
        print(f"[SUCCESS] TERMINÉ - {mouvements_crees} mouvements manuels créés, Total: {total_mouvements}")
        
        return JsonResponse({
            'success': True,
            'message': f'MANUEL: {mouvements_crees} mouvements créés! Total: {total_mouvements}',
            'total_mouvements': total_mouvements,
            'nouveaux_mouvements': mouvements_crees
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_consultation_mouvements(request):
    """Test simple pour vérifier les mouvements existants"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        print("[SEARCH] TEST CONSULTATION MOUVEMENTS...")
        print(f"[TARGET] Agence: {agence.nom_agence}")
        
        # Vérifier tous les mouvements de l'agence
        tous_mouvements = MouvementStock.objects.filter(agence=agence)
        print(f"[CHART] Total mouvements dans l'agence: {tous_mouvements.count()}")
        
        if tous_mouvements.count() > 0:
            print("[LIST] PREMIERS MOUVEMENTS:")
            for i, mvt in enumerate(tous_mouvements[:5]):
                print(f"  {i+1}. {mvt.date_mouvement} - {mvt.article.designation} - {mvt.type_mouvement} - {mvt.numero_piece}")
        
        # Vérifier les articles avec mouvements
        articles_avec_mouvements = Article.objects.filter(
            agence=agence,
            mouvementstock__isnull=False
        ).distinct()
        print(f"[PACKAGE] Articles avec mouvements: {articles_avec_mouvements.count()}")
        
        for article in articles_avec_mouvements[:3]:
            mouvements_article = MouvementStock.objects.filter(agence=agence, article=article)
            print(f"  - {article.designation}: {mouvements_article.count()} mouvements")
        
        return JsonResponse({
            'success': True,
            'message': f'Test terminé - {tous_mouvements.count()} mouvements trouvés',
            'total_mouvements': tous_mouvements.count(),
            'articles_avec_mouvements': articles_avec_mouvements.count()
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_mouvements_session(request):
    """Récupérer les données de mouvements depuis la session"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les données depuis la session
        mouvements_data = request.session.get('mouvements_data', [])
        article_info = request.session.get('article_info', {})
        
        print(f"[CHART] RÉCUPÉRATION SESSION:")
        print(f"  - Mouvements en session: {len(mouvements_data)}")
        print(f"  - Article info: {article_info}")
        
        return JsonResponse({
            'success': True,
            'mouvements': mouvements_data,
            'article_info': article_info
        })
        
    except Exception as e:
        print(f"[ERREUR] ERREUR: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)})

# ===== GESTION DES COMPTES UTILISATEURS =====

def login_comptes(request):
    """Page de connexion pour le module gestion des comptes"""
    if request.user.is_authenticated and request.method == 'GET':
        logout(request)
        request.session.flush()

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        login(request, user)
                        request.session['agence_id'] = compte.agence.id_agence
                        request.session['agence_nom'] = compte.agence.nom_agence
                        messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                        return redirect('gestion_comptes')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Aucun compte actif trouvé pour cet utilisateur.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')

    return render(request, 'supermarket/comptes/login.html')

@login_required
def logout_comptes(request):
    """Déconnexion pour le module comptes"""
    logout(request)
    return redirect('index')

@login_required
@require_comptes_access
def gestion_comptes(request):
    """Vue pour lister tous les comptes utilisateurs"""
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    comptes = Compte.objects.all().order_by('nom', 'prenom')
    
    context = {
        'comptes': comptes,
        'agence': agence,
    }
    return render(request, 'supermarket/comptes/gestion_comptes.html', context)

@login_required
@require_comptes_access
def creer_compte(request):
    """Vue pour créer un nouveau compte utilisateur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        numero_compte = request.POST.get('numero_compte')
        type_compte = request.POST.get('type_compte')
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        telephone = request.POST.get('telephone')
        email = request.POST.get('email')
        agence_id = request.POST.get('agence')
        
        try:
            # Récupérer l'agence sélectionnée
            if agence_id:
                agence_selectionnee = Agence.objects.get(id_agence=agence_id)
            else:
                agence_selectionnee = agence  # Utiliser l'agence de l'utilisateur connecté par défaut
            
            # Créer l'utilisateur Django
            user = User.objects.create_user(username=username, password=password, email=email)
            
            # Créer le compte
            compte = Compte.objects.create(
                user=user,
                numero_compte=numero_compte,
                type_compte=type_compte,
                nom=nom,
                prenom=prenom,
                telephone=telephone,
                email=email,
                agence=agence_selectionnee
            )
            
            messages.success(request, f'Compte créé avec succès pour {compte.nom_complet} - Agence: {agence_selectionnee.nom_agence}')
            return redirect('gestion_comptes')
        except Agence.DoesNotExist:
            messages.error(request, 'Agence sélectionnée non trouvée.')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du compte: {str(e)}')
    
    agences = Agence.objects.all().order_by('nom_agence')
    type_compte_choices = Compte._meta.get_field('type_compte').choices
    context = {
        'agences': agences,
        'agence': agence,
        'type_compte_choices': type_compte_choices,
    }
    return render(request, 'supermarket/comptes/creer_compte.html', context)

@login_required
@require_comptes_access
def detail_compte(request, compte_id):
    """Vue pour afficher les détails d'un compte"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
        return redirect('gestion_comptes')
    
    context = {
        'compte': compte,
        'agence': agence,
    }
    return render(request, 'supermarket/comptes/detail_compte.html', context)

@login_required
@require_comptes_access
def modifier_compte(request, compte_id):
    """Vue pour modifier un compte utilisateur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
        return redirect('gestion_comptes')
    
    if request.method == 'POST':
        compte.numero_compte = request.POST.get('numero_compte', compte.numero_compte)
        compte.type_compte = request.POST.get('type_compte', compte.type_compte)
        compte.nom = request.POST.get('nom', compte.nom)
        compte.prenom = request.POST.get('prenom', compte.prenom)
        compte.telephone = request.POST.get('telephone', compte.telephone)
        compte.email = request.POST.get('email', compte.email)
        compte.actif = request.POST.get('actif') == 'on'
        
        # Mettre à jour l'utilisateur Django si nécessaire
        if request.POST.get('username'):
            compte.user.username = request.POST.get('username')
            compte.user.save()
        
        if request.POST.get('password'):
            compte.user.set_password(request.POST.get('password'))
            compte.user.save()
        
        compte.save()
        messages.success(request, f'Compte modifié avec succès pour {compte.nom_complet}')
        return redirect('detail_compte', compte_id=compte.id)
    
    agences = Agence.objects.all()
    type_compte_choices = Compte._meta.get_field('type_compte').choices
    context = {
        'compte': compte,
        'agences': agences,
        'agence': agence,
        'type_compte_choices': type_compte_choices,
    }
    return render(request, 'supermarket/comptes/modifier_compte.html', context)

@login_required
@require_comptes_access
def supprimer_compte(request, compte_id):
    """Vue pour supprimer un compte utilisateur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
        nom_complet = compte.nom_complet
        compte.user.delete()  # Cela supprimera aussi le compte grâce à CASCADE
        messages.success(request, f'Compte {nom_complet} supprimé avec succès')
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la suppression: {str(e)}')
    
    return redirect('gestion_comptes')

@login_required
@require_comptes_access
def activer_desactiver_compte(request, compte_id):
    """Vue pour activer ou désactiver un compte"""
    if request.method != 'POST':
        messages.error(request, 'Méthode non autorisée.')
        return redirect('gestion_comptes')
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
        
        # Inverser le statut actif/inactif
        compte.actif = not compte.actif
        compte.save()
        
        statut = "activé" if compte.actif else "désactivé"
        messages.success(request, f'Compte {compte.nom_complet} {statut} avec succès.')
        
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
    except Exception as e:
        messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    return redirect('gestion_comptes')

@login_required
@require_comptes_access
def reinitialiser_mot_de_passe(request, compte_id):
    """Vue pour réinitialiser le mot de passe d'un compte"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_comptes')
    
    try:
        compte = Compte.objects.get(id=compte_id, agence=agence)
    except Compte.DoesNotExist:
        messages.error(request, 'Compte non trouvé.')
        return redirect('gestion_comptes')
    
    if request.method == 'POST':
        nouveau_mot_de_passe = request.POST.get('nouveau_mot_de_passe')
        confirmer_mot_de_passe = request.POST.get('confirmer_mot_de_passe')
        
        if nouveau_mot_de_passe and nouveau_mot_de_passe == confirmer_mot_de_passe:
            compte.user.set_password(nouveau_mot_de_passe)
            compte.user.save()
            messages.success(request, f'Mot de passe réinitialisé avec succès pour {compte.nom_complet}')
            return redirect('detail_compte', compte_id=compte.id)
        else:
            messages.error(request, 'Les mots de passe ne correspondent pas.')
    
    context = {
        'compte': compte,
        'agence': agence,
    }
    return render(request, 'supermarket/comptes/reinitialiser_mot_de_passe.html', context)
# ==================== RAPPORT LIVREUR ====================

@login_required
@require_commandes_feature('rapport_livreur')
def rapport_livreur(request):
    """Vue pour afficher le formulaire de sélection pour le rapport livreur"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('dashboard_commandes')
    
    compte = get_user_compte(request)
    if not compte:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return redirect('login_commandes')
    
    # Si c'est un livreur, ne montrer que lui-même
    if compte.type_compte == 'livreur':
        livreur = get_user_livreur(request)
        if livreur:
            livreurs = Livreur.objects.filter(id=livreur.id, agence=agence, actif=True)
        else:
            messages.error(request, 'Aucun livreur associé à votre compte.')
            return redirect('dashboard_commandes')
    else:
        # Admin/gérant voient tous les livreurs
        livreurs = Livreur.objects.filter(agence=agence, actif=True).order_by('nom', 'prenom')
    
    context = {
        'livreurs': livreurs,
        'agence': agence,
        'compte': compte,
    }
    return render(request, 'supermarket/commandes/rapport_livreur.html', context)

@login_required
def generer_rapport_livreur(request):
    """Vue AJAX pour générer les statistiques du rapport livreur"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        agence = get_user_agence(request)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Agence non trouvée: {str(e)}'})
    
    try:
        # Récupérer les paramètres
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        livreur_id = request.POST.get('livreur')
        
        # Vérification des paramètres obligatoires
        if not date_debut or not date_fin:
            return JsonResponse({'success': False, 'error': 'Les dates de début et de fin sont obligatoires'})
        
        if not livreur_id:
            return JsonResponse({'success': False, 'error': 'Le livreur est obligatoire'})
        
        # Convertir les dates
        from datetime import datetime
        date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
        
        # Récupérer le livreur
        try:
            livreur = Livreur.objects.get(id=livreur_id, agence=agence)
        except Livreur.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Livreur non trouvé'})
        
        # Vérifier que si c'est un livreur, il ne peut générer que son propre rapport
        compte = get_user_compte(request)
        if compte and compte.type_compte == 'livreur':
            livreur_utilisateur = get_user_livreur(request)
            if not livreur_utilisateur or livreur_utilisateur.id != livreur.id:
                return JsonResponse({'success': False, 'error': 'Vous ne pouvez générer que votre propre rapport.'})
        
        # Récupérer les livraisons du livreur dans la période
        livraisons = Livraison.objects.filter(
            livreur=livreur,
            agence=agence,
            date_livraison__gte=date_debut_obj,
            date_livraison__lte=date_fin_obj
        ).select_related('commande', 'commande__client', 'livreur').order_by('date_livraison', 'heure_depart')
        
        # Calculer les statistiques
        total_livraisons = livraisons.count()
        livrees = livraisons.filter(etat_livraison='livree').count()
        partielles = livraisons.filter(etat_livraison='livree_partiellement').count()
        non_livrees = livraisons.filter(etat_livraison='pas_livree').count()
        
        # Préparer les données détaillées des livraisons
        livraisons_detaillees = []
        for livraison in livraisons:
            # Récupérer les commandes associées
            # Si la livraison a une commande directe, utiliser son client et heure
            # Sinon, utiliser les informations de la livraison
            client_livraison = None
            heure_livraison = None
            
            if livraison.commande:
                client_livraison = livraison.commande.client
                heure_livraison = livraison.commande.heure
            else:
                # Si pas de commande directe, chercher par date et zone
                # On ne peut pas récupérer les commandes sans client, donc on utilise la commande si elle existe
                pass
            
            # Récupérer toutes les commandes du même groupe (même date, même heure, même client)
            if client_livraison and heure_livraison:
                commandes_groupe = Commande.objects.filter(
                    date=livraison.date_livraison,
                    heure=heure_livraison,
                    client=client_livraison,
                    agence=agence
                ).select_related('article', 'client').order_by('article__designation')
            elif livraison.commande:
                # Si on a une commande directe, utiliser seulement celle-ci
                commandes_groupe = [livraison.commande]
            else:
                # Pas de commandes associées
                commandes_groupe = []
            
            # Récupérer les quantités livrées si partiellement livrée
            quantites_livrees = {}
            raison_non_livraison = ""
            if livraison.etat_livraison == 'livree_partiellement' and livraison.notes:
                try:
                    import json
                    quantites_livrees = json.loads(livraison.notes)
                    # Extraire la raison de non-livraison si disponible
                    if isinstance(quantites_livrees, dict):
                        # Chercher une clé 'raison' ou dans les valeurs
                        for key, value in quantites_livrees.items():
                            if isinstance(value, dict) and 'raison' in value:
                                raison_non_livraison = value.get('raison', '')
                                break
                except (json.JSONDecodeError, TypeError):
                    pass
            
            # Préparer les articles de la livraison
            articles_livraison = []
            for cmd in commandes_groupe:
                quantite_commandee = float(cmd.quantite)
                quantite_livree = quantite_commandee
                quantite_non_livree = 0
                
                # Si partiellement livrée, récupérer les quantités depuis notes
                if livraison.etat_livraison == 'livree_partiellement' and str(cmd.id) in quantites_livrees:
                    quantite_livree = float(quantites_livrees[str(cmd.id)].get('quantite_livree', quantite_commandee))
                    quantite_non_livree = float(quantites_livrees[str(cmd.id)].get('quantite_non_livree', 0))
                elif livraison.etat_livraison == 'livree':
                    quantite_livree = quantite_commandee
                    quantite_non_livree = 0
                elif livraison.etat_livraison == 'pas_livree':
                    quantite_livree = 0
                    quantite_non_livree = quantite_commandee
                
                articles_livraison.append({
                    'reference': cmd.article.reference_article,
                    'designation': cmd.article.designation,
                    'quantite_commandee': quantite_commandee,
                    'quantite_livree': quantite_livree,
                    'quantite_non_livree': quantite_non_livree,
                    'unite': cmd.article.unite_vente or 'unité',
                    'prix_unitaire': float(cmd.prix_total / cmd.quantite) if cmd.quantite > 0 else 0,
                    'prix_total': float(cmd.prix_total),
                })
            
            livraisons_detaillees.append({
                'id': livraison.id,
                'date': livraison.date_livraison.strftime('%d/%m/%Y'),
                'date_iso': livraison.date_livraison.strftime('%Y-%m-%d'),
                'heure_depart': livraison.heure_depart.strftime('%H:%M') if livraison.heure_depart else '',
                'heure_arrivee': livraison.heure_arrivee.strftime('%H:%M') if livraison.heure_arrivee else '',
                'client': livraison.commande.client.intitule if livraison.commande and livraison.commande.client else '',
                'telephone': livraison.commande.client.telephone if livraison.commande and livraison.commande.client else '',
                'zone': livraison.zone or '',
                'livreur': str(livreur),
                'etat': livraison.etat_livraison,
                'etat_display': livraison.get_etat_livraison_display(),
                'articles': articles_livraison,
                'raison_non_livraison': raison_non_livraison,
            })
        
        # Stocker dans la session pour l'export Excel
        request.session['rapport_livreur'] = {
            'date_debut': date_debut,
            'date_fin': date_fin,
            'livreur_id': livreur_id,
            'livreur_nom': str(livreur),
            'total_livraisons': total_livraisons,
            'livrees': livrees,
            'partielles': partielles,
            'non_livrees': non_livrees,
            'livraisons_detaillees': livraisons_detaillees,
        }
        
        return JsonResponse({
            'success': True,
            'total': total_livraisons,
            'livrees': livrees,
            'partielles': partielles,
            'non_livrees': non_livrees,
            'livreur': str(livreur),
            'livraisons': livraisons_detaillees,
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({'success': False, 'error': f'Erreur lors de la génération: {str(e)}\n{traceback.format_exc()}'})

@login_required
def export_rapport_livreur_excel(request):
    """Vue pour exporter le rapport livreur en Excel"""
    try:
        agence = get_user_agence(request)
    except:
        return JsonResponse({'success': False, 'error': 'Agence non trouvée'})
    
    try:
        # Récupérer les données depuis la session
        rapport_data = request.session.get('rapport_livreur')
        if not rapport_data:
            return JsonResponse({'success': False, 'error': 'Aucune donnée de rapport trouvée. Veuillez générer le rapport d\'abord.'})
        
        date_debut = rapport_data['date_debut']
        date_fin = rapport_data['date_fin']
        livreur_nom = rapport_data['livreur_nom']
        total_livraisons = rapport_data['total_livraisons']
        livrees = rapport_data['livrees']
        partielles = rapport_data['partielles']
        non_livrees = rapport_data['non_livrees']
        livraisons_detaillees = rapport_data['livraisons_detaillees']
        
        # Créer le workbook Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Rapport Livreur {date_debut} - {date_fin}"
        
        # Style turquoise pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="06beb6", end_color="06beb6", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # En-tête du document
        ws.merge_cells('A1:H1')
        ws['A1'] = f"RAPPORT LIVREUR - {agence.nom_agence}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A2:H2')
        ws['A2'] = f"Période: du {date_debut} au {date_fin}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A3:H3')
        ws['A3'] = f"Livreur: {livreur_nom}"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].alignment = Alignment(horizontal="center")
        
        ws.merge_cells('A4:H4')
        ws['A4'] = f"Résumé: Total: {total_livraisons} | Livrées: {livrees} | Partielles: {partielles} | Non livrées: {non_livrees}"
        ws['A4'].font = Font(bold=True, size=11)
        ws['A4'].alignment = Alignment(horizontal="center")
        ws['A4'].fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
        
        # Tableau principal des livraisons
        row = 6
        headers = ['Date', 'Heure Départ', 'Heure Arrivée', 'Client', 'Téléphone', 'Zone', 'Livreur', 'État']
        
        # En-têtes du tableau
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row += 1
        
        # Données des livraisons
        for livraison in livraisons_detaillees:
            ws.cell(row=row, column=1, value=livraison['date']).border = thin_border
            ws.cell(row=row, column=2, value=livraison['heure_depart']).border = thin_border
            ws.cell(row=row, column=3, value=livraison['heure_arrivee']).border = thin_border
            ws.cell(row=row, column=4, value=livraison['client']).border = thin_border
            ws.cell(row=row, column=5, value=livraison['telephone']).border = thin_border
            ws.cell(row=row, column=6, value=livraison['zone']).border = thin_border
            ws.cell(row=row, column=7, value=livraison['livreur']).border = thin_border
            
            # État avec couleur
            etat_cell = ws.cell(row=row, column=8, value=livraison['etat_display'])
            etat_cell.border = thin_border
            etat_cell.alignment = Alignment(horizontal="center")
            
            # Couleur selon l'état
            if livraison['etat'] == 'livree':
                etat_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                etat_cell.font = Font(bold=True, color="28a745")
            elif livraison['etat'] == 'livree_partiellement':
                etat_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                etat_cell.font = Font(bold=True, color="ffc107")
            elif livraison['etat'] == 'pas_livree':
                etat_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                etat_cell.font = Font(bold=True, color="dc3545")
            
            row += 1
            
            # Section ARTICLES pour cette livraison
            ws.merge_cells(f'A{row}:H{row}')
            ws.cell(row=row, column=1, value="ARTICLES:")
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            row += 1
            
            # En-têtes des articles
            article_headers = ['Référence', 'Désignation', 'Qté Commandée', 'Qté Livrée', 'Qté Non Livrée', 'Unité', 'Prix Unitaire', 'Prix Total']
            for col, header in enumerate(article_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
            
            row += 1
            
            # Articles de la livraison
            for article in livraison['articles']:
                ws.cell(row=row, column=1, value=article['reference']).border = thin_border
                ws.cell(row=row, column=2, value=article['designation']).border = thin_border
                ws.cell(row=row, column=3, value=float(article['quantite_commandee'])).border = thin_border
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                ws.cell(row=row, column=4, value=float(article['quantite_livree'])).border = thin_border
                ws.cell(row=row, column=4).number_format = '#,##0.00'
                ws.cell(row=row, column=5, value=float(article['quantite_non_livree'])).border = thin_border
                ws.cell(row=row, column=5).number_format = '#,##0.00'
                ws.cell(row=row, column=6, value=article['unite']).border = thin_border
                ws.cell(row=row, column=7, value=float(article['prix_unitaire'])).border = thin_border
                ws.cell(row=row, column=7).number_format = '#,##0.00'
                ws.cell(row=row, column=8, value=float(article['prix_total'])).border = thin_border
                ws.cell(row=row, column=8).number_format = '#,##0.00'
                row += 1
            
            # Raison non-livraison si partiellement livrée
            if livraison['etat'] == 'livree_partiellement' and livraison.get('raison_non_livraison'):
                ws.merge_cells(f'A{row}:H{row}')
                raison_cell = ws.cell(row=row, column=1, value=f"Raison non-livraison: {livraison['raison_non_livraison']}")
                raison_cell.font = Font(bold=True, color="dc3545")
                raison_cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                row += 1
            
            row += 1  # Espacement entre les livraisons
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 12  # Date
        ws.column_dimensions['B'].width = 12  # Heure Départ
        ws.column_dimensions['C'].width = 12  # Heure Arrivée
        ws.column_dimensions['D'].width = 25  # Client
        ws.column_dimensions['E'].width = 15  # Téléphone
        ws.column_dimensions['F'].width = 15  # Zone
        ws.column_dimensions['G'].width = 20  # Livreur
        ws.column_dimensions['H'].width = 18  # État
        
        # Réponse HTTP avec le fichier Excel
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Rapport_Livreur_{livreur_nom.replace(' ', '_')}_{date_debut}_{date_fin}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        return JsonResponse({'success': False, 'error': f'Erreur lors de l\'export: {str(e)}\n{traceback.format_exc()}'})

# ==================== PAGES DE TRAVAIL PERSONNALISÉES ====================

@login_required
@require_commandes_feature('gestion_commande')
def travail_acm(request):
    """Page de travail personnalisée pour les ACM - Affiche seulement leurs données"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    compte = get_user_compte(request)
    if not compte or compte.type_compte != 'acm':
        messages.error(request, 'Accès réservé aux ACM.')
        return redirect('login_commandes')
    
    # Filtrer les données pour cet ACM uniquement
    commandes_queryset = Commande.objects.filter(agence=agence, created_by=compte.user)
    suivi_actions_queryset = SuiviClientAction.objects.filter(agence=agence, created_by=compte.user)
    
    # Statistiques des commandes
    total_commandes = commandes_queryset.count()
    commandes_en_attente = commandes_queryset.filter(etat_commande='en_attente').count()
    commandes_validees = commandes_queryset.filter(etat_commande='validee').count()
    commandes_livrees = commandes_queryset.filter(etat_commande='livree').count()
    
    # Commandes récentes (10 dernières)
    commandes_recents = commandes_queryset.select_related('client', 'article').order_by('-date', '-heure')[:10]
    
    # Actions de suivi récentes (10 dernières)
    actions_recents = suivi_actions_queryset.select_related('client').order_by('-date_action', '-heure_appel')[:10]
    
    # Zones des clients suivis par cet ACM
    zones = Client.objects.filter(
        agence=agence,
        actions_suivi__created_by=compte.user
    ).values_list('zone', flat=True).distinct().exclude(zone__isnull=True).exclude(zone='').order_by('zone')
    
    context = {
        'agence': agence,
        'compte': compte,
        'nom_utilisateur': compte.nom_complet,
        'total_commandes': total_commandes,
        'commandes_en_attente': commandes_en_attente,
        'commandes_en_livraison': 0,  # Pas de livraisons pour ACM
        'commandes_livrees': commandes_livrees,
        'commandes_annulees': 0,
        'total_livraisons': 0,
        'livraisons_en_cours': 0,
        'total_factures': 0,
        'total_statistiques': 0,
        'commandes_recents': commandes_recents,
        'is_acm_view': True,  # Flag pour identifier la vue ACM
    }
    return render(request, 'supermarket/commandes/dashboard.html', context)

@login_required
@require_commandes_feature('definir_etat_livraison')
def travail_livreur(request):
    """Page de travail personnalisée pour les livreurs - Affiche seulement leurs livraisons"""
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_commandes')
    
    compte = get_user_compte(request)
    if not compte or compte.type_compte != 'livreur':
        messages.error(request, 'Accès réservé aux livreurs.')
        return redirect('login_commandes')
    
    livreur = get_user_livreur(request)
    if not livreur:
        # Créer automatiquement un livreur pour ce compte s'il n'existe pas
        try:
            livreur = Livreur.objects.create(
                nom=compte.nom,
                prenom=compte.prenom,
                telephone=compte.telephone,
                email=compte.email,
                agence=agence,
                compte=compte,
                actif=True
            )
            messages.success(request, f'Livreur créé automatiquement pour votre compte : {livreur.nom} {livreur.prenom}')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du livreur : {str(e)}')
            return redirect('login_commandes')
    
    # Filtrer les livraisons pour ce livreur uniquement
    livraisons_queryset = Livraison.objects.filter(agence=agence, livreur=livreur)
    
    # Statistiques des livraisons
    total_livraisons = livraisons_queryset.count()
    livraisons_planifiees = livraisons_queryset.filter(etat_livraison='planifiee').count()
    livraisons_en_preparation = livraisons_queryset.filter(etat_livraison='en_preparation').count()
    livraisons_en_cours = livraisons_queryset.filter(etat_livraison='en_cours').count()
    livraisons_confirmees = livraisons_queryset.filter(etat_livraison='confirmee').count()
    livraisons_livrees = livraisons_queryset.filter(etat_livraison='livree').count()
    livraisons_partielles = livraisons_queryset.filter(etat_livraison='livree_partiellement').count()
    livraisons_non_livrees = livraisons_queryset.filter(etat_livraison='pas_livree').count()
    
    # Livraisons à traiter aujourd'hui (confirmées ou en cours)
    from django.utils import timezone
    livraisons_aujourdhui = livraisons_queryset.filter(
        date_livraison=timezone.now().date(),
        etat_livraison__in=['confirmee', 'en_cours', 'planifiee', 'en_preparation']
    ).select_related('commande', 'commande__client').order_by('heure_depart', 'ordre_livraison')
    
    # Livraisons récentes (10 dernières)
    livraisons_recents = livraisons_queryset.select_related('commande', 'commande__client').order_by('-date_livraison', '-heure_depart')[:10]
    
    context = {
        'agence': agence,
        'compte': compte,
        'livreur': livreur,
        'nom_utilisateur': compte.nom_complet,
        'total_commandes': 0,  # Pas de commandes pour livreur
        'commandes_en_attente': 0,
        'commandes_en_livraison': 0,
        'commandes_livrees': 0,
        'commandes_annulees': 0,
        'total_livraisons': total_livraisons,
        'livraisons_en_cours': livraisons_en_cours,
        'total_factures': 0,
        'total_statistiques': 0,
        'commandes_recents': [],
        'livraisons_recents': livraisons_recents,
        'livraisons_aujourdhui': livraisons_aujourdhui,
        'is_livreur_view': True,  # Flag pour identifier la vue livreur
    }
    return render(request, 'supermarket/commandes/dashboard.html', context)

# ==================== MODULE ANALYSE FINANCIÈRE ====================

def login_financier(request):
    """Page de connexion pour le module Analyse Financière - Réservé aux analystes financières et administrateurs"""
    # Toujours afficher la page de connexion, même si l'utilisateur est déjà connecté
    if request.user.is_authenticated and request.method == 'GET':
        logout(request)
        request.session.flush()
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        # Vérifier que le type de compte est analyste_financiere ou admin
                        if compte.type_compte not in ['analyste_financiere', 'admin']:
                            type_display = compte.get_type_compte_display()
                            messages.error(
                                request, 
                                f'Accès refusé. Ce module est réservé uniquement aux analystes financières et aux administrateurs. '
                                f'Votre type de compte ({type_display}) n\'est pas autorisé.'
                            )
                            return render(request, 'supermarket/financier/login.html')
                        
                        login(request, user)
                        # Stocker l'agence dans la session
                        request.session['agence_id'] = compte.agence.id_agence
                        request.session['agence_nom'] = compte.agence.nom_agence
                        messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                        # Rediriger vers la page demandée (next) ou vers le dashboard
                        next_url = request.GET.get('next', 'dashboard_financier')
                        return redirect(next_url)
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Compte non trouvé ou inactif.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')
    
    return render(request, 'supermarket/financier/login.html')

def logout_financier(request):
    """Déconnexion du module Analyse Financière"""
    logout(request)
    request.session.flush()
    messages.success(request, 'Vous avez été déconnecté avec succès.')
    return redirect('index')

def check_comptable_access(request):
    """Vérifier que l'utilisateur est un analyste financière ou admin"""
    if not request.user.is_authenticated:
        # Rediriger vers la page de login financier avec le paramètre next
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        login_url = reverse('login_financier')
        return None, HttpResponseRedirect(f'{login_url}?next={request.path}')
    
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
        if compte.type_compte not in ['analyste_financiere', 'admin']:
            messages.error(request, 'Accès refusé. Ce module est réservé aux analystes financières et administrateurs.')
            return None, redirect('index')
        return compte, None
    except Compte.DoesNotExist:
        messages.error(request, 'Votre compte n\'est pas configuré correctement.')
        return None, redirect('index')

def dashboard_financier(request):
    """Dashboard principal du module Analyse Financière"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_financier')
    
    # Calculer les statistiques financières pour TOUTES les agences
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count, Avg, F
    
    # Période : 30 derniers jours
    date_debut = timezone.now().date() - timedelta(days=30)
    
    # Chiffre d'affaires (ventes) - TOUTES LES AGENCES
    ca_total = FactureVente.objects.filter(
        date__gte=date_debut
    ).aggregate(total=Sum('nette_a_payer'))['total'] or Decimal('0')
    
    # Dépenses (module Dépense) - TOUTES LES AGENCES
    # Convertir date_debut en datetime pour le filtre
    from datetime import time as time_class
    date_debut_dt = datetime.combine(date_debut, time_class.min)
    depenses_total = Depense.objects.filter(
        date__gte=date_debut_dt
    ).aggregate(total=Sum('montant'))['total'] or Decimal('0')
    
    # Calculer la marge totale (somme de toutes les marges de toutes les agences) - TOUTES LES AGENCES
    # Optimisation : Calcul direct en base de données avec agrégation au lieu de boucle Python
    # La marge = (Prix de vente - Prix d'achat) × Quantité pour chaque ligne de vente
    from django.db.models import ExpressionWrapper, DecimalField
    
    # Calculer la marge totale directement en SQL pour toutes les agences
    marge_totale_result = LigneFactureVente.objects.filter(
        facture_vente__date__gte=date_debut,
        article__isnull=False
    ).aggregate(
        total_marge=Sum(
            ExpressionWrapper(
                (F('prix_unitaire') - F('article__prix_achat')) * F('quantite'),
                output_field=DecimalField(max_digits=15, decimal_places=2)
            )
        )
    )['total_marge'] or Decimal('0')
    
    marge_totale = Decimal(str(marge_totale_result))
    
    # Résultat net (Bénéfice) = Marge - Dépenses
    # La marge est le bénéfice brut, donc Résultat = Marge - Dépenses
    resultat_net = marge_totale - depenses_total
    
    # Statistiques supplémentaires - TOUTES LES AGENCES
    nb_ventes = FactureVente.objects.filter(date__gte=date_debut).count()
    nb_depenses = Depense.objects.filter(date__gte=date_debut_dt).count()
    
    # CA moyen par vente
    ca_moyen = ca_total / nb_ventes if nb_ventes > 0 else Decimal('0')
    
    # Évolution sur 30 derniers jours (pour graphique) - TOUTES LES AGENCES
    # Optimisation : Utiliser des agrégations groupées au lieu de boucles avec requêtes individuelles
    date_fin_evolution = timezone.now().date()
    date_debut_evolution = date_fin_evolution - timedelta(days=30)
    
    # Récupérer le CA groupé par date (une seule requête SQL)
    ca_par_date = FactureVente.objects.filter(
        date__range=[date_debut_evolution, date_fin_evolution]
    ).values('date').annotate(
        total_ca=Sum('nette_a_payer')
    ).order_by('date')
    
    # Récupérer les dépenses groupées par date (une seule requête SQL)
    depenses_par_date = Depense.objects.filter(
        date__gte=datetime.combine(date_debut_evolution, time_class.min),
        date__lte=datetime.combine(date_fin_evolution, time_class.max)
    ).extra(
        select={'date_only': "DATE(date)"}
    ).values('date_only').annotate(
        total_depenses=Sum('montant')
    ).order_by('date_only')
    
    # Récupérer les marges groupées par date (une seule requête SQL)
    marges_par_date = LigneFactureVente.objects.filter(
        facture_vente__date__range=[date_debut_evolution, date_fin_evolution],
        article__isnull=False
    ).values('facture_vente__date').annotate(
        total_marge=Sum(
            ExpressionWrapper(
                (F('prix_unitaire') - F('article__prix_achat')) * F('quantite'),
                output_field=DecimalField(max_digits=15, decimal_places=2)
            )
        )
    ).order_by('facture_vente__date')
    
    # Créer des dictionnaires pour accès rapide
    ca_dict = {str(item['date']): float(item['total_ca'] or 0) for item in ca_par_date}
    depenses_dict = {str(item['date_only']): float(item['total_depenses'] or 0) for item in depenses_par_date}
    marges_dict = {str(item['facture_vente__date']): float(item['total_marge'] or 0) for item in marges_par_date}
    
    # Construire les listes pour les 30 derniers jours
    evolution_ca = []
    evolution_depenses = []
    evolution_resultat = []
    labels_jours = []
    
    for i in range(30, 0, -1):
        date_jour = date_fin_evolution - timedelta(days=i)
        date_jour_str = str(date_jour)
        labels_jours.append(date_jour.strftime('%d/%m'))
        
        ca_jour = ca_dict.get(date_jour_str, 0.0)
        dep_jour = depenses_dict.get(date_jour_str, 0.0)
        marge_jour = marges_dict.get(date_jour_str, 0.0)
        resultat_jour = marge_jour - dep_jour
        
        evolution_ca.append(ca_jour)
        evolution_depenses.append(dep_jour)
        evolution_resultat.append(resultat_jour)
    
    # Top 5 articles les plus vendus - TOUTES LES AGENCES
    top_articles = LigneFactureVente.objects.filter(
        facture_vente__date__gte=date_debut
    ).values('article__designation').annotate(
        total_vente=Sum(F('quantite') * F('prix_unitaire')),
        quantite_totale=Sum('quantite')
    ).order_by('-total_vente')[:5]
    
    context = {
        'agence': agence,
        'compte': compte,
        'ca_total': float(ca_total),
        'depenses_total': float(depenses_total),
        'marge_totale': float(marge_totale),
        'resultat_net': float(resultat_net),
        'nb_ventes': nb_ventes,
        'nb_achats': nb_depenses,
        'ca_moyen': float(ca_moyen),
        'pourcentage_marge': float((marge_totale / ca_total * 100) if ca_total > 0 else 0),
        'evolution_ca': evolution_ca,
        'evolution_depenses': evolution_depenses,
        'evolution_resultat': evolution_resultat,
        'labels_jours': labels_jours,
        'top_articles': top_articles,
        'date_debut': date_debut,
        'date_fin': timezone.now().date(),
    }
    return render(request, 'supermarket/financier/dashboard_financier.html', context)

def suivi_statistique(request):
    """Vue pour le suivi statistique - Toutes les agences avec regroupement par marge"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_financier')
    
    # Récupérer toutes les agences
    agences = Agence.objects.all().order_by('nom_agence')
    
    context = {
        'agence': agence,
        'compte': compte,
        'agences': agences,
    }
    return render(request, 'supermarket/financier/suivi_statistique.html', context)

def etat_depense(request):
    """Vue pour l'état des dépenses - Toutes les agences"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_financier')
    
    # Récupérer toutes les agences
    agences = Agence.objects.all().order_by('nom_agence')
    
    context = {
        'agence': agence,
        'compte': compte,
        'agences': agences,
    }
    return render(request, 'supermarket/financier/etat_depense.html', context)

def etat_tresorerie(request):
    """Vue pour l'état de trésorerie - Toutes les agences"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_financier')
    
    # Récupérer toutes les agences sauf "Agence Principale" et "Super Market Principal"
    agences = Agence.objects.exclude(
        nom_agence__in=['Agence Principale', 'Super Market Principal']
    ).order_by('nom_agence')
    
    context = {
        'agence': agence,
        'compte': compte,
        'agences': agences,
    }
    return render(request, 'supermarket/financier/etat_tresorerie.html', context)

def generer_etat_tresorerie(request):
    """Générer l'état de trésorerie pour toutes les agences sur une période donnée"""
    # Vérifier l'authentification pour les requêtes AJAX
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Non authentifié. Veuillez vous reconnecter.'}, status=401)
    
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
        if compte.type_compte not in ['analyste_financiere', 'comptable', 'assistant_comptable', 'admin']:
            return JsonResponse({'success': False, 'error': 'Accès refusé.'}, status=403)
    except Compte.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')
        
        if not date_debut_str or not date_fin_str:
            return JsonResponse({'success': False, 'error': 'Les dates sont obligatoires'})
        
        from datetime import datetime
        from django.db.models import Sum, F
        from decimal import Decimal
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        
        # Récupérer toutes les agences sauf "Agence Principale" et "Super Market Principal"
        agences = Agence.objects.exclude(
            nom_agence__in=['Agence Principale', 'Super Market Principal']
        ).order_by('nom_agence')
        
        # Trésorerie par agence
        tresorerie_par_agence = {}
        total_global = {
            'ca': Decimal('0'),
            'banque': Decimal('0'),
            'caisse': Decimal('0'),
            'om': Decimal('0'),
            'momo': Decimal('0'),
            'sav': Decimal('0'),
            'total_disponible': Decimal('0')
        }
        
        for agence in agences:
            # Récupérer la dernière trésorerie de cette agence dans la période
            # On prend le dernier jour de la période pour chaque agence
            treso = Tresorerie.objects.filter(
                agence=agence,
                date__lte=date_fin
            ).order_by('-date').first()
            
            # Calculer le CA pour cette agence du DERNIER JOUR de la période (date_fin)
            # et non le cumulé sur toute la période
            ventes_jour = LigneFactureVente.objects.filter(
                facture_vente__agence=agence,
                facture_vente__date=date_fin
            )
            ca_agence = ventes_jour.aggregate(total=Sum(F('quantite') * F('prix_unitaire')))['total'] or Decimal('0')
            
            # Variables pour stocker les valeurs Decimal avant conversion en float
            banque_val = Decimal('0')
            caisse_val = Decimal('0')
            om_val = Decimal('0')
            momo_val = Decimal('0')
            sav_val = Decimal('0')
            total_dispo_val = Decimal('0')
            
            if treso:
                # Utiliser les soldes finaux de la dernière trésorerie (déjà en Decimal)
                banque_val = treso.solde_banque_final
                caisse_val = treso.solde_caisse_final
                om_val = treso.solde_om_final
                momo_val = treso.solde_momo_final
                sav_val = treso.solde_sav_final
                total_dispo_val = treso.total_disponible
                
                tresorerie_par_agence[agence.id_agence] = {
                    'agence_nom': agence.nom_agence,
                    'ca': float(ca_agence),
                    'banque': float(banque_val),
                    'caisse': float(caisse_val),
                    'om': float(om_val),
                    'momo': float(momo_val),
                    'sav': float(sav_val),
                    'total_disponible': float(total_dispo_val),
                    'date_derniere_maj': treso.date.strftime('%d/%m/%Y')
                }
            else:
                # Aucune trésorerie trouvée pour cette agence
                tresorerie_par_agence[agence.id_agence] = {
                    'agence_nom': agence.nom_agence,
                    'ca': float(ca_agence),
                    'banque': 0.0,
                    'caisse': 0.0,
                    'om': 0.0,
                    'momo': 0.0,
                    'sav': 0.0,
                    'total_disponible': 0.0,
                    'date_derniere_maj': 'Aucune'
                }
            
            # Ajouter aux totaux globaux (utiliser directement les Decimal)
            total_global['ca'] += ca_agence
            total_global['banque'] += banque_val
            total_global['caisse'] += caisse_val
            total_global['om'] += om_val
            total_global['momo'] += momo_val
            total_global['sav'] += sav_val
            total_global['total_disponible'] += total_dispo_val
        
        # Convertir les totaux en float pour JSON
        total_global_float = {k: float(v) for k, v in total_global.items()}
        
        # Stocker dans la session pour l'export Excel
        request.session['etat_tresorerie'] = {
            'date_debut': date_debut_str,
            'date_fin': date_fin_str,
            'tresorerie_par_agence': tresorerie_par_agence,
            'total_global': total_global_float
        }
        
        return JsonResponse({
            'success': True,
            'data': {
                'tresorerie_par_agence': tresorerie_par_agence,
                'total_global': total_global_float,
                'date_debut': date_debut_str,
                'date_fin': date_fin_str
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erreur: {str(e)}'})

def export_etat_tresorerie_excel(request):
    """Exporter l'état de trésorerie en Excel"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    try:
        # Récupérer les données depuis la session
        etat_data = request.session.get('etat_tresorerie')
        
        if not etat_data:
            return JsonResponse({'success': False, 'error': 'Aucun état généré'})
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé'})
        
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"État Trésorerie {etat_data['date_debut']} - {etat_data['date_fin']}"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        agence_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ca_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
        
        row = 1
        
        # Titre
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = f"ÉTAT DE TRÉSORERIE - Période du {etat_data['date_debut']} au {etat_data['date_fin']}"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
        row += 2
        
        # En-têtes des colonnes
        headers = ['Agence', 'Chiffre d\'Affaire', 'Banque', 'Caisse', 'Orange Money', 'MTN Money', 'SAV', 'Total Disponible']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Données par agence
        for agence_id, agence_data in etat_data['tresorerie_par_agence'].items():
            ws.cell(row=row, column=1, value=agence_data['agence_nom']).fill = agence_fill
            ws.cell(row=row, column=2, value=agence_data['ca']).number_format = '#,##0.00'
            ws.cell(row=row, column=3, value=agence_data['banque']).number_format = '#,##0.00'
            ws.cell(row=row, column=4, value=agence_data['caisse']).number_format = '#,##0.00'
            ws.cell(row=row, column=5, value=agence_data['om']).number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=agence_data['momo']).number_format = '#,##0.00'
            ws.cell(row=row, column=7, value=agence_data['sav']).number_format = '#,##0.00'
            ws.cell(row=row, column=8, value=agence_data['total_disponible']).number_format = '#,##0.00'
            row += 1
        
        # Ligne de total
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").fill = total_fill
        ws.cell(row=row, column=1).font = Font(bold=True)
        total_global = etat_data['total_global']
        ws.cell(row=row, column=2, value=total_global['ca']).number_format = '#,##0.00'
        ws.cell(row=row, column=2).fill = total_fill
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_global['banque']).number_format = '#,##0.00'
        ws.cell(row=row, column=3).fill = total_fill
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_global['caisse']).number_format = '#,##0.00'
        ws.cell(row=row, column=4).fill = total_fill
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_global['om']).number_format = '#,##0.00'
        ws.cell(row=row, column=5).fill = total_fill
        ws.cell(row=row, column=5).font = Font(bold=True)
        ws.cell(row=row, column=6, value=total_global['momo']).number_format = '#,##0.00'
        ws.cell(row=row, column=6).fill = total_fill
        ws.cell(row=row, column=6).font = Font(bold=True)
        ws.cell(row=row, column=7, value=total_global['sav']).number_format = '#,##0.00'
        ws.cell(row=row, column=7).fill = total_fill
        ws.cell(row=row, column=7).font = Font(bold=True)
        ws.cell(row=row, column=8, value=total_global['total_disponible']).number_format = '#,##0.00'
        ws.cell(row=row, column=8).fill = total_fill
        ws.cell(row=row, column=8).font = Font(bold=True)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 30
        for col in range(2, 9):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        # Réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Etat_Tresorerie_{etat_data['date_debut']}_{etat_data['date_fin']}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erreur export: {str(e)}'})

def etat_resultat(request):
    """Vue pour l'état de résultat - Toutes les agences"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    agence = get_user_agence(request)
    if not agence:
        messages.error(request, 'Votre compte n\'est pas correctement lié à une agence.')
        return redirect('login_financier')
    
    # Récupérer toutes les agences
    agences = Agence.objects.all().order_by('nom_agence')
    
    context = {
        'agence': agence,
        'compte': compte,
        'agences': agences,
    }
    return render(request, 'supermarket/financier/etat_resultat.html', context)

def generer_etat_depense(request):
    """Générer l'état des dépenses pour toutes les agences sur une période donnée"""
    # Vérifier l'authentification pour les requêtes AJAX
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Non authentifié. Veuillez vous reconnecter.'}, status=401)
    
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
        if compte.type_compte not in ['analyste_financiere', 'admin']:
            return JsonResponse({'success': False, 'error': 'Accès refusé. Ce module est réservé aux analystes financières et administrateurs.'}, status=403)
    except Compte.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')
        
        if not date_debut_str or not date_fin_str:
            return JsonResponse({'success': False, 'error': 'Les dates sont obligatoires'})
        
        from datetime import datetime
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        
        # Récupérer toutes les agences
        agences = Agence.objects.all().order_by('nom_agence')
        
        # Dépenses par agence
        depenses_par_agence = {}
        total_general = Decimal('0')
        
        for agence in agences:
            # Convertir les dates en datetime pour le filtre (start et end de la journée)
            from datetime import time as time_class
            date_debut_dt = datetime.combine(date_debut, time_class.min)
            date_fin_dt = datetime.combine(date_fin, time_class.max)
            
            # Récupérer les dépenses de cette agence pour la période
            depenses_agence = Depense.objects.filter(
                agence=agence,
                date__range=[date_debut_dt, date_fin_dt]
            ).order_by('date')
            
            # Préparer les données des dépenses
            liste_depenses = []
            total_agence = Decimal('0')
            
            for depense in depenses_agence:
                # Sérialiser la date en string ISO pour JSON
                if depense.date:
                    # Convertir datetime avec timezone en datetime naive puis en string
                    if hasattr(depense.date, 'replace'):
                        date_naive = depense.date.replace(tzinfo=None)
                        date_str = date_naive.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        date_str = str(depense.date)
                else:
                    date_str = None
                    
                liste_depenses.append({
                    'date': date_str,
                    'libelle': depense.libelle,
                    'montant': float(depense.montant),
                })
                total_agence += depense.montant
            
            if liste_depenses:
                depenses_par_agence[agence.id_agence] = {
                    'nom_agence': agence.nom_agence,
                    'depenses': liste_depenses,
                    'total': float(total_agence),
                }
                total_general += total_agence
        
        # Stocker dans la session pour l'export
        request.session['etat_depense'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'depenses_par_agence': depenses_par_agence,
            'total_general': float(total_general),
        }
        
        return JsonResponse({
            'success': True,
            'depenses_par_agence': depenses_par_agence,
            'total_general': float(total_general),
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erreur: {str(e)}'})

def export_etat_depense_excel(request):
    """Exporter l'état des dépenses en Excel"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    try:
        # Récupérer les données depuis la session
        etat_data = request.session.get('etat_depense')
        
        if not etat_data:
            return JsonResponse({'success': False, 'error': 'Aucun état généré'})
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé'})
        
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"État Dépenses {etat_data['date_debut']} - {etat_data['date_fin']}"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        agence_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        row = 1
        
        # En-têtes des colonnes
        headers = ['Date', 'Libellé', 'Montant (FCFA)']
        
        # Par agence
        for agence_id, agence_data in etat_data['depenses_par_agence'].items():
            # En-tête Agence
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = f"AGENCE: {agence_data['nom_agence']}"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].fill = agence_fill
            ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            
            # En-têtes du tableau
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row += 1
            
            # Dépenses de l'agence
            for depense in agence_data['depenses']:
                # Date (colonne A) - depense['date'] est déjà une string formatée 'YYYY-MM-DD HH:MM:SS'
                date_value = depense.get('date', '')
                if date_value:
                    try:
                        # Parser la string pour créer un objet datetime pour Excel
                        from datetime import datetime
                        # Parser le format 'YYYY-MM-DD HH:MM:SS' ou 'YYYY-MM-DD'
                        date_str_clean = date_value.split(' ')[0] if ' ' in date_value else date_value
                        date_obj = datetime.strptime(date_str_clean, '%Y-%m-%d')
                        ws.cell(row=row, column=1, value=date_obj).number_format = 'DD/MM/YYYY HH:MM'
                    except (ValueError, AttributeError):
                        # En cas d'erreur de parsing, utiliser la string telle quelle
                        ws.cell(row=row, column=1, value=date_value)
                else:
                    ws.cell(row=row, column=1, value="")
                ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
                
                # Libellé (colonne B)
                ws.cell(row=row, column=2, value=depense['libelle']).alignment = Alignment(vertical="center")
                
                # Montant (colonne C)
                cell_montant = ws.cell(row=row, column=3, value=depense['montant'])
                cell_montant.number_format = '#,##0.00'
                cell_montant.alignment = Alignment(horizontal="right", vertical="center")
                
                row += 1
            
            # Total de l'agence
            ws.merge_cells(f'A{row}:B{row}')
            cell_total_agence_label = ws.cell(row=row, column=1, value=f"TOTAL {agence_data['nom_agence']}")
            cell_total_agence_label.font = Font(bold=True, size=11)
            cell_total_agence_label.alignment = Alignment(horizontal="right", vertical="center")
            
            cell_total_agence = ws.cell(row=row, column=3, value=agence_data['total'])
            cell_total_agence.font = Font(bold=True, size=11)
            cell_total_agence.number_format = '#,##0.00'
            cell_total_agence.alignment = Alignment(horizontal="right", vertical="center")
            cell_total_agence.fill = total_fill
            
            row += 1
            row += 1  # Ligne vide entre les agences
        
        # Total général
        ws.merge_cells(f'A{row}:B{row}')
        cell_total_general_label = ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL")
        cell_total_general_label.font = Font(bold=True, size=12)
        cell_total_general_label.alignment = Alignment(horizontal="right", vertical="center")
        
        cell_total_general = ws.cell(row=row, column=3, value=etat_data['total_general'])
        cell_total_general.font = Font(bold=True, size=12)
        cell_total_general.number_format = '#,##0.00'
        cell_total_general.alignment = Alignment(horizontal="right", vertical="center")
        cell_total_general.fill = total_fill
        
        # Ajuster largeurs colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        
        # Réponse HTTP
        filename = f"Etat_Depenses_{etat_data['date_debut']}_{etat_data['date_fin']}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erreur export: {str(e)}'})

def generer_etat_resultat(request):
    """Générer l'état des résultats pour toutes les agences sur une période donnée"""
    # Vérifier l'authentification pour les requêtes AJAX
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'Non authentifié. Veuillez vous reconnecter.'}, status=401)
    
    try:
        compte = Compte.objects.get(user=request.user, actif=True)
        if compte.type_compte not in ['comptable', 'admin']:
            return JsonResponse({'success': False, 'error': 'Accès refusé. Ce module est réservé aux comptables.'}, status=403)
    except Compte.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Votre compte n\'est pas configuré correctement.'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')
        
        if not date_debut_str or not date_fin_str:
            return JsonResponse({'success': False, 'error': 'Les dates sont obligatoires'})
        
        from datetime import datetime, time
        from django.db.models import ExpressionWrapper, DecimalField
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        
        # Récupérer toutes les agences sauf "Agence Principale" et "Super Market Principal"
        agences = Agence.objects.exclude(
            nom_agence__in=['Agence Principale', 'Super Market Principal']
        ).order_by('nom_agence')
        
        # Résultats par agence
        resultats_par_agence = {}
        total_ca_general = Decimal('0')
        total_depenses_general = Decimal('0')
        total_marge_general = Decimal('0')
        total_resultat_general = Decimal('0')
        
        # Convertir les dates en datetime pour les filtres
        date_debut_dt = datetime.combine(date_debut, time.min)
        date_fin_dt = datetime.combine(date_fin, time.max)
        
        for agence in agences:
            # Chiffre d'affaires de l'agence (ventes)
            ca_agence = FactureVente.objects.filter(
                agence=agence,
                date__range=[date_debut, date_fin]
            ).aggregate(total=Sum('nette_a_payer'))['total'] or Decimal('0')
            
            # Dépenses de l'agence
            depenses_agence = Depense.objects.filter(
                agence=agence,
                date__range=[date_debut_dt, date_fin_dt]
            ).aggregate(total=Sum('montant'))['total'] or Decimal('0')
            
            # Marge (Bénéfice brut) de l'agence
            # Calculer la marge directement en SQL pour cette agence
            marge_agence_result = LigneFactureVente.objects.filter(
                facture_vente__agence=agence,
                facture_vente__date__range=[date_debut, date_fin],
                article__isnull=False
            ).aggregate(
                total_marge=Sum(
                    ExpressionWrapper(
                        (F('prix_unitaire') - F('article__prix_achat')) * F('quantite'),
                        output_field=DecimalField(max_digits=15, decimal_places=2)
                    )
                )
            )['total_marge'] or Decimal('0')
            
            marge_agence = Decimal(str(marge_agence_result))
            
            # Résultat = Bénéfice (Marge) - Dépenses
            resultat_agence = marge_agence - depenses_agence
            
            # Stocker les résultats même si tout est à zéro (pour afficher toutes les agences)
            resultats_par_agence[agence.id_agence] = {
                'nom_agence': agence.nom_agence,
                'ca': float(ca_agence),
                'depenses': float(depenses_agence),
                'marge': float(marge_agence),
                'resultat': float(resultat_agence),
            }
            
            total_ca_general += ca_agence
            total_depenses_general += depenses_agence
            total_marge_general += marge_agence
            total_resultat_general += resultat_agence
        
        # Stocker dans la session pour l'export
        request.session['etat_resultat'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'resultats_par_agence': resultats_par_agence,
            'total_ca': float(total_ca_general),
            'total_depenses': float(total_depenses_general),
            'total_marge': float(total_marge_general),
            'total_resultat': float(total_resultat_general),
        }
        
        return JsonResponse({
            'success': True,
            'resultats_par_agence': resultats_par_agence,
            'total_ca': float(total_ca_general),
            'total_depenses': float(total_depenses_general),
            'total_marge': float(total_marge_general),
            'total_resultat': float(total_resultat_general),
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erreur: {str(e)}'})

def export_etat_resultat_excel(request):
    """Exporter l'état des résultats en Excel"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    try:
        # Récupérer les données depuis la session
        etat_data = request.session.get('etat_resultat')
        
        if not etat_data:
            return JsonResponse({'success': False, 'error': 'Aucun état généré'})
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé'})
        
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"État Résultats {etat_data['date_debut']} - {etat_data['date_fin']}"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        agence_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        agence_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        total_font = Font(bold=True, size=12)
        grand_total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        grand_total_font = Font(bold=True, color="FFFFFF", size=13)
        
        row = 1
        
        # Titre du document
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f"ÉTAT DES RÉSULTATS - Période du {etat_data['date_debut']} au {etat_data['date_fin']}"
        ws[f'A{row}'].font = Font(bold=True, size=14, color="2C3E50")
        ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
        row += 2
        
        # En-têtes des colonnes
        headers = ['Agence', 'Chiffre d\'Affaires (FCFA)', 'Dépenses (FCFA)', 'Marge/Bénéfice (FCFA)', 'Résultat Net (FCFA)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        # Données par agence
        for agence_id, agence_data in etat_data['resultats_par_agence'].items():
            # Nom de l'agence
            ws.cell(row=row, column=1, value=agence_data['nom_agence']).font = agence_font
            ws.cell(row=row, column=1).fill = agence_fill
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
            
            # CA
            cell_ca = ws.cell(row=row, column=2, value=agence_data['ca'])
            cell_ca.number_format = '#,##0.00'
            cell_ca.alignment = Alignment(horizontal="right", vertical="center")
            
            # Dépenses
            cell_dep = ws.cell(row=row, column=3, value=agence_data['depenses'])
            cell_dep.number_format = '#,##0.00'
            cell_dep.alignment = Alignment(horizontal="right", vertical="center")
            
            # Marge/Bénéfice
            cell_marge = ws.cell(row=row, column=4, value=agence_data['marge'])
            cell_marge.number_format = '#,##0.00'
            cell_marge.alignment = Alignment(horizontal="right", vertical="center")
            
            # Résultat Net
            cell_resultat = ws.cell(row=row, column=5, value=agence_data['resultat'])
            cell_resultat.number_format = '#,##0.00'
            cell_resultat.alignment = Alignment(horizontal="right", vertical="center")
            # Colorer en vert si positif, rouge si négatif
            if agence_data['resultat'] >= 0:
                cell_resultat.font = Font(bold=True, color="006100")
            else:
                cell_resultat.font = Font(bold=True, color="C00000")
            
            row += 1
        
        # Ligne de séparation
        row += 1
        
        # TOTAL GÉNÉRAL
        ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL").font = grand_total_font
        ws.cell(row=row, column=1).fill = grand_total_fill
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right", vertical="center")
        
        # Total CA
        cell_total_ca = ws.cell(row=row, column=2, value=etat_data['total_ca'])
        cell_total_ca.font = grand_total_font
        cell_total_ca.fill = grand_total_fill
        cell_total_ca.number_format = '#,##0.00'
        cell_total_ca.alignment = Alignment(horizontal="right", vertical="center")
        
        # Total Dépenses
        cell_total_dep = ws.cell(row=row, column=3, value=etat_data['total_depenses'])
        cell_total_dep.font = grand_total_font
        cell_total_dep.fill = grand_total_fill
        cell_total_dep.number_format = '#,##0.00'
        cell_total_dep.alignment = Alignment(horizontal="right", vertical="center")
        
        # Total Marge
        total_marge_general = etat_data.get('total_marge', sum(ag['marge'] for ag in etat_data['resultats_par_agence'].values()))
        cell_total_marge = ws.cell(row=row, column=4, value=total_marge_general)
        cell_total_marge.font = grand_total_font
        cell_total_marge.fill = grand_total_fill
        cell_total_marge.number_format = '#,##0.00'
        cell_total_marge.alignment = Alignment(horizontal="right", vertical="center")
        
        # Total Résultat
        cell_total_resultat = ws.cell(row=row, column=5, value=etat_data['total_resultat'])
        cell_total_resultat.font = grand_total_font
        cell_total_resultat.fill = grand_total_fill
        cell_total_resultat.number_format = '#,##0.00'
        cell_total_resultat.alignment = Alignment(horizontal="right", vertical="center")
        if etat_data['total_resultat'] >= 0:
            cell_total_resultat.font = Font(bold=True, color="FFFFFF", size=13)
        else:
            cell_total_resultat.font = Font(bold=True, color="FFFFFF", size=13)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        
        # Réponse HTTP
        filename = f"Etat_Resultats_{etat_data['date_debut']}_{etat_data['date_fin']}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erreur export: {str(e)}'})

def generer_suivi_statistique(request):
    """Générer les statistiques de toutes les agences avec regroupement par catégorie de marge"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        date_debut_str = request.POST.get('date_debut')
        date_fin_str = request.POST.get('date_fin')
        
        if not date_debut_str or not date_fin_str:
            return JsonResponse({'success': False, 'error': 'Les dates sont obligatoires'})
        
        from datetime import datetime
        date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
        
        # Récupérer toutes les agences
        agences = Agence.objects.all().order_by('nom_agence')
        
        # Statistiques par agence et par jour
        statistiques_par_agence = {}
        
        for agence in agences:
            # Pour chaque jour de la période
            stats_journalieres = {}
            current_date = date_debut
            
            while current_date <= date_fin:
                # Récupérer les ventes de cette agence pour ce jour
                ventes_jour = LigneFactureVente.objects.filter(
                    facture_vente__agence=agence,
                    facture_vente__date=current_date
                ).select_related('article', 'facture_vente')
                
                # Catégories de marge
                bonne_marge = []      # ≥ 10%
                marge_mediocre = []   # 4% à 9%
                marge_faible = []     # < 4%
                
                # Articles déjà traités pour éviter les doublons par jour
                articles_traites = {}
                
                for vente in ventes_jour:
                    if not vente.article:
                        continue
                    
                    article_id = vente.article.id
                    prix_vente = Decimal(str(vente.prix_unitaire))
                    prix_achat = Decimal(str(vente.article.prix_achat))
                    quantite = Decimal(str(vente.quantite))
                    
                    # Calculer la marge en pourcentage
                    if prix_vente > 0:
                        marge_unitaire = prix_vente - prix_achat
                        pourcentage_marge = (marge_unitaire / prix_vente) * 100
                    else:
                        pourcentage_marge = Decimal('0')
                    
                    # Regrouper par article (sommer les quantités)
                    if article_id not in articles_traites:
                        articles_traites[article_id] = {
                            'article': vente.article,
                            'quantite_totale': Decimal('0'),
                            'chiffre_affaires': Decimal('0'),
                            'marge_totale': Decimal('0'),
                            'pourcentage_marge': pourcentage_marge,
                        }
                    
                    articles_traites[article_id]['quantite_totale'] += quantite
                    articles_traites[article_id]['chiffre_affaires'] += prix_vente * quantite
                    articles_traites[article_id]['marge_totale'] += marge_unitaire * quantite
                
                # Classer les articles par catégorie de marge
                for article_id, stats in articles_traites.items():
                    marge_pct = float(stats['pourcentage_marge'])
                    article_data = {
                        'reference': stats['article'].reference_article,
                        'designation': stats['article'].designation,
                        'quantite': float(stats['quantite_totale']),
                        'chiffre_affaires': float(stats['chiffre_affaires']),
                        'marge': float(stats['marge_totale']),
                        'pourcentage_marge': marge_pct,
                    }
                    
                    if marge_pct >= 10:
                        bonne_marge.append(article_data)
                    elif marge_pct >= 4:
                        marge_mediocre.append(article_data)
                    else:
                        marge_faible.append(article_data)
                
                if bonne_marge or marge_mediocre or marge_faible:
                    stats_journalieres[str(current_date)] = {
                        'bonne_marge': bonne_marge,
                        'marge_mediocre': marge_mediocre,
                        'marge_faible': marge_faible,
                    }
                
                # Jour suivant
                from datetime import timedelta
                current_date += timedelta(days=1)
            
            # N'inclure l'agence que si elle a des ventes sur la période
            if stats_journalieres:
                statistiques_par_agence[agence.id_agence] = {
                    'nom_agence': agence.nom_agence,
                    'stats_journalieres': stats_journalieres,
                }
        
        # Stocker dans la session pour l'export
        request.session['suivi_statistique'] = {
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
            'statistiques_par_agence': statistiques_par_agence,
        }
        
        return JsonResponse({
            'success': True,
            'statistiques': statistiques_par_agence,
            'date_debut': str(date_debut),
            'date_fin': str(date_fin),
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erreur: {str(e)}'})

def export_suivi_statistique_excel(request):
    """Exporter les statistiques de suivi en Excel"""
    compte, redirect_response = check_comptable_access(request)
    if redirect_response:
        return redirect_response
    
    try:
        # Récupérer les statistiques depuis la session
        stats_data = request.session.get('suivi_statistique')
        
        if not stats_data:
            return JsonResponse({'success': False, 'error': 'Aucune statistique générée'})
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return JsonResponse({'success': False, 'error': 'Module openpyxl non installé'})
        
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Suivi Statistique {stats_data['date_debut']} - {stats_data['date_fin']}"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        bonne_marge_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        mediocre_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        faible_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        agence_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row = 1
        
        # En-têtes des colonnes (sans "Catégorie Marge")
        headers = ['Référence', 'Désignation', 'Quantité', 'Chiffre d\'Affaires (FCFA)', 'Marge (FCFA)', 'Marge (%)']
        
        # Par agence
        for agence_id, agence_data in stats_data['statistiques_par_agence'].items():
            # En-tête Agence
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f"AGENCE: {agence_data['nom_agence']}"
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].fill = agence_fill
            ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
            row += 1
            
            # Par jour
            for date_jour, stats_jour in sorted(agence_data['stats_journalieres'].items()):
                # Date
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'] = f"Date: {date_jour}"
                ws[f'A{row}'].font = Font(bold=True, size=10)
                ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
                row += 1
                
                # Bonne marge (≥10%)
                bonne_marge_articles = stats_jour.get('bonne_marge', [])
                if bonne_marge_articles:
                    # Ligne d'en-tête de catégorie
                    ws.merge_cells(f'A{row}:F{row}')
                    cell_cat = ws.cell(row=row, column=1, value="Bonne Marge (≥10%)")
                    cell_cat.fill = bonne_marge_fill
                    cell_cat.font = Font(bold=True, size=10)
                    cell_cat.alignment = Alignment(horizontal="left", vertical="center")
                    row += 1
                    
                    # En-têtes du tableau
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    row += 1
                    
                    # Articles de bonne marge
                    for article in bonne_marge_articles:
                        # Référence (colonne A)
                        ws.cell(row=row, column=1, value=article['reference']).alignment = Alignment(vertical="center")
                        
                        # Désignation (colonne B)
                        ws.cell(row=row, column=2, value=article['designation']).alignment = Alignment(vertical="center")
                        
                        # Quantité (colonne C)
                        cell_qte = ws.cell(row=row, column=3, value=article['quantite'])
                        cell_qte.number_format = '#,##0.000'
                        cell_qte.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Chiffre d'Affaires (colonne D)
                        cell_ca = ws.cell(row=row, column=4, value=article['chiffre_affaires'])
                        cell_ca.number_format = '#,##0.00'
                        cell_ca.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Marge (colonne E)
                        cell_marge = ws.cell(row=row, column=5, value=article['marge'])
                        cell_marge.number_format = '#,##0.00'
                        cell_marge.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Marge % (colonne F)
                        ws.cell(row=row, column=6, value=f"{article['pourcentage_marge']:.2f}%").alignment = Alignment(horizontal="right", vertical="center")
                        
                        row += 1
                
                # Marge médiocre (4-9%)
                marge_mediocre_articles = stats_jour.get('marge_mediocre', [])
                if marge_mediocre_articles:
                    # Ligne d'en-tête de catégorie
                    ws.merge_cells(f'A{row}:F{row}')
                    cell_cat = ws.cell(row=row, column=1, value="Marge Médiocre (4-9%)")
                    cell_cat.fill = mediocre_fill
                    cell_cat.font = Font(bold=True, size=10)
                    cell_cat.alignment = Alignment(horizontal="left", vertical="center")
                    row += 1
                    
                    # En-têtes du tableau
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    row += 1
                    
                    # Articles de marge médiocre
                    for article in marge_mediocre_articles:
                        # Référence (colonne A)
                        ws.cell(row=row, column=1, value=article['reference']).alignment = Alignment(vertical="center")
                        
                        # Désignation (colonne B)
                        ws.cell(row=row, column=2, value=article['designation']).alignment = Alignment(vertical="center")
                        
                        # Quantité (colonne C)
                        cell_qte = ws.cell(row=row, column=3, value=article['quantite'])
                        cell_qte.number_format = '#,##0.000'
                        cell_qte.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Chiffre d'Affaires (colonne D)
                        cell_ca = ws.cell(row=row, column=4, value=article['chiffre_affaires'])
                        cell_ca.number_format = '#,##0.00'
                        cell_ca.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Marge (colonne E)
                        cell_marge = ws.cell(row=row, column=5, value=article['marge'])
                        cell_marge.number_format = '#,##0.00'
                        cell_marge.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Marge % (colonne F)
                        ws.cell(row=row, column=6, value=f"{article['pourcentage_marge']:.2f}%").alignment = Alignment(horizontal="right", vertical="center")
                        
                        row += 1
                
                # Marge faible (<4%)
                marge_faible_articles = stats_jour.get('marge_faible', [])
                if marge_faible_articles:
                    # Ligne d'en-tête de catégorie
                    ws.merge_cells(f'A{row}:F{row}')
                    cell_cat = ws.cell(row=row, column=1, value="Marge Faible (<4%)")
                    cell_cat.fill = faible_fill
                    cell_cat.font = Font(bold=True, size=10)
                    cell_cat.alignment = Alignment(horizontal="left", vertical="center")
                    row += 1
                    
                    # En-têtes du tableau
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    row += 1
                    
                    # Articles de marge faible
                    for article in marge_faible_articles:
                        # Référence (colonne A)
                        ws.cell(row=row, column=1, value=article['reference']).alignment = Alignment(vertical="center")
                        
                        # Désignation (colonne B)
                        ws.cell(row=row, column=2, value=article['designation']).alignment = Alignment(vertical="center")
                        
                        # Quantité (colonne C)
                        cell_qte = ws.cell(row=row, column=3, value=article['quantite'])
                        cell_qte.number_format = '#,##0.000'
                        cell_qte.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Chiffre d'Affaires (colonne D)
                        cell_ca = ws.cell(row=row, column=4, value=article['chiffre_affaires'])
                        cell_ca.number_format = '#,##0.00'
                        cell_ca.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Marge (colonne E)
                        cell_marge = ws.cell(row=row, column=5, value=article['marge'])
                        cell_marge.number_format = '#,##0.00'
                        cell_marge.alignment = Alignment(horizontal="right", vertical="center")
                        
                        # Marge % (colonne F)
                        ws.cell(row=row, column=6, value=f"{article['pourcentage_marge']:.2f}%").alignment = Alignment(horizontal="right", vertical="center")
                        
                        row += 1
                
                # Calculer les totaux pour cette date
                total_quantite = 0
                total_chiffre_affaires = 0
                total_marge = 0
                
                # Récupérer tous les articles de toutes les catégories
                tous_articles = bonne_marge_articles + marge_mediocre_articles + marge_faible_articles
                
                for article in tous_articles:
                    total_quantite += float(article['quantite'])
                    total_chiffre_affaires += float(article['chiffre_affaires'])
                    total_marge += float(article['marge'])
                
                # Calculer le pourcentage de marge global
                pourcentage_marge_global = 0
                if total_chiffre_affaires > 0:
                    pourcentage_marge_global = (total_marge / total_chiffre_affaires) * 100
                
                # Ajouter la ligne TOTAL GÉNÉRAL si on a des articles
                if tous_articles:
                    # Ligne vide avant le total
                    row += 1
                    
                    # Ligne TOTAL GÉNÉRAL
                    # Colonne A: "TOTAL GÉNÉRAL"
                    cell_total_label = ws.cell(row=row, column=1, value="TOTAL GÉNÉRAL")
                    cell_total_label.font = Font(bold=True, size=11)
                    cell_total_label.alignment = Alignment(horizontal="left", vertical="center")
                    
                    # Colonne B: vide (Désignation)
                    ws.cell(row=row, column=2, value="")
                    
                    # Colonne C: Quantité totale
                    cell_qte_total = ws.cell(row=row, column=3, value=total_quantite)
                    cell_qte_total.number_format = '#,##0.000'
                    cell_qte_total.font = Font(bold=True)
                    cell_qte_total.alignment = Alignment(horizontal="right", vertical="center")
                    
                    # Colonne D: Chiffre d'Affaires total
                    cell_ca_total = ws.cell(row=row, column=4, value=total_chiffre_affaires)
                    cell_ca_total.number_format = '#,##0.00'
                    cell_ca_total.font = Font(bold=True)
                    cell_ca_total.alignment = Alignment(horizontal="right", vertical="center")
                    
                    # Colonne E: Marge totale
                    cell_marge_total = ws.cell(row=row, column=5, value=total_marge)
                    cell_marge_total.number_format = '#,##0.00'
                    cell_marge_total.font = Font(bold=True)
                    cell_marge_total.alignment = Alignment(horizontal="right", vertical="center")
                    
                    # Colonne F: Pourcentage de marge global
                    cell_pct_total = ws.cell(row=row, column=6, value=f"{pourcentage_marge_global:.2f}%")
                    cell_pct_total.font = Font(bold=True)
                    cell_pct_total.alignment = Alignment(horizontal="right", vertical="center")
                    
                    row += 1
                
                row += 1  # Ligne vide entre les jours
            
            row += 1  # Ligne vide entre les agences
        
        # Ajuster largeurs colonnes (6 colonnes: A=Reference, B=Designation, C=Quantite, D=CA, E=Marge, F=Marge%)
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        
        # Réponse HTTP
        filename = f"Suivi_Statistique_{stats_data['date_debut']}_{stats_data['date_fin']}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': f'Erreur export: {str(e)}'})

# ==================== MODULE COMPTABILITÉ ====================

def login_comptabiliter(request):
    """Page de connexion pour le module Comptabilité"""
    # Toujours afficher la page de connexion, même si l'utilisateur est déjà connecté
    if request.user.is_authenticated and request.method == 'GET':
        logout(request)
        request.session.flush()
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                # Vérifier que l'utilisateur a un compte lié à une agence
                try:
                    compte = Compte.objects.get(user=user, actif=True)
                    if compte.agence:
                        # Vérifier que l'utilisateur est un comptable, assistant_comptable ou admin
                        if compte.type_compte not in ['comptable', 'assistant_comptable', 'admin']:
                            messages.error(request, 'Accès refusé. Ce module est réservé aux comptables, assistants comptables et aux administrateurs.')
                        else:
                            login(request, user)
                            # Stocker l'agence dans la session
                            request.session['agence_id'] = compte.agence.id_agence
                            request.session['agence_nom'] = compte.agence.nom_agence
                            messages.success(request, f'Connexion réussie ! Bienvenue {compte.nom_complet}')
                            return redirect('dashboard_comptabiliter')
                    else:
                        messages.error(request, 'Votre compte n\'est pas lié à une agence.')
                except Compte.DoesNotExist:
                    messages.error(request, 'Aucun compte actif trouvé pour cet utilisateur.')
            else:
                messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
        else:
            messages.error(request, 'Veuillez remplir tous les champs.')
    
    return render(request, 'supermarket/comptabiliter/login_comptabiliter.html')

def logout_comptabiliter(request):
    """Déconnexion du module Comptabilité"""
    logout(request)
    request.session.flush()
    messages.success(request, 'Vous avez été déconnecté avec succès.')
    return redirect('index')

@login_required
def dashboard_comptabiliter(request):
    """Dashboard principal du module de comptabilité/trésorerie"""
    # Imports nécessaires
    from django.db.models import Sum, F, DecimalField
    from django.utils.timezone import localtime
    from .models import Tresorerie, ChiffreAffaire, Compte, LigneFactureVente, Depense
    
    try:
        # 1. Récupérer l'agence
        agence = get_user_agence(request)
        if not agence:
            messages.error(request, "Votre compte n'est pas configuré correctement (Pas d'agence).")
            # On continue pour ne pas planter, mais l'affichage sera vide
        
        # 2. Gestion Date / Heure (LOCALE)
        now = timezone.now()
        local_now = localtime(now) # Heure du Cameroun
        today = local_now.date()

        # 3. CALCUL DU CA (Ventes du jour)
        # On calcule en direct depuis les lignes de facture
        ventes_du_jour = LigneFactureVente.objects.filter(
            facture_vente__agence=agence,
            facture_vente__date=today
        )
        
        resultat_ca = ventes_du_jour.aggregate(
            total=Sum(F('quantite') * F('prix_unitaire'), output_field=DecimalField())
        )['total']
        
        ca_du_jour = resultat_ca if resultat_ca is not None else 0

        # 4. CALCUL DES DÉPENSES (Journée entière)
        # On définit le début de la journée à 00:00:00
        debut_journee = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
        
        depenses_du_jour = Depense.objects.filter(
            agence=agence,
            date__gte=debut_journee  # Tout ce qui est après minuit
        ).aggregate(Sum('montant'))['montant__sum'] or 0


        historique_ca = ChiffreAffaire.objects.filter(
            agence=agence
        ).order_by('-date')[:5]

        # 5. Récupérer la Trésorerie (Dernier solde connu)
        last_treso = Tresorerie.objects.filter(agence=agence).order_by('-date').first()

        # 6. Récupérer le nom de l'utilisateur
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except:
            nom_utilisateur = request.user.username

        # 7. Envoi au Template
        context = {
            'agence': agence,
            'today': today,
            'ca_du_jour': ca_du_jour,
            'depenses_du_jour': depenses_du_jour,
            'treso': last_treso,
            'historique_ca': historique_ca,
            'nom_utilisateur': nom_utilisateur,
        }
        return render(request, 'supermarket/comptabiliter/dashboard.html', context)
        
    except Exception as e:
        # Gestion d'erreur globale pour ne pas planter la page
        messages.error(request, f'Erreur Dashboard: {str(e)}')
        # Fallback pour éviter l'écran blanc
        return render(request, 'supermarket/comptabiliter/dashboard.html', {
            'agence': None,
            'nom_utilisateur': request.user.username,
            'ca_du_jour': 0, 
            'depenses_du_jour': 0,
            'treso': None
        })

@login_required
def creer_depense(request): 
    """Vue pour créer une nouvelle dépense"""
    if request.method == 'POST':
        try:
            
            libelle = request.POST.get('libelle')
            date = request.POST.get('date')
            montant = request.POST.get('montant')
            
          
            if not all([libelle, date, montant]):
                messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                return redirect('creer_depense')
            
           
            depense = Depense.objects.create(
                libelle=libelle, 
                date=date,
                montant=montant,
            )
            
            messages.success(request, 'Dépense enregistrée avec succès!')
            
            return redirect('dashboard_comptabiliter') 
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de la dépense: {str(e)}')
            return redirect('creer_depense')
    
    # GET - Afficher le formulaire
    
    context = {
        # 'agences': agences,
    }
    # Chemin corrigé : comptabilite (pas comptabiliter)
    return render(request, 'supermarket/comptabiliter/creer-depense.html', context)


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

# Assurez-vous d'importer vos modèles et fonctions utilitaires
from .models import Depense, Compte, Agence
# Si get_user_agence est dans un autre fichier (ex: utils.py), importez-le. 
# Sinon, assurez-vous qu'il est défini ou accessible.
# from .utils import get_user_agence 

@login_required
def creer_depense(request): 
    # On importe le bon modèle
    from .models import Beneficiaire, Depense
    
    try:
        agence = get_user_agence(request)
        if not agence: raise ValueError
    except:
        return redirect('dashboard_comptabiliter')

    if request.method == 'POST':
        try:
            libelle = request.POST.get('libelle')
            date = request.POST.get('date')
            montant = request.POST.get('montant')
            nom_beneficiaire = request.POST.get('nom_employe') # On garde le name HTML
            
            if not all([libelle, date, montant, nom_beneficiaire]):
                messages.error(request, 'Champs obligatoires manquants.')
                return redirect('creer_depense')
            
            # --- CRÉATION AUTOMATIQUE DU BÉNÉFICIAIRE ---
            nom_propre = nom_beneficiaire.strip().upper()
            
            # On utilise Beneficiaire ici, pas Employe
            beneficiaire_obj, created = Beneficiaire.objects.get_or_create(
                agence=agence,
                nom_complet=nom_propre
            )

            Depense.objects.create(
                libelle=libelle, 
                date=date,
                montant=montant,
                agence=agence,
                beneficiaire=beneficiaire_obj # <-- Changement ici
            )
            
            messages.success(request, 'Dépense enregistrée !')
            return redirect('dashboard_comptabiliter')
            
        except Exception as e:
            messages.error(request, f'Erreur : {str(e)}')
            return redirect('creer_depense')
    
    # On récupère les bénéficiaires existants pour la liste déroulante
    beneficiaires_existants = Beneficiaire.objects.filter(agence=agence).order_by('nom_complet')
    
    return render(request, 'supermarket/comptabiliter/creer-depense.html', {
        'employes_existants': beneficiaires_existants # On garde la même variable pour le template
    })


@login_required
def consulter_depenses(request):
    """Affiche les dépenses avec filtre par Employé et export Excel complet"""
    # --- CORRECTION 1 : On importe Beneficiaire et HttpResponse ---
    from .models import Depense, Compte, Beneficiaire
    from django.db.models import Sum
    from django.utils.timezone import localtime
    from django.http import HttpResponse # Indispensable pour l'export Excel
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime

    # 1. Agence & Utilisateur
    try:
        agence = get_user_agence(request)
        if not agence: return redirect('dashboard_comptabiliter')
        
        try:
            compte = Compte.objects.get(user=request.user, actif=True)
            nom_utilisateur = compte.nom_complet
        except:
            nom_utilisateur = request.user.username
    except:
        return redirect('login_comptabiliter')

    # 2. Récupération des filtres
    date_debut_str = request.GET.get('date_debut')
    date_fin_str = request.GET.get('date_fin')
    
    # --- FILTRE DATE ---
    now = timezone.now()
    depenses = Depense.objects.filter(agence=agence).order_by('-date')

    if date_debut_str and date_fin_str:
        try:
            d_debut = timezone.make_aware(datetime.strptime(date_debut_str, '%Y-%m-%d').replace(hour=0, minute=0))
            d_fin = timezone.make_aware(datetime.strptime(date_fin_str, '%Y-%m-%d').replace(hour=23, minute=59))
            depenses = depenses.filter(date__range=[d_debut, d_fin])
            periode_info = f"Du {date_debut_str} au {date_fin_str}"
        except ValueError:
            periode_info = "Erreur Date"
    else:
        # Par défaut : Aujourd'hui minuit
        d_debut = now.replace(hour=0, minute=0, second=0, microsecond=0)
        depenses = depenses.filter(date__gte=d_debut)
        periode_info = "Aujourd'hui"

    # --- FILTRE BÉNÉFICIAIRE (Nouveau système) ---
    beneficiaire_id = request.GET.get('employe_id') # On garde le name du HTML 'employe_id'
    
    # On récupère la liste pour le menu déroulant
    tous_beneficiaires = Beneficiaire.objects.filter(agence=agence).order_by('nom_complet')

    if beneficiaire_id:
        # On filtre sur le champ 'beneficiaire' du modèle Depense
        depenses = depenses.filter(beneficiaire_id=beneficiaire_id)

    # Calcul Total
    total_general = depenses.aggregate(Sum('montant'))['montant__sum'] or 0

    # 3. EXPORT EXCEL
    if request.GET.get('export') == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Depenses_{agence.nom_agence}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dépenses"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="06beb6", end_color="06beb6", fill_type="solid")
        
        # En-têtes
        headers = ['Date', 'Bénéficiaire', 'Libellé', 'Montant']
        ws.append(headers)
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font

        row = 2
        for d in depenses:
            # Gestion de la date naive pour Excel
            d_date = d.date.replace(tzinfo=None) if d.date else ""
            
            # --- CORRECTION 2 : Utiliser d.beneficiaire et non d.employe ---
            nom_emp = d.beneficiaire.nom_complet if d.beneficiaire else "Non assigné"
            
            ws.append([d_date, nom_emp, d.libelle, d.montant])
            row += 1
            
        # Total en bas
        ws.cell(row=row, column=3, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_general).font = Font(bold=True)

        wb.save(response)
        return response

    context = {
        'depenses': depenses,
        'total_general': total_general or 0,
        'date_debut': date_debut_str,
        'date_fin': date_fin_str,
        'periode_info': periode_info,
        'tous_employes': tous_beneficiaires, # On passe les bénéficiaires au template
        'employe_id': int(beneficiaire_id) if beneficiaire_id else '',
        'agence': agence,
        'nom_utilisateur': nom_utilisateur,
    }
    return render(request, 'supermarket/comptabiliter/consulter_depenses.html', context)   
    
@login_required
def suivi_chiffre_affaire(request):
    # --- IMPORTS NÉCESSAIRES ---

    
    # 1. RECUPERATION AGENCE (Sécurisée)
    try:
        agence = get_user_agence(request)
    except Exception as e:
        messages.error(request, "Erreur d'identification de l'agence.")
        return redirect('dashboard_comptabiliter')

    # Gestion date pour l'affichage (Input date)
    date_str = request.POST.get('date') or request.GET.get('date')
    if date_str:
        date_a_afficher = date_str
    else:
        date_a_afficher = timezone.now().date()

    # 2. TRAITEMENT DU POST (Calcul & Enregistrement)
    if request.method == "POST":
        try:
            montant_previsionnel = request.POST.get('montant_previsionnel', 500000)

            # Calcul du CA du jour via les factures
            ventes_jour = LigneFactureVente.objects.filter(
                facture_vente__agence=agence,
                facture_vente__date=date_a_afficher
            )
            calcul_ca = ventes_jour.aggregate(total=Sum(F('quantite') * F('prix_unitaire')))['total'] or 0
            montant_realise = Decimal(calcul_ca)

            # Enregistrement avec liaison AGENCE
            ChiffreAffaire.objects.update_or_create(
                date=date_a_afficher,
                agence=agence,
                defaults={
                    'montant_realise': montant_realise,
                    'montant_previsionnel': montant_previsionnel
                }
            )
            messages.success(request, "Mise à jour effectuée.")

        except Exception as e:
            print(f"ERREUR POST: {e}") 
            messages.error(request, f"Erreur : {e}")

    # 3. RÉCUPÉRATION DES DONNÉES (Pour l'historique et l'Excel)
    # On filtre uniquement sur l'agence connectée
    donnees_ca = ChiffreAffaire.objects.filter(agence=agence).order_by('-date')

   # 4. EXPORT EXCEL (Format Rapport Journalier - 3 Colonnes)
    if request.GET.get('export') == 'excel':
        # Récupération des données spécifiques au jour affiché
        try:
            ca_export = ChiffreAffaire.objects.get(date=date_a_afficher, agence=agence)
            cap = ca_export.montant_previsionnel
            car = ca_export.montant_realise
            canr = ca_export.montant_non_realise
            p_car = f"{ca_export.pourcent_car}%"
            p_canr = f"{ca_export.pourcent_canr}%"
        except ChiffreAffaire.DoesNotExist:
            # Valeurs par défaut si pas de calcul
            cap = car = canr = 0
            p_car = p_canr = "0%"

        # Configuration Nom Fichier
        nom_agence_safe = agence.nom_agence.replace(" ", "_")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Rapport_CA_{nom_agence_safe}_{date_a_afficher}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"CA {date_a_afficher}"

        # --- STYLES ---
        # Titre
        title_font = Font(bold=True, size=14, color="2c3e50")
        subtitle_font = Font(bold=True, size=11, color="7f8c8d")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        # En-têtes Colonnes
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Couleurs des lignes de données (pour faire joli comme sur le site)
        fill_cap = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid") # Bleu clair
        fill_car = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid") # Vert clair
        fill_canr = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid") # Rouge clair

        # --- EN-TÊTE DU FICHIER (AGENCE & DATE) ---
        # Ligne 1 : Nom de l'agence
        ws.merge_cells('A1:C1')
        ws['A1'] = f"AGENCE : {agence.nom_agence.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        # Ligne 2 : Date
        ws.merge_cells('A2:C2')
        ws['A2'] = f"Rapport du : {date_a_afficher}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = center_align

        # Ligne 3 : Vide
        ws.append([])

        # --- TABLEAU 3 COLONNES ---
        
        # Ligne 4 : En-têtes des colonnes
        headers = ["Type de Chiffre d'Affaire", "Montant (FCFA)", "Pourcentage"]
        ws.append(headers)
        
        # Appliquer le style aux en-têtes (Ligne 4)
        for col in range(1, 4):
            cell = ws.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Ligne 5 : CAP (Objectif)
        ws.append(["C.A. Prévisionnel (CAP)", cap, "100%"])
        # Style de la ligne 5
        for col in range(1, 4):
            ws.cell(row=5, column=col).fill = fill_cap

        # Ligne 6 : CAR (Réalisé)
        ws.append(["C.A. Réalisé (CAR)", car, p_car])
        # Style de la ligne 6
        for col in range(1, 4):
            ws.cell(row=6, column=col).fill = fill_car

        # Ligne 7 : CANR (Manque)
        ws.append(["C.A. Non Réalisé (CANR)", canr, p_canr])
        # Style de la ligne 7
        for col in range(1, 4):
            ws.cell(row=7, column=col).fill = fill_canr

        # --- MISE EN FORME FINALE ---
        # Élargir les colonnes pour que ce soit lisible
        ws.column_dimensions['A'].width = 35  # Colonne Type large
        ws.column_dimensions['B'].width = 20  # Colonne Montant
        ws.column_dimensions['C'].width = 15  # Colonne %

        # Centrer les montants et pourcentages
        for row in range(5, 8):
            ws.cell(row=row, column=2).alignment = center_align
            ws.cell(row=row, column=3).alignment = center_align

        wb.save(response)
        return response

    # 5. OBJET DU JOUR (Pour affichage tableau 3 colonnes)
    try:
        ca_du_jour = ChiffreAffaire.objects.get(date=date_a_afficher, agence=agence)
    except ChiffreAffaire.DoesNotExist:
        ca_du_jour = None

    context = {
        'historique_ca': donnees_ca,
        'ca_du_jour': ca_du_jour,
        'date_a_afficher': date_a_afficher,
        'agence': agence,
       
        
    }

    return render(request, 'supermarket/comptabiliter/suivi-chiffre-affaire.html', context)



# supermarket/views.py

# supermarket/views.py

@login_required
def suivi_tresorerie(request):
    """
    Saisie journalière de la trésorerie.
    Affiche le CA calculé depuis les ventes du jour et les soldes en temps réel.
    Mise à jour automatique des soldes lors des transactions.
    """
    from .models import Tresorerie, ChiffreAffaire
    
    # 1. Agence
    try:
        agence = get_user_agence(request)
        if not agence: raise ValueError
    except:
        messages.error(request, "Agence non identifiée.")
        return redirect('dashboard_comptabilite')

    date_a_afficher = timezone.now().date()

    # --- CALCUL DU CA DEPUIS LES VENTES DU JOUR (PROBLÈME 1 RÉSOLU) ---
    # Calculer le CA directement depuis les FactureVente du jour
    ventes_jour = LigneFactureVente.objects.filter(
        facture_vente__agence=agence,
        facture_vente__date=date_a_afficher
    )
    ca_du_jour = ventes_jour.aggregate(total=Sum(F('quantite') * F('prix_unitaire')))['total'] or Decimal('0')
    
    # Créer un objet CA temporaire pour l'affichage
    class CAObject:
        def __init__(self, montant):
            self.montant_realise = montant
    ca_obj = CAObject(ca_du_jour)

    # --- LOGIQUE DE REPORT DE SOLDE (ROLLING BALANCE) ---
    # IMPORTANT : Le solde initial = solde final du jour précédent (solde actuel)
    # Il reste FIXE pour toute la journée et est déterminé une seule fois au début
    
    # 1. Récupérer la trésorerie du jour si elle existe (pour les dépôts/retraits actuels)
    try:
        treso = Tresorerie.objects.get(agence=agence, date=date_a_afficher)
    except Tresorerie.DoesNotExist:
        treso = None
    
    # 2. DÉTERMINER LE SOLDE INITIAL (toujours depuis le jour précédent, même si treso existe)
    # Chercher le DERNIER JOUR ENREGISTRÉ pour déterminer le solde initial
    last_treso = Tresorerie.objects.filter(agence=agence, date__lt=date_a_afficher).order_by('-date').first()
    
    if last_treso:
        # On prend le SOLDE FINAL du jour précédent comme INITIAL d'aujourd'hui (solde actuel)
        valeurs_initiales = {
            'banque': float(last_treso.solde_banque_final),
            'caisse': float(last_treso.solde_caisse_final),
            'om': float(last_treso.solde_om_final),
            'momo': float(last_treso.solde_momo_final),
            'sav': float(last_treso.solde_sav_final)
        }
    else:
        # C'est la toute première fois qu'on utilise le logiciel
        valeurs_initiales = {'banque': 0, 'caisse': 0, 'om': 0, 'momo': 0, 'sav': 0}

    # 5. Enregistrement (POST) avec mise à jour automatique des soldes (PROBLÈME 2 RÉSOLU)
    if request.method == "POST":
        try:
            def get_val(name):
                """Récupère une valeur POST et la convertit en Decimal de manière sécurisée"""
                val = request.POST.get(name, '').strip()
                if not val or val == '':
                    return Decimal('0')
                try:
                    # Remplacer les virgules par des points pour la conversion
                    val = val.replace(',', '.')
                    return Decimal(str(val))
                except (ValueError, TypeError, Exception):
                    return Decimal('0')

            # Récupérer les valeurs saisies depuis le formulaire
            banque_initial = get_val('banque_initial')
            banque_depot_saisi = get_val('banque_depot')  # Nouveau montant saisi
            banque_retrait_saisi = get_val('banque_retrait')  # Nouveau montant saisi
            
            # Vérifier si une trésorerie existe déjà pour aujourd'hui
            treso_existante = Tresorerie.objects.filter(agence=agence, date=date_a_afficher).first()
            
            if treso_existante:
                # Si elle existe, AJOUTER les nouveaux montants aux montants existants
                # Seulement si un montant > 0 est saisi, sinon conserver la valeur existante
                banque_depot = treso_existante.banque_depot + banque_depot_saisi if banque_depot_saisi > 0 else treso_existante.banque_depot
                banque_retrait = treso_existante.banque_retrait + banque_retrait_saisi if banque_retrait_saisi > 0 else treso_existante.banque_retrait
                
                caisse_entree_saisi = get_val('caisse_entree')
                caisse_sortie_saisi = get_val('caisse_sortie')
                caisse_entree = treso_existante.caisse_entree + caisse_entree_saisi if caisse_entree_saisi > 0 else treso_existante.caisse_entree
                caisse_sortie = treso_existante.caisse_sortie + caisse_sortie_saisi if caisse_sortie_saisi > 0 else treso_existante.caisse_sortie
                
                om_depot_saisi = get_val('om_depot')
                om_retrait_saisi = get_val('om_retrait')
                om_depot = treso_existante.om_depot + om_depot_saisi if om_depot_saisi > 0 else treso_existante.om_depot
                om_retrait = treso_existante.om_retrait + om_retrait_saisi if om_retrait_saisi > 0 else treso_existante.om_retrait
                
                momo_depot_saisi = get_val('momo_depot')
                momo_retrait_saisi = get_val('momo_retrait')
                momo_depot = treso_existante.momo_depot + momo_depot_saisi if momo_depot_saisi > 0 else treso_existante.momo_depot
                momo_retrait = treso_existante.momo_retrait + momo_retrait_saisi if momo_retrait_saisi > 0 else treso_existante.momo_retrait
                
                sav_entree_saisi = get_val('sav_entree')
                sav_sortie_saisi = get_val('sav_sortie')
                sav_entree = treso_existante.sav_entree + sav_entree_saisi if sav_entree_saisi > 0 else treso_existante.sav_entree
                sav_sortie = treso_existante.sav_sortie + sav_sortie_saisi if sav_sortie_saisi > 0 else treso_existante.sav_sortie
            else:
                # Si elle n'existe pas, utiliser directement les valeurs saisies
                banque_depot = banque_depot_saisi
                banque_retrait = banque_retrait_saisi
                caisse_entree = get_val('caisse_entree')
                caisse_sortie = get_val('caisse_sortie')
                om_depot = get_val('om_depot')
                om_retrait = get_val('om_retrait')
                momo_depot = get_val('momo_depot')
                momo_retrait = get_val('momo_retrait')
                sav_entree = get_val('sav_entree')
                sav_sortie = get_val('sav_sortie')
            
            # Récupérer les valeurs initiales (pour compatibilité avec le code existant)
            caisse_initial = get_val('caisse_initial')
            om_initial = get_val('om_initial')
            momo_initial = get_val('momo_initial')
            sav_initial = get_val('sav_initial')

            # IMPORTANT : Le solde initial reste FIXE (celui calculé au début de la journée)
            # Si une trésorerie existe déjà, on conserve son solde initial
            # Sinon, on utilise les valeurs_initiales calculées depuis le jour précédent
            # Seuls les dépôts/retraits peuvent être modifiés
            
            if treso_existante:
                # Si elle existe, on conserve les soldes initiaux existants (fixes pour la journée)
                banque_initial_val = treso_existante.banque_initial
                caisse_initial_val = treso_existante.caisse_initial
                om_initial_val = treso_existante.om_initial
                momo_initial_val = treso_existante.momo_initial
                sav_initial_val = treso_existante.sav_initial
            else:
                # Si elle n'existe pas, on utilise les valeurs initiales calculées
                banque_initial_val = Decimal(str(valeurs_initiales['banque']))
                caisse_initial_val = Decimal(str(valeurs_initiales['caisse']))
                om_initial_val = Decimal(str(valeurs_initiales['om']))
                momo_initial_val = Decimal(str(valeurs_initiales['momo']))
                sav_initial_val = Decimal(str(valeurs_initiales['sav']))
            
            # Créer ou mettre à jour la trésorerie
            treso, created = Tresorerie.objects.update_or_create(
                agence=agence,
                date=date_a_afficher,
                defaults={
                    # Les soldes initiaux sont FIXES (conservés s'ils existent, sinon calculés)
                    'banque_initial': banque_initial_val,
                    'caisse_initial': caisse_initial_val,
                    'om_initial': om_initial_val,
                    'momo_initial': momo_initial_val,
                    'sav_initial': sav_initial_val,
                    
                    # Seuls les dépôts/retraits peuvent être modifiés
                    'banque_depot': banque_depot,
                    'banque_retrait': banque_retrait,
                    'caisse_entree': caisse_entree,
                    'caisse_sortie': caisse_sortie,
                    'om_depot': om_depot,
                    'om_retrait': om_retrait,
                    'momo_depot': momo_depot,
                    'momo_retrait': momo_retrait,
                    'sav_entree': sav_entree,
                    'sav_sortie': sav_sortie,
                }
            )
            
            # Mise à jour automatique : mettre à jour les soldes initiaux du jour suivant
            # avec les soldes finaux calculés du jour actuel
            date_suivante = date_a_afficher + timedelta(days=1)
            Tresorerie.objects.update_or_create(
                agence=agence,
                date=date_suivante,
                defaults={
                    'banque_initial': treso.solde_banque_final,
                    'caisse_initial': treso.solde_caisse_final,
                    'om_initial': treso.solde_om_final,
                    'momo_initial': treso.solde_momo_final,
                    'sav_initial': treso.solde_sav_final,
                    # Les dépôts/retraits restent à 0 pour le jour suivant
                    'banque_depot': 0,
                    'banque_retrait': 0,
                    'caisse_entree': 0,
                    'caisse_sortie': 0,
                    'om_depot': 0,
                    'om_retrait': 0,
                    'momo_depot': 0,
                    'momo_retrait': 0,
                    'sav_entree': 0,
                    'sav_sortie': 0,
                }
            )
            
            # Récupérer les valeurs finales pour le message
            treso.refresh_from_db()
            messages.success(
                request, 
                f"Trésorerie mise à jour avec succès ! "
                f"Banque: {treso.banque_depot} FCFA dépôt, "
                f"Solde final: {treso.solde_banque_final} FCFA. "
                f"Vous pouvez continuer à modifier les dépôts/retraits. "
                f"Le solde initial ({treso.banque_initial} FCFA) reste fixe."
            )
            # Redirection vers le mode visualisation pour voir les résultats
            from django.urls import reverse
            # Rediriger vers la page sans le paramètre edit pour revenir en mode visualisation
            return redirect(reverse('suivi_tresorerie'))
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f"Erreur lors de l'enregistrement: {str(e)}")
            # En cas d'erreur, rester sur la page pour afficher le message

    # Debug : vérifier les valeurs initiales
    print(f"DEBUG - Valeurs initiales: {valeurs_initiales}")
    print(f"DEBUG - Treso existe: {treso is not None}")
    if treso:
        print(f"DEBUG - Treso banque_initial: {treso.banque_initial}")
    
    # Calculer le total global pour l'affichage
    total_global = sum(valeurs_initiales.values())
    
    context = {
        'treso': treso,
        'init': valeurs_initiales,
        'total_global': total_global,
        'date_a_afficher': date_a_afficher,
        'agence': agence,
        'ca': ca_obj 
    }
    return render(request, 'supermarket/comptabiliter/suivi_tresorerie.html', context)

@login_required
def consulter_tresorerie(request):
    """
    Par défaut : 30 derniers jours.
    Si filtre : 1 seule journée (Affichage + Export).
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.db.models import Sum
    from .models import Tresorerie, ChiffreAffaire
    
    # 1. Identification Agence
    try:
        agence = get_user_agence(request)
    except:
        return redirect('dashboard_comptabilite')
    
    # --- 2. LOGIQUE HYBRIDE (30 jours OU 1 jour) ---
    date_filter = request.GET.get('date') # On cherche une date unique
    now = timezone.now().date()
    
    if date_filter:
        # CAS FILTRÉ : On force le début et la fin à la MÊME date
        date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
        date_debut = date_obj
        date_fin = date_obj
        mode_filtre = True
    else:
        # CAS PAR DÉFAUT : Période de 30 jours
        date_fin = now
        date_debut = now - timedelta(days=30)
        mode_filtre = False

    # 3. Récupération des données pour l'affichage HTML
    # Utiliser date__lte pour inclure la date de fin (aujourd'hui)
    # IMPORTANT: Utiliser date__lte avec date_fin pour inclure aujourd'hui
    historique = Tresorerie.objects.filter(
        agence=agence, 
        date__gte=date_debut,
        date__lte=date_fin
    ).order_by('-date')
    
    # Debug: vérifier si aujourd'hui est inclus
    print(f"DEBUG HISTORIQUE - date_debut: {date_debut}, date_fin: {date_fin}, now: {now}")
    print(f"DEBUG HISTORIQUE - Nombre d'entrées: {historique.count()}")

    # Ajout du CA temporaire pour l'affichage tableau HTML (calculé depuis les ventes)
    for t in historique:
        # Calculer le CA directement depuis les FactureVente du jour
        ventes_jour = LigneFactureVente.objects.filter(
            facture_vente__agence=agence,
            facture_vente__date=t.date
        )
        ca_du_jour = ventes_jour.aggregate(total=Sum(F('quantite') * F('prix_unitaire')))['total'] or Decimal('0')
        t.ca_temp = ca_du_jour

    # 4. EXPORT EXCEL (Votre code Excel reste identique, il s'adapte aux dates)
    if request.GET.get('export') == 'excel':
        
        # Calculs du CA depuis les ventes
        # IMPORTANT : Utiliser exactement la même logique que le tableau HTML
        # Si une seule journée dans l'historique, utiliser le CA de cette journée (déjà calculé)
        # Sinon, calculer le total cumulé
        
        # Convertir le queryset en liste pour pouvoir vérifier la longueur
        historique_list = list(historique)
        
        if len(historique_list) == 1:
            # Une seule journée : utiliser le CA déjà calculé pour cette journée (comme dans le tableau HTML)
            total_ca = historique_list[0].ca_temp if hasattr(historique_list[0], 'ca_temp') else Decimal('0')
        elif mode_filtre and date_debut == date_fin:
            # Mode filtre avec une seule date : calculer pour cette date uniquement
            ventes_periode = LigneFactureVente.objects.filter(
                facture_vente__agence=agence,
                facture_vente__date=date_debut
            )
            total_ca = ventes_periode.aggregate(total=Sum(F('quantite') * F('prix_unitaire')))['total'] or Decimal('0')
        else:
            # Période de plusieurs jours : calculer le total cumulé
            ventes_periode = LigneFactureVente.objects.filter(
                facture_vente__agence=agence,
                facture_vente__date__range=[date_debut, date_fin]
            )
            total_ca = ventes_periode.aggregate(total=Sum(F('quantite') * F('prix_unitaire')))['total'] or Decimal('0')

        # Derniers Soldes
        last_treso = Tresorerie.objects.filter(
            agence=agence, 
            date__lte=date_fin
        ).order_by('-date').first()

        if last_treso:
            banque = last_treso.solde_banque_final
            caisse = last_treso.solde_caisse_final
            om = last_treso.solde_om_final
            momo = last_treso.solde_momo_final
            sav = last_treso.solde_sav_final
            total_dispo = last_treso.total_disponible
            date_arrete = last_treso.date
        else:
            banque = caisse = om = momo = sav = total_dispo = 0
            date_arrete = date_fin

        # --- CRÉATION EXCEL ---
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Nom du fichier adapté
        if mode_filtre:
            filename = f"Bilan_Jour_{agence.nom_agence}_{date_debut}.xlsx"
            titre_excel = f"BILAN JOURNALIER : {date_debut}"
        else:
            filename = f"Bilan_Periode_{agence.nom_agence}_{date_debut}_{date_fin}.xlsx"
            titre_excel = f"Période du {date_debut} au {date_fin}"
            
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bilan"

        # Styles (Votre code existant)
        title_font = Font(bold=True, size=16, color="2c3e50")
        subtitle_font = Font(size=12, italic=True, color="7f8c8d")
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        ca_fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
        total_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        total_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # Remplissage
        ws.merge_cells('A1:B1')
        ws['A1'] = f"BILAN : {agence.nom_agence.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:B2')
        ws['A2'] = titre_excel # Utilisation du titre dynamique
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = center_align

        ws.append([])

        ws.append(["LIBELLÉ", "MONTANT (FCFA)"])
        ws['A4'].fill = header_fill; ws['A4'].font = header_font
        ws['B4'].fill = header_fill; ws['B4'].font = header_font; ws['B4'].alignment = center_align

        # CA
        libelle_ca = "Chiffre d'Affaire (Journée)" if mode_filtre else "Chiffre d'Affaire CUMULÉ"
        ws.append([libelle_ca, total_ca])
        ws['A5'].fill = ca_fill; ws['B5'].fill = ca_fill; ws['B5'].font = Font(bold=True)

        ws.append(["", ""])

        # Trésorerie
        ws.append([f"Solde Banque (au {date_arrete})", banque])
        ws.append([f"Solde Caisse (au {date_arrete})", caisse])
        ws.append([f"Solde Orange (au {date_arrete})", om])
        ws.append([f"Solde MTN (au {date_arrete})", momo])
        ws.append([f"Solde SAV (au {date_arrete})", sav])

        ws.append(["", ""])

        # Total
        row_total = ws.max_row + 1
        ws.append(["TOTAL DISPONIBLE", total_dispo])
        ws.cell(row=row_total, column=1).fill = total_fill; ws.cell(row=row_total, column=1).font = total_font
        ws.cell(row=row_total, column=2).fill = total_fill; ws.cell(row=row_total, column=2).font = total_font

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        for row in range(5, ws.max_row + 1):
            ws.cell(row=row, column=2).alignment = right_align
            ws.cell(row=row, column=2).number_format = '#,##0'

        wb.save(response)
        return response

    context = {
        'historique': historique,
        # On renvoie la date unique si on filtre, sinon None
        'date_filter': date_filter if mode_filtre else '', 
        'agence': agence
    }
    return render(request, 'supermarket/comptabiliter/consulter_tresorerie.html', context)